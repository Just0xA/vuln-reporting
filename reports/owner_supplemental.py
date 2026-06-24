"""
reports/owner_supplemental.py — Combined Owner/Application supplemental Excel + CSV writer.

Produces a flat ``Owner | Application | Open Findings | Asset Count`` workbook
for analyst drill-down and tagging-cleanup worklist (SEG-03, D-09, D-10).

``Unassigned`` Owner rows double as the "needs Owner assignment" worklist.

PII rules (D-10, D-11):
  - Written only to ``output/`` (gitignored) — never to ``data/trend/``.
  - May be emailed internally; MUST NOT be committed or fed to AI/Claude.

Design follows ``reports/unscanned_assets.py`` Excel/CSV pattern.
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from reports.modules.board_report_utils import extract_owner
from utils.open_count import open_findings_at

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants (matching unscanned_assets.py palette)
# ---------------------------------------------------------------------------

_FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # dark navy
_FILL_ALT    = PatternFill("solid", fgColor="F5F5F5")   # zebra stripe

_SUPPLEMENTAL_COLS: list[tuple[str, str]] = [
    ("Owner",         "owner"),
    ("Application",   "application"),
    ("Open Findings", "open_count"),
    ("Asset Count",   "asset_count"),
]

_EXCEL_FILENAME = "owner_segmentation.xlsx"
_CSV_FILENAME   = "owner_segmentation.csv"
_TAB_NAME       = "Owner Assignment"


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _safe_cell_value(val: object) -> object:
    """
    Convert a raw DataFrame value to something safe for openpyxl.

    - pandas NA / NaT / None → empty string
    - Lists/tuples → comma-joined string
    - CSV-injection guard: prefix values starting with ``= + - @`` with
      a leading apostrophe (T-13-09).
    - Everything else → as-is
    """
    if isinstance(val, (list, tuple)):
        val = ", ".join(str(v) for v in val if v)
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if val is None:
        return ""
    s = str(val)
    if s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return val


def _build_owner_app_df(
    assets_df: pd.DataFrame,
    vulns_df: pd.DataFrame,
    report_date: Optional[datetime] = None,
) -> pd.DataFrame:
    """
    Build the flat (owner, application) aggregation frame.

    Columns: owner, application, open_count, asset_count.

    - Per (owner, application) group: distinct asset UUIDs and vuln row count.
    - No SLA/severity computation (RESEARCH open-question 3 — minimal v1).
    """
    if assets_df.empty:
        return pd.DataFrame(columns=["owner", "application", "open_count", "asset_count"])

    enriched = extract_owner(assets_df)

    # Dedup on asset_uuid (keep-first owner/application) so one physical asset
    # contributes to exactly ONE (owner, application) row.  This is the same
    # CR-01 keep-first pattern used by the open-count path below — applying it
    # once here makes both the asset-count and open-count paths consistent.
    enriched_deduped = (
        enriched[["asset_uuid", "owner", "application"]]
        .drop_duplicates("asset_uuid")
    )

    # Asset count per (owner, application) — built from deduped frame so a
    # duplicate asset_uuid is counted exactly once (Phase-13 WR-01 fix).
    asset_counts = (
        enriched_deduped
        .groupby(["owner", "application"], dropna=False)["asset_uuid"]
        .nunique()
        .reset_index(name="asset_count")
    )

    if vulns_df.empty or "asset_uuid" not in vulns_df.columns:
        asset_counts["open_count"] = 0
        return asset_counts[["owner", "application", "open_count", "asset_count"]]

    # WR-02: filter to the open set at report_date before aggregating, so the
    # supplemental "Open Findings" column ties out to the owner trend snapshot.
    # When report_date is None, count raw export rows (back-compat for callers
    # that do not thread a date).
    open_vulns = open_findings_at(vulns_df, report_date) if report_date is not None else vulns_df

    # CR-01: reuse the already-deduped frame for the uuid→owner index so that
    # the open-count path uses the same first-row-wins attribution as asset_counts.
    uuid_to_owner = enriched_deduped.set_index("asset_uuid")

    # Open findings count per asset_uuid, then join owner/application via enriched.
    # Use .assign() for fillna casts (F-DTYPE convention — avoids CoW chained assignment).
    # Break the join result into a named frame before .assign() so pandas CoW tracking
    # does not see chained mutation through an intermediate slice.
    vuln_owner = open_vulns[["asset_uuid"]].copy().join(uuid_to_owner, on="asset_uuid", how="left")
    vuln_owner = vuln_owner.assign(
        owner=vuln_owner["owner"].fillna("Unassigned"),
        application=vuln_owner["application"].fillna(""),
    )

    open_counts = (
        vuln_owner
        .groupby(["owner", "application"], dropna=False)
        .size()
        .reset_index(name="open_count")
    )

    result = asset_counts.merge(open_counts, on=["owner", "application"], how="left")
    result = result.assign(open_count=result["open_count"].fillna(0).astype(int))
    result = result.sort_values(["owner", "application"]).reset_index(drop=True)
    return result[["owner", "application", "open_count", "asset_count"]]


def _write_excel_tab(
    wb: openpyxl.Workbook,
    df: pd.DataFrame,
) -> None:
    """Write the 'Owner Assignment' tab into workbook wb."""
    ws = wb.create_sheet(_TAB_NAME)

    if df.empty:
        ws["A1"] = "No data for this run."
        ws["A1"].font = Font(italic=True, color="888888")
        return

    # Header row
    for col_idx, (header, _) in enumerate(_SUPPLEMENTAL_COLS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # Data rows — zebra stripe
    for row_idx in range(len(df)):
        row_series = df.iloc[row_idx]
        alt = (row_idx % 2 == 1)
        for col_idx, (_, field) in enumerate(_SUPPLEMENTAL_COLS, start=1):
            raw = row_series.get(field) if field in df.columns else None
            val = _safe_cell_value(raw)
            cell           = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left")
            if alt:
                cell.fill = _FILL_ALT

    ws.freeze_panes = "A2"


def _write_csv(csv_path: Path, df: pd.DataFrame) -> None:
    """Write the flat supplemental CSV with all columns quoted, utf-8-sig encoding."""
    fieldnames = [col for _, col in _SUPPLEMENTAL_COLS]
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(
            fh, fieldnames=fieldnames, extrasaction="ignore", quoting=csv.QUOTE_ALL
        )
        writer.writeheader()
        for _, row in df.iterrows():
            record = {
                field: _safe_cell_value(row.get(field) if field in df.columns else None)
                for field in fieldnames
            }
            writer.writerow(record)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_owner_supplemental(
    assets_df:   pd.DataFrame,
    vulns_df:    pd.DataFrame,
    output_dir:  Path,
    report_date: Optional[datetime] = None,
) -> dict:
    """
    Write the combined Owner/Application supplemental Excel and CSV.

    Produces one flat tab ``Owner Assignment`` with columns:
    ``Owner | Application | Open Findings | Asset Count``.

    ``Unassigned`` owner rows serve as a tagging-cleanup worklist.

    Files are written to ``output_dir`` only (gitignored ``output/``).
    Never writes to ``data/trend/`` (D-11).

    Parameters
    ----------
    assets_df : pd.DataFrame
        In-scope assets DataFrame from ``fetch_all_assets()``.
    vulns_df : pd.DataFrame
        Open findings DataFrame for the same scope.
    output_dir : Path
        Run output directory.  Created if missing.
    report_date : datetime, optional
        Point-in-time reference for the open-set filter (WR-02).  When provided,
        "Open Findings" counts only findings open at this date via
        ``open_findings_at()``, so the supplemental ties out to the owner trend
        snapshot from the same run.  When ``None``, counts raw export rows
        (back-compat for callers that do not thread a date).

    Returns
    -------
    dict
        ``{"supplemental_excel": Path, "supplemental_csv": Path}``
        Both paths point to files inside ``output_dir``.
    """
    output_dir = Path(output_dir)
    # CR-S4: fail-fast guard — refuse an output_dir that resolves inside
    # data/trend/ to prevent PII row-level data from landing in the no-PII
    # trend store (project_pii_rule_is_ai_not_email policy).
    _resolved = output_dir.resolve()
    _trend_resolved = (Path(__file__).resolve().parent.parent / "data" / "trend").resolve()
    if _resolved == _trend_resolved or _trend_resolved in _resolved.parents:
        raise ValueError(
            f"owner_supplemental: output_dir {str(output_dir)!r} resolves inside "
            "data/trend/ — PII output-path policy violation "
            "(project_pii_rule_is_ai_not_email)."
        )
    output_dir.mkdir(parents=True, exist_ok=True)

    df = _build_owner_app_df(assets_df, vulns_df, report_date=report_date)

    # Excel
    wb = openpyxl.Workbook()
    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)
    _write_excel_tab(wb, df)
    excel_path = output_dir / _EXCEL_FILENAME
    wb.save(str(excel_path))
    logger.info("owner_supplemental: Excel written → %s", excel_path)

    # CSV
    csv_path = output_dir / _CSV_FILENAME
    _write_csv(csv_path, df)
    logger.info("owner_supplemental: CSV written → %s", csv_path)

    return {
        "supplemental_excel": excel_path,
        "supplemental_csv":   csv_path,
    }
