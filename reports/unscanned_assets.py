"""
reports/unscanned_assets.py — Unscanned / Not-On-Time Asset Report.

Lists all assets that are not scanning on time relative to a configurable
scan window (default 30 days), using the same deduplication and licensed-
asset logic as the board_summary Scan Coverage SLA metric so counts can be
reconciled directly.

Assets are split into two not-on-time categories:

    1. Overdue Licensed  — last_licensed_scan_date IS NOT NULL but older
                           than the scan window.  These have been licensed-
                           scanned before, but not recently enough.
    2. No Licensed Scan  — last_licensed_scan_date IS NULL.  Tenable has
                           no record of a licensed scan for these assets.

Both categories are written to the Excel workbook.  The Summary tab shows
counts that reconcile directly with the board Scan Coverage SLA metric:

    Scan Coverage % = On-Time / (On-Time + Overdue Licensed)

    (No Licensed Scan assets are excluded from both numerator and
    denominator — they are shown separately for investigative purposes.)

Usage (standalone)
------------------
    python reports/unscanned_assets.py
    python reports/unscanned_assets.py --tag-category "Owner" --tag-value "Network Defense"
    python reports/unscanned_assets.py --scan-window-days 45 --output-dir output/scan_gap

Via run_all.py (delivery group)
--------------------------------
    Registered as slug "unscanned_assets".
    Optional group-config key:
        scan_window_days: 30   # integer, defaults to 30

Outputs
-------
    unscanned_assets.xlsx  — 3 tabs: Summary | Overdue Licensed | No Licensed Scan
    unscanned_assets.csv   — flat file combining both not-on-time categories
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import CACHE_DIR, OUTPUT_DIR
from data.fetchers import fetch_all_assets
from reports.modules.board_report_utils import deduplicate_assets_by_name

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_DEFAULT_SCAN_WINDOW_DAYS: int = 30
_EXCEL_FILENAME = "unscanned_assets.xlsx"
_CSV_FILENAME   = "unscanned_assets.csv"

# openpyxl fills
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # dark navy
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")   # light red
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")   # light amber
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")   # light green
_FILL_ALT    = PatternFill("solid", fgColor="F5F5F5")   # zebra stripe

# Column specs: (Header label, DataFrame column name)
_OVERDUE_COLS: list[tuple[str, str]] = [
    ("Hostname",               "hostname"),
    ("IP Address",             "ipv4"),
    ("FQDN",                   "fqdn"),
    ("Operating System",       "operating_system"),
    ("Last Licensed Scan",     "last_licensed_scan_date"),
    ("Days Since Lic. Scan",   "days_since_licensed_scan"),
    ("Last Seen",              "last_seen"),
    ("Days Since Last Seen",   "days_since_last_seen"),
    ("Last Scan Time",         "last_scan_time"),
    ("Source",                 "source_name"),
    ("Tags",                   "tags_str"),
    ("Asset UUID",             "asset_uuid"),
]

_UNLICENSED_COLS: list[tuple[str, str]] = [
    ("Hostname",             "hostname"),
    ("IP Address",           "ipv4"),
    ("FQDN",                 "fqdn"),
    ("Operating System",     "operating_system"),
    ("Last Seen",            "last_seen"),
    ("Days Since Last Seen", "days_since_last_seen"),
    ("First Seen",           "first_seen"),
    ("Has Plugin Results",   "has_plugin_results"),
    ("Source",               "source_name"),
    ("Tags",                 "tags_str"),
    ("Asset UUID",           "asset_uuid"),
]

_DATE_FIELDS = frozenset({
    "last_licensed_scan_date", "last_seen", "last_scan_time", "first_seen"
})

_COL_WIDTHS: dict[str, int] = {
    "Hostname":               30,
    "IP Address":             16,
    "FQDN":                   30,
    "Operating System":       22,
    "Last Licensed Scan":     20,
    "Days Since Lic. Scan":   22,
    "Last Seen":              16,
    "Days Since Last Seen":   20,
    "Last Scan Time":         20,
    "First Seen":             16,
    "Has Plugin Results":     18,
    "Source":                 22,
    "Tags":                   42,
    "Asset UUID":             38,
}


# ===========================================================================
# Public API — called by run_all.py
# ===========================================================================

def run_report(
    tio,
    run_id: str,
    *,
    tag_category:     Optional[str]      = None,
    tag_value:        Optional[str]      = None,
    output_dir:       Optional[Path]     = None,
    generated_at:     Optional[datetime] = None,
    cache_dir:        Optional[Path]     = None,
    scan_window_days: int                = _DEFAULT_SCAN_WINDOW_DAYS,
) -> dict:
    """
    Generate the Unscanned / Not-On-Time Asset Report.

    Parameters
    ----------
    tio : TenableIO
        Authenticated Tenable client.
    run_id : str
        Cache key (typically YYYY-MM-DD).
    tag_category : str, optional
        Tenable tag category to scope the report (e.g. ``"Owner"``).
    tag_value : str, optional
        Tag value paired with ``tag_category``.  Both must be non-empty
        to apply a filter; otherwise all assets are included.
    output_dir : Path, optional
        Directory for output files.  Defaults to
        ``OUTPUT_DIR / "unscanned_assets"``.
    generated_at : datetime, optional
        UTC-aware report timestamp.  Defaults to UTC now.
    cache_dir : Path, optional
        Parquet cache directory.
    scan_window_days : int
        Recency window in days.  Assets whose ``last_licensed_scan_date``
        is older than this value are classified as overdue.  Must match
        the board Scan Coverage SLA window (default 30) for counts to
        reconcile.

    Returns
    -------
    dict
        ``{"pdf": None, "excel": path_or_none, "csv": path_or_none,
           "charts": [], "metrics": {counts...}}``
    """
    if generated_at is None:
        generated_at = datetime.now(tz=timezone.utc)
    if cache_dir is None:
        cache_dir = CACHE_DIR / datetime.now().strftime("%Y-%m-%d")
    if output_dir is None:
        output_dir = OUTPUT_DIR / "unscanned_assets"

    output_dir = Path(output_dir)
    cache_dir  = Path(cache_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)

    scope_log = (
        f"{tag_category}={tag_value}" if tag_category and tag_value else "all assets"
    )
    logger.info(
        "unscanned_assets: starting (scope=%s, window=%dd, run_id=%s)",
        scope_log, scan_window_days, run_id,
    )

    # ------------------------------------------------------------------ #
    # Fetch                                                                #
    # ------------------------------------------------------------------ #
    logger.info("unscanned_assets: fetching assets …")
    assets_df = fetch_all_assets(tio, cache_dir)
    logger.info("unscanned_assets: %d raw records.", len(assets_df))

    # ------------------------------------------------------------------ #
    # Tag filter                                                           #
    # ------------------------------------------------------------------ #
    if tag_category and tag_value:
        assets_df = _filter_by_tag(assets_df, tag_category, tag_value)
        logger.info(
            "unscanned_assets: '%s=%s' → %d assets.",
            tag_category, tag_value, len(assets_df),
        )

    # ------------------------------------------------------------------ #
    # Deduplicate — same logic as Scan Coverage SLA module                #
    # ------------------------------------------------------------------ #
    assets_df = deduplicate_assets_by_name(assets_df)
    logger.info("unscanned_assets: %d after dedup.", len(assets_df))

    # ------------------------------------------------------------------ #
    # Coerce date columns to UTC-aware Timestamps                         #
    # ------------------------------------------------------------------ #
    assets_df = assets_df.copy()
    for col in ("last_licensed_scan_date", "last_seen", "last_scan_time", "first_seen"):
        if col in assets_df.columns:
            assets_df.loc[:, col] = pd.to_datetime(
                assets_df[col], utc=True, errors="coerce"
            )

    # ------------------------------------------------------------------ #
    # Compute cutoff                                                       #
    # ------------------------------------------------------------------ #
    if hasattr(generated_at, "tzinfo") and generated_at.tzinfo is not None:
        rd_ts = pd.Timestamp(generated_at).tz_convert("UTC")
    else:
        rd_ts = pd.Timestamp(generated_at, tz="UTC")
    cutoff = rd_ts - pd.Timedelta(days=scan_window_days)

    # ------------------------------------------------------------------ #
    # Split into three sets                                                #
    # ------------------------------------------------------------------ #
    lsd = "last_licensed_scan_date"
    if lsd not in assets_df.columns:
        assets_df.loc[:, lsd] = pd.NaT

    licensed_mask    = assets_df[lsd].notna()
    on_time_mask     = licensed_mask & (assets_df[lsd] >= cutoff)
    overdue_mask     = licensed_mask & (assets_df[lsd] <  cutoff)
    unlicensed_mask  = ~licensed_mask

    on_time_df    = assets_df[on_time_mask].copy().reset_index(drop=True)
    overdue_df    = assets_df[overdue_mask].copy().reset_index(drop=True)
    unlicensed_df = assets_df[unlicensed_mask].copy().reset_index(drop=True)

    # ------------------------------------------------------------------ #
    # Add computed age columns                                             #
    # ------------------------------------------------------------------ #
    overdue_df    = _add_age_columns(overdue_df,    rd_ts)
    unlicensed_df = _add_age_columns(unlicensed_df, rd_ts)

    # Sort: most overdue / longest-unseen first
    if not overdue_df.empty:
        overdue_df = overdue_df.sort_values(
            "days_since_licensed_scan", ascending=False, na_position="last"
        ).reset_index(drop=True)
    if not unlicensed_df.empty:
        unlicensed_df = unlicensed_df.sort_values(
            "days_since_last_seen", ascending=False, na_position="last"
        ).reset_index(drop=True)

    metrics = {
        "total_assets":     len(assets_df),
        "on_time":          len(on_time_df),
        "overdue_licensed": len(overdue_df),
        "unlicensed":       len(unlicensed_df),
        "scan_window_days": scan_window_days,
        "cutoff_date":      cutoff.strftime("%Y-%m-%d"),
    }
    logger.info(
        "unscanned_assets: total=%d  on_time=%d  overdue=%d  unlicensed=%d",
        metrics["total_assets"], metrics["on_time"],
        metrics["overdue_licensed"], metrics["unlicensed"],
    )

    # ------------------------------------------------------------------ #
    # Excel                                                                #
    # ------------------------------------------------------------------ #
    excel_path: Optional[Path] = None
    try:
        wb = openpyxl.Workbook()
        if wb.worksheets:
            wb.remove(wb.worksheets[0])

        _write_summary_tab(wb, metrics, generated_at, scope_log)
        _write_data_tab(wb, "Overdue Licensed", overdue_df, _OVERDUE_COLS)
        _write_data_tab(wb, "No Licensed Scan",  unlicensed_df, _UNLICENSED_COLS)

        excel_file  = output_dir / _EXCEL_FILENAME
        wb.save(str(excel_file))
        excel_path  = excel_file
        logger.info("unscanned_assets: Excel → %s", excel_file)
    except Exception as exc:  # noqa: BLE001
        logger.error("unscanned_assets: Excel failed: %s", exc, exc_info=True)

    # ------------------------------------------------------------------ #
    # CSV                                                                  #
    # ------------------------------------------------------------------ #
    csv_path: Optional[Path] = None
    try:
        csv_file = output_dir / _CSV_FILENAME
        _write_csv(csv_file, overdue_df, unlicensed_df)
        csv_path = csv_file
        logger.info("unscanned_assets: CSV → %s", csv_file)
    except Exception as exc:  # noqa: BLE001
        logger.error("unscanned_assets: CSV failed: %s", exc, exc_info=True)

    return {
        "pdf":     None,
        "excel":   excel_path,
        "csv":     csv_path,
        "charts":  [],
        "metrics": metrics,
    }


# ===========================================================================
# Private helpers
# ===========================================================================

def _filter_by_tag(
    assets_df:    pd.DataFrame,
    tag_category: str,
    tag_value:    str,
) -> pd.DataFrame:
    """Exact-token match on the semicolon-delimited ``tags`` column."""
    if "tags" not in assets_df.columns:
        logger.warning("_filter_by_tag: 'tags' column absent — returning unfiltered.")
        return assets_df
    target = f"{tag_category}={tag_value}"

    def _has(tags_str: object) -> bool:
        if not isinstance(tags_str, str) or not tags_str.strip():
            return False
        return any(t.strip() == target for t in tags_str.split(";"))

    return assets_df[assets_df["tags"].apply(_has)].copy().reset_index(drop=True)


def _add_age_columns(df: pd.DataFrame, rd_ts: pd.Timestamp) -> pd.DataFrame:
    """
    Add two integer age columns (in days) to ``df``:

    - ``days_since_licensed_scan`` — days since ``last_licensed_scan_date``
    - ``days_since_last_seen``     — days since ``last_seen``

    Null dates produce ``pd.NA`` (nullable integer).
    """
    df = df.copy()

    lsd = "last_licensed_scan_date"
    if lsd in df.columns:
        df.loc[:, "days_since_licensed_scan"] = (
            (rd_ts - df[lsd])
            .dt.days
            .where(df[lsd].notna(), other=pd.NA)
            .astype("Int64")
        )
    else:
        df.loc[:, "days_since_licensed_scan"] = pd.NA

    ls = "last_seen"
    if ls in df.columns:
        df.loc[:, "days_since_last_seen"] = (
            (rd_ts - df[ls])
            .dt.days
            .where(df[ls].notna(), other=pd.NA)
            .astype("Int64")
        )
    else:
        df.loc[:, "days_since_last_seen"] = pd.NA

    return df


def _fmt_date(val: object) -> str:
    """Format a pandas Timestamp / datetime-like value as ``YYYY-MM-DD``."""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if val is None:
        return ""
    try:
        return pd.Timestamp(val).strftime("%Y-%m-%d")
    except Exception:  # noqa: BLE001
        return str(val)


def _safe_cell_value(val: object, field: str) -> object:
    """
    Convert a raw DataFrame value to something safe to write into openpyxl.

    - Date fields → ``YYYY-MM-DD`` string
    - Lists/tuples → comma-joined string
    - pandas NA / NaT / float NaN → empty string
    - Everything else → as-is
    """
    if field in _DATE_FIELDS:
        return _fmt_date(val)
    if isinstance(val, (list, tuple)):
        return ", ".join(str(v) for v in val if v)
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if val is None:
        return ""
    return val


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------

def _write_summary_tab(
    wb:          openpyxl.Workbook,
    metrics:     dict,
    generated_at: datetime,
    scope_log:   str,
) -> None:
    """Write the Summary tab with reconciliation counts and run metadata."""
    ws = wb.create_sheet("Summary")

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"]      = "Unscanned / Not-On-Time Asset Report — Summary"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")

    # Metadata rows
    meta_rows: list[tuple[str, object, str | None]] = [
        ("Generated",    generated_at.strftime("%Y-%m-%d %H:%M UTC"), None),
        ("Scope",        scope_log,                                    None),
        ("Scan Window",  f"{metrics['scan_window_days']} days  (cutoff: {metrics['cutoff_date']})", None),
    ]
    for row_idx, (label, value, _) in enumerate(meta_rows, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    # Counts table
    header_row = 7
    ws.cell(row=header_row, column=1, value="Category").font   = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=1).fill                     = _FILL_HEADER
    ws.cell(row=header_row, column=2, value="Asset Count").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=2).fill                     = _FILL_HEADER
    ws.cell(row=header_row, column=3, value="Notes").font      = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=3).fill                     = _FILL_HEADER

    total = metrics["total_assets"] or 1   # avoid ZeroDivisionError

    count_rows = [
        (
            "On Time (scanned within window)",
            metrics["on_time"],
            f"{metrics['on_time'] / total * 100:.1f}%  ← Scan Coverage SLA numerator",
            _FILL_GREEN,
        ),
        (
            "Overdue Licensed (has scan date, but stale)",
            metrics["overdue_licensed"],
            "Licensed but not scanned recently — investigate in Tenable",
            _FILL_RED,
        ),
        (
            "No Licensed Scan (unlicensed)",
            metrics["unlicensed"],
            "Excluded from Scan Coverage SLA metric (not in denominator)",
            _FILL_AMBER,
        ),
        (
            "Total Deduplicated Assets",
            metrics["total_assets"],
            "On Time + Overdue Licensed + No Licensed Scan",
            None,
        ),
    ]

    for offset, (label, count, note, fill) in enumerate(count_rows, start=1):
        r = header_row + offset
        label_cell = ws.cell(row=r, column=1, value=label)
        count_cell = ws.cell(row=r, column=2, value=count)
        note_cell  = ws.cell(row=r, column=3, value=note)
        label_cell.font = Font(bold=True if "Total" in label else False)
        count_cell.alignment = Alignment(horizontal="right")
        if fill:
            label_cell.fill = fill
            count_cell.fill = fill

    # Reconciliation note
    note_row = header_row + len(count_rows) + 2
    ws.merge_cells(f"A{note_row}:C{note_row}")
    ws[f"A{note_row}"] = (
        "Scan Coverage SLA % = On Time ÷ (On Time + Overdue Licensed)  "
        "— unlicensed assets are excluded from both numerator and denominator."
    )
    ws[f"A{note_row}"].font = Font(italic=True, color="555555", size=8)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60


def _write_data_tab(
    wb:       openpyxl.Workbook,
    tab_name: str,
    df:       pd.DataFrame,
    col_spec: list[tuple[str, str]],
) -> None:
    """
    Write a data tab (Overdue Licensed or No Licensed Scan).

    Rows are zebra-striped for readability.  Header row is pinned (freeze panes).
    """
    ws = wb.create_sheet(tab_name)

    if df.empty:
        ws["A1"] = f"No assets in this category for this run."
        ws["A1"].font = Font(italic=True, color="888888")
        return

    # Header
    for col_idx, (header, _) in enumerate(col_spec, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # Data
    for row_idx in range(len(df)):
        row_series = df.iloc[row_idx]
        alt = (row_idx % 2 == 1)

        for col_idx, (header, field) in enumerate(col_spec, start=1):
            raw = row_series.get(field) if field in df.columns else None
            val = _safe_cell_value(raw, field)

            cell           = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left")
            if alt:
                cell.fill = _FILL_ALT

    # Column widths
    for col_idx, (header, _) in enumerate(col_spec, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            _COL_WIDTHS.get(header, 18)
        )

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# CSV writer
# ---------------------------------------------------------------------------

def _write_csv(
    csv_file:     Path,
    overdue_df:   pd.DataFrame,
    unlicensed_df: pd.DataFrame,
) -> None:
    """
    Write a flat CSV combining both not-on-time categories.

    Includes a ``Category`` column so the two groups are distinguishable
    when imported into a ticketing or tracking system.
    """
    fieldnames = [
        "Category",
        "hostname", "ipv4", "fqdn", "operating_system",
        "last_licensed_scan_date", "days_since_licensed_scan",
        "last_seen", "days_since_last_seen",
        "last_scan_time", "source_name", "tags_str", "asset_uuid",
    ]

    def _iter_rows(df: pd.DataFrame, category: str):
        for _, row in df.iterrows():
            record = {"Category": category}
            for col in fieldnames[1:]:
                raw = row.get(col) if col in df.columns else None
                record[col] = _safe_cell_value(raw, col)
            yield record

    with open(csv_file, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for record in _iter_rows(overdue_df,    "Overdue Licensed"):
            writer.writerow(record)
        for record in _iter_rows(unlicensed_df, "No Licensed Scan"):
            writer.writerow(record)


# ===========================================================================
# CLI entry point
# ===========================================================================

def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Unscanned / Not-On-Time Asset Report — "
            "lists assets not scanning within the configured window."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # All assets, default 30-day window
  python reports/unscanned_assets.py

  # Scoped to a tag
  python reports/unscanned_assets.py --tag-category "Owner" --tag-value "Network Defense"

  # 45-day window, custom output directory
  python reports/unscanned_assets.py --scan-window-days 45 --output-dir output/scan_gap
        """,
    )
    parser.add_argument("--tag-category", metavar="CATEGORY",
                        help="Tenable tag category to scope report")
    parser.add_argument("--tag-value",    metavar="VALUE",
                        help="Tag value paired with --tag-category")
    parser.add_argument("--scan-window-days", metavar="N", type=int,
                        default=_DEFAULT_SCAN_WINDOW_DAYS,
                        help=f"Scan recency window in days (default: {_DEFAULT_SCAN_WINDOW_DAYS})")
    parser.add_argument("--output-dir",   metavar="PATH",
                        help="Output directory (default: output/unscanned_assets/)")
    parser.add_argument("--run-id",       metavar="ID", default=None,
                        help="Parquet cache key (default: today's local date YYYY-MM-DD)")
    parser.add_argument("--no-email",     action="store_true",
                        help="Informational — email is managed by run_all.py for this report")
    return parser


def main() -> int:
    """CLI entry point."""
    from dotenv import load_dotenv  # noqa: PLC0415
    from config import LOG_DIR, LOG_LEVEL  # noqa: PLC0415

    logging.basicConfig(
        level=getattr(logging, LOG_LEVEL, logging.INFO),
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(LOG_DIR / "app.log", encoding="utf-8"),
        ],
    )

    load_dotenv()

    args   = _build_arg_parser().parse_args()
    run_id = args.run_id or datetime.now().strftime("%Y-%m-%d")
    output = Path(args.output_dir) if args.output_dir else None

    try:
        from tenable_client import get_client  # noqa: PLC0415
        tio = get_client()
    except SystemExit:
        raise
    except Exception as exc:
        logger.error("Tenable connection failed: %s", exc, exc_info=True)
        return 1

    result = run_report(
        tio              = tio,
        run_id           = run_id,
        tag_category     = args.tag_category,
        tag_value        = args.tag_value,
        output_dir       = output,
        scan_window_days = args.scan_window_days,
    )

    m = result.get("metrics", {})
    print(f"Excel: {result.get('excel') or '(not generated — see logs)'}")
    print(f"CSV:   {result.get('csv')   or '(not generated — see logs)'}")
    print()
    print(f"Total deduplicated assets : {m.get('total_assets', 0):,}")
    print(f"On time                   : {m.get('on_time', 0):,}")
    print(f"Overdue licensed          : {m.get('overdue_licensed', 0):,}")
    print(f"No licensed scan          : {m.get('unlicensed', 0):,}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
