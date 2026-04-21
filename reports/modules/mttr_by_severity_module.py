"""
reports/modules/mttr_by_severity_module.py — Mean Time to Remediate (MTTR) by Severity.

Average days from first discovery to remediation for closed vulnerabilities,
per severity tier.

The preferred days_to_fix source is the ``time_taken_to_fix`` field (seconds
converted to days) provided directly by Tenable.  When absent, the fallback is
``(last_fixed - first_found).days``.

Usage
-----
::

    from reports.modules import registry
    from reports.modules.base import ModuleConfig

    mod  = registry.get("mttr_by_severity")()
    data = mod.compute(fixed_vulns_df, assets_df, report_date,
                       ModuleConfig("mttr_by_severity"))
    print(data.metrics["critical_mttr"])  # e.g. 12.3  (days)

    # Require at least 5 fixed findings to compute MTTR for a severity:
    data = mod.compute(
        fixed_vulns_df, assets_df, report_date,
        ModuleConfig("mttr_by_severity", options={"min_sample_size": 5}),
    )
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SLA_DAYS
from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.chart_utils import draw_gauge
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# Severity display order
_SEVERITIES: tuple[str, ...] = ("critical", "high", "medium", "low")

# SLA comparison thresholds
_WITHIN_SLA_LIMIT = 1.0
_NEAR_LIMIT       = 1.25

# Status labels — matching the existing management_summary.py labels
_STATUS_WITHIN = "Within SLA"
_STATUS_NEAR   = "Near SLA Limit"
_STATUS_EXCEED = "Exceeding SLA"
_STATUS_NODATA = "No Data"

# Status colours
_COLOR_GREEN  = "#388e3c"
_COLOR_AMBER  = "#fbc02d"
_COLOR_RED    = "#d32f2f"
_COLOR_GREY   = "#9E9E9E"

# Excel status fills
_FILL_GREEN = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED   = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY  = PatternFill("solid", fgColor="F5F5F5")


def _status_from_ratio(ratio: float) -> tuple[str, str]:
    """Return (status_label, hex_color) from the MTTR / SLA ratio."""
    if ratio <= _WITHIN_SLA_LIMIT:
        return _STATUS_WITHIN, _COLOR_GREEN
    if ratio <= _NEAR_LIMIT:
        return _STATUS_NEAR, _COLOR_AMBER
    return _STATUS_EXCEED, _COLOR_RED


def _status_fill(status: Optional[str]) -> PatternFill:
    """Return the Excel fill for a status string."""
    if status == _STATUS_WITHIN:
        return _FILL_GREEN
    if status == _STATUS_NEAR:
        return _FILL_AMBER
    if status == _STATUS_EXCEED:
        return _FILL_RED
    return _FILL_GREY


@register_module
class MTTRBySeverityModule(BaseModule):
    """
    Compute Mean Time to Remediate (MTTR) for closed vulnerabilities per severity.

    Calculation (per finding):
        Preferred:  ``days_to_fix = time_taken_to_fix / 86400``
                    when ``time_taken_to_fix`` is not null and > 0
        Fallback:   ``days_to_fix = (last_fixed - first_found).days``
                    clipped to ≥ 0

    Findings where neither source yields a usable value are excluded from
    the MTTR calculation.

    The ``vulns_df`` passed to this module must include both fixed and open
    findings (the compute method filters to FIXED state internally).
    Alternatively, pass a pre-filtered DataFrame of only fixed findings.

    Supported options
    -----------------
    min_sample_size : int, optional
        Minimum number of remediated findings required to compute MTTR for a
        severity tier.  Tiers with fewer findings are reported as None / No Data.
        Default: 1.
    """

    MODULE_ID         = "mttr_by_severity"
    DISPLAY_NAME      = "Mean Time to Remediate (MTTR) by Severity"
    DESCRIPTION       = ("Average days from first discovery to remediation "
                         "for closed vulnerabilities, per severity tier.")
    REQUIRED_DATA     = ["vulns"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute per-severity MTTR from fixed/remediated vulnerability findings.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  May contain all states —
            this method filters to FIXED rows internally.  Must have
            ``state``, ``severity``, ``time_taken_to_fix``, ``last_fixed``,
            and ``first_found`` columns (or a subset; missing columns are handled).
        assets_df : pd.DataFrame
            Not used; accepted for interface compatibility.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Options: ``min_sample_size`` (int, default 1).
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            min_sample = int(config.options.get("min_sample_size", 1))

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            # ---- Filter to fixed/remediated findings ----
            if vulns_df.empty or "state" not in vulns_df.columns:
                fixed_df = pd.DataFrame()
            else:
                fixed_mask = (
                    vulns_df["state"].str.lower().isin({"fixed"}) |
                    vulns_df.get("last_fixed", pd.Series(dtype=object)).notna()
                )
                fixed_df = vulns_df[fixed_mask].copy()

            if fixed_df.empty:
                return self._build_result(
                    per_sev_mttr   = {sev: None for sev in _SEVERITIES},
                    per_sev_status = {sev: _STATUS_NODATA for sev in _SEVERITIES},
                    per_sev_color  = {sev: _COLOR_GREY for sev in _SEVERITIES},
                    per_sev_sample = {sev: 0 for sev in _SEVERITIES},
                    sample_size    = 0,
                    computed_at    = computed_at,
                    config         = config,
                )

            # ---- Compute days_to_fix vectorised ----
            ttf_valid = (
                fixed_df["time_taken_to_fix"].notna() &
                (pd.to_numeric(fixed_df["time_taken_to_fix"], errors="coerce") > 0)
                if "time_taken_to_fix" in fixed_df.columns
                else pd.Series(False, index=fixed_df.index)
            )
            ttf_days = (
                pd.to_numeric(fixed_df["time_taken_to_fix"], errors="coerce").div(86400)
                if "time_taken_to_fix" in fixed_df.columns
                else pd.Series(float("nan"), index=fixed_df.index)
            )

            _nat_series = pd.Series(
                [pd.NaT] * len(fixed_df), index=fixed_df.index, dtype="object"
            )
            last_fixed_ts  = pd.to_datetime(
                fixed_df["last_fixed"]  if "last_fixed"  in fixed_df.columns else _nat_series,
                utc=True, errors="coerce",
            )
            first_found_ts = pd.to_datetime(
                fixed_df["first_found"] if "first_found" in fixed_df.columns else _nat_series,
                utc=True, errors="coerce",
            )
            date_diff_days = (last_fixed_ts - first_found_ts).dt.days.clip(lower=0)

            fixed_df = fixed_df.assign(
                days_to_fix=ttf_days.where(ttf_valid, other=date_diff_days)
            )
            # Drop rows where we cannot compute a days_to_fix value
            fixed_df = fixed_df[
                fixed_df["days_to_fix"].notna() & (fixed_df["days_to_fix"] >= 0)
            ]

            sample_size = len(fixed_df)

            # ---- Per-severity MTTR ----
            per_sev_mttr:   dict[str, Optional[float]] = {}
            per_sev_status: dict[str, str]              = {}
            per_sev_color:  dict[str, str]              = {}
            per_sev_sample: dict[str, int]              = {}

            for sev in _SEVERITIES:
                sla = SLA_DAYS[sev]
                sev_df = fixed_df[fixed_df["severity"].str.lower() == sev]
                n      = len(sev_df)
                per_sev_sample[sev] = n

                if n < min_sample:
                    per_sev_mttr[sev]   = None
                    per_sev_status[sev] = _STATUS_NODATA
                    per_sev_color[sev]  = _COLOR_GREY
                else:
                    mttr = round(float(sev_df["days_to_fix"].mean()), 1)
                    per_sev_mttr[sev] = mttr
                    status, color = _status_from_ratio(mttr / sla)
                    per_sev_status[sev] = status
                    per_sev_color[sev]  = color

            return self._build_result(
                per_sev_mttr   = per_sev_mttr,
                per_sev_status = per_sev_status,
                per_sev_color  = per_sev_color,
                per_sev_sample = per_sev_sample,
                sample_size    = sample_size,
                computed_at    = computed_at,
                config         = config,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # Internal builder
    # ------------------------------------------------------------------

    def _build_result(
        self,
        per_sev_mttr:   dict[str, Optional[float]],
        per_sev_status: dict[str, str],
        per_sev_color:  dict[str, str],
        per_sev_sample: dict[str, int],
        sample_size:    int,
        computed_at:    str,
        config:         ModuleConfig,
    ) -> ModuleData:
        """Assemble ModuleData from computed MTTR values."""

        # Overall MTTR = mean of non-None per-severity MTTRs
        valid_mttrs = [v for v in per_sev_mttr.values() if v is not None]
        overall_mttr: Optional[float] = (
            round(sum(valid_mttrs) / len(valid_mttrs), 1) if valid_mttrs else None
        )

        metrics: dict[str, Any] = {
            "overall_mttr":   overall_mttr,
            "critical_mttr":  per_sev_mttr.get("critical"),
            "high_mttr":      per_sev_mttr.get("high"),
            "medium_mttr":    per_sev_mttr.get("medium"),
            "low_mttr":       per_sev_mttr.get("low"),
            "critical_status": per_sev_status.get("critical", _STATUS_NODATA),
            "high_status":     per_sev_status.get("high",     _STATUS_NODATA),
            "medium_status":   per_sev_status.get("medium",   _STATUS_NODATA),
            "low_status":      per_sev_status.get("low",      _STATUS_NODATA),
            "sample_size":     sample_size,
            "critical_sample": per_sev_sample.get("critical", 0),
            "high_sample":     per_sev_sample.get("high",     0),
            "medium_sample":   per_sev_sample.get("medium",   0),
            "low_sample":      per_sev_sample.get("low",      0),
        }

        table_data = [
            {
                "severity":    sev.capitalize(),
                "mttr_days":   per_sev_mttr.get(sev),
                "sla_days":    SLA_DAYS[sev],
                "status":      per_sev_status.get(sev, _STATUS_NODATA),
                "sample_size": per_sev_sample.get(sev, 0),
            }
            for sev in _SEVERITIES
        ]

        # Build summary text per severity
        parts: list[str] = []
        for sev in _SEVERITIES:
            mttr = per_sev_mttr.get(sev)
            sla  = SLA_DAYS[sev]
            if mttr is None:
                parts.append(
                    f"Insufficient remediated data for {sev.capitalize()} MTTR calculation."
                )
            else:
                parts.append(
                    f"{sev.capitalize()} vulnerabilities are being remediated in an average "
                    f"of {mttr:.1f} days (SLA target: {sla} days)."
                )
        summary_text = " ".join(parts)

        return ModuleData(
            module_id    = self.MODULE_ID,
            display_name = self.DISPLAY_NAME,
            metrics      = metrics,
            table_data   = table_data,
            chart_data   = {
                "per_severity": {sev: per_sev_mttr.get(sev) for sev in _SEVERITIES},
                "sla_targets":  dict(SLA_DAYS),
                "statuses":     {sev: per_sev_status.get(sev) for sev in _SEVERITIES},
            },
            summary_text = summary_text,
            metadata     = {
                "days_to_fix_method": (
                    "time_taken_to_fix (seconds converted to days) where available, "
                    "otherwise (last_fixed - first_found).days"
                ),
                "sla_windows":  dict(SLA_DAYS),
                "computed_at":  computed_at,
            },
            error        = None,
        )

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render four per-severity MTTR gauges side by side plus an SLA reference table.

        Returns an error callout if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m = data.metrics
        gauges_html = ""

        for sev in _SEVERITIES:
            mttr    = m.get(f"{sev}_mttr")
            sla     = float(SLA_DAYS[sev])
            status  = m.get(f"{sev}_status", _STATUS_NODATA)
            color   = (
                _COLOR_GREEN if status == _STATUS_WITHIN else
                _COLOR_AMBER if status == _STATUS_NEAR else
                _COLOR_RED   if status == _STATUS_EXCEED else
                _COLOR_GREY
            )
            thresholds = [
                (sla,        color),
                (sla * 1.25, _COLOR_AMBER),
                (sla * 2,    _COLOR_RED),
            ]

            if mttr is None:
                gauge_inner = (
                    f'<div style="text-align:center;padding:16pt;">'
                    f'<span style="color:#9E9E9E;font-size:9pt;">No Data</span>'
                    f"</div>"
                )
            else:
                b64 = draw_gauge(
                    value=mttr,
                    min_val=0,
                    max_val=sla * 2,
                    thresholds=thresholds,
                    title=f"MTTR — {sev.capitalize()}",
                    unit="d",
                    reference_line=sla,
                    reference_label="SLA",
                )
                gauge_inner = (
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="width:100%;max-width:160pt;">'
                )

            gauges_html += (
                f'<div style="display:inline-block;text-align:center;'
                f'width:23%;margin:0 1%;">'
                f"{gauge_inner}"
                f"</div>"
            )

        # SLA reference table
        table_rows = ""
        for row in data.table_data:
            mttr_str = f"{row['mttr_days']:.1f}d" if row["mttr_days"] is not None else "N/A"
            table_rows += (
                f"<tr>"
                f"<td>{row['severity']}</td>"
                f"<td>{mttr_str}</td>"
                f"<td>{row['sla_days']}d</td>"
                f"<td>{row['status']}</td>"
                f"</tr>"
            )

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <div style="text-align:center;margin-bottom:8pt;">
    {gauges_html}
  </div>
  <table class="data-table" style="width:100%;margin-top:8pt;">
    <thead>
      <tr>
        <th>Severity</th><th>MTTR</th><th>SLA Target</th><th>Status</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>
  <p class="explanatory-text">
    Mean Time to Remediate (MTTR) measures the average number of days between
    a vulnerability being first discovered and being confirmed as fixed. This
    is compared against our SLA targets. Green indicates we are meeting targets.
    Red means remediation is taking longer than agreed and may require escalation.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a single "MTTR" tab with per-severity MTTR vs. SLA comparison.

        Returns ``[]`` on error; writes an error row if ``data.error`` is set.
        """
        tab_name = "MTTR"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            # ---- Headers ----
            headers = ["Severity", "MTTR (Days)", "SLA Target (Days)", "Variance (Days)",
                       "Status", "Sample Size"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            # ---- Data rows ----
            for row_idx, row in enumerate(data.table_data, start=2):
                mttr     = row["mttr_days"]
                sla      = row["sla_days"]
                variance = round(mttr - sla, 1) if mttr is not None else None

                ws.cell(row=row_idx, column=1, value=row["severity"])
                ws.cell(row=row_idx, column=2, value=mttr if mttr is not None else "N/A")
                ws.cell(row=row_idx, column=3, value=sla)
                ws.cell(row=row_idx, column=4,
                        value=variance if variance is not None else "N/A")

                status_cell = ws.cell(row=row_idx, column=5, value=row["status"])
                status_cell.fill = _status_fill(row["status"])

                ws.cell(row=row_idx, column=6, value=row["sample_size"])

            # ---- Column widths ----
            widths = [14, 14, 18, 16, 16, 14]
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_kpis
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return MTTR for Critical and High tiers for email KPI tiles.
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m = data.metrics
        crit = m.get("critical_mttr")
        high = m.get("high_mttr")
        return {
            "MTTR — Critical": f"{crit:.1f} days" if crit is not None else "N/A",
            "MTTR — High":     f"{high:.1f} days" if high is not None else "N/A",
        }

    # ------------------------------------------------------------------
    # validate_config
    # ------------------------------------------------------------------

    def validate_config(self, config: ModuleConfig) -> list[str]:
        """Validate the optional ``min_sample_size`` option."""
        errors: list[str] = []
        val = config.options.get("min_sample_size")
        if val is not None:
            try:
                if int(val) < 1:
                    errors.append(
                        "mttr_by_severity: 'min_sample_size' must be >= 1, "
                        f"got {val}"
                    )
            except (TypeError, ValueError):
                errors.append(
                    f"mttr_by_severity: 'min_sample_size' must be an integer, "
                    f"got {type(val).__name__}"
                )
        return errors

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "MTTR": (
                    "mean(days_to_fix) for remediated findings per severity tier"
                ),
                "days_to_fix (primary)": (
                    "time_taken_to_fix field (seconds) / 86400 — provided directly by "
                    "Tenable in the vulnerability export"
                ),
                "days_to_fix (fallback)": (
                    "(last_fixed - first_found).days — used when time_taken_to_fix is "
                    "absent. Note: this is an approximation as last_fixed reflects the "
                    "last scan date, not the exact remediation date."
                ),
                "SLA comparison": (
                    f"mttr / sla_days. {_STATUS_WITHIN} <= {_WITHIN_SLA_LIMIT}, "
                    f"{_STATUS_NEAR} <= {_NEAR_LIMIT}, {_STATUS_EXCEED} > {_NEAR_LIMIT}"
                ),
                "Data Source": (
                    "tio.exports.vulns() — findings where state == 'fixed' or "
                    "last_fixed is not null"
                ),
                "Limitation": (
                    "MTTR accuracy depends on scan frequency. Vulnerabilities remediated "
                    "between scans may show a longer MTTR than the actual fix time."
                ),
            },
        }
