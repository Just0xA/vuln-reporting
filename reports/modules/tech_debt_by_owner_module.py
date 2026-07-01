"""
reports/modules/tech_debt_by_owner_module.py — Tech Debt by Owner.

Quantifies each asset owner's accumulated backlog of overdue, unfixed
Critical + High findings — a per-owner cut of the same "tech debt" concept
already surfaced on the cover-page KPI strip.

v1 scope (LOCKED per CONTEXT.md 260701-da9)
--------------------------------------------
Tech debt = count of **overdue open Critical + High findings** per owner.
No debt-dimension classifier (missing-patch / legacy / misconfig) — that
fuller framing is explicitly OUT of v1 scope, noted as future work.

Owner derivation
-----------------
``composed_report.py`` does not enrich ``vulns_df`` with tag columns, so this
module derives owner itself: parse the ``Owner`` tag category (configurable
via the ``owner_category`` option) from ``assets_df["tags"]``, build an
``asset_uuid -> owner`` map, and join onto vulns via ``asset_uuid``. Assets
with no Owner tag — or vulns whose ``asset_uuid`` has no matching asset —
bucket under ``"(Unassigned)"``.

Severity source
----------------
VPR-derived severity (``config.vpr_to_severity``) with a native-severity
fallback when VPR is null, per the project's SLA policy.

Overdue definition
-------------------
``today - first_found > SLA_DAYS[severity]`` AND ``state in {open, reopened}``
— mirrors ``utils/sla_calculator.py`` semantics (a finding whose severity has
no SLA_DAYS entry, e.g. info, is never overdue).

RAG thresholds (per owner + cover-page strip)
-----------------------------------------------
Overdue Crit+High count: 0 → green; 1..amber_max (default 4) → yellow;
> amber_max → red. Overridable via ``green_max`` / ``amber_max`` options.
Strip status = worst per-owner status; headline = total overdue Crit+High
across all owners; ``"no_data"`` when zero findings are in scope.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from reports.modules.board_report_utils import extract_owner
from reports.modules import (
    safe_pct, safe_int, safe_format,
    build_rag_strip_entry, NO_DATA_HEADLINE, NO_DATA_DRIVER,
    STATUS_COLOR, STATUS_LABEL,
)
from config import SLA_DAYS, vpr_to_severity

logger = logging.getLogger(__name__)

# ── Open states ───────────────────────────────────────────────────────────────
_OPEN_STATES: frozenset[str] = frozenset({"open", "reopened"})

# ── Severities counted toward tech debt (v1 scope) ────────────────────────────
_DEBT_SEVERITIES: frozenset[str] = frozenset({"critical", "high"})

_UNASSIGNED_LABEL = "(Unassigned)"

# ── RAG threshold defaults (option-overridable) ───────────────────────────────
_DEFAULT_GREEN_MAX = 0
_DEFAULT_AMBER_MAX = 4


# ── Owner derivation helper (importable for tests) ────────────────────────────

def build_owner_map(
    assets_df:        pd.DataFrame,
    owner_category:   str = "Owner",
    unassigned_label: str = _UNASSIGNED_LABEL,
) -> dict[str, str]:
    """
    Build an ``asset_uuid -> owner`` map from ``assets_df["tags"]``.

    When ``owner_category`` is the default ``"Owner"``, delegates to the
    canonical ``board_report_utils.extract_owner()`` parser. For any other
    category, mirrors its semantics with an inline single pass: split on
    ``";"``, partition each token on ``"="``, case-insensitive category
    match, multiple values joined with ``" | "``. Missing/blank Owner tag
    (or a missing ``tags``/``asset_uuid`` column) maps to ``unassigned_label``.

    Fail-soft: never raises. Missing columns return an empty map (or an
    all-unassigned map when only ``tags`` is missing but ``asset_uuid`` is
    present).

    Parameters
    ----------
    assets_df : pd.DataFrame
        Asset DataFrame with ``asset_uuid`` and (ideally) ``tags`` columns.
    owner_category : str
        Tag category name to extract as owner. Default ``"Owner"``.
    unassigned_label : str
        Label used when an asset carries no matching tag value.

    Returns
    -------
    dict[str, str]
        ``asset_uuid -> owner`` mapping. Empty dict if ``asset_uuid`` column
        is absent.
    """
    if assets_df is None or assets_df.empty or "asset_uuid" not in assets_df.columns:
        return {}

    if owner_category == "Owner":
        enriched = extract_owner(assets_df, unassigned_label=unassigned_label)
        return dict(zip(enriched["asset_uuid"], enriched["owner"]))

    # Inline parse for a non-default owner_category
    if "tags" not in assets_df.columns:
        logger.warning(
            "build_owner_map: 'tags' column not present — all assets "
            "mapped to %r.",
            unassigned_label,
        )
        return {uid: unassigned_label for uid in assets_df["asset_uuid"]}

    category_cf = owner_category.strip().casefold()

    def _parse(tags_val: Any) -> str:
        if not isinstance(tags_val, str) or not tags_val.strip():
            return unassigned_label
        values: list[str] = []
        for token in tags_val.split(";"):
            token = token.strip()
            if not token or "=" not in token:
                continue
            cat, _, val = token.partition("=")
            val_s = val.strip()
            if not val_s:
                continue
            if cat.strip().casefold() == category_cf:
                values.append(val_s)
        return " | ".join(sorted(set(values))) if values else unassigned_label

    owners = assets_df["tags"].apply(_parse)
    return dict(zip(assets_df["asset_uuid"], owners))


def _rag_bucket(count: int, green_max: int, amber_max: int) -> str:
    """Classify an owner's overdue Crit+High count into green/yellow/red."""
    if count <= green_max:
        return "green"
    if count <= amber_max:
        return "yellow"
    return "red"


# ── Module ────────────────────────────────────────────────────────────────────

@register_module
class TechDebtByOwnerModule(BaseModule):
    """
    Per-owner backlog of overdue, unfixed Critical + High findings.

    Owner is derived from ``assets_df["tags"]`` (Owner tag category,
    configurable) and joined onto ``vulns_df`` via ``asset_uuid``. Assets
    with no Owner tag, or vulns with no matching asset, bucket under
    ``"(Unassigned)"``.

    Per-owner risk = count of overdue open Critical + High findings
    (VPR-derived severity, native fallback when VPR is null; overdue per
    ``utils/sla_calculator.py`` semantics: ``days_open > SLA_DAYS[severity]``).

    Supported options
    ------------------
    owner_category : str
        Tag category to treat as "owner". Default ``"Owner"``.
    green_max : int
        Overdue Crit+High count <= this value is green. Default ``0``.
    amber_max : int
        Overdue Crit+High count <= this value (and > green_max) is yellow;
        above it is red. Default ``4``.
    """

    MODULE_ID         = "tech_debt_by_owner"
    DISPLAY_NAME      = "Tech Debt by Owner"
    DESCRIPTION       = (
        "Per-owner backlog of overdue, unfixed Critical + High findings — "
        "who is carrying the most accumulated remediation debt."
    )
    REQUIRED_DATA     = ["vulns", "assets"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute per-owner overdue Critical + High counts and RAG buckets.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame. Must have ``state``,
            ``asset_uuid``, ``first_found``, ``vpr_score`` columns
            (``severity`` used as native fallback).
        assets_df : pd.DataFrame
            Tag-filtered asset DataFrame with ``asset_uuid`` and ``tags``.
        report_date : datetime
            Report run timestamp (UTC-aware); used as the overdue reference.
        config : ModuleConfig
            Module configuration; consumes ``owner_category``, ``green_max``,
            ``amber_max`` options.

        Returns
        -------
        ModuleData
            On success: ``error`` is None and all data fields are populated.
            On failure: ``error`` is set and data fields hold safe empty
            defaults.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            options        = config.options or {}
            owner_category = options.get("owner_category", "Owner")
            green_max      = int(options.get("green_max", _DEFAULT_GREEN_MAX))
            amber_max      = int(options.get("amber_max", _DEFAULT_AMBER_MAX))

            # 1. Restrict to open states
            if vulns_df.empty or "state" not in vulns_df.columns:
                open_df = pd.DataFrame(
                    columns=list(vulns_df.columns) if not vulns_df.empty
                    else ["state", "asset_uuid", "first_found", "vpr_score", "severity"]
                )
            else:
                open_df = vulns_df[
                    vulns_df["state"].str.lower().isin(_OPEN_STATES)
                ].copy()

            if len(open_df) == 0:
                kept_df = open_df
            else:
                # 2. Derive severity per row: VPR-primary, native fallback
                if "vpr_score" in open_df.columns:
                    vpr_col = open_df["vpr_score"]
                else:
                    vpr_col = pd.Series([None] * len(open_df), index=open_df.index)
                if "severity" in open_df.columns:
                    native_col = open_df["severity"].fillna("info").str.lower()
                else:
                    native_col = pd.Series(["info"] * len(open_df), index=open_df.index)

                derived_severity = pd.Series(
                    [
                        vpr_to_severity(vpr, fallback=native)
                        for vpr, native in zip(vpr_col, native_col)
                    ],
                    index=open_df.index,
                )

                # 3. Overdue per row (vectorized) — matches sla_calculator semantics
                if "first_found" in open_df.columns:
                    first_found = pd.to_datetime(
                        open_df["first_found"], utc=True, errors="coerce"
                    )
                else:
                    first_found = pd.Series(
                        pd.NaT, index=open_df.index, dtype="datetime64[ns, UTC]"
                    )

                report_ts = (
                    pd.Timestamp(report_date).tz_convert("UTC")
                    if getattr(report_date, "tzinfo", None) is not None
                    else pd.Timestamp(report_date, tz="UTC")
                )
                days_open   = (report_ts - first_found).dt.days
                sla_days    = derived_severity.map(SLA_DAYS)
                is_overdue  = (
                    first_found.notna()
                    & sla_days.notna()
                    & (days_open > sla_days)
                )

                open_df = open_df.assign(
                    _severity=derived_severity,
                    _is_overdue=is_overdue,
                    _days_open=days_open,
                )

                # 4. Keep overdue Crit+High rows
                keep_mask = (
                    open_df["_is_overdue"]
                    & open_df["_severity"].isin(_DEBT_SEVERITIES)
                )
                kept_df = open_df[keep_mask].copy()

            # Owner map + join
            owner_map = build_owner_map(assets_df, owner_category, _UNASSIGNED_LABEL)

            if len(kept_df) > 0:
                if "asset_uuid" in kept_df.columns:
                    owners = kept_df["asset_uuid"].map(owner_map).fillna(_UNASSIGNED_LABEL)
                else:
                    owners = pd.Series([_UNASSIGNED_LABEL] * len(kept_df), index=kept_df.index)
                kept_df = kept_df.assign(_owner=owners)

            # 5. Per-owner aggregate
            owner_rows: list[dict[str, Any]] = []
            if len(kept_df) > 0:
                for owner, grp in kept_df.groupby("_owner"):
                    overdue_critical = int((grp["_severity"] == "critical").sum())
                    overdue_high     = int((grp["_severity"] == "high").sum())
                    total            = overdue_critical + overdue_high
                    owner_rows.append({
                        "owner":            owner,
                        "overdue_critical": overdue_critical,
                        "overdue_high":     overdue_high,
                        "total":            total,
                        "rag_status":       _rag_bucket(total, green_max, amber_max),
                    })
                owner_rows.sort(key=lambda r: r["total"], reverse=True)

            # 6. Strip status + headline
            total_overdue_ch = sum(r["total"] for r in owner_rows)
            if total_overdue_ch == 0 or not owner_rows:
                strip_status = "no_data"
                headline_str = NO_DATA_HEADLINE
            else:
                statuses = {r["rag_status"] for r in owner_rows}
                if "red" in statuses:
                    strip_status = "red"
                elif "yellow" in statuses:
                    strip_status = "yellow"
                else:
                    strip_status = "green"
                headline_str = safe_int(total_overdue_ch)

            rag_strip = build_rag_strip_entry(
                self.DISPLAY_NAME, headline_str, strip_status
            )

            # 7. Driver narrative
            if total_overdue_ch == 0 or not owner_rows:
                driver = NO_DATA_DRIVER
            else:
                top_owner = owner_rows[0]
                driver = (
                    f"{top_owner['owner']} carries the largest overdue "
                    f"Crit+High backlog ({safe_int(top_owner['total'])} findings) "
                    f"of {safe_int(total_overdue_ch)} total across all owners."
                )

            # 8. Analyst rows — per-finding drill-down
            analyst_rows: list[tuple[str, pd.DataFrame]] = []
            if len(kept_df) > 0:
                cols_present = [c for c in ("plugin_id", "plugin_name") if c in kept_df.columns]
                analyst_df = pd.DataFrame({
                    "owner":        kept_df["_owner"],
                    "days_open":    kept_df["_days_open"],
                    "severity":     kept_df["_severity"],
                })
                for col in cols_present:
                    analyst_df[col] = kept_df[col].values
                analyst_rows = [("Tech Debt Detail", analyst_df)]

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            metrics: dict[str, Any] = {
                "total_overdue_ch": total_overdue_ch,
                "owner_rows":       owner_rows,
                "owner_category":   owner_category,
                "green_max":        green_max,
                "amber_max":        amber_max,
            }

            summary_text = (
                NO_DATA_DRIVER if total_overdue_ch == 0 else
                f"{safe_int(total_overdue_ch)} overdue Critical+High findings "
                f"across {len(owner_rows)} owner(s)."
            )

            return ModuleData(
                module_id        = self.MODULE_ID,
                display_name     = self.DISPLAY_NAME,
                metrics          = metrics,
                table_data       = [],
                chart_data       = {},
                summary_text     = summary_text,
                metadata         = {
                    "computed_at":     computed_at,
                    "owner_category":  owner_category,
                    "thresholds":      {"green_max": green_max, "amber_max": amber_max},
                    "filter":          (
                        "state in {open,reopened} AND overdue AND "
                        "severity in {critical,high}"
                    ),
                    "severity_source": "VPR primary, native fallback",
                },
                driver_narrative = driver,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render an Owner | Overdue Critical | Overdue High | Total | RAG table.

        Returns an error callout when ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m          = data.metrics
        owner_rows = m.get("owner_rows", [])

        rows_html = ""
        for row in owner_rows:
            rag = row.get("rag_status", "no_data")
            bg  = STATUS_COLOR.get(rag, STATUS_COLOR["no_data"])
            rows_html += (
                f"<tr>"
                f'<td style="padding:4px 8px;">{row.get("owner", "")}</td>'
                f'<td style="text-align:right;padding:4px 8px;">{safe_int(row.get("overdue_critical"))}</td>'
                f'<td style="text-align:right;padding:4px 8px;">{safe_int(row.get("overdue_high"))}</td>'
                f'<td style="text-align:right;padding:4px 8px;font-weight:bold;">{safe_int(row.get("total"))}</td>'
                f'<td style="background-color:{bg};color:#ffffff;font-weight:bold;'
                f'text-align:center;padding:4px 8px;">{STATUS_LABEL.get(rag, "No Data")}</td>'
                f"</tr>"
            )

        total_overdue_ch = m.get("total_overdue_ch", 0)

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p style="font-size:8pt;margin-bottom:6pt;">{data.driver_narrative}</p>
  <table style="width:100%;border-collapse:collapse;font-size:9pt;">
    <thead>
      <tr style="background-color:#eeeeee;">
        <th style="text-align:left;padding:4px 8px;">Owner</th>
        <th style="text-align:right;padding:4px 8px;">Overdue Critical</th>
        <th style="text-align:right;padding:4px 8px;">Overdue High</th>
        <th style="text-align:right;padding:4px 8px;">Total</th>
        <th style="text-align:center;padding:4px 8px;">RAG</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
    <tfoot>
      <tr style="border-top:2px solid #333;">
        <td style="font-weight:bold;padding:4px 8px;">Total</td>
        <td></td>
        <td></td>
        <td style="text-align:right;font-weight:bold;padding:4px 8px;">{safe_int(total_overdue_ch)}</td>
        <td></td>
      </tr>
    </tfoot>
  </table>
  <p style="font-size:7.5pt;color:#555;margin-top:4pt;">
    Tech debt = overdue open Critical + High findings (VPR-derived severity,
    native fallback when VPR is null). RAG: green=0, yellow=1-{m.get("amber_max", _DEFAULT_AMBER_MAX)},
    red&gt;{m.get("amber_max", _DEFAULT_AMBER_MAX)}.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a single "Tech Debt by Owner" tab.

        Returns ``[]`` on error.
        """
        tab_name = "Tech Debt by Owner"
        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            headers = ["Owner", "Overdue Critical", "Overdue High", "Total", "RAG"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            m          = data.metrics
            owner_rows = m.get("owner_rows", [])

            row_idx = 2
            for row in owner_rows:
                rag = row.get("rag_status", "no_data")
                ws.cell(row=row_idx, column=1, value=row.get("owner", ""))
                ws.cell(row=row_idx, column=2, value=row.get("overdue_critical", 0))
                ws.cell(row=row_idx, column=3, value=row.get("overdue_high", 0))
                ws.cell(row=row_idx, column=4, value=row.get("total", 0))
                rag_cell = ws.cell(row=row_idx, column=5, value=STATUS_LABEL.get(rag, "No Data"))
                hex_color = STATUS_COLOR.get(rag, STATUS_COLOR["no_data"]).lstrip("#")
                rag_cell.fill = PatternFill("solid", fgColor=hex_color)
                rag_cell.font = Font(bold=True, color="FFFFFF")
                row_idx += 1

            # Footer
            total_overdue_ch = m.get("total_overdue_ch", 0)
            footer_cell = ws.cell(row=row_idx, column=1, value="Total")
            footer_cell.font = Font(bold=True)
            ws.cell(row=row_idx, column=4, value=total_overdue_ch).font = Font(bold=True)

            ws.column_dimensions[get_column_letter(1)].width = 22
            ws.column_dimensions[get_column_letter(2)].width = 16
            ws.column_dimensions[get_column_letter(3)].width = 14
            ws.column_dimensions[get_column_letter(4)].width = 10
            ws.column_dimensions[get_column_letter(5)].width = 12

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Inline-CSS email panel for the composed email body.

        Returns ``""`` when ``data.error`` is set.
        """
        if data.error:
            return ""

        m          = data.metrics
        owner_rows = m.get("owner_rows", [])

        rows_html = ""
        for row in owner_rows:
            rag = row.get("rag_status", "no_data")
            bg  = STATUS_COLOR.get(rag, STATUS_COLOR["no_data"])
            rows_html += (
                f'<tr>'
                f'<td style="padding:4px 8px;border:1px solid #ddd;">{row.get("owner", "")}</td>'
                f'<td style="text-align:right;padding:4px 8px;border:1px solid #ddd;">'
                f'{safe_int(row.get("overdue_critical"))}</td>'
                f'<td style="text-align:right;padding:4px 8px;border:1px solid #ddd;">'
                f'{safe_int(row.get("overdue_high"))}</td>'
                f'<td style="text-align:right;padding:4px 8px;font-weight:bold;border:1px solid #ddd;">'
                f'{safe_int(row.get("total"))}</td>'
                f'<td style="background-color:{bg};color:#ffffff;font-weight:bold;'
                f'text-align:center;padding:4px 8px;border:1px solid #ddd;">'
                f'{STATUS_LABEL.get(rag, "No Data")}</td>'
                f'</tr>'
            )

        total_overdue_ch = m.get("total_overdue_ch", 0)
        driver           = data.driver_narrative or ""

        return (
            f'<div style="font-family:Arial,sans-serif;margin-bottom:16px;">'
            f'<h3 style="margin:0 0 6px 0;font-size:13px;color:#333;">'
            f'{self.DISPLAY_NAME}</h3>'
            f'<p style="margin:0 0 8px 0;font-size:11px;color:#555;">{driver}</p>'
            f'<table style="width:100%;border-collapse:collapse;font-size:11px;">'
            f'<thead><tr style="background-color:#eeeeee;">'
            f'<th style="text-align:left;padding:4px 8px;border:1px solid #ddd;">Owner</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">Overdue Critical</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">Overdue High</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">Total</th>'
            f'<th style="text-align:center;padding:4px 8px;border:1px solid #ddd;">RAG</th>'
            f'</tr></thead>'
            f'<tbody>{rows_html}</tbody>'
            f'<tfoot><tr style="border-top:2px solid #333;">'
            f'<td style="font-weight:bold;padding:4px 8px;border:1px solid #ddd;">Total</td>'
            f'<td style="border:1px solid #ddd;"></td>'
            f'<td style="border:1px solid #ddd;"></td>'
            f'<td style="text-align:right;font-weight:bold;padding:4px 8px;border:1px solid #ddd;">'
            f'{safe_int(total_overdue_ch)}</td>'
            f'<td style="border:1px solid #ddd;"></td>'
            f'</tr></tfoot>'
            f'</table>'
            f'</div>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        Return the flat per-finding DataFrame for the analyst workbook.

        Returns ``[]`` on error or zero rows.
        """
        if data.error:
            return []
        return list(data.analyst_rows) if data.analyst_rows else []

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Tech debt definition (v1)": (
                    "Count of overdue open Critical + High findings per owner. "
                    "Overdue = days_open > SLA_DAYS[severity] AND first_found "
                    "not null; open states = {'open', 'reopened'}. No "
                    "debt-dimension classifier (missing-patch/legacy/misconfig) "
                    "— out of v1 scope."
                ),
                "Severity source": (
                    "VPR-derived via config.vpr_to_severity(vpr_score, "
                    "fallback=native_severity) — VPR primary, native Tenable "
                    "severity fallback only when VPR is null."
                ),
                "Owner derivation": (
                    "Parsed from assets_df['tags'] (semicolon-delimited "
                    "'Category=Value' string) for the configured "
                    "owner_category (default 'Owner'); asset_uuid -> owner map "
                    "joined onto vulns. Multiple values on one asset join with "
                    "' | '. Assets with no matching tag, or vulns whose "
                    "asset_uuid has no matching asset, bucket under "
                    "'(Unassigned)'."
                ),
                "RAG thresholds": (
                    "Per-owner: count <= green_max (default 0) -> green; "
                    "<= amber_max (default 4) -> yellow; else red. "
                    "Strip status = worst per-owner status (red > yellow > "
                    "green); 'no_data' when total_overdue_ch == 0."
                ),
                "Open states": (
                    "state in {'open', 'reopened'} — fixed/remediated findings "
                    "excluded upstream."
                ),
            },
        }
