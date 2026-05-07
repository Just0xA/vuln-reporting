"""
reports/modules/composer.py — Report composition utilities.

ReportComposer orchestrates module execution and assembles their
outputs into complete PDF HTML, Excel workbooks, and email KPI dicts.

Typical usage inside a report script
-------------------------------------
::

    from reports.modules import ReportComposer
    from reports.modules.base import ModuleConfig

    composer = ReportComposer(
        vulns_df=vulns_df,
        assets_df=assets_df,
        report_date=generated_at,
        module_configs=[
            ModuleConfig("sla_summary"),
            ModuleConfig("asset_risk", options={"top_n": 25}),
        ],
    )

    results       = composer.run_all()
    pdf_html      = composer.assemble_pdf(results)
    tab_names     = composer.assemble_excel(results, workbook)
    kpis          = composer.collect_email_kpis(results)
    errors        = composer.get_error_summary(results)

Design principles
-----------------
- Module failures are isolated: one module raising or returning an
  error does not prevent other modules from running.
- Order is preserved: modules execute and appear in output in the
  exact order given in ``module_configs``.
- The composer owns no metric logic — it is purely an orchestrator.
- ``assemble_pdf()`` wraps HTML fragments in a minimal but complete
  WeasyPrint-ready document including page CSS and a metadata footer.
- ``assemble_excel()`` appends a ``_Metadata`` tab after all module
  tabs to record run parameters and module audit info.
"""

from __future__ import annotations

import html
import logging
import re
import traceback
from datetime import datetime, timezone
from typing import Any, Optional

import pandas as pd

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import registry

logger = logging.getLogger(__name__)

# Whitelist for ``slug`` arguments accepted by public-API methods that
# interpolate the slug into a filesystem path (WR-04). Letters, digits,
# underscore, and hyphen only — no path separators, no traversal.
_SAFE_SLUG_RE = re.compile(r"^[A-Za-z0-9_\-]+$")

# ---------------------------------------------------------------------------
# PDF document scaffolding
# ---------------------------------------------------------------------------

_PDF_DOCTYPE = "<!DOCTYPE html>"

_PDF_CSS = """
<style>
  /* ── Page setup ─────────────────────────────────────────────────── */
  @page {
    size: A4 landscape;
    margin: 15mm 12mm 18mm 12mm;
    @bottom-center {
      content: "Page " counter(page) " of " counter(pages);
      font-size: 8pt;
      color: #666;
    }
  }

  /* ── Base typography ─────────────────────────────────────────────── */
  body {
    font-family: "Helvetica Neue", Helvetica, Arial, sans-serif;
    font-size: 9pt;
    color: #1a1a1a;
    margin: 0;
    padding: 0;
  }

  /* ── Module sections ─────────────────────────────────────────────── */
  .module-section {
    margin-bottom: 12mm;
  }

  .module-section + .module-section {
    page-break-before: auto;
  }

  .page-break {
    page-break-before: always;
  }

  /* ── Headings ────────────────────────────────────────────────────── */
  .section-heading {
    font-size: 13pt;
    font-weight: bold;
    color: #1F3864;
    border-bottom: 1.5pt solid #1F3864;
    padding-bottom: 2mm;
    margin-top: 0;
    margin-bottom: 4mm;
  }

  .subsection-heading {
    font-size: 10pt;
    font-weight: bold;
    color: #2e4a7a;
    margin-top: 4mm;
    margin-bottom: 2mm;
  }

  /* ── Narrative text ──────────────────────────────────────────────── */
  .explanatory-text {
    font-size: 8.5pt;
    color: #444;
    margin-bottom: 3mm;
    line-height: 1.4;
  }

  /* ── Error callout ───────────────────────────────────────────────── */
  .error-box {
    background: #fff3cd;
    border: 1pt solid #ffc107;
    border-left: 4pt solid #e65100;
    padding: 3mm 4mm;
    margin: 3mm 0;
    font-size: 8.5pt;
    color: #333;
    border-radius: 2pt;
  }

  /* ── Data tables ─────────────────────────────────────────────────── */
  .data-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 8pt;
    margin-bottom: 4mm;
  }

  .data-table th {
    background: #1F3864;
    color: #ffffff;
    padding: 2mm 3mm;
    text-align: left;
    font-weight: bold;
  }

  .data-table td {
    padding: 1.5mm 3mm;
    border-bottom: 0.5pt solid #ddd;
    vertical-align: top;
  }

  .data-table tr:nth-child(even) td {
    background: #f5f7fa;
  }

  /* ── KPI tiles (inline use in modules) ───────────────────────────── */
  .kpi-row {
    display: table;
    width: 100%;
    margin-bottom: 5mm;
    border-spacing: 3mm;
  }

  .kpi-tile {
    display: table-cell;
    background: #f0f4ff;
    border: 0.5pt solid #c5cfe8;
    border-radius: 3pt;
    padding: 3mm 4mm;
    text-align: center;
    vertical-align: middle;
    min-width: 30mm;
  }

  .kpi-value {
    font-size: 16pt;
    font-weight: bold;
    color: #1F3864;
    display: block;
  }

  .kpi-label {
    font-size: 7.5pt;
    color: #555;
    display: block;
    margin-top: 1mm;
  }

  /* ── Report header/footer ────────────────────────────────────────── */
  .report-header {
    border-bottom: 2pt solid #1F3864;
    padding-bottom: 3mm;
    margin-bottom: 6mm;
  }

  .report-title {
    font-size: 16pt;
    font-weight: bold;
    color: #1F3864;
    margin: 0;
  }

  .report-subtitle {
    font-size: 9pt;
    color: #555;
    margin: 1mm 0 0 0;
  }

  .report-footer {
    border-top: 0.5pt solid #ccc;
    padding-top: 2mm;
    margin-top: 8mm;
    font-size: 7.5pt;
    color: #777;
  }

  /* ── Cover / title page (Phase 3 D-01: unified — band + RAG strip) ─ */
  .report-cover {
    page-break-after: always;
    text-align: center;
    padding-top: 18mm;            /* tightened from 48mm — RAG cells now share the page */
  }

  .cover-title {
    font-size: 22pt;
    font-weight: bold;
    color: #1F3864;
    margin: 0 0 5mm 0;
  }

  .cover-subtitle {
    font-size: 11pt;
    color: #555;
    margin: 0 0 10mm 0;
  }

  .cover-divider {
    border: none;
    border-top: 2pt solid #1F3864;
    width: 50%;
    margin: 0 auto 10mm auto;
  }

  .cover-meta {
    font-size: 9pt;
    color: #666;
    line-height: 2.2;
  }

  /* ── Unified-cover RAG strip (Phase 3 D-01) ────────────────────── */
  /* The .rag-strip wrapper is now nested INSIDE .report-cover (which
     already carries page-break-after:always), so the inner element
     no longer needs to force a break of its own — doing so would
     produce an empty extra page after the cover. */
  .rag-strip {
    padding: 4mm 0 0 0;          /* tighter than the prior page-2 standalone */
  }

  .rag-strip-header {
    font-size: 18pt;
    font-weight: bold;
    color: #1F3864;
    text-align: center;
    margin: 0 0 12mm 0;
  }

  .rag-cell-row {
    /* Flexbox: centers the row on the page (justify-content:center),
       gives every cell the same height regardless of label wrapping
       (align-items:stretch), and supports D-03's "1–12 modules,
       wrapping past 4 per row" via flex-wrap. */
    display: flex;
    flex-wrap: wrap;
    justify-content: center;
    align-items: stretch;
    gap: 4mm;
  }

  .rag-cell {
    /* Fixed mm flex-basis (NOT calc(% − mm)) — WeasyPrint's flex
       implementation does not always resolve mixed % / mm calc()
       basis values and falls back to shrink-to-content when it
       can't, leaving cells sized to label width.
       Page geometry (A4 landscape, @page margin 12mm both sides):
         content width = 273mm; 4 cells × 62mm + 3 gaps × 4mm = 260mm;
         13mm total horizontal gutter (6.5mm each side) — visually
         centered. */
    flex: 0 0 62mm;
    width: 62mm;
    min-height: 55mm;
    box-sizing: border-box;
    display: flex;
    flex-direction: column;
    background: #ffffff;
    border: 0.5pt solid #c5cfe8;
    border-radius: 3pt;
    text-align: center;
    padding: 6mm 4mm 0 4mm;
  }

  .rag-cell-label {
    font-size: 9pt;
    color: #555;
    margin: 0 0 2mm 0;
    line-height: 1.2;
    min-height: 22pt;            /* reserve 2-line space — uniform top */
  }

  .rag-cell-value {
    font-size: 22pt;
    font-weight: bold;
    color: #1a1a1a;
    margin: 0 0 4mm 0;
    line-height: 1.0;
    flex: 1 1 auto;              /* fills middle; pushes band to bottom */
    display: flex;
    align-items: center;
    justify-content: center;
  }

  .rag-cell-band {
    margin: 0 -4mm 0 -4mm;       /* extend band to cell edges */
    padding: 3mm 2mm;
    color: #ffffff;
    font-size: 10pt;
    font-weight: bold;
    text-align: center;
    border-radius: 0 0 3pt 3pt;
  }

  .rag-cell-icon {
    display: inline-block;
    margin-right: 2mm;
    font-size: 11pt;
    line-height: 1.0;
  }

  .rag-cell-rag-label {
    display: inline-block;
    vertical-align: middle;
  }
</style>
"""

_PDF_UNIFIED_COVER_TEMPLATE = """
<div class="report-cover">
  <p class="cover-title">{title}</p>
  <p class="cover-subtitle">{subtitle}</p>
  <hr class="cover-divider">
  <div class="cover-meta">
    <p style="margin:0 0 2mm 0;">Generated: {generated_at}</p>
    <p style="margin:0 0 4mm 0;">Sections: {module_list}</p>
  </div>
  <div class="rag-strip">
    <h2 class="rag-strip-header">{header}</h2>
    <div class="rag-cell-row">
{cells_html}
    </div>
  </div>
</div>
"""

# ── W3 Phase 2 symbol-alias safety net (Plan 03-01 Task 2 step 9) ──
# The Phase 2 regression test (tests/test_phase2_composer_pipeline.py)
# was written against the pre-rename names. Keep these aliases so any
# indirect reference (string-based assertion, inspect.getsource scan,
# snapshot probe) continues to resolve. Plan 03-06 owns final removal
# if/when the test is rebaselined to the new names.
_PDF_RAG_STRIP_TEMPLATE = _PDF_UNIFIED_COVER_TEMPLATE


# ===========================================================================
# ReportComposer
# ===========================================================================

class ReportComposer:
    """
    Orchestrates module execution and output assembly for composed reports.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Normalized, tag-filtered vulnerability DataFrame.  Passed
        unchanged to every module's ``compute()`` call.
    assets_df : pd.DataFrame
        Normalized, tag-filtered asset DataFrame.
    report_date : datetime
        UTC-aware datetime for the report run.  Used for age calculations
        and display timestamps.
    module_configs : list[ModuleConfig]
        Ordered list of module configurations.  Modules execute and
        appear in output in this order.
    **kwargs
        Additional context forwarded to every module's ``compute()``
        call (e.g. ``trend_history=df`` for time-series modules).
    """

    def __init__(
        self,
        vulns_df:       pd.DataFrame,
        assets_df:      pd.DataFrame,
        report_date:    Any,
        module_configs: list[ModuleConfig],
        **kwargs:       Any,
    ) -> None:
        self._vulns_df       = vulns_df
        self._assets_df      = assets_df
        self._report_date    = report_date
        self._module_configs = module_configs
        self._kwargs         = kwargs

        # Validate configs up front — log warnings for unknown/misconfigured
        # modules but do not abort; run_all() will surface per-module errors.
        self._warn_invalid_configs()

    # ------------------------------------------------------------------
    # Module execution
    # ------------------------------------------------------------------

    def run_all(self) -> list[ModuleData]:
        """
        Execute ``compute()`` on all configured modules in order.

        A failure in one module (exception or returned error) does not
        stop subsequent modules.  Each failure is logged and represented
        in the results list as a ``ModuleData`` with ``error`` set.

        Returns
        -------
        list[ModuleData]
            One entry per ``ModuleConfig``, in the same order.
            Entries with ``error`` set indicate failed modules.
        """
        results: list[ModuleData] = []

        for config in self._module_configs:
            data = self.run_module(config.module_id, config)
            results.append(data)

        success_count = sum(1 for r in results if r.error is None)
        fail_count    = len(results) - success_count
        logger.info(
            "ReportComposer.run_all: %d/%d modules succeeded.",
            success_count, len(results),
        )
        if fail_count:
            logger.warning(
                "ReportComposer.run_all: %d module(s) failed: %s",
                fail_count,
                [r.module_id for r in results if r.error],
            )

        return results

    def run_module(
        self,
        module_id: str,
        config:    Optional[ModuleConfig] = None,
    ) -> ModuleData:
        """
        Execute a single module by ID and return its ``ModuleData``.

        If ``config`` is ``None``, a default ``ModuleConfig`` is created
        for the module with no options.

        Any unhandled exception from the module's ``compute()`` call is
        caught here so the composer never propagates module failures to
        the caller.

        Parameters
        ----------
        module_id : str
            The ``MODULE_ID`` to look up in the registry.
        config : ModuleConfig, optional
            Configuration for this run.  Defaults to
            ``ModuleConfig(module_id)``.

        Returns
        -------
        ModuleData
            Populated on success; ``error`` field set on any failure.
        """
        if config is None:
            config = ModuleConfig(module_id=module_id)

        # --- Resolve module class from registry ---
        mod_class = registry.get(module_id)
        if mod_class is None:
            err = (
                f"Module '{module_id}' is not registered. "
                f"Registered modules: {sorted(registry._modules.keys())}"
            )
            logger.error("ReportComposer.run_module: %s", err)
            return _error_data(module_id, err)

        # --- Validate config options ---
        try:
            instance      = mod_class()
            config_errors = instance.validate_config(config)
        except Exception as exc:  # noqa: BLE001
            err = f"validate_config() raised: {exc}"
            logger.error(
                "ReportComposer.run_module [%s]: %s\n%s",
                module_id, err, traceback.format_exc(),
            )
            return _error_data(module_id, err)

        if config_errors:
            err = (
                f"Module '{module_id}' config validation failed: "
                + "; ".join(config_errors)
            )
            logger.error("ReportComposer.run_module: %s", err)
            return _error_data(module_id, err)

        # --- Execute compute() ---
        logger.debug(
            "ReportComposer.run_module: calling compute() on '%s'.",
            module_id,
        )
        try:
            data = instance.compute(
                vulns_df    = self._vulns_df,
                assets_df   = self._assets_df,
                report_date = self._report_date,
                config      = config,
                **self._kwargs,
            )
        except Exception as exc:  # noqa: BLE001
            err = (
                f"compute() raised an unhandled exception: "
                f"{type(exc).__name__}: {exc}"
            )
            logger.error(
                "ReportComposer.run_module [%s]: %s\n%s",
                module_id, err, traceback.format_exc(),
            )
            return _error_data(module_id, err)

        # --- Sanity-check return type ---
        if not isinstance(data, ModuleData):
            err = (
                f"compute() returned {type(data).__name__!r} "
                f"instead of ModuleData."
            )
            logger.error("ReportComposer.run_module [%s]: %s", module_id, err)
            return _error_data(module_id, err)

        if data.error:
            logger.warning(
                "ReportComposer.run_module [%s]: module reported error: %s",
                module_id, data.error,
            )
        else:
            logger.debug(
                "ReportComposer.run_module [%s]: compute() succeeded. "
                "metrics=%s",
                module_id, list(data.metrics.keys()),
            )

        return data

    # ------------------------------------------------------------------
    # PDF assembly
    # ------------------------------------------------------------------

    def assemble_pdf(
        self,
        results:  list[ModuleData],
        page_css: str = "",
        title:    str = "Vulnerability Management Report",
        subtitle: str = "",
    ) -> str:
        """
        Assemble module HTML sections into a complete WeasyPrint-ready
        HTML document.

        Each module's ``render_pdf_section()`` is called in results
        order.  Empty strings (modules that don't support PDF output)
        are silently skipped.  A page-break ``<div>`` is inserted
        between sections.

        The document includes:
        - Base stylesheet (``_PDF_CSS``) plus any caller-supplied CSS
        - Report header (title + subtitle)
        - Module sections separated by page-break hints
        - Report footer with generation timestamp and module list

        Parameters
        ----------
        results : list[ModuleData]
            Output of ``run_all()``.
        page_css : str
            Additional CSS to append after the base stylesheet.
            Useful for report-specific overrides.
        title : str
            Report title shown in the header band.
        subtitle : str
            Subtitle line (scope, date range, tag filter, etc.).

        Returns
        -------
        str
            Complete HTML string ready for ``weasyprint.HTML(string=...)``.
        """
        sections: list[str] = []

        for i, data in enumerate(results):
            mod_class = registry.get(data.module_id)
            if mod_class is None:
                logger.warning(
                    "ReportComposer.assemble_pdf: module '%s' not in registry "
                    "— skipping PDF section.",
                    data.module_id,
                )
                continue

            try:
                config   = self._config_for(data.module_id)
                instance = mod_class()
                html_section = instance.render_pdf_section(data, config)
            except Exception as exc:  # noqa: BLE001
                logger.error(
                    "ReportComposer.assemble_pdf [%s]: render_pdf_section() "
                    "raised: %s\n%s",
                    data.module_id, exc, traceback.format_exc(),
                )
                # WR-02: HTML-escape display name + exception message so they
                # cannot break out of the placeholder div even when the
                # exception args carry markup-bearing content.
                safe_name = html.escape(str(data.display_name), quote=True)
                safe_exc  = html.escape(str(exc),               quote=True)
                html_section = (
                    f'<div class="error-box">'
                    f'<strong>{safe_name}</strong>: '
                    f'PDF render failed — {safe_exc}'
                    f'</div>'
                )

            if not html_section or not html_section.strip():
                continue  # Module doesn't support PDF output

            # Insert page break before all sections except the first
            if sections:
                sections.append('<div class="page-break"></div>')
            sections.append(html_section)

        # Build generated_at string
        generated_at_str = (
            self._report_date.strftime("%Y-%m-%d %H:%M UTC")
            if hasattr(self._report_date, "strftime")
            else str(self._report_date)
        )

        # Subtitle fallback (scope only — generated_at appears on the cover separately)
        if not subtitle:
            subtitle = f"Scope: All Assets  |  Generated {generated_at_str}"

        # Human-readable section list for the cover page
        module_list_str = ", ".join(d.display_name for d in results)

        # Phase 3 D-01: page 1 is now the unified cover (title + scope +
        # generated + sections + RAG strip cells). The legacy thin cover
        # template constant has been deleted and the separate page-2 RAG
        # strip has been collapsed into this one page.
        cover = self._build_unified_cover_page(
            results,
            title             = title,
            subtitle          = subtitle,
            generated_at_str  = generated_at_str,
            module_list_str   = module_list_str,
        )

        body = "\n".join(sections) if sections else (
            '<p class="explanatory-text">No module output to display.</p>'
        )

        return "\n".join([
            _PDF_DOCTYPE,
            "<html>",
            "<head>",
            '<meta charset="utf-8">',
            f"<title>{title}</title>",
            _PDF_CSS,
            f"<style>{page_css}</style>" if page_css else "",
            "</head>",
            "<body>",
            cover,
            body,
            "</body>",
            "</html>",
        ])

    # ------------------------------------------------------------------
    # Unified RAG-strip cover (Phase 3 D-01 — supersedes Phase 2 page-2 strip)
    # ------------------------------------------------------------------

    def _build_unified_cover_page(
        self,
        results: list[ModuleData],
        *,
        title:             str,
        subtitle:          str,
        generated_at_str:  str,
        module_list_str:   str,
    ) -> str:
        """
        Build the unified page-1 cover: header band + RAG strip cells.

        Combines the legacy thin cover (title + subtitle + generated +
        sections list) and the Phase 2 page-2 RAG strip into a single
        page-1 cover (D-01). Page 2+ are the per-module sections.

        Iterates results in ``_module_configs`` order, calls each
        module's ``render_rag_strip_entry()`` (Phase 1 contract), and
        renders one cell per module: label-top + headline-value-middle
        + RAG-colored band-bottom containing the status-icon shape AND
        the rag_label text (D-04 / D-08).

        Per-module exception isolation mirrors ``assemble_pdf()``'s
        existing pattern at composer.py:522-533 (Phase 2 D-28). Modules
        whose class is missing from the registry, whose render method
        raises, or whose entry returns an empty dict all collapse to a
        gray "No Data" placeholder cell — never skipped, so the strip
        always shows one cell per configured module (D-06).

        T-03-01 Mitigation: every interpolated string (title, subtitle,
        generated_at_str, module_list_str, label, headline_value,
        rag_label) is HTML-escaped via ``html.escape(..., quote=True)``
        before f-string composition.

        Parameters
        ----------
        results : list[ModuleData]
            Output of ``run_all()``.
        title, subtitle : str, keyword-only
            Cover header band text.
        generated_at_str : str, keyword-only
            Pre-formatted timestamp string (UTC).
        module_list_str : str, keyword-only
            Comma-joined display names for the "Sections:" line.

        Returns
        -------
        str
            Complete ``<div class="report-cover">...</div>`` block
            produced from ``_PDF_UNIFIED_COVER_TEMPLATE``. Always
            non-empty so the page is rendered even when every module
            returned no data.
        """
        from reports.modules.rag_utils import (  # noqa: PLC0415
            STATUS_COLOR, STATUS_LABEL, STATUS_ICON, NO_DATA_HEADLINE,
        )

        def _gray_placeholder(label: str) -> dict:
            return {
                "label":          label,
                "headline_value": NO_DATA_HEADLINE,
                "rag_color":      STATUS_COLOR["no_data"],
                "rag_label":      STATUS_LABEL["no_data"],
            }

        def _icon_for(rag_color_hex: str, module_id: str = "") -> str:
            # Reverse-lookup the status key from the rag_color hex so the
            # icon palette stays loosely coupled to whatever palette the
            # module returned. Fallback to no_data if the color is unknown,
            # but log a warning first so module developers see the contract
            # drift in logs immediately (WR-01).
            for key, hex_str in STATUS_COLOR.items():
                if str(hex_str).lower() == str(rag_color_hex).lower():
                    return STATUS_ICON.get(key, STATUS_ICON["no_data"])
            logger.warning(
                "ReportComposer._build_unified_cover_page [%s]: rag_color %r is not in "
                "STATUS_COLOR — falling back to no_data icon.",
                module_id or "unknown", rag_color_hex,
            )
            return STATUS_ICON["no_data"]

        cells: list[str] = []

        for data in results:
            mod_class = registry.get(data.module_id)
            if mod_class is None:
                logger.warning(
                    "ReportComposer._build_unified_cover_page: module '%s' not "
                    "in registry — emitting gray placeholder cell.",
                    data.module_id,
                )
                cell_dict = _gray_placeholder(data.display_name or data.module_id)
            else:
                try:
                    config   = self._config_for(data.module_id)
                    instance = mod_class()
                    cell_dict = instance.render_rag_strip_entry(data, config)
                except Exception as exc:  # noqa: BLE001
                    logger.error(
                        "ReportComposer._build_unified_cover_page [%s]: "
                        "render_rag_strip_entry() raised: %s\n%s",
                        data.module_id, exc, traceback.format_exc(),
                    )
                    cell_dict = _gray_placeholder(data.display_name or data.module_id)

            # Defensive: if a module returned a malformed dict, fall back to gray.
            if not isinstance(cell_dict, dict) or not all(
                k in cell_dict for k in ("label", "headline_value", "rag_color", "rag_label")
            ):
                logger.warning(
                    "ReportComposer._build_unified_cover_page [%s]: "
                    "render_rag_strip_entry returned malformed dict %r — "
                    "using gray placeholder.",
                    data.module_id, cell_dict,
                )
                cell_dict = _gray_placeholder(data.display_name or data.module_id)

            # WR-01: HTML-escape every module-supplied text field before
            # interpolating into the cover-page HTML. Three concrete failure
            # modes this guards against — display names containing '&', '<',
            # '>'; CSS injection via rag_color; and silent palette drift
            # (handled by the rag_color whitelist below + _icon_for warning).
            label          = html.escape(str(cell_dict["label"]),          quote=False)
            headline_value = html.escape(str(cell_dict["headline_value"]), quote=False)
            rag_label      = html.escape(str(cell_dict["rag_label"]),      quote=False)

            # rag_color is a CSS color — validate strictly against the known
            # palette to defend the style="..." attribute from injection.
            rag_color_raw = str(cell_dict["rag_color"]).strip()
            _palette_lc = {v.lower() for v in STATUS_COLOR.values()}
            if rag_color_raw.lower() not in _palette_lc:
                logger.warning(
                    "ReportComposer._build_unified_cover_page [%s]: rag_color %r is not in "
                    "STATUS_COLOR — substituting no_data gray.",
                    data.module_id, rag_color_raw,
                )
                rag_color_raw = STATUS_COLOR["no_data"]
            rag_color = rag_color_raw
            icon      = _icon_for(rag_color, data.module_id)

            cells.append(
                '    <div class="rag-cell">\n'
                f'      <div class="rag-cell-label">{label}</div>\n'
                f'      <div class="rag-cell-value">{headline_value}</div>\n'
                f'      <div class="rag-cell-band" '
                f'style="background-color: {rag_color};">\n'
                f'        <span class="rag-cell-icon">{icon}</span>'
                f'<span class="rag-cell-rag-label">{rag_label}</span>\n'
                '      </div>\n'
                '    </div>'
            )

        cells_html = "\n".join(cells)

        # T-03-01: HTML-escape every interpolated string for the cover band.
        safe_title         = html.escape(str(title),            quote=True)
        safe_subtitle      = html.escape(str(subtitle),         quote=True)
        safe_generated_at  = html.escape(str(generated_at_str), quote=True)
        safe_module_list   = html.escape(str(module_list_str),  quote=True)

        return _PDF_UNIFIED_COVER_TEMPLATE.format(
            title         = safe_title,
            subtitle      = safe_subtitle,
            generated_at  = safe_generated_at,
            module_list   = safe_module_list,
            header        = "Risk Status Summary",
            cells_html    = cells_html,
        )

    # Class-level alias for the renamed method (W3 safety net).
    # Preserves any test or smoke caller that still references the
    # pre-Phase-3 method name.
    _build_rag_strip_page = _build_unified_cover_page

    # ------------------------------------------------------------------
    # Excel assembly
    # ------------------------------------------------------------------

    def assemble_excel(
        self,
        results:  list[ModuleData],
        workbook: Any,
    ) -> list[str]:
        """
        Call ``render_excel_tabs()`` on all modules and collect tab names.

        Modules that return an empty list (no Excel support) are silently
        skipped.  A ``_Metadata`` tab is appended as the final tab
        containing run parameters and per-module audit info.

        Parameters
        ----------
        results : list[ModuleData]
            Output of ``run_all()``.
        workbook : openpyxl.Workbook
            The workbook to write into.  The caller owns the workbook
            lifecycle — this method does not save or close it.

        Returns
        -------
        list[str]
            All worksheet names added across all modules, including
            ``_Metadata``.
        """
        all_tab_names: list[str] = []

        for data in results:
            mod_class = registry.get(data.module_id)
            if mod_class is None:
                logger.warning(
                    "ReportComposer.assemble_excel: module '%s' not in "
                    "registry — skipping.",
                    data.module_id,
                )
                continue

            try:
                config    = self._config_for(data.module_id)
                instance  = mod_class()
                tab_names = instance.render_excel_tabs(data, workbook, config)
            except Exception as exc:  # noqa: BLE001
                logger.error(
                    "ReportComposer.assemble_excel [%s]: render_excel_tabs() "
                    "raised: %s\n%s",
                    data.module_id, exc, traceback.format_exc(),
                )
                tab_names = _write_error_tab(
                    workbook,
                    tab_name  = f"{data.display_name[:25]} Err",
                    module_id = data.module_id,
                    error     = str(exc),
                )

            all_tab_names.extend(tab_names)

        # Append metadata tab
        meta_tab = _write_metadata_tab(
            workbook     = workbook,
            results      = results,
            report_date  = self._report_date,
            module_configs = self._module_configs,
        )
        all_tab_names.append(meta_tab)

        logger.info(
            "ReportComposer.assemble_excel: wrote %d tab(s): %s",
            len(all_tab_names), all_tab_names,
        )
        return all_tab_names

    # ------------------------------------------------------------------
    # Analyst workbook assembly (Phase 2 D-16..D-21, D-25, D-28, COMPOSER-03)
    # ------------------------------------------------------------------

    def assemble_analyst_workbook(
        self,
        results:     list[ModuleData],
        output_path: Any,
        *,
        slug:        str  = "",
        scope_label: str  = "",
        generate:    bool = True,
    ) -> Optional[Any]:
        """
        Write a separate analyst-detail ``.xlsx`` workbook.

        Iterates ``results`` in ``_module_configs`` order, calls each
        module's ``render_analyst_tabs()`` (Phase 1 contract), and
        writes one tab per ``(sheet_name, DataFrame)`` tuple. Empty
        DataFrames are silently skipped (D-20 contributing rule).
        Sheet-name collisions are auto-suffixed ``_2``/``_3`` with
        Excel's 31-char limit respected (D-18).

        After all module tabs are written, a ``_Metadata`` tab is
        appended carrying Report (slug), Generated (UTC), Scope, and
        Modules per D-19. Per-module failures (raised from
        ``render_analyst_tabs`` per D-28) are recorded in the
        ``_Metadata`` tab's Failures subsection.

        All-empty workbook handling (D-20): if every module returns
        ``[]`` or only empty DataFrames, no file is written and the
        method returns ``None``. The caller's ``run_full_pipeline()``
        sets ``analyst_excel: None`` in its bundle and
        ``email_sender.py`` only attaches non-None paths.

        Phase 4 opt-out hook (D-25): when ``generate=False``, the
        method returns ``None`` immediately without iterating modules
        or importing openpyxl. Phase 4 wires
        ``generate = group_config.get('analyst_detail', True)`` at
        the report-script level.

        Parameters
        ----------
        results : list[ModuleData]
            Output of ``run_all()``.
        output_path : Path
            Destination ``.xlsx`` path. The caller computes the
            filename per D-16 (``{slug}_{date}_analyst.xlsx``);
            this method writes the bytes to that path.
        slug : str, keyword-only
            Report slug for the ``_Metadata`` Report row.
        scope_label : str, keyword-only
            Pre-formatted scope string for the ``_Metadata`` Scope
            row (e.g. ``"Application = UC Engineering"`` or
            ``"All Assets"``). Empty string → Scope row reads
            ``"All Assets"``.
        generate : bool, keyword-only
            Phase 4 opt-out hook (D-25). Always ``True`` in Phase 2.

        Returns
        -------
        pathlib.Path or None
            ``output_path`` on success.
            ``None`` when ``generate=False`` OR when every module
            returned ``[]`` / empty DataFrames (D-20).
        """
        if not generate:
            logger.info(
                "ReportComposer.assemble_analyst_workbook: generate=False — "
                "skipping analyst workbook (D-25 opt-out)."
            )
            return None

        # WR-04: slug is also written to the _Metadata Report row and is
        # public-callable; reject path-traversal-style values up front.
        # Empty slug stays permitted (the default) since this method does
        # not interpolate slug into any filesystem path itself.
        if slug and not _SAFE_SLUG_RE.match(slug):
            raise ValueError(
                f"assemble_analyst_workbook: slug {slug!r} must match "
                f"{_SAFE_SLUG_RE.pattern} (letters, digits, underscore, "
                "hyphen only — no path separators)."
            )

        # Defer openpyxl import to keep module-level imports lean (CONVENTIONS.md)
        import openpyxl  # noqa: PLC0415

        # Collect all (sheet_name, df) tuples first so we can decide whether
        # to write a file at all (D-20 all-empty fallback).
        collected:  list[tuple[str, pd.DataFrame]] = []
        used_names: set[str] = set()
        failures:   list[tuple[str, str]] = []

        for data in results:
            mod_class = registry.get(data.module_id)
            if mod_class is None:
                logger.warning(
                    "ReportComposer.assemble_analyst_workbook: module '%s' "
                    "not in registry — recording failure and skipping.",
                    data.module_id,
                )
                failures.append((data.module_id, "module not registered"))
                continue

            try:
                config   = self._config_for(data.module_id)
                instance = mod_class()
                tabs     = instance.render_analyst_tabs(data, config)
            except Exception as exc:  # noqa: BLE001
                logger.error(
                    "ReportComposer.assemble_analyst_workbook [%s]: "
                    "render_analyst_tabs() raised: %s\n%s",
                    data.module_id, exc, traceback.format_exc(),
                )
                failures.append((data.module_id, f"{type(exc).__name__}: {exc}"))
                continue

            # Defensive: a buggy module returning a non-list collapses to a
            # recorded failure rather than crashing the iteration.
            if not isinstance(tabs, list):
                logger.warning(
                    "ReportComposer.assemble_analyst_workbook [%s]: "
                    "render_analyst_tabs() returned %r (not a list) — "
                    "recording failure and skipping.",
                    data.module_id, type(tabs).__name__,
                )
                failures.append((data.module_id, f"render_analyst_tabs returned non-list ({type(tabs).__name__})"))
                continue

            for entry in tabs:
                # Each entry must be (sheet_name, df). Defensive shape check.
                if (
                    not isinstance(entry, tuple)
                    or len(entry) != 2
                    or not isinstance(entry[0], str)
                ):
                    logger.warning(
                        "ReportComposer.assemble_analyst_workbook [%s]: "
                        "skipping malformed analyst-tab entry %r.",
                        data.module_id, entry,
                    )
                    continue

                sheet_name, df = entry
                if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                    continue   # D-20 contributing rule — empty df contributes no tab

                # WR-03: _unique_sheet_name raises after 98 collision
                # attempts; the prior call site let that propagate out and
                # break D-28 fail-soft. Catch + record into failures so the
                # batch keeps running and the audit lands in _Metadata.
                # WR-05: _unique_sheet_name now mutates ``used_names``
                # itself; no manual add() needed at the call site.
                try:
                    unique = _unique_sheet_name(sheet_name, used_names)
                except ValueError as exc:
                    logger.error(
                        "ReportComposer.assemble_analyst_workbook [%s]: "
                        "could not allocate unique sheet name for %r — "
                        "recording failure.",
                        data.module_id, sheet_name,
                    )
                    failures.append(
                        (data.module_id, f"sheet-name allocation failed: {exc}")
                    )
                    continue
                collected.append((unique, df))

        # D-20: all-empty workbook → no file written
        if not collected:
            logger.info(
                "ReportComposer.assemble_analyst_workbook: every module "
                "returned [] or empty DataFrames — skipping analyst "
                "workbook (failures=%d).",
                len(failures),
            )
            return None

        # ── Open workbook, write tabs, append _Metadata, save ────────────
        wb = openpyxl.Workbook()
        if wb.worksheets:
            wb.remove(wb.worksheets[0])   # mirrors board_summary.py:248-250

        for sheet_name, df in collected:
            ws = wb.create_sheet(sheet_name)
            # Header row
            for col_idx, col in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col_idx, value=str(col))
            # Data rows
            # Gap 03-UAT.md #1 — openpyxl's _bind_value accepts None and
            # np.nan but raises on pd.NA (StringDtype null) and pd.NaT
            # (datetime null). All four Phase 3 board modules coerce text
            # columns through .astype("string") (StringDtype produces
            # pd.NA, not np.nan) and have nullable Int64 / datetime64
            # columns. Coerce every pandas-null sentinel to None at this
            # chokepoint so the analyst workbook renders empty cells
            # rather than crashing the batch.
            #
            # Additionally: openpyxl rejects tz-aware datetimes with
            # ``TypeError: Excel does not support timezones in datetimes``.
            # The Phase 3 modules emit ``last_licensed_scan_date`` and
            # ``last_seen`` as ``datetime64[ns, UTC]`` straight into the
            # analyst_df. Strip tzinfo at this same chokepoint so openpyxl
            # writes a naive datetime (Excel has no concept of timezone
            # so the UTC instant is the only meaningful representation).
            #
            # Locked by check_11 in tests/test_phase2_composer_pipeline.py
            # — committed RED in the prior commit, GREEN at this commit.
            for row_idx, row in enumerate(df.itertuples(index=False), start=2):
                for col_idx, val in enumerate(row, start=1):
                    if pd.isna(val):
                        cell_value = None
                    elif isinstance(val, pd.Timestamp) and val.tzinfo is not None:
                        cell_value = val.tz_convert("UTC").tz_localize(None).to_pydatetime()
                    elif hasattr(val, "tzinfo") and val.tzinfo is not None:
                        cell_value = val.replace(tzinfo=None)
                    else:
                        cell_value = val
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # _Metadata tab last (D-19)
        _write_analyst_metadata_tab(
            wb,
            slug         = slug,
            generated_at = self._report_date,
            scope_label  = scope_label,
            module_ids   = [c.module_id for c in self._module_configs],
            failures     = failures,
        )

        wb.save(str(output_path))
        logger.info(
            "ReportComposer.assemble_analyst_workbook: wrote %d analyst tab(s) "
            "+ _Metadata to %s (failures=%d).",
            len(collected), output_path, len(failures),
        )
        return output_path

    # ------------------------------------------------------------------
    # Email KPI collection
    # ------------------------------------------------------------------

    def collect_email_kpis(
        self,
        results: list[ModuleData],
    ) -> dict[str, str]:
        """
        Collect KPI tiles from all modules for the email body.

        Calls ``render_email_kpis()`` on each module and merges the
        results into a single flat dict.  On key collision, later modules
        overwrite earlier ones — put higher-priority modules last in
        ``module_configs`` if ordering matters.

        Modules that returned errors contribute no KPI tiles.

        Parameters
        ----------
        results : list[ModuleData]

        Returns
        -------
        dict[str, str]
            Merged ``{label: value}`` dict for email KPI tile rendering.
        """
        merged: dict[str, str] = {}

        for data in results:
            mod_class = registry.get(data.module_id)
            if mod_class is None:
                continue

            try:
                config   = self._config_for(data.module_id)
                instance = mod_class()
                kpis     = instance.render_email_kpis(data, config)
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "ReportComposer.collect_email_kpis [%s]: "
                    "render_email_kpis() raised: %s",
                    data.module_id, exc,
                )
                continue

            merged.update(kpis)

        return merged

    # ------------------------------------------------------------------
    # Email body assembly (Phase 2 D-09..D-14, COMPOSER-02)
    # ------------------------------------------------------------------

    def assemble_email_body(
        self,
        results: list[ModuleData],
    ) -> str:
        """
        Assemble per-module email panels into a panels-only HTML fragment.

        Iterates ``results`` in ``_module_configs`` order, calls each
        module's ``render_email_panel()`` (Phase 1 contract), and
        concatenates the non-empty HTML fragments. The returned string
        is a fragment — no ``<html>``, ``<head>``, ``<body>``, scope
        banner, SLA table, or footer (D-09); the wrapping shell stays
        in ``templates/report_email.html``.

        Per-module exception isolation mirrors ``assemble_pdf()``'s
        existing pattern at composer.py:522-533 (D-28). One module's
        render failure inserts a visible error placeholder ``<div>``
        in the panel position but never aborts assembly.

        Modules that return ``""`` or whitespace are silently skipped
        (D-14) — mirrors ``assemble_pdf()``'s skip-empty rule at
        composer.py:535. Un-migrated modules whose ``render_email_panel``
        is the no-op ``""`` default contribute nothing to the body (their
        cover-page strip cell still renders separately as gray "No Data").

        Modules absent from the registry are silently skipped (mirrors
        ``collect_email_kpis()``'s behavior at composer.py:691-692).

        Per D-13 the composer never calls ``draw_gauge()`` itself —
        ``render_email_panel()`` (Phase 3) embeds gauges as
        ``data:image/png;base64,...`` inside the returned HTML fragment.

        Parameters
        ----------
        results : list[ModuleData]
            Output of ``run_all()``.

        Returns
        -------
        str
            Concatenated panel HTML fragments joined by ``"\\n"``.
            Empty string when every panel is empty (the email template's
            ``{% if module_panels_html %}`` conditional then falls through
            to the legacy KPI tiles section).
        """
        panels: list[str] = []

        for data in results:
            mod_class = registry.get(data.module_id)
            if mod_class is None:
                logger.warning(
                    "ReportComposer.assemble_email_body: module '%s' not "
                    "in registry — skipping panel.",
                    data.module_id,
                )
                continue

            try:
                config   = self._config_for(data.module_id)
                instance = mod_class()
                panel_html = instance.render_email_panel(data, config)
            except Exception as exc:  # noqa: BLE001
                logger.error(
                    "ReportComposer.assemble_email_body [%s]: "
                    "render_email_panel() raised: %s\n%s",
                    data.module_id, exc, traceback.format_exc(),
                )
                # WR-02: HTML-escape display name + exception message so they
                # cannot break out of the placeholder div in Outlook / Gmail /
                # Apple Mail when exception args carry markup-bearing content.
                safe_name = html.escape(str(data.display_name), quote=True)
                safe_exc  = html.escape(str(exc),               quote=True)
                panel_html = (
                    '<div style="border:1px solid #d32f2f; '
                    'background:#FFF3CD; color:#5D4037; '
                    'padding:8px 12px; margin:6px 0; '
                    'font-family:Arial,Helvetica,sans-serif; font-size:10pt;">'
                    f'<strong>{safe_name}</strong>: '
                    f'email panel render failed — {safe_exc}'
                    '</div>'
                )

            if not panel_html or not panel_html.strip():
                continue   # D-14 — skip empty / whitespace-only panels

            panels.append(panel_html)

        # T-03-02 contract: assemble_email_body only concatenates
        # pre-escaped HTML fragments produced by per-module
        # render_email_panel implementations (Plans 03-02..05). The
        # per-module HTML escaping is owned there; this method does
        # not re-escape and trusts the fragments to be safe.
        return "\n".join(panels)

    # ------------------------------------------------------------------
    # CID inline-image collection (Phase 3 D-04, COMPOSER-EMAIL)
    # ------------------------------------------------------------------

    def collect_email_inline_images(
        self,
        results: list[ModuleData],
    ) -> list[dict[str, str]]:
        """
        Collect per-module CID inline-image entries for the email panels.

        Each module that overrides ``render_email_panel`` is expected to
        expose its gauge PNG via ``data.metadata["email_gauge_b64"]`` —
        the base64 PNG bytes (no data URI prefix). This method scans
        every module's metadata, builds a list of CID entries, and returns
        them for inclusion in the bundle's ``email_inline_images`` slot.

        The cid string is ``f"{data.module_id}_gauge"``. The module_id is
        constrained to ``[A-Za-z0-9_-]+`` by the registry; cid header
        safety is enforced again at email-attach time (T-03-04).

        Parameters
        ----------
        results : list[ModuleData]

        Returns
        -------
        list[dict]
            One ``{"cid": str, "b64_png": str}`` entry per module that
            produced a non-empty ``email_gauge_b64`` metadata value, in
            registration order. Modules that did not populate this field
            (un-migrated modules, or empty-data modules per D-15) emit
            nothing — the empty-panel placeholder has no gauge image.
        """
        entries: list[dict[str, str]] = []
        for data in results:
            b64 = (data.metadata or {}).get("email_gauge_b64", "")
            if isinstance(b64, str) and b64.strip():
                entries.append({
                    "cid":     f"{data.module_id}_gauge",
                    "b64_png": b64,
                })
        return entries

    # ------------------------------------------------------------------
    # Audit info collection
    # ------------------------------------------------------------------

    def collect_audit_info(
        self,
        results: list[ModuleData],
    ) -> list[dict]:
        """
        Collect audit/calculation metadata from all modules.

        Merges ``get_audit_info()`` from the module class with
        ``ModuleData.metadata`` from the compute run so the audit record
        contains both the static calculation description and the dynamic
        run-time values (row counts, filter parameters, etc.).

        Parameters
        ----------
        results : list[ModuleData]

        Returns
        -------
        list[dict]
            One audit dict per module, in results order.
        """
        audit_records: list[dict] = []

        for data in results:
            mod_class = registry.get(data.module_id)

            static_info: dict = {}
            if mod_class is not None:
                try:
                    static_info = mod_class().get_audit_info()
                except Exception as exc:  # noqa: BLE001
                    logger.warning(
                        "ReportComposer.collect_audit_info [%s]: "
                        "get_audit_info() raised: %s",
                        data.module_id, exc,
                    )

            record = {
                **static_info,
                "run_metadata": data.metadata,
                "run_error":    data.error,
            }
            audit_records.append(record)

        return audit_records

    # ------------------------------------------------------------------
    # Error summary
    # ------------------------------------------------------------------

    def get_error_summary(
        self,
        results: list[ModuleData],
    ) -> list[str]:
        """
        Return error messages from all failed modules.

        Parameters
        ----------
        results : list[ModuleData]

        Returns
        -------
        list[str]
            One entry per failed module: ``"module_id: error message"``.
            Empty list if all modules succeeded.
        """
        return [
            f"{r.module_id}: {r.error}"
            for r in results
            if r.error
        ]

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _config_for(self, module_id: str) -> ModuleConfig:
        """Return the ModuleConfig for ``module_id``, or a default."""
        for cfg in self._module_configs:
            if cfg.module_id == module_id:
                return cfg
        return ModuleConfig(module_id=module_id)

    def _warn_invalid_configs(self) -> None:
        """Log warnings for any module IDs not found in the registry."""
        ids = [c.module_id for c in self._module_configs]
        _, invalid = registry.validate_module_list(ids)
        for mid in invalid:
            logger.warning(
                "ReportComposer: module '%s' is configured but not "
                "registered. It will produce an error result when run.",
                mid,
            )

    # ------------------------------------------------------------------
    # Full pipeline orchestrator (Phase 2 D-22..D-27, COMPOSER-04)
    # ------------------------------------------------------------------

    def run_full_pipeline(
        self,
        results:           list[ModuleData],
        output_dir:        Any,
        *,
        slug:              str,
        report_date:       Any  = None,
        generate_analyst:  bool = True,
        pdf_title:         str  = "Vulnerability Management Report",
        pdf_subtitle:      str  = "",
        scope_label:       str  = "",
    ) -> dict[str, Any]:
        """
        Drive the four render channels and return a typed bundle dict.

        High-level orchestrator that calls the per-channel methods
        internally:

        - ``assemble_pdf(results, ...)``  → ``bundle["pdf_html"]``
          (Plan 02-01 page-2 RAG strip is inserted automatically.)
        - ``assemble_excel(results, wb)`` → ``bundle["excel_workbook"]``
        - ``assemble_analyst_workbook(results, ..., generate=generate_analyst)``
          → ``bundle["analyst_workbook_path"]`` (``None`` when D-20 all-empty
          fallback hits or when ``generate_analyst=False``).
        - ``assemble_email_body(results)`` → ``bundle["email_body_html"]``
        - ``collect_email_inline_images(results)`` → ``bundle["email_inline_images"]``
          (Phase 3 D-04 — base64 gauge PNGs decoded into MIMEImage parts
          by ``delivery/email_sender.py``.)
        - ``collect_email_kpis(results)`` → ``bundle["email_kpis"]``
          (Existing legacy channel kept for un-migrated callers per D-23.)
        - ``{r.module_id: r.metrics for r in results}`` → ``bundle["metrics"]``
        - ``get_error_summary(results)`` → ``bundle["errors"]``

        Per-channel methods stay public as building blocks (D-23).
        Tests, debug scripts, and à-la-carte callers can use them
        directly; ``run_full_pipeline()`` is the convenience layer.

        Module ordering across all channels is driven by the
        ``_module_configs`` list passed to ``__init__`` (D-27).

        Phase 4 opt-out hook (D-25): ``generate_analyst`` is forwarded
        to ``assemble_analyst_workbook(generate=...)``. Phase 2 always
        calls with ``True``; Phase 4 wires
        ``generate_analyst = group_config.get('analyst_detail', True)``
        at the report-script level.

        Parameters
        ----------
        results : list[ModuleData]
            Output of ``run_all()``.
        output_dir : Path
            Directory where the analyst workbook is written. The PDF
            and main Excel are returned in-memory (caller writes them).
        slug : str, keyword-only
            Report slug used for the analyst filename
            (``{slug}_{date}_analyst.xlsx`` per D-16) and the
            ``_Metadata`` Report row (D-19).
        report_date : Any, keyword-only
            Date-like for the analyst filename. Falls back to
            ``self._report_date`` when ``None``.
        generate_analyst : bool, keyword-only
            Phase 4 opt-out hook. Always ``True`` in Phase 2.
        pdf_title : str, keyword-only
            Forwarded to ``assemble_pdf(title=...)``.
        pdf_subtitle : str, keyword-only
            Forwarded to ``assemble_pdf(subtitle=...)``.
        scope_label : str, keyword-only
            Forwarded to ``assemble_analyst_workbook(scope_label=...)``
            for the ``_Metadata`` Scope row (D-19).

        Returns
        -------
        dict[str, Any]
            Bundle dict with exactly these eight keys:
            ``pdf_html`` (str),
            ``excel_workbook`` (openpyxl.Workbook),
            ``analyst_workbook_path`` (Path | None),
            ``email_body_html`` (str),
            ``email_inline_images`` (list[dict]) — D-04 CID gauge entries:
                one ``{"cid": "{module_id}_gauge", "b64_png": "<base64 PNG bytes>"}``
                per migrated module. Decoded into ``MIMEImage`` parts by
                ``delivery/email_sender.py``.
            ``email_kpis`` (dict[str, str]),
            ``metrics`` (dict[str, dict]),
            ``errors`` (list[str]).
        """
        # WR-04: slug is interpolated into ``analyst_filename`` below
        # (``f"{slug}_{date_str}_analyst.xlsx"``). Without validation a caller
        # passing ``slug="../escape"`` would land the workbook outside
        # ``output_dir``. Whitelist letters/digits/underscore/hyphen only.
        if not _SAFE_SLUG_RE.match(slug):
            raise ValueError(
                f"run_full_pipeline: slug {slug!r} must match "
                f"{_SAFE_SLUG_RE.pattern} (letters, digits, underscore, "
                "hyphen only — no path separators)."
            )

        # Defer openpyxl import to keep module-level imports lean (CONVENTIONS.md).
        import openpyxl  # noqa: PLC0415
        from pathlib import Path  # noqa: PLC0415

        bundle: dict[str, Any] = {
            "pdf_html":              "",
            "excel_workbook":        None,
            "analyst_workbook_path": None,
            "email_body_html":       "",
            "email_inline_images":   [],   # Phase 3 D-04 — list[{"cid", "b64_png"}]
            "email_kpis":            {},
            "metrics":               {},
            "errors":                [],
        }

        # ── PDF (page-2 RAG strip is inserted internally by Plan 02-01) ──
        bundle["pdf_html"] = self.assemble_pdf(
            results,
            title    = pdf_title,
            subtitle = pdf_subtitle,
        )

        # ── Main Excel (caller writes the bytes) ─────────────────────────
        wb = openpyxl.Workbook()
        if wb.worksheets:
            wb.remove(wb.worksheets[0])   # mirrors board_summary.py:248-250
        self.assemble_excel(results, wb)
        bundle["excel_workbook"] = wb

        # ── Analyst workbook (separate file, written here) ───────────────
        eff_report_date = report_date if report_date is not None else self._report_date
        date_str = (
            eff_report_date.strftime("%Y-%m-%d")
            if hasattr(eff_report_date, "strftime")
            else str(eff_report_date)
        )
        analyst_filename = f"{slug}_{date_str}_analyst.xlsx"   # D-16
        bundle["analyst_workbook_path"] = self.assemble_analyst_workbook(
            results,
            Path(output_dir) / analyst_filename,
            slug         = slug,
            scope_label  = scope_label,
            generate     = generate_analyst,
        )

        # ── Email body fragment (panels-only per D-09) ───────────────────
        bundle["email_body_html"] = self.assemble_email_body(results)

        # ── Email inline gauge images (Phase 3 D-04) ────────────────────
        bundle["email_inline_images"] = self.collect_email_inline_images(results)

        # ── Email KPIs (existing legacy channel — kept per D-23) ─────────
        bundle["email_kpis"] = self.collect_email_kpis(results)

        # ── Per-module metrics dict ──────────────────────────────────────
        bundle["metrics"] = {r.module_id: r.metrics for r in results}

        # ── Aggregated error list ────────────────────────────────────────
        bundle["errors"] = self.get_error_summary(results)

        return bundle


# ===========================================================================
# Module-level helpers (not part of the public API)
# ===========================================================================

def _error_data(module_id: str, error: str) -> ModuleData:
    """Return a failed ModuleData for a module that could not be run."""
    return ModuleData(
        module_id    = module_id,
        display_name = module_id,
        metrics      = {},
        table_data   = [],
        chart_data   = {},
        summary_text = "",
        metadata     = {},
        error        = error,
    )


def _write_error_tab(
    workbook:  Any,
    tab_name:  str,
    module_id: str,
    error:     str,
) -> list[str]:
    """
    Write a minimal error tab to ``workbook`` when render_excel_tabs()
    raises an unhandled exception.

    Returns the list ``[tab_name]`` so callers can extend their tab list.
    """
    try:
        ws      = workbook.create_sheet(tab_name[:31])   # Excel 31-char limit
        ws["A1"] = "Module"
        ws["B1"] = module_id
        ws["A2"] = "Error"
        ws["B2"] = error
        ws["A3"] = "Action"
        ws["B3"] = "Check application logs for details."
        return [tab_name[:31]]
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "_write_error_tab: could not write error tab for '%s': %s",
            module_id, exc,
        )
        return []


def _write_metadata_tab(
    workbook:       Any,
    results:        list[ModuleData],
    report_date:    Any,
    module_configs: list[ModuleConfig],
) -> str:
    """
    Append a ``_Metadata`` tab to ``workbook`` with run parameters and
    per-module audit info.

    Returns the name of the tab that was written (``"_Metadata"``).
    """
    TAB_NAME = "_Metadata"

    try:
        ws = workbook.create_sheet(TAB_NAME)

        # ── Run summary ──────────────────────────────────────────────
        ws["A1"] = "Report Metadata"
        ws["A2"] = "Generated At"
        ws["B2"] = (
            report_date.strftime("%Y-%m-%d %H:%M UTC")
            if hasattr(report_date, "strftime")
            else str(report_date)
        )
        ws["A3"] = "Modules Run"
        ws["B3"] = len(results)
        ws["A4"] = "Modules Failed"
        ws["B4"] = sum(1 for r in results if r.error)

        # ── Per-module summary ───────────────────────────────────────
        ws["A6"] = "Module ID"
        ws["B6"] = "Display Name"
        ws["C6"] = "Status"
        ws["D6"] = "Error"

        for row_idx, data in enumerate(results, start=7):
            ws.cell(row=row_idx, column=1, value=data.module_id)
            ws.cell(row=row_idx, column=2, value=data.display_name)
            ws.cell(row=row_idx, column=3,
                    value="OK" if data.error is None else "FAILED")
            ws.cell(row=row_idx, column=4, value=data.error or "")

    except Exception as exc:  # noqa: BLE001
        logger.error(
            "_write_metadata_tab: could not write metadata tab: %s", exc
        )

    return TAB_NAME


# ===========================================================================
# Phase 2 analyst workbook helpers (COMPOSER-03)
# ===========================================================================

def _unique_sheet_name(name: str, used: set[str]) -> str:
    """
    Return a unique Excel-31-char-safe sheet name for ``name``.

    If ``name[:31]`` is not in ``used``, return it. Otherwise append
    ``_2``, ``_3``, ... while keeping the total length <= 31 by
    truncating the base name to leave room for the suffix
    (Phase 2 D-18).

    WR-05 fix — this helper now mutates ``used`` itself, adding the
    returned name to the set before returning. Previously the caller
    was responsible for ``used.add(name)``, which was fragile: a future
    contributor could forget the call or interleave it with other
    state changes. The mutation is now atomic with the lookup.

    Collision semantics note (acknowledged for v1):
    When two long sheet names share their first 31 characters but
    differ later, only the auto-suffix counter (``_2``, ``_3``, ...)
    discriminates them — exactly Excel's own behavior. Acceptable for
    v1 since board modules use short, distinct sheet names. v2 may
    revisit if module-supplied sheet names start colliding past index
    30 in practice.

    Parameters
    ----------
    name : str
        Desired sheet name. Will be truncated to fit Excel's 31-char
        worksheet name limit.
    used : set[str]
        Set of sheet names already taken in the target workbook.
        Mutated in-place: the returned name is added before return.

    Returns
    -------
    str
        A unique sheet name <= 31 characters. Already added to ``used``.

    Raises
    ------
    ValueError
        If a unique name cannot be generated within 98 suffix attempts
        (suffixes _2 through _99 inclusive — defensive guard against
        pathological inputs).
    """
    base = name[:31]
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 100):
        suffix    = f"_{i}"
        max_base  = 31 - len(suffix)
        candidate = name[:max_base] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError(
        f"_unique_sheet_name: could not generate unique sheet name "
        f"from {name!r} after 98 attempts (suffixes _2 through _99)."
    )


def _write_analyst_metadata_tab(
    workbook,
    *,
    slug:         str,
    generated_at: Any,
    scope_label:  str,
    module_ids:   list[str],
    failures:     list[tuple[str, str]],
) -> str:
    """
    Append the analyst workbook ``_Metadata`` tab.

    Per Phase 2 D-19, the tab carries exactly four canonical rows
    (Report / Generated / Scope / Modules) in a two-column key/value
    layout that mirrors :func:`_write_metadata_tab`. Per-tab row counts
    and run duration are deliberately excluded — those bleed runtime
    concerns the analyst workbook is not the right surface for.

    When ``failures`` is non-empty, a "Failures" subsection is appended
    listing ``module_id : error_message`` rows (Phase 2 D-28). Failed
    modules are still listed in the canonical Modules row so the
    audit trail is complete.

    Parameters
    ----------
    workbook : openpyxl.Workbook
    slug : str
        Report slug (e.g. ``"board_summary"``).
    generated_at : Any
        Datetime-like with ``.strftime`` or any value that ``str()``
        will format sensibly.
    scope_label : str
        Pre-formatted scope string (e.g. ``"Application = UC Engineering"``
        or ``"All Assets"``). Computed by the caller — composer does
        not assemble scope strings itself.
    module_ids : list[str]
        Module IDs in ``_module_configs`` order. Comma-joined into
        the Modules row.
    failures : list[tuple[str, str]]
        Per-module render failures. Empty list when all modules
        succeeded.

    Returns
    -------
    str
        The tab name written (always ``"_Metadata"``).
    """
    TAB_NAME = "_Metadata"

    try:
        ws = workbook.create_sheet(TAB_NAME)

        gen_str = (
            generated_at.strftime("%Y-%m-%d %H:%M UTC")
            if hasattr(generated_at, "strftime")
            else str(generated_at)
        )

        # ── D-19 canonical rows (key/value, two columns) ─────────────
        ws["A1"] = "Report"
        ws["B1"] = slug
        ws["A2"] = "Generated"
        ws["B2"] = gen_str
        ws["A3"] = "Scope"
        ws["B3"] = scope_label or "All Assets"
        ws["A4"] = "Modules"
        ws["B4"] = ", ".join(module_ids)

        # ── Failures subsection (D-28 audit trail) ───────────────────
        if failures:
            ws["A6"] = "Failures"
            ws["A7"] = "Module ID"
            ws["B7"] = "Error"
            for row_idx, (module_id, err) in enumerate(failures, start=8):
                ws.cell(row=row_idx, column=1, value=module_id)
                ws.cell(row=row_idx, column=2, value=err)

    except Exception as exc:  # noqa: BLE001
        logger.error(
            "_write_analyst_metadata_tab: could not write metadata tab: %s",
            exc,
        )

    return TAB_NAME
