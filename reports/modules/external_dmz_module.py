"""
reports/modules/external_dmz_module.py — External / DMZ Exposure module.

Counts open vulnerability findings (Critical/High/Medium) on externally-scoped
assets (Location=External/DMZ tagged + public-IP-untagged gap assets) using the
Phase-14 ``external_scope()`` substrate from ``utils/external_scope.py``.

This module is **CURRENT-SNAPSHOT ONLY** (Phase-14 D-03).  No MoM trend branch,
no ``trend_snapshots`` kwarg, and no entry in ``_MODULES_NEEDING_TREND_SNAPSHOTS``
(EXT-TREND-01 deferred to v1.5).

Scope classification
--------------------
``external_scope(assets_df)`` is called inline in ``compute()`` — no new kwargs
gate needed.  It returns ``(scoped_df, mismatches_df)``.  When ``scoped_df``
is empty (internal-only group), the module returns a valid gray "no_data" strip
cell and ``error=None`` — this is a normal state, NOT an error.

Mismatch analyst tab (Pitfall 11 / D-11)
-----------------------------------------
``analyst_rows`` contains one tab "External Scope Mismatches" with the LOCKED
schema: ``asset_uuid``, ``ip_address``, ``owner_tag``, ``untagged_reason``,
``finding_count``.  No per-finding plugin/CVE/severity columns are included.
``finding_count`` is an aggregate groupby-size per asset.

RAG thresholds (D-15-07)
------------------------
Green = 0 external Critical findings.
Yellow = 1–5 external Critical findings.
Red   = >5 external Critical findings.
Overridable via ``config.options``:
  ``green_ext_crit_threshold``   — default 0
  ``yellow_ext_crit_threshold``  — default 5
"""

from __future__ import annotations

import logging
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.board_report_utils import extract_owner
from reports.modules.format_utils import safe_format, safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR,
    STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)
from reports.modules.registry import register_module
from utils.external_scope import external_scope

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# RAG thresholds — overridable via module_options (D-15-07)
# ---------------------------------------------------------------------------
_DEFAULT_GREEN_EXT_CRIT  = 0    # 0 Critical external findings = green
_DEFAULT_YELLOW_EXT_CRIT = 5    # 1-5 = yellow; >5 = red

# ---------------------------------------------------------------------------
# Excel fills
# ---------------------------------------------------------------------------
_FILL_HEADER = PatternFill("solid", fgColor="E3F2FD")
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY   = PatternFill("solid", fgColor="F5F5F5")


def _rag_fill(status: str) -> PatternFill:
    if status == "green":
        return _FILL_GREEN
    if status == "yellow":
        return _FILL_AMBER
    if status == "red":
        return _FILL_RED
    return _FILL_GREY


# ===========================================================================
# Module
# ===========================================================================

@register_module
class ExternalDmzModule(BaseModule):
    """
    Open vulnerability counts (Critical/High/Medium) on externally-scoped assets.

    Scope comes from Phase-14 ``external_scope(assets_df)`` called inline in
    ``compute()`` — no new kwargs gate.  Includes a per-owner external table and
    an analyst mismatch tab for public-IP-untagged gap assets.

    CURRENT-SNAPSHOT ONLY (Phase-14 D-03): no trend branch, no trend_snapshots
    kwarg, not in _MODULES_NEEDING_TREND_SNAPSHOTS.

    Supported options
    -----------------
    green_ext_crit_threshold  : int  — default 0  (0 critical = green)
    yellow_ext_crit_threshold : int  — default 5  (1-5 = yellow; >5 = red)
    """

    MODULE_ID         = "external_dmz"
    DISPLAY_NAME      = "External / DMZ Exposure"
    DESCRIPTION       = "Open vulnerability counts on externally-scoped assets (current snapshot)."
    REQUIRED_DATA     = ["vulns", "assets"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute External/DMZ exposure counts and mismatch analyst list.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Expected columns include
            ``asset_uuid`` and ``severity``.
        assets_df : pd.DataFrame
            Asset DataFrame passed to ``external_scope()`` for classification.
            Expected columns: ``asset_uuid``, ``tags``, ``ipv4``.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Options: ``green_ext_crit_threshold`` (int),
            ``yellow_ext_crit_threshold`` (int).
        **kwargs
            Accepted but not used — external_dmz is current-snapshot only.
            Presence of ``trend_snapshots`` in kwargs is silently ignored
            (Phase-14 D-03).
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d, assets_df rows: %d",
            self._log_prefix(),
            len(vulns_df),
            len(assets_df),
        )

        try:
            green_threshold  = int(config.options.get("green_ext_crit_threshold",  _DEFAULT_GREEN_EXT_CRIT))
            yellow_threshold = int(config.options.get("yellow_ext_crit_threshold", _DEFAULT_YELLOW_EXT_CRIT))

            # ----------------------------------------------------------------
            # External scope classification (Phase-14 substrate, inline call)
            # ----------------------------------------------------------------
            scoped_assets_df, mismatches_df = external_scope(assets_df)

            # ----------------------------------------------------------------
            # QUAL-03 / valid state: internal-only group → gray cell, not error
            # ----------------------------------------------------------------
            if scoped_assets_df.empty:
                no_data_rag = build_rag_strip_entry(
                    display_name       = self.DISPLAY_NAME,
                    headline_value_str = NO_DATA_HEADLINE,
                    status             = "no_data",
                )
                return ModuleData(
                    module_id        = self.MODULE_ID,
                    display_name     = self.DISPLAY_NAME,
                    metrics          = {
                        "ext_critical": 0,
                        "ext_high":     0,
                        "ext_medium":   0,
                        "rag_status":   "no_data",
                    },
                    table_data       = [],
                    chart_data       = {},
                    summary_text     = "No external-scope assets in scope.",
                    metadata         = {},
                    driver_narrative = NO_DATA_DRIVER,
                    analyst_rows     = [
                        ("External Scope Mismatches", pd.DataFrame(
                            columns=["asset_uuid", "ip_address", "owner_tag",
                                     "untagged_reason", "finding_count"]
                        )),
                    ],
                    rag_strip        = no_data_rag,
                    error            = None,
                )

            # ----------------------------------------------------------------
            # Scope vulns to external assets
            # ----------------------------------------------------------------
            scoped_uuids = set(scoped_assets_df["asset_uuid"].dropna())

            if not vulns_df.empty and "asset_uuid" in vulns_df.columns:
                ext_vulns_df = vulns_df[vulns_df["asset_uuid"].isin(scoped_uuids)].copy()
            else:
                ext_vulns_df = pd.DataFrame(columns=["asset_uuid", "severity"])

            # ----------------------------------------------------------------
            # Count by severity (CoW-safe — no column assignment on slice)
            # ----------------------------------------------------------------
            def _count_sev(sev: str) -> int:
                if ext_vulns_df.empty or "severity" not in ext_vulns_df.columns:
                    return 0
                return int(
                    (ext_vulns_df["severity"].astype(str).str.lower() == sev).sum()
                )

            ext_critical = _count_sev("critical")
            ext_high     = _count_sev("high")
            ext_medium   = _count_sev("medium")

            # ----------------------------------------------------------------
            # Mismatch analyst tab — finding_count via groupby + map (CoW-safe)
            # Pitfall 11: ONLY locked schema columns — no plugin/CVE/per-sev
            # ----------------------------------------------------------------
            if not ext_vulns_df.empty and "asset_uuid" in ext_vulns_df.columns:
                ext_counts = (
                    ext_vulns_df.groupby("asset_uuid", dropna=False)
                    .size()
                    .rename("finding_count")
                )
            else:
                ext_counts = pd.Series(dtype=int)

            if not mismatches_df.empty:
                mismatches_with_counts = mismatches_df.assign(
                    finding_count=(
                        mismatches_df["asset_uuid"]
                        .map(ext_counts)
                        .fillna(0)
                        .astype(int)
                    )
                )
            else:
                # Zero-row mismatch frame: add finding_count column with locked schema
                mismatches_with_counts = pd.DataFrame(
                    columns=["asset_uuid", "ip_address", "owner_tag",
                             "untagged_reason", "finding_count"]
                )

            # ----------------------------------------------------------------
            # Owner cut via extract_owner (for per-Owner external table)
            # ----------------------------------------------------------------
            enriched_assets = extract_owner(scoped_assets_df)
            uuid_to_owner   = dict(
                zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
            )

            if not ext_vulns_df.empty:
                ext_vulns_with_owner = ext_vulns_df.assign(
                    owner=ext_vulns_df["asset_uuid"]
                    .map(uuid_to_owner)
                    .fillna("Unassigned")
                )
                owner_counts: dict[str, int] = (
                    ext_vulns_with_owner["owner"].value_counts().to_dict()
                )
            else:
                owner_counts = {}

            # ----------------------------------------------------------------
            # RAG strip from external Critical count (lower_is_better)
            # ----------------------------------------------------------------
            status = rag_status_from_value(
                float(ext_critical),
                green_threshold=float(green_threshold),
                yellow_threshold=float(yellow_threshold),
                direction="lower_is_better",
            )

            headline = f"{ext_critical} Critical / {ext_high} High / {ext_medium} Medium"
            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = headline,
                status             = status,
            )

            # ----------------------------------------------------------------
            # Driver narrative (for render_email_panel)
            # ----------------------------------------------------------------
            top_owners = sorted(owner_counts.items(), key=lambda x: x[1], reverse=True)[:3]
            top_owners_str = (
                ", ".join(f"{o} ({c})" for o, c in top_owners) if top_owners else "none"
            )
            driver_narrative = (
                f"{ext_critical} Critical, {ext_high} High, {ext_medium} Medium "
                f"on externally-scoped assets; top owners: {top_owners_str}."
            )

            # ----------------------------------------------------------------
            # Per-owner table rows (for PDF + Excel summary)
            # ----------------------------------------------------------------
            table_data = [
                {"owner": owner, "count": count}
                for owner, count in sorted(
                    owner_counts.items(), key=lambda x: x[1], reverse=True
                )
            ]

            # ----------------------------------------------------------------
            # analyst_rows — "External Scope Mismatches" tab
            # Locked schema: asset_uuid, ip_address, owner_tag,
            #                untagged_reason, finding_count
            # ----------------------------------------------------------------
            analyst_rows: list[tuple[str, pd.DataFrame]] = [
                ("External Scope Mismatches", mismatches_with_counts),
            ]

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "ext_critical": ext_critical,
                    "ext_high":     ext_high,
                    "ext_medium":   ext_medium,
                    "rag_status":   status,
                    "owner_counts": owner_counts,
                },
                table_data   = table_data,
                chart_data   = {
                    "owners": list(owner_counts.keys()),
                    "counts": list(owner_counts.values()),
                },
                summary_text = (
                    f"External/DMZ exposure: {ext_critical} Critical, "
                    f"{ext_high} High, {ext_medium} Medium open findings on "
                    f"{len(scoped_assets_df)} external-scope asset(s)."
                ),
                metadata     = {
                    "scoped_asset_count": len(scoped_assets_df),
                    "mismatch_count":     len(mismatches_with_counts),
                    "green_threshold":    green_threshold,
                    "yellow_threshold":   yellow_threshold,
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render External/DMZ exposure counts and per-owner table as HTML.

        Returns an error callout if ``data.error`` is set.
        Returns a "no external scope" notice if summary indicates no assets.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m           = data.metrics
        ext_crit    = safe_int(m.get("ext_critical"))
        ext_high    = safe_int(m.get("ext_high"))
        ext_medium  = safe_int(m.get("ext_medium"))

        # Per-owner table rows
        table_rows = ""
        for row in data.table_data:
            table_rows += (
                f"<tr><td>{row['owner']}</td>"
                f"<td style='text-align:right;'>{row['count']}</td></tr>"
            )

        owner_table = ""
        if table_rows:
            owner_table = f"""
<table class="data-table" style="width:100%;margin-top:8pt;">
  <thead>
    <tr><th>Owner</th><th style="text-align:right;">Finding Count</th></tr>
  </thead>
  <tbody>
    {table_rows}
  </tbody>
</table>"""

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p>
    <strong>Critical:</strong> {ext_crit} &nbsp;
    <strong>High:</strong> {ext_high} &nbsp;
    <strong>Medium:</strong> {ext_medium}
  </p>
  {owner_table}
  <p class="explanatory-text">
    External/DMZ exposure counts open findings on assets classified as
    Location=External or Location=DMZ, plus public-IP-untagged gap assets.
    Current snapshot only — MoM trend deferred to v1.5.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write an "External DMZ" tab with severity counts and owner breakdown.

        Returns ``[]`` on exception; writes an error row if ``data.error`` is set.
        """
        tab_name = "External DMZ"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            m = data.metrics

            # Summary block
            ws["A1"] = "External / DMZ Exposure"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A2"] = "Critical"
            ws["B2"] = m.get("ext_critical", 0)
            ws["A3"] = "High"
            ws["B3"] = m.get("ext_high", 0)
            ws["A4"] = "Medium"
            ws["B4"] = m.get("ext_medium", 0)
            ws["A5"] = "RAG Status"
            ws["B5"] = STATUS_LABEL.get(m.get("rag_status", "no_data"), "No Data")

            # Per-owner table starting at row 7
            if data.table_data:
                headers = ["Owner", "Finding Count"]
                for col_idx, header in enumerate(headers, start=1):
                    cell      = ws.cell(row=7, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = _FILL_HEADER

                for row_idx, row in enumerate(data.table_data, start=8):
                    ws.cell(row=row_idx, column=1, value=row["owner"])
                    ws.cell(row=row_idx, column=2, value=row["count"])

            # Column widths
            ws.column_dimensions[get_column_letter(1)].width = 28
            ws.column_dimensions[get_column_letter(2)].width = 18

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01 — NOT render_email_kpis)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Returns ``""`` when ``data.error`` is set or ``driver_narrative`` is empty.
        Uses only ``safe_int`` — never raw f-string on possibly-None values.
        """
        if data.error or not data.driver_narrative:
            return ""

        m          = data.metrics
        ext_crit   = safe_int(m.get("ext_critical"))
        ext_high   = safe_int(m.get("ext_high"))
        ext_medium = safe_int(m.get("ext_medium"))

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#FFF3E0;border-left:4px solid #E65100;">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;">Critical: {ext_crit} &nbsp; High: {ext_high} &nbsp; Medium: {ext_medium}</span><br>
      <em style="font-size:11px;color:#555;">{data.driver_narrative}</em>
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns ``data.analyst_rows`` when valid; ``[]`` on error or no rows.
        "External Scope Mismatches" tab: ONLY locked columns
        asset_uuid, ip_address, owner_tag, untagged_reason, finding_count
        (QUAL-05 — no plugin/CVE/per-severity; Pitfall 11).
        """
        if data.error or not data.analyst_rows:
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03, honors data.rag_strip)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present;
        falls back to a gray "No Data" cell if empty or error.
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip
