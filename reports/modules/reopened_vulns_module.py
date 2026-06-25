"""
reports/modules/reopened_vulns_module.py — Reopened Vulnerabilities module.

Counts findings currently in ``state == REOPENED`` (re-emerged after a fix),
computes a reopen rate when the fixed-findings export is available, segments
by Owner tag, and provides an analyst drill-down with
``plugin_id``, ``resurfaced_date``, ``reopen_lag_days``, and ``owner``.

This is the PATHFINDER module for Phase 15 (D-15-09): it proves the full
four-channel render contract and empty-data guard before the other four
Phase-15 modules copy the shape.

Primary filter
--------------
``vulns_df["state"].astype(str).str.upper() == "REOPENED"``
(current snapshot — NOT via ``open_findings_at()``)

Rate denominator
----------------
``reopened_count / (reopened_count + fixed_count)``

Requires the ``fixed_vulns_df`` kwarg when available; degrades gracefully
to count-only when absent, with an explicit disclosure note.

Owner cut
---------
``extract_owner(assets_df)`` joined on ``asset_uuid``; rows with no matching
asset → ``"Unassigned"``.

Supported options
-----------------
green_rate_threshold  : float  — default 5.0  (reopen rate % for green)
yellow_rate_threshold : float  — default 10.0 (reopen rate % for yellow)
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.board_report_utils import extract_owner
from reports.modules.format_utils import safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_HEADLINE,
    build_rag_strip_entry,
    rag_status_from_value,
)
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# RAG thresholds — overridable via module_options (D-15-07)
# ---------------------------------------------------------------------------
_DEFAULT_GREEN_RATE  = 5.0    # reopen rate % below which is green
_DEFAULT_YELLOW_RATE = 10.0   # reopen rate % below which is yellow

# ---------------------------------------------------------------------------
# Excel fills
# ---------------------------------------------------------------------------
_FILL_HEADER = PatternFill("solid", fgColor="E3F2FD")
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY   = PatternFill("solid", fgColor="F5F5F5")


# ===========================================================================
# Module
# ===========================================================================

@register_module
class ReopenedVulnsModule(BaseModule):
    """
    Count and rate of findings in state==REOPENED (resurfaced after a fix).

    Primary filter: vulns_df["state"].str.upper() == "REOPENED"
    Rate denominator: reopened_count / (reopened_count + fixed_count)
      - Requires fixed_vulns_df kwarg when available; degrades to count-only if absent.
    Owner cut: extract_owner(assets_df) joined on asset_uuid.
    Analyst drill-down: plugin_id, resurfaced_date, reopen_lag_days, owner per finding.

    Supported options
    -----------------
    green_rate_threshold  : float  — default 5.0  (reopen rate % for green)
    yellow_rate_threshold : float  — default 10.0 (reopen rate % for yellow)
    """

    MODULE_ID         = "reopened_vulns"
    DISPLAY_NAME      = "Reopened Vulnerabilities"
    DESCRIPTION       = "Count and rate of findings that re-emerged after being marked fixed."
    REQUIRED_DATA     = ["vulns", "assets"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute reopened-vulnerability count, rate, and Owner cut.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Must include ``state``,
            ``resurfaced_date``, ``last_fixed``, ``plugin_id``, ``asset_uuid``,
            and ``severity`` columns.
        assets_df : pd.DataFrame
            Asset DataFrame passed to ``extract_owner()`` for the Owner cut.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Options: ``green_rate_threshold`` (float), ``yellow_rate_threshold`` (float).
        **kwargs
            ``fixed_vulns_df`` (pd.DataFrame, optional) — used as the rate
            denominator.  When absent the rate is ``None`` and a disclosure note
            is set in ``metadata["rate_disclosure"]``.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            green_threshold  = float(config.options.get("green_rate_threshold",  _DEFAULT_GREEN_RATE))
            yellow_threshold = float(config.options.get("yellow_rate_threshold", _DEFAULT_YELLOW_RATE))

            # QUAL-03 — empty-data guard
            if vulns_df.empty or "state" not in vulns_df.columns:
                return self._empty_result("No vulnerability data in scope.", config)

            # ----------------------------------------------------------------
            # Filter to REOPENED (current snapshot — NOT open_findings_at)
            # ----------------------------------------------------------------
            reopened_df = vulns_df[
                vulns_df["state"].astype(str).str.upper() == "REOPENED"
            ]

            reopened_count = len(reopened_df)

            # ----------------------------------------------------------------
            # Reopen-lag computation via .assign() (QUAL-03 / CoW)
            # ----------------------------------------------------------------
            if not reopened_df.empty:
                reopened_df = reopened_df.assign(
                    resurfaced_ts=pd.to_datetime(
                        reopened_df["resurfaced_date"], utc=True, errors="coerce"
                    ),
                    last_fixed_ts=pd.to_datetime(
                        reopened_df["last_fixed"], utc=True, errors="coerce"
                    ),
                )
                reopened_df = reopened_df.assign(
                    reopen_lag_days=(
                        reopened_df["resurfaced_ts"] - reopened_df["last_fixed_ts"]
                    ).dt.days
                )
            else:
                # Ensure columns exist even on zero-row slice
                reopened_df = reopened_df.assign(
                    resurfaced_ts=pd.Series(dtype="datetime64[ns, UTC]"),
                    last_fixed_ts=pd.Series(dtype="datetime64[ns, UTC]"),
                    reopen_lag_days=pd.Series(dtype="float64"),
                )

            # ----------------------------------------------------------------
            # Owner cut via extract_owner
            # ----------------------------------------------------------------
            enriched_assets = extract_owner(assets_df)
            uuid_to_owner   = dict(
                zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
            )

            if not reopened_df.empty:
                reopened_df = reopened_df.assign(
                    owner=reopened_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                )
            else:
                reopened_df = reopened_df.assign(
                    owner=pd.Series(dtype="object")
                )

            # ----------------------------------------------------------------
            # Owner count breakdown
            # ----------------------------------------------------------------
            owner_counts: dict[str, int] = (
                reopened_df["owner"].value_counts().to_dict()
                if not reopened_df.empty
                else {}
            )

            # ----------------------------------------------------------------
            # Reopen rate — optional (requires fixed_vulns_df kwarg)
            # ----------------------------------------------------------------
            fixed_vulns_df = kwargs.get("fixed_vulns_df")
            has_rate       = False
            reopen_rate: Optional[float] = None
            rate_disclosure = ""

            if fixed_vulns_df is not None and not fixed_vulns_df.empty:
                # Denominator: findings that were fixed and stayed fixed
                # (not part of the current REOPENED population)
                fixed_count = len(fixed_vulns_df)
                denom       = reopened_count + fixed_count
                if denom > 0:
                    reopen_rate = round(reopened_count / denom * 100.0, 2)
                    has_rate    = True
            else:
                rate_disclosure = "Rate omitted — fixed export unavailable."
                logger.debug(
                    "%s fixed_vulns_df absent — reopen rate not computed.",
                    self._log_prefix(),
                )

            # ----------------------------------------------------------------
            # RAG strip
            # ----------------------------------------------------------------
            rag_value = reopen_rate  # None when rate unavailable → no_data
            status    = rag_status_from_value(
                rag_value,
                green_threshold=green_threshold,
                yellow_threshold=yellow_threshold,
                direction="lower_is_better",
            )
            headline = (
                f"{reopened_count} ({safe_pct(reopen_rate)})"
                if has_rate
                else str(reopened_count)
            )
            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = headline,
                status             = status,
            )

            # ----------------------------------------------------------------
            # Driver narrative (for render_email_panel)
            # ----------------------------------------------------------------
            top_owners = sorted(owner_counts.items(), key=lambda x: x[1], reverse=True)[:3]
            top_owners_str = ", ".join(f"{o} ({c})" for o, c in top_owners) if top_owners else "none"
            if has_rate:
                driver_narrative = (
                    f"{safe_pct(reopen_rate)} reopen rate; "
                    f"top owners: {top_owners_str}."
                )
            else:
                driver_narrative = (
                    f"{reopened_count} reopened finding(s); "
                    f"top owners: {top_owners_str}. "
                    f"({rate_disclosure})"
                )

            # ----------------------------------------------------------------
            # Analyst rows — QUAL-05: only plugin_id, resurfaced_date,
            # reopen_lag_days, owner — no hostnames/IPs/asset_uuid
            # ----------------------------------------------------------------
            if not reopened_df.empty:
                detail_cols = [c for c in ["plugin_id", "resurfaced_date", "reopen_lag_days", "owner"]
                               if c in reopened_df.columns]
                detail_df = reopened_df[detail_cols].copy().reset_index(drop=True)

                owner_df = (
                    reopened_df.groupby("owner", dropna=False)
                    .size()
                    .rename("reopened_count")
                    .reset_index()
                    .sort_values("reopened_count", ascending=False)
                    .reset_index(drop=True)
                )
            else:
                detail_df = pd.DataFrame(
                    columns=["plugin_id", "resurfaced_date", "reopen_lag_days", "owner"]
                )
                owner_df  = pd.DataFrame(columns=["owner", "reopened_count"])

            analyst_rows: list[tuple[str, pd.DataFrame]] = [
                ("Reopened Detail", detail_df),
                ("By Owner", owner_df),
            ]

            # ----------------------------------------------------------------
            # table_data (per-owner rows for PDF table)
            # ----------------------------------------------------------------
            table_data = [
                {"owner": owner, "count": count}
                for owner, count in sorted(
                    owner_counts.items(), key=lambda x: x[1], reverse=True
                )
            ]

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "reopened_count": reopened_count,
                    "reopen_rate":    reopen_rate,
                    "has_rate":       has_rate,
                    "owner_counts":   owner_counts,
                    "rag_status":     status,
                },
                table_data   = table_data,
                chart_data   = {
                    "owners":  list(owner_counts.keys()),
                    "counts":  list(owner_counts.values()),
                },
                summary_text = (
                    f"{reopened_count} finding(s) currently in REOPENED state. "
                    + (
                        f"Reopen rate: {safe_pct(reopen_rate)} of fixed+reopened population."
                        if has_rate
                        else rate_disclosure
                    )
                ),
                metadata     = {
                    "rate_available":   has_rate,
                    "rate_disclosure":  rate_disclosure,
                    "green_threshold":  green_threshold,
                    "yellow_threshold": yellow_threshold,
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a summary section with reopened count, rate, and per-owner table.

        Returns an error callout if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m           = data.metrics
        count_str   = safe_int(m.get("reopened_count"))
        rate_str    = safe_pct(m.get("reopen_rate"))
        has_rate    = m.get("has_rate", False)
        disclosure  = data.metadata.get("rate_disclosure", "")

        rate_html = (
            f"<p><strong>Reopen Rate:</strong> {rate_str} of fixed+reopened population</p>"
            if has_rate
            else f'<p style="color:#9E9E9E;font-size:9pt;">{disclosure}</p>'
        )

        # Per-owner table rows
        table_rows = ""
        for row in data.table_data:
            table_rows += (
                f"<tr><td>{row['owner']}</td>"
                f"<td style='text-align:right;'>{row['count']}</td></tr>"
            )

        owner_table = ""
        if table_rows:
            owner_table = f"""
<table class="data-table" style="width:100%;margin-top:8pt;">
  <thead>
    <tr><th>Owner</th><th style="text-align:right;">Reopened Count</th></tr>
  </thead>
  <tbody>
    {table_rows}
  </tbody>
</table>"""

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p><strong>Reopened Findings:</strong> {count_str}</p>
  {rate_html}
  {owner_table}
  <p class="explanatory-text">
    A reopened vulnerability is one that was previously marked as fixed and has
    since re-emerged in a scan. A high reopen rate may indicate incomplete
    remediation or patch rollback. Owner breakdown shows which teams have the
    most recurring findings.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a ``Reopened Vulns`` tab with owner breakdown.

        Returns ``[]`` on exception; writes an error row if ``data.error`` is set.
        """
        tab_name = "Reopened Vulns"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            m          = data.metrics
            count      = m.get("reopened_count", 0)
            rate       = m.get("reopen_rate")
            has_rate   = m.get("has_rate", False)
            disclosure = data.metadata.get("rate_disclosure", "")

            # Summary block
            ws["A1"] = "Reopened Vulnerabilities"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A2"] = "Reopened Count"
            ws["B2"] = count
            ws["A3"] = "Reopen Rate"
            ws["B3"] = f"{safe_pct(rate)}" if has_rate else f"N/A — {disclosure}"
            ws["A3"].font = Font(italic=True)

            # Per-owner table starting at row 5
            headers = ["Owner", "Reopened Count"]
            for col_idx, header in enumerate(headers, start=1):
                cell      = ws.cell(row=5, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _FILL_HEADER

            for row_idx, row in enumerate(data.table_data, start=6):
                ws.cell(row=row_idx, column=1, value=row["owner"])
                ws.cell(row=row_idx, column=2, value=row["count"])

            # Column widths
            ws.column_dimensions[get_column_letter(1)].width = 28
            ws.column_dimensions[get_column_letter(2)].width = 18

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01 — NOT render_email_kpis)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Returns ``""`` when ``data.error`` is set or ``driver_narrative`` is empty.
        Uses only ``safe_pct`` / ``safe_int`` — never raw f-string on possibly-None values.
        """
        if data.error or not data.driver_narrative:
            return ""

        m          = data.metrics
        count_str  = safe_int(m.get("reopened_count"))
        rate_str   = safe_pct(m.get("reopen_rate"))
        has_rate   = m.get("has_rate", False)
        disclosure = data.metadata.get("rate_disclosure", "")

        rate_line = (
            f"{count_str} reopened findings ({rate_str} of fixed+reopened)"
            if has_rate
            else f"{count_str} reopened findings (rate unavailable — {disclosure})"
        )

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#F3E5F5;border-left:4px solid #7B1FA2;">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;">{rate_line}</span><br>
      <em style="font-size:11px;color:#555;">{data.driver_narrative}</em>
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns ``data.analyst_rows`` when data is valid; ``[]`` on error or no rows.
        Columns: ``plugin_id``, ``resurfaced_date``, ``reopen_lag_days``, ``owner``
        (QUAL-05 — no hostnames/IPs/asset_uuid).
        """
        if data.error or not data.analyst_rows:
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03, honors data.rag_strip)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present;
        falls back to a gray "No Data" cell if empty or error.
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip
