"""
reports/modules/program_health_module.py — Program Health Overview module.

Computes 4 velocity signals (Open-Critical MoM, Net Velocity, SLA Posture,
MTTR) from the S1 severity-dimension trend snapshot substrate, derives the
OD-5 composite RAG, and provides an Owner velocity table.

Signal re-derivation design (D-17-02 — definition parity)
----------------------------------------------------------
All four signals read pre-computed snapshot fields so they NEVER disagree
with new_vs_remediated or mttr_trend:

  Signal 1 — Open-Critical MoM  : snap.get("critical")
  Signal 2 — Net Velocity       : snap.get("new_findings_count") - snap.get("fixed_findings_count")
  Signal 3 — SLA Posture        : snap.get("sla_rate_crit_high")         [Plan 17-01 field]
  Signal 4 — MTTR Overall       : snap.get("mttr_overall_days")

Net-velocity direction (RESEARCH Open Question 1): "delta of deltas" —
  curr_net_delta < prev_net_delta means the backlog is shrinking faster →
  improving. Binary; no flat band.

Cold-start MTTR tile (RESEARCH Open Question 2): always "—" because no
  fixed_vulns_df is threaded to this module (D-17-01: not in
  _MODULES_NEEDING_FIXED_VULNS). The snapshot field is used for MoM; the
  current-value tile requires fixed_vulns_df which is absent.

Render channels (Plan 17-03)
-----------------------------
Four render channels implemented: PDF sparkline row + Owner velocity table
(D-17-09), email 4-tile KPI panel + narrative (D-17-06), Excel "Program
Health" + "Owner Velocity" tabs, analyst "PH — Owner Detail" tab (QUAL-05),
and the RAG strip entry (CONTRACT-03).

_render_sparkline_b64 helper: matplotlib Agg, figsize=(2.0,1.2) @120dpi,
plt.close(fig) mandatory (T-17-08 figure-leak mitigation).

OD-5 composite RAG rule
------------------------
green  : all 4 signals green
amber  : 2–3 signals green
red    : 0–1 signals green
cap    : any missing signal → composite capped at amber (D-17-06);
         the cap is structural in _composite_rag_od5 and cannot be
         bypassed via module_options.

Threat mitigations
------------------
T-17-04 : validate_config() coerces module_options numerics; bad values
          log WARNING and fall back to default (Security V5).
T-17-05 : analyst_rows carry only owner-name + aggregate counts + MoM
          delta — no UUIDs, IPs, hostnames, or plugin IDs (QUAL-05).
T-17-06 : snap.get() throughout (never []); whole compute wrapped in
          try/except → _empty_result (fail-soft batch).
T-17-07 : owner/tag strings escaped with html.escape() before interpolation
          into HTML/PDF markup — never raw f-string interpolated.
T-17-08 : _render_sparkline_b64 calls plt.close(fig) per figure —
          prevents matplotlib figure accumulation across many groups.
"""

from __future__ import annotations

import base64
import html
import io
import logging
from typing import Any, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SLA_DAYS
from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from utils.sla_calculator import compute_sla_rate_crit_high
from reports.modules.board_report_utils import extract_owner
from reports.modules.format_utils import safe_format, safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR,
    STATUS_LABEL,
    build_rag_strip_entry,
)
from reports.modules.registry import register_module
from utils.open_count import open_findings_at

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Module-level pure helper functions
# ---------------------------------------------------------------------------

def _composite_rag_od5(
    signal_statuses: list[str],
    green_min: int = 4,
    amber_min: int = 2,
) -> tuple[str, bool]:
    """
    Compute the OD-5 composite RAG status from a list of per-signal statuses.

    Parameters
    ----------
    signal_statuses : list[str]
        Each element is one of: "green", "amber", "red", "missing".
    green_min : int
        Minimum green-signal count for a Green composite. Default 4.
    amber_min : int
        Minimum green-signal count for an Amber composite. Default 2.

    Returns
    -------
    tuple[str, bool]
        (composite_status, data_incomplete) where composite_status is
        "green" / "amber" / "red" and data_incomplete is True when any
        signal is "missing".

    Notes
    -----
    D-17-06 structural cap: if any signal is "missing" AND the raw
    composite would be "green", it is capped to "amber" (data_incomplete=True).
    This cap is structural — it cannot be bypassed via module_options.
    """
    green_count = sum(1 for s in signal_statuses if s == "green")
    has_missing = any(s == "missing" for s in signal_statuses)

    # Raw composite by green-count thresholds (OD-5)
    if green_count >= green_min:
        raw = "green"
    elif green_count >= amber_min:
        raw = "amber"
    else:
        raw = "red"

    # D-17-06 structural cap: missing signal prevents Green
    if has_missing and raw == "green":
        return ("amber", True)

    return (raw, has_missing)


def _signal_direction(
    curr: Optional[float],
    prev: Optional[float],
    higher_is_better: bool,
    flat_band: float = 0.0,
) -> str:
    """
    Classify MoM signal direction as "green" / "amber" / "red" / "missing".

    Parameters
    ----------
    curr : float or None
        Current-period value.
    prev : float or None
        Previous-period value.
    higher_is_better : bool
        True for SLA Posture (higher rate = better);
        False for Open-Critical, Net Velocity, MTTR (lower = better).
    flat_band : float
        Absolute tolerance for "flat" (amber). Default 0.0.
        A delta within [-flat_band, +flat_band] inclusive is "flat".

    Returns
    -------
    str
        "missing" when either value is None.
        "amber"   when abs(delta) <= flat_band.
        "green"   when direction improved.
        "red"     when direction worsened.
    """
    if curr is None or prev is None or pd.isna(curr) or pd.isna(prev):
        return "missing"
    delta = curr - prev
    if abs(delta) <= flat_band:
        return "amber"
    improved = (delta > 0) if higher_is_better else (delta < 0)
    return "green" if improved else "red"


# ---------------------------------------------------------------------------
# WR-05: Metadata keys written by capture_snapshot() into owner-dimension
# snapshots.  Kept as a module-level constant so the prev-snap reader and
# any future writers reference the same set and cannot silently drift.
# Any snapshot key NOT in this set and typed int is treated as an owner count.
# ---------------------------------------------------------------------------
_OWNER_SNAPSHOT_METADATA_KEYS: frozenset[str] = frozenset({
    "month",
    "tag_filter",
    "asset_count",
    "on_time_asset_count",
    "reopened_count",
    "accepted_count",
    "recast_count",
    "new_findings_count",
    "fixed_findings_count",
    "mttr_overall_days",
    "mttr_by_severity",
    "mttr_by_owner",
    "sla_rate_crit_high",
    "generated_at",
})


# ---------------------------------------------------------------------------
# Module
# ---------------------------------------------------------------------------

@register_module
class ProgramHealthModule(BaseModule):
    """
    Program Health Overview — 4 velocity signals + composite RAG.

    Reads the S1 severity-dimension trend snapshot substrate to derive:
      1. Open-Critical MoM direction
      2. Net Velocity direction (inflow - outflow MoM delta)
      3. SLA Posture direction (% Crit+High within SLA, reopened-aware)
      4. MTTR Overall direction

    Composite RAG: Green = all 4 green; Amber = 2–3 green; Red = 0–1 green.
    Missing-signal Amber cap (D-17-06) is structural.

    Owner Velocity: current Crit+High open counts per owner with MoM delta
    and >20% rise outlier flag (D-17-09). Degrades gracefully when the
    owner-dimension snapshot has insufficient history.

    This module's compute() is the sole source of truth for the four-channel
    render contract (Plan 17-03). All metrics, table_data, analyst_rows,
    rag_strip, and driver_narrative are populated here; renderers are pure.
    """

    MODULE_ID         = "program_health"
    DISPLAY_NAME      = "Program Health Overview"
    DESCRIPTION       = (
        "Four-signal composite program health RAG — Open-Critical MoM, "
        "Net Velocity, SLA Posture, and MTTR — with Owner velocity table."
    )
    REQUIRED_DATA     = ["vulns", "assets", "trend_snapshots"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # _build_cold_start_result
    # ------------------------------------------------------------------

    def _build_cold_start_result(
        self,
        vulns_df: pd.DataFrame,
        report_date: Any,
        config: ModuleConfig,
    ) -> ModuleData:
        """
        Return a coherent cold-start ModuleData for < 2 snapshots (D-17-08).

        Diverges from the mttr_trend analog at ONE point:
          rag_strip status = "yellow" (NOT "no_data") with
          headline_value_str = "Trend Being Established".

        Current-value Open-Critical and SLA posture are computed from the
        live vulns_df so cold-start tiles show real numbers (not "—").
        MTTR current tile is None ("—") — no fixed_vulns_df is provided
        (D-17-01 / RESEARCH Open Question 2).
        """
        # Current-value Open-Critical count (reopened-aware)
        curr_open_crit: Optional[int] = None
        curr_sla_rate: Optional[float] = None

        if not vulns_df.empty:
            try:
                open_df = open_findings_at(vulns_df, report_date)
                if not open_df.empty and "severity" in open_df.columns:
                    # Open-Critical count
                    curr_open_crit = int(
                        (open_df["severity"].str.lower() == "critical").sum()
                    )
                    # SLA posture: Crit+High within SLA
                    # D-05 / WR-01: shared helper excludes NaT first_found from both sides
                    if "first_found" in open_df.columns:
                        curr_sla_rate = compute_sla_rate_crit_high(open_df, report_date, SLA_DAYS)
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "%s cold-start current-value computation failed: %s",
                    self._log_prefix(), exc,
                )

        metrics = {
            "cold_start":        True,
            "composite_rag":     "amber",
            "data_incomplete":   True,
            "open_crit_current": curr_open_crit,
            "sla_rate_current":  curr_sla_rate,
            "mttr_current":      None,
            "net_delta_current": None,
            # D-3: surfaced current intake/fixed counts — None in cold-start
            "new_current":       None,
            "fixed_current":     None,
            # D-3: current-sign status — flat in cold-start (no net data)
            "net_velocity_status_current": "flat",
            # Signal statuses all missing in cold-start
            "signal_open_crit_status":   "missing",
            "signal_net_velocity_status": "missing",
            "signal_sla_rate_status":    "missing",
            "signal_mttr_status":        "missing",
            "missing_signal_names":      [
                "Open Critical count", "Net Velocity",
                "SLA Posture", "MTTR",
            ],
        }

        driver_narrative = (
            "Program health trend being established — "
            "month-over-month direction available from next snapshot."
        )

        summary_text = (
            f"Program Health Overview — cold start. "
            f"Current Open Critical: {safe_int(curr_open_crit)}. "
            f"SLA Posture: {safe_pct(curr_sla_rate)}."
        )

        return ModuleData(
            module_id        = self.MODULE_ID,
            display_name     = self.DISPLAY_NAME,
            metrics          = metrics,
            table_data       = [],
            chart_data       = {},
            summary_text     = summary_text,
            metadata         = {"cold_start": True},
            driver_narrative = driver_narrative,
            analyst_rows     = [],
            rag_strip        = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = "Trend Being Established",
                status             = "yellow",   # D-17-08 — NOT "no_data"
            ),
            error            = None,
        )

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute the 4-signal composite program health RAG.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered open vulnerability DataFrame.
        assets_df : pd.DataFrame
            Asset DataFrame for extract_owner() Owner velocity cut.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Options (all overridable via module_options YAML):
              green_count_min  (int, default 4)  — signals needed for Green.
              amber_count_min  (int, default 2)  — signals needed for Amber.
              open_crit_flat_abs (int, default 5) — ±N = flat for Open-Crit.
              sla_rate_flat_pct (float, default 2.0) — ±N% = flat for SLA.
              mttr_flat_days   (float, default 1.0) — ±N days = flat for MTTR.
              owner_outlier_pct (float, default 20.0) — % rise threshold.
              tag_category, tag_value — injected by composed_report.
        **kwargs
            ``trend_snapshots`` (dict) — severity-dimension read_trend() result:
            ``{"snapshots": [...], "insufficient_data": bool}``.
            Delivered via _MODULES_NEEDING_TREND_SNAPSHOTS. Absent /
            insufficient → cold-start (QUAL-01).
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            # ----------------------------------------------------------------
            # QUAL-03 — empty-data guard
            # ----------------------------------------------------------------
            if vulns_df.empty:
                logger.warning(
                    "%s vulns_df empty — returning _empty_result.", self._log_prefix()
                )
                return self._empty_result("vulns_df is empty", config)

            # ----------------------------------------------------------------
            # Read module_options. The composer runs validate_config() before
            # compute() (rejecting non-coercible values), so the int()/float()
            # reads below are safe; each also carries a per-key default. Any
            # residual bad value is caught by this method's fail-soft wrapper.
            # ----------------------------------------------------------------
            green_count_min   = int(config.options.get("green_count_min",   4))
            amber_count_min   = int(config.options.get("amber_count_min",   2))
            open_crit_flat    = int(config.options.get("open_crit_flat_abs", 5))
            sla_rate_flat     = float(config.options.get("sla_rate_flat_pct", 2.0))
            mttr_flat         = float(config.options.get("mttr_flat_days",    1.0))
            owner_outlier_pct = float(config.options.get("owner_outlier_pct", 20.0))
            tag_category      = config.options.get("tag_category")
            tag_value         = config.options.get("tag_value")

            # ----------------------------------------------------------------
            # QUAL-01 — cold-start guard
            # ----------------------------------------------------------------
            trend_snapshots = kwargs.get("trend_snapshots")
            cold_start = (
                trend_snapshots is None
                or trend_snapshots.get("insufficient_data", True)
            )
            if cold_start:
                logger.info(
                    "%s trend_snapshots absent or insufficient — cold start.",
                    self._log_prefix(),
                )
                return self._build_cold_start_result(vulns_df, report_date, config)

            snapshots: list[dict] = trend_snapshots.get("snapshots", [])
            if not snapshots:
                return self._build_cold_start_result(vulns_df, report_date, config)

            # ----------------------------------------------------------------
            # Deduplicate snapshots by calendar month (D-16-08 pattern)
            # VERBATIM from mttr_trend_module lines 646–651
            # ----------------------------------------------------------------
            by_month: dict[str, dict] = {}
            for snap in snapshots:
                m = snap.get("month", "")
                if m not in by_month or snap.get("generated_at", "") > by_month[m].get("generated_at", ""):
                    by_month[m] = snap
            snapshots_deduped = [by_month[m] for m in sorted(by_month)]

            if len(snapshots_deduped) < 2:
                logger.info(
                    "%s only %d deduped snapshot(s) — cold start.",
                    self._log_prefix(), len(snapshots_deduped),
                )
                return self._build_cold_start_result(vulns_df, report_date, config)

            curr = snapshots_deduped[-1]
            prev = snapshots_deduped[-2]

            # ----------------------------------------------------------------
            # Signal 1 — Open-Critical MoM (higher_is_better=False)
            # ----------------------------------------------------------------
            curr_crit = curr.get("critical")
            prev_crit = prev.get("critical")
            sig1_status = _signal_direction(
                curr_crit, prev_crit,
                higher_is_better=False,
                flat_band=float(open_crit_flat),
            )

            # ----------------------------------------------------------------
            # Signal 2 — Net Velocity MoM (higher_is_better=False)
            # Direction: "delta of deltas" — curr_net < prev_net → improving
            # Net velocity is binary; no flat band (flat_band=0.0)
            # ----------------------------------------------------------------
            curr_new = curr.get("new_findings_count")
            curr_fix = curr.get("fixed_findings_count")
            prev_new = prev.get("new_findings_count")
            prev_fix = prev.get("fixed_findings_count")

            curr_net_delta: Optional[float] = None
            prev_net_delta: Optional[float] = None
            if curr_new is not None and curr_fix is not None:
                curr_net_delta = float(curr_new) - float(curr_fix)
            if prev_new is not None and prev_fix is not None:
                prev_net_delta = float(prev_new) - float(prev_fix)

            sig2_status = _signal_direction(
                curr_net_delta, prev_net_delta,
                higher_is_better=False,
                flat_band=0.0,
            )

            # ----------------------------------------------------------------
            # D-7: current-sign Net Velocity status — derived from sign(curr_net_delta).
            # Computed here (before signal_statuses list) so composite RAG, green_count,
            # and narrative all use the SAME status the tile shows.
            # net < 0 → "green" (fixed > intake); net > 0 → "red"; net == 0 → "flat".
            # Missing (curr_net_delta is None) → "missing" so the missing-signal cap
            # (D-17-06) still applies correctly for a genuinely absent Net Velocity.
            # sig2_status (delta-of-deltas sparkline trend) is kept for sparkline
            # annotation only but is no longer used for the composite.
            # ----------------------------------------------------------------
            if curr_net_delta is None:
                net_velocity_status_current = "missing"
            elif curr_net_delta < 0:
                net_velocity_status_current = "green"
            elif curr_net_delta > 0:
                net_velocity_status_current = "red"
            else:
                net_velocity_status_current = "flat"

            # ----------------------------------------------------------------
            # Signal 3 — SLA Posture (higher_is_better=True)
            # Reads sla_rate_crit_high from snapshot (Plan 17-01 field)
            # ----------------------------------------------------------------
            curr_sla = curr.get("sla_rate_crit_high")
            prev_sla = prev.get("sla_rate_crit_high")
            sig3_status = _signal_direction(
                curr_sla, prev_sla,
                higher_is_better=True,
                flat_band=sla_rate_flat,
            )

            # ----------------------------------------------------------------
            # Signal 4 — MTTR Overall (higher_is_better=False)
            # ----------------------------------------------------------------
            curr_mttr = curr.get("mttr_overall_days")
            prev_mttr = prev.get("mttr_overall_days")
            sig4_status = _signal_direction(
                curr_mttr, prev_mttr,
                higher_is_better=False,
                flat_band=mttr_flat,
            )

            # D-7: use current-sign net velocity status for the composite slot (slot 1),
            # not the delta-of-deltas sig2_status. sig2_status is still available for
            # sparkline MoM annotation if needed; it is no longer used for the composite.
            signal_statuses = [sig1_status, net_velocity_status_current, sig3_status, sig4_status]
            signal_names    = [
                "Open Critical count", "Net Velocity", "SLA Posture", "MTTR"
            ]

            # ----------------------------------------------------------------
            # Composite RAG (OD-5 + D-17-06 missing cap)
            # ----------------------------------------------------------------
            composite_rag, data_incomplete = _composite_rag_od5(
                signal_statuses,
                green_min=green_count_min,
                amber_min=amber_count_min,
            )

            missing_signal_names = [
                name for name, status in zip(signal_names, signal_statuses)
                if status == "missing"
            ]
            green_count = sum(1 for s in signal_statuses if s == "green")

            # ----------------------------------------------------------------
            # Current-value SLA posture (reopened-aware, D-17-03/QUAL-02)
            # For the current-value tile — uses live vulns_df, not snapshot
            # ----------------------------------------------------------------
            curr_sla_tile: Optional[float] = None
            try:
                open_df_live = open_findings_at(vulns_df, report_date)
                if not open_df_live.empty and "severity" in open_df_live.columns:
                    # D-05 / WR-01: shared helper excludes NaT first_found from both sides
                    curr_sla_tile = compute_sla_rate_crit_high(open_df_live, report_date, SLA_DAYS)
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "%s SLA tile computation failed: %s", self._log_prefix(), exc
                )

            # Current-value Open-Critical
            curr_open_crit_tile: Optional[int] = None
            if curr_crit is not None:
                curr_open_crit_tile = int(curr_crit)

            # ----------------------------------------------------------------
            # Sparkline series for Plan 03 renderers
            # ----------------------------------------------------------------
            open_crit_series = [
                s.get("critical") for s in snapshots_deduped
            ]
            sla_series = [
                s.get("sla_rate_crit_high") for s in snapshots_deduped
            ]
            mttr_series = [
                s.get("mttr_overall_days") for s in snapshots_deduped
            ]
            net_velocity_series = [
                (
                    # WR-04: use s.get(key) with explicit None check so a genuinely-absent
                    # key produces a sparkline gap (None), not a misleading 0.
                    (float(s.get("new_findings_count")) - float(s.get("fixed_findings_count")))
                    if s.get("new_findings_count") is not None and s.get("fixed_findings_count") is not None
                    else None
                )
                for s in snapshots_deduped
            ]
            month_labels = [s.get("month", "") for s in snapshots_deduped]

            # ----------------------------------------------------------------
            # Owner Velocity (D-17-09)
            # Read owner-dimension snapshot inside compute (Option A)
            # ----------------------------------------------------------------
            owner_rows:            list[dict] = []
            owner_mom_suppressed:  bool = True   # True = suppress MoM Delta column
            owner_insufficient_note: bool = False

            try:
                from data.trend_store import (  # noqa: PLC0415
                    read_trend, _sanitise_tag_for_filename,
                )
                _tag_filter = _sanitise_tag_for_filename(tag_category, tag_value)
                owner_trend = read_trend(
                    dimension  = "owner",
                    tag_filter = _tag_filter,
                    months     = 13,
                )
                owner_insufficient = owner_trend.get("insufficient_data", True)

                # Open Crit+High per owner from live vulns_df (QUAL-02)
                open_df_owner = open_findings_at(vulns_df, report_date)
                enriched = extract_owner(assets_df)
                uuid_to_owner = dict(
                    zip(enriched["asset_uuid"], enriched["owner"])
                )

                if not open_df_owner.empty and "severity" in open_df_owner.columns:
                    ch_owner = open_df_owner[
                        open_df_owner["severity"].str.lower().isin(["critical", "high"])
                    ]
                    if not ch_owner.empty:
                        # CoW-safe: assign to a local Series, never to a column
                        owner_series = ch_owner["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                        curr_owner_counts: dict[str, int] = (
                            owner_series.value_counts().to_dict()
                        )
                    else:
                        curr_owner_counts = {}
                else:
                    curr_owner_counts = {}

                # Previous-month owner counts from snapshot
                prev_owner_counts: dict[str, int] = {}
                if not owner_insufficient:
                    owner_snaps = owner_trend.get("snapshots", [])
                    # Dedup owner snapshots by month
                    o_by_month: dict[str, dict] = {}
                    for osnap in owner_snaps:
                        m = osnap.get("month", "")
                        if m not in o_by_month or osnap.get("generated_at", "") > o_by_month[m].get("generated_at", ""):
                            o_by_month[m] = osnap
                    o_deduped = [o_by_month[m] for m in sorted(o_by_month)]
                    if len(o_deduped) >= 2:
                        prev_snap = o_deduped[-2]
                        # Owner snapshot stores owner-keyed counts at top level.
                        # WR-05: use the module-level constant so reader and writers
                        # cannot drift independently.
                        for key, val in prev_snap.items():
                            if key not in _OWNER_SNAPSHOT_METADATA_KEYS and isinstance(val, int):
                                prev_owner_counts[key] = val
                        owner_mom_suppressed = False
                    else:
                        owner_insufficient_note = True
                else:
                    owner_insufficient_note = True

                # D-5: in-scope total Open Crit+High for share_pct denominator
                in_scope_total = sum(curr_owner_counts.values())

                # D-5: per-owner asset count — count distinct assets per owner
                # extract_owner() always adds an "owner" column; dedup on asset_uuid
                # first so one asset with two Owner tags is not double-counted.
                enriched_deduped = enriched.drop_duplicates(subset=["asset_uuid"])
                owner_asset_counts: dict[str, int] = {}
                if "owner" in enriched_deduped.columns and not enriched_deduped.empty:
                    owner_asset_counts = (
                        enriched_deduped.groupby("owner")["asset_uuid"]
                        .count()
                        .to_dict()
                    )

                # Build owner_rows
                all_owners = sorted(
                    set(curr_owner_counts) | set(prev_owner_counts)
                )
                for owner_name in all_owners:
                    curr_cnt = curr_owner_counts.get(owner_name, 0)
                    prev_cnt = prev_owner_counts.get(owner_name, 0) if not owner_mom_suppressed else None
                    if prev_cnt is not None and prev_cnt > 0:
                        mom_delta = curr_cnt - prev_cnt
                        mom_delta_pct = (mom_delta / prev_cnt) * 100
                        outlier = mom_delta_pct > owner_outlier_pct
                    elif prev_cnt is not None and prev_cnt == 0 and curr_cnt > 0:
                        mom_delta = curr_cnt
                        mom_delta_pct = None
                        outlier = True
                    else:
                        mom_delta = None
                        mom_delta_pct = None
                        outlier = False

                    # D-5: share_pct — owner's share of in-scope Open Crit+High
                    # Guard divide-by-zero: zero total → None (renders "—")
                    share_pct: Optional[float] = None
                    if in_scope_total > 0:
                        share_pct = (curr_cnt / in_scope_total) * 100.0

                    owner_rows.append({
                        "owner":            owner_name,
                        "open_crit_high":   curr_cnt,
                        "prev_open":        prev_cnt,
                        "mom_delta":        mom_delta,
                        "mom_delta_pct":    mom_delta_pct,
                        "outlier":          outlier,
                        # D-5: new columns
                        "share_pct":        share_pct,
                        "asset_count":      owner_asset_counts.get(owner_name, 0),
                    })

                # Sort descending by current open count
                owner_rows.sort(key=lambda r: r["open_crit_high"], reverse=True)

            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "%s owner velocity computation failed: %s",
                    self._log_prefix(), exc,
                )
                owner_rows = []
                owner_mom_suppressed = True
                owner_insufficient_note = False

            # ----------------------------------------------------------------
            # RAG strip headline (UI-SPEC Copywriting Contract)
            # ----------------------------------------------------------------
            rag_strip_status = composite_rag  # "green" / "amber" / "red"
            # Map amber → "yellow" for build_rag_strip_entry API
            rag_key = "yellow" if rag_strip_status == "amber" else rag_strip_status
            headline_str = f"{green_count} / 4 On Track"
            if data_incomplete:
                headline_str += " (incomplete)"

            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = headline_str,
                status             = rag_key,
            )

            # ----------------------------------------------------------------
            # Driver narrative (UI-SPEC templates)
            # ----------------------------------------------------------------
            if composite_rag == "green":
                # Find best signal
                green_names = [
                    n for n, s in zip(signal_names, signal_statuses) if s == "green"
                ]
                top_signal = green_names[0] if green_names else "all signals"
                driver_narrative = (
                    f"The program improved on {green_count} of 4 indicators this month. "
                    f"{top_signal} shows improvement."
                )
            elif composite_rag == "amber":
                lag_names = [
                    n for n, s in zip(signal_names, signal_statuses)
                    if s in ("amber", "red")
                ]
                lag_str = lag_names[0] if lag_names else "one or more signals"
                driver_narrative = (
                    f"The program held steady on {green_count} of 4 indicators this month. "
                    f"{lag_str} warrants attention."
                )
            else:
                # Red
                bad_names = [
                    n for n, s in zip(signal_names, signal_statuses)
                    if s in ("amber", "red")
                ]
                bottom_str = ", ".join(bad_names[:2]) if bad_names else "multiple signals"
                driver_narrative = (
                    f"The program worsened on {4 - green_count} of 4 indicators this month. "
                    f"Immediate focus recommended on {bottom_str}."
                )

            if missing_signal_names:
                driver_narrative += (
                    f" Note: {', '.join(missing_signal_names)} data incomplete this period."
                )

            # ----------------------------------------------------------------
            # Summary text
            # ----------------------------------------------------------------
            composite_label = STATUS_LABEL.get(rag_key, rag_strip_status.title())
            summary_text = (
                f"Program Health Overview — {headline_str}. "
                f"Composite: {composite_label}. "
                f"SLA Posture: {safe_pct(curr_sla_tile)}. "
                f"Open Critical: {safe_int(curr_open_crit_tile)}."
            )

            # ----------------------------------------------------------------
            # Metrics dict (consumed by renderers)
            # ----------------------------------------------------------------
            metrics = {
                "composite_rag":     composite_rag,
                "data_incomplete":   data_incomplete,
                "cold_start":        False,
                "green_count":       green_count,
                "missing_signal_names": missing_signal_names,
                # Signal current values
                "open_crit_current":    curr_open_crit_tile,
                "net_delta_current":    curr_net_delta,
                "sla_rate_current":     curr_sla_tile,     # live, reopened-aware
                "mttr_current":         curr_mttr,          # from snapshot
                # D-3: surfaced intake/fixed counts for current-month annotation
                "new_current":          curr_new,
                "fixed_current":        curr_fix,
                # D-3: current-sign status (drives arrow/annotation, not sparkline)
                "net_velocity_status_current": net_velocity_status_current,
                # Signal statuses (MoM sparkline/trend — unchanged)
                "signal_open_crit_status":    sig1_status,
                "signal_net_velocity_status": sig2_status,
                "signal_sla_rate_status":     sig3_status,
                "signal_mttr_status":         sig4_status,
                # Previous-period values (for MoM arrows)
                "open_crit_prev":    prev_crit,
                "net_delta_prev":    prev_net_delta,
                "sla_rate_prev":     prev_sla,
                "mttr_prev":         prev_mttr,
                # Options passthrough for renderers
                "owner_mom_suppressed":    owner_mom_suppressed,
                "owner_insufficient_note": owner_insufficient_note,
                # Sparkline series
                "sparkline_months":        month_labels,
                "sparkline_open_crit":     open_crit_series,
                "sparkline_net_velocity":  net_velocity_series,
                "sparkline_sla_rate":      sla_series,
                "sparkline_mttr":          mttr_series,
            }

            # ----------------------------------------------------------------
            # Analyst rows (CONTRACT-02, QUAL-05)
            # Aggregate-only: owner tag name + counts + delta; no UUIDs/IPs
            # ----------------------------------------------------------------
            analyst_df = pd.DataFrame([
                {
                    "Owner":                r["owner"],
                    "Open Crit+High (curr)": r["open_crit_high"],
                    "Open Crit+High (prev)": r["prev_open"],
                    "MoM Delta":            r["mom_delta"],
                    "MoM Delta %":          (
                        round(r["mom_delta_pct"], 1)
                        if r["mom_delta_pct"] is not None else None
                    ),
                    "Outlier":              r["outlier"],
                }
                for r in owner_rows
            ]) if owner_rows else pd.DataFrame(
                columns=["Owner", "Open Crit+High (curr)", "Open Crit+High (prev)",
                         "MoM Delta", "MoM Delta %", "Outlier"]
            )
            # WR-05: drop rows where the current count is NaN (malformed snapshot data)
            # and ensure count columns are Python int, not numpy int64, for clean JSON/Excel output.
            if not analyst_df.empty:
                analyst_df = analyst_df.dropna(subset=["Open Crit+High (curr)"])
                analyst_df = analyst_df.assign(
                    **{
                        "Open Crit+High (curr)": analyst_df["Open Crit+High (curr)"].astype(int),
                    }
                )
            analyst_rows_list: list[tuple[str, pd.DataFrame]] = [
                ("PH — Owner Detail", analyst_df),
            ]

            return ModuleData(
                module_id        = self.MODULE_ID,
                display_name     = self.DISPLAY_NAME,
                metrics          = metrics,
                table_data       = owner_rows,
                chart_data       = {
                    "months":          month_labels,
                    "open_crit":       open_crit_series,
                    "net_velocity":    net_velocity_series,
                    "sla_rate":        sla_series,
                    "mttr":            mttr_series,
                },
                summary_text     = summary_text,
                metadata         = {
                    "snapshots_used": len(snapshots_deduped),
                    "tag_filter": f"{tag_category}={tag_value}" if tag_category and tag_value else "all_assets",
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows_list,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # _render_sparkline_b64  (T-17-08: plt.close mandatory)
    # ------------------------------------------------------------------

    def _render_sparkline_b64(
        self,
        values:          list,
        signal_label:    str,
        current_val_str: str,
        mom_arrow:       str,
        arrow_color:     str,
        line_color:      str,
    ) -> str:
        """
        Render a mini sparkline to a base64-encoded PNG string.

        Parameters
        ----------
        values : list
            Numeric series (may contain None — filtered to non-None pairs).
        signal_label : str
            Short label displayed as the chart title (e.g. "Open Critical").
        current_val_str : str
            Pre-formatted current value displayed in the title
            (e.g. "47", "87.3%", "—").
        mom_arrow : str
            Unicode arrow symbol (▼ / ▲ / —).
        arrow_color : str
            Hex color for the title (matches MoM arrow semantic color).
        line_color : str
            Hex color for the sparkline line.

        Returns
        -------
        str
            Base64-encoded PNG string (no data-URI prefix).

        Notes
        -----
        T-17-08: ``plt.close(fig)`` is called unconditionally to prevent
        matplotlib figure accumulation across many group runs.
        figsize=(2.0, 1.2) at dpi=120 per UI-SPEC §PDF sparkline spec.
        """
        fig, ax = plt.subplots(figsize=(2.0, 1.2))
        try:
            # Filter out None entries while preserving x-positions
            x_vals = [i for i, v in enumerate(values) if v is not None]
            y_vals = [v for v in values if v is not None]
            if len(y_vals) >= 2:
                ax.plot(x_vals, y_vals, color=line_color, linewidth=1.5)
                # Shade area under the line
                ax.fill_between(x_vals, y_vals, alpha=0.15, color=line_color)
            elif len(y_vals) == 1:
                ax.scatter(x_vals, y_vals, color=line_color, s=20)
            ax.set_title(
                f"{signal_label}\n{current_val_str} {mom_arrow}",
                fontsize=7,
                color=arrow_color,
                pad=2,
            )
            ax.axis("off")
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
            buf.seek(0)
            return base64.b64encode(buf.getvalue()).decode()
        finally:
            plt.close(fig)  # T-17-08: mandatory — prevents figure accumulation

    # ------------------------------------------------------------------
    # render_pdf_section  (D-17-09: 4 sparklines + Owner velocity table)
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render PDF section: 4-sparkline row (D-17-09) + Owner velocity table.

        Error guard: returns "" when ``data.error`` is set.
        Cold-start guard: replaces sparkline row with cold-start notice;
        Owner table still renders (current-snapshot counts, no MoM Delta).

        Sparkline line colors (UI-SPEC §Color — locked):
          Open-Critical  #d32f2f
          Net Velocity   #1976d2
          SLA Posture    #388e3c
          MTTR           #f57c00

        MoM arrows and colors:
          ▼ improved  #388e3c
          ▲ worsened  #d32f2f
          — flat/missing #9E9E9E

        T-17-07: all owner/tag strings escaped with html.escape() before
        interpolation into markup.
        """
        if data.error:
            return ""

        m = data.metrics
        cold_start = m.get("cold_start", False)

        # Locked per-signal line colors (UI-SPEC §Color)
        _LINE_COLORS = ["#d32f2f", "#1976d2", "#388e3c", "#f57c00"]
        _SIGNAL_LABELS = ["Open Critical", "Net Velocity", "SLA Posture", "MTTR"]
        _COLOR_IMPROVED = "#388e3c"
        _COLOR_WORSENED = "#d32f2f"
        _COLOR_FLAT     = "#9E9E9E"

        def _mom_arrow_and_color(status: str, higher_is_better: bool) -> tuple[str, str]:
            """Return (arrow_symbol, color) from signal status."""
            if status == "green":
                # Green = improved. Arrow direction: ▼ for lower-is-better, ▲ for higher-is-better
                arrow = "▲" if higher_is_better else "▼"
                return arrow, _COLOR_IMPROVED
            elif status == "red":
                arrow = "▼" if higher_is_better else "▲"
                return arrow, _COLOR_WORSENED
            return "—", _COLOR_FLAT

        # Build explanatory paragraph
        if cold_start:
            summary_p = (
                '<p class="explanatory-text">'
                "Month-over-month trend being established — available from next snapshot."
                "</p>"
            )
        else:
            summary_p = (
                f'<p class="explanatory-text">'
                f"{html.escape(data.summary_text)}"
                f"</p>"
            )

        # ------------------------------------------------------------------
        # Sparkline row (or cold-start notice)
        # ------------------------------------------------------------------
        # D-2: verbatim approved captions (spec lines 40-45 — copy EXACTLY)
        _CAPTION_OPEN_CRITICAL = (
            "open Critical-VPR findings — lower is better; ▼ green = falling"
        )
        _CAPTION_NET_VELOCITY = (
            "new findings minus fixed this window (intake − fixed) — negative is good"
            " (you fixed more than came in); ▼ green when fixed > intake, ▲ red when intake > fixed"
        )
        _CAPTION_SLA_POSTURE = (
            "% of open Critical + High findings still within SLA — higher is better;"
            " ▲ green = rising"
        )
        _CAPTION_MTTR = (
            "average days to remediate (rolling 30-day) — lower is better; ▼ green = falling"
        )
        _CHART_CAPTIONS = [
            _CAPTION_OPEN_CRITICAL,
            _CAPTION_NET_VELOCITY,
            _CAPTION_SLA_POSTURE,
            _CAPTION_MTTR,
        ]

        if cold_start:
            sparkline_row = (
                '<p style="color:#9E9E9E;font-style:italic;">'
                "Month-over-month trend being established — available from next snapshot."
                "</p>"
            )
        else:
            # D-8: Net Velocity tile — derive arrow/color from current-sign status.
            # The tile headline shows ONLY "net {net}" (no intake/fixed in the tile).
            # The "in {new} / fixed {fix}" breakdown moves to the caption row below.
            nv_status_curr = m.get("net_velocity_status_current", "flat")
            if nv_status_curr == "green":
                nv_arrow, nv_arrow_color = "▼", _COLOR_IMPROVED
            elif nv_status_curr == "red":
                nv_arrow, nv_arrow_color = "▲", _COLOR_WORSENED
            else:
                nv_arrow, nv_arrow_color = "—", _COLOR_FLAT

            curr_new = m.get("new_current")
            curr_fix = m.get("fixed_current")
            curr_net = m.get("net_delta_current")
            if curr_new is not None and curr_fix is not None and curr_net is not None:
                new_str = f"{int(curr_new):,}"
                fix_str = f"{int(curr_fix):,}"
                net_sign = "+" if curr_net >= 0 else ""
                net_str = f"{net_sign}{int(curr_net):,}"
                # D-8: tile headline = net value only (uniform with other tiles)
                nv_tile_str = f"net {net_str}"
                # D-8: caption row carries the in/fixed breakdown (all three numbers shown per D-3)
                nv_breakdown_str = f"in {new_str} / fixed {fix_str}"
            else:
                nv_tile_str = "—"
                nv_breakdown_str = ""

            # D-4: MTTR caption — append "establishing from monthly snapshots" when None
            mttr_curr = m.get("mttr_current")
            if mttr_curr is not None:
                mttr_curr_str = safe_format(mttr_curr, ".0f") + " d"
                mttr_caption = _CAPTION_MTTR
            else:
                mttr_curr_str = "—"
                mttr_caption = _CAPTION_MTTR + " — establishing from monthly snapshots"

            # D-8: sparklines_data tuple: (values, label, curr_str, arrow, arrow_color, line_color, caption, extra_caption)
            # extra_caption is appended in the caption row (used for NV breakdown).
            # D-8 single arrow: for Net Velocity, pass mom_arrow="" to _render_sparkline_b64
            # so the PNG title is "Net Velocity\nnet -7" with NO appended arrow;
            # the arrow appears only once — in the annotation div beside the tile.
            oc_arrow, oc_arrow_color = _mom_arrow_and_color(m.get("signal_open_crit_status", "missing"), False)
            sla_arrow, sla_arrow_color = _mom_arrow_and_color(m.get("signal_sla_rate_status", "missing"), True)
            mttr_arrow, mttr_arrow_color = _mom_arrow_and_color(m.get("signal_mttr_status", "missing"), False)

            sparklines_data = [
                (
                    m.get("sparkline_open_crit", []),
                    "Open Critical",
                    safe_int(m.get("open_crit_current")),
                    oc_arrow,
                    oc_arrow_color,
                    "#d32f2f",
                    _CAPTION_OPEN_CRITICAL,
                    "",                           # no extra caption breakdown
                ),
                (
                    m.get("sparkline_net_velocity", []),
                    "Net Velocity",
                    nv_tile_str,                  # D-8: net value only in tile
                    nv_arrow,
                    nv_arrow_color,
                    "#1976d2",
                    _CAPTION_NET_VELOCITY,
                    nv_breakdown_str,             # D-8: breakdown in caption row
                ),
                (
                    m.get("sparkline_sla_rate", []),
                    "SLA Posture",
                    safe_pct(m.get("sla_rate_current")),
                    sla_arrow,
                    sla_arrow_color,
                    "#388e3c",
                    _CAPTION_SLA_POSTURE,
                    "",
                ),
                (
                    m.get("sparkline_mttr", []),
                    "MTTR",
                    mttr_curr_str,
                    mttr_arrow,
                    mttr_arrow_color,
                    "#f57c00",
                    mttr_caption,
                    "",
                ),
            ]

            # D-8: two-row layout.
            # Top row: 4 tile cells (sparkline PNG + headline value + arrow annotation).
            #   Each tile is uniform height/font; no definition text inside the tile.
            # Caption row: 4 caption cells aligned under each chart with definition text
            #   and (for Net Velocity) the in/fixed breakdown.
            tile_cells_html = ""
            caption_cells_html = ""
            for values, label, curr_str, arrow, arrow_color, line_color, caption, extra_caption in sparklines_data:
                # D-8 single arrow: for Net Velocity pass mom_arrow="" so _render_sparkline_b64
                # does NOT append a second arrow in the PNG title. The arrow is shown once in
                # the annotation div.
                png_mom_arrow = "" if label == "Net Velocity" else arrow
                try:
                    b64 = self._render_sparkline_b64(
                        values          = values,
                        signal_label    = label,
                        current_val_str = curr_str,
                        mom_arrow       = png_mom_arrow,
                        arrow_color     = arrow_color,
                        line_color      = line_color,
                    )
                    img_tag = f'<img src="data:image/png;base64,{b64}" style="width:100%;max-width:160pt;">'
                except Exception as exc:  # noqa: BLE001
                    logger.warning("%s sparkline render failed for %s: %s", self._log_prefix(), label, exc)
                    img_tag = f'<span style="color:#9E9E9E;font-size:8pt;">{html.escape(label)}: chart unavailable</span>'

                # Annotation: arrow symbol under the PNG (one per tile, uniform font)
                arrow_annotation = (
                    f'<div style="font-size:8pt;color:{arrow_color};font-weight:bold;margin-top:1pt;">'
                    f"{html.escape(arrow)}"
                    f"</div>"
                )

                tile_cells_html += (
                    f'<div style="display:table-cell;text-align:center;'
                    f'width:25%;padding:0 4pt;vertical-align:top;">'
                    f"{img_tag}"
                    f"{arrow_annotation}"
                    f"</div>"
                )

                # Caption row: definition text + optional breakdown (NV in/fixed)
                breakdown_html = ""
                if extra_caption:
                    breakdown_html = (
                        f'<div style="font-size:7pt;color:{arrow_color};font-weight:bold;'
                        f'margin-top:1pt;">'
                        f"{html.escape(extra_caption)}"
                        f"</div>"
                    )
                caption_html = (
                    f'<div style="font-size:7pt;color:#666;margin-top:1pt;text-align:left;">'
                    f'<strong>{html.escape(label)}</strong> — {html.escape(caption)}'
                    f"</div>"
                    f"{breakdown_html}"
                )

                caption_cells_html += (
                    f'<div style="display:table-cell;text-align:left;'
                    f'width:25%;padding:0 4pt;vertical-align:top;">'
                    f"{caption_html}"
                    f"</div>"
                )

            sparkline_row = (
                f'<p style="font-size:8.5pt;color:#444;margin-bottom:4pt;">'
                f"Month-over-month signal trends</p>"
                # Top row: 4 uniform tiles (sparkline + headline + arrow)
                f'<div style="display:table;width:100%;margin-bottom:4pt;">'
                f"{tile_cells_html}"
                f"</div>"
                # Caption row: 4 cells with definitions + NV breakdown
                f'<div style="display:table;width:100%;margin-bottom:8pt;">'
                f"{caption_cells_html}"
                f"</div>"
            )

        # ------------------------------------------------------------------
        # Owner velocity table
        # ------------------------------------------------------------------
        owner_mom_suppressed  = m.get("owner_mom_suppressed", True)
        owner_insufficient_note = m.get("owner_insufficient_note", False)
        owner_rows = data.table_data or []

        if cold_start:
            # Cold-start: current-only — D-5 columns minus MoM (no history yet)
            owner_header = (
                "<thead><tr>"
                "<th>Owner</th>"
                "<th style='text-align:right;'>Open Crit+High</th>"
                "<th style='text-align:right;'>Share %</th>"
                "<th style='text-align:right;'>Assets</th>"
                "</tr></thead>"
            )
            if owner_rows:
                owner_body = "".join(
                    f"<tr>"
                    f"<td>{html.escape(str(r.get('owner', '')))}</td>"
                    f"<td style='text-align:right;'>{safe_int(r.get('open_crit_high'))}</td>"
                    f"<td style='text-align:right;'>{safe_pct(r.get('share_pct'))}</td>"
                    f"<td style='text-align:right;'>{safe_int(r.get('asset_count'))}</td>"
                    f"</tr>"
                    for r in owner_rows
                )
            else:
                owner_body = (
                    "<tr><td colspan='4' style='color:#9E9E9E;'>"
                    "No owner data available.</td></tr>"
                )
            owner_note = ""
        else:
            # Normal: D-5 six columns — Owner | Open Crit+High | Share % | Assets | MoM Delta | MoM Delta %
            if owner_mom_suppressed:
                # Snapshot data insufficient — suppress MoM Delta / MoM Delta % columns
                owner_header = (
                    "<thead><tr>"
                    "<th>Owner</th>"
                    "<th style='text-align:right;'>Open Crit+High</th>"
                    "<th style='text-align:right;'>Share %</th>"
                    "<th style='text-align:right;'>Assets</th>"
                    "</tr></thead>"
                )
                owner_body = "".join(
                    f"<tr>"
                    f"<td>{html.escape(str(r.get('owner', '')))}</td>"
                    f"<td style='text-align:right;'>{safe_int(r.get('open_crit_high'))}</td>"
                    f"<td style='text-align:right;'>{safe_pct(r.get('share_pct'))}</td>"
                    f"<td style='text-align:right;'>{safe_int(r.get('asset_count'))}</td>"
                    f"</tr>"
                    for r in owner_rows
                ) if owner_rows else (
                    "<tr><td colspan='4' style='color:#9E9E9E;'>"
                    "No owner data available.</td></tr>"
                )
                owner_note = (
                    '<p style="color:#9E9E9E;font-style:italic;font-size:8pt;">'
                    "Owner month-over-month trend being established.</p>"
                    if owner_insufficient_note else ""
                )
            else:
                # Full D-5 six columns
                owner_header = (
                    "<thead><tr>"
                    "<th>Owner</th>"
                    "<th style='text-align:right;'>Open Crit+High</th>"
                    "<th style='text-align:right;'>Share %</th>"
                    "<th style='text-align:right;'>Assets</th>"
                    "<th style='text-align:right;'>MoM Delta</th>"
                    "<th style='text-align:right;'>MoM Delta %</th>"
                    "</tr></thead>"
                )
                owner_body_parts = []
                for r in (owner_rows or []):
                    outlier = r.get("outlier", False)
                    mom_delta = r.get("mom_delta")
                    mom_delta_pct = r.get("mom_delta_pct")
                    mom_str = (
                        ("+" if mom_delta >= 0 else "") + safe_format(mom_delta, ".0f")
                    ) if mom_delta is not None else "—"
                    # Outlier marker appended to MoM Delta % (▲ Outlier in red)
                    if outlier:
                        mom_pct_str = (
                            (("+" if mom_delta_pct >= 0 else "") + safe_format(mom_delta_pct, ".1f") + "%")
                            if mom_delta_pct is not None else ""
                        ) + ' <span style="color:#d32f2f;font-weight:bold;">&#9650; Outlier</span>'
                    else:
                        mom_pct_str = (
                            ("+" if mom_delta_pct >= 0 else "") + safe_format(mom_delta_pct, ".1f") + "%"
                        ) if mom_delta_pct is not None else "—"
                    owner_body_parts.append(
                        f"<tr>"
                        f"<td>{html.escape(str(r.get('owner', '')))}</td>"
                        f"<td style='text-align:right;'>{safe_int(r.get('open_crit_high'))}</td>"
                        f"<td style='text-align:right;'>{safe_pct(r.get('share_pct'))}</td>"
                        f"<td style='text-align:right;'>{safe_int(r.get('asset_count'))}</td>"
                        f"<td style='text-align:right;'>{mom_str}</td>"
                        f"<td style='text-align:right;'>{mom_pct_str}</td>"
                        f"</tr>"
                    )
                owner_body = "".join(owner_body_parts) if owner_body_parts else (
                    "<tr><td colspan='6' style='color:#9E9E9E;'>"
                    "No owner data available.</td></tr>"
                )
                owner_note = ""

        # D-1: page-break before Owner table so it begins on page 2
        owner_table = f"""
<div class="page-break"></div>
  <h3 class="subsection-heading">Owner Velocity — Open Critical + High</h3>
  <table class="data-table" style="width:100%;margin-top:4pt;">
    {owner_header}
    <tbody>
      {owner_body}
    </tbody>
  </table>
  {owner_note}"""

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  {summary_p}
  {sparkline_row}
  {owner_table}
</div>"""

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01 — 4-tile KPI row + narrative)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Normal: header bg #FFF3E0 with RAG-color left border; 4-tile KPI
        row (Open Critical / Net Velocity / SLA Posture (Crit+High) /
        MTTR (30-day)); driver_narrative; missing-signal note if any.

        Cold-start: header bg #F5F5F5, #757575 left border; current-value
        tiles; "trend being established" notice; no NaN%.

        Returns "" on error. Inline CSS only (Outlook/Gmail/Apple Mail).
        T-17-07: html.escape() on all tag-derived strings.
        """
        if data.error:
            return ""

        m = data.metrics
        cold_start = m.get("cold_start", False)

        # ------------------------------------------------------------------
        # Cold-start branch (per mttr_trend_module pattern 1285–1294)
        # ------------------------------------------------------------------
        if cold_start:
            open_crit_str  = f"{safe_int(m.get('open_crit_current'))} open"
            sla_str        = safe_pct(m.get("sla_rate_current"))
            cold_tiles = (
                f'<div style="display:table;width:100%;margin:6px 0;">'
                f'<div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">'
                f'<strong style="font-size:13px;color:#1F3864;">{safe_int(m.get("open_crit_current"))}</strong>'
                f'<div style="font-size:10px;color:#757575;">Open Critical</div>'
                f"</div>"
                f'<div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">'
                f'<strong style="font-size:13px;color:#1F3864;">—</strong>'
                f'<div style="font-size:10px;color:#757575;">Net Velocity</div>'
                f"</div>"
                f'<div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">'
                f'<strong style="font-size:13px;color:#1F3864;">{sla_str}</strong>'
                f'<div style="font-size:10px;color:#757575;">SLA Posture (Crit+High)</div>'
                f"</div>"
                f'<div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">'
                f'<strong style="font-size:13px;color:#1F3864;">—</strong>'
                f'<div style="font-size:10px;color:#757575;">MTTR (30-day)</div>'
                f"</div>"
                f"</div>"
            )
            return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#F5F5F5;border-left:4px solid #757575;">
      <strong style="font-size:13px;">{html.escape(self.DISPLAY_NAME)}</strong><br>
      <span style="font-size:12px;color:#757575;">
        Program health trend being established.
      </span>
      {cold_tiles}
      <em style="font-size:11px;color:#555;">{html.escape(data.driver_narrative or '')}</em>
    </td>
  </tr>
</table>"""

        # ------------------------------------------------------------------
        # Normal render
        # ------------------------------------------------------------------
        composite_rag  = m.get("composite_rag", "amber")
        data_incomplete = m.get("data_incomplete", False)
        green_count    = m.get("green_count", 0)

        # Map composite_rag → rag_utils key ("amber" → "yellow")
        rag_key   = "yellow" if composite_rag == "amber" else composite_rag
        rag_color = STATUS_COLOR.get(rag_key, STATUS_COLOR["no_data"])
        composite_label = STATUS_LABEL.get(rag_key, composite_rag.title())

        # MoM arrow helpers
        _COLOR_IMPROVED = "#388e3c"
        _COLOR_WORSENED = "#d32f2f"
        _COLOR_FLAT     = "#9E9E9E"

        def _arrow_html(status: str, higher_is_better: bool) -> str:
            if status == "green":
                sym   = "▲" if higher_is_better else "▼"
                color = _COLOR_IMPROVED
            elif status == "red":
                sym   = "▼" if higher_is_better else "▲"
                color = _COLOR_WORSENED
            else:
                sym, color = "—", _COLOR_FLAT
            return f'<span style="color:{color};font-weight:bold;">{sym}</span>'

        # D-3: Net Velocity — intake/fixed/net display using current-sign status
        nv_status_curr = m.get("net_velocity_status_current", "flat")
        if nv_status_curr == "green":
            nv_curr_arrow_sym, nv_curr_color = "▼", _COLOR_IMPROVED
        elif nv_status_curr == "red":
            nv_curr_arrow_sym, nv_curr_color = "▲", _COLOR_WORSENED
        else:
            nv_curr_arrow_sym, nv_curr_color = "—", _COLOR_FLAT

        curr_new_e = m.get("new_current")
        curr_fix_e = m.get("fixed_current")
        curr_net_e = m.get("net_delta_current")
        if curr_new_e is not None and curr_fix_e is not None and curr_net_e is not None:
            net_sign = "+" if curr_net_e >= 0 else ""
            net_vel_val = (
                f"in {int(curr_new_e):,} / fixed {int(curr_fix_e):,}"
                f" · net {net_sign}{int(curr_net_e):,} {nv_curr_arrow_sym}"
            )
            nv_arrow_html = f'<span style="color:{nv_curr_color};font-weight:bold;">{nv_curr_arrow_sym}</span>'
        else:
            net_vel_val = "—"
            nv_arrow_html = f'<span style="color:{_COLOR_FLAT};font-weight:bold;">—</span>'

        # D-4: MTTR tile — always "—"; append establishing caption when None
        mttr_curr = m.get("mttr_current")
        if mttr_curr is not None:
            mttr_val = safe_format(mttr_curr, ".0f") + " d"
            mttr_caption_note = ""
        else:
            mttr_val = "—"
            mttr_caption_note = (
                '<div style="font-size:9px;color:#757575;font-style:italic;">'
                "establishing from monthly snapshots</div>"
            )

        # Tile values
        open_crit_val = f"{safe_int(m.get('open_crit_current'))} open"
        sla_val  = safe_pct(m.get("sla_rate_current"))

        # Arrow HTML per signal (MoM sparkline direction — unchanged)
        arrow1 = _arrow_html(m.get("signal_open_crit_status",    "missing"), False)
        arrow3 = _arrow_html(m.get("signal_sla_rate_status",     "missing"), True)
        arrow4 = _arrow_html(m.get("signal_mttr_status",         "missing"), False)

        # D-2: per-tile definition text (inline-CSS, Outlook-safe, no <style> blocks)
        _DEF_OPEN_CRIT = "open Critical-VPR findings — lower is better; &#9660; green = falling"
        _DEF_NET_VEL   = (
            "new findings minus fixed this window (intake &#8722; fixed) — "
            "negative is good (you fixed more than came in)"
        )
        _DEF_SLA       = "% of open Crit+High still within SLA — higher is better; &#9650; green = rising"
        _DEF_MTTR      = "average days to remediate (rolling 30-day) — lower is better; &#9660; green = falling"

        def _tile_def(text: str) -> str:
            return (
                f'<div style="font-size:9px;color:#9E9E9E;margin-top:2px;">{text}</div>'
            )

        tiles_html = f"""
<div style="display:table;width:100%;margin:6px 0;">
  <div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">
    <strong style="font-size:13px;color:#1F3864;">{html.escape(open_crit_val)}</strong>
    {arrow1}
    <div style="font-size:10px;color:#757575;">Open Critical</div>
    {_tile_def(_DEF_OPEN_CRIT)}
  </div>
  <div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">
    <strong style="font-size:13px;color:{nv_curr_color};">{html.escape(net_vel_val)}</strong>
    <div style="font-size:10px;color:#757575;">Net Velocity</div>
    {_tile_def(_DEF_NET_VEL)}
  </div>
  <div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">
    <strong style="font-size:13px;color:#1F3864;">{html.escape(sla_val)}</strong>
    {arrow3}
    <div style="font-size:10px;color:#757575;">SLA Posture (Crit+High)</div>
    {_tile_def(_DEF_SLA)}
  </div>
  <div style="display:table-cell;width:25%;padding:4px 8px;text-align:center;">
    <strong style="font-size:13px;color:#1F3864;">{html.escape(mttr_val)}</strong>
    {arrow4}
    <div style="font-size:10px;color:#757575;">MTTR (30-day)</div>
    {mttr_caption_note}
    {_tile_def(_DEF_MTTR)}
  </div>
</div>"""

        # Missing-signal note
        missing_names = m.get("missing_signal_names", [])
        missing_note = ""
        if missing_names and data_incomplete:
            names_str = ", ".join(html.escape(n) for n in missing_names)
            missing_note = (
                f'<span style="font-size:10px;color:#757575;">'
                f"Note: {names_str} data incomplete this period.</span>"
            )

        narrative_html = (
            f'<em style="font-size:11px;color:#555;">'
            f"{html.escape(data.driver_narrative or '')}</em>"
        )

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#FFF3E0;border-left:4px solid {rag_color};">
      <strong style="font-size:13px;">{html.escape(self.DISPLAY_NAME)}</strong><br>
      <span style="font-size:12px;">
        {html.escape(str(green_count))} of 4 signals on track &mdash; {html.escape(composite_label)}
      </span>
      {tiles_html}
      {narrative_html}<br>
      {missing_note}
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_excel_tabs  ("Program Health" + "Owner Velocity")
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write "Program Health" summary tab and "Owner Velocity" tab.

        Program Health: A1 title 13pt bold; A2 subtitle italic #757575;
        composite status + 4 signal current values/arrows.
        Owner Velocity: aggregate-only (QUAL-05 — no asset-level fields).
        Returns [] on exception.
        """
        tab_ph = "Program Health"
        tab_ov = "Owner Velocity"

        try:
            # ----------------------------------------------------------------
            # Tab 1 — Program Health summary
            # ----------------------------------------------------------------
            ws_ph = workbook.create_sheet(tab_ph)

            if data.error:
                ws_ph["A1"] = "Error"
                ws_ph["B1"] = data.error
                ws_ov = workbook.create_sheet(tab_ov)
                ws_ov["A1"] = "Error"
                return [tab_ph, tab_ov]

            m          = data.metrics
            cold_start = m.get("cold_start", False)

            ws_ph["A1"] = "Program Health Overview"
            ws_ph["A1"].font = Font(bold=True, size=13)

            if cold_start:
                ws_ph["A2"] = "Trend being established — month-over-month direction available from next snapshot."
            else:
                ws_ph["A2"] = data.summary_text or ""
            ws_ph["A2"].font = Font(italic=True, color="757575")

            # D-6: readable header fill — match mttr_trend tab's readable style
            # PatternFill("solid", fgColor="E3F2FD") + Font(bold=True)
            # (replaces the old dark 1F3864 fill with black text)
            _FILL_HDR = PatternFill("solid", fgColor="E3F2FD")

            signal_rows = [
                ("Signal", "Current Value", "MoM Direction"),
                ("Open Critical",         safe_int(m.get("open_crit_current")),  m.get("signal_open_crit_status",    "—")),
                ("Net Velocity",
                 (("+" if (m.get("net_delta_current") or 0) >= 0 else "")
                  + safe_format(m.get("net_delta_current"), ".0f"))
                 if m.get("net_delta_current") is not None else "—",
                 m.get("signal_net_velocity_status", "—")),
                ("SLA Posture (Crit+High)", safe_pct(m.get("sla_rate_current")),  m.get("signal_sla_rate_status",    "—")),
                ("MTTR (30-day)",
                 safe_format(m.get("mttr_current"), ".0f") + " d"
                 if m.get("mttr_current") is not None else "—",
                 m.get("signal_mttr_status", "—")),
            ]

            for r_idx, (label, curr_val, direction) in enumerate(signal_rows, start=4):
                ws_ph.cell(row=r_idx, column=1, value=label)
                ws_ph.cell(row=r_idx, column=2, value=curr_val)
                ws_ph.cell(row=r_idx, column=3, value=direction)
                if r_idx == 4:  # header row — readable light-blue fill + bold
                    for c in range(1, 4):
                        ws_ph.cell(row=r_idx, column=c).font = Font(bold=True)
                        ws_ph.cell(row=r_idx, column=c).fill = _FILL_HDR

            for col_idx, w in enumerate([28, 18, 18], start=1):
                ws_ph.column_dimensions[get_column_letter(col_idx)].width = w

            # D-6: definitions block — 4 approved caption lines below the signal rows
            _DEFS_START_ROW = 4 + len(signal_rows) + 1   # one blank row gap
            _DEFINITIONS = [
                ("Open Critical",      "open Critical-VPR findings — lower is better; down-arrow green = falling"),
                ("Net Velocity",       "new findings minus fixed this window (intake - fixed) — negative is good (you fixed more than came in)"),
                ("SLA Posture",        "% of open Critical + High findings still within SLA — higher is better; up-arrow green = rising"),
                ("MTTR",               "average days to remediate (rolling 30-day) — lower is better; down-arrow green = falling"),
            ]
            for d_idx, (metric, definition) in enumerate(_DEFINITIONS, start=_DEFS_START_ROW):
                ws_ph.cell(row=d_idx, column=1, value=metric).font = Font(bold=True, color="444444")
                ws_ph.cell(row=d_idx, column=2, value=definition)

            # ----------------------------------------------------------------
            # Tab 2 — Owner Velocity (aggregate-only, QUAL-05)
            # ----------------------------------------------------------------
            ws_ov = workbook.create_sheet(tab_ov)
            ws_ov["A1"] = "Owner Velocity — Open Critical + High"
            ws_ov["A1"].font = Font(bold=True, size=13)

            owner_rows     = data.table_data or []
            mom_suppressed = m.get("owner_mom_suppressed", True)

            # D-5/D-6: Owner Velocity columns match PDF table
            # Base columns always present; MoM Delta / MoM Delta % added when not suppressed
            if mom_suppressed:
                ov_headers = ["Owner", "Open Crit+High", "Share %", "Assets"]
                ov_widths  = [28, 18, 12, 12]
            else:
                ov_headers = ["Owner", "Open Crit+High", "Share %", "Assets", "MoM Delta", "MoM Delta %"]
                ov_widths  = [28, 18, 12, 12, 14, 14]

            for c_idx, hdr in enumerate(ov_headers, start=1):
                cell = ws_ov.cell(row=3, column=c_idx, value=hdr)
                cell.font = Font(bold=True)
                cell.fill = _FILL_HDR

            for r_idx, row in enumerate(owner_rows, start=4):
                ws_ov.cell(row=r_idx, column=1, value=str(row.get("owner", "")))
                ws_ov.cell(row=r_idx, column=2, value=row.get("open_crit_high"))
                # Share % as display string (safe_pct handles None → "—")
                ws_ov.cell(row=r_idx, column=3, value=safe_pct(row.get("share_pct")))
                ws_ov.cell(row=r_idx, column=4, value=safe_int(row.get("asset_count")))
                if not mom_suppressed:
                    mom_delta = row.get("mom_delta")
                    mom_delta_pct = row.get("mom_delta_pct")
                    ws_ov.cell(row=r_idx, column=5, value=mom_delta)
                    outlier_note = " (▲ Outlier)" if row.get("outlier") else ""
                    ws_ov.cell(
                        row=r_idx, column=6,
                        value=(
                            (("+" if mom_delta_pct >= 0 else "") + f"{mom_delta_pct:.1f}%{outlier_note}")
                            if mom_delta_pct is not None else "—"
                        ),
                    )

            for c_idx, w in enumerate(ov_widths, start=1):
                ws_ov.column_dimensions[get_column_letter(c_idx)].width = w

            return [tab_ph, tab_ov]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02 — aggregate-only, QUAL-05)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns [("PH — Owner Detail", df)] when data available.
        Returns [] on error or cold-start (no MoM data to show).
        QUAL-05: columns are aggregate-only — Owner tag name + aggregate
        Crit+High counts + MoM delta/%/outlier. No asset UUIDs, IPs,
        hostnames, or plugin IDs.
        """
        if data.error or not data.analyst_rows:
            return []
        if data.metrics.get("cold_start"):
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present.
        Falls back to a gray "No Data" cell on error or empty strip.
        Amber composite maps to STATUS_COLOR["yellow"] (#f57c00) — never the
        Medium-severity color (which is reserved for severity tables only).
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip

    # ------------------------------------------------------------------
    # validate_config
    # ------------------------------------------------------------------

    def validate_config(self, config: ModuleConfig) -> list[str]:
        """
        Validate the program_health module_options.

        Per the four-channel contract, ``validate_config`` returns a
        ``list[str]`` of error messages (empty list = valid); the composer
        treats a non-empty return as a config-validation failure. The int
        thresholds (``green_count_min``, ``amber_count_min``,
        ``open_crit_flat_abs``) and float thresholds (``sla_rate_flat_pct``,
        ``mttr_flat_days``, ``owner_outlier_pct``) are checked for coercibility
        here (T-17-04 threshold-injection mitigation); ``compute()`` re-reads
        them defensively with ``int()``/``float()`` and per-key defaults.
        """
        _int_keys   = ("green_count_min", "amber_count_min", "open_crit_flat_abs")
        _float_keys = ("sla_rate_flat_pct", "mttr_flat_days", "owner_outlier_pct")

        errors: list[str] = []

        for key in _int_keys:
            val = config.options.get(key)
            if val is not None:
                try:
                    int(val)
                except (TypeError, ValueError):
                    errors.append(
                        f"program_health: '{key}' must be an integer, "
                        f"got {type(val).__name__}"
                    )

        for key in _float_keys:
            val = config.options.get(key)
            if val is not None:
                try:
                    float(val)
                except (TypeError, ValueError):
                    errors.append(
                        f"program_health: '{key}' must be a number, "
                        f"got {type(val).__name__}"
                    )

        return errors

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Signal 1 — Open-Critical MoM": (
                    "snap.get('critical') curr vs prev. higher_is_better=False. "
                    "flat_band=open_crit_flat_abs (default 5). "
                    "Reads S1 severity snapshot; definition parity with trend_store (D-17-02)."
                ),
                "Signal 2 — Net Velocity MoM": (
                    "curr_net_delta = new_findings_count - fixed_findings_count. "
                    "Direction = 'delta of deltas': curr_net < prev_net → improving. "
                    "Binary; flat_band=0 (no flat zone per Net Velocity contract). "
                    "higher_is_better=False."
                ),
                "Signal 3 — SLA Posture MoM": (
                    "snap.get('sla_rate_crit_high') — % Crit+High within config.SLA_DAYS, "
                    "computed reopened-aware by capture_trend_snapshot.py (Plan 17-01). "
                    "higher_is_better=True. flat_band=sla_rate_flat_pct (default 2.0)."
                ),
                "Signal 4 — MTTR Overall MoM": (
                    "snap.get('mttr_overall_days') — sample-weighted MTTR from snapshot. "
                    "higher_is_better=False. flat_band=mttr_flat_days (default 1.0)."
                ),
                "OD-5 Composite RAG": (
                    "Green = green_count >= green_count_min (default 4). "
                    "Amber = green_count >= amber_count_min (default 2). "
                    "Red = green_count < amber_count_min."
                ),
                "D-17-06 Missing-Signal Amber Cap": (
                    "Structural in _composite_rag_od5: any 'missing' signal AND "
                    "raw composite == 'green' → capped to 'amber' (data_incomplete=True). "
                    "Cannot be bypassed via module_options."
                ),
                "Current SLA Posture tile": (
                    "Computed from live vulns_df via open_findings_at() (reopened-aware). "
                    "open_findings_at(vulns_df, report_date) → filter Crit+High → "
                    "vectorized days_open <= config.SLA_DAYS[severity] (D-17-03/QUAL-02)."
                ),
                "Owner Velocity (D-17-09)": (
                    "Current Crit+High open counts from live vulns_df via extract_owner. "
                    "Previous from owner-dimension snapshot (read_trend dimension='owner'). "
                    "Outlier flag: MoM rise > owner_outlier_pct% (default 20%). "
                    "Degrades gracefully when owner trend insufficient_data=True."
                ),
                "Cold-start (D-17-08)": (
                    "< 2 deduped snapshots → composite='amber', rag_strip status='yellow', "
                    "headline='Trend Being Established'. Current Open-Crit and SLA tiles "
                    "show live values from vulns_df. MTTR tile = '—' (no fixed_vulns_df, "
                    "per D-17-01 / RESEARCH Open Question 2)."
                ),
            },
        }


# ---------------------------------------------------------------------------
# CLI smoke test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    from reports.modules.registry import discover, get_module

    discover()
    m = get_module("program_health")
    assert m is not None, "program_health not found in registry"
    assert m.MODULE_ID == "program_health"

    # Pure function smoke
    assert _composite_rag_od5(["green"] * 4) == ("green", False), "4-green → green"
    assert _composite_rag_od5(["green", "green", "green", "missing"]) == ("amber", True), "3+missing → amber"
    assert _composite_rag_od5(["green", "red", "red", "red"]) == ("red", False), "1 green → red"
    assert _signal_direction(40.0, 50.0, False, 0.0) == "green", "lower = improved for False"
    assert _signal_direction(None, 50.0, False, 0.0) == "missing", "None → missing"
    assert _signal_direction(50.0, 48.0, False, 5.0) == "amber", "within flat band"
    assert _signal_direction(90.0, 80.0, True, 2.0) == "green", "higher = improved for True"

    print("program_health module smoke test PASSED")
    sys.exit(0)
