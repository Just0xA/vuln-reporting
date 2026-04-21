"""
reports/modules/total_vulns_by_severity_module.py — Total Open Vulnerabilities by Severity.

Counts all open vulnerability findings grouped by VPR-derived severity tier.
Findings in the ``open`` or ``reopened`` state are included; informational
findings are excluded upstream by the fetcher and do not appear in vulns_df.

Usage
-----
::

    from reports.modules import registry
    from reports.modules.base import ModuleConfig

    mod  = registry.get("total_vulns_by_severity")()
    data = mod.compute(vulns_df, assets_df, report_date,
                       ModuleConfig("total_vulns_by_severity"))
    print(data.metrics)
    # {"critical": 12, "high": 47, "medium": 203, "low": 88,
    #  "total": 350, "total_label": "350"}
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SEVERITY_COLORS, SLA_DAYS
from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# Open vulnerability states — matches the definition used across all report scripts
_OPEN_STATES: frozenset[str] = frozenset({"open", "reopened"})

# Severity display order (descending risk)
_SEVERITIES: tuple[str, ...] = ("critical", "high", "medium", "low")

# Text colours for KPI tiles (white text on dark backgrounds, black on light)
_TILE_TEXT_COLORS: dict[str, str] = {
    "critical": "#ffffff",
    "high":     "#ffffff",
    "medium":   "#000000",
    "low":      "#000000",
}


@register_module
class TotalVulnsBySeverityModule(BaseModule):
    """
    Count all open vulnerability findings grouped by VPR-derived severity tier.

    The ``severity`` column in vulns_df is assumed to already reflect
    VPR-derived severity (produced by ``vpr_to_severity()`` in the fetcher).
    Informational findings are excluded upstream by the fetcher and will not
    appear in vulns_df.

    Supported options
    -----------------
    None — this module accepts no options.
    """

    MODULE_ID         = "total_vulns_by_severity"
    DISPLAY_NAME      = "Total Open Vulnerabilities by Severity"
    DESCRIPTION       = ("Counts all open vulnerability findings grouped by "
                         "VPR-derived severity tier.")
    REQUIRED_DATA     = ["vulns"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Count open findings per severity tier.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Must have ``state`` and
            ``severity`` columns.  ``severity`` must already be VPR-derived
            (lowercase: ``"critical"``, ``"high"``, ``"medium"``, ``"low"``).
        assets_df : pd.DataFrame
            Not used by this module; accepted for interface compatibility.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Module configuration (no options consumed).

        Returns
        -------
        ModuleData
            On success: ``error`` is None and all data fields are populated.
            On failure: ``error`` is set and data fields hold safe empty defaults.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            if vulns_df.empty or "state" not in vulns_df.columns:
                counts = {sev: 0 for sev in _SEVERITIES}
            else:
                open_df   = vulns_df[vulns_df["state"].str.lower().isin(_OPEN_STATES)]
                sev_lower = open_df["severity"].str.lower()
                counts    = {
                    sev: int((sev_lower == sev).sum())
                    for sev in _SEVERITIES
                }

            total       = sum(counts.values())
            total_label = f"{total:,}"

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            metrics = {
                **counts,
                "total":       total,
                "total_label": total_label,
            }

            chart_data = {
                "severity_counts": {**counts},
                "colors": {
                    "critical": SEVERITY_COLORS["critical"],
                    "high":     SEVERITY_COLORS["high"],
                    "medium":   SEVERITY_COLORS["medium"],
                    "low":      SEVERITY_COLORS["low"],
                },
            }

            summary_text = (
                f"There are currently {total:,} open vulnerabilities: "
                f"{counts['critical']:,} Critical, {counts['high']:,} High, "
                f"{counts['medium']:,} Medium, and {counts['low']:,} Low."
            )

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = metrics,
                table_data   = [],
                chart_data   = chart_data,
                summary_text = summary_text,
                metadata     = {
                    "filter_applied":  "state in {open, reopened}, excluding Informational",
                    "severity_source": "VPR score via vpr_to_severity()",
                    "computed_at":     computed_at,
                },
                error        = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render four KPI tiles (one per severity) plus a total count line.

        Returns an error callout if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m = data.metrics
        tiles_html = ""
        for sev in _SEVERITIES:
            count = m.get(sev, 0)
            bg    = SEVERITY_COLORS.get(sev, "#9E9E9E")
            fg    = _TILE_TEXT_COLORS.get(sev, "#000000")
            label = sev.capitalize()
            tiles_html += (
                f'<div class="kpi-tile" '
                f'style="background-color:{bg};">'
                f'<span class="kpi-value" style="color:{fg};">{count:,}</span>'
                f'<span class="kpi-label" style="color:{fg};">{label}</span>'
                f"</div>"
            )

        total = m.get("total", 0)

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <div class="kpi-row">
    {tiles_html}
  </div>
  <p style="text-align:center;font-size:8pt;margin-top:4pt;">
    Total Open: <strong>{total:,}</strong>
  </p>
  <p class="explanatory-text">
    This section shows the total number of open security vulnerabilities
    categorised by severity based on Tenable&rsquo;s Vulnerability Priority
    Rating (VPR) score. Critical vulnerabilities (VPR 9.0&ndash;10.0) require
    remediation within 15 days. High (VPR 7.0&ndash;8.9) within 30 days.
    Lower numbers in Critical and High indicate a stronger security posture.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a single "Vuln Summary" tab with per-severity counts and
        percentage of total.

        Returns ``[]`` on error; writes an error row if ``data.error`` is set.
        """
        tab_name = "Vuln Summary"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            # ---- Headers ----
            headers = ["Severity", "Count", "% of Total"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            m     = data.metrics
            total = m.get("total", 0)

            # ---- Data rows (one per severity) ----
            for row_idx, sev in enumerate(_SEVERITIES, start=2):
                count = m.get(sev, 0)
                pct   = round(count / total * 100, 1) if total > 0 else 0.0

                # Severity label cell — coloured fill
                sev_cell = ws.cell(row=row_idx, column=1, value=sev.capitalize())
                hex_color = SEVERITY_COLORS.get(sev, "#9E9E9E").lstrip("#")
                sev_cell.fill = PatternFill("solid", fgColor=hex_color)
                fg = _TILE_TEXT_COLORS.get(sev, "#000000").lstrip("#")
                sev_cell.font = Font(bold=True, color=fg)

                ws.cell(row=row_idx, column=2, value=count)
                ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")

            # ---- Total row ----
            total_row = len(_SEVERITIES) + 2
            total_cell = ws.cell(row=total_row, column=1, value="Total")
            total_cell.font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=total).font = Font(bold=True)
            ws.cell(row=total_row, column=3, value="100.0%").font = Font(bold=True)

            # ---- Column widths ----
            ws.column_dimensions[get_column_letter(1)].width = 14
            ws.column_dimensions[get_column_letter(2)].width = 10
            ws.column_dimensions[get_column_letter(3)].width = 12

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_kpis
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return Critical, High, Medium, and Total open counts for email tiles.
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m = data.metrics
        return {
            "Critical Open": str(m.get("critical", 0)),
            "High Open":     str(m.get("high", 0)),
            "Medium Open":   str(m.get("medium", 0)),
            "Total Open":    str(m.get("total", 0)),
        }

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Total Open Vulns": (
                    "Count of rows in vulns_df where state in {'open', 'reopened'}, "
                    "excluding findings where vpr_score is null and raw severity == 'info'."
                ),
                "Severity Assignment": (
                    "Derived from VPR score using vpr_to_severity() in config.py: "
                    "Critical = VPR 9.0–10.0, High = VPR 7.0–8.9, "
                    "Medium = VPR 4.0–6.9, Low = VPR 0.1–3.9. "
                    "Findings with no VPR score fall back to Tenable native severity field."
                ),
                "Data Source": (
                    "tio.exports.vulns() bulk export, cached as parquet at data/cache/."
                ),
                "Exclusions": (
                    "Informational findings excluded upstream by the fetcher. "
                    "Fixed/closed findings excluded (state not in {'open', 'reopened'})."
                ),
            },
        }
