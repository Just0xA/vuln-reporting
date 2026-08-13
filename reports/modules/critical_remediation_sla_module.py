"""
reports/modules/critical_remediation_sla_module.py — Critical Vulnerability Remediation SLA.

Measures how much of the Critical remediation workload met its 15-day SLA,
counting BOTH criticals fixed in the last 30 days AND criticals still open
past their 15-day clock.

quick-260805-ezo reformulation
------------------------------
Cohort (D-04)
    ``vpr_severity == "critical"`` (board-local VPR-only tiering — no
    native-CVSS fallback) AND not risk-managed (ACCEPTED / RECASTED).  The
    cohort filter is applied to BOTH the open and the fixed population.

Clock (D-05)
    ``clock_start = COALESCE(resurfaced_date, first_found)``.  Tenable's
    ``time_taken_to_fix`` is deliberately NOT consulted — it measures from the
    original discovery across a reopen and therefore inflates the duration of
    every reopened finding.  Same clock as ``mttr_trend_module`` (D-16-02);
    both modules ship in the same PDF and must agree.

Populations (``rd`` = report_date)
    A  fixed_in_window : ``state == FIXED AND last_fixed >= rd - 30d``
    B  open_past_due   : open/reopened, ``last_found >= rd - 30d``, asset
                         on-time, ``(rd - clock_start).days > 15``
    C  open_not_due    : as B but ``(rd - clock_start).days <= 15``

Scoping asymmetry (D-06)
    The asset-level on-time gate applies to the OPEN side only.  It was
    removed from the FIXED side because it dropped credit for fixes on assets
    that were decommissioned (and therefore stopped being scanned) after the
    remediation landed.  The OPEN side additionally carries a finding-level
    staleness guard (``last_found >= rd - 30d``) so a finding no scanner has
    seen recently cannot be reported as overdue.

Metric (D-07)
    ``compliant   = |A where days_to_fix <= 15|``
    ``fixed_late  = |A where days_to_fix > 15|``
    ``breached    = fixed_late + |B|``
    ``denominator = compliant + breached``
    ``sla_pct     = compliant / denominator * 100``  (None when denominator 0)
    ``total_critical_open = |B| + |C|``

    C is excluded from the denominator: those findings have not yet had their
    chance to meet or miss the SLA.

Module ID:    critical_remediation_sla
Display Name: Critical Vulnerability Remediation SLA

SLA thresholds (board-defined):
    Green:  remediation_sla_pct >= 95%
    Yellow: remediation_sla_pct >= 85%  and < 95%
    Red:    remediation_sla_pct <  85%

Data sources:
    vulns_df         — open / reopened findings (from fetch_all_vulnerabilities)
    fixed_vulns_df   — fixed findings (from fetch_fixed_vulnerabilities),
                       passed via **kwargs["fixed_vulns_df"] by the caller
    assets_df        — full asset inventory (from fetch_all_assets)

Fixed-vuln dependency
---------------------
This module needs BOTH open AND fixed vulnerability data.  The standard
``vulns_df`` parameter contains only open/reopened findings.  Fixed findings
must be supplied by the caller (board_summary.py) as a kwarg::

    composer = ReportComposer(
        vulns_df=open_vulns_df,
        assets_df=assets_df,
        report_date=report_date,
        module_configs=module_configs,
        fixed_vulns_df=fetch_fixed_vulnerabilities(tio, cache_dir),
    )

When ``fixed_vulns_df`` is absent or empty the metric returns ``"no_data"``.
"""

from __future__ import annotations

import html
import logging
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SLA_DAYS
from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from reports.modules.board_pdf_layout import two_column_metric_section
from reports.modules.board_report_utils import (
    add_vpr_severity,
    compute_per_bu_breakdown,
    exclude_risk_managed,
    extract_owner,
    identify_on_time_assets,
    sla_status_from_thresholds,
    ON_TIME_WINDOW_DAYS,
    VPR_NONE_LABEL,
)
from reports.modules.chart_utils import draw_gauge
from reports.modules.format_utils import safe_int, safe_pct
from reports.modules.rag_utils import (
    STATUS_COLOR,
    STATUS_LABEL,
    STATUS_ICON,
    NO_DATA_HEADLINE,
    NO_DATA_DRIVER,
    build_rag_strip_entry,
    rag_status_from_value,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

_GREEN_THRESHOLD  = 95.0   # >= green
_YELLOW_THRESHOLD = 85.0   # >= yellow; < green
_DIRECTION        = "higher_is_better"

# Gauge colour zones: 0–85 red | 85–95 amber | 95–100 green
# WR-04 fix — see scan_coverage_sla_module.py for rationale: gauge amber
# band aligned with rag_utils.STATUS_COLOR['yellow'] (#f57c00) so the
# gauge, status badge, and RAG strip use the same orange.
_GAUGE_THRESHOLDS = [
    (_YELLOW_THRESHOLD, "#d32f2f"),
    (_GREEN_THRESHOLD,  "#f57c00"),
    (100.0,             "#388e3c"),
]

_STATUS_COLOR: dict[str, str] = {
    "green":   "#388e3c",
    "yellow":  "#f57c00",
    "red":     "#d32f2f",
    "no_data": "#757575",
}

_STATUS_LABEL: dict[str, str] = {
    "green":   "On Target",
    "yellow":  "At Risk",
    "red":     "Off Target",
    "no_data": "No Data",
}

# Excel fills (openpyxl RGB, no leading #)
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_YELLOW = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")

# SLA for critical findings (days) — sourced from config.SLA_DAYS
_CRITICAL_SLA_DAYS: int = SLA_DAYS["critical"]   # 15

# quick-260805-ezo — fixed row order for the "VPR Severity Distribution"
# analyst tab. The `none` row is the entire point of the tab, so it is always
# emitted even when the count is zero.
_VPR_DISTRIBUTION_TIERS: tuple[str, ...] = (
    "critical", "high", "medium", "low", VPR_NONE_LABEL,
)

# States that count as "currently open" for the B / C populations.
_OPEN_STATES: frozenset[str] = frozenset({"OPEN", "REOPENED"})


# ===========================================================================
# Module class
# ===========================================================================

@register_module
class CriticalRemediationSLAModule(BaseModule):
    """
    Percentage of Critical vulns fixed within their 15-day SLA in the last 30 days.

    Only findings on assets that were scanned on time (within the last 30 days)
    are included.  This ensures the denominator reflects a recently-validated
    asset population, not stale or decommissioned systems.

    Supported options
    -----------------
    None — this module accepts no configurable options.
    """

    MODULE_ID         = "critical_remediation_sla"
    DISPLAY_NAME      = "Critical Vulnerability Remediation SLA"
    DESCRIPTION       = (
        "Percentage of Critical vulnerabilities fixed within their 15-day SLA "
        "during the last 30 days, scoped to assets scanned on time."
    )
    REQUIRED_DATA     = ["vulns", "assets", "fixed_vulns"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute()
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute Critical remediation SLA compliance across four populations.

        The cohort is VPR-only Critical (``vpr_severity == "critical"``) with
        risk-managed (ACCEPTED / RECASTED) findings excluded from BOTH the
        open and the fixed frame (D-04).  Duration is measured from
        ``COALESCE(resurfaced_date, first_found)`` (D-05).  The denominator
        counts criticals fixed in the last 30 days plus criticals still open
        past their 15-day clock; criticals still inside that clock are
        excluded (D-07).

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Open / reopened findings from fetch_all_vulnerabilities().
            Source of the ``open_past_due`` (B) and ``open_not_due`` (C)
            populations and of the VPR-distribution analyst tab.
        assets_df : pd.DataFrame
            Full asset inventory from fetch_all_assets().
            Used to derive the on-time asset set (OPEN side only) and the
            owner labels for the per-owner breakdown.
        report_date : datetime
            UTC-aware report run timestamp.
        config : ModuleConfig
            Module configuration (no options consumed).
        **kwargs
            ``fixed_vulns_df`` (pd.DataFrame, optional): fixed findings from
            fetch_fixed_vulnerabilities().  Source of the ``compliant`` and
            ``fixed_late`` components.  When absent or empty the metric can
            still be computed from the open side alone; ``"no_data"`` is
            returned only when the denominator is zero.

        Returns
        -------
        ModuleData
            ``error`` is None on success; set to an error string on failure.
        """
        logger.debug(
            "%s compute() — vulns_df=%d rows, assets_df=%d rows",
            self._log_prefix(), len(vulns_df), len(assets_df),
        )

        try:
            fixed_vulns_df: pd.DataFrame = kwargs.get(
                "fixed_vulns_df", pd.DataFrame()
            )
            # Exclude risk-managed (ACCEPTED/RECASTED) findings from BOTH
            # populations up front — they must not inflate/deflate the SLA
            # metric (quick-260722-lx9; extended to the OPEN side by
            # quick-260805-ezo D-04, matching high_risk_assets_module.py and
            # aged_vulns_assets_module.py).  quick-260813-ga2 — gated by the
            # include_risk_managed module option (default False = today's
            # behavior); both populations must never diverge.
            include_risk_managed = bool(
                config.options.get("include_risk_managed", False)
            )
            if not include_risk_managed:
                vulns_df       = exclude_risk_managed(vulns_df)
                fixed_vulns_df = exclude_risk_managed(fixed_vulns_df)

            # quick-260813-ga2 — auditor-facing scope string, precomputed so no
            # inline conditional expression is needed in the f-strings below
            # (Hard Rule 6).
            risk_managed_scope = (
                "Includes risk-managed findings (severity_modification_type in "
                "{ACCEPTED, RECASTED})"
                if include_risk_managed else
                "Excludes risk-managed findings (severity_modification_type in "
                "{ACCEPTED, RECASTED})"
            )

            # quick-260805-ezo D-04 — cohort is the board-local VPR-only
            # Critical tier; a finding with no VPR score is "none", never
            # promoted from its native-CVSS severity string.
            vulns_df       = add_vpr_severity(vulns_df)
            fixed_vulns_df = add_vpr_severity(fixed_vulns_df)

            # Phase 3 — explicit empty-input guard. When assets_df has no
            # rows or lacks required columns we cannot derive an on-time
            # set; bail out with the contract-fields populated.
            if assets_df.empty or "asset_uuid" not in assets_df.columns:
                return ModuleData(
                    module_id    = self.MODULE_ID,
                    display_name = self.DISPLAY_NAME,
                    metrics      = {},
                    table_data   = [],
                    chart_data   = {},
                    summary_text = "",
                    metadata     = {"email_gauge_b64": ""},
                    error            = None,
                    driver_narrative = NO_DATA_DRIVER,
                    analyst_rows     = [],
                    rag_strip        = build_rag_strip_entry(
                        display_name       = self.DISPLAY_NAME,
                        headline_value_str = safe_pct(None),
                        status             = "no_data",
                    ),
                )

            # ---- Step 1: derive on-time asset set ----
            on_time, _ = identify_on_time_assets(assets_df, report_date)
            on_time_uuids = set(on_time["asset_uuid"].dropna())

            # ---- Step 2: build UTC-aware 30-day window ----
            if hasattr(report_date, "tzinfo") and report_date.tzinfo is not None:
                rd_ts = pd.Timestamp(report_date).tz_convert("UTC")
            else:
                rd_ts = pd.Timestamp(report_date, tz="UTC")
            thirty_days_ago = rd_ts - pd.Timedelta(days=ON_TIME_WINDOW_DAYS)

            # ---- Step 3: population A — criticals FIXED in the window ----
            # quick-260805-ezo D-06 — no asset-level on-time gate here. The
            # gate used to drop credit for fixes on assets that stopped being
            # scanned (decommissioned) after the remediation landed.
            fixed_crit = _filter_critical_vpr(fixed_vulns_df)

            if (
                fixed_crit.empty
                or "state" not in fixed_crit.columns
                or "last_fixed" not in fixed_crit.columns
            ):
                fixed_in_window = fixed_crit.iloc[0:0].copy()
            else:
                state_upper   = fixed_crit["state"].astype(str).str.upper()
                last_fixed_ts = _coerce_ts(fixed_crit, "last_fixed")
                fixed_mask    = (
                    (state_upper == "FIXED")
                    & last_fixed_ts.notna()
                    & (last_fixed_ts >= thirty_days_ago)
                )
                fixed_in_window = fixed_crit[fixed_mask].copy()

            # days_to_fix — reopened-aware COALESCE clock (D-05 / D-16-02).
            # Hard Rule 5: .assign() only, never .loc[:, col] = ... on a slice.
            if fixed_in_window.empty:
                fixed_in_window = fixed_in_window.assign(
                    days_to_fix=pd.Series(dtype="float64")
                )
            else:
                days_to_fix = (
                    _coerce_ts(fixed_in_window, "last_fixed")
                    - _compute_clock_start(fixed_in_window)
                ).dt.days.clip(lower=0)
                fixed_in_window = fixed_in_window.assign(days_to_fix=days_to_fix)

            dtf              = fixed_in_window["days_to_fix"]
            compliant_mask   = dtf.notna() & (dtf <= _CRITICAL_SLA_DAYS)
            fixed_late_mask  = dtf.notna() & (dtf > _CRITICAL_SLA_DAYS)
            compliant        = int(compliant_mask.sum())
            fixed_late       = int(fixed_late_mask.sum())

            # ---- Step 4: populations B and C — criticals still OPEN ----
            open_in_scope = _filter_open_in_scope(
                _filter_critical_vpr(vulns_df), on_time_uuids, thirty_days_ago,
            )

            if open_in_scope.empty:
                open_age_days   = pd.Series(dtype="float64")
                past_due_mask   = pd.Series(dtype=bool)
            else:
                open_age_days = (
                    rd_ts - _compute_clock_start(open_in_scope)
                ).dt.days
                # QT-02 — a NaT clock_start yields NaN age; such a finding is
                # NOT counted as a failure, it falls into C (not yet due) and
                # still appears in Total Critical Open.
                past_due_mask = (
                    open_age_days.notna()
                    & (open_age_days > _CRITICAL_SLA_DAYS)
                )

            open_past_due = int(past_due_mask.sum())
            open_not_due  = int(len(open_in_scope) - open_past_due)

            # ---- Step 5: the metric (D-07 / D-08) ----
            breached            = fixed_late + open_past_due
            denominator         = compliant + breached
            total_critical_open = open_past_due + open_not_due

            if denominator == 0:
                remediation_sla_pct = None
                logger.debug(
                    "%s denominator is zero (no criticals fixed in the window "
                    "and none open past their SLA) — returning no_data.",
                    self._log_prefix(),
                )
            else:
                remediation_sla_pct = round(compliant / denominator * 100, 1)

            # sla_status_from_thresholds() maps None -> "no_data" itself.
            status = sla_status_from_thresholds(
                remediation_sla_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,
            )

            # ---- Step 6: per-owner breakdown over all four components ----
            bu_breakdown = _compute_bu_breakdown(
                fixed_in_window = fixed_in_window,
                compliant_mask  = compliant_mask,
                fixed_late_mask = fixed_late_mask,
                open_in_scope   = open_in_scope,
                past_due_mask   = past_due_mask,
                assets_df       = assets_df,
            )
            table_data = bu_breakdown.to_dict("records")

            # ---- Step 7: narrative summary ----
            summary_text = _build_summary(
                remediation_sla_pct, compliant, fixed_late,
                open_past_due, open_not_due, denominator, status,
            )

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            # ===== Phase 3 contract fields =====

            # Phase 3 D-10/D-11 — analyst rows for Critical Remediation SLA.
            # quick-260805-ezo — the detail tab now unions BOTH kinds of
            # missed-SLA finding: A rows fixed late AND B rows still open past
            # their clock. D-13 — finding-level rows; NO dedup.
            analyst_rows_payload: list = []

            missed_detail_df = _build_missed_detail(
                fixed_late_rows = fixed_in_window[fixed_late_mask],
                open_past_due_rows = (
                    open_in_scope[past_due_mask]
                    if not open_in_scope.empty
                    else open_in_scope
                ),
                rd_ts = rd_ts,
            )
            if not missed_detail_df.empty:
                analyst_rows_payload.append(
                    ("Critical Remediation Detail", missed_detail_df)
                )

            # quick-260805-ezo D-10 — VPR distribution of the OPEN population.
            # Scope: on-time assets + risk-managed excluded ONLY; the 30-day
            # finding-level staleness guard is deliberately NOT applied here so
            # the tab reconciles against the operator's console counts. The
            # `none` row is the entire point of the tab, so it is emitted even
            # when the detail tab above is empty.
            analyst_rows_payload.append((
                "VPR Severity Distribution",
                _build_vpr_distribution(vulns_df, on_time_uuids),
            ))

            # Item 3 — driver narrative disclosing all four components.
            if denominator > 0 or total_critical_open > 0:
                driver = (
                    f"{safe_int(compliant)} of {safe_int(denominator)} criticals "
                    f"met the {_CRITICAL_SLA_DAYS}-day SLA; "
                    f"{safe_int(fixed_late)} fixed late, "
                    f"{safe_int(open_past_due)} still overdue, "
                    f"{safe_int(open_not_due)} not yet due."
                )
            else:
                driver = NO_DATA_DRIVER

            # Phase 3 D-04 — email gauge base64
            # W5 — `remediation_sla_pct` is the verified local in compute() scope.
            if remediation_sla_pct is not None:
                try:
                    email_gauge_b64 = draw_gauge(
                        value      = remediation_sla_pct,
                        thresholds = _GAUGE_THRESHOLDS,
                        title      = self.DISPLAY_NAME,
                        unit       = "%",
                        figsize    = (2.4, 1.6),
                    )
                except Exception as exc:  # noqa: BLE001
                    logger.warning(
                        "%s email gauge render failed: %s",
                        self._log_prefix(), exc,
                    )
                    email_gauge_b64 = ""
            else:
                email_gauge_b64 = ""

            # Phase 3 D-05/D-08/D-09 — RAG strip cell (option-2 pure construction)
            _status_for_strip = rag_status_from_value(
                remediation_sla_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,
            )
            rag_strip_payload = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = safe_pct(remediation_sla_pct),
                status             = _status_for_strip,
            )

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                # quick-260805-ezo QT-01 — the four-component metrics contract.
                # total_open_last_month / total_fixed_last_month /
                # fixed_within_sla are REMOVED (no external reader).
                metrics      = {
                    "remediation_sla_pct": remediation_sla_pct,
                    "compliant":           compliant,
                    "fixed_late":          fixed_late,
                    "open_past_due":       open_past_due,
                    "open_not_due":        open_not_due,
                    "breached":            breached,
                    "denominator":         denominator,
                    "total_critical_open": total_critical_open,
                    "status":              status,
                },
                table_data   = table_data,
                chart_data   = {
                    "value":      remediation_sla_pct,
                    "thresholds": {
                        "green":  _GREEN_THRESHOLD,
                        "yellow": _YELLOW_THRESHOLD,
                    },
                    "direction":  _DIRECTION,
                    "top_5":      bu_breakdown.head(5).to_dict("records"),
                },
                summary_text = summary_text,
                metadata     = {
                    # quick-260805-ezo — window / scope / clock strings rewritten.
                    "window": (
                        f"Fixed: last_fixed >= report_date − {ON_TIME_WINDOW_DAYS}d; "
                        f"Open: last_found >= report_date − {ON_TIME_WINDOW_DAYS}d "
                        "(finding-level staleness guard)"
                    ),
                    "critical_sla_days": _CRITICAL_SLA_DAYS,
                    "sla_source":        (
                        "Board-defined thresholds "
                        f"(Green ≥{_GREEN_THRESHOLD}%, "
                        f"Amber ≥{_YELLOW_THRESHOLD}%, Red <{_YELLOW_THRESHOLD}%)"
                    ),
                    "on_time_scope": (
                        "The asset-level on-time gate (last_licensed_scan_date "
                        f">= report_date − {ON_TIME_WINDOW_DAYS} days) applies to "
                        "the OPEN population only; the FIXED population carries "
                        "no asset-level gate."
                    ),
                    "days_to_fix_source": (
                        "(last_fixed − COALESCE(resurfaced_date, first_found)).days, "
                        "clipped at 0; time_taken_to_fix is deliberately NOT used"
                    ),
                    "computed_at":         computed_at,
                    "email_gauge_b64":     email_gauge_b64,
                    "risk_managed_scope":  risk_managed_scope,
                },
                error            = None,
                driver_narrative = driver,
                analyst_rows     = analyst_rows_payload,
                rag_strip        = rag_strip_payload,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            empty = self._empty_result(str(exc), config)
            # Phase 3 — guarantee the email_gauge_b64 metadata key is present
            # (composer.collect_email_inline_images probes for "" vs missing).
            empty.metadata = {**(empty.metadata or {}), "email_gauge_b64": ""}
            return empty

    # ------------------------------------------------------------------
    # render_pdf_section()
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a full-page PDF section for WeasyPrint.

        Layout (two-column row via ``two_column_metric_section`` to avoid the
        page bleed a stacked explanation caused):
        - Section heading (full width)
        - Left column: gauge with 95 / 85 colour zones, status badge, and the
          five bold support numbers disclosing every metric component
          (Compliant | Fixed late | Still open, overdue | Not yet due |
          Total Critical Open) — quick-260805-ezo Item 3.  They stay in ONE
          table row: a second row re-introduces the page bleed this layout
          was created to fix.
        - Right column: explanatory paragraph (left-aligned)
        - Below, full width: Top-5 worst-performing BUs table
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m                   = data.metrics
        sla_pct             = m.get("remediation_sla_pct")
        compliant           = m.get("compliant", 0)
        fixed_late          = m.get("fixed_late", 0)
        open_past_due       = m.get("open_past_due", 0)
        open_not_due        = m.get("open_not_due", 0)
        total_critical_open = m.get("total_critical_open", 0)
        status              = m.get("status", "no_data")

        # ---- Gauge ----
        gauge_value = sla_pct if sla_pct is not None else 0.0
        try:
            gauge_b64  = draw_gauge(
                value      = gauge_value,
                min_val    = 0,
                max_val    = 100,
                thresholds = _GAUGE_THRESHOLDS,
                title      = "Critical Remediation SLA %",
                unit       = "%",
                figsize    = (5, 3),
            )
            gauge_html = (
                f'<div style="text-align:center; margin-bottom:4mm;">'
                f'<img src="data:image/png;base64,{gauge_b64}" '
                f'style="width:46%; max-width:320px;" '
                f'alt="Critical Remediation SLA gauge">'
                f'</div>'
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("%s PDF gauge render failed: %s", self._log_prefix(), exc)
            pct_display = f"{gauge_value:.1f}%"  # safe: gauge_value guaranteed non-None by line 343 (None coalesced to 0.0)
            gauge_html = (
                f'<p style="text-align:center; font-size:20pt; font-weight:bold; '
                f'color:{_STATUS_COLOR.get(status, "#333")};">{pct_display}</p>'
            )

        # ---- Status badge ----
        status_color = _STATUS_COLOR.get(status, "#757575")
        status_label = _STATUS_LABEL.get(status, status)
        pct_display  = (
            f"{sla_pct:.1f}%" if sla_pct is not None else "N/A"
        )
        status_html = (
            f'<p style="text-align:center; font-size:10pt; font-weight:bold; '
            f'color:{status_color}; margin:0 0 5mm 0;">'
            f'SLA Compliance: {pct_display} &nbsp;&middot;&nbsp; {status_label}'
            f'</p>'
        )

        # ---- Five support numbers (quick-260805-ezo Item 3) ----
        # ONE row, narrower cells + smaller type. A second row re-creates the
        # page bleed recorded in the module docstring.
        _cell_css = (
            "text-align:center; padding:1.5mm 2mm; vertical-align:middle;"
        )
        _div_css  = _cell_css + " border-right:0.5pt solid #ddd;"
        _num_css  = "font-size:11pt; font-weight:bold;"
        _lbl_css  = "font-size:6pt; color:#777;"
        support_html = f"""
<table style="width:100%; margin:0 auto 6mm auto; border-collapse:collapse;">
  <tr>
    <td style="{_div_css}">
      <span style="{_num_css} color:#388e3c;">{compliant:,}</span>
      <br><span style="{_lbl_css}">Compliant (&le;{_CRITICAL_SLA_DAYS}d)</span>
    </td>
    <td style="{_div_css}">
      <span style="{_num_css} color:#d32f2f;">{fixed_late:,}</span>
      <br><span style="{_lbl_css}">Fixed late (&gt;{_CRITICAL_SLA_DAYS}d)</span>
    </td>
    <td style="{_div_css}">
      <span style="{_num_css} color:#d32f2f;">{open_past_due:,}</span>
      <br><span style="{_lbl_css}">Still open, overdue</span>
    </td>
    <td style="{_div_css}">
      <span style="{_num_css} color:#555;">{open_not_due:,}</span>
      <br><span style="{_lbl_css}">Not yet due</span>
    </td>
    <td style="{_cell_css}">
      <span style="{_num_css} color:#1F3864;">{total_critical_open:,}</span>
      <br><span style="{_lbl_css}">Total Critical Open</span>
    </td>
  </tr>
</table>"""

        # ---- Top 5 BU table ----
        top5 = data.chart_data.get("top_5", [])
        if top5:
            rows_html = ""
            for row in top5:
                bu_pct  = float(row.get("percentage", 0.0))
                bu_name = str(row.get("owner", ""))
                bu_num  = int(row.get("numerator",    0))
                bu_den  = int(row.get("denominator",  0))
                row_bg  = _row_bg(bu_pct, _GREEN_THRESHOLD, _YELLOW_THRESHOLD)
                rows_html += (
                    f'<tr style="background:{row_bg};">'
                    f'<td style="padding:1.5mm 3mm;">{bu_name}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_num:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_den:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm; '
                    f'font-weight:bold;">{bu_pct:.1f}%</td>'  # safe: bu_pct non-None per compute_per_bu_breakdown contract (zero-denom BUs excluded)
                    f'</tr>'
                )
            bu_table_html = f"""
<h3 class="subsection-heading">Top 5 Worst-Performing Owners</h3>
<table class="data-table">
  <thead>
    <tr>
      <th>Owner</th>
      <th style="text-align:right;">Compliant</th>
      <th style="text-align:right;">Denominator</th>
      <th style="text-align:right;">SLA Compliance %</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>"""
        else:
            bu_table_html = (
                '<p class="explanatory-text" style="color:#888; font-style:italic;">'
                'No owner breakdown available '
                '(no critical findings fixed in the last 30 days and none open '
                'past their SLA, or assets lack Owner tags).'
                '</p>'
            )

        # ---- Explanatory paragraph ----
        # Pre-format module-level threshold constants (never None) so the
        # multi-line HTML f-string below contains no inline format specs.
        green_str  = format(_GREEN_THRESHOLD,  ".0f")
        yellow_str = format(_YELLOW_THRESHOLD, ".0f")
        # quick-260813-ga2 — the exclusion claim follows the module option
        # instead of being asserted unconditionally.
        include_risk_managed = bool(config.options.get("include_risk_managed", False))
        risk_managed_clause = (
            "including risk-accepted and recast findings"
            if include_risk_managed else
            "excluding risk-accepted and recast findings"
        )
        explain_html = f"""
<p class="explanatory-text">
  <strong>What this measures:</strong> Of the Critical vulnerabilities
  (VPR 9.0&ndash;10.0, {risk_managed_clause}) whose
  {_CRITICAL_SLA_DAYS}-day clock has expired, what percentage were remediated in time?
  The {_CRITICAL_SLA_DAYS}-day clock runs from the resurface date when a finding has
  reopened, otherwise from first discovery.  The denominator counts criticals fixed in
  the last {ON_TIME_WINDOW_DAYS} days PLUS criticals still open past their
  {_CRITICAL_SLA_DAYS}-day SLA; criticals still inside their {_CRITICAL_SLA_DAYS}-day
  clock are excluded because they have not yet had the chance to meet or miss it.
  Open findings are scoped to assets scanned in the last {ON_TIME_WINDOW_DAYS} days AND
  to findings a scanner has actually seen in the last {ON_TIME_WINDOW_DAYS} days.
  Board target is &ge;{green_str}% (green).  &ge;{yellow_str}% is at-risk (amber).
  Below {yellow_str}% is off-target (red).
</p>"""

        return two_column_metric_section(
            heading_html=f'<h2 class="section-heading">{self.DISPLAY_NAME}</h2>',
            left_html=f"{gauge_html}{status_html}{support_html}",
            explanation_html=explain_html,
            full_width_html=bu_table_html,
        )

    # ------------------------------------------------------------------
    # render_excel_tabs()
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write the "Critical Remediation SLA" tab into ``workbook``.

        Returns
        -------
        list[str]
            ``["Critical Remediation SLA"]`` on success, ``[]`` on error.
        """
        tab_name = "Critical Remediation SLA"
        try:
            # D-16 — Phase 3 zero-row standardisation. If both metrics and
            # table_data are empty AND there is no error, emit a single
            # standard placeholder cell at A1 instead of a fully-empty sheet.
            # This is a behavior change from pre-Phase-3 — surface in the SUMMARY.
            empty_metrics = not (
                data.metrics
                and any(v is not None for v in data.metrics.values())
            )
            empty_tables = not data.table_data
            if empty_metrics and empty_tables and not data.error:
                ws = workbook.create_sheet(tab_name)
                ws["A1"] = "No data in scope"
                ws["A1"].font = Font(bold=True, color="666666")
                return [tab_name]

            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            # quick-260805-ezo Item 3 — the KPI block discloses all four
            # components plus Total Critical Open and the denominator.
            m                   = data.metrics
            sla_pct             = m.get("remediation_sla_pct")
            compliant           = m.get("compliant", 0)
            fixed_late          = m.get("fixed_late", 0)
            open_past_due       = m.get("open_past_due", 0)
            open_not_due        = m.get("open_not_due", 0)
            denominator         = m.get("denominator", 0)
            total_critical_open = m.get("total_critical_open", 0)
            status              = m.get("status", "no_data")

            # quick-260813-ga2 — the exclusion claim in the auditor-facing
            # Scope: row follows the module option.
            include_risk_managed = bool(
                config.options.get("include_risk_managed", False)
            )
            risk_managed_clause = (
                "including risk-accepted / recast findings"
                if include_risk_managed else
                "excluding risk-accepted / recast findings"
            )

            # ---- Overall KPI block ----
            _xl_title(ws, "A1", "Critical Vulnerability Remediation SLA — 30-Day Window")

            status_color_hex = _STATUS_COLOR.get(status, "#757575").lstrip("#")
            pct_str = (
                f"{sla_pct:.1f}%" if sla_pct is not None else "N/A"
            )

            _xl_kv(ws, 3, "SLA Compliance (30d):", pct_str,
                   value_font=Font(bold=True, size=12, color=status_color_hex))
            _xl_kv(ws, 4, "Status:",
                   _STATUS_LABEL.get(status, status),
                   value_font=Font(bold=True, color=status_color_hex))
            _xl_kv(ws, 5, f"Compliant (fixed ≤{_CRITICAL_SLA_DAYS}d):",
                   f"{compliant:,}")
            _xl_kv(ws, 6, f"Fixed late (>{_CRITICAL_SLA_DAYS}d):",
                   f"{fixed_late:,}")
            _xl_kv(ws, 7, f"Still open, overdue (>{_CRITICAL_SLA_DAYS}d):",
                   f"{open_past_due:,}")
            _xl_kv(ws, 8,
                   f"Not yet due (≤{_CRITICAL_SLA_DAYS}d) "
                   "(excluded from calculation):",
                   f"{open_not_due:,}")
            _xl_kv(ws, 9, "Total Critical Open:", f"{total_critical_open:,}")
            _xl_kv(ws, 10, "Denominator:", f"{denominator:,}")
            _xl_kv(ws, 11, "Window:",
                   f"Fixed: last_fixed >= report_date − {ON_TIME_WINDOW_DAYS}d  |  "
                   f"Open: last_found >= report_date − {ON_TIME_WINDOW_DAYS}d "
                   "(finding-level staleness guard)")
            _xl_kv(ws, 12, "SLA Thresholds:",
                   f"Green ≥{_GREEN_THRESHOLD:.0f}%  |  "  # safe: module-level int constant, never None
                   f"Amber ≥{_YELLOW_THRESHOLD:.0f}%  |  "  # safe: module-level int constant, never None
                   f"Red <{_YELLOW_THRESHOLD:.0f}%")  # safe: module-level int constant, never None
            _xl_kv(ws, 13, "Scope:",
                   f"Cohort: VPR 9.0–10.0 Critical, {risk_managed_clause}.  "
                   "The asset-level on-time gate "
                   f"(last_licensed_scan_date within {ON_TIME_WINDOW_DAYS}d) "
                   "applies to the OPEN side only.")

            # ---- Owner breakdown table ----
            header_row = 15
            headers = [
                "Owner", "Compliant", "Fixed late", "Still open, overdue",
                "Not yet due", "Denominator", "SLA Compliance %",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell           = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = _FILL_HEADER
                cell.alignment = Alignment(horizontal="center")

            for row_offset, row in enumerate(data.table_data or [], start=1):
                data_row = header_row + row_offset
                bu_pct   = float(row.get("percentage",  0.0))
                bu_fill  = _xl_fill(bu_pct, _GREEN_THRESHOLD, _YELLOW_THRESHOLD)

                ws.cell(row=data_row, column=1,
                        value=str(row.get("owner", ""))).alignment = (
                    Alignment(horizontal="left")
                )
                # quick-260805-ezo — the per-owner component columns.
                # "compliant" and "numerator" carry the same count; the
                # component key is used so the sheet reads consistently with
                # the KPI block above.
                for _col_idx, _key in (
                    (2, "compliant"),
                    (3, "fixed_late"),
                    (4, "open_past_due"),
                    (5, "open_not_due"),
                    (6, "denominator"),
                ):
                    ws.cell(row=data_row, column=_col_idx,
                            value=int(row.get(_key, 0))).alignment = (
                        Alignment(horizontal="right")
                    )
                pct_cell           = ws.cell(row=data_row, column=7,
                                             value=f"{bu_pct:.1f}%")  # safe: bu_pct non-None per compute_per_bu_breakdown contract
                pct_cell.fill      = bu_fill
                pct_cell.font      = Font(bold=True)
                pct_cell.alignment = Alignment(horizontal="center")

            # ---- Column widths ----
            ws.column_dimensions[get_column_letter(1)].width = 32
            ws.column_dimensions[get_column_letter(2)].width = 12
            ws.column_dimensions[get_column_letter(3)].width = 12
            ws.column_dimensions[get_column_letter(4)].width = 20
            ws.column_dimensions[get_column_letter(5)].width = 13
            ws.column_dimensions[get_column_letter(6)].width = 13
            ws.column_dimensions[get_column_letter(7)].width = 18

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel() — Phase 3 D-02 horizontal-split panel
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Build the per-module email panel — horizontal split layout (D-02).

        Layout: 620px-wide table, 150px gauge cell on the left + 430px text
        cell on the right. Inline CSS only. Outlook-safe ``<table>`` shell
        with explicit ``width=""`` attributes per project email conventions.

        Empty-data behavior (D-15): when ``data.error`` or no
        ``email_gauge_b64`` is available, returns the same 620px shell with
        a gray "No data" placeholder block where the gauge would be.

        Returns
        -------
        str
            Inline-CSS HTML fragment. Returns ``""`` only on a render
            exception (caught locally — never raised out).
        """
        try:
            # Empty-data placeholder per D-15
            b64 = (data.metadata or {}).get("email_gauge_b64", "")
            if data.error or not isinstance(b64, str) or not b64.strip():
                return self._render_empty_email_panel()

            pct       = (
                data.metrics.get("remediation_sla_pct")
                if data.metrics else None
            )
            headline  = safe_pct(pct)
            status    = (data.metrics or {}).get("status", "no_data")
            rag_color = STATUS_COLOR.get(status, STATUS_COLOR["no_data"])
            rag_label = STATUS_LABEL.get(status, STATUS_LABEL["no_data"])
            icon      = STATUS_ICON.get(status,  STATUS_ICON["no_data"])

            cid       = f"{self.MODULE_ID}_gauge"
            # T-03-03-01 — html-escape every module-supplied string
            label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
            driver_esc = html.escape(str(data.driver_narrative or ""), quote=True)

            return (
                '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
                'style="width:620px; max-width:620px; margin:8px 0; '
                'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
                '<tr>'
                '  <td width="150" style="padding:12px; vertical-align:middle; '
                '      text-align:center;">'
                f'    <img src="cid:{cid}" alt="" width="120" height="120" '
                '         style="display:block; margin:0 auto;" />'
                '  </td>'
                '  <td width="430" style="padding:12px; vertical-align:middle;">'
                f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
                f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{headline}</div>'
                f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
                f'{icon} {html.escape(rag_label)}</div>'
                f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
                f'{driver_esc}</div>'
                '  </td>'
                '</tr>'
                '</table>'
            )
        except Exception as exc:    # noqa: BLE001
            logger.error(
                "%s render_email_panel raised: %s",
                self._log_prefix() if hasattr(self, "_log_prefix") else self.MODULE_ID,
                exc,
            )
            return ""

    def _render_empty_email_panel(self) -> str:
        """Return the D-15 gray 'No Data' placeholder panel (620px wide)."""
        label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
        driver_esc = html.escape(NO_DATA_DRIVER, quote=True)
        rag_color  = STATUS_COLOR["no_data"]
        rag_label  = STATUS_LABEL["no_data"]
        icon       = STATUS_ICON["no_data"]
        return (
            '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
            'style="width:620px; max-width:620px; margin:8px 0; '
            'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
            '<tr>'
            '  <td width="150" style="padding:12px; vertical-align:middle; '
            '      text-align:center; background:#f5f5f5; color:#999;">'
            '    <div style="font-size:10pt;">No data</div>'
            '  </td>'
            '  <td width="430" style="padding:12px; vertical-align:middle;">'
            f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
            f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{NO_DATA_HEADLINE}</div>'
            f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
            f'{icon} {html.escape(rag_label)}</div>'
            f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
            f'{driver_esc}</div>'
            '  </td>'
            '</tr>'
            '</table>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs() — Phase 3 D-14 single-tab list
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, "pd.DataFrame"]]:
        """
        Return the module's pre-built analyst rows (D-14 single-tab list).

        Source: ``data.analyst_rows`` populated inside ``compute()``.
        Empty-data: returns ``[]`` (no tab written for this module).
        """
        try:
            if data.error or not data.analyst_rows:
                return []
            return list(data.analyst_rows)
        except Exception as exc:   # noqa: BLE001
            logger.error("[%s] render_analyst_tabs raised: %s", self.MODULE_ID, exc)
            return []

    # ------------------------------------------------------------------
    # render_email_kpis()
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return three KPI tiles for the HTML email body.

        quick-260805-ezo — repointed onto the QT-01 metric keys. The
        CONTRACT-01 email panel and the CONTRACT-03 RAG strip are unchanged.

        Returns
        -------
        dict[str, str]
            Keys: "Crit Remediation SLA", "Crit Compliant (30d)",
            "Crit Breached".
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m   = data.metrics
        pct = m.get("remediation_sla_pct")
        return {
            "Crit Remediation SLA":  f"{pct:.1f}%" if pct is not None else "N/A",
            "Crit Compliant (30d)":  f"{m.get('compliant', 0):,}",
            "Crit Breached":         f"{m.get('breached', 0):,}",
        }

    # ------------------------------------------------------------------
    # get_audit_info()
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            # quick-260805-ezo — every entry rewritten for the four-component
            # formula, the reopened-aware clock, and the asymmetric scoping.
            "calculations": {
                "cohort": (
                    "vpr_severity == 'critical' (board-local VPR-only tiering, "
                    "VPR 9.0–10.0, NO native-CVSS fallback) AND "
                    "severity_modification_type NOT IN (ACCEPTED, RECASTED). "
                    "Applied to BOTH the open and the fixed population "
                    "(quick-260805-ezo D-04; previously fixed-only) — UNLESS the "
                    "include_risk_managed module option is set, in which case "
                    "the severity_modification_type filter is skipped and the "
                    "full population (including ACCEPTED/RECASTED) is used "
                    "(quick-260813-ga2)."
                ),
                "days_to_fix": (
                    "(last_fixed − clock_start).days, clipped at 0, where "
                    "clock_start = COALESCE(resurfaced_date, first_found). "
                    "Tenable's time_taken_to_fix field is deliberately NOT used: "
                    "it measures from the original discovery straight through a "
                    "reopen and inflates the duration of every reopened finding. "
                    "Same clock as mttr_trend_module (D-16-02, quick-260805-ezo D-05)."
                ),
                "compliant": (
                    f"Count of fixed_in_window rows with days_to_fix <= "
                    f"{_CRITICAL_SLA_DAYS}. fixed_in_window = cohort rows where "
                    f"state == 'FIXED' AND last_fixed >= report_date − "
                    f"{ON_TIME_WINDOW_DAYS} days."
                ),
                "fixed_late": (
                    f"Count of fixed_in_window rows with days_to_fix > "
                    f"{_CRITICAL_SLA_DAYS}. Rows whose days_to_fix cannot be "
                    "computed (NaT clock_start or NaT last_fixed) fall into "
                    "neither compliant nor fixed_late and therefore never enter "
                    "the denominator."
                ),
                "open_past_due": (
                    f"Count of in-scope open/reopened cohort rows where "
                    f"(report_date − clock_start).days > {_CRITICAL_SLA_DAYS}."
                ),
                "open_not_due": (
                    f"Count of in-scope open/reopened cohort rows where "
                    f"(report_date − clock_start).days <= {_CRITICAL_SLA_DAYS}, "
                    "plus rows whose clock_start is NaT (QT-02 — a finding whose "
                    "overdue-ness cannot be computed is not counted as a failure). "
                    "EXCLUDED from the denominator: these findings have not yet "
                    "had the chance to meet or miss the SLA."
                ),
                "breached": "fixed_late + open_past_due.",
                "denominator": "compliant + breached.",
                "remediation_sla_pct": (
                    "compliant / denominator × 100, rounded to 1 decimal. "
                    "None when denominator == 0 (status → 'no_data'), even when "
                    "open_not_due > 0."
                ),
                "total_critical_open": (
                    "open_past_due + open_not_due — the in-scope open Critical "
                    "population, disclosed alongside the percentage so the board "
                    "sees the absolute workload behind the ratio."
                ),
                "open_scope": (
                    "TWO gates. (1) Asset-level on-time gate: "
                    "last_licensed_scan_date IS NOT NULL AND >= report_date − "
                    f"{ON_TIME_WINDOW_DAYS} days. (2) Finding-level staleness "
                    f"guard: last_found >= report_date − {ON_TIME_WINDOW_DAYS} "
                    "days, so a finding no scanner has actually seen recently "
                    "cannot be reported as overdue."
                ),
                "fixed_scope": (
                    f"last_fixed >= report_date − {ON_TIME_WINDOW_DAYS} days. "
                    "NO asset-level on-time gate (quick-260805-ezo D-06): the "
                    "gate dropped credit for fixes on assets that were "
                    "decommissioned — and therefore stopped being scanned — "
                    "after the remediation landed."
                ),
                "owner_breakdown": (
                    "compute_per_bu_breakdown(higher_is_better=True) via "
                    "_compute_bu_breakdown() over the union of the fixed-in-window, "
                    "open-past-due and open-not-due rows. Per-owner: numerator = "
                    "compliant; denominator = compliant + fixed_late + "
                    "open_past_due (open_not_due contributes its component count "
                    "but never the denominator). Owner is mapped from the Owner "
                    "tag on the FULL asset frame — the fixed side has no "
                    "asset-level gate. affected = denominator − numerator. "
                    "Primary sort: affected DESC (most missed-SLA criticals first). "
                    "Secondary sort: percentage ASC (worst compliance % among ties)."
                ),
            },
        }


# ===========================================================================
# Module-private helpers
# ===========================================================================


def _coerce_ts(df: pd.DataFrame, column: str) -> "pd.Series":
    """
    Return ``df[column]`` coerced to ``datetime64[ns, UTC]``, NaT-safe.

    Mirrors the defensive pattern in ``mttr_trend_module.py`` (D-16-02,
    Pitfall A): when the column is absent an all-NaT object Series of the
    right length is coerced instead, so the caller can always subtract two
    tz-aware series without a KeyError or a tz-mismatch.
    """
    source = (
        df[column]
        if column in df.columns
        else pd.Series([pd.NaT] * len(df), index=df.index, dtype="object")
    )
    return pd.to_datetime(source, utc=True, errors="coerce")


def _compute_clock_start(df: pd.DataFrame) -> "pd.Series":
    """
    Return the reopened-aware SLA clock start for each row.

    ``clock_start = COALESCE(resurfaced_date, first_found)`` (D-05).

    quick-260805-ezo — Tenable's ``time_taken_to_fix`` is deliberately NOT
    consulted: it measures from the ORIGINAL discovery straight through a
    reopen, so a finding that was fixed, resurfaced, and re-fixed within days
    is reported as months old. Both this module and ``mttr_trend_module``
    (D-16-02) ship in the same board PDF and must share one clock.

    Parameters
    ----------
    df : pd.DataFrame
        Findings frame. ``resurfaced_date`` and/or ``first_found`` may be
        absent — each missing column is treated as all-NaT.

    Returns
    -------
    pd.Series
        ``datetime64[ns, UTC]`` series aligned with ``df.index``. NaT where
        neither date is available.
    """
    resurfaced_ts  = _coerce_ts(df, "resurfaced_date")
    first_found_ts = _coerce_ts(df, "first_found")
    return resurfaced_ts.where(resurfaced_ts.notna(), other=first_found_ts)


def _filter_critical_vpr(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return the VPR-Critical rows of ``df`` — cohort filter only, no asset gate.

    quick-260805-ezo D-04/D-06 — the cohort is ``vpr_severity == "critical"``
    (board-local VPR-only tiering, no native-CVSS fallback). No asset-level
    scoping is applied here so the FIXED and OPEN sides can differ: the caller
    layers the on-time asset gate onto the OPEN side only.

    Returns an empty frame when ``df`` is empty or lacks the required columns
    (fail-soft — no exception).
    """
    if df.empty:
        return df.copy()

    if "asset_uuid" not in df.columns or "vpr_severity" not in df.columns:
        logger.warning(
            "_filter_critical_vpr: missing asset_uuid / vpr_severity column "
            "— returning an empty frame."
        )
        return pd.DataFrame()

    return df[df["vpr_severity"] == "critical"].copy()


def _filter_open_in_scope(
    open_crit:       pd.DataFrame,
    on_time_uuids:   set,
    thirty_days_ago: "pd.Timestamp",
) -> pd.DataFrame:
    """
    Scope the VPR-Critical open cohort down to populations B ∪ C.

    Two gates apply on the OPEN side (D-06):

    1. **Asset-level on-time gate** — ``asset_uuid`` must be in the on-time
       scanned set.
    2. **Finding-level staleness guard** — ``last_found >= report_date − 30d``,
       so a finding no scanner has actually seen recently cannot be reported
       as overdue.

    Fail-soft: an absent ``state`` column means the frame came from the
    open-findings fetch, so every row is treated as open. An absent
    ``last_found`` column means there is no staleness signal at all — a
    warning is logged and the guard is skipped rather than silently zeroing
    the metric. A *present* ``last_found`` that is NaT excludes the row (its
    freshness cannot be established).
    """
    if open_crit.empty:
        return open_crit.copy()

    if "state" in open_crit.columns:
        state_mask = (
            open_crit["state"].astype(str).str.upper().isin(_OPEN_STATES)
        )
    else:
        state_mask = pd.Series(True, index=open_crit.index)

    if "last_found" in open_crit.columns:
        last_found_ts = _coerce_ts(open_crit, "last_found")
        fresh_mask    = last_found_ts.notna() & (last_found_ts >= thirty_days_ago)
    else:
        logger.warning(
            "_filter_open_in_scope: column 'last_found' not present — the "
            "finding-level staleness guard is skipped for this run."
        )
        fresh_mask = pd.Series(True, index=open_crit.index)

    mask = (
        state_mask
        & open_crit["asset_uuid"].isin(on_time_uuids)
        & fresh_mask
    )
    return open_crit[mask].copy()


def _build_missed_detail(
    fixed_late_rows:    pd.DataFrame,
    open_past_due_rows: pd.DataFrame,
    rd_ts:              "pd.Timestamp",
) -> pd.DataFrame:
    """
    Build the "Critical Remediation Detail" analyst tab (A-late ∪ B).

    Both populations "missed the 15-day SLA", so they share one tab
    (quick-260805-ezo Item 4):

    - **A rows (fixed late):** ``days overdue = days_to_fix − 15``
    - **B rows (still open):** ``days overdue = (report_date − clock_start).days − 15``

    ``remediation due_date`` is derived from ``clock_start + 15d`` for BOTH
    populations — not ``first_found + 15d``, because the clock now restarts on
    a resurface (D-05).

    Returns an empty DataFrame when neither population has rows.
    """
    def _project(rows: pd.DataFrame, days_overdue: "pd.Series") -> pd.DataFrame:
        """
        Reduce a population to the columns the tab actually renders.

        [Rule 1] quick-260805-ezo — projecting BEFORE the concat keeps the two
        populations' column sets identical and drops the irrelevant all-NA
        columns (``resurfaced_date`` on a never-reopened fixed frame,
        ``last_fixed`` on the open frame) that made ``pd.concat`` emit the
        "concatenation with empty or all-NA entries is deprecated" FutureWarning.
        """
        return pd.DataFrame({
            "hostname":      rows.get("hostname",    pd.Series(index=rows.index, dtype="object")),
            "plugin_name":   rows.get("plugin_name", pd.Series(index=rows.index, dtype="object")),
            "plugin_id":     rows.get("plugin_id",   pd.Series(index=rows.index, dtype="object")),
            "tags":          rows.get("tags",        pd.Series(index=rows.index, dtype="object")),
            "first_found":   _coerce_ts(rows, "first_found"),
            "_clock_start":  _compute_clock_start(rows),
            "_days_overdue": days_overdue.astype("Float64"),
        })

    frames: list[pd.DataFrame] = []

    if not fixed_late_rows.empty:
        frames.append(_project(
            fixed_late_rows,
            fixed_late_rows["days_to_fix"] - _CRITICAL_SLA_DAYS,
        ))

    if not open_past_due_rows.empty:
        frames.append(_project(
            open_past_due_rows,
            (rd_ts - _compute_clock_start(open_past_due_rows)).dt.days
            - _CRITICAL_SLA_DAYS,
        ))

    if not frames:
        return pd.DataFrame(columns=[
            "asset", "plugin", "days overdue", "first_found",
            "owner_tag", "remediation due_date",
        ])

    missed = pd.concat(frames, ignore_index=True)

    # Derive owner_tag for each finding row via extract_owner. extract_owner
    # operates on an assets-style DataFrame; findings carry a 'tags' column in
    # the same "Cat=Val;Cat=Val" format, so passing 'missed' directly produces
    # an 'owner' column we rename to 'owner_tag' for the drill-down display.
    missed_with_owner = extract_owner(missed)
    missed = missed.assign(
        owner_tag=missed_with_owner["owner"].values,
        plugin=missed.apply(
            lambda r: f"{r.get('plugin_name', '')} ({r.get('plugin_id', '')})",
            axis=1,
        ),
        **{
            "days overdue": (
                missed["_days_overdue"].clip(lower=0).round().astype("Int64")
            ),
            "remediation due_date": (
                missed["_clock_start"] + pd.Timedelta(days=_CRITICAL_SLA_DAYS)
            ),
        },
    )

    analyst_df = missed.reindex(columns=[
        "hostname",
        "plugin",
        "days overdue",
        "first_found",
        "owner_tag",
        "remediation due_date",
    ]).rename(columns={"hostname": "asset"})

    # D-11 — sort by days overdue desc
    analyst_df = analyst_df.sort_values(
        "days overdue", ascending=False, na_position="last",
    ).reset_index(drop=True)

    # T-03-03-02 / T-ezo-02 — CSV-formula injection guard on the text columns.
    # This runs AFTER the A ∪ B union so the B rows are guarded too.
    #
    # [Rule 1] quick-260805-ezo — built via .assign() rather than
    # `analyst_df.loc[:, col] = ...` (Hard Rule 5). When a source frame lacks
    # `hostname`, reindex produces an all-NaN float64 `asset` column and the
    # .loc setter fired a pandas "Setting an item of incompatible dtype is
    # deprecated" FutureWarning that becomes an error under pandas 3.0.
    def _guard(s):
        return (
            ("'" + s)
            if isinstance(s, str) and s[:1] in ("=", "+", "-", "@")
            else s
        )

    return analyst_df.assign(**{
        _col: analyst_df[_col].astype("string").map(_guard)
        for _col in ("asset", "plugin", "owner_tag")
    })


def _build_vpr_distribution(
    vulns_df:      pd.DataFrame,
    on_time_uuids: set,
) -> pd.DataFrame:
    """
    Build the "VPR Severity Distribution" analyst tab.

    Counts OPEN findings by board-local ``vpr_severity`` tier, scoped to
    on-time assets with risk-managed findings already excluded. The 30-day
    finding-level staleness guard is deliberately NOT applied — this scope is
    what reconciles against the operator's console counts.

    Returns
    -------
    pd.DataFrame
        Columns ``VPR Severity`` / ``Open Findings``; six rows in fixed order
        (critical, high, medium, low, none, TOTAL), zero-filled where a tier
        is absent. All values are integers and all labels are fixed literals,
        so no CSV-formula-injection guard is required (T-ezo-01).
    """
    counts = {tier: 0 for tier in _VPR_DISTRIBUTION_TIERS}

    if (
        not vulns_df.empty
        and "vpr_severity" in vulns_df.columns
        and "asset_uuid" in vulns_df.columns
    ):
        in_scope = vulns_df[vulns_df["asset_uuid"].isin(on_time_uuids)]
        observed = in_scope["vpr_severity"].value_counts()
        for tier in _VPR_DISTRIBUTION_TIERS:
            counts[tier] = int(observed.get(tier, 0))

    rows = [
        {"VPR Severity": tier, "Open Findings": counts[tier]}
        for tier in _VPR_DISTRIBUTION_TIERS
    ]
    rows.append({
        "VPR Severity":  "TOTAL",
        "Open Findings": int(sum(counts.values())),
    })
    return pd.DataFrame(rows, columns=["VPR Severity", "Open Findings"])


def _compute_bu_breakdown(
    fixed_in_window: pd.DataFrame,
    compliant_mask:  "pd.Series[bool]",
    fixed_late_mask: "pd.Series[bool]",
    open_in_scope:   pd.DataFrame,
    past_due_mask:   "pd.Series[bool]",
    assets_df:       pd.DataFrame,
) -> pd.DataFrame:
    """
    Build the per-owner breakdown across all four metric components.

    quick-260805-ezo — the frame is the union of A (fixed in window), B (open
    past due) and C (open, not yet due). Each row carries four boolean
    component flags. The denominator mask deliberately EXCLUDES C, matching
    the headline formula (D-07): C rows contribute to their component count
    but never to the SLA percentage.

    Owner is mapped from the FULL asset frame, not just the on-time subset —
    the fixed side no longer carries an asset-level gate (D-06), so its rows
    can belong to assets that are not on-time-scanned.

    Returns
    -------
    pd.DataFrame
        ``compute_per_bu_breakdown`` output (owner / numerator / denominator /
        percentage / affected) plus per-owner sums of the four components:
        ``compliant`` / ``fixed_late`` / ``open_past_due`` / ``open_not_due``.
        Empty (with the full column set) when neither population has rows.
    """
    component_cols = [
        "compliant", "fixed_late", "open_past_due", "open_not_due",
    ]
    empty_columns = [
        "owner", "numerator", "denominator", "percentage", "affected",
    ] + component_cols

    frames: list[pd.DataFrame] = []

    if not fixed_in_window.empty:
        frames.append(pd.DataFrame({
            "asset_uuid":    fixed_in_window["asset_uuid"].values,
            "compliant":     compliant_mask.reindex(
                fixed_in_window.index, fill_value=False).values,
            "fixed_late":    fixed_late_mask.reindex(
                fixed_in_window.index, fill_value=False).values,
            "open_past_due": False,
            "open_not_due":  False,
        }))

    if not open_in_scope.empty:
        past_due = past_due_mask.reindex(
            open_in_scope.index, fill_value=False
        ).values
        frames.append(pd.DataFrame({
            "asset_uuid":    open_in_scope["asset_uuid"].values,
            "compliant":     False,
            "fixed_late":    False,
            "open_past_due": past_due,
            "open_not_due":  ~past_due,
        }))

    if not frames:
        return pd.DataFrame(columns=empty_columns)

    fw = pd.concat(frames, ignore_index=True)

    enriched_assets = extract_owner(assets_df)
    uuid_to_owner   = dict(
        zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
    )
    fw = fw.assign(
        owner=fw["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
    )

    # Masks are derived from fw itself, so they are index-aligned by
    # construction; compute_per_bu_breakdown still reindexes defensively
    # (the CR-01 guard) so a future refactor cannot silently zero them.
    numerator_mask   = fw["compliant"]
    denominator_mask = (
        fw["compliant"] | fw["fixed_late"] | fw["open_past_due"]
    )

    breakdown = compute_per_bu_breakdown(
        fw, numerator_mask, denominator_mask, higher_is_better=True,
    )

    component_sums = (
        fw.groupby("owner", dropna=False)[component_cols]
        .sum()
        .astype(int)
        .reset_index()
    )
    breakdown = breakdown.merge(component_sums, on="owner", how="left")
    return breakdown.assign(**{
        col: breakdown[col].fillna(0).astype(int) for col in component_cols
    })


def _build_summary(
    sla_pct:       float | None,
    compliant:     int,
    fixed_late:    int,
    open_past_due: int,
    open_not_due:  int,
    denominator:   int,
    status:        str,
) -> str:
    """Build a plain-language narrative sentence for the email body."""
    if sla_pct is None:
        if open_not_due == 0:
            return (
                "No Critical vulnerabilities were fixed in the last 30 days and "
                "none are open past their SLA — remediation SLA compliance "
                "cannot be computed."
            )
        return (
            f"{safe_int(open_not_due)} open Critical findings are still inside "
            f"their {_CRITICAL_SLA_DAYS}-day SLA and none have been fixed or "
            "breached yet — remediation SLA compliance cannot be computed."
        )

    status_label = _STATUS_LABEL.get(status, status)
    # WR-07 fix — use safe_pct() / safe_int() instead of inline f-string
    # format specs on possibly-None values, so a future refactor that breaks
    # the early-return guard above cannot crash this line.
    return (
        f"Critical remediation SLA compliance is {safe_pct(sla_pct)} — "
        f"{safe_int(compliant)} of {safe_int(denominator)} Critical "
        f"vulnerabilities met the {_CRITICAL_SLA_DAYS}-day SLA "
        f"({safe_int(fixed_late)} fixed late, {safe_int(open_past_due)} still "
        f"open past due; {safe_int(open_not_due)} not yet due are excluded). "
        f"Status: {status_label}."
    )


def _row_bg(pct: float, green_threshold: float, yellow_threshold: float) -> str:
    """Light HTML background colour for a BU table row (higher-is-better)."""
    if pct >= green_threshold:
        return "#E8F5E9"
    if pct >= yellow_threshold:
        return "#FFF8E1"
    return "#FFEBEE"


def _xl_fill(
    pct: float,
    green_threshold: float,
    yellow_threshold: float,
) -> PatternFill:
    """openpyxl PatternFill for a SLA-% cell (higher-is-better)."""
    if pct >= green_threshold:
        return _FILL_GREEN
    if pct >= yellow_threshold:
        return _FILL_YELLOW
    return _FILL_RED


def _xl_title(ws, cell_ref: str, value: str) -> None:
    ws[cell_ref]      = value
    ws[cell_ref].font = Font(bold=True, size=12)


def _xl_kv(ws, row: int, label: str, value: str,
            value_font: Font | None = None) -> None:
    lc      = ws.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True)
    vc      = ws.cell(row=row, column=2, value=value)
    if value_font is not None:
        vc.font = value_font
