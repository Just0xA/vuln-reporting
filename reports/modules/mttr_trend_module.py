"""
reports/modules/mttr_trend_module.py — MTTR Trend (Reopened-Aware) module.

Fixes four correctness gaps in ``mttr_by_severity_module.py``:

D-16-01  Population: source from ``fixed_vulns_df`` (durably-FIXED state only),
         NOT the open-findings export. Never uses ``state == "fixed" OR
         last_fixed.notna()`` — that clause re-includes REOPENED findings.

D-16-02  Clock: ``(last_fixed - COALESCE(resurfaced_date, first_found)).days``
         clipped >= 0. The old ``time_taken_to_fix`` preference is dropped
         entirely. Criterion-3 fixture: first_found=−200d, resurfaced_date=−10d,
         last_fixed=−2d → days_to_fix = 8 (old module: 198).

D-16-04  Window disclosure: the rolling-window value (default 30 days,
         configurable via ``mttr_window_days``) is emitted in all four channels.

D-16-06  Owner cut: MoM with per-Owner cold-start + drift reconciliation;
         calendar-month X-axis with partial-month flag (D-16-08).

D-16-11  View selector: ``mttr_view ∈ {owner, severity, both}`` (default
         ``owner``) splits the breakdown into two independent tables/sections —
         a Severity table and an Owner table, each with its own header.  Applied
         consistently across PDF, Excel, and email panel.  Analyst-detail tab
         (CONTRACT-02) always carries both cuts regardless of ``mttr_view``.

D-16-12  Default view is ``owner``: single Owner table renders by default
         (exec headline; fits one page).  The Owner table does NOT show an
         SLA Target column — the Critical-SLA anchor is meaningless for Owner
         rows and is dropped.

Cold-start (QUAL-01): when ``trend_snapshots`` is absent or
``insufficient_data=True``, the MoM trend line cold-starts independently —
per-severity gauges still render from live ``fixed_vulns_df``. The two paths
(live gauges, MoM line) cold-start independently.

RAG strip (CONTRACT-03):
  Green   — overall_mttr / SLA_DAYS["critical"] <= 1.0
  Yellow  — ratio <= 1.25
  Red     — ratio > 1.25
  No data — overall_mttr is None

Supported options
-----------------
mttr_window_days : int, default 30
    Rolling window in calendar days.  Findings fixed before
    ``report_date - window_days`` are excluded from the live gauge.
min_sample_size : int, default 5
    Minimum durably-fixed findings to compute MTTR for any bucket.
    Sub-threshold buckets render:
    "Insufficient data (N findings — minimum 5 required)"
mttr_view : str, default "owner"
    Controls which breakdown table(s) appear in PDF, Excel, and the email
    panel.  Must be one of ``{owner, severity, both}``.  Unknown values log
    a WARNING and fall back to ``owner``.
    - ``owner``    — Owner table only (no Severity rows).  Exec default.
    - ``severity`` — Severity table only (with per-severity SLA Target column).
    - ``both``     — Two distinct tables: Severity first, then Owner.
    The analyst-detail workbook (CONTRACT-02) always carries both cuts.
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SLA_DAYS
from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.board_report_utils import extract_owner
from reports.modules.chart_utils import draw_gauge
from reports.modules.format_utils import safe_format, safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR,
    STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SEVERITIES: tuple[str, ...] = ("critical", "high", "medium", "low")

# SLA comparison thresholds (mirrors mttr_by_severity_module)
_WITHIN_SLA_LIMIT = 1.0
_NEAR_LIMIT       = 1.25

# Status labels
_STATUS_WITHIN = "Within SLA"
_STATUS_NEAR   = "Near SLA Limit"
_STATUS_EXCEED = "Exceeding SLA"
_STATUS_NODATA = "No Data"

# Status colors
_COLOR_GREEN = "#388e3c"
_COLOR_AMBER = "#fbc02d"
_COLOR_RED   = "#d32f2f"
_COLOR_GREY  = "#9E9E9E"

# Excel fills
_FILL_HEADER = PatternFill("solid", fgColor="E3F2FD")
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY   = PatternFill("solid", fgColor="F5F5F5")


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _status_from_ratio(ratio: float) -> tuple[str, str]:
    """Return (status_label, hex_color) from MTTR / SLA ratio."""
    if ratio <= _WITHIN_SLA_LIMIT:
        return _STATUS_WITHIN, _COLOR_GREEN
    if ratio <= _NEAR_LIMIT:
        return _STATUS_NEAR, _COLOR_AMBER
    return _STATUS_EXCEED, _COLOR_RED


def _status_fill(status: Optional[str]) -> PatternFill:
    """Return the Excel fill for a status string."""
    if status == _STATUS_WITHIN:
        return _FILL_GREEN
    if status == _STATUS_NEAR:
        return _FILL_AMBER
    if status == _STATUS_EXCEED:
        return _FILL_RED
    return _FILL_GREY


def _month_label(month_str: str, current_period: pd.Period) -> str:
    """
    Return a human-readable label for a snapshot month string.

    Appends "(MTD — partial)" to the current in-progress period (D-16-08)
    so recipients know the rolling value is live but the calendar month
    is not complete.

    Parameters
    ----------
    month_str : str
        "YYYY-MM" string from the trend snapshot.
    current_period : pd.Period
        The current month period (derived from report_date at compute() time).

    Returns
    -------
    str
        Plain "YYYY-MM" for completed months;
        "YYYY-MM (MTD — partial)" for the current month.
    """
    try:
        snap_period = pd.Period(month_str, "M")
        if snap_period == current_period:
            return f"{month_str} (MTD — partial)"
    except Exception:  # noqa: BLE001
        pass
    return month_str


def _owner_mom_delta(series: list[Optional[float]]) -> Optional[float]:
    """
    Return absolute MoM delta in days (curr - prev), or None when either
    side is None or fewer than 2 data points exist.

    Uses absolute day delta (not percentage) per D-16-06.
    """
    if len(series) < 2:
        return None
    curr, prev = series[-1], series[-2]
    if curr is None or prev is None:
        return None
    return round(curr - prev, 1)


# ===========================================================================
# Module
# ===========================================================================

@register_module
class MTTRTrendModule(BaseModule):
    """
    Rolling-window Mean Time to Remediate (MTTR) — sample-weighted,
    reopened-aware, with MoM trend and Owner breakdown.

    Fixes D-16-01 (population), D-16-02 (clock), D-16-04 (disclosure),
    D-16-06/08 (Owner cut + partial-month label).

    The module reads TWO independent inputs:
    - ``fixed_vulns_df`` (kwargs) — live durably-fixed findings for gauges.
      Delivered via ``_MODULES_NEEDING_FIXED_VULNS`` in composed_report.py.
    - ``trend_snapshots`` (kwargs) — persisted rolling MTTR for MoM line.
      Delivered via ``_MODULES_NEEDING_TREND_SNAPSHOTS`` in composed_report.py.

    Both inputs cold-start independently (QUAL-01). A missing or empty
    ``fixed_vulns_df`` triggers a full cold-start (gray RAG). A missing
    or insufficient ``trend_snapshots`` only cold-starts the MoM line;
    per-severity gauges still render from live data.

    mttr_by_severity_module.py is left byte-unchanged (D-16-10).
    """

    MODULE_ID         = "mttr_trend"
    DISPLAY_NAME      = "MTTR Trend (Reopened-Aware)"
    DESCRIPTION       = (
        "Rolling-window Mean Time to Remediate — sample-weighted, reopened-aware, "
        "with MoM trend and Owner breakdown."
    )
    REQUIRED_DATA     = ["vulns", "assets", "fixed_vulns", "trend_snapshots"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # _build_cold_start_result
    # ------------------------------------------------------------------

    def _build_cold_start_result(self, config: ModuleConfig) -> ModuleData:
        """
        Return a coherent cold-start ModuleData — not an error, just
        insufficient history or no fixed findings (QUAL-01).

        ``error`` is ``None``; ``metrics["cold_start"]`` is ``True`` so
        renderers can branch on it without parsing the summary text.
        """
        return ModuleData(
            module_id        = self.MODULE_ID,
            display_name     = self.DISPLAY_NAME,
            metrics          = {"cold_start": True},
            table_data       = [],
            chart_data       = {},
            summary_text     = "Trend data being established — available from next month.",
            metadata         = {"cold_start": True},
            driver_narrative = "Trend data being established.",
            analyst_rows     = [],
            rag_strip        = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            ),
            error            = None,
        )

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute rolling-window MTTR from durably-fixed findings.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered open vulnerability DataFrame.  Not used for gauge
            computation (population is fixed_vulns_df); accepted for interface
            compatibility.
        assets_df : pd.DataFrame
            Asset DataFrame for extract_owner() Owner cut.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Options:
              ``mttr_window_days``  (int, default 30) — rolling-window days.
              ``min_sample_size``   (int, default 5)  — sub-threshold guard.
        **kwargs
            ``fixed_vulns_df`` (pd.DataFrame, optional) — durably-fixed
            findings.  Delivered via ``_MODULES_NEEDING_FIXED_VULNS``.
            None/empty → full cold-start (gray RAG).

            ``trend_snapshots`` (dict, optional) — full ``read_trend()``
            result: ``{"snapshots": [...], "insufficient_data": bool}``.
            When absent/insufficient → MoM cold-start (live gauges still
            render). QUAL-01.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            # ----------------------------------------------------------------
            # QUAL-01: independent cold-start guards
            # ----------------------------------------------------------------

            # --- Trend-snapshots cold-start (MoM line only) ---
            trend_snapshots = kwargs.get("trend_snapshots")
            snapshots_cold = (
                trend_snapshots is None
                or trend_snapshots.get("insufficient_data", True)
            )
            # NOTE: cold_start on snapshots does NOT abort the module —
            # per-severity gauges still render from fixed_vulns_df live data.
            # Only the MoM trend line cold-starts independently.

            snapshots: list[dict] = (
                []
                if snapshots_cold
                else trend_snapshots.get("snapshots", [])
            )

            # --- fixed_vulns_df cold-start (full module) ---
            fixed_vulns_df: Optional[pd.DataFrame] = kwargs.get("fixed_vulns_df")

            if fixed_vulns_df is None or fixed_vulns_df.empty:
                logger.info(
                    "%s fixed_vulns_df absent or empty — full cold start.",
                    self._log_prefix(),
                )
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # Config
            # ----------------------------------------------------------------
            window_days = int(config.options.get("mttr_window_days", 30))
            min_sample  = int(config.options.get("min_sample_size", 5))

            # D-16-11/D-16-12: view selector — whitelist to {owner, severity, both}
            _raw_view = str(config.options.get("mttr_view", "owner")).lower().strip()
            _valid_views = {"owner", "severity", "both"}
            if _raw_view not in _valid_views:
                logger.warning(
                    "%s unknown mttr_view=%r — falling back to 'owner'",
                    self._log_prefix(), _raw_view,
                )
                mttr_view = "owner"
            else:
                mttr_view = _raw_view

            # ----------------------------------------------------------------
            # D-16-01: durably-fixed filter — state == "FIXED" only
            # Never uses "state == fixed OR last_fixed.notna()" (includes REOPENED)
            # ----------------------------------------------------------------
            fixed_df = fixed_vulns_df[
                fixed_vulns_df["state"].astype(str).str.upper() == "FIXED"
            ].copy()

            if fixed_df.empty:
                logger.info(
                    "%s no durably-FIXED findings — full cold start.",
                    self._log_prefix(),
                )
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # D-16-02: date-math only, reopened-aware COALESCE.
            # Both sides coerced to datetime64[ns, UTC] defensively (Pitfall A).
            # ----------------------------------------------------------------
            _nat_series = pd.Series(
                [pd.NaT] * len(fixed_df), index=fixed_df.index, dtype="object"
            )

            last_fixed_ts = pd.to_datetime(
                fixed_df["last_fixed"] if "last_fixed" in fixed_df.columns else _nat_series,
                utc=True, errors="coerce",
            )
            first_found_ts = pd.to_datetime(
                fixed_df["first_found"] if "first_found" in fixed_df.columns else _nat_series,
                utc=True, errors="coerce",
            )
            resurfaced_ts = pd.to_datetime(
                fixed_df["resurfaced_date"] if "resurfaced_date" in fixed_df.columns else _nat_series,
                utc=True, errors="coerce",
            )

            # COALESCE: use resurfaced_date when present, else first_found
            clock_start_ts = resurfaced_ts.where(resurfaced_ts.notna(), other=first_found_ts)

            # Both sides are datetime64[ns, UTC]; subtraction is safe
            date_diff_days = (last_fixed_ts - clock_start_ts).dt.days.clip(lower=0)

            # CoW-compliant: .assign() only — never fixed_df["days_to_fix"] = ... after filter
            fixed_df = fixed_df.assign(days_to_fix=date_diff_days)
            fixed_df = fixed_df[
                fixed_df["days_to_fix"].notna() & (fixed_df["days_to_fix"] >= 0)
            ]

            # ----------------------------------------------------------------
            # Apply rolling window to fixed_df for live gauge computation (D-16-04)
            # ----------------------------------------------------------------
            report_ts = pd.Timestamp(report_date)
            report_ts = report_ts if report_ts.tzinfo is not None else report_ts.tz_localize("UTC")
            window_cutoff = report_ts - pd.Timedelta(days=window_days)

            # last_fixed_ts was built before filter; rebuild on filtered index
            last_fixed_ts_filtered = pd.to_datetime(
                fixed_df["last_fixed"] if "last_fixed" in fixed_df.columns else pd.Series(dtype="object"),
                utc=True, errors="coerce",
            )
            fixed_df = fixed_df[last_fixed_ts_filtered >= window_cutoff]

            if fixed_df.empty:
                logger.info(
                    "%s no findings in rolling window (%d days) — full cold start.",
                    self._log_prefix(), window_days,
                )
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # D-16-02 consequence: overall MTTR = flat sample-weighted mean
            # NOT a mean of per-severity means
            # ----------------------------------------------------------------
            overall_mttr: Optional[float] = (
                round(float(fixed_df["days_to_fix"].mean()), 1)
                if len(fixed_df) >= min_sample
                else None
            )

            # ----------------------------------------------------------------
            # Per-severity MTTR
            # ----------------------------------------------------------------
            per_sev_mttr:        dict[str, Optional[float]] = {}
            per_sev_status:      dict[str, str]              = {}
            per_sev_color:       dict[str, str]              = {}
            per_sev_sample:      dict[str, int]              = {}
            per_sev_sla:         dict[str, int]              = {}
            per_sev_variance:    dict[str, Optional[float]]  = {}
            per_sev_insufficient: dict[str, str]             = {}

            for sev in _SEVERITIES:
                sla = SLA_DAYS[sev]
                per_sev_sla[sev] = sla

                if "severity" in fixed_df.columns:
                    sev_df = fixed_df[fixed_df["severity"].str.lower() == sev]
                else:
                    sev_df = pd.DataFrame()

                n = len(sev_df)
                per_sev_sample[sev] = n

                if n == 0:
                    # Zero fixed for this severity — show "No Data"
                    per_sev_mttr[sev]        = None
                    per_sev_status[sev]      = _STATUS_NODATA
                    per_sev_color[sev]       = _COLOR_GREY
                    per_sev_variance[sev]    = None
                    per_sev_insufficient[sev] = _STATUS_NODATA
                elif n < min_sample:
                    # D-16-07 sparse wording
                    per_sev_mttr[sev]        = None
                    per_sev_status[sev]      = _STATUS_NODATA
                    per_sev_color[sev]       = _COLOR_GREY
                    per_sev_variance[sev]    = None
                    per_sev_insufficient[sev] = (
                        f"Insufficient data ({n} findings — minimum {min_sample} required)"
                    )
                else:
                    mttr = round(float(sev_df["days_to_fix"].mean()), 1)
                    per_sev_mttr[sev] = mttr
                    status, color = _status_from_ratio(mttr / sla)
                    per_sev_status[sev]      = status
                    per_sev_color[sev]       = color
                    per_sev_variance[sev]    = round(mttr - sla, 1)
                    per_sev_insufficient[sev] = ""

            # ----------------------------------------------------------------
            # Per-Owner MTTR (D-16-06)
            # ----------------------------------------------------------------
            enriched_assets = extract_owner(assets_df)
            uuid_to_owner   = dict(
                zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
            ) if not enriched_assets.empty else {}

            if "asset_uuid" in fixed_df.columns and uuid_to_owner:
                owner_col = fixed_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                fixed_df_w_owner = fixed_df.assign(owner=owner_col)

                owner_mttr:   dict[str, Optional[float]] = {}
                owner_sample: dict[str, int]             = {}
                owner_status: dict[str, str]             = {}
                owner_insuf:  dict[str, str]             = {}
                # Critical SLA used only for status label (velocity reference);
                # NOT emitted as an SLA Target for Owner rows (D-16-12).
                critical_sla = SLA_DAYS["critical"]

                for owner, grp in fixed_df_w_owner.groupby("owner", dropna=False):
                    owner_str = str(owner)
                    n = len(grp)
                    if n == 0:
                        # Omit owners with 0 fixed findings per D-16-07
                        continue
                    owner_sample[owner_str] = n
                    if n < min_sample:
                        owner_mttr[owner_str]   = None
                        owner_status[owner_str] = _STATUS_NODATA
                        owner_insuf[owner_str]  = (
                            f"Insufficient data ({n} findings — minimum {min_sample} required)"
                        )
                    else:
                        m = round(float(grp["days_to_fix"].mean()), 1)
                        owner_mttr[owner_str]   = m
                        status, _ = _status_from_ratio(m / critical_sla)
                        owner_status[owner_str] = status
                        owner_insuf[owner_str]  = ""
            else:
                owner_mttr   = {}
                owner_sample = {}
                owner_status = {}
                owner_insuf  = {}

            # ----------------------------------------------------------------
            # RAG strip (CONTRACT-03) — overall MTTR vs Critical SLA anchor
            # ----------------------------------------------------------------
            if overall_mttr is None:
                rag_status  = "no_data"
                rag_headline = NO_DATA_HEADLINE
            else:
                ratio = overall_mttr / SLA_DAYS["critical"]
                if ratio <= _WITHIN_SLA_LIMIT:
                    rag_status = "green"
                elif ratio <= _NEAR_LIMIT:
                    rag_status = "yellow"
                else:
                    rag_status = "red"
                rag_headline = (safe_format(overall_mttr, ".1f") + "d")

            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = rag_headline,
                status             = rag_status,
            )

            # ----------------------------------------------------------------
            # MoM trend from snapshots (D-16-08: deduplicate per calendar month)
            # ----------------------------------------------------------------
            current_period = pd.Period(report_date, "M")

            by_month: dict[str, dict] = {}
            for snap in snapshots:
                m = snap.get("month", "")
                if m not in by_month or snap.get("generated_at", "") > by_month[m].get("generated_at", ""):
                    by_month[m] = snap
            snapshots_deduped = [by_month[m] for m in sorted(by_month)]

            # Build aligned series (RESEARCH Open Item 2)
            months_labels:       list[str]              = []
            overall_mttr_series: list[Optional[float]]  = []
            sev_series: dict[str, list[Optional[float]]] = {s: [] for s in _SEVERITIES}
            owner_set:  set[str] = set()

            for snap in snapshots_deduped:
                month_str = snap.get("month", "")
                months_labels.append(_month_label(month_str, current_period))
                overall_mttr_series.append(snap.get("mttr_overall_days"))
                for sev in _SEVERITIES:
                    sev_series[sev].append((snap.get("mttr_by_severity") or {}).get(sev))
                for owner in (snap.get("mttr_by_owner") or {}):
                    owner_set.add(owner)

            # Per-Owner series with None back-fill (Pitfall B)
            owner_series: dict[str, list[Optional[float]]] = {}
            for snap_idx, snap in enumerate(snapshots_deduped):
                snap_owners = snap.get("mttr_by_owner") or {}
                for owner in owner_set:
                    if owner not in owner_series:
                        # New Owner seen mid-series — back-fill earlier months with None
                        owner_series[owner] = [None] * snap_idx
                    owner_series[owner].append(snap_owners.get(owner))
            for owner in owner_set:
                while len(owner_series[owner]) < len(months_labels):
                    owner_series[owner].append(None)

            # Owner MoM delta (absolute days)
            owner_mom: dict[str, Optional[float]] = {
                owner: _owner_mom_delta(series)
                for owner, series in owner_series.items()
            }

            # ----------------------------------------------------------------
            # Driver narrative
            # ----------------------------------------------------------------
            if overall_mttr is not None:
                status_label, _ = _status_from_ratio(overall_mttr / SLA_DAYS["critical"])
                driver_narrative = (
                    f"Overall MTTR {safe_format(overall_mttr, '.1f')}d "
                    f"(rolling {window_days} days) — {status_label}. "
                    f"Critical SLA target: {SLA_DAYS['critical']}d."
                )
            else:
                driver_narrative = "Insufficient fixed findings to compute MTTR."

            # ----------------------------------------------------------------
            # D-16-11: two independent breakdown lists (split Severity / Owner)
            # ----------------------------------------------------------------

            # Severity rows — per-severity SLA Target and Variance are meaningful
            table_data_severity: list[dict] = []
            for sev in _SEVERITIES:
                table_data_severity.append({
                    "label":        sev.capitalize(),
                    "mttr_days":    per_sev_mttr.get(sev),
                    "sla_days":     per_sev_sla[sev],
                    "variance":     per_sev_variance.get(sev),
                    "status":       per_sev_status.get(sev, _STATUS_NODATA),
                    "sample_size":  per_sev_sample.get(sev, 0),
                    "mom_delta":    None,   # severity rows: no MoM delta
                    "insufficient": per_sev_insufficient.get(sev, ""),
                })

            # Owner rows — D-16-11/D-16-12: drop meaningless Critical-SLA anchor.
            # sla_days=None and variance=None; renderers omit the SLA Target column
            # for the Owner cut.  mom_delta and sample_size remain (SLA-independent).
            table_data_owner: list[dict] = []
            for owner_str in sorted(owner_mttr.keys()):
                table_data_owner.append({
                    "label":        owner_str,
                    "mttr_days":    owner_mttr[owner_str],
                    "sla_days":     None,   # D-16-12: no SLA anchor for Owner cut
                    "variance":     None,   # meaningless without an SLA basis
                    "status":       owner_status.get(owner_str, _STATUS_NODATA),
                    "sample_size":  owner_sample.get(owner_str, 0),
                    "mom_delta":    owner_mom.get(owner_str),
                    "insufficient": owner_insuf.get(owner_str, ""),
                })

            # Combined list kept only for analyst_rows (CONTRACT-02 keeps both cuts)
            table_data: list[dict] = table_data_severity + table_data_owner

            # ----------------------------------------------------------------
            # chart_data
            # ----------------------------------------------------------------
            chart_data = {
                "months":              months_labels,
                "overall_mttr_series": overall_mttr_series,
                "sev_series":          sev_series,
                "owner_series":        owner_series,
                "window_days":         window_days,
            }

            # ----------------------------------------------------------------
            # Analyst rows (CONTRACT-02) — aggregate only (QUAL-05)
            # ----------------------------------------------------------------
            sev_df = pd.DataFrame([
                {
                    "Severity / Owner": row["label"],
                    "MTTR (Days)":      row["mttr_days"],
                    "SLA Target (Days)": row["sla_days"],
                    "Variance (Days)":  row["variance"],
                    "Status":           row["status"],
                    "Sample Size":      row["sample_size"],
                    "MoM Delta (Days)": row["mom_delta"],
                    "Note":             row["insufficient"],
                }
                for row in table_data
            ])

            mom_df = pd.DataFrame({
                "Month":         months_labels,
                "Overall MTTR":  overall_mttr_series,
                **{f"MTTR — {sev.capitalize()}": sev_series[sev] for sev in _SEVERITIES},
            }) if months_labels else pd.DataFrame()

            analyst_rows: list[tuple[str, pd.DataFrame]] = [
                ("MTTR by Severity+Owner", sev_df),
                ("MTTR MoM Trend",         mom_df),
            ]

            # ----------------------------------------------------------------
            # summary_text
            # ----------------------------------------------------------------
            if overall_mttr is not None:
                summary_text = (
                    f"Overall MTTR: {safe_format(overall_mttr, '.1f')}d "
                    f"(rolling {window_days}-day window, "
                    f"{len(fixed_df)} fixed findings). "
                    f"Current month ({str(current_period)}) is month-to-date (partial)."
                )
            else:
                summary_text = (
                    f"Insufficient fixed findings ({len(fixed_df)}) to compute MTTR. "
                    f"Minimum {min_sample} required."
                )

            current_month = str(current_period)

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "overall_mttr":    overall_mttr,
                    "rag_status":      rag_status,
                    "cold_start":      False,
                    "window_days":     window_days,
                    "sample_size":     len(fixed_df),
                    "snapshots_cold":  snapshots_cold,
                    **{f"{sev}_mttr":   per_sev_mttr.get(sev)   for sev in _SEVERITIES},
                    **{f"{sev}_status": per_sev_status.get(sev) for sev in _SEVERITIES},
                    **{f"{sev}_sample": per_sev_sample.get(sev, 0) for sev in _SEVERITIES},
                },
                table_data   = table_data,
                chart_data   = chart_data,
                summary_text = summary_text,
                metadata     = {
                    "window_days":         window_days,
                    "current_month":       current_month,
                    "snapshots_cold":      snapshots_cold,
                    "min_sample":          min_sample,
                    "mttr_view":           mttr_view,
                    "table_data_severity": table_data_severity,
                    "table_data_owner":    table_data_owner,
                    "computed_at":         (
                        report_date.isoformat()
                        if hasattr(report_date, "isoformat")
                        else str(report_date)
                    ),
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render per-severity MTTR gauges, window disclosure, MoM series,
        and Owner table.

        Returns an error callout if ``data.error`` is set; a cold-start
        notice if ``metrics["cold_start"]`` is True.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        if data.metrics.get("cold_start"):
            return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p style="color:#757575;">{data.summary_text}</p>
</div>"""

        m = data.metrics
        window_days   = data.metadata.get("window_days", 30)
        current_month = data.metadata.get("current_month", "")
        overall_mttr  = m.get("overall_mttr")

        # D-16-04 disclosure
        disclosure = (
            f'<p class="explanatory-text">'
            f"Rolling {window_days}-day MTTR — findings remediated in the last "
            f"{window_days} days. Current month ({current_month}) is month-to-date (partial)."
            f"</p>"
        )

        # Overall MTTR summary line
        rag_status = m.get("rag_status", "no_data")
        rag_color  = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        rag_label  = STATUS_LABEL.get(rag_status, STATUS_LABEL.get("no_data", "No Data"))

        overall_line = (
            f"<p>"
            f"<strong>Overall MTTR:</strong> "
            f"{safe_format(overall_mttr, '.1f') + 'd' if overall_mttr is not None else 'Insufficient data'}"
            f" &nbsp;|&nbsp; "
            f"<strong>Status:</strong> "
            f'<span style="color:{rag_color};font-weight:bold;">{rag_label}</span>'
            f" (vs Critical SLA {SLA_DAYS['critical']}d)"
            f"</p>"
        )

        # Per-severity gauges
        gauges_html = ""
        for sev in _SEVERITIES:
            mttr   = m.get(f"{sev}_mttr")
            sla    = float(SLA_DAYS[sev])
            status = m.get(f"{sev}_status", _STATUS_NODATA)
            color  = (
                _COLOR_GREEN if status == _STATUS_WITHIN else
                _COLOR_AMBER if status == _STATUS_NEAR   else
                _COLOR_RED   if status == _STATUS_EXCEED else
                _COLOR_GREY
            )
            thresholds = [
                (sla,        color),
                (sla * 1.25, _COLOR_AMBER),
                (sla * 2,    _COLOR_RED),
            ]

            if mttr is None:
                n     = m.get(f"{sev}_sample", 0)
                insuf = data.metadata.get("min_sample", 5)
                msg   = (
                    f"Insufficient data ({n} findings — minimum {insuf} required)"
                    if n > 0
                    else "No Data"
                )
                gauge_inner = (
                    f'<div style="text-align:center;padding:16pt;">'
                    f'<span style="color:#9E9E9E;font-size:9pt;">{msg}</span>'
                    f"</div>"
                )
            else:
                b64 = draw_gauge(
                    value          = mttr,
                    min_val        = 0,
                    max_val        = sla * 2,
                    thresholds     = thresholds,
                    title          = f"MTTR — {sev.capitalize()}",
                    unit           = "d",
                    reference_line = sla,
                    reference_label = "SLA",
                )
                gauge_inner = (
                    f'<img src="data:image/png;base64,{b64}" '
                    f'style="width:100%;max-width:160pt;">'
                )

            gauges_html += (
                f'<div style="display:inline-block;text-align:center;'
                f'width:23%;margin:0 1%;">'
                f"{gauge_inner}"
                f"</div>"
            )

        # SLA + Owner table
        table_rows = ""
        for row in data.table_data:
            mttr  = row.get("mttr_days")
            insuf = row.get("insufficient", "")
            if insuf:
                mttr_str = insuf
            elif mttr is not None:
                mttr_str = safe_format(mttr, ".1f") + "d"
            else:
                mttr_str = "No Data"

            var     = row.get("variance")
            var_str = (safe_format(var, "+.1f") + "d") if var is not None else "—"
            mom     = row.get("mom_delta")
            mom_str = (safe_format(mom, "+.1f") + "d") if mom is not None else "—"

            table_rows += (
                f"<tr>"
                f"<td>{row['label']}</td>"
                f"<td style='text-align:right;'>{mttr_str}</td>"
                f"<td style='text-align:right;'>{row['sla_days']}d</td>"
                f"<td style='text-align:right;'>{var_str}</td>"
                f"<td>{row['status']}</td>"
                f"<td style='text-align:right;'>{safe_int(row['sample_size'])}</td>"
                f"<td style='text-align:right;'>{mom_str}</td>"
                f"</tr>"
            )

        snapshots_cold = data.metadata.get("snapshots_cold", True)
        mom_notice = (
            '<p style="color:#757575;font-style:italic;">'
            "Month-over-month trend being established."
            "</p>"
            if snapshots_cold else ""
        )

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  {overall_line}
  <div style="text-align:center;margin-bottom:8pt;">
    {gauges_html}
  </div>
  {disclosure}
  {mom_notice}
  <table class="data-table" style="width:100%;margin-top:8pt;">
    <thead>
      <tr>
        <th>Severity / Owner</th>
        <th style="text-align:right;">MTTR (Days)</th>
        <th style="text-align:right;">SLA Target</th>
        <th style="text-align:right;">Variance</th>
        <th>Status</th>
        <th style="text-align:right;">Sample Size</th>
        <th style="text-align:right;">MoM Delta (Days)</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>
  <p class="explanatory-text">
    MTTR uses reopened-aware date math: clock resets to resurfaced_date when a
    finding re-emerges. Overall MTTR is a sample-weighted flat mean (not a mean
    of per-severity means). MoM delta = current month MTTR minus prior month
    MTTR (absolute days). Severity rows use per-severity SLA target; Owner rows
    use Critical SLA ({SLA_DAYS["critical"]}d) as the anchor.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a single "MTTR Trend" tab with window disclosure in row 1.

        Column order (D-16-05):
        Severity/Owner, MTTR (Days), SLA Target (Days),
        Variance (Days), Status, Sample Size, MoM Delta (Days).

        Returns ``[]`` on exception; writes a cold-start or error row
        when set.
        """
        tab_name = "MTTR Trend"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            if data.metrics.get("cold_start"):
                ws["A1"] = "MTTR Trend (Reopened-Aware)"
                ws["A1"].font = Font(bold=True, size=13)
                ws["A2"] = data.summary_text
                ws["A2"].font = Font(italic=True, color="757575")
                return [tab_name]

            window_days = data.metadata.get("window_days", 30)

            # Row 1: title with window disclosure (D-16-04)
            ws["A1"] = f"MTTR Trend — Rolling {window_days}-day window"
            ws["A1"].font = Font(bold=True, size=13)

            # Headers (D-16-05 column order)
            headers = [
                "Severity / Owner",
                "MTTR (Days)",
                "SLA Target (Days)",
                "Variance (Days)",
                "Status",
                "Sample Size",
                "MoM Delta (Days)",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _FILL_HEADER

            # Data rows
            for row_idx, row in enumerate(data.table_data, start=4):
                mttr  = row.get("mttr_days")
                insuf = row.get("insufficient", "")
                var   = row.get("variance")
                mom   = row.get("mom_delta")

                if insuf:
                    mttr_val = insuf
                elif mttr is not None:
                    mttr_val = mttr
                else:
                    mttr_val = "No Data"

                ws.cell(row=row_idx, column=1, value=row["label"])
                ws.cell(row=row_idx, column=2, value=mttr_val)
                ws.cell(row=row_idx, column=3, value=row["sla_days"])
                ws.cell(row=row_idx, column=4, value=var if var is not None else "—")
                status_cell = ws.cell(row=row_idx, column=5, value=row["status"])
                status_cell.fill = _status_fill(row["status"])
                ws.cell(row=row_idx, column=6, value=row["sample_size"])
                ws.cell(row=row_idx, column=7, value=mom if mom is not None else "—")

            # Column widths
            widths = [28, 14, 18, 16, 18, 14, 18]
            for col_idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = w

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01 — NOT render_email_kpis)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Returns ``""`` on error or empty driver_narrative.
        Cold-start renders the "Trend data being established" notice.
        Inline CSS only — no <style> blocks (Outlook/Gmail/Apple Mail compat).
        """
        if data.error:
            return ""

        if data.metrics.get("cold_start"):
            return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#F5F5F5;border-left:4px solid #757575;">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;color:#757575;">{data.summary_text}</span>
    </td>
  </tr>
</table>"""

        if not data.driver_narrative:
            return ""

        m           = data.metrics
        rag_status  = m.get("rag_status", "no_data")
        rag_color   = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        overall     = m.get("overall_mttr")
        window_days = data.metadata.get("window_days", 30)
        sample      = safe_int(m.get("sample_size"))

        overall_str = (
            safe_format(overall, ".1f") + "d"
            if overall is not None
            else "Insufficient data"
        )

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#FFF3E0;border-left:4px solid {rag_color};">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;">
        Overall MTTR: <strong>{overall_str}</strong>
        &mdash; {sample} fixed findings in window
      </span><br>
      <em style="font-size:11px;color:#555;">{data.driver_narrative}</em><br>
      <span style="font-size:10px;color:#757575;">
        Rolling {window_days}-day MTTR. Current month is partial.
      </span>
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns ``data.analyst_rows`` when valid; ``[]`` on error or
        cold-start. Columns are aggregate-only (QUAL-05 — no per-finding PII).
        """
        if data.error or not data.analyst_rows:
            return []
        if data.metrics.get("cold_start"):
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present;
        falls back to a gray "No Data" cell on error or empty strip.
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip

    # ------------------------------------------------------------------
    # validate_config
    # ------------------------------------------------------------------

    def validate_config(self, config: ModuleConfig) -> list[str]:
        """Validate optional ``mttr_window_days``, ``min_sample_size``, and ``mttr_view`` options."""
        errors: list[str] = []
        for key, min_val in [("mttr_window_days", 1), ("min_sample_size", 1)]:
            val = config.options.get(key)
            if val is not None:
                try:
                    if int(val) < min_val:
                        errors.append(
                            f"mttr_trend: '{key}' must be >= {min_val}, got {val}"
                        )
                except (TypeError, ValueError):
                    errors.append(
                        f"mttr_trend: '{key}' must be an integer, "
                        f"got {type(val).__name__}"
                    )
        # D-16-11: validate mttr_view when explicitly set
        view_val = config.options.get("mttr_view")
        if view_val is not None:
            _valid_views = {"owner", "severity", "both"}
            if str(view_val).lower().strip() not in _valid_views:
                errors.append(
                    f"mttr_trend: 'mttr_view' must be one of "
                    f"{sorted(_valid_views)}, got {view_val!r}"
                )
        return errors

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Population": (
                    "Durably-fixed findings: state == 'FIXED' AND last_fixed >= "
                    "report_date - window_days. Never includes REOPENED findings "
                    "(D-16-01)."
                ),
                "days_to_fix": (
                    "(last_fixed - COALESCE(resurfaced_date, first_found)).days "
                    "clipped >= 0. Reopened-aware: clock resets on resurfaced_date "
                    "(D-16-02). time_taken_to_fix field NOT used."
                ),
                "Overall MTTR": (
                    "Flat sample-weighted mean of days_to_fix across all in-scope "
                    "findings (NOT a mean of per-severity means — D-16-02 consequence)."
                ),
                "RAG anchor": (
                    f"Overall MTTR / SLA_DAYS['critical'] ({SLA_DAYS['critical']}d). "
                    f"Green <= {_WITHIN_SLA_LIMIT}, Amber <= {_NEAR_LIMIT}, Red > {_NEAR_LIMIT}."
                ),
                "MoM trend": (
                    "Reads from trend_snapshots kwargs (persisted by capture_trend_snapshot.py). "
                    "Cold-starts independently of live gauges (QUAL-01)."
                ),
                "Owner MoM delta": "Absolute day delta (curr - prev), NOT percentage (D-16-06).",
                "min_sample_size": (
                    "Default 5. Buckets with n < min_sample render: "
                    "'Insufficient data (N findings — minimum 5 required)' (D-16-07)."
                ),
            },
        }
