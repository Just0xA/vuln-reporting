"""
reports/modules/tag_severity_share_module.py — Tag Severity Share vs Environment.

Expresses the VPR severity distribution of a selected Tenable tag's open
findings as a share of the **entire environment's** open-finding total.

Each severity percentage is:

    pct[sev] = tag_count[sev] / env_vuln_total * 100

The five severity rows (critical/high/medium/low/none) therefore sum to the
tag's overall share of the environment, NOT to 100%.

Severity is VPR-PURE with an explicit "None" bucket — no native-severity
fallback (locked by design decision D3 in the spec).  See ``get_audit_info()``
and ``docs/tag_severity_share_calculations.md`` for the intentional divergence
from ``config.vpr_to_severity``.

Denominator: ``env_vuln_total`` is the unfiltered environment open-finding count
forwarded by ``composed_report.py`` via ``**kwargs`` (gated on the
``_MODULES_NEEDING_ENV_TOTAL`` frozenset).  When this module is run in isolation
without the gate (e.g. unit tests), pass ``env_vuln_total=N`` directly to
``compute()``.

RAG strip: This is a *share* metric, not an SLA gate.  The strip headline is the
tag's total share of the environment expressed as a percentage.  Status is
``"no_data"`` when the tag has zero open findings, otherwise a fixed
``"yellow"`` (informational / at-attention) because share alone does not indicate
pass/fail without a threshold — and the spec deliberately does not lock one.
This is documented here so auditors do not interpret the yellow colour as a
warning threshold breach.
"""

from __future__ import annotations

import logging
import math
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from reports.modules import (
    safe_pct, safe_int, safe_format,
    build_rag_strip_entry, NO_DATA_HEADLINE, NO_DATA_DRIVER,
)

logger = logging.getLogger(__name__)

# ── Open states ──────────────────────────────────────────────────────────────
_OPEN_STATES: frozenset[str] = frozenset({"open", "reopened"})

# ── Severity display order (descending risk, then None last) ─────────────────
_SEVERITIES: tuple[str, ...] = ("critical", "high", "medium", "low", "none")

# ── VPR range boundaries (mirrors VPR_SEVERITY_MAP in config.py, VPR-pure) ───
# None bucket: vpr_score is null/NaN OR == 0.0  (D2, D3 — no native fallback)
_VPR_RANGES: tuple[tuple[float, float, str], ...] = (
    (9.0,  10.0, "critical"),
    (7.0,   8.9, "high"),
    (4.0,   6.9, "medium"),
    (0.1,   3.9, "low"),
)

# ── Severity colours (reuse project palette; import lazily to avoid cycles) ──
_SEV_COLORS: dict[str, str] = {
    "critical": "#d32f2f",
    "high":     "#f57c00",
    "medium":   "#fbc02d",
    "low":      "#388e3c",
    "none":     "#9e9e9e",
}
_SEV_TEXT_COLORS: dict[str, str] = {
    "critical": "#ffffff",
    "high":     "#ffffff",
    "medium":   "#000000",
    "low":      "#ffffff",
    "none":     "#ffffff",
}


# ── VPR-pure bucket helper ────────────────────────────────────────────────────

def _bucket_severity(vpr_score: Any) -> str:
    """
    Map a raw vpr_score value to one of the five severity tiers.

    Returns ``"none"`` when vpr_score is:
    - ``None``
    - ``float('nan')`` / ``pd.NA`` / any ``pd.isna()``-truthy value
    - ``0.0`` (matches Tenable GUI "None" per D2)
    - an empty string or non-numeric value

    Does NOT call ``config.vpr_to_severity`` — no native-severity fallback
    (D3).
    """
    if vpr_score is None:
        return "none"
    try:
        if pd.isna(vpr_score):
            return "none"
    except (TypeError, ValueError):
        pass
    try:
        score = float(vpr_score)
    except (TypeError, ValueError):
        return "none"
    if score == 0.0:
        return "none"
    for lo, hi, label in _VPR_RANGES:
        if lo <= score <= hi:
            return label
    return "none"


# ── Module ────────────────────────────────────────────────────────────────────

@register_module
class TagSeverityShareModule(BaseModule):
    """
    VPR severity distribution of a tag's open findings as a share of the
    entire environment.

    The denominator is the **environment grand total** (all open findings,
    all assets, all severities) supplied via ``env_vuln_total`` in ``**kwargs``.

    Severity derivation: VPR-pure with an explicit None bucket — null/NaN/0.0
    VPR scores land in "None", regardless of the Tenable native severity field.
    This is an intentional divergence from ``config.vpr_to_severity`` which
    applies a native-severity fallback; it is documented in
    ``docs/tag_severity_share_calculations.md``.

    Supported options
    -----------------
    None — this module accepts no options.
    """

    MODULE_ID         = "tag_severity_share"
    DISPLAY_NAME      = "Tag Severity Share vs Environment"
    DESCRIPTION       = (
        "VPR severity distribution of a Tenable tag's open findings, "
        "expressed as a share of the entire environment's open-finding total."
    )
    REQUIRED_DATA     = ["vulns"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute per-severity counts and their share of the environment total.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Must have ``state`` and
            ``vpr_score`` columns.
        assets_df : pd.DataFrame
            Not used by this module; accepted for interface compatibility.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Module configuration (no options consumed).
        **kwargs : Any
            Must include ``env_vuln_total`` (int) — the unfiltered environment
            open-finding count forwarded by ``composed_report.py``.

        Returns
        -------
        ModuleData
            On success: ``error`` is None and all data fields are populated.
            On failure: ``error`` is set and data fields hold safe empty defaults.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            env_vuln_total = int(kwargs.get("env_vuln_total", 0) or 0)

            # Restrict to open states
            if vulns_df.empty or "state" not in vulns_df.columns:
                open_df = pd.DataFrame(columns=list(vulns_df.columns) if not vulns_df.empty else ["state", "vpr_score"])
            else:
                open_df = vulns_df[vulns_df["state"].str.lower().isin(_OPEN_STATES)].copy()

            tag_total = len(open_df)

            # Bucket each finding by VPR-pure severity
            if tag_total > 0 and "vpr_score" in open_df.columns:
                open_df = open_df.assign(
                    _sev_bucket=open_df["vpr_score"].map(_bucket_severity)
                )
                counts = {
                    sev: int((open_df["_sev_bucket"] == sev).sum())
                    for sev in _SEVERITIES
                }
            else:
                counts = {sev: 0 for sev in _SEVERITIES}

            # Per-severity pct = tag_count[sev] / env_vuln_total * 100 (D1)
            pcts: dict[str, float] = {}
            for sev in _SEVERITIES:
                pcts[sev] = (
                    counts[sev] / env_vuln_total * 100.0
                    if env_vuln_total > 0 else 0.0
                )

            tag_share_pct = (
                tag_total / env_vuln_total * 100.0
                if env_vuln_total > 0 else 0.0
            )

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            metrics: dict[str, Any] = {
                "env_vuln_total":   env_vuln_total,
                "tag_total":        tag_total,
                "tag_share_pct":    tag_share_pct,
            }
            for sev in _SEVERITIES:
                metrics[f"{sev}_count"] = counts[sev]
                metrics[f"{sev}_pct"]   = pcts[sev]

            # ── driver narrative ──────────────────────────────────────────
            if tag_total == 0:
                driver = NO_DATA_DRIVER
            else:
                top_sev = max(
                    (s for s in ("critical", "high", "medium", "low", "none")),
                    key=lambda s: counts[s],
                )
                driver = (
                    f"This tag accounts for {safe_pct(tag_share_pct)} of all "
                    f"environment findings; {safe_int(counts[top_sev])} are "
                    f"{top_sev.capitalize()}."
                )

            # ── RAG strip (informational) ─────────────────────────────────
            # Status: "no_data" when tag has no findings; "yellow" otherwise
            # (share metric — no threshold locked by spec).
            if tag_total == 0:
                strip_status = "no_data"
                headline_str = NO_DATA_HEADLINE
            else:
                strip_status = "yellow"
                headline_str = safe_pct(tag_share_pct)

            rag_strip = build_rag_strip_entry(
                self.DISPLAY_NAME, headline_str, strip_status
            )

            # ── analyst rows ──────────────────────────────────────────────
            analyst_df = pd.DataFrame()
            if tag_total > 0 and "_sev_bucket" in open_df.columns:
                analyst_df = open_df.drop(columns=["_sev_bucket"], errors="ignore").assign(
                    vpr_severity_bucket=open_df["_sev_bucket"]
                )

            analyst_rows = (
                [(f"{self.DISPLAY_NAME[:28]}", analyst_df)]
                if not analyst_df.empty else []
            )

            summary_text = (
                f"Tag has {tag_total:,} open findings "
                f"({safe_pct(tag_share_pct)} of environment total "
                f"of {env_vuln_total:,})."
            )

            return ModuleData(
                module_id        = self.MODULE_ID,
                display_name     = self.DISPLAY_NAME,
                metrics          = metrics,
                table_data       = [],
                chart_data       = {},
                summary_text     = summary_text,
                metadata         = {
                    "env_vuln_total":  env_vuln_total,
                    "tag_total":       tag_total,
                    "filter_applied":  "state in {open, reopened}",
                    "severity_source": "VPR-pure: vpr_score null/NaN/0.0 → None; no native fallback (D3)",
                    "computed_at":     computed_at,
                },
                driver_narrative = driver,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a severity-vs-environment share table plus a tag-share footer.

        Returns an error callout when ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m = data.metrics
        rows_html = ""
        for sev in _SEVERITIES:
            count = m.get(f"{sev}_count", 0)
            pct   = m.get(f"{sev}_pct", 0.0)
            label = "None" if sev == "none" else sev.capitalize()
            bg    = _SEV_COLORS.get(sev, "#9e9e9e")
            fg    = _SEV_TEXT_COLORS.get(sev, "#ffffff")
            rows_html += (
                f"<tr>"
                f'<td style="background-color:{bg};color:{fg};font-weight:bold;'
                f'padding:4px 8px;">{label}</td>'
                f'<td style="text-align:right;padding:4px 8px;">{safe_int(count)}</td>'
                f'<td style="text-align:right;padding:4px 8px;">{safe_pct(pct)}</td>'
                f"</tr>"
            )

        env_total = m.get("env_vuln_total", 0)
        tag_total = m.get("tag_total", 0)
        tag_share = m.get("tag_share_pct", 0.0)

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p style="font-size:8pt;margin-bottom:6pt;">{data.driver_narrative}</p>
  <table style="width:100%;border-collapse:collapse;font-size:9pt;">
    <thead>
      <tr style="background-color:#eeeeee;">
        <th style="text-align:left;padding:4px 8px;">Severity</th>
        <th style="text-align:right;padding:4px 8px;">Tag Count</th>
        <th style="text-align:right;padding:4px 8px;">% of Env Total</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
    <tfoot>
      <tr style="border-top:2px solid #333;">
        <td style="font-weight:bold;padding:4px 8px;">Tag Total</td>
        <td style="text-align:right;font-weight:bold;padding:4px 8px;">{safe_int(tag_total)}</td>
        <td style="text-align:right;font-weight:bold;padding:4px 8px;">{safe_pct(tag_share)} of env</td>
      </tr>
    </tfoot>
  </table>
  <p style="font-size:7.5pt;color:#555;margin-top:4pt;">
    Environment total (all assets, all severities, open states): {safe_int(env_total)} findings.
    Severity derived from VPR score — null/NaN/0.0 VPR = "None" (no native-severity fallback).
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a single "Sev Share" tab with severity, tag count, and % of env.

        Returns ``[]`` on error.
        """
        tab_name = "Sev Share"
        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            headers = ["Severity", "Tag Count", "% of Env Total"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            m = data.metrics
            for row_idx, sev in enumerate(_SEVERITIES, start=2):
                count = m.get(f"{sev}_count", 0)
                pct   = m.get(f"{sev}_pct", 0.0)
                label = "None" if sev == "none" else sev.capitalize()

                sev_cell = ws.cell(row=row_idx, column=1, value=label)
                hex_color = _SEV_COLORS.get(sev, "#9e9e9e").lstrip("#")
                sev_cell.fill = PatternFill("solid", fgColor=hex_color)
                fg_hex = _SEV_TEXT_COLORS.get(sev, "#ffffff").lstrip("#")
                sev_cell.font = Font(bold=True, color=fg_hex)

                ws.cell(row=row_idx, column=2, value=count)
                ws.cell(row=row_idx, column=3, value=safe_pct(pct))

            # Footer row
            footer_row = len(_SEVERITIES) + 2
            tag_total = m.get("tag_total", 0)
            tag_share = m.get("tag_share_pct", 0.0)
            env_total = m.get("env_vuln_total", 0)

            footer_cell = ws.cell(row=footer_row, column=1, value="Tag Total")
            footer_cell.font = Font(bold=True)
            ws.cell(row=footer_row, column=2, value=tag_total).font = Font(bold=True)
            ws.cell(row=footer_row, column=3, value=safe_pct(tag_share)).font = Font(bold=True)

            # Env total note
            note_row = footer_row + 1
            ws.cell(row=note_row, column=1, value="Environment total (open)")
            ws.cell(row=note_row, column=2, value=env_total)
            ws.cell(row=note_row, column=3, value="all assets / all severities")

            ws.column_dimensions[get_column_letter(1)].width = 16
            ws.column_dimensions[get_column_letter(2)].width = 12
            ws.column_dimensions[get_column_letter(3)].width = 16

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Inline-CSS email panel for the composed email body.

        Returns ``""`` when ``data.error`` is set.
        """
        if data.error:
            return ""

        m = data.metrics

        rows_html = ""
        for sev in _SEVERITIES:
            count = m.get(f"{sev}_count", 0)
            pct   = m.get(f"{sev}_pct", 0.0)
            label = "None" if sev == "none" else sev.capitalize()
            bg    = _SEV_COLORS.get(sev, "#9e9e9e")
            fg    = _SEV_TEXT_COLORS.get(sev, "#ffffff")
            rows_html += (
                f'<tr>'
                f'<td style="background-color:{bg};color:{fg};font-weight:bold;'
                f'padding:4px 8px;border:1px solid #ddd;">{label}</td>'
                f'<td style="text-align:right;padding:4px 8px;border:1px solid #ddd;">'
                f'{safe_int(count)}</td>'
                f'<td style="text-align:right;padding:4px 8px;border:1px solid #ddd;">'
                f'{safe_pct(pct)}</td>'
                f'</tr>'
            )

        tag_total = m.get("tag_total", 0)
        tag_share = m.get("tag_share_pct", 0.0)
        env_total = m.get("env_vuln_total", 0)
        driver    = data.driver_narrative or ""

        return (
            f'<div style="font-family:Arial,sans-serif;margin-bottom:16px;">'
            f'<h3 style="margin:0 0 6px 0;font-size:13px;color:#333;">'
            f'{self.DISPLAY_NAME}</h3>'
            f'<p style="margin:0 0 8px 0;font-size:11px;color:#555;">{driver}</p>'
            f'<table style="width:100%;border-collapse:collapse;font-size:11px;">'
            f'<thead><tr style="background-color:#eeeeee;">'
            f'<th style="text-align:left;padding:4px 8px;border:1px solid #ddd;">Severity</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">Tag Count</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">% of Env</th>'
            f'</tr></thead>'
            f'<tbody>{rows_html}</tbody>'
            f'<tfoot><tr style="border-top:2px solid #333;">'
            f'<td style="font-weight:bold;padding:4px 8px;border:1px solid #ddd;">Tag Total</td>'
            f'<td style="text-align:right;font-weight:bold;padding:4px 8px;border:1px solid #ddd;">'
            f'{safe_int(tag_total)}</td>'
            f'<td style="text-align:right;font-weight:bold;padding:4px 8px;border:1px solid #ddd;">'
            f'{safe_pct(tag_share)} of env</td>'
            f'</tr></tfoot>'
            f'</table>'
            f'<p style="font-size:9px;color:#888;margin-top:4px;">'
            f'Env total: {safe_int(env_total)} open findings (all assets).</p>'
            f'</div>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        Return the flat per-finding DataFrame for the analyst workbook.

        Returns ``[]`` on error or zero rows.
        """
        if data.error:
            return []
        return list(data.analyst_rows) if data.analyst_rows else []

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Denominator": (
                    "env_vuln_total = count of all open findings in the UNFILTERED "
                    "vulns_df (state in {'open','reopened'}), forwarded by "
                    "composed_report.py before the tag filter is applied."
                ),
                "Per-severity pct": (
                    "tag_count[sev] / env_vuln_total * 100.0; zero when env_vuln_total == 0 "
                    "(divide-by-zero guard)."
                ),
                "tag_share_pct": (
                    "tag_total / env_vuln_total * 100.0; zero when env_vuln_total == 0. "
                    "Equals the sum of all five per-severity pcts."
                ),
                "Severity buckets (VPR-pure, D3)": (
                    "Derived from raw vpr_score ONLY — no native-severity fallback. "
                    "null/NaN/0.0 → 'None'; 0.1–3.9 → Low; 4.0–6.9 → Medium; "
                    "7.0–8.9 → High; 9.0–10.0 → Critical. "
                    "INTENTIONAL DIVERGENCE: config.vpr_to_severity applies a native "
                    "fallback for null VPR; this module does not. See D3 in spec."
                ),
                "None bucket": (
                    "vpr_score null/NaN OR == 0.0 maps to 'None', matching the Tenable "
                    "GUI 'None' label. Findings with no VPR are NOT promoted to their "
                    "native severity tier (divergence from spike convention and from "
                    "config.vpr_to_severity used elsewhere in the suite)."
                ),
                "Open states": (
                    "state in {'open', 'reopened'} — informational findings excluded "
                    "upstream by the fetcher."
                ),
                "No-tag degenerate case (D7)": (
                    "When no tag filter is applied, tag = environment ⇒ severity shares "
                    "sum to ~100% (the tag IS the environment)."
                ),
                "RAG status": (
                    "Fixed 'yellow' (informational) when tag_total > 0; 'no_data' when "
                    "tag_total == 0. No threshold applied — share alone does not indicate "
                    "pass/fail without a locked threshold (spec D-intentional)."
                ),
            },
        }
