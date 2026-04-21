"""
reports/modules/patch_compliance_rate_module.py — Patch Compliance Rate.

Measures the percentage of open vulnerabilities that are within their
required SLA remediation window.

Usage
-----
::

    from reports.modules import registry
    from reports.modules.base import ModuleConfig

    mod  = registry.get("patch_compliance_rate")()
    data = mod.compute(vulns_df, assets_df, report_date,
                       ModuleConfig("patch_compliance_rate"))
    print(data.metrics["overall_rate"])   # e.g. 87.3

    # Scoped to critical + high only:
    data = mod.compute(
        vulns_df, assets_df, report_date,
        ModuleConfig("patch_compliance_rate",
                     options={"severity_filter": ["critical", "high"]}),
    )
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import SLA_DAYS
from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.chart_utils import draw_gauge
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# Open vulnerability states
_OPEN_STATES: frozenset[str] = frozenset({"open", "reopened"})

# Severity display order
_SEVERITIES: tuple[str, ...] = ("critical", "high", "medium", "low")

# Compliance rate thresholds
_THRESHOLD_GREEN = 90.0
_THRESHOLD_AMBER = 75.0

# Gauge threshold list for draw_gauge — red / amber / green zones
_COMPLIANCE_THRESHOLDS: list[tuple[float, str]] = [
    (_THRESHOLD_AMBER, "#d32f2f"),   # 0–75 red
    (_THRESHOLD_GREEN, "#fbc02d"),   # 75–90 amber
    (100.0,            "#388e3c"),   # 90–100 green
]

# Severity tile colours for small per-severity rate display
_SEV_COLORS: dict[str, str] = {
    "critical": "#d32f2f",
    "high":     "#f57c00",
    "medium":   "#fbc02d",
    "low":      "#388e3c",
}

# Excel fill colours for compliance rate cells
_FILL_GREEN = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED   = PatternFill("solid", fgColor="FFCDD2")


def _rate_fill(rate: Optional[float]) -> PatternFill:
    """Return the appropriate Excel fill for a compliance rate value."""
    if rate is None:
        return PatternFill()   # no fill
    if rate >= _THRESHOLD_GREEN:
        return _FILL_GREEN
    if rate >= _THRESHOLD_AMBER:
        return _FILL_AMBER
    return _FILL_RED


@register_module
class PatchComplianceRateModule(BaseModule):
    """
    Measure the percentage of open vulnerabilities within their SLA window.

    A finding is Within SLA when:
        days_open = (report_date - first_found).days <= SLA_DAYS[severity]

    When ``first_found`` is NaN, ``days_open`` is treated as 0
    (always within SLA), matching the established behaviour in management_summary.py.

    Supported options
    -----------------
    severity_filter : list[str], optional
        Limit computation to the listed severity tiers.
        Default: ``["critical", "high", "medium", "low"]``.
        Example: ``{"severity_filter": ["critical", "high"]}``
        When set, overall_rate is computed only from the listed severities.
    """

    MODULE_ID         = "patch_compliance_rate"
    DISPLAY_NAME      = "Patch Compliance Rate"
    DESCRIPTION       = ("Measures the percentage of open vulnerabilities "
                         "that are within their required SLA remediation window.")
    REQUIRED_DATA     = ["vulns"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute overall and per-severity patch compliance rates.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame. Must have ``state``,
            ``severity``, and ``first_found`` columns.
        assets_df : pd.DataFrame
            Not used; accepted for interface compatibility.
        report_date : datetime
            UTC-aware reference date for days_open calculation.
        config : ModuleConfig
            Options: ``severity_filter`` (list[str]).
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            # ---- Resolve severity filter from config ----
            sev_filter_raw = config.options.get("severity_filter")
            if sev_filter_raw:
                active_sevs = tuple(
                    s.lower() for s in sev_filter_raw
                    if s.lower() in _SEVERITIES
                )
                if not active_sevs:
                    logger.warning(
                        "%s severity_filter contains no valid severities: %r. "
                        "Falling back to all severities.",
                        self._log_prefix(), sev_filter_raw,
                    )
                    active_sevs = _SEVERITIES
            else:
                active_sevs = _SEVERITIES

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            # ---- Empty DataFrame guard ----
            if vulns_df.empty or "state" not in vulns_df.columns:
                return self._build_result(
                    overall_rate=None,
                    per_sev_rates={sev: None for sev in active_sevs},
                    per_sev_counts={sev: {"within_sla": 0, "total": 0} for sev in active_sevs},
                    within_total=0,
                    total_open=0,
                    active_sevs=active_sevs,
                    computed_at=computed_at,
                    config=config,
                )

            # ---- Filter to open state and active severities ----
            open_df = vulns_df[
                vulns_df["state"].str.lower().isin(_OPEN_STATES) &
                vulns_df["severity"].str.lower().isin(active_sevs)
            ].copy()

            total_open = len(open_df)

            if total_open == 0:
                return self._build_result(
                    overall_rate=None,
                    per_sev_rates={sev: None for sev in active_sevs},
                    per_sev_counts={sev: {"within_sla": 0, "total": 0} for sev in active_sevs},
                    within_total=0,
                    total_open=0,
                    active_sevs=active_sevs,
                    computed_at=computed_at,
                    config=config,
                )

            # ---- Compute days_open and within-SLA mask ----
            report_ts   = pd.Timestamp(report_date)
            first_found = pd.to_datetime(open_df["first_found"], utc=True, errors="coerce")
            days_open   = (report_ts - first_found).dt.days.fillna(0).astype(int)
            sla_series  = open_df["severity"].str.lower().map(SLA_DAYS)
            within_mask = days_open <= sla_series

            within_total = int(within_mask.sum())
            overall_rate = round(within_total / total_open * 100, 1)

            # ---- Per-severity breakdown ----
            per_sev_rates: dict[str, Optional[float]] = {}
            per_sev_counts: dict[str, dict] = {}

            for sev in active_sevs:
                sev_mask   = open_df["severity"].str.lower() == sev
                sev_total  = int(sev_mask.sum())
                sev_within = int((within_mask & sev_mask).sum())

                per_sev_counts[sev] = {"within_sla": sev_within, "total": sev_total}
                per_sev_rates[sev]  = (
                    round(sev_within / sev_total * 100, 1) if sev_total > 0 else None
                )

            return self._build_result(
                overall_rate=overall_rate,
                per_sev_rates=per_sev_rates,
                per_sev_counts=per_sev_counts,
                within_total=within_total,
                total_open=total_open,
                active_sevs=active_sevs,
                computed_at=computed_at,
                config=config,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # Internal builder
    # ------------------------------------------------------------------

    def _build_result(
        self,
        overall_rate:    Optional[float],
        per_sev_rates:   dict[str, Optional[float]],
        per_sev_counts:  dict[str, dict],
        within_total:    int,
        total_open:      int,
        active_sevs:     tuple[str, ...],
        computed_at:     str,
        config:          ModuleConfig,
    ) -> ModuleData:
        """Assemble ModuleData from computed values."""

        metrics: dict[str, Any] = {
            "overall_rate":  overall_rate,
            "critical_rate": per_sev_rates.get("critical"),
            "high_rate":     per_sev_rates.get("high"),
            "medium_rate":   per_sev_rates.get("medium"),
            "low_rate":      per_sev_rates.get("low"),
            "within_sla":    within_total,
            "overdue":       total_open - within_total,
            "total_open":    total_open,
        }

        table_data = [
            {
                "severity":  sev.capitalize(),
                "total":     per_sev_counts.get(sev, {}).get("total", 0),
                "within_sla": per_sev_counts.get(sev, {}).get("within_sla", 0),
                "overdue":   (
                    per_sev_counts.get(sev, {}).get("total", 0)
                    - per_sev_counts.get(sev, {}).get("within_sla", 0)
                ),
                "rate":      per_sev_rates.get(sev),
            }
            for sev in _SEVERITIES   # always all four for consistent table structure
        ]

        if overall_rate is None:
            summary_text = "No open vulnerabilities found."
        else:
            summary_text = (
                f"Overall patch compliance is {overall_rate:.1f}% — "
                f"{within_total:,} of {total_open:,} open vulnerabilities "
                f"are within their remediation SLA target."
            )

        return ModuleData(
            module_id    = self.MODULE_ID,
            display_name = self.DISPLAY_NAME,
            metrics      = metrics,
            table_data   = table_data,
            chart_data   = {
                "overall_rate":  overall_rate,
                "per_severity":  {sev: per_sev_rates.get(sev) for sev in _SEVERITIES},
                "thresholds":    {"green": _THRESHOLD_GREEN, "amber": _THRESHOLD_AMBER},
            },
            summary_text = summary_text,
            metadata     = {
                "sla_windows": dict(SLA_DAYS),
                "calculation": "within_sla / total_open × 100",
                "computed_at": computed_at,
            },
            error        = None,
        )

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render one large compliance gauge plus four per-severity rate tiles.

        Returns an error callout if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m            = data.metrics
        overall_rate = m.get("overall_rate")

        # ---- Main gauge ----
        if overall_rate is None:
            gauge_html = '<div class="gauge-na" style="text-align:center;font-size:10pt;padding:20pt;">N/A — No open vulnerabilities</div>'
        else:
            b64 = draw_gauge(
                value=overall_rate,
                min_val=0,
                max_val=100,
                thresholds=_COMPLIANCE_THRESHOLDS,
                title="Overall Patch Compliance",
                unit="%",
            )
            gauge_html = (
                f'<div style="text-align:center;">'
                f'<img src="data:image/png;base64,{b64}" '
                f'style="width:100%;max-width:220pt;">'
                f"</div>"
            )

        # ---- Per-severity rate tiles ----
        sev_tiles = ""
        for sev in _SEVERITIES:
            rate = m.get(f"{sev}_rate")
            rate_str = f"{rate:.1f}%" if rate is not None else "N/A"
            color = _SEV_COLORS.get(sev, "#9E9E9E")
            sev_tiles += (
                f'<div style="display:inline-block;text-align:center;'
                f'width:22%;margin:0 1%;padding:4pt;">'
                f'<span style="font-size:11pt;font-weight:bold;color:{color};">'
                f"{rate_str}</span><br/>"
                f'<span style="font-size:7pt;color:#555;">{sev.capitalize()}</span>'
                f"</div>"
            )

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  {gauge_html}
  <div style="text-align:center;margin-top:6pt;">
    {sev_tiles}
  </div>
  <p class="explanatory-text">
    Patch Compliance Rate measures the percentage of open vulnerabilities that
    are still within their required remediation timeframe (15 days for Critical,
    30 days for High, 90 days for Medium, 180 days for Low). A rate below 90%
    means a significant portion of vulnerabilities have exceeded their target
    remediation date.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a "Patch Compliance" tab with per-severity rates and colour coding.

        Returns ``[]`` on error; writes an error row if ``data.error`` is set.
        """
        tab_name = "Patch Compliance"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            # ---- Headers ----
            headers = ["Severity", "Total Open", "Within SLA", "Overdue", "Compliance Rate"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            m = data.metrics

            # ---- Data rows ----
            for row_idx, row in enumerate(data.table_data, start=2):
                ws.cell(row=row_idx, column=1, value=row["severity"])
                ws.cell(row=row_idx, column=2, value=row["total"])
                ws.cell(row=row_idx, column=3, value=row["within_sla"])
                ws.cell(row=row_idx, column=4, value=row["overdue"])

                rate     = row["rate"]
                rate_str = f"{rate:.1f}%" if rate is not None else "N/A"
                rate_cell = ws.cell(row=row_idx, column=5, value=rate_str)
                rate_cell.fill = _rate_fill(rate)

            # ---- Total row ----
            total_row = len(data.table_data) + 2
            ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=m.get("total_open", 0)).font = Font(bold=True)
            ws.cell(row=total_row, column=3, value=m.get("within_sla", 0)).font = Font(bold=True)
            ws.cell(row=total_row, column=4, value=m.get("overdue", 0)).font = Font(bold=True)

            overall = m.get("overall_rate")
            overall_str  = f"{overall:.1f}%" if overall is not None else "N/A"
            overall_cell = ws.cell(row=total_row, column=5, value=overall_str)
            overall_cell.font = Font(bold=True)
            overall_cell.fill = _rate_fill(overall)

            # ---- Column widths ----
            widths = [14, 12, 12, 10, 18]
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_kpis
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return overall compliance rate, within-SLA count, and overdue count.
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m    = data.metrics
        rate = m.get("overall_rate")
        return {
            "Patch Compliance": f"{rate:.1f}%" if rate is not None else "N/A",
            "Within SLA":       str(m.get("within_sla", 0)),
            "Overdue":          str(m.get("overdue", 0)),
        }

    # ------------------------------------------------------------------
    # validate_config
    # ------------------------------------------------------------------

    def validate_config(self, config: ModuleConfig) -> list[str]:
        """Validate the optional ``severity_filter`` option."""
        errors: list[str] = []
        sev_filter = config.options.get("severity_filter")
        if sev_filter is not None:
            if not isinstance(sev_filter, list):
                errors.append(
                    f"patch_compliance_rate: 'severity_filter' must be a list, "
                    f"got {type(sev_filter).__name__}"
                )
            else:
                invalid = [s for s in sev_filter if s.lower() not in _SEVERITIES]
                if invalid:
                    errors.append(
                        f"patch_compliance_rate: unknown severity value(s) in "
                        f"severity_filter: {invalid}. Valid: {list(_SEVERITIES)}"
                    )
        return errors

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Compliance Rate": (
                    "(count of open findings where days_open <= SLA days for their severity) "
                    "/ (total open findings) × 100"
                ),
                "days_open": (
                    "(report_date - first_found).days where first_found is the UTC datetime "
                    "the finding was first observed by Tenable. NaN first_found is treated "
                    "as days_open = 0 (always within SLA)."
                ),
                "SLA Windows": (
                    f"Critical={SLA_DAYS['critical']}d, High={SLA_DAYS['high']}d, "
                    f"Medium={SLA_DAYS['medium']}d, Low={SLA_DAYS['low']}d "
                    "— defined in config.SLA_DAYS"
                ),
                "Data Source": (
                    "tio.exports.vulns() bulk export, state in {'open', 'reopened'} only."
                ),
                "Exclusions": (
                    "Fixed findings excluded. Informational findings excluded upstream by fetcher."
                ),
            },
        }
