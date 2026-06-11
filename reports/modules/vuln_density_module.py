"""
reports/modules/vuln_density_module.py — Vulnerability Density MoM trend module.

Displays open vulns-per-asset trended month-over-month.  Each historical
density point divides by THAT snapshot's own ``on_time_asset_count``
(Phase-14 D-01/D-02; new field from plan 15-02) — never the current
``len(assets_df)`` (Pitfall 4 — retroactive drift).

Denominator source
------------------
- **Historical points:** ``snapshot["on_time_asset_count"]`` — the count
  frozen at capture time.  Snapshots missing this field (older captures
  pre-15-02) are SKIPPED for density computation; they do not crash and
  they do not silently produce wrong values (D-15-06 backward-compat).
- **Current-run denominator:** ``count_on_time_assets(assets_df, report_date)``
  (D-14 None-sentinel).  When this returns None, the module returns
  ``_empty_result`` — no division is attempted.

Drift detection (success criterion 2)
--------------------------------------
When the last two usable snapshots' ``on_time_asset_count`` differ by more
than 10%, ``metadata["denom_drift_flag"]`` is set to ``True`` so downstream
email/PDF renderers can surface a "fleet size changed" note.

Cold-start (QUAL-01)
--------------------
When ``trend_snapshots`` is absent, ``insufficient_data=True``, or all
snapshots lack a valid ``on_time_asset_count``, returns a coherent
``_build_cold_start_result()`` — ``error=None``, ``metrics["cold_start"]=True``,
no ``NaN`` in any channel.

Partial-month label (D-15-08)
------------------------------
The current in-progress month is labeled "<YYYY-MM> (MTD — partial)" in all
four channels.

RAG defaults (D-15-07)
-----------------------
  Green   — density  < 2.0 vulns/asset
  Yellow  — 2.0 <= density < 4.0
  Red     — density >= 4.0
Thresholds are overridable via ``config.options`` (``green_density_threshold``,
``yellow_density_threshold``); direction is ``"lower_is_better"``.

Supported options
-----------------
green_density_threshold  : float  (default 2.0)
yellow_density_threshold : float  (default 4.0)
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.board_report_utils import extract_owner
from reports.modules.format_utils import safe_format, safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR,
    STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level RAG threshold defaults (D-15-07)
# ---------------------------------------------------------------------------
_DEFAULT_GREEN_DENSITY:  float = 2.0   # vulns/asset
_DEFAULT_YELLOW_DENSITY: float = 4.0   # vulns/asset

# ---------------------------------------------------------------------------
# Excel fills
# ---------------------------------------------------------------------------
_FILL_HEADER = PatternFill("solid", fgColor="E8F5E9")
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY   = PatternFill("solid", fgColor="F5F5F5")


def _rag_fill(status: str) -> PatternFill:
    if status == "green":
        return _FILL_GREEN
    if status == "yellow":
        return _FILL_AMBER
    if status == "red":
        return _FILL_RED
    return _FILL_GREY


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _month_label(month_str: str, current_period: pd.Period) -> str:
    """
    Return a human-readable label for a snapshot month string.

    Appends "(MTD — partial)" to the current in-progress period (D-15-08)
    so recipients know the point is not a full month.

    Parameters
    ----------
    month_str : str
        "YYYY-MM" string from the trend snapshot.
    current_period : pd.Period
        The current month period (derived from report_date at compute() time).

    Returns
    -------
    str
        Plain "YYYY-MM" for completed months; "YYYY-MM (MTD — partial)" for
        the current month.
    """
    try:
        snap_period = pd.Period(month_str, "M")
        if snap_period == current_period:
            return f"{month_str} (MTD — partial)"
    except Exception:  # noqa: BLE001
        pass
    return month_str


def _open_count_from_snap(snap: dict) -> int:
    """Sum the four severity buckets in a snapshot to get total open count."""
    return sum(snap.get(sev, 0) or 0 for sev in ("critical", "high", "medium", "low"))


# ===========================================================================
# Module
# ===========================================================================

@register_module
class VulnDensityModule(BaseModule):
    """
    Vulnerability Density — open vulns per on-time-scanned licensed asset,
    trended month-over-month.

    Each historical point uses THAT snapshot's own ``on_time_asset_count``
    denominator (D-01/D-02, Pitfall 4).  Denominator MoM change >10% is
    flagged so readers can distinguish real exposure changes from fleet-size
    changes.

    Cold-start: when fewer than 2 snapshots have a usable ``on_time_asset_count``
    field, returns a coherent cold-start ModuleData (QUAL-01).

    RAG: Green < 2.0 vulns/asset; Yellow 2.0–4.0; Red > 4.0 (D-15-07).
    """

    MODULE_ID         = "vuln_density"
    DISPLAY_NAME      = "Vulnerability Density"
    DESCRIPTION       = "Open vulns per on-time-scanned licensed asset, trended MoM."
    REQUIRED_DATA     = ["vulns", "assets", "trend_snapshots"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # _build_cold_start_result
    # ------------------------------------------------------------------

    def _build_cold_start_result(self, config: ModuleConfig) -> ModuleData:
        """
        Return a coherent cold-start ModuleData — not an error, just
        insufficient history (QUAL-01).

        ``error`` is ``None`` (cold-start is a valid operational state,
        not a failure); ``metrics["cold_start"]`` is ``True`` so renderers
        can branch on it without parsing the summary text.
        """
        return ModuleData(
            module_id        = self.MODULE_ID,
            display_name     = self.DISPLAY_NAME,
            metrics          = {"cold_start": True},
            table_data       = [],
            chart_data       = {},
            summary_text     = "Trend data being established — available from next month.",
            metadata         = {"cold_start": True},
            driver_narrative = "Trend data being established.",
            analyst_rows     = [],
            rag_strip        = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            ),
            error            = None,
        )

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute MoM vulnerability density from trend snapshots.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Used for Owner-cut context
            via extract_owner.  Requires columns: first_found, asset_uuid, state.
        assets_df : pd.DataFrame
            Asset DataFrame for current-run denominator via
            count_on_time_assets() and Owner cut via extract_owner().
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            ``options`` may contain ``green_density_threshold`` and
            ``yellow_density_threshold`` (D-15-07).
        **kwargs
            ``trend_snapshots`` (dict, optional) — full ``read_trend()`` result:
            ``{"snapshots": [...], "insufficient_data": bool}``.
            When absent or ``insufficient_data=True``, returns cold-start result
            (QUAL-01).
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d, assets_df rows: %d",
            self._log_prefix(), len(vulns_df), len(assets_df),
        )

        try:
            # ----------------------------------------------------------------
            # QUAL-01 — cold-start guard (insufficient trend history)
            # ----------------------------------------------------------------
            trend_snapshots = kwargs.get("trend_snapshots")
            cold_start = (
                trend_snapshots is None
                or trend_snapshots.get("insufficient_data", True)
            )
            if cold_start:
                logger.info(
                    "%s trend_snapshots absent or insufficient_data=True — cold start.",
                    self._log_prefix(),
                )
                return self._build_cold_start_result(config)

            snapshots: list[dict] = trend_snapshots.get("snapshots", [])
            if not snapshots:
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # Current-run denominator (D-14 None-sentinel — Pitfall 4 guard)
            # count_on_time_assets returns None if no on-time licensed assets.
            # ----------------------------------------------------------------
            from utils.asset_count import count_on_time_assets  # noqa: PLC0415

            current_denom = count_on_time_assets(assets_df, report_date)
            if current_denom is None:
                logger.warning(
                    "%s count_on_time_assets returned None — no on-time licensed assets "
                    "in scope.  Returning empty result (no division).",
                    self._log_prefix(),
                )
                return self._empty_result(
                    "No on-time-scanned licensed assets in scope.", config
                )

            # ----------------------------------------------------------------
            # Current period for partial-month label (D-15-08)
            # ----------------------------------------------------------------
            current_period = pd.Period(report_date, "M")

            # ----------------------------------------------------------------
            # RAG thresholds — module_options-overridable (D-15-07)
            # ----------------------------------------------------------------
            green_threshold  = float(
                config.options.get("green_density_threshold",  _DEFAULT_GREEN_DENSITY)
            )
            yellow_threshold = float(
                config.options.get("yellow_density_threshold", _DEFAULT_YELLOW_DENSITY)
            )

            # ----------------------------------------------------------------
            # Per-snapshot density computation
            # CRITICAL: use snap["on_time_asset_count"] — NEVER len(assets_df)
            # for historical points (Pitfall 4 — retroactive drift).
            # Snapshots with on_time_asset_count None or 0 are SKIPPED (D-15-06).
            # ----------------------------------------------------------------
            months_labels:    list[str]   = []
            density_series:   list[float] = []
            open_counts:      list[int]   = []
            denom_series:     list[int]   = []   # usable denoms for drift check
            table_data:       list[dict]  = []

            for snap in snapshots:
                month_str  = snap.get("month", "")
                label      = _month_label(month_str, current_period)
                open_count = _open_count_from_snap(snap)

                snap_denom = snap.get("on_time_asset_count")
                if not snap_denom:
                    # None or 0 → skip density for this point (cold-start D-15-06);
                    # still add a row to table_data with "—" density
                    months_labels.append(label)
                    open_counts.append(open_count)
                    table_data.append({
                        "month":       label,
                        "open_count":  open_count,
                        "asset_count": snap.get("on_time_asset_count"),
                        "density":     None,
                    })
                    continue

                density = open_count / snap_denom
                months_labels.append(label)
                open_counts.append(open_count)
                density_series.append(density)
                denom_series.append(snap_denom)
                table_data.append({
                    "month":       label,
                    "open_count":  open_count,
                    "asset_count": snap_denom,
                    "density":     round(density, 2),
                })

            # ----------------------------------------------------------------
            # Cold-start: fewer than 2 usable density points
            # ----------------------------------------------------------------
            if len(density_series) < 1:
                logger.info(
                    "%s no usable on_time_asset_count in snapshots — cold start.",
                    self._log_prefix(),
                )
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # Denominator drift flag (success criterion 2)
            # Compare last two USABLE denoms; abs delta > 10% → flag True.
            # ----------------------------------------------------------------
            denom_drift_flag = False
            if len(denom_series) >= 2:
                prev_denom = denom_series[-2]
                curr_denom = denom_series[-1]
                denom_shift_pct = abs(curr_denom - prev_denom) / max(prev_denom, 1) * 100.0
                if denom_shift_pct > 10.0:
                    denom_drift_flag = True
                    logger.info(
                        "%s denominator drift %.1f%% (prev=%d curr=%d) — flagging.",
                        self._log_prefix(), denom_shift_pct, prev_denom, curr_denom,
                    )

            # ----------------------------------------------------------------
            # Current density (last usable point in density_series)
            # ----------------------------------------------------------------
            current_density = density_series[-1]

            # ----------------------------------------------------------------
            # RAG status (lower_is_better, D-15-07)
            # ----------------------------------------------------------------
            rag_status = rag_status_from_value(
                current_density, green_threshold, yellow_threshold,
                direction="lower_is_better",
            )
            rag_headline = safe_format(current_density, ".1f")

            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = f"{rag_headline} vulns/asset",
                status             = rag_status,
            )

            # ----------------------------------------------------------------
            # Owner cut — current open findings per owner
            # ----------------------------------------------------------------
            from utils.open_count import open_findings_at  # noqa: PLC0415

            open_df = open_findings_at(vulns_df, report_date)
            enriched_assets = extract_owner(assets_df)
            uuid_to_owner = dict(
                zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
            )
            if not open_df.empty and "asset_uuid" in open_df.columns:
                owner_col = open_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                owner_counts: dict[str, int] = owner_col.value_counts().to_dict()
            else:
                owner_counts = {}

            # Per-Owner density (aggregate counts only — QUAL-05)
            owner_df_rows: list[dict] = []
            for owner, cnt in owner_counts.items():
                owner_density = cnt / current_denom if current_denom else None
                owner_df_rows.append({
                    "owner":         owner,
                    "open_count":    cnt,
                    "density":       round(owner_density, 2) if owner_density is not None else None,
                    "asset_count":   current_denom,
                })
            owner_density_df = pd.DataFrame(
                owner_df_rows,
                columns=["owner", "open_count", "density", "asset_count"],
            ) if owner_df_rows else pd.DataFrame(
                columns=["owner", "open_count", "density", "asset_count"]
            )

            # ----------------------------------------------------------------
            # chart_data
            # ----------------------------------------------------------------
            chart_data = {
                "months":    months_labels,
                "density":   [
                    r["density"] if r["density"] is not None else 0.0
                    for r in table_data
                ],
                "open_count": open_counts,
            }

            # ----------------------------------------------------------------
            # Driver narrative (CONTRACT-01)
            # ----------------------------------------------------------------
            rag_color_label = STATUS_LABEL.get(rag_status, "")
            driver_narrative = (
                f"Current density: {safe_format(current_density, '.2f')} vulns/asset "
                f"({rag_color_label}). "
                + (
                    "Fleet-size change >10% detected — interpret density trend with caution. "
                    if denom_drift_flag else ""
                )
                + (
                    f"Top owners: {', '.join(list(owner_counts.keys())[:3])}."
                    if owner_counts else "No open findings in scope."
                )
            )

            # ----------------------------------------------------------------
            # Summary text
            # ----------------------------------------------------------------
            summary_text = (
                f"Vulnerability Density over {len(density_series)} usable month(s). "
                f"Current: {safe_format(current_density, '.2f')} vulns/asset. "
                + (
                    "Fleet-size drift >10% flagged. "
                    if denom_drift_flag else ""
                )
            )

            # ----------------------------------------------------------------
            # Analyst rows (CONTRACT-02 — aggregate only, QUAL-05)
            # ----------------------------------------------------------------
            mom_df = pd.DataFrame(table_data)
            analyst_rows: list[tuple[str, pd.DataFrame]] = [
                ("Density by Month", mom_df),
                ("Density by Owner", owner_density_df),
            ]

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "density_series":   density_series,
                    "current_density":  current_density,
                    "rag_status":       rag_status,
                    "cold_start":       False,
                    "owner_counts":     owner_counts,
                    "months":           len(density_series),
                },
                table_data   = table_data,
                chart_data   = chart_data,
                summary_text = summary_text,
                metadata     = {
                    "denom_drift_flag": denom_drift_flag,
                    "current_period":   str(current_period),
                    "current_denom":    current_denom,
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a density summary section with MoM trend table.

        Returns an error callout if ``data.error`` is set, and a
        cold-start notice if ``metrics["cold_start"]`` is True.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        if data.metrics.get("cold_start"):
            return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p style="color:#757575;">{data.summary_text}</p>
</div>"""

        m           = data.metrics
        rag_status  = m.get("rag_status", "no_data")
        rag_color   = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        rag_label   = STATUS_LABEL.get(rag_status, STATUS_LABEL["no_data"])
        curr_d      = m.get("current_density")
        drift_flag  = data.metadata.get("denom_drift_flag", False)
        drift_note  = (
            '<p style="color:#f57c00;"><strong>Note:</strong> Fleet size changed '
            '&gt;10% between last two periods — interpret density trend with caution.</p>'
            if drift_flag else ""
        )

        table_rows = ""
        for row in data.table_data:
            dens_str = safe_format(row.get("density"), ".2f") if row.get("density") is not None else "—"
            table_rows += (
                f"<tr>"
                f"<td>{row['month']}</td>"
                f"<td style='text-align:right;'>{safe_int(row['open_count'])}</td>"
                f"<td style='text-align:right;'>{safe_int(row.get('asset_count'))}</td>"
                f"<td style='text-align:right;'>{dens_str}</td>"
                f"</tr>"
            )

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p>
    <strong>RAG Status:</strong>
    <span style="color:{rag_color};font-weight:bold;">{rag_label}</span>
    &nbsp;|&nbsp;
    <strong>Current Density:</strong> {safe_format(curr_d, '.2f')} vulns/asset
  </p>
  {drift_note}
  <table class="data-table" style="width:100%;margin-top:8pt;">
    <thead>
      <tr>
        <th>Month</th>
        <th style="text-align:right;">Open Vulns</th>
        <th style="text-align:right;">On-Time Assets</th>
        <th style="text-align:right;">Density</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>
  <p class="explanatory-text">
    Density = open vulns / on-time-scanned licensed assets.
    Each month uses its own asset count (historical point is not retroactively adjusted).
    Months labeled "(MTD — partial)" are in-progress.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a ``Vuln Density`` tab with MoM density trend.

        Returns ``[]`` on exception; writes an error/cold-start row if set.
        """
        tab_name = "Vuln Density"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            if data.metrics.get("cold_start"):
                ws["A1"] = self.DISPLAY_NAME
                ws["A1"].font = Font(bold=True, size=13)
                ws["A2"] = data.summary_text
                ws["A2"].font = Font(italic=True, color="757575")
                return [tab_name]

            ws["A1"] = "Vulnerability Density — Monthly Trend"
            ws["A1"].font = Font(bold=True, size=13)

            drift_flag = data.metadata.get("denom_drift_flag", False)
            if drift_flag:
                ws["A2"] = (
                    "Note: Fleet size changed >10% between last two periods — "
                    "interpret trend with caution."
                )
                ws["A2"].font = Font(italic=True, color="F57C00")

            headers = ["Month", "Open Vulns", "On-Time Assets", "Density (vulns/asset)"]
            header_row = 4 if drift_flag else 3
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _FILL_HEADER

            for row_idx, row in enumerate(data.table_data, start=header_row + 1):
                dens = row.get("density")
                dens_str = f"{dens:.2f}" if dens is not None else "—"
                ws.cell(row=row_idx, column=1, value=row["month"])
                ws.cell(row=row_idx, column=2, value=row.get("open_count", 0))
                ws.cell(row=row_idx, column=3, value=row.get("asset_count"))
                ws.cell(row=row_idx, column=4, value=dens_str)

                rag_status = data.metrics.get("rag_status", "no_data")
                ws.cell(row=row_idx, column=4).fill = _rag_fill(rag_status)

            widths = [30, 14, 18, 22]
            for col_idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = w

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Returns ``""`` when ``data.error`` is set.
        Cold-start renders the "Trend data being established" notice.
        Uses only ``safe_format`` / ``safe_pct`` / ``safe_int`` — never
        raw f-string format specs on possibly-None values.
        """
        if data.error:
            return ""

        if data.metrics.get("cold_start"):
            return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#F5F5F5;border-left:4px solid #757575;">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;color:#757575;">{data.summary_text}</span>
    </td>
  </tr>
</table>"""

        if not data.driver_narrative:
            return ""

        m           = data.metrics
        rag_status  = m.get("rag_status", "no_data")
        rag_color   = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        curr_d      = m.get("current_density")
        months      = safe_int(m.get("months"))
        drift_flag  = data.metadata.get("denom_drift_flag", False)
        drift_note  = (
            "<br><em style='font-size:11px;color:#f57c00;'>"
            "Fleet-size change &gt;10% detected — interpret trend with caution.</em>"
            if drift_flag else ""
        )

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#E8F5E9;border-left:4px solid {rag_color};">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;">
        {months} month(s) of trend &mdash; current density: {safe_format(curr_d, '.2f')} vulns/asset
      </span><br>
      <em style="font-size:11px;color:#555;">{data.driver_narrative}</em>
      {drift_note}
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns ``data.analyst_rows`` when data is valid; ``[]`` on error or
        cold-start (no detail to drill down into).
        Columns are aggregate-only (QUAL-05 — no per-finding PII).
        """
        if data.error or not data.analyst_rows:
            return []
        if data.metrics.get("cold_start"):
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present;
        falls back to a gray "No Data" cell if empty or error.
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip
