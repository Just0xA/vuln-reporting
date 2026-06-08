"""
reports/modules/vuln_type_distribution_module.py — Vulnerability Type Distribution.

Classifies open findings within a Tenable tag into four CPE type buckets:

    Application / OS / Hardware / Other

using the VTD-01 family-override classifier (spike 001, ``vuln-metric-substrate.md``).

Classification logic
--------------------
1. If ``plugin_family`` matches the OS distro/bulletin regex → **OS**
   (family override first — covers Linux "Local Security Checks" that carry
   ``cpe:/a:`` app CPEs but are OS-team work; and Microsoft Bulletins → OS).
2. Else if any CPE part letter is ``a`` → **Application**  (a > o > h precedence)
3. Else if any CPE part letter is ``o`` → **OS**
4. Else if any CPE part letter is ``h`` → **Hardware**
5. Else → **Other**

Coverage: ~99.2% of Critical+High findings classify on real Tenable data.

Denominator
-----------
Within-tag counts and percentages (D5): pct = count / tag_total * 100.
This is NOT a share of the environment (cf. tag_severity_share_module).

Hardware row
------------
The Hardware row/tile is **omitted** from all rendered output when its count
is 0 (typical in production environments per spike findings).

RAG strip
---------
Headline = Application share of tag findings (% within tag).  Status is
``"no_data"`` when tag has zero findings, otherwise ``"yellow"`` (informational
snapshot; no threshold locked by spec).
"""

from __future__ import annotations

import logging
import re
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from reports.modules import (
    safe_pct, safe_int, safe_format,
    build_rag_strip_entry, NO_DATA_HEADLINE, NO_DATA_DRIVER,
)

logger = logging.getLogger(__name__)

# ── Open states ───────────────────────────────────────────────────────────────
_OPEN_STATES: frozenset[str] = frozenset({"open", "reopened"})

# ── VTD-01 classifier config ──────────────────────────────────────────────────
# CPE part letter, both 2.2 (cpe:/a:...) and 2.3 (cpe:2.3:a:...)
_CPE_PART = re.compile(r"cpe:(?:2\.3:|/)([aoh])[:/]", re.IGNORECASE)

PART_TO_BUCKET: dict[str, str] = {
    "a": "Application",
    "o": "OS",
    "h": "Hardware",
}

# Families whose findings are OS-team work regardless of CPE part letter.
# Microsoft Bulletin → OS (decision from requestor per vuln-metric-substrate.md).
OS_FAMILY = re.compile(
    r"local security checks|red hat|centos|oracle linux|ubuntu|debian|"
    r"suse|amazon linux|rocky|alma|fedora|microsoft bulletin",
    re.IGNORECASE,
)

# Display order for rows/tiles
_BUCKETS: tuple[str, ...] = ("Application", "OS", "Hardware", "Other")

# Bucket colours for Excel/PDF (informational palette)
_BUCKET_COLORS: dict[str, str] = {
    "Application": "#1976d2",  # Info blue
    "OS":          "#388e3c",  # Low green
    "Hardware":    "#f57c00",  # High orange
    "Other":       "#9e9e9e",  # Grey
}
_BUCKET_TEXT: dict[str, str] = {
    "Application": "#ffffff",
    "OS":          "#ffffff",
    "Hardware":    "#ffffff",
    "Other":       "#ffffff",
}


# ── Classifier (importable for unit tests per plan) ───────────────────────────

def classify(plugin_family: str, cpe: str) -> str:
    """
    Classify a single finding into Application / OS / Hardware / Other.

    Parameters
    ----------
    plugin_family : str
        The Tenable plugin family string (e.g. "Red Hat Local Security Checks").
        May be ``None`` or empty — treated as no family.
    cpe : str
        Comma-joined CPE string from ``vulns_df["cpe"]`` (see fetchers.py:343).
        May be ``None`` or empty — treated as no CPE.

    Returns
    -------
    str
        One of ``"Application"``, ``"OS"``, ``"Hardware"``, ``"Other"``.

    Notes
    -----
    - Family override fires first (step 1 in VTD-01).
    - CPE part-letter precedence: a > o > h (mixed-CPE rows).
    - Fallback label is ``"Other"`` (spec D4) — NOT ``"Unclassified"``.
    """
    # Step 1: family override
    if plugin_family and OS_FAMILY.search(plugin_family):
        return "OS"

    # Step 2-4: CPE part letter with a>o>h precedence
    cpe_str  = cpe or ""
    parts    = {m.lower() for m in _CPE_PART.findall(cpe_str)}
    for p in ("a", "o", "h"):
        if p in parts:
            return PART_TO_BUCKET[p]

    # Step 5: fallback
    return "Other"


# ── Module ────────────────────────────────────────────────────────────────────

@register_module
class VulnTypeDistributionModule(BaseModule):
    """
    CPE type distribution of a Tenable tag's open findings.

    Uses the VTD-01 family-override classifier:
    - ``plugin_family`` OS/distro/Microsoft-Bulletin override first.
    - Then CPE part-letter ``a`` → Application, ``o`` → OS, ``h`` → Hardware.
    - Else → Other (D4 fallback label).
    - a > o > h precedence for mixed-CPE rows.

    Percentages are within-tag (denominator = tag_total, not environment).
    Hardware row/tile is omitted from rendered output when count == 0.

    Supported options
    -----------------
    None — this module accepts no options.
    """

    MODULE_ID         = "vuln_type_distribution"
    DISPLAY_NAME      = "Vulnerability Type Distribution"
    DESCRIPTION       = (
        "CPE type distribution (Application / OS / Hardware / Other) "
        "of a tag's open findings using the VTD-01 family-override classifier."
    )
    REQUIRED_DATA     = ["vulns"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Classify open findings and compute within-tag type distribution.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Must have ``state``,
            ``plugin_family``, and ``cpe`` columns.
        assets_df : pd.DataFrame
            Not used; accepted for interface compatibility.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Module configuration (no options consumed).

        Returns
        -------
        ModuleData
            On success: ``error`` is None and all data fields are populated.
            On failure: ``error`` is set and data fields hold safe empty defaults.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            # Restrict to open states
            if vulns_df.empty or "state" not in vulns_df.columns:
                open_df = pd.DataFrame(
                    columns=list(vulns_df.columns) if not vulns_df.empty
                    else ["state", "plugin_family", "cpe"]
                )
            else:
                open_df = vulns_df[
                    vulns_df["state"].str.lower().isin(_OPEN_STATES)
                ].copy()

            tag_total = len(open_df)

            # Classify each row
            if tag_total > 0:
                fam_col = open_df["plugin_family"] if "plugin_family" in open_df.columns else pd.Series([""] * tag_total)
                cpe_col = open_df["cpe"]           if "cpe" in open_df.columns else pd.Series([""] * tag_total)
                open_df = open_df.assign(
                    _vuln_type=pd.Series(
                        [classify(str(f) if pd.notna(f) else "",
                                  str(c) if pd.notna(c) else "")
                         for f, c in zip(fam_col, cpe_col)],
                        index=open_df.index,
                    )
                )
                counts: dict[str, int] = {
                    bucket: int((open_df["_vuln_type"] == bucket).sum())
                    for bucket in _BUCKETS
                }
            else:
                counts = {bucket: 0 for bucket in _BUCKETS}

            # Within-tag pct (D5)
            pcts: dict[str, float] = {}
            for bucket in _BUCKETS:
                pcts[bucket] = (
                    counts[bucket] / tag_total * 100.0
                    if tag_total > 0 else 0.0
                )

            hide_hardware = (counts["Hardware"] == 0)

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            metrics: dict[str, Any] = {
                "tag_total":     tag_total,
                "hide_hardware": hide_hardware,
            }
            for bucket in _BUCKETS:
                metrics[f"{bucket.lower()}_count"] = counts[bucket]
                metrics[f"{bucket.lower()}_pct"]   = pcts[bucket]

            # ── driver narrative ──────────────────────────────────────────
            if tag_total == 0:
                driver = NO_DATA_DRIVER
            else:
                dominant = max(_BUCKETS, key=lambda b: counts[b])
                app_share = safe_pct(pcts["Application"])
                driver = (
                    f"Of {safe_int(tag_total)} open findings, "
                    f"Application = {app_share}; "
                    f"dominant type: {dominant}."
                )

            # ── RAG strip (informational) ─────────────────────────────────
            if tag_total == 0:
                strip_status = "no_data"
                headline_str = NO_DATA_HEADLINE
            else:
                strip_status = "yellow"
                headline_str = safe_pct(pcts["Application"])  # Application share as headline

            rag_strip = build_rag_strip_entry(
                self.DISPLAY_NAME, headline_str, strip_status
            )

            # ── analyst rows ──────────────────────────────────────────────
            analyst_rows: list[tuple[str, pd.DataFrame]] = []
            if tag_total > 0 and "_vuln_type" in open_df.columns:
                analyst_df = open_df.drop(columns=["_vuln_type"], errors="ignore").assign(
                    vuln_type=open_df["_vuln_type"]
                )
                analyst_rows = [("VulnType Detail", analyst_df)]

            summary_text = (
                f"Of {tag_total:,} open findings in scope: "
                + ", ".join(
                    f"{b} {safe_pct(pcts[b])}"
                    for b in _BUCKETS
                    if not (b == "Hardware" and hide_hardware)
                )
                + "."
            )

            return ModuleData(
                module_id        = self.MODULE_ID,
                display_name     = self.DISPLAY_NAME,
                metrics          = metrics,
                table_data       = [],
                chart_data       = {},
                summary_text     = summary_text,
                metadata         = {
                    "tag_total":      tag_total,
                    "hide_hardware":  hide_hardware,
                    "classifier":     "VTD-01 family-override → CPE part a>o>h → Other",
                    "denominator":    "within-tag (tag_total, not environment)",
                    "computed_at":    computed_at,
                },
                driver_narrative = driver,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a Type | Count | % within Tag table, omitting Hardware when 0.

        Returns an error callout when ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m             = data.metrics
        hide_hardware = m.get("hide_hardware", False)

        rows_html = ""
        for bucket in _BUCKETS:
            if bucket == "Hardware" and hide_hardware:
                continue
            count = m.get(f"{bucket.lower()}_count", 0)
            pct   = m.get(f"{bucket.lower()}_pct", 0.0)
            bg    = _BUCKET_COLORS.get(bucket, "#9e9e9e")
            fg    = _BUCKET_TEXT.get(bucket, "#ffffff")
            rows_html += (
                f"<tr>"
                f'<td style="background-color:{bg};color:{fg};font-weight:bold;'
                f'padding:4px 8px;">{bucket}</td>'
                f'<td style="text-align:right;padding:4px 8px;">{safe_int(count)}</td>'
                f'<td style="text-align:right;padding:4px 8px;">{safe_pct(pct)}</td>'
                f"</tr>"
            )

        tag_total = m.get("tag_total", 0)

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p style="font-size:8pt;margin-bottom:6pt;">{data.driver_narrative}</p>
  <table style="width:100%;border-collapse:collapse;font-size:9pt;">
    <thead>
      <tr style="background-color:#eeeeee;">
        <th style="text-align:left;padding:4px 8px;">Type</th>
        <th style="text-align:right;padding:4px 8px;">Count</th>
        <th style="text-align:right;padding:4px 8px;">% within Tag</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
    <tfoot>
      <tr style="border-top:2px solid #333;">
        <td style="font-weight:bold;padding:4px 8px;">Tag Total</td>
        <td style="text-align:right;font-weight:bold;padding:4px 8px;">{safe_int(tag_total)}</td>
        <td style="text-align:right;font-weight:bold;padding:4px 8px;">100.0%</td>
      </tr>
    </tfoot>
  </table>
  <p style="font-size:7.5pt;color:#555;margin-top:4pt;">
    Classifier: plugin_family OS/distro/Microsoft-Bulletin override first; then CPE
    part a (Application) &gt; o (OS) &gt; h (Hardware); else Other.
    Denominator is within-tag (not vs environment).
    {"Hardware omitted (count = 0)." if hide_hardware else ""}
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a single "Vuln Type" tab, omitting Hardware when count == 0.

        Returns ``[]`` on error.
        """
        tab_name = "Vuln Type"
        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            headers = ["Type", "Count", "% within Tag"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            m             = data.metrics
            hide_hardware = m.get("hide_hardware", False)

            row_idx = 2
            for bucket in _BUCKETS:
                if bucket == "Hardware" and hide_hardware:
                    continue
                count = m.get(f"{bucket.lower()}_count", 0)
                pct   = m.get(f"{bucket.lower()}_pct", 0.0)

                type_cell = ws.cell(row=row_idx, column=1, value=bucket)
                hex_color = _BUCKET_COLORS.get(bucket, "#9e9e9e").lstrip("#")
                type_cell.fill = PatternFill("solid", fgColor=hex_color)
                fg_hex = _BUCKET_TEXT.get(bucket, "#ffffff").lstrip("#")
                type_cell.font = Font(bold=True, color=fg_hex)

                ws.cell(row=row_idx, column=2, value=count)
                ws.cell(row=row_idx, column=3, value=safe_pct(pct))
                row_idx += 1

            # Footer
            tag_total = m.get("tag_total", 0)
            footer_cell = ws.cell(row=row_idx, column=1, value="Tag Total")
            footer_cell.font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=tag_total).font = Font(bold=True)
            ws.cell(row=row_idx, column=3, value="100.0%").font = Font(bold=True)

            ws.column_dimensions[get_column_letter(1)].width = 16
            ws.column_dimensions[get_column_letter(2)].width = 10
            ws.column_dimensions[get_column_letter(3)].width = 15

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Inline-CSS email panel for the composed email body.

        Returns ``""`` when ``data.error`` is set.
        """
        if data.error:
            return ""

        m             = data.metrics
        hide_hardware = m.get("hide_hardware", False)

        rows_html = ""
        for bucket in _BUCKETS:
            if bucket == "Hardware" and hide_hardware:
                continue
            count = m.get(f"{bucket.lower()}_count", 0)
            pct   = m.get(f"{bucket.lower()}_pct", 0.0)
            bg    = _BUCKET_COLORS.get(bucket, "#9e9e9e")
            fg    = _BUCKET_TEXT.get(bucket, "#ffffff")
            rows_html += (
                f'<tr>'
                f'<td style="background-color:{bg};color:{fg};font-weight:bold;'
                f'padding:4px 8px;border:1px solid #ddd;">{bucket}</td>'
                f'<td style="text-align:right;padding:4px 8px;border:1px solid #ddd;">'
                f'{safe_int(count)}</td>'
                f'<td style="text-align:right;padding:4px 8px;border:1px solid #ddd;">'
                f'{safe_pct(pct)}</td>'
                f'</tr>'
            )

        tag_total = m.get("tag_total", 0)
        driver    = data.driver_narrative or ""

        return (
            f'<div style="font-family:Arial,sans-serif;margin-bottom:16px;">'
            f'<h3 style="margin:0 0 6px 0;font-size:13px;color:#333;">'
            f'{self.DISPLAY_NAME}</h3>'
            f'<p style="margin:0 0 8px 0;font-size:11px;color:#555;">{driver}</p>'
            f'<table style="width:100%;border-collapse:collapse;font-size:11px;">'
            f'<thead><tr style="background-color:#eeeeee;">'
            f'<th style="text-align:left;padding:4px 8px;border:1px solid #ddd;">Type</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">Count</th>'
            f'<th style="text-align:right;padding:4px 8px;border:1px solid #ddd;">% within Tag</th>'
            f'</tr></thead>'
            f'<tbody>{rows_html}</tbody>'
            f'<tfoot><tr style="border-top:2px solid #333;">'
            f'<td style="font-weight:bold;padding:4px 8px;border:1px solid #ddd;">Tag Total</td>'
            f'<td style="text-align:right;font-weight:bold;padding:4px 8px;border:1px solid #ddd;">'
            f'{safe_int(tag_total)}</td>'
            f'<td style="text-align:right;font-weight:bold;padding:4px 8px;border:1px solid #ddd;">'
            f'100.0%</td>'
            f'</tr></tfoot>'
            f'</table>'
            f'</div>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        Return the flat per-finding DataFrame for the analyst workbook.

        Returns ``[]`` on error or zero rows.
        """
        if data.error:
            return []
        return list(data.analyst_rows) if data.analyst_rows else []

    # ------------------------------------------------------------------
    # get_audit_info
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "Classifier (VTD-01)": (
                    "1. plugin_family matches OS_FAMILY regex (Linux distros + "
                    "Microsoft Bulletin families) → 'OS' regardless of CPE. "
                    "2. CPE part 'a' → 'Application'. "
                    "3. CPE part 'o' → 'OS'. "
                    "4. CPE part 'h' → 'Hardware'. "
                    "5. No match → 'Other'. "
                    "Precedence for mixed-CPE rows: a > o > h."
                ),
                "Family override rationale": (
                    "Linux 'Local Security Checks' (and distro-specific variants) "
                    "carry app CPEs (cpe:/a:<package>) but represent OS-team patch "
                    "work. Without the override, ~6% of OS volume lands on the "
                    "Application bucket (measured on real Tenable data, spike 001)."
                ),
                "Coverage": (
                    "~99.2% of Critical+High findings classify to Application/OS/Hardware; "
                    "~0.8% fall through to Other."
                ),
                "Denominator (D5)": (
                    "Within-tag: pct = count / tag_total * 100.0. "
                    "Divide-by-zero guard → 0.0 when tag_total == 0."
                ),
                "Hardware-hidden rule": (
                    "When Hardware count == 0, the row/tile is omitted from all "
                    "rendered output. Hardware is ~0 in most real production environments."
                ),
                "Open states": (
                    "state in {'open', 'reopened'} — informational findings excluded "
                    "upstream by the fetcher."
                ),
            },
        }
