"""
reports/modules/aged_vulns_assets_module.py — Aged Vulnerability Assets metric module.

Measures the percentage of on-time-scanned assets that carry at least one
Medium, High, or Critical vulnerability open for more than 90 days.  Assets
with long-standing unresolved findings indicate a persistent remediation gap
and represent enduring exposure even at lower severities.

Module ID:    aged_vulns_assets
Display Name: Aged Vulnerability Assets

SLA thresholds (board-defined, lower is better):
    Green:  aged_assets_pct <= 2%
    Yellow: aged_assets_pct <= 5%  (and > 2%)
    Red:    aged_assets_pct >  5%

Denominator: all deduplicated on-time-scanned assets (last_licensed_scan_date
             within the last 30 days).
Numerator:   subset with >= 1 Medium/High/Critical finding open > 90 days.
"""

from __future__ import annotations

import html
import logging
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from reports.modules.board_pdf_layout import two_column_metric_section
from config import RISK_WEIGHTS
from reports.modules.board_report_utils import (
    compute_bu_risk_scores,
    compute_per_bu_breakdown,
    deduplicate_assets_by_name,
    extract_owner,
    identify_on_time_assets,
    sla_status_from_thresholds,
    ON_TIME_WINDOW_DAYS,
)
from reports.modules.chart_utils import draw_gauge
from reports.modules.format_utils import safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR as _RAG_STATUS_COLOR,
    STATUS_ICON,
    STATUS_LABEL as _RAG_STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

_GREEN_THRESHOLD  = 2.0    # <= green (lower is better)
_YELLOW_THRESHOLD = 5.0    # <= yellow; > green
_DIRECTION        = "lower_is_better"

#: Findings open longer than this many days qualify as "aged".
_AGED_DAYS_THRESHOLD: int = 90

#: Minimum aged findings required for an asset to be included in the numerator.
_MIN_AGED_COUNT: int = 1

#: Severity tiers included in the aged-vuln scan.
_AGED_SEVERITIES: frozenset[str] = frozenset({"critical", "high", "medium"})

# draw_gauge threshold list: (upper_bound, colour)
# 0–2 green | 2–5 amber | 5–100 red
# WR-04 fix — gauge amber band aligned with rag_utils.STATUS_COLOR['yellow']
# (#f57c00) so the gauge, status badge, and RAG strip use the same orange.
_GAUGE_THRESHOLDS = [
    (_GREEN_THRESHOLD,  "#388e3c"),
    (_YELLOW_THRESHOLD, "#f57c00"),
    (100.0,             "#d32f2f"),
]

_STATUS_COLOR: dict[str, str] = {
    "green":   "#388e3c",
    "yellow":  "#f57c00",
    "red":     "#d32f2f",
    "no_data": "#757575",
}

_STATUS_LABEL: dict[str, str] = {
    "green":   "On Target",
    "yellow":  "At Risk",
    "red":     "Off Target",
    "no_data": "No Data",
}

_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_YELLOW = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")


# ===========================================================================
# Module class
# ===========================================================================

@register_module
class AgedVulnsAssetsModule(BaseModule):
    """
    Percentage of on-time-scanned assets with >= 1 Med/High/Crit vuln open > 90 days.

    Lower is better.  Per-owner breakdown sorted worst-first (highest risk score
    at the top) to surface owners with the deepest aging backlog.

    Supported options
    -----------------
    None — this module accepts no configurable options.
    """

    MODULE_ID         = "aged_vulns_assets"
    DISPLAY_NAME      = "Aged Vulnerability Assets"
    DESCRIPTION       = (
        f"Percentage of on-time-scanned assets with >={_MIN_AGED_COUNT} "
        f"Medium/High/Critical vulnerability open >{_AGED_DAYS_THRESHOLD} days."
    )
    REQUIRED_DATA     = ["vulns", "assets"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute()
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute aged-vulnerability asset percentage and per-BU breakdown.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Open / reopened findings from fetch_all_vulnerabilities().
            Expected columns: asset_uuid, severity (VPR-derived), first_found.
        assets_df : pd.DataFrame
            Full asset inventory from fetch_all_assets().
            Required columns: asset_uuid, hostname, last_seen,
            last_licensed_scan_date, tags.
        report_date : datetime
            UTC-aware report run timestamp.
        config : ModuleConfig
            Module configuration (no options consumed).
        **kwargs
            Accepted but not used.

        Returns
        -------
        ModuleData
            ``error`` is None on success; set on failure.
        """
        logger.debug(
            "%s compute() — vulns_df=%d rows, assets_df=%d rows",
            self._log_prefix(), len(vulns_df), len(assets_df),
        )

        try:
            # ---- Step 1: derive on-time asset set ----
            on_time, _ = identify_on_time_assets(assets_df, report_date)
            # [Rule 1] Defensive — when assets_df is empty (or has no
            # asset_uuid column at all), identify_on_time_assets returns an
            # empty frame whose columns may not include asset_uuid; using
            # `on_time["asset_uuid"]` would KeyError and skip the no_data
            # early-return below. Probe via `in on_time.columns` instead so
            # the no_data early-return path always wins on empty input.
            if "asset_uuid" in on_time.columns:
                on_time_uuids = set(on_time["asset_uuid"].dropna())
            else:
                on_time_uuids = set()
            total_on_time = len(on_time)

            if total_on_time == 0:
                logger.warning(
                    "%s no on-time assets found — returning no_data.",
                    self._log_prefix(),
                )
                return ModuleData(
                    module_id    = self.MODULE_ID,
                    display_name = self.DISPLAY_NAME,
                    metrics      = {
                        "aged_assets_pct":   None,
                        "aged_assets_count": 0,
                        "total_on_time":     0,
                        "status":            "no_data",
                    },
                    table_data   = [],
                    chart_data   = {"value": None, "top_5": []},
                    summary_text = (
                        "No on-time-scanned assets were found — "
                        "aged vulnerability asset percentage cannot be computed."
                    ),
                    metadata     = {**_build_metadata(report_date), "email_gauge_b64": ""},
                    # ── Phase 3 contract fields (D-07/D-15) ──
                    driver_narrative = NO_DATA_DRIVER,
                    analyst_rows     = [],
                    rag_strip        = build_rag_strip_entry(
                        display_name       = self.DISPLAY_NAME,
                        headline_value_str = NO_DATA_HEADLINE,
                        status             = "no_data",
                    ),
                    error        = None,
                )

            # ---- Step 2: build UTC-aware report timestamp ----
            if hasattr(report_date, "tzinfo") and report_date.tzinfo is not None:
                rd_ts = pd.Timestamp(report_date).tz_convert("UTC")
            else:
                rd_ts = pd.Timestamp(report_date, tz="UTC")

            # ---- Step 3: identify assets with aged Med/High/Crit findings ----
            aged_uuids, aged_findings = _find_aged_assets(vulns_df, on_time_uuids, rd_ts)
            aged_assets_count = len(aged_uuids)

            # ---- Step 4: overall metric ----
            aged_assets_pct = round(aged_assets_count / total_on_time * 100, 1)
            status = sla_status_from_thresholds(
                aged_assets_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,
            )

            # ---- Step 5: per-owner breakdown ----
            enriched       = extract_owner(on_time)
            numerator_mask = enriched["asset_uuid"].isin(aged_uuids)
            denom_mask     = pd.Series(True, index=enriched.index)

            bu_breakdown = compute_per_bu_breakdown(
                enriched, numerator_mask, denom_mask,
                higher_is_better=False,
            )

            # ---- Step 5a: compute owner risk scores and re-sort ----
            bu_risk = compute_bu_risk_scores(
                vulns_df         = vulns_df,
                qualifying_uuids = aged_uuids,
                enriched         = enriched,
                severities       = frozenset({"critical", "high", "medium"}),
                weights          = RISK_WEIGHTS,
            )
            bu_breakdown = bu_breakdown.merge(
                bu_risk.rename("risk_score").reset_index(),
                on="owner",
                how="left",
            )
            # F-DTYPE (Plan 03-07 Task 3): use .assign() rather than
            # df[col]= or .loc[:, col]= setters — see board_report_utils
            # F-DTYPE comment for rationale. .assign() replaces the
            # column on a fresh frame; preserves int64 dtype AND avoids
            # the pandas 3.0 ChainedAssignmentError FutureWarning.
            bu_breakdown = bu_breakdown.assign(
                risk_score=bu_breakdown["risk_score"].fillna(0).astype(int),
            )
            bu_breakdown = bu_breakdown.sort_values(
                "risk_score", ascending=False,
            ).reset_index(drop=True)
            table_data = bu_breakdown.to_dict("records")

            # ---- Step 6: narrative summary ----
            summary_text = _build_summary(
                aged_assets_pct, aged_assets_count, total_on_time, status,
            )

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            # ============================================================
            # Phase 3 — populate the three new ModuleData fields
            # (D-05/D-06/D-07/D-10/D-11/D-12/D-13/D-14/T-03-05-01..03)
            # ============================================================

            # ---- Step 7a: analyst rows DataFrame (D-10/D-11/D-12/D-13) ----
            # Source: aged_findings (the filtered frame where
            # days_open > _AGED_DAYS_THRESHOLD).
            # D-12 — single tab with worst_severity column (no per-severity sub-tabs).
            # D-13 — apply asset-level dedup via deduplicate_assets_by_name.
            if aged_uuids and not aged_findings.empty:
                aged_subset = aged_findings[
                    aged_findings["asset_uuid"].isin(aged_uuids)
                ].copy()

                # Per-asset aggregation
                def _join_plugins(s):  # noqa: PLC0415
                    # contributing_plugins: alphabetical-sorted unique plugin
                    # names per D-12 caveat. Deterministic via sorted(set(...)).
                    return ", ".join(
                        sorted({str(v) for v in s.dropna().tolist() if str(v).strip()})
                    )

                grouped = (
                    aged_subset
                    .groupby("asset_uuid", as_index=False)
                    .agg(
                        oldest_finding_age_days = ("days_open", "max"),
                        count_of_aged_findings  = ("plugin_id", "count"),
                        contributing_plugins    = ("plugin_name", _join_plugins),
                        worst_severity          = ("severity", lambda s: _worst_severity(set(s))),
                    )
                )

                # W6 — JOIN real (hostname, owner, last_seen) from
                # assets_df. `deduplicate_assets_by_name` REQUIRES the
                # `last_seen` column AND uses it to break duplicate-hostname
                # ties (board_report_utils.py:94, 102-107). We project the
                # REAL last_seen from assets_df rather than injecting a
                # pd.NaT placeholder — placeholders make dedup
                # nondeterministic when multiple rows share a hostname.
                asset_cols = assets_df.copy()
                if "owner" not in asset_cols.columns:
                    asset_cols = extract_owner(asset_cols)
                if "last_seen" not in asset_cols.columns:
                    # Defensive — fetch_all_assets() guarantees this column,
                    # but log if the upstream contract is ever broken.
                    asset_cols = asset_cols.assign(last_seen=pd.NaT)
                    logger.warning(
                        "%s assets_df missing 'last_seen' column — falling back to NaT. "
                        "deduplicate_assets_by_name dedup-tie behavior may be nondeterministic.",
                        self._log_prefix(),
                    )
                asset_cols = (
                    asset_cols[["asset_uuid", "hostname", "owner", "last_seen"]]
                    .drop_duplicates("asset_uuid")
                )
                analyst_df = grouped.merge(asset_cols, on="asset_uuid", how="left")

                # Apply asset-level dedup with REAL last_seen — most-recent
                # row wins on duplicate hostnames (deterministic).
                analyst_df = deduplicate_assets_by_name(analyst_df)

                # last_seen is no longer needed in the analyst output — drop it.
                if "last_seen" in analyst_df.columns:
                    analyst_df = analyst_df.drop(columns=["last_seen"])

                analyst_df = analyst_df.reindex(columns=[
                    "hostname",
                    "owner",
                    "oldest_finding_age_days",
                    "count_of_aged_findings",
                    "contributing_plugins",
                    "worst_severity",
                ])
                # D-11 — sort by oldest_finding_age_days desc
                analyst_df = analyst_df.sort_values(
                    "oldest_finding_age_days", ascending=False, na_position="last",
                ).reset_index(drop=True)

                # T-03-05-02 — CSV-formula injection guard (text columns)
                for _col in ("hostname", "owner", "contributing_plugins", "worst_severity"):
                    analyst_df.loc[:, _col] = analyst_df[_col].astype("string").map(
                        lambda s: ("'" + s)
                        if isinstance(s, str) and s[:1] in ("=", "+", "-", "@")
                        else s
                    )

                analyst_rows_payload: list = [("Aged Vulns Detail", analyst_df)]
            else:
                analyst_rows_payload = []

            # ---- Step 7b: driver narrative (D-06) ----
            # Template (locked in plan 03-05):
            #   "{count} assets carry at least one Med+ vuln open
            #    >{_AGED_DAYS_THRESHOLD} days; oldest finding: {oldest_age}
            #    days; worst BU: {worst_bu_name} with {worst_bu_count} assets."
            if aged_assets_count > 0 and analyst_rows_payload:
                df_for_driver = analyst_rows_payload[0][1]
                oldest_age = int(df_for_driver["oldest_finding_age_days"].max())
                bu_counts = (
                    df_for_driver
                    .groupby("owner", dropna=False, as_index=False)
                    .size()
                    .rename(columns={"size": "asset_count"})
                )
                bu_counts = bu_counts.assign(
                    owner=bu_counts["owner"].fillna("Unassigned").replace("", "Unassigned")
                )
                bu_counts = bu_counts.sort_values(
                    ["asset_count", "owner"], ascending=[False, True],
                )
                worst_bu_name  = str(bu_counts.iloc[0]["owner"])
                worst_bu_count = int(bu_counts.iloc[0]["asset_count"])
                driver = (
                    f"{safe_int(aged_assets_count)} assets carry at least one Med+ vuln open "
                    f">{_AGED_DAYS_THRESHOLD} days; oldest finding: {safe_int(oldest_age)} days; "
                    f"worst Owner: {worst_bu_name} with {safe_int(worst_bu_count)} assets."
                )
            else:
                driver = NO_DATA_DRIVER

            # ---- Step 7c: email gauge base64 (D-04) ----
            if aged_assets_pct is not None:
                try:
                    email_gauge_b64 = draw_gauge(
                        value      = aged_assets_pct,
                        thresholds = _GAUGE_THRESHOLDS,
                        title      = self.DISPLAY_NAME,
                        unit       = "%",
                        figsize    = (2.4, 1.6),
                    )
                except Exception as exc:    # noqa: BLE001
                    logger.warning(
                        "%s email-gauge draw_gauge failed: %s",
                        self._log_prefix(), exc,
                    )
                    email_gauge_b64 = ""
            else:
                email_gauge_b64 = ""

            # ---- Step 7d: rag_strip dict (lower_is_better — T-03-05-03) ----
            _status_for_strip = rag_status_from_value(
                aged_assets_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,    # "lower_is_better" — T-03-05-03
            )
            rag_strip_payload = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = safe_pct(aged_assets_pct),
                status             = _status_for_strip,
            )

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "aged_assets_pct":   aged_assets_pct,
                    "aged_assets_count": aged_assets_count,
                    "total_on_time":     total_on_time,
                    "status":            status,
                },
                table_data   = table_data,
                chart_data   = {
                    "value":      aged_assets_pct,
                    "thresholds": {
                        "green":  _GREEN_THRESHOLD,
                        "yellow": _YELLOW_THRESHOLD,
                    },
                    "direction":  _DIRECTION,
                    "top_5":      bu_breakdown.head(5).to_dict("records"),
                },
                summary_text = summary_text,
                metadata     = {
                    **_build_metadata(report_date),
                    "computed_at":     computed_at,
                    "email_gauge_b64": email_gauge_b64,
                },
                # ── Phase 3 contract fields (D-05/D-06/D-10..D-14) ──
                driver_narrative = driver,
                analyst_rows     = analyst_rows_payload,
                rag_strip        = rag_strip_payload,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section()
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a full-page PDF section for WeasyPrint.

        Layout (two-column row via ``two_column_metric_section`` to avoid the
        page bleed a stacked explanation caused):
        - Section heading (full width)
        - Left column: gauge (green near 0%, red near 100%), status badge
          showing aged-assets % and on-target label, and two bold support
          numbers (Aged Assets | Total On-Time Assets)
        - Right column: explanatory paragraph (left-aligned)
        - Below, full width: Top-5 worst-performing BUs table (highest % first)

        Returns an error callout div if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m                 = data.metrics
        aged_assets_pct   = m.get("aged_assets_pct")
        aged_assets_count = m.get("aged_assets_count", 0)
        total_on_time     = m.get("total_on_time", 0)
        status            = m.get("status", "no_data")

        # ---- Gauge ----
        gauge_value = aged_assets_pct if aged_assets_pct is not None else 0.0
        try:
            gauge_b64  = draw_gauge(
                value      = gauge_value,
                min_val    = 0,
                max_val    = 100,
                thresholds = _GAUGE_THRESHOLDS,
                title      = "Aged Vuln Assets %",
                unit       = "%",
                figsize    = (5, 3),
            )
            gauge_html = (
                f'<div style="text-align:center; margin-bottom:4mm;">'
                f'<img src="data:image/png;base64,{gauge_b64}" '
                f'style="width:46%; max-width:320px;" '
                f'alt="Aged Vulnerability Assets gauge">'
                f'</div>'
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("%s PDF gauge render failed: %s", self._log_prefix(), exc)
            pct_display = f"{gauge_value:.1f}%"  # safe: gauge_value guaranteed non-None by line 315 (None coalesced to 0.0)
            gauge_html = (
                f'<p style="text-align:center; font-size:20pt; font-weight:bold; '
                f'color:{_STATUS_COLOR.get(status, "#333")};">{pct_display}</p>'
            )

        # ---- Status badge ----
        status_color = _STATUS_COLOR.get(status, "#757575")
        status_label = _STATUS_LABEL.get(status, status)
        pct_display  = (
            f"{aged_assets_pct:.1f}%" if aged_assets_pct is not None else "N/A"
        )
        status_html = (
            f'<p style="text-align:center; font-size:10pt; font-weight:bold; '
            f'color:{status_color}; margin:0 0 5mm 0;">'
            f'Aged Assets: {pct_display} &nbsp;&middot;&nbsp; {status_label}'
            f'</p>'
        )

        # ---- Supporting numbers ----
        aged_color   = _STATUS_COLOR.get(status, "#555")
        support_html = f"""
<table style="width:52%; margin:0 auto 6mm auto; border-collapse:collapse;">
  <tr>
    <td style="text-align:center; padding:2mm 8mm;
               border-right:0.5pt solid #ddd; vertical-align:middle;">
      <span style="font-size:14pt; font-weight:bold;
             color:{aged_color};">{aged_assets_count:,}</span>
      <br><span style="font-size:7.5pt; color:#555;">Assets with Aged Vulns</span>
    </td>
    <td style="text-align:center; padding:2mm 8mm; vertical-align:middle;">
      <span style="font-size:14pt; font-weight:bold;
             color:#1F3864;">{total_on_time:,}</span>
      <br><span style="font-size:7.5pt; color:#555;">Total On-Time Assets</span>
    </td>
  </tr>
</table>"""

        # ---- Top 5 BU table (worst performers = highest %) ----
        top5 = data.chart_data.get("top_5", [])
        if top5:
            rows_html = ""
            for row in top5:
                bu_name  = str(row.get("owner", ""))
                bu_num   = int(row.get("numerator",    0))
                bu_den   = int(row.get("denominator",  0))
                bu_score = int(row.get("risk_score",   0))
                rows_html += (
                    f'<tr>'
                    f'<td style="padding:1.5mm 3mm;">{bu_name}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_num:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_den:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_score:,}</td>'
                    f'</tr>'
                )
            bu_table_html = f"""
<h3 class="subsection-heading">Top 5 Worst-Performing Owners</h3>
<table class="data-table">
  <thead>
    <tr>
      <th>Owner</th>
      <th style="text-align:right;">Assets with Aged Vulns</th>
      <th style="text-align:right;">On-Time Assets</th>
      <th style="text-align:right;">Risk Score</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>"""
        else:
            bu_table_html = (
                '<p class="explanatory-text" style="color:#888; font-style:italic;">'
                'No owner breakdown available — '
                'assets may lack Owner tags or no aged findings were found.'
                '</p>'
            )

        # ---- Explanatory paragraph ----
        # Pre-format module-level threshold constants (never None) so the
        # multi-line HTML f-string below contains no inline format specs.
        green_str  = format(_GREEN_THRESHOLD,  ".0f")
        yellow_str = format(_YELLOW_THRESHOLD, ".0f")
        explain_html = f"""
<p class="explanatory-text">
  <strong>What this measures:</strong> The percentage of assets scanned on time
  (licensed scan within the last {ON_TIME_WINDOW_DAYS} days) that carry at least one
  Medium, High, or Critical vulnerability (VPR&nbsp;&ge;4.0) open for more than
  {_AGED_DAYS_THRESHOLD} days.  Long-standing unresolved findings — even at medium
  severity — represent persistent, accepted risk exposure and can indicate systemic
  remediation gaps.  Board target is &le;{green_str}% (green).
  &le;{yellow_str}% is at-risk (amber).  Above {yellow_str}%
  is off-target (red).  Owner breakdown uses the Tenable
  &ldquo;Owner&rdquo; tag category.
</p>"""

        return two_column_metric_section(
            heading_html=f'<h2 class="section-heading">{self.DISPLAY_NAME}</h2>',
            left_html=f"{gauge_html}{status_html}{support_html}",
            explanation_html=explain_html,
            full_width_html=bu_table_html,
        )

    # ------------------------------------------------------------------
    # render_excel_tabs()
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write the "Aged Vuln Assets" tab into ``workbook``.

        Tab contents:
        - Overall KPI block (aged %, status, supporting numbers, thresholds)
        - Full per-BU breakdown table sorted worst-first (highest % at top)

        Returns
        -------
        list[str]
            ``["Aged Vuln Assets"]`` on success, ``[]`` on error.
        """
        tab_name = "Aged Vuln Assets"
        try:
            # D-16 — Phase 3 zero-row standardisation. If both metrics and
            # table_data are empty AND there is no error, emit a single
            # placeholder cell at A1 instead of a fully-empty sheet.
            # This is a behavior change from pre-Phase-3 — surface in SUMMARY.
            empty_metrics = not (
                data.metrics and any(v is not None for v in data.metrics.values())
            )
            empty_tables = not data.table_data
            if empty_metrics and empty_tables and not data.error:
                ws = workbook.create_sheet(tab_name)
                ws["A1"]      = "No data in scope"
                ws["A1"].font = Font(bold=True, color="666666")
                return [tab_name]

            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            m                 = data.metrics
            aged_assets_pct   = m.get("aged_assets_pct")
            aged_assets_count = m.get("aged_assets_count", 0)
            total_on_time     = m.get("total_on_time", 0)
            status            = m.get("status", "no_data")

            # ---- Overall KPI block ----
            _xl_title(ws, "A1", "Aged Vulnerability Assets — Overall Summary")

            status_color_hex = _STATUS_COLOR.get(status, "#757575").lstrip("#")
            pct_str = (
                f"{aged_assets_pct:.1f}%" if aged_assets_pct is not None else "N/A"
            )

            _xl_kv(ws, 3, "Aged Assets %:", pct_str,
                   value_font=Font(bold=True, size=12, color=status_color_hex))
            _xl_kv(ws, 4, "Status:",
                   _STATUS_LABEL.get(status, status),
                   value_font=Font(bold=True, color=status_color_hex))
            _xl_kv(ws, 5, "Assets with Aged Vulns:", f"{aged_assets_count:,}")
            _xl_kv(ws, 6, "Total On-Time Assets:", f"{total_on_time:,}")
            _xl_kv(ws, 7, "Aged Definition:",
                   f">={_MIN_AGED_COUNT} Medium/High/Critical vuln open >{_AGED_DAYS_THRESHOLD} days")
            _xl_kv(ws, 8, "Scope:",
                   "On-time-scanned assets only (last_licensed_scan_date within last 30 days)")
            _xl_kv(ws, 9, "SLA Thresholds (lower is better):",
                   f"Green <={_GREEN_THRESHOLD:.0f}%  |  "  # safe: module-level int constant, never None
                   f"Amber <={_YELLOW_THRESHOLD:.0f}%  |  "  # safe: module-level int constant, never None
                   f"Red >{_YELLOW_THRESHOLD:.0f}%")  # safe: module-level int constant, never None

            # ---- Owner breakdown table (starts at row 11, worst first) ----
            header_row = 11
            headers = [
                "Owner", "Assets with Aged Vulns", "On-Time Assets", "Risk Score"
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell           = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = _FILL_HEADER
                cell.alignment = Alignment(horizontal="center")

            for row_offset, row in enumerate(data.table_data or [], start=1):
                data_row = header_row + row_offset

                ws.cell(row=data_row, column=1,
                        value=str(row.get("owner", ""))).alignment = (
                    Alignment(horizontal="left")
                )
                ws.cell(row=data_row, column=2,
                        value=int(row.get("numerator",   0))).alignment = (
                    Alignment(horizontal="right")
                )
                ws.cell(row=data_row, column=3,
                        value=int(row.get("denominator", 0))).alignment = (
                    Alignment(horizontal="right")
                )
                score_cell           = ws.cell(row=data_row, column=4,
                                               value=int(row.get("risk_score", 0)))
                score_cell.alignment = Alignment(horizontal="right")

            # ---- Column widths ----
            ws.column_dimensions[get_column_letter(1)].width = 32
            ws.column_dimensions[get_column_letter(2)].width = 22
            ws.column_dimensions[get_column_letter(3)].width = 18
            ws.column_dimensions[get_column_letter(4)].width = 12

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_kpis()
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return three KPI tiles for the HTML email body.

        Returns
        -------
        dict[str, str]
            Keys: "Aged Assets %", "Assets w/ Aged Vulns", "On-Time Assets".
            Returns empty dict if ``data.error`` is set.
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m   = data.metrics
        pct = m.get("aged_assets_pct")
        return {
            "Aged Assets %":       f"{pct:.1f}%" if pct is not None else "N/A",
            "Assets w/ Aged Vulns": f"{m.get('aged_assets_count', 0):,}",
            "On-Time Assets":       f"{m.get('total_on_time', 0):,}",
        }

    # ------------------------------------------------------------------
    # render_email_panel() — Phase 3 D-02 horizontal-split panel
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Build the per-module email panel — horizontal split layout (D-02).

        Layout: 620px-wide table, 150px gauge cell on the left + 430px text
        cell on the right. Inline CSS only. Outlook-safe ``<table>`` shell
        with explicit ``width=""`` attributes per project email conventions.

        Empty-data behavior (D-15): when ``data.error`` or no
        ``email_gauge_b64`` is available, returns the same 620px shell with
        a gray "No data" placeholder block where the gauge would be.

        Returns
        -------
        str
            Inline-CSS HTML fragment. Returns ``""`` only on a render
            exception (caught by the composer's per-module exception
            isolation pattern).
        """
        try:
            # Empty-data placeholder per D-15
            b64 = (data.metadata or {}).get("email_gauge_b64", "")
            if data.error or not isinstance(b64, str) or not b64.strip():
                return self._render_empty_email_panel()

            pct       = data.metrics.get("aged_assets_pct") if data.metrics else None
            headline  = safe_pct(pct)
            status    = (data.metrics or {}).get("status", "no_data")
            rag_color = _RAG_STATUS_COLOR.get(status, _RAG_STATUS_COLOR["no_data"])
            rag_label = _RAG_STATUS_LABEL.get(status, _RAG_STATUS_LABEL["no_data"])
            icon      = STATUS_ICON.get(status,        STATUS_ICON["no_data"])

            cid       = f"{self.MODULE_ID}_gauge"
            # T-03-05-01 — html-escape every module-supplied string
            label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
            driver_esc = html.escape(str(data.driver_narrative or ""), quote=True)

            return (
                '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
                'style="width:620px; max-width:620px; margin:8px 0; '
                'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
                '<tr>'
                '  <td width="150" style="padding:12px; vertical-align:middle; '
                '      text-align:center;">'
                f'    <img src="cid:{cid}" alt="" width="120" height="120" '
                '         style="display:block; margin:0 auto;" />'
                '  </td>'
                '  <td width="430" style="padding:12px; vertical-align:middle;">'
                f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
                f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{headline}</div>'
                f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
                f'{icon} {html.escape(rag_label)}</div>'
                f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
                f'{driver_esc}</div>'
                '  </td>'
                '</tr>'
                '</table>'
            )
        except Exception as exc:    # noqa: BLE001
            logger.error(
                "%s render_email_panel raised: %s",
                self._log_prefix(), exc,
            )
            return ""

    def _render_empty_email_panel(self) -> str:
        """Return the D-15 gray 'No Data' placeholder panel."""
        label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
        driver_esc = html.escape(NO_DATA_DRIVER, quote=True)
        rag_color  = _RAG_STATUS_COLOR["no_data"]
        rag_label  = _RAG_STATUS_LABEL["no_data"]
        icon       = STATUS_ICON["no_data"]
        return (
            '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
            'style="width:620px; max-width:620px; margin:8px 0; '
            'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
            '<tr>'
            '  <td width="150" style="padding:12px; vertical-align:middle; '
            '      text-align:center; background:#f5f5f5; color:#999;">'
            '    <div style="font-size:10pt;">No data</div>'
            '  </td>'
            '  <td width="430" style="padding:12px; vertical-align:middle;">'
            f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
            f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{NO_DATA_HEADLINE}</div>'
            f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
            f'{icon} {html.escape(rag_label)}</div>'
            f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
            f'{driver_esc}</div>'
            '  </td>'
            '</tr>'
            '</table>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs() — Phase 3 D-14 single-tab list
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, "pd.DataFrame"]]:
        """
        Return the module's pre-built analyst rows (D-14 single-tab list).

        Source: ``data.analyst_rows`` populated inside ``compute()``.
        Empty-data: returns ``[]`` (no tab written for this module).
        """
        try:
            if data.error or not data.analyst_rows:
                return []
            return list(data.analyst_rows)
        except Exception as exc:    # noqa: BLE001
            logger.error("[%s] render_analyst_tabs raised: %s", self.MODULE_ID, exc)
            return []

    # ------------------------------------------------------------------
    # get_audit_info()
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "aged_assets_pct": (
                    "aged_assets_count / total_on_time × 100, rounded to 1 decimal. "
                    "None → 'no_data' when total_on_time == 0."
                ),
                "aged_assets_count": (
                    f"Count of on-time assets that have >= {_MIN_AGED_COUNT} "
                    f"Medium/High/Critical finding with days_open > {_AGED_DAYS_THRESHOLD}."
                ),
                "total_on_time": (
                    "Count of deduplicated assets where last_licensed_scan_date IS NOT NULL "
                    f"AND >= report_date − {ON_TIME_WINDOW_DAYS} days."
                ),
                "days_open": (
                    "(report_date − first_found).days.  Findings with null first_found "
                    "produce NaT/NaN and are treated as not aged."
                ),
                "severity_filter": (
                    "severity IN ('critical', 'high', 'medium').  Severity is VPR-derived "
                    "as produced by fetch_all_vulnerabilities()."
                ),
                "owner_breakdown": (
                    "compute_per_bu_breakdown(higher_is_better=False) on on-time assets "
                    "enriched with Owner tag. "
                    "Numerator = assets with aged finding(s) per owner; "
                    "denominator = all on-time assets per owner. "
                    "affected = numerator (raw aged-asset count). "
                    "Primary sort: risk_score DESC (highest risk score at the top)."
                ),
                # WR-06 — explicit audit-info entry documenting the
                # intentional broadening of compute_bu_risk_scores(): the
                # weighted Risk Score column counts ALL open findings on
                # an aged asset (Critical/High/Medium), not only the
                # findings whose age > 90 days that triggered the asset's
                # "aged" qualification. This makes the Risk Score a
                # holistic asset-risk indicator rather than a strict
                # aged-only count. See compute_bu_risk_scores() docstring
                # in board_report_utils.py and the runbook section
                # 'Risk Score broadening' in docs/board_summary_calculations.md.
                "risk_score": (
                    "Sum of weighted (severity_weight × open_finding_count) "
                    "across ALL open Critical/High/Medium findings on each "
                    "aged-qualifying asset, then summed per BU. Intentionally "
                    "includes non-aged findings on aged assets so the score "
                    "reflects the asset's holistic risk posture, not just the "
                    "findings that caused it to qualify."
                ),
            },
        }


# ===========================================================================
# Module-private helpers
# ===========================================================================

def _worst_severity(severities: set[str]) -> str:
    """
    Return the worst severity from a set, ordered: critical > high > medium.

    Severities not in {"critical", "high", "medium"} are ignored. An empty
    or all-unknown set yields ``""``.

    Parameters
    ----------
    severities : set[str]
        A set of severity strings (case-insensitive). Non-string values and
        severities outside the Med/High/Crit triad are ignored.

    Returns
    -------
    str
        ``"critical"``, ``"high"``, ``"medium"``, or ``""``.
    """
    sevs = {str(s).strip().lower() for s in severities if isinstance(s, str)}
    for tier in ("critical", "high", "medium"):
        if tier in sevs:
            return tier
    return ""


def _find_aged_assets(
    vulns_df:      pd.DataFrame,
    on_time_uuids: set,
    rd_ts:         pd.Timestamp,
) -> tuple[set, pd.DataFrame]:
    """
    Return the set of on-time asset UUIDs with >= 1 aged Med/High/Crit finding,
    along with the filtered aged-findings DataFrame.

    A finding is aged when (report_date − first_found).days > _AGED_DAYS_THRESHOLD.
    Findings with null first_found are treated as 0 days old (not aged).

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Open / reopened findings.
    on_time_uuids : set
        UUIDs of on-time-scanned assets.
    rd_ts : pd.Timestamp
        UTC-aware report timestamp.

    Returns
    -------
    tuple[set, pd.DataFrame]
        ``(aged_uuids, aged_findings)``
        - ``aged_uuids``: UUIDs of on-time assets with at least one qualifying
          aged finding.
        - ``aged_findings``: DataFrame slice of ``vulns_df`` (the
          ``relevant[aged_mask]`` subset) with a precomputed ``days_open``
          column. Phase 3 — used by ``compute()`` to build the per-asset
          analyst tab without re-deriving the filter. Empty DataFrame when
          no aged findings exist.
    """
    empty_frame = vulns_df.iloc[0:0].copy() if not vulns_df.empty else pd.DataFrame()

    if vulns_df.empty:
        return set(), empty_frame

    required = {"asset_uuid", "severity", "first_found"}
    if not required.issubset(vulns_df.columns):
        missing = required - set(vulns_df.columns)
        logger.warning(
            "_find_aged_assets: missing columns %s — returning empty set.", missing
        )
        return set(), empty_frame

    # Filter to on-time assets + qualifying severities
    on_time_mask  = vulns_df["asset_uuid"].isin(on_time_uuids)
    severity_mask = vulns_df["severity"].str.lower().isin(_AGED_SEVERITIES)
    relevant      = vulns_df[on_time_mask & severity_mask].copy()

    if relevant.empty:
        return set(), empty_frame

    # Compute days_open; NaT → NaN → aged_mask = False (not aged)
    days_open = (rd_ts - relevant["first_found"]).dt.days
    aged_mask = days_open > _AGED_DAYS_THRESHOLD

    aged = relevant[aged_mask].copy()

    if aged.empty:
        return set(), empty_frame

    # Attach days_open onto the aged frame so compute() can use it directly
    # for the oldest_finding_age_days aggregation in the analyst tab.
    aged.loc[:, "days_open"] = days_open[aged_mask].astype("Int64")

    aged_uuids = set(aged["asset_uuid"].dropna().unique())
    return aged_uuids, aged


def _build_metadata(report_date: Any) -> dict:
    """Return the standard metadata block for this module."""
    return {
        "aged_definition":   (
            f">={_MIN_AGED_COUNT} Medium/High/Critical finding open "
            f">{_AGED_DAYS_THRESHOLD} days on an on-time-scanned asset."
        ),
        "severity_scope":    "Medium (VPR 4.0–6.9), High (VPR 7.0–8.9), Critical (VPR 9.0–10.0)",
        "denominator_scope": (
            f"On-time-scanned assets: last_licensed_scan_date IS NOT NULL "
            f"AND >= report_date − {ON_TIME_WINDOW_DAYS} days."
        ),
        "sla_source":        (
            f"Board-defined thresholds "
            f"(Green <={_GREEN_THRESHOLD}%, "
            f"Amber <={_YELLOW_THRESHOLD}%, Red >{_YELLOW_THRESHOLD}%, "
            f"direction=lower_is_better)"
        ),
        "window":            f"Last {ON_TIME_WINDOW_DAYS} days from report_date",
    }


def _build_summary(
    aged_assets_pct:   float | None,
    aged_assets_count: int,
    total_on_time:     int,
    status:            str,
) -> str:
    """Build a plain-language narrative sentence for the email body."""
    if aged_assets_pct is None:
        return (
            "No on-time-scanned assets were found — "
            "aged vulnerability asset percentage cannot be computed."
        )
    status_label = _STATUS_LABEL.get(status, status)
    # WR-07 fix — use safe_pct() instead of an inline f-string format
    # spec on a possibly-None value. The early-return guard above
    # currently makes aged_assets_pct non-None here, but safe_pct()
    # makes the rule mechanical so a future refactor that breaks the
    # guard cannot crash this line.
    return (
        f"{safe_pct(aged_assets_pct)} of on-time-scanned assets have aged vulnerabilities — "
        f"{aged_assets_count:,} of {total_on_time:,} assets carry at least one "
        f"Medium/High/Critical finding open >{_AGED_DAYS_THRESHOLD} days. "
        f"Status: {status_label}."
    )


def _row_bg(pct: float) -> str:
    """Light HTML background-color for a BU table row (lower-is-better)."""
    if pct <= _GREEN_THRESHOLD:
        return "#E8F5E9"
    if pct <= _YELLOW_THRESHOLD:
        return "#FFF8E1"
    return "#FFEBEE"


def _xl_fill(pct: float) -> PatternFill:
    """openpyxl PatternFill for an aged-% cell (lower-is-better)."""
    if pct <= _GREEN_THRESHOLD:
        return _FILL_GREEN
    if pct <= _YELLOW_THRESHOLD:
        return _FILL_YELLOW
    return _FILL_RED


def _xl_title(ws, cell_ref: str, value: str) -> None:
    ws[cell_ref]      = value
    ws[cell_ref].font = Font(bold=True, size=12)


def _xl_kv(ws, row: int, label: str, value: str,
            value_font: Font | None = None) -> None:
    lc      = ws.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True)
    vc      = ws.cell(row=row, column=2, value=value)
    if value_font is not None:
        vc.font = value_font
