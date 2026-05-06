"""
reports/modules/high_risk_assets_module.py — High-Risk Assets metric module.

Measures the percentage of on-time-scanned assets that carry 10 or more
Critical or High vulnerabilities that have been open for more than 30 days.
A high concentration of such assets signals a systemic remediation backlog
that exposes the organisation to sustained elevated risk.

Module ID:    high_risk_assets
Display Name: High-Risk Assets

SLA thresholds (board-defined, lower is better):
    Green:  high_risk_pct <= 0.5%
    Yellow: high_risk_pct <= 1.0%  (and > 0.5%)
    Red:    high_risk_pct >  1.0%

Denominator: all deduplicated on-time-scanned assets (last_licensed_scan_date
             within the last 30 days).
Numerator:   subset where the count of Critical/High open findings with
             days_open > 30 is >= 10.
"""

from __future__ import annotations

import html
import logging
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from config import RISK_WEIGHTS
from reports.modules.board_report_utils import (
    compute_bu_risk_scores,
    compute_per_bu_breakdown,
    deduplicate_assets_by_name,
    extract_business_unit,
    identify_on_time_assets,
    sla_status_from_thresholds,
    ON_TIME_WINDOW_DAYS,
)
from reports.modules.chart_utils import draw_gauge
from reports.modules.format_utils import safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR as _RAG_STATUS_COLOR,
    STATUS_ICON,
    STATUS_LABEL as _RAG_STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

_GREEN_THRESHOLD  = 0.5    # <= green (lower is better)
_YELLOW_THRESHOLD = 1.0    # <= yellow; > green
_DIRECTION        = "lower_is_better"

#: Minimum number of aged Critical/High findings required for an asset to be
#: classified as "high-risk".
_HIGH_RISK_COUNT: int = 10

#: Findings open longer than this many days qualify as "aged".
_AGED_DAYS_THRESHOLD: int = 30

# draw_gauge threshold list: each tuple = (upper_bound, colour)
# For lower_is_better the gauge reads left=green, right=red:
#   0–0.5  green | 0.5–1.0 amber | 1.0–100 red
_GAUGE_THRESHOLDS = [
    (_GREEN_THRESHOLD,  "#388e3c"),   # 0 – 0.5  green
    (_YELLOW_THRESHOLD, "#fbc02d"),   # 0.5 – 1.0 amber
    (100.0,             "#d32f2f"),   # 1.0 – 100 red
]

# Status display properties
_STATUS_COLOR: dict[str, str] = {
    "green":   "#388e3c",
    "yellow":  "#f57c00",
    "red":     "#d32f2f",
    "no_data": "#757575",
}

_STATUS_LABEL: dict[str, str] = {
    "green":   "On Target",
    "yellow":  "At Risk",
    "red":     "Off Target",
    "no_data": "No Data",
}

# Excel fill colours (no leading #, openpyxl RGB format)
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")  # light green
_FILL_YELLOW = PatternFill("solid", fgColor="FFF9C4")  # light amber
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")  # light red
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")  # dark navy


# ===========================================================================
# Module class
# ===========================================================================

@register_module
class HighRiskAssetsModule(BaseModule):
    """
    Percentage of on-time-scanned assets with >= 10 Critical/High vulns open > 30 days.

    Lower is better.  The per-BU breakdown table shows worst performers first
    (highest percentage of high-risk assets at the top) so the PDF and Excel
    surfaces the business units with the most acute remediation backlog.

    Supported options
    -----------------
    None — this module accepts no configurable options.
    """

    MODULE_ID         = "high_risk_assets"
    DISPLAY_NAME      = "High-Risk Assets"
    DESCRIPTION       = (
        f"Percentage of on-time-scanned assets carrying "
        f">={_HIGH_RISK_COUNT} Critical/High vulnerabilities open >{_AGED_DAYS_THRESHOLD} days."
    )
    REQUIRED_DATA     = ["vulns", "assets"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute()
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute the high-risk asset percentage and per-BU breakdown.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Open / reopened findings from fetch_all_vulnerabilities().
            Expected columns: asset_uuid, severity (VPR-derived), first_found.
        assets_df : pd.DataFrame
            Full asset inventory from fetch_all_assets().
            Required columns: asset_uuid, hostname, last_seen,
            last_licensed_scan_date, tags.
        report_date : datetime
            UTC-aware report run timestamp.
        config : ModuleConfig
            Module configuration (no options consumed).
        **kwargs
            Accepted but not used.

        Returns
        -------
        ModuleData
            ``error`` is None on success; set on failure.
        """
        logger.debug(
            "%s compute() — vulns_df=%d rows, assets_df=%d rows",
            self._log_prefix(), len(vulns_df), len(assets_df),
        )

        try:
            # ---- Step 1: derive on-time asset set ----
            on_time, _ = identify_on_time_assets(assets_df, report_date)
            # [Rule 1] Defensive — when assets_df is empty (or has no
            # asset_uuid column at all), identify_on_time_assets returns an
            # empty frame whose columns may not include asset_uuid; using
            # `on_time["asset_uuid"]` would KeyError and skip the no_data
            # early-return below. Probe via `.get` instead so the no_data
            # early-return path always wins on empty input.
            if "asset_uuid" in on_time.columns:
                on_time_uuids = set(on_time["asset_uuid"].dropna())
            else:
                on_time_uuids = set()
            total_on_time = len(on_time)

            if total_on_time == 0:
                logger.warning(
                    "%s no on-time assets found — returning no_data.",
                    self._log_prefix(),
                )
                return ModuleData(
                    module_id    = self.MODULE_ID,
                    display_name = self.DISPLAY_NAME,
                    metrics      = {
                        "high_risk_pct":   None,
                        "high_risk_count": 0,
                        "total_on_time":   0,
                        "status":          "no_data",
                    },
                    table_data   = [],
                    chart_data   = {"value": None, "top_5": []},
                    summary_text = (
                        "No on-time-scanned assets were found — "
                        "high-risk asset percentage cannot be computed."
                    ),
                    metadata     = {**_build_metadata(report_date), "email_gauge_b64": ""},
                    # ── Phase 3 contract fields (D-07/D-15) ──
                    driver_narrative = NO_DATA_DRIVER,
                    analyst_rows     = [],
                    rag_strip        = build_rag_strip_entry(
                        display_name       = self.DISPLAY_NAME,
                        headline_value_str = NO_DATA_HEADLINE,
                        status             = "no_data",
                    ),
                    error        = None,
                )

            # ---- Step 2: build UTC-aware report timestamp ----
            if hasattr(report_date, "tzinfo") and report_date.tzinfo is not None:
                rd_ts = pd.Timestamp(report_date).tz_convert("UTC")
            else:
                rd_ts = pd.Timestamp(report_date, tz="UTC")

            # ---- Step 3: filter vulns to on-time assets + Critical/High ----
            high_risk_uuids, aged_counts_per_asset, aged_findings = _find_high_risk_assets(
                vulns_df, on_time_uuids, rd_ts,
            )
            high_risk_count = len(high_risk_uuids)

            # ---- Step 4: overall metric ----
            high_risk_pct = round(high_risk_count / total_on_time * 100, 1)
            status = sla_status_from_thresholds(
                high_risk_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,
            )

            # ---- Step 5: per-BU breakdown ----
            enriched       = extract_business_unit(on_time)
            numerator_mask = enriched["asset_uuid"].isin(high_risk_uuids)
            denom_mask     = pd.Series(True, index=enriched.index)

            bu_breakdown = compute_per_bu_breakdown(
                enriched, numerator_mask, denom_mask,
                higher_is_better=False,
            )

            # ---- Step 5a: compute BU risk scores and re-sort ----
            bu_risk = compute_bu_risk_scores(
                vulns_df         = vulns_df,
                qualifying_uuids = high_risk_uuids,
                enriched         = enriched,
                severities       = frozenset({"critical", "high"}),
                weights          = RISK_WEIGHTS,
            )
            bu_breakdown = bu_breakdown.merge(
                bu_risk.rename("risk_score").reset_index(),
                on="business_unit",
                how="left",
            )
            bu_breakdown["risk_score"] = bu_breakdown["risk_score"].fillna(0).astype(int)
            bu_breakdown = bu_breakdown.sort_values(
                "risk_score", ascending=False,
            ).reset_index(drop=True)
            table_data = bu_breakdown.to_dict("records")

            # ---- Step 6: narrative summary ----
            summary_text = _build_summary(
                high_risk_pct, high_risk_count, total_on_time, status,
            )

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            # ============================================================
            # Phase 3 — populate the three new ModuleData fields
            # (D-05/D-06/D-07/D-10/D-11/D-13/D-14/T-03-04-01..03)
            # ============================================================

            # ---- Step 7a: analyst rows DataFrame (D-10/D-11/D-13) ----
            # Source: the high_risk_uuids subset + aged_findings frame from
            # _find_high_risk_assets. D-13 — apply asset-level dedup.
            if high_risk_uuids and not aged_findings.empty:
                # Filter to the high-risk asset subset
                aged_high_risk = aged_findings[
                    aged_findings["asset_uuid"].isin(high_risk_uuids)
                ].copy()

                # Group by asset_uuid: count of aged Crit/High + sorted unique
                # plugin IDs joined as ", ".join(str(p) for p in sorted(unique_ids)).
                # Deterministic numeric sort where possible; fall back to lexical.
                def _join_ids(s):  # noqa: PLC0415
                    # Deterministic sort: numeric ascending where possible,
                    # fall back to lexical for non-int IDs.
                    unique_vals = set(s.dropna().tolist())
                    try:
                        int_ids = [int(v) for v in unique_vals]
                        return ", ".join(str(v) for v in sorted(set(int_ids)))
                    except Exception:    # noqa: BLE001
                        return ", ".join(sorted(set(map(str, unique_vals))))

                grouped = (
                    aged_high_risk
                    .groupby("asset_uuid", as_index=False)
                    .agg(
                        crit_high_open_count    = ("plugin_id", "count"),
                        contributing_finding_ids= ("plugin_id", _join_ids),
                    )
                )

                # W6 — JOIN real (hostname, business_unit, last_seen) from
                # assets_df. `deduplicate_assets_by_name` REQUIRES the
                # `last_seen` column AND uses it to break duplicate-hostname
                # ties (board_report_utils.py:94, 102-107). We project the
                # REAL last_seen from assets_df rather than injecting a
                # pd.NaT placeholder — placeholders make dedup
                # nondeterministic when multiple rows share a hostname.
                asset_cols = assets_df.copy()
                if "business_unit" not in asset_cols.columns:
                    asset_cols = extract_business_unit(asset_cols)
                if "last_seen" not in asset_cols.columns:
                    # Defensive — fetch_all_assets() guarantees this column,
                    # but log if the upstream contract is ever broken.
                    asset_cols = asset_cols.assign(last_seen=pd.NaT)
                    logger.warning(
                        "%s assets_df missing 'last_seen' column — falling back to NaT. "
                        "deduplicate_assets_by_name dedup-tie behavior may be nondeterministic.",
                        self._log_prefix(),
                    )
                asset_cols = (
                    asset_cols[["asset_uuid", "hostname", "business_unit", "last_seen"]]
                    .drop_duplicates("asset_uuid")
                )
                analyst_df = grouped.merge(asset_cols, on="asset_uuid", how="left")

                # Apply asset-level dedup with REAL last_seen — most-recent
                # row wins on duplicate hostnames (deterministic).
                analyst_df = deduplicate_assets_by_name(analyst_df)

                # last_seen is no longer needed in the analyst output — drop it.
                if "last_seen" in analyst_df.columns:
                    analyst_df = analyst_df.drop(columns=["last_seen"])

                analyst_df = analyst_df.reindex(columns=[
                    "hostname",
                    "business_unit",
                    "crit_high_open_count",
                    "contributing_finding_ids",
                ])
                # D-11 — sort by crit_high_open_count desc
                analyst_df = analyst_df.sort_values(
                    "crit_high_open_count", ascending=False, na_position="last",
                ).reset_index(drop=True)

                # T-03-04-02 — CSV-formula injection guard (text columns)
                for _col in ("hostname", "business_unit", "contributing_finding_ids"):
                    analyst_df[_col] = analyst_df[_col].astype("string").map(
                        lambda s: ("'" + s)
                        if isinstance(s, str) and s[:1] in ("=", "+", "-", "@")
                        else s
                    )

                analyst_rows_payload: list = [("High-Risk Assets Detail", analyst_df)]
            else:
                analyst_rows_payload = []

            # ---- Step 7b: driver narrative (D-06) ----
            # Template (locked in plan 03-04):
            #   "{count} assets crossed the high-risk threshold (>={_HIGH_RISK_COUNT}
            #    Crit/High open >{_AGED_DAYS_THRESHOLD}d); worst BU: {worst_bu_name}
            #    with {worst_bu_count} assets."
            if high_risk_count > 0 and analyst_rows_payload:
                bu_counts = (
                    analyst_rows_payload[0][1]
                    .groupby("business_unit", dropna=False, as_index=False)
                    .size()
                    .rename(columns={"size": "asset_count"})
                )
                bu_counts["business_unit"] = (
                    bu_counts["business_unit"].fillna("Untagged").replace("", "Untagged")
                )
                bu_counts = bu_counts.sort_values(
                    ["asset_count", "business_unit"], ascending=[False, True],
                )
                worst_bu_name  = str(bu_counts.iloc[0]["business_unit"])
                worst_bu_count = int(bu_counts.iloc[0]["asset_count"])
                driver = (
                    f"{safe_int(high_risk_count)} assets crossed the high-risk threshold "
                    f"(>={_HIGH_RISK_COUNT} Crit/High open >{_AGED_DAYS_THRESHOLD}d); "
                    f"worst BU: {worst_bu_name} with {safe_int(worst_bu_count)} assets."
                )
            else:
                driver = NO_DATA_DRIVER

            # ---- Step 7c: email gauge base64 (D-04) ----
            if high_risk_pct is not None:
                try:
                    email_gauge_b64 = draw_gauge(
                        value      = high_risk_pct,
                        thresholds = _GAUGE_THRESHOLDS,
                        title      = self.DISPLAY_NAME,
                        unit       = "%",
                        figsize    = (2.4, 1.6),
                    )
                except Exception as exc:    # noqa: BLE001
                    logger.warning(
                        "%s email-gauge draw_gauge failed: %s",
                        self._log_prefix(), exc,
                    )
                    email_gauge_b64 = ""
            else:
                email_gauge_b64 = ""

            # ---- Step 7d: rag_strip dict (lower_is_better — T-03-04-03) ----
            _status_for_strip = rag_status_from_value(
                high_risk_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,    # "lower_is_better" — T-03-04-03
            )
            rag_strip_payload = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = safe_pct(high_risk_pct),
                status             = _status_for_strip,
            )

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "high_risk_pct":   high_risk_pct,
                    "high_risk_count": high_risk_count,
                    "total_on_time":   total_on_time,
                    "status":          status,
                },
                table_data   = table_data,
                chart_data   = {
                    "value":      high_risk_pct,
                    "thresholds": {
                        "green":  _GREEN_THRESHOLD,
                        "yellow": _YELLOW_THRESHOLD,
                    },
                    "direction":  _DIRECTION,
                    # Top 5 BUs by affected asset count (most high-risk assets first)
                    "top_5":      bu_breakdown.head(5).to_dict("records"),
                },
                summary_text = summary_text,
                metadata     = {
                    **_build_metadata(report_date),
                    "computed_at":     computed_at,
                    "email_gauge_b64": email_gauge_b64,
                },
                # ── Phase 3 contract fields (D-05/D-06/D-10..D-14) ──
                driver_narrative = driver,
                analyst_rows     = analyst_rows_payload,
                rag_strip        = rag_strip_payload,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section()
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a full-page PDF section for WeasyPrint.

        Layout (top to bottom):
        1. Section heading
        2. Gauge (centred) — green near 0%, red near 100%
        3. Status badge showing high-risk % and on-target label
        4. Two bold support numbers: High-Risk Assets | Total On-Time Assets
        5. Top-5 worst-performing BUs table (highest % first)
        6. Explanatory paragraph

        Returns an error callout div if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m               = data.metrics
        high_risk_pct   = m.get("high_risk_pct")
        high_risk_count = m.get("high_risk_count", 0)
        total_on_time   = m.get("total_on_time", 0)
        status          = m.get("status", "no_data")

        # ---- Gauge ----
        gauge_value = high_risk_pct if high_risk_pct is not None else 0.0
        try:
            gauge_b64  = draw_gauge(
                value      = gauge_value,
                min_val    = 0,
                max_val    = 100,
                thresholds = _GAUGE_THRESHOLDS,
                title      = "High-Risk Assets %",
                unit       = "%",
                figsize    = (5, 3),
            )
            gauge_html = (
                f'<div style="text-align:center; margin-bottom:4mm;">'
                f'<img src="data:image/png;base64,{gauge_b64}" '
                f'style="width:46%; max-width:320px;" '
                f'alt="High-Risk Assets gauge">'
                f'</div>'
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("%s PDF gauge render failed: %s", self._log_prefix(), exc)
            pct_display = f"{gauge_value:.1f}%"  # safe: gauge_value guaranteed non-None by line 321 (None coalesced to 0.0)
            gauge_html = (
                f'<p style="text-align:center; font-size:20pt; font-weight:bold; '
                f'color:{_STATUS_COLOR.get(status, "#333")};">{pct_display}</p>'
            )

        # ---- Status badge ----
        status_color = _STATUS_COLOR.get(status, "#757575")
        status_label = _STATUS_LABEL.get(status, status)
        pct_display  = (
            f"{high_risk_pct:.1f}%" if high_risk_pct is not None else "N/A"
        )
        status_html = (
            f'<p style="text-align:center; font-size:10pt; font-weight:bold; '
            f'color:{status_color}; margin:0 0 5mm 0;">'
            f'High-Risk Assets: {pct_display} &nbsp;&middot;&nbsp; {status_label}'
            f'</p>'
        )

        # ---- Supporting numbers ----
        risk_color  = _STATUS_COLOR.get(status, "#555")
        support_html = f"""
<table style="width:52%; margin:0 auto 6mm auto; border-collapse:collapse;">
  <tr>
    <td style="text-align:center; padding:2mm 8mm;
               border-right:0.5pt solid #ddd; vertical-align:middle;">
      <span style="font-size:14pt; font-weight:bold;
             color:{risk_color};">{high_risk_count:,}</span>
      <br><span style="font-size:7.5pt; color:#555;">High-Risk Assets</span>
    </td>
    <td style="text-align:center; padding:2mm 8mm; vertical-align:middle;">
      <span style="font-size:14pt; font-weight:bold;
             color:#1F3864;">{total_on_time:,}</span>
      <br><span style="font-size:7.5pt; color:#555;">Total On-Time Assets</span>
    </td>
  </tr>
</table>"""

        # ---- Top 5 BU table (worst performers = highest %) ----
        top5 = data.chart_data.get("top_5", [])
        if top5:
            rows_html = ""
            for row in top5:
                bu_name  = str(row.get("business_unit", ""))
                bu_num   = int(row.get("numerator",    0))
                bu_den   = int(row.get("denominator",  0))
                bu_score = int(row.get("risk_score",   0))
                rows_html += (
                    f'<tr>'
                    f'<td style="padding:1.5mm 3mm;">{bu_name}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_num:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_den:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_score:,}</td>'
                    f'</tr>'
                )
            bu_table_html = f"""
<h3 class="subsection-heading">Top 5 Worst-Performing Business Units</h3>
<table class="data-table">
  <thead>
    <tr>
      <th>Business Unit</th>
      <th style="text-align:right;">High-Risk Assets</th>
      <th style="text-align:right;">On-Time Assets</th>
      <th style="text-align:right;">Risk Score</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>"""
        else:
            bu_table_html = (
                '<p class="explanatory-text" style="color:#888; font-style:italic;">'
                'No business-unit breakdown available — '
                'assets may lack Application tags or no high-risk assets were found.'
                '</p>'
            )

        # ---- Explanatory paragraph ----
        # Pre-format module-level threshold constants (never None) so the
        # multi-line HTML f-string below contains no inline format specs.
        green_str  = format(_GREEN_THRESHOLD,  ".1f")
        yellow_str = format(_YELLOW_THRESHOLD, ".1f")
        explain_html = f"""
<p class="explanatory-text">
  <strong>What this measures:</strong> The percentage of assets scanned on time
  (licensed scan within the last {ON_TIME_WINDOW_DAYS} days) that carry
  &ge;{_HIGH_RISK_COUNT} Critical or High vulnerabilities (VPR&nbsp;7.0&ndash;10.0) that
  have been open for more than {_AGED_DAYS_THRESHOLD} days.  These assets represent
  the highest sustained risk exposure — they are actively managed yet have significant
  unresolved findings well past normal triage timelines.  Board target is
  &le;{green_str}% (green).  &le;{yellow_str}% is at-risk (amber).
  Above {yellow_str}% is off-target (red).  Business-unit breakdown uses
  the Tenable &ldquo;Application&rdquo; tag category.
</p>"""

        return f"""
<div class="module-section">
  <h2 class="section-heading">{self.DISPLAY_NAME}</h2>
  {gauge_html}
  {status_html}
  {support_html}
  {bu_table_html}
  {explain_html}
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs()
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write the "High-Risk Assets" tab into ``workbook``.

        Tab contents:
        - Overall KPI block (high-risk %, status, supporting numbers, thresholds)
        - Full per-BU breakdown table sorted worst-first (highest % at top)

        Returns
        -------
        list[str]
            ``["High-Risk Assets"]`` on success, ``[]`` on error.
        """
        tab_name = "High-Risk Assets"
        try:
            # D-16 — Phase 3 zero-row standardisation. If both metrics and
            # table_data are empty AND there is no error, emit a single
            # placeholder cell at A1 instead of a fully-empty sheet.
            # This is a behavior change from pre-Phase-3 — surface in SUMMARY.
            empty_metrics = not (
                data.metrics and any(v is not None for v in data.metrics.values())
            )
            empty_tables = not data.table_data
            if empty_metrics and empty_tables and not data.error:
                ws = workbook.create_sheet(tab_name)
                ws["A1"]      = "No data in scope"
                ws["A1"].font = Font(bold=True, color="666666")
                return [tab_name]

            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            m               = data.metrics
            high_risk_pct   = m.get("high_risk_pct")
            high_risk_count = m.get("high_risk_count", 0)
            total_on_time   = m.get("total_on_time", 0)
            status          = m.get("status", "no_data")

            # ---- Overall KPI block ----
            _xl_title(ws, "A1", "High-Risk Assets — Overall Summary")

            status_color_hex = _STATUS_COLOR.get(status, "#757575").lstrip("#")
            pct_str = (
                f"{high_risk_pct:.1f}%" if high_risk_pct is not None else "N/A"
            )

            _xl_kv(ws, 3, "High-Risk Asset %:", pct_str,
                   value_font=Font(bold=True, size=12, color=status_color_hex))
            _xl_kv(ws, 4, "Status:",
                   _STATUS_LABEL.get(status, status),
                   value_font=Font(bold=True, color=status_color_hex))
            _xl_kv(ws, 5, "High-Risk Assets:", f"{high_risk_count:,}")
            _xl_kv(ws, 6, "Total On-Time Assets:", f"{total_on_time:,}")
            _xl_kv(ws, 7, "High-Risk Definition:",
                   f">={_HIGH_RISK_COUNT} Critical/High vulns open >{_AGED_DAYS_THRESHOLD} days")
            _xl_kv(ws, 8, "Scope:",
                   "On-time-scanned assets only (last_licensed_scan_date within last 30 days)")
            _xl_kv(ws, 9, "SLA Thresholds (lower is better):",
                   f"Green <={_GREEN_THRESHOLD:.1f}%  |  "  # safe: module-level float constant, never None
                   f"Amber <={_YELLOW_THRESHOLD:.1f}%  |  "  # safe: module-level float constant, never None
                   f"Red >{_YELLOW_THRESHOLD:.1f}%")  # safe: module-level float constant, never None

            # ---- BU breakdown table (starts at row 11, worst first) ----
            header_row = 11
            headers = [
                "Business Unit", "High-Risk Assets", "On-Time Assets", "Risk Score"
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell           = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = _FILL_HEADER
                cell.alignment = Alignment(horizontal="center")

            for row_offset, row in enumerate(data.table_data or [], start=1):
                data_row = header_row + row_offset

                ws.cell(row=data_row, column=1,
                        value=str(row.get("business_unit", ""))).alignment = (
                    Alignment(horizontal="left")
                )
                ws.cell(row=data_row, column=2,
                        value=int(row.get("numerator",   0))).alignment = (
                    Alignment(horizontal="right")
                )
                ws.cell(row=data_row, column=3,
                        value=int(row.get("denominator", 0))).alignment = (
                    Alignment(horizontal="right")
                )
                score_cell           = ws.cell(row=data_row, column=4,
                                               value=int(row.get("risk_score", 0)))
                score_cell.alignment = Alignment(horizontal="right")

            # ---- Column widths ----
            ws.column_dimensions[get_column_letter(1)].width = 32
            ws.column_dimensions[get_column_letter(2)].width = 18
            ws.column_dimensions[get_column_letter(3)].width = 18
            ws.column_dimensions[get_column_letter(4)].width = 14

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_kpis()
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return three KPI tiles for the HTML email body.

        Returns
        -------
        dict[str, str]
            Keys: "High-Risk Assets %", "High-Risk Assets", "On-Time Assets".
            Returns empty dict if ``data.error`` is set.
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m   = data.metrics
        pct = m.get("high_risk_pct")
        return {
            "High-Risk Assets %": f"{pct:.1f}%" if pct is not None else "N/A",
            "High-Risk Assets":   f"{m.get('high_risk_count', 0):,}",
            "On-Time Assets":     f"{m.get('total_on_time', 0):,}",
        }

    # ------------------------------------------------------------------
    # render_email_panel() — Phase 3 D-02 horizontal-split panel
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Build the per-module email panel — horizontal split layout (D-02).

        Layout: 620px-wide table, 150px gauge cell on the left + 430px text
        cell on the right. Inline CSS only. Outlook-safe ``<table>`` shell
        with explicit ``width=""`` attributes per project email conventions.

        Empty-data behavior (D-15): when ``data.error`` or no
        ``email_gauge_b64`` is available, returns the same 620px shell with
        a gray "No data" placeholder block where the gauge would be.

        Returns
        -------
        str
            Inline-CSS HTML fragment. Returns ``""`` only on a render
            exception (caught by the composer's per-module exception
            isolation pattern).
        """
        try:
            # Empty-data placeholder per D-15
            b64 = (data.metadata or {}).get("email_gauge_b64", "")
            if data.error or not isinstance(b64, str) or not b64.strip():
                return self._render_empty_email_panel()

            pct       = data.metrics.get("high_risk_pct") if data.metrics else None
            headline  = safe_pct(pct)
            status    = (data.metrics or {}).get("status", "no_data")
            rag_color = _RAG_STATUS_COLOR.get(status, _RAG_STATUS_COLOR["no_data"])
            rag_label = _RAG_STATUS_LABEL.get(status, _RAG_STATUS_LABEL["no_data"])
            icon      = STATUS_ICON.get(status,        STATUS_ICON["no_data"])

            cid       = f"{self.MODULE_ID}_gauge"
            # T-03-04-01 — html-escape every module-supplied string
            label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
            driver_esc = html.escape(str(data.driver_narrative or ""), quote=True)

            return (
                '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
                'style="width:620px; max-width:620px; margin:8px 0; '
                'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
                '<tr>'
                '  <td width="150" style="padding:12px; vertical-align:middle; '
                '      text-align:center;">'
                f'    <img src="cid:{cid}" alt="" width="120" height="120" '
                '         style="display:block; margin:0 auto;" />'
                '  </td>'
                '  <td width="430" style="padding:12px; vertical-align:middle;">'
                f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
                f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{headline}</div>'
                f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
                f'{icon} {html.escape(rag_label)}</div>'
                f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
                f'{driver_esc}</div>'
                '  </td>'
                '</tr>'
                '</table>'
            )
        except Exception as exc:    # noqa: BLE001
            logger.error(
                "%s render_email_panel raised: %s",
                self._log_prefix(), exc,
            )
            return ""

    def _render_empty_email_panel(self) -> str:
        """Return the D-15 gray 'No Data' placeholder panel."""
        label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
        driver_esc = html.escape(NO_DATA_DRIVER, quote=True)
        rag_color  = _RAG_STATUS_COLOR["no_data"]
        rag_label  = _RAG_STATUS_LABEL["no_data"]
        icon       = STATUS_ICON["no_data"]
        return (
            '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
            'style="width:620px; max-width:620px; margin:8px 0; '
            'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
            '<tr>'
            '  <td width="150" style="padding:12px; vertical-align:middle; '
            '      text-align:center; background:#f5f5f5; color:#999;">'
            '    <div style="font-size:10pt;">No data</div>'
            '  </td>'
            '  <td width="430" style="padding:12px; vertical-align:middle;">'
            f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
            f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{NO_DATA_HEADLINE}</div>'
            f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
            f'{icon} {html.escape(rag_label)}</div>'
            f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
            f'{driver_esc}</div>'
            '  </td>'
            '</tr>'
            '</table>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs() — Phase 3 D-14 single-tab list
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, "pd.DataFrame"]]:
        """
        Return the module's pre-built analyst rows (D-14 single-tab list).

        Source: ``data.analyst_rows`` populated inside ``compute()``.
        Empty-data: returns ``[]`` (no tab written for this module).
        """
        try:
            if data.error or not data.analyst_rows:
                return []
            return list(data.analyst_rows)
        except Exception as exc:    # noqa: BLE001
            logger.error("[%s] render_analyst_tabs raised: %s", self.MODULE_ID, exc)
            return []

    # ------------------------------------------------------------------
    # get_audit_info()
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "high_risk_pct": (
                    "high_risk_count / total_on_time × 100, rounded to 1 decimal. "
                    "None → 'no_data' when total_on_time == 0."
                ),
                "high_risk_count": (
                    f"Count of on-time assets where the number of open Critical or High "
                    f"findings with days_open > {_AGED_DAYS_THRESHOLD} is "
                    f">= {_HIGH_RISK_COUNT}."
                ),
                "total_on_time": (
                    "Count of deduplicated assets where last_licensed_scan_date IS NOT NULL "
                    f"AND >= report_date − {ON_TIME_WINDOW_DAYS} days."
                ),
                "days_open": (
                    "(report_date − first_found).days.  Findings with null first_found "
                    "produce NaT/NaN and are treated as 0 days (not counted as aged)."
                ),
                "severity_filter": (
                    "severity IN ('critical', 'high').  Severity is VPR-derived "
                    "(vpr_to_severity from config.py) as produced by fetch_all_vulnerabilities()."
                ),
                "BU_breakdown": (
                    "compute_per_bu_breakdown(higher_is_better=False) on on-time assets "
                    "enriched with Application tag. "
                    "Numerator = high-risk assets per BU; denominator = all on-time assets per BU. "
                    "affected = numerator (raw high-risk count). "
                    "Primary sort: affected DESC (largest absolute problem first). "
                    "Secondary sort: percentage DESC (worst % among ties)."
                ),
            },
        }


# ===========================================================================
# Module-private helpers
# ===========================================================================

def _find_high_risk_assets(
    vulns_df:      pd.DataFrame,
    on_time_uuids: set,
    rd_ts:         pd.Timestamp,
) -> tuple[set, pd.Series, pd.DataFrame]:
    """
    Identify on-time assets that qualify as "high-risk".

    An asset is high-risk when it has >= ``_HIGH_RISK_COUNT`` Critical or High
    findings that have been open for more than ``_AGED_DAYS_THRESHOLD`` days.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Open / reopened findings.
    on_time_uuids : set
        UUIDs of on-time-scanned assets.
    rd_ts : pd.Timestamp
        UTC-aware report timestamp used to compute days_open.

    Returns
    -------
    tuple[set, pd.Series, pd.DataFrame]
        ``(high_risk_uuids, aged_counts_per_asset, aged_findings)``
        - ``high_risk_uuids``: set of asset_uuids classified as high-risk.
        - ``aged_counts_per_asset``: pd.Series indexed by asset_uuid with the
          count of aged Critical/High findings per asset (all on-time assets,
          not just high-risk ones).
        - ``aged_findings``: DataFrame slice of ``vulns_df`` containing the
          relevant aged Critical/High findings (the ``relevant[aged_mask]``
          subset). Phase 3 — used by ``compute()`` to produce per-asset
          contributing_finding_ids for the analyst tab. Empty DataFrame
          (with vulns_df.columns) when no aged findings exist.
    """
    empty_frame = vulns_df.iloc[0:0].copy() if not vulns_df.empty else pd.DataFrame()

    if vulns_df.empty:
        return set(), pd.Series(dtype=int), empty_frame

    required = {"asset_uuid", "severity", "first_found"}
    if not required.issubset(vulns_df.columns):
        missing = required - set(vulns_df.columns)
        logger.warning(
            "_find_high_risk_assets: missing columns %s — returning empty set.", missing
        )
        return set(), pd.Series(dtype=int), empty_frame

    # Filter to on-time assets + Critical/High severity
    on_time_mask  = vulns_df["asset_uuid"].isin(on_time_uuids)
    severity_mask = vulns_df["severity"].str.lower().isin(["critical", "high"])
    relevant      = vulns_df[on_time_mask & severity_mask].copy()

    if relevant.empty:
        return set(), pd.Series(dtype=int), empty_frame

    # Compute days_open; NaT first_found → NaN days → treated as 0 (not aged)
    days_open = (rd_ts - relevant["first_found"]).dt.days
    aged_mask = days_open > _AGED_DAYS_THRESHOLD

    aged = relevant[aged_mask].copy()

    if aged.empty:
        return set(), pd.Series(dtype=int), empty_frame

    # Count aged Critical/High findings per asset
    aged_counts = aged.groupby("asset_uuid").size()

    # Assets that meet or exceed the high-risk threshold
    high_risk_uuids = set(aged_counts[aged_counts >= _HIGH_RISK_COUNT].index)

    return high_risk_uuids, aged_counts, aged


def _build_metadata(report_date: Any) -> dict:
    """Return the standard metadata block for this module."""
    return {
        "high_risk_definition":  (
            f">={_HIGH_RISK_COUNT} Critical/High findings open "
            f">{_AGED_DAYS_THRESHOLD} days on an on-time-scanned asset."
        ),
        "severity_scope":        "Critical (VPR 9.0–10.0) and High (VPR 7.0–8.9)",
        "denominator_scope":     (
            f"On-time-scanned assets: last_licensed_scan_date IS NOT NULL "
            f"AND >= report_date − {ON_TIME_WINDOW_DAYS} days."
        ),
        "sla_source":            (
            f"Board-defined thresholds "
            f"(Green <={_GREEN_THRESHOLD}%, "
            f"Amber <={_YELLOW_THRESHOLD}%, Red >{_YELLOW_THRESHOLD}%, "
            f"direction=lower_is_better)"
        ),
        "window":                f"Last {ON_TIME_WINDOW_DAYS} days from report_date",
    }


def _build_summary(
    high_risk_pct:   float | None,
    high_risk_count: int,
    total_on_time:   int,
    status:          str,
) -> str:
    """Build a plain-language narrative sentence for the email body."""
    if high_risk_pct is None:
        return (
            "No on-time-scanned assets were found — "
            "high-risk asset percentage cannot be computed."
        )
    status_label = _STATUS_LABEL.get(status, status)
    return (
        f"{high_risk_pct:.1f}% of on-time-scanned assets are high-risk — "  # safe: high_risk_pct guarded by line 716 early-return on None
        f"{high_risk_count:,} of {total_on_time:,} assets have "
        f">={_HIGH_RISK_COUNT} Critical/High vulnerabilities open "
        f">{_AGED_DAYS_THRESHOLD} days. "
        f"Status: {status_label}."
    )


def _row_bg(pct: float) -> str:
    """Light HTML background-color for a BU table row (lower-is-better)."""
    if pct <= _GREEN_THRESHOLD:
        return "#E8F5E9"   # light green
    if pct <= _YELLOW_THRESHOLD:
        return "#FFF8E1"   # light amber
    return "#FFEBEE"       # light red


def _xl_fill(pct: float) -> PatternFill:
    """openpyxl PatternFill for a high-risk-% cell (lower-is-better)."""
    if pct <= _GREEN_THRESHOLD:
        return _FILL_GREEN
    if pct <= _YELLOW_THRESHOLD:
        return _FILL_YELLOW
    return _FILL_RED


def _xl_title(ws, cell_ref: str, value: str) -> None:
    ws[cell_ref]      = value
    ws[cell_ref].font = Font(bold=True, size=12)


def _xl_kv(ws, row: int, label: str, value: str,
            value_font: Font | None = None) -> None:
    lc      = ws.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True)
    vc      = ws.cell(row=row, column=2, value=value)
    if value_font is not None:
        vc.font = value_font
