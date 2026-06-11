"""
reports/modules/new_vs_remediated_module.py — New vs Remediated MoM trend module.

Displays monthly inflow vs outflow with a net-delta line and an Owner cut.
Inflow is the LOCKED accurate definition (D-15-01): first_found in month M OR
resurfaced_date in month M. Inflow is displayed as TWO stacked components
(D-15-02): net-new (first_found-based) and resurfaced (resurfaced_date-based).

Outflow (Option B / D-15-06): the per-month remediated count is sourced from
the AGGREGATE snapshot field ``fixed_findings_count`` written by
``capture_snapshot()`` (plan 15-02). This module is intentionally NOT in
``_MODULES_NEEDING_FIXED_VULNS`` — ``fixed_vulns_df=None`` is the NORMAL
production path.

Cold-start (QUAL-01): when ``trend_snapshots`` is absent or
``insufficient_data=True``, returns a coherent "Trend data being established"
ModuleData (error=None, cold_start=True in metrics). Never renders NaN%.

Partial-month label (D-15-08): the current in-progress month is labeled
"<YYYY-MM> (MTD — partial)" in all four channels.

Open-count context (QUAL-02): uses ``open_findings_at()`` for any point-in-time
open count — never drops REOPENED findings.

RAG strip (D-15-07):
  Green   — backlog shrinking (net_delta < 0)
  Amber   — flat (net_delta == 0 or backlog data insufficient)
  Red     — backlog growing (net_delta > 0)

Supported options
-----------------
(none for v1; thresholds are directional based on net_delta sign)
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.board_report_utils import extract_owner
from reports.modules.format_utils import safe_format, safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR,
    STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Excel fills
# ---------------------------------------------------------------------------
_FILL_HEADER = PatternFill("solid", fgColor="E3F2FD")
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY   = PatternFill("solid", fgColor="F5F5F5")


def _rag_fill(status: str) -> PatternFill:
    if status == "green":
        return _FILL_GREEN
    if status == "yellow":
        return _FILL_AMBER
    if status == "red":
        return _FILL_RED
    return _FILL_GREY


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _safe_mom_delta(curr: Optional[int], prev: Optional[int]) -> str:
    """
    Return MoM net-delta as "N/A" when prior period is None or zero.

    Never raises ZeroDivisionError; never computes % inline in render methods
    (Pitfall 5).

    Parameters
    ----------
    curr : int or None
        Current-period value.
    prev : int or None
        Prior-period value.  None or 0 → "N/A".

    Returns
    -------
    str
        Formatted percentage string (e.g. "+12.5%", "-3.2%") or "N/A".
    """
    if curr is None or prev is None or prev == 0:
        return "N/A"
    pct = (curr - prev) / prev * 100.0
    sign = "+" if pct >= 0 else ""
    return f"{sign}{pct:.1f}%"


def _month_label(month_str: str, current_period: pd.Period) -> str:
    """
    Return a human-readable label for a snapshot month string.

    Appends "(MTD — partial)" to the current in-progress period (D-15-08)
    so recipients know the bar is not a full month.

    Parameters
    ----------
    month_str : str
        "YYYY-MM" string from the trend snapshot.
    current_period : pd.Period
        The current month period (derived from report_date at compute() time).

    Returns
    -------
    str
        Plain "YYYY-MM" for completed months; "YYYY-MM (MTD — partial)" for
        the current month.
    """
    try:
        snap_period = pd.Period(month_str, "M")
        if snap_period == current_period:
            return f"{month_str} (MTD — partial)"
    except Exception:  # noqa: BLE001
        pass
    return month_str


# ===========================================================================
# Module
# ===========================================================================

@register_module
class NewVsRemediatedModule(BaseModule):
    """
    Monthly New vs Remediated velocity — stacked inflow (net-new + resurfaced)
    vs outflow (snapshot fixed_findings_count), with Owner cut.

    Inflow definition (D-15-01, locked — do not simplify):
      net_new    = first_found in month M
      resurfaced = resurfaced_date in month M AND NOT first_found in month M
      total_inflow = net_new + resurfaced

    Outflow definition (Option B / D-15-06):
      outflow = trend_snapshots[...]["fixed_findings_count"]
      (aggregate from capture_snapshot in plan 15-02)
      When absent/None in a snapshot → "Trend data being established" notice
      for that month's outflow, never silent zero.

    Cold-start (QUAL-01): insufficient trend history → coherent cold-start
    ModuleData; error=None; metrics["cold_start"]=True.

    RAG: Green if last-month net_delta < 0 (backlog shrinking);
         Amber if zero or insufficient; Red if growing (> 0).
    """

    MODULE_ID         = "new_vs_remediated"
    DISPLAY_NAME      = "New vs Remediated"
    DESCRIPTION       = "Monthly inflow (stacked net-new + resurfaced) vs outflow with net delta and Owner cut."
    REQUIRED_DATA     = ["vulns", "assets", "trend_snapshots"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # _build_cold_start_result
    # ------------------------------------------------------------------

    def _build_cold_start_result(self, config: ModuleConfig) -> ModuleData:
        """
        Return a coherent cold-start ModuleData — not an error, just
        insufficient history (QUAL-01).

        ``error`` is ``None`` (cold-start is a valid operational state,
        not a failure); ``metrics["cold_start"]`` is ``True`` so renderers
        can branch on it without parsing the summary text.
        """
        return ModuleData(
            module_id        = self.MODULE_ID,
            display_name     = self.DISPLAY_NAME,
            metrics          = {"cold_start": True},
            table_data       = [],
            chart_data       = {},
            summary_text     = "Trend data being established — available from next month.",
            metadata         = {"cold_start": True},
            driver_narrative = "Trend data being established.",
            analyst_rows     = [],
            rag_strip        = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            ),
            error            = None,
        )

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute monthly new vs remediated velocity from trend snapshots.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame.  Used for Owner-cut context
            and open-count via open_findings_at (QUAL-02).  Requires columns:
            first_found, resurfaced_date, asset_uuid, state.
        assets_df : pd.DataFrame
            Asset DataFrame for extract_owner() Owner cut.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            (No options in v1.)
        **kwargs
            ``trend_snapshots`` (dict, optional) — full ``read_trend()`` result:
            ``{"snapshots": [...], "insufficient_data": bool}``.
            When absent or ``insufficient_data=True``, returns cold-start result
            (QUAL-01).  ``fixed_vulns_df`` is intentionally NOT consumed — outflow
            comes from the aggregate snapshot field (Option B / D-15-06).
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            # ----------------------------------------------------------------
            # QUAL-01 — cold-start guard
            # ----------------------------------------------------------------
            trend_snapshots = kwargs.get("trend_snapshots")
            cold_start = (
                trend_snapshots is None
                or trend_snapshots.get("insufficient_data", True)
            )
            if cold_start:
                logger.info(
                    "%s trend_snapshots absent or insufficient_data=True — cold start.",
                    self._log_prefix(),
                )
                return self._build_cold_start_result(config)

            snapshots: list[dict] = trend_snapshots.get("snapshots", [])
            if not snapshots:
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # QUAL-03 — empty-data guard on vulns_df
            # ----------------------------------------------------------------
            if vulns_df.empty or "state" not in vulns_df.columns:
                logger.warning(
                    "%s vulns_df empty or missing state — cold start.", self._log_prefix()
                )
                return self._build_cold_start_result(config)

            # ----------------------------------------------------------------
            # Current period for partial-month label (D-15-08)
            # ----------------------------------------------------------------
            current_period = pd.Period(report_date, "M")

            # ----------------------------------------------------------------
            # Normalise date columns (expect already normalised by fetcher,
            # but coerce defensively per QUAL-03)
            # ----------------------------------------------------------------
            # WR-02: strip tz before the later .dt.to_period("M") calls — pandas 3.x
            # emits a UserWarning when to_period() drops tz info from a tz-aware
            # series (mirrors the accepted_recast_module fix).
            ff_ts = pd.to_datetime(
                vulns_df["first_found"], utc=True, errors="coerce"
            ).dt.tz_localize(None)
            rs_ts = pd.to_datetime(
                vulns_df.get("resurfaced_date", pd.Series(dtype="object")),
                utc=True, errors="coerce",
            ).dt.tz_localize(None)

            # ----------------------------------------------------------------
            # Per-month inflow computation (D-15-01/02)
            # ----------------------------------------------------------------
            months_labels: list[str]     = []
            net_new_counts: list[int]    = []
            resurfaced_counts: list[int] = []
            outflow_counts: list[Optional[int]] = []
            net_delta_counts: list[Optional[int]] = []
            outflow_cold: list[bool]     = []  # True when fixed_findings_count absent

            for snap in snapshots:
                month_str = snap.get("month", "")
                label     = _month_label(month_str, current_period)
                months_labels.append(label)

                # Inflow — two components (D-15-01/02)
                try:
                    month_period = pd.Period(month_str, "M")
                except Exception:  # noqa: BLE001
                    net_new_counts.append(0)
                    resurfaced_counts.append(0)
                    outflow_counts.append(None)
                    net_delta_counts.append(None)
                    outflow_cold.append(True)
                    continue

                net_new_mask = ff_ts.dt.to_period("M") == month_period
                # Resurfaced: resurfaced_date in month M AND NOT net_new
                # (exclude first_found==resurfaced_date edge to avoid double-count)
                rs_in_month = (
                    rs_ts.notna()
                    & (rs_ts.dt.to_period("M") == month_period)
                )
                resurfaced_mask = rs_in_month & ~net_new_mask

                nn_count  = int(net_new_mask.sum())
                rs_count  = int(resurfaced_mask.sum())
                total_inflow = nn_count + rs_count
                net_new_counts.append(nn_count)
                resurfaced_counts.append(rs_count)

                # Outflow — Option B: from snapshot aggregate (D-15-06)
                fixed_count = snap.get("fixed_findings_count")
                if fixed_count is None:
                    # Older snapshot without this field — cold-start notice for outflow
                    outflow_counts.append(None)
                    net_delta_counts.append(None)
                    outflow_cold.append(True)
                else:
                    outflow_counts.append(int(fixed_count))
                    net_delta_counts.append(total_inflow - int(fixed_count))
                    outflow_cold.append(False)

            # ----------------------------------------------------------------
            # Owner cut for context (extract_owner + open_findings_at QUAL-02)
            # ----------------------------------------------------------------
            from utils.open_count import open_findings_at  # noqa: PLC0415

            open_df = open_findings_at(vulns_df, report_date)
            enriched_assets = extract_owner(assets_df)
            uuid_to_owner = dict(
                zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
            )
            if not open_df.empty:
                owner_col = open_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                owner_counts: dict[str, int] = owner_col.value_counts().to_dict()
            else:
                owner_counts = {}

            # ----------------------------------------------------------------
            # RAG strip — based on last completed month net_delta (D-15-07)
            # Find last month with a valid (non-None) net_delta
            # ----------------------------------------------------------------
            last_valid_delta: Optional[int] = None
            for nd in reversed(net_delta_counts):
                if nd is not None:
                    last_valid_delta = nd
                    break

            if last_valid_delta is None:
                rag_status = "no_data"
                rag_headline = NO_DATA_HEADLINE
            elif last_valid_delta < 0:
                rag_status = "green"
                rag_headline = f"{last_valid_delta:+d}"
            elif last_valid_delta == 0:
                rag_status = "yellow"
                rag_headline = "0"
            else:
                rag_status = "red"
                rag_headline = f"+{last_valid_delta}"

            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = rag_headline,
                status             = rag_status,
            )

            # ----------------------------------------------------------------
            # Driver narrative (for render_email_panel CONTRACT-01)
            # ----------------------------------------------------------------
            if last_valid_delta is not None:
                trend_word = (
                    "shrinking" if last_valid_delta < 0
                    else ("stable" if last_valid_delta == 0 else "growing")
                )
                driver_narrative = (
                    f"Backlog {trend_word}; last month net delta {rag_headline}. "
                    f"Top owners: {', '.join(list(owner_counts.keys())[:3]) or 'none'}."
                )
            else:
                driver_narrative = "Trend data being established."

            # ----------------------------------------------------------------
            # Summary text
            # ----------------------------------------------------------------
            if months_labels:
                any_outflow_cold = any(outflow_cold)
                summary_text = (
                    f"New vs Remediated over {len(months_labels)} month(s). "
                    + (
                        "Some months lack outflow data (trend data being established)."
                        if any_outflow_cold
                        else f"Last month net delta: {rag_headline}."
                    )
                )
            else:
                summary_text = "Trend data being established — available from next month."

            # ----------------------------------------------------------------
            # table_data — per-month summary rows for PDF table
            # ----------------------------------------------------------------
            table_data = []
            for i, label in enumerate(months_labels):
                nn  = net_new_counts[i] if i < len(net_new_counts) else 0
                rs  = resurfaced_counts[i] if i < len(resurfaced_counts) else 0
                out = outflow_counts[i] if i < len(outflow_counts) else None
                nd  = net_delta_counts[i] if i < len(net_delta_counts) else None
                cold = outflow_cold[i] if i < len(outflow_cold) else True
                table_data.append({
                    "month":       label,
                    "net_new":     nn,
                    "resurfaced":  rs,
                    "total_inflow": nn + rs,
                    "outflow":     safe_int(out) if not cold else "—",
                    "net_delta":   safe_int(nd) if not cold else "—",
                })

            # ----------------------------------------------------------------
            # chart_data — drives stacked PDF bar + delta line (D-15-02)
            # ----------------------------------------------------------------
            chart_data = {
                "months":          months_labels,
                "net_new":         net_new_counts,
                "resurfaced":      resurfaced_counts,
                "outflow":         [
                    o if o is not None else 0 for o in outflow_counts
                ],
                "net_delta":       [
                    nd if nd is not None else 0 for nd in net_delta_counts
                ],
                "outflow_cold":    outflow_cold,  # True entries = "—" in display
            }

            # ----------------------------------------------------------------
            # Analyst rows (CONTRACT-02) — aggregate per-month per-Owner
            # Aggregate-only: no per-finding rows (QUAL-05 PII boundary)
            # ----------------------------------------------------------------
            mom_df = pd.DataFrame(table_data)
            owner_df = (
                open_df.assign(
                    owner=open_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                )
                .groupby("owner", dropna=False)
                .size()
                .rename("open_count")
                .reset_index()
                .sort_values("open_count", ascending=False)
                .reset_index(drop=True)
            ) if not open_df.empty else pd.DataFrame(columns=["owner", "open_count"])

            analyst_rows: list[tuple[str, pd.DataFrame]] = [
                ("Monthly Summary", mom_df),
                ("Open by Owner",   owner_df),
            ]

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "months":           len(months_labels),
                    "last_net_delta":   last_valid_delta,
                    "rag_status":       rag_status,
                    "cold_start":       False,
                    "owner_counts":     owner_counts,
                },
                table_data   = table_data,
                chart_data   = chart_data,
                summary_text = summary_text,
                metadata     = {
                    "any_outflow_cold": any(outflow_cold) if outflow_cold else False,
                    "current_period":   str(current_period),
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a summary section with monthly inflow/outflow table.

        Returns an error callout if ``data.error`` is set, and a
        cold-start notice if ``metrics["cold_start"]`` is True.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        if data.metrics.get("cold_start"):
            return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p style="color:#757575;">{data.summary_text}</p>
</div>"""

        # Build table rows — show net-new / resurfaced split (D-15-02)
        table_rows = ""
        for row in data.table_data:
            table_rows += (
                f"<tr>"
                f"<td>{row['month']}</td>"
                f"<td style='text-align:right;'>{row['net_new']}</td>"
                f"<td style='text-align:right;'>{row['resurfaced']}</td>"
                f"<td style='text-align:right;'>{row['total_inflow']}</td>"
                f"<td style='text-align:right;'>{row['outflow']}</td>"
                f"<td style='text-align:right;'>{row['net_delta']}</td>"
                f"</tr>"
            )

        m = data.metrics
        last_nd = m.get("last_net_delta")
        rag_status = m.get("rag_status", "no_data")
        rag_color = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        rag_label = STATUS_LABEL.get(rag_status, STATUS_LABEL["no_data"])

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p>
    <strong>RAG Status:</strong>
    <span style="color:{rag_color};font-weight:bold;">{rag_label}</span>
    &nbsp;|&nbsp;
    <strong>Last Month Net Delta:</strong> {safe_int(last_nd)}
  </p>
  <table class="data-table" style="width:100%;margin-top:8pt;">
    <thead>
      <tr>
        <th>Month</th>
        <th style="text-align:right;">Net-New</th>
        <th style="text-align:right;">Resurfaced</th>
        <th style="text-align:right;">Total Inflow</th>
        <th style="text-align:right;">Outflow</th>
        <th style="text-align:right;">Net Delta</th>
      </tr>
    </thead>
    <tbody>
      {table_rows}
    </tbody>
  </table>
  <p class="explanatory-text">
    Inflow = net-new (first seen this month) + resurfaced (re-emerged this month).
    Outflow = findings remediated this month (from trend snapshot).
    Net Delta = Total Inflow − Outflow; negative means backlog shrinking.
    Months labeled "(MTD — partial)" are in-progress.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write a ``New vs Remediated`` tab with monthly summary.

        Returns ``[]`` on exception; writes an error/cold-start row if set.
        """
        tab_name = "New vs Remediated"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            if data.metrics.get("cold_start"):
                ws["A1"] = "New vs Remediated"
                ws["A1"].font = Font(bold=True, size=13)
                ws["A2"] = data.summary_text
                ws["A2"].font = Font(italic=True, color="757575")
                return [tab_name]

            ws["A1"] = "New vs Remediated — Monthly Summary"
            ws["A1"].font = Font(bold=True, size=13)

            headers = [
                "Month", "Net-New", "Resurfaced", "Total Inflow", "Outflow", "Net Delta",
            ]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _FILL_HEADER

            for row_idx, row in enumerate(data.table_data, start=4):
                ws.cell(row=row_idx, column=1, value=row["month"])
                ws.cell(row=row_idx, column=2, value=row["net_new"])
                ws.cell(row=row_idx, column=3, value=row["resurfaced"])
                ws.cell(row=row_idx, column=4, value=row["total_inflow"])
                ws.cell(row=row_idx, column=5, value=str(row["outflow"]))
                ws.cell(row=row_idx, column=6, value=str(row["net_delta"]))

            # Column widths
            widths = [28, 12, 14, 16, 14, 14]
            for col_idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = w

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01 — NOT render_email_kpis)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Returns ``""`` when ``data.error`` is set or ``driver_narrative`` is empty.
        Cold-start renders the "Trend data being established" notice.
        Uses only ``safe_pct`` / ``safe_int`` — never raw f-string on possibly-None.
        """
        if data.error:
            return ""

        if data.metrics.get("cold_start"):
            return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#F5F5F5;border-left:4px solid #757575;">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;color:#757575;">{data.summary_text}</span>
    </td>
  </tr>
</table>"""

        if not data.driver_narrative:
            return ""

        m = data.metrics
        rag_status = m.get("rag_status", "no_data")
        rag_color  = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        last_nd    = m.get("last_net_delta")
        months     = safe_int(m.get("months"))

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#F3E5F5;border-left:4px solid {rag_color};">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;">
        {months} month(s) of trend &mdash; last month net delta: {safe_int(last_nd)}
      </span><br>
      <em style="font-size:11px;color:#555;">{data.driver_narrative}</em>
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns ``data.analyst_rows`` when data is valid; ``[]`` on error or
        cold-start (no detail to drill down into).
        Columns are aggregate-only (QUAL-05 — no per-finding PII).
        """
        if data.error or not data.analyst_rows:
            return []
        if data.metrics.get("cold_start"):
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03, honors data.rag_strip)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present;
        falls back to a gray "No Data" cell if empty or error.
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip
