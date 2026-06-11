"""
reports/modules/accepted_recast_module.py — Accepted & Recast exception posture module.

Tracks risk-management exceptions — ACCEPTED and RECASTED findings are tracked
SEPARATELY (never aggregated — Pitfall 6b). Expired-rule findings are excluded
from current counts and flagged "pending re-evaluation" (Pitfall 6a). Finding
counts drive the headline; rule-level detail goes to the analyst tab only
(Pitfall 6c). Current-vs-prior-month MoM delta sourced from trend snapshots.

Primary classification
----------------------
``severity_modification_type`` field on vulns_df (uppercase coerced):
  - ACCEPTED  → accepted_df
  - RECASTED  → recasted_df
  - ""/"NONE"/anything else → unmodified; excluded from both counts

Expiry cross-check (Pitfall 6a)
--------------------------------
``recast_rules_df`` kwarg (from fetch_recast_rules):
  columns: rule_id, rule_name, plugin_id (int nullable), action (RECAST|ACCEPT),
           new_severity, original_severity, expires_at, created_at.
  When present+non-empty: exclude findings whose recast_rule_uuid maps to an
  expired rule (expires_at < report_date) and surface them as
  "pending re-evaluation".
  When None: log warning, skip cross-check, finding-level counts still computed.

Rate denominator
----------------
Total open findings (state in {OPEN, REOPENED}).

MoM delta (QUAL-01 / Pitfall 5)
---------------------------------
Read accepted_count/recast_count from trend_snapshots prior month.
If trend_snapshots absent, insufficient_data=True, or prior month absent
→ OMIT delta arrow entirely (no "▲ 0%", no NaN%).

RAG thresholds (D-15-07)
-------------------------
Exception rate = (accepted + recasted) / total_open * 100.
  Green  ≤ _DEFAULT_GREEN_EXCEPTION_RATE  (5.0%)
  Yellow ≤ _DEFAULT_YELLOW_EXCEPTION_RATE (15.0%)
  Red    >  15.0%
Overridable via config.options.

Supported options
-----------------
green_exception_rate  : float  — default 5.0
yellow_exception_rate : float  — default 15.0
"""

from __future__ import annotations

import logging
from typing import Any, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.board_report_utils import extract_owner
from reports.modules.format_utils import safe_format, safe_int, safe_pct
from reports.modules.rag_utils import (
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
    STATUS_COLOR,
    STATUS_LABEL,
    build_rag_strip_entry,
    rag_status_from_value,
)
from reports.modules.registry import register_module

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# RAG thresholds — overridable via module_options (D-15-07)
# ---------------------------------------------------------------------------
_DEFAULT_GREEN_EXCEPTION_RATE  = 5.0   # exception rate % below which is green
_DEFAULT_YELLOW_EXCEPTION_RATE = 15.0  # exception rate % below which is yellow

# ---------------------------------------------------------------------------
# Excel fills
# ---------------------------------------------------------------------------
_FILL_HEADER = PatternFill("solid", fgColor="E3F2FD")
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")
_FILL_AMBER  = PatternFill("solid", fgColor="FFF9C4")
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")
_FILL_GREY   = PatternFill("solid", fgColor="F5F5F5")


def _rag_fill(status: str) -> PatternFill:
    if status == "green":
        return _FILL_GREEN
    if status == "yellow":
        return _FILL_AMBER
    if status == "red":
        return _FILL_RED
    return _FILL_GREY


# ---------------------------------------------------------------------------
# Safe MoM delta helper (QUAL-01 / Pitfall 5)
# Returns a delta-arrow string or None when prior period is absent/zero.
# Never returns a "▲ 0%" or NaN string.
# ---------------------------------------------------------------------------

def _safe_delta_arrow(curr: Optional[int], prev: Optional[int]) -> Optional[str]:
    """
    Compute MoM delta arrow string, or None when prior is unavailable.

    Parameters
    ----------
    curr : int or None
        Current-period finding count.
    prev : int or None
        Prior-period finding count.  None → omit delta entirely.

    Returns
    -------
    str or None
        E.g. "▲ +3", "▼ -2", "→ 0", or None when prior is missing.
    """
    if curr is None or prev is None:
        return None
    delta = curr - prev
    if delta > 0:
        return f"▲ +{delta}"
    if delta < 0:
        return f"▼ {delta}"
    return "→ 0"


# ===========================================================================
# Module
# ===========================================================================

@register_module
class AcceptedRecastModule(BaseModule):
    """
    Accepted & Recast exception posture with MoM delta and expiry awareness.

    ACCEPTED and RECASTED findings are tracked SEPARATELY (Pitfall 6b — never
    silently aggregated). Expired-rule findings excluded via recast_rules_df
    cross-check (Pitfall 6a). Headline = finding counts; rule-level detail in
    analyst tab only (Pitfall 6c).

    Severity classification:
      severity_modification_type.str.upper().isin({"ACCEPTED"}) → accepted
      severity_modification_type.str.upper().isin({"RECASTED"}) → recasted
      "" / "NONE" / other → unmodified; excluded from both (Pitfall 6b)

    Rate denominator = total open findings (state in {OPEN, REOPENED}).
    Exception rate   = (accepted + recasted) / total_open * 100.
    RAG: lower_is_better; green≤5%, yellow≤15%, red>15% (overridable).

    MoM delta (QUAL-01): sourced from trend_snapshots accepted_count /
    recast_count fields; prior-month absent → omit delta (no "▲ 0%", no NaN%).
    """

    MODULE_ID         = "accepted_recast"
    DISPLAY_NAME      = "Accepted & Recast"
    DESCRIPTION       = "Risk-management exception posture — ACCEPTED and RECASTED tracked separately with MoM delta and expiry awareness."
    REQUIRED_DATA     = ["vulns", "assets", "recast_rules", "trend_snapshots"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute accepted/recast exception counts, expiry-exclusion, MoM delta,
        Owner cut, and RAG status.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Tag-filtered vulnerability DataFrame. Must include
            ``severity_modification_type``, ``recast_rule_uuid``, ``state``,
            ``asset_uuid`` columns.
        assets_df : pd.DataFrame
            Asset DataFrame for extract_owner() Owner cut.
        report_date : datetime
            Report run timestamp (UTC-aware).
        config : ModuleConfig
            Options: ``green_exception_rate`` (float),
            ``yellow_exception_rate`` (float).
        **kwargs
            ``recast_rules_df`` (pd.DataFrame or None) — from
            fetch_recast_rules(); used for expiry cross-check (Pitfall 6a).
            When None, cross-check is skipped with a warning.

            ``trend_snapshots`` (dict or None) — from read_trend(); used for
            MoM delta. When absent or insufficient_data=True, delta is omitted.
        """
        logger.debug(
            "%s compute() — vulns_df rows: %d",
            self._log_prefix(), len(vulns_df),
        )

        try:
            green_threshold  = float(config.options.get("green_exception_rate",  _DEFAULT_GREEN_EXCEPTION_RATE))
            yellow_threshold = float(config.options.get("yellow_exception_rate", _DEFAULT_YELLOW_EXCEPTION_RATE))

            # ----------------------------------------------------------------
            # QUAL-03 — empty-data guard: zero exceptions is valid, not an error
            # ----------------------------------------------------------------
            if vulns_df.empty or "state" not in vulns_df.columns:
                # Return a coherent zero-exception result rather than _empty_result,
                # because empty scope is a valid operational state for this module.
                return self._build_zero_exception_result(config, green_threshold, yellow_threshold)

            # ----------------------------------------------------------------
            # Classification (Pitfall 6b): use .isin(); "" and "NONE" excluded
            # ----------------------------------------------------------------
            mod_type = vulns_df["severity_modification_type"].astype(str).str.upper()
            accepted_mask = mod_type.isin({"ACCEPTED"})
            recasted_mask = mod_type.isin({"RECASTED"})

            accepted_df = vulns_df[accepted_mask].copy()
            recasted_df = vulns_df[recasted_mask].copy()

            # ----------------------------------------------------------------
            # Rate denominator: total open findings (OPEN or REOPENED)
            # ----------------------------------------------------------------
            open_mask  = vulns_df["state"].astype(str).str.upper().isin({"OPEN", "REOPENED"})
            total_open = int(open_mask.sum())

            # ----------------------------------------------------------------
            # Expiry cross-check (Pitfall 6a)
            # ----------------------------------------------------------------
            recast_rules_df: Optional[pd.DataFrame] = kwargs.get("recast_rules_df")
            pending_reeval_count = 0

            if recast_rules_df is not None and not recast_rules_df.empty:
                today = pd.Timestamp(report_date).tz_convert("UTC")
                expired_mask = (
                    recast_rules_df["expires_at"].notna()
                    & (pd.to_datetime(recast_rules_df["expires_at"], utc=True) < today)
                )
                expired_ids: set = set(
                    recast_rules_df.loc[expired_mask, "rule_id"]
                )
                if expired_ids and "recast_rule_uuid" in accepted_df.columns:
                    pending_accepted = accepted_df[accepted_df["recast_rule_uuid"].isin(expired_ids)]
                    pending_reeval_count += len(pending_accepted)
                    accepted_df = accepted_df[~accepted_df["recast_rule_uuid"].isin(expired_ids)]
                if expired_ids and "recast_rule_uuid" in recasted_df.columns:
                    pending_recasted = recasted_df[recasted_df["recast_rule_uuid"].isin(expired_ids)]
                    pending_reeval_count += len(pending_recasted)
                    recasted_df = recasted_df[~recasted_df["recast_rule_uuid"].isin(expired_ids)]
            else:
                logger.warning(
                    "%s recast_rules_df absent — expiry cross-check skipped.",
                    self._log_prefix(),
                )

            # ----------------------------------------------------------------
            # Separate counts (Pitfall 6b — NEVER aggregate silently)
            # ----------------------------------------------------------------
            accepted_count = len(accepted_df)
            recast_count   = len(recasted_df)
            total_exceptions = accepted_count + recast_count

            # ----------------------------------------------------------------
            # Exception rate
            # ----------------------------------------------------------------
            exception_rate: Optional[float] = None
            if total_open > 0:
                exception_rate = round(total_exceptions / total_open * 100.0, 2)

            # ----------------------------------------------------------------
            # Owner cut via extract_owner
            # ----------------------------------------------------------------
            enriched_assets = extract_owner(assets_df)
            uuid_to_owner   = dict(
                zip(enriched_assets["asset_uuid"], enriched_assets["owner"])
            )

            if not accepted_df.empty and "asset_uuid" in accepted_df.columns:
                accepted_df = accepted_df.assign(
                    owner=accepted_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                )
            elif not accepted_df.empty:
                accepted_df = accepted_df.assign(owner=pd.Series("Unassigned", index=accepted_df.index, dtype="object"))

            if not recasted_df.empty and "asset_uuid" in recasted_df.columns:
                recasted_df = recasted_df.assign(
                    owner=recasted_df["asset_uuid"].map(uuid_to_owner).fillna("Unassigned")
                )
            elif not recasted_df.empty:
                recasted_df = recasted_df.assign(owner=pd.Series("Unassigned", index=recasted_df.index, dtype="object"))

            owner_accepted: dict[str, int] = (
                accepted_df["owner"].value_counts().to_dict()
                if not accepted_df.empty and "owner" in accepted_df.columns
                else {}
            )
            owner_recasted: dict[str, int] = (
                recasted_df["owner"].value_counts().to_dict()
                if not recasted_df.empty and "owner" in recasted_df.columns
                else {}
            )

            # ----------------------------------------------------------------
            # MoM delta (QUAL-01 / Pitfall 5):
            # Read accepted_count / recast_count from trend_snapshots.
            # If absent, insufficient_data, or prior month missing → None (omit arrow).
            # ----------------------------------------------------------------
            accepted_delta_str: Optional[str] = None
            recast_delta_str:   Optional[str] = None

            trend_snapshots = kwargs.get("trend_snapshots")
            if (
                trend_snapshots is not None
                and not trend_snapshots.get("insufficient_data", True)
            ):
                snapshots: list[dict] = trend_snapshots.get("snapshots", [])
                # Find prior completed month snapshot (last snapshot before current month)
                report_ts     = pd.Timestamp(report_date).tz_convert("UTC")
                # Strip timezone before to_period to avoid UserWarning in pandas 3.x
                current_month = report_ts.tz_localize(None).to_period("M")

                prior_snap: Optional[dict] = None
                for snap in reversed(snapshots):
                    month_str = snap.get("month", "")
                    try:
                        snap_period = pd.Period(month_str, "M")
                        if snap_period < current_month:
                            prior_snap = snap
                            break
                    except Exception:  # noqa: BLE001
                        continue

                if prior_snap is not None:
                    prev_accepted = prior_snap.get("accepted_count")
                    prev_recast   = prior_snap.get("recast_count")
                    accepted_delta_str = _safe_delta_arrow(accepted_count, prev_accepted)
                    recast_delta_str   = _safe_delta_arrow(recast_count,   prev_recast)
                # else: prior_snap absent → leave deltas as None (no arrow)
            # else: no snapshots → leave deltas as None (QUAL-01)

            # ----------------------------------------------------------------
            # RAG strip
            # ----------------------------------------------------------------
            status = rag_status_from_value(
                exception_rate,
                green_threshold=green_threshold,
                yellow_threshold=yellow_threshold,
                direction="lower_is_better",
            )
            if total_exceptions == 0:
                rag_headline = "0 exceptions"
                status = "green"
            elif exception_rate is not None:
                rag_headline = f"{total_exceptions} ({safe_pct(exception_rate)})"
            else:
                rag_headline = str(total_exceptions)

            rag_strip = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = rag_headline,
                status             = status,
            )

            # ----------------------------------------------------------------
            # Driver narrative
            # ----------------------------------------------------------------
            top_owners_all: dict[str, int] = {}
            for owner, cnt in list(owner_accepted.items()) + list(owner_recasted.items()):
                top_owners_all[owner] = top_owners_all.get(owner, 0) + cnt
            top_3 = sorted(top_owners_all.items(), key=lambda x: x[1], reverse=True)[:3]
            top_owners_str = ", ".join(f"{o} ({c})" for o, c in top_3) if top_3 else "none"

            if total_exceptions == 0:
                driver_narrative = "0 managed exceptions in scope."
            else:
                parts = [
                    f"{accepted_count} accepted, {recast_count} recasted"
                ]
                if exception_rate is not None:
                    parts.append(f"{safe_pct(exception_rate)} of open findings")
                if pending_reeval_count > 0:
                    parts.append(f"{pending_reeval_count} pending re-evaluation")
                driver_narrative = (
                    "; ".join(parts)
                    + f". Top owners: {top_owners_str}."
                )

            # ----------------------------------------------------------------
            # Summary text
            # ----------------------------------------------------------------
            summary_parts = [
                f"{accepted_count} accepted and {recast_count} recasted exception(s) currently active."
            ]
            if exception_rate is not None:
                summary_parts.append(f"Exception rate: {safe_pct(exception_rate)} of open findings.")
            if pending_reeval_count > 0:
                summary_parts.append(
                    f"{pending_reeval_count} finding(s) pending re-evaluation (expired rules)."
                )
            summary_text = " ".join(summary_parts)

            # ----------------------------------------------------------------
            # table_data (per-owner rows for PDF)
            # ----------------------------------------------------------------
            table_data = []
            for owner, cnt in sorted(top_owners_all.items(), key=lambda x: x[1], reverse=True):
                table_data.append({
                    "owner":          owner,
                    "accepted_count": owner_accepted.get(owner, 0),
                    "recast_count":   owner_recasted.get(owner, 0),
                    "total":          cnt,
                })

            # ----------------------------------------------------------------
            # Analyst rows (Pitfall 6c): per-RULE detail, not per-finding rows.
            # Columns: rule_id, action, plugin_id, original_severity,
            #          new_severity, expires_at, created_at, finding_count,
            #          filter_summary (via _summarize_filter — no inline parse)
            # ----------------------------------------------------------------
            rule_rows_df, owner_df = self._build_analyst_rows(
                accepted_df, recasted_df, recast_rules_df,
                owner_accepted, owner_recasted,
            )

            analyst_rows: list[tuple[str, pd.DataFrame]] = [
                ("Rule Detail", rule_rows_df),
                ("By Owner",    owner_df),
            ]

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "accepted_count":      accepted_count,
                    "recast_count":        recast_count,
                    "total_exceptions":    total_exceptions,
                    "exception_rate":      exception_rate,
                    "total_open":          total_open,
                    "pending_reeval":      pending_reeval_count,
                    "rag_status":          status,
                    "accepted_delta":      accepted_delta_str,
                    "recast_delta":        recast_delta_str,
                    "owner_accepted":      owner_accepted,
                    "owner_recasted":      owner_recasted,
                },
                table_data   = table_data,
                chart_data   = {
                    "labels": ["Accepted", "Recasted"],
                    "counts": [accepted_count, recast_count],
                },
                summary_text     = summary_text,
                metadata         = {
                    "green_threshold":  green_threshold,
                    "yellow_threshold": yellow_threshold,
                    "expiry_checked":   recast_rules_df is not None,
                    "pending_reeval":   pending_reeval_count,
                },
                driver_narrative = driver_narrative,
                analyst_rows     = analyst_rows,
                rag_strip        = rag_strip,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # _build_zero_exception_result
    # ------------------------------------------------------------------

    def _build_zero_exception_result(
        self,
        config:           ModuleConfig,
        green_threshold:  float,
        yellow_threshold: float,
    ) -> ModuleData:
        """
        Return a coherent ModuleData for zero-exception / empty-scope input.

        Zero exceptions is a valid operational state (QUAL-03): error=None,
        RAG=green, driver_narrative="0 managed exceptions in scope."
        """
        rag_strip = build_rag_strip_entry(
            display_name       = self.DISPLAY_NAME,
            headline_value_str = "0 exceptions",
            status             = "green",
        )
        return ModuleData(
            module_id        = self.MODULE_ID,
            display_name     = self.DISPLAY_NAME,
            metrics          = {
                "accepted_count":   0,
                "recast_count":     0,
                "total_exceptions": 0,
                "exception_rate":   None,
                "total_open":       0,
                "pending_reeval":   0,
                "rag_status":       "green",
                "accepted_delta":   None,
                "recast_delta":     None,
                "owner_accepted":   {},
                "owner_recasted":   {},
            },
            table_data       = [],
            chart_data       = {"labels": ["Accepted", "Recasted"], "counts": [0, 0]},
            summary_text     = "0 accepted and 0 recasted exception(s) currently active.",
            metadata         = {
                "green_threshold":  green_threshold,
                "yellow_threshold": yellow_threshold,
                "expiry_checked":   False,
                "pending_reeval":   0,
            },
            driver_narrative = "0 managed exceptions in scope.",
            analyst_rows     = [
                ("Rule Detail", pd.DataFrame(columns=[
                    "rule_id", "action", "plugin_id", "original_severity",
                    "new_severity", "expires_at", "created_at",
                    "finding_count", "filter_summary",
                ])),
                ("By Owner", pd.DataFrame(columns=["owner", "accepted_count", "recast_count"])),
            ],
            rag_strip        = rag_strip,
            error            = None,
        )

    # ------------------------------------------------------------------
    # _build_analyst_rows  (Pitfall 6c — rule-level detail only)
    # ------------------------------------------------------------------

    def _build_analyst_rows(
        self,
        accepted_df:    pd.DataFrame,
        recasted_df:    pd.DataFrame,
        recast_rules_df: Optional[pd.DataFrame],
        owner_accepted: dict[str, int],
        owner_recasted: dict[str, int],
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Build analyst-tab DataFrames:

        - Rule Detail: per-rule rows with finding_count and filter_summary.
          Uses _summarize_filter() for filter display (no inline tree parse).
          Accepts None plugin_id (Pitfall 6c).
        - By Owner: accepted_count + recast_count per owner.

        Returns
        -------
        (rule_rows_df, owner_df)
        """
        _rule_cols = [
            "rule_id", "action", "plugin_id", "original_severity",
            "new_severity", "expires_at", "created_at",
            "finding_count", "filter_summary",
        ]

        try:
            if recast_rules_df is None or recast_rules_df.empty:
                # No rule detail available — build summary from finding UUIDs only
                rows = []
                for action, df in [("ACCEPT", accepted_df), ("RECAST", recasted_df)]:
                    if df.empty:
                        continue
                    uuid_col = "recast_rule_uuid" if "recast_rule_uuid" in df.columns else None
                    if uuid_col:
                        for rule_uuid, grp in df.groupby(uuid_col, dropna=False):
                            rows.append({
                                "rule_id":           str(rule_uuid),
                                "action":            action,
                                "plugin_id":         None,
                                "original_severity": None,
                                "new_severity":      None,
                                "expires_at":        None,
                                "created_at":        None,
                                "finding_count":     len(grp),
                                "filter_summary":    "No filter",
                            })
                rule_rows_df = pd.DataFrame(rows, columns=_rule_cols) if rows else pd.DataFrame(columns=_rule_cols)
            else:
                # Cross-reference rules with finding counts (Pitfall 6c)
                from data.fetchers import _summarize_filter  # noqa: PLC0415

                # Build rule_uuid → finding_count maps
                acc_uuid_counts: dict[str, int] = {}
                rec_uuid_counts: dict[str, int] = {}
                if not accepted_df.empty and "recast_rule_uuid" in accepted_df.columns:
                    acc_uuid_counts = accepted_df["recast_rule_uuid"].value_counts().to_dict()
                if not recasted_df.empty and "recast_rule_uuid" in recasted_df.columns:
                    rec_uuid_counts = recasted_df["recast_rule_uuid"].value_counts().to_dict()

                rows = []
                for _, rule in recast_rules_df.iterrows():
                    rule_id   = rule.get("rule_id", "")
                    action    = str(rule.get("action", "")).upper()
                    plugin_id = rule.get("plugin_id")  # nullable int — leave as-is
                    uuid_counts = acc_uuid_counts if action == "ACCEPT" else rec_uuid_counts
                    finding_count = uuid_counts.get(str(rule_id), 0)

                    # _summarize_filter: no inline tree parse (Pitfall 6c)
                    filter_dict   = rule.get("filter", {}) if isinstance(rule.get("filter"), dict) else {}
                    filter_summary = _summarize_filter(filter_dict)

                    rows.append({
                        "rule_id":           rule_id,
                        "action":            action,
                        "plugin_id":         plugin_id,
                        "original_severity": rule.get("original_severity"),
                        "new_severity":      rule.get("new_severity"),
                        "expires_at":        rule.get("expires_at"),
                        "created_at":        rule.get("created_at"),
                        "finding_count":     finding_count,
                        "filter_summary":    filter_summary,
                    })

                rule_rows_df = pd.DataFrame(rows, columns=_rule_cols) if rows else pd.DataFrame(columns=_rule_cols)

        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "%s _build_analyst_rows failed: %s — returning empty Rule Detail tab.",
                self._log_prefix(), exc,
            )
            rule_rows_df = pd.DataFrame(columns=_rule_cols)

        # By Owner tab
        all_owners = set(owner_accepted.keys()) | set(owner_recasted.keys())
        if all_owners:
            owner_rows = [
                {
                    "owner":          o,
                    "accepted_count": owner_accepted.get(o, 0),
                    "recast_count":   owner_recasted.get(o, 0),
                }
                for o in sorted(all_owners, key=lambda x: -(owner_accepted.get(x, 0) + owner_recasted.get(x, 0)))
            ]
            owner_df = pd.DataFrame(owner_rows)
        else:
            owner_df = pd.DataFrame(columns=["owner", "accepted_count", "recast_count"])

        return rule_rows_df, owner_df

    # ------------------------------------------------------------------
    # render_pdf_section
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a summary section with separate accepted/recasted counts,
        exception rate, pending re-evaluation notice, and per-owner table.

        Returns an error callout if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m              = data.metrics
        accepted       = safe_int(m.get("accepted_count"))
        recasted       = safe_int(m.get("recast_count"))
        rate_str       = safe_pct(m.get("exception_rate"))
        pending        = m.get("pending_reeval", 0)
        rag_status     = m.get("rag_status", "no_data")
        rag_color      = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        rag_label      = STATUS_LABEL.get(rag_status, STATUS_LABEL["no_data"])
        accepted_delta = m.get("accepted_delta")
        recast_delta   = m.get("recast_delta")

        acc_delta_html = (
            f" <small style='color:#757575;'>({accepted_delta})</small>"
            if accepted_delta else ""
        )
        rec_delta_html = (
            f" <small style='color:#757575;'>({recast_delta})</small>"
            if recast_delta else ""
        )

        pending_html = ""
        if pending > 0:
            pending_html = (
                f"<p style='color:#f57c00;'>"
                f"<strong>{safe_int(pending)}</strong> finding(s) pending re-evaluation "
                f"(linked to expired rules — review for re-classification)."
                f"</p>"
            )

        # Per-owner table rows
        table_rows = ""
        for row in data.table_data:
            table_rows += (
                f"<tr>"
                f"<td>{row['owner']}</td>"
                f"<td style='text-align:right;'>{row['accepted_count']}</td>"
                f"<td style='text-align:right;'>{row['recast_count']}</td>"
                f"<td style='text-align:right;'>{row['total']}</td>"
                f"</tr>"
            )

        owner_table = ""
        if table_rows:
            owner_table = f"""
<table class="data-table" style="width:100%;margin-top:8pt;">
  <thead>
    <tr>
      <th>Owner</th>
      <th style="text-align:right;">Accepted</th>
      <th style="text-align:right;">Recasted</th>
      <th style="text-align:right;">Total</th>
    </tr>
  </thead>
  <tbody>
    {table_rows}
  </tbody>
</table>"""

        return f"""
<div class="module-section">
  <h2 class="section-heading">{data.display_name}</h2>
  <p>
    <strong>RAG Status:</strong>
    <span style="color:{rag_color};font-weight:bold;">{rag_label}</span>
    &nbsp;|&nbsp;
    <strong>Exception Rate:</strong> {rate_str}
  </p>
  <p>
    <strong>Accepted Findings:</strong> {accepted}{acc_delta_html}
    &nbsp;&nbsp;
    <strong>Recasted Findings:</strong> {recasted}{rec_delta_html}
  </p>
  {pending_html}
  {owner_table}
  <p class="explanatory-text">
    <em>Accepted</em> exceptions: findings where the risk has been formally accepted
    by the business. <em>Recasted</em> exceptions: findings whose severity has been
    downgraded via a recast rule. Exception rate = (accepted + recasted) / total open
    findings. Findings linked to expired rules are flagged for re-evaluation.
  </p>
</div>"""

    # ------------------------------------------------------------------
    # render_excel_tabs
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write an ``Accepted & Recast`` summary tab.

        Returns ``[]`` on exception; writes an error row if ``data.error`` is set.
        """
        tab_name = "Accepted & Recast"

        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            m = data.metrics

            ws["A1"] = "Accepted & Recast Exception Posture"
            ws["A1"].font = Font(bold=True, size=13)

            ws["A2"] = "Accepted Findings"
            ws["B2"] = m.get("accepted_count", 0)
            ws["C2"] = m.get("accepted_delta") or ""

            ws["A3"] = "Recasted Findings"
            ws["B3"] = m.get("recast_count", 0)
            ws["C3"] = m.get("recast_delta") or ""

            ws["A4"] = "Exception Rate"
            exc_rate = m.get("exception_rate")
            ws["B4"] = f"{safe_pct(exc_rate)}" if exc_rate is not None else "—"

            pending = m.get("pending_reeval", 0)
            ws["A5"] = "Pending Re-evaluation"
            ws["B5"] = pending
            if pending > 0:
                ws["A5"].font = Font(color="F57C00")
                ws["B5"].font = Font(color="F57C00")

            # Per-owner table at row 7
            headers = ["Owner", "Accepted", "Recasted", "Total"]
            for col_idx, header in enumerate(headers, start=1):
                cell      = ws.cell(row=7, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = _FILL_HEADER

            rag_status = m.get("rag_status", "no_data")
            for row_idx, row in enumerate(data.table_data, start=8):
                ws.cell(row=row_idx, column=1, value=row["owner"])
                ws.cell(row=row_idx, column=2, value=row["accepted_count"])
                ws.cell(row=row_idx, column=3, value=row["recast_count"])
                ws.cell(row=row_idx, column=4, value=row["total"])
                fill = _rag_fill(rag_status)
                for col_idx in range(1, 5):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

            widths = [28, 14, 14, 14, 20]
            for col_idx, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = w

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel  (CONTRACT-01 — NOT render_email_kpis)
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        CONTRACT-01: modular email body panel.

        Returns ``""`` when ``data.error`` is set or ``driver_narrative`` is empty.
        Uses only ``safe_pct`` / ``safe_int`` — never raw f-string on possibly-None.
        Distinguishes Accepted vs Recasted with separate counts (Pitfall 6b).
        """
        if data.error or not data.driver_narrative:
            return ""

        m              = data.metrics
        accepted       = safe_int(m.get("accepted_count"))
        recasted       = safe_int(m.get("recast_count"))
        rate_str       = safe_pct(m.get("exception_rate"))
        rag_status     = m.get("rag_status", "no_data")
        rag_color      = STATUS_COLOR.get(rag_status, STATUS_COLOR["no_data"])
        accepted_delta = m.get("accepted_delta")
        recast_delta   = m.get("recast_delta")
        pending        = m.get("pending_reeval", 0)

        acc_delta_str = f" ({accepted_delta})" if accepted_delta else ""
        rec_delta_str = f" ({recast_delta})"   if recast_delta   else ""

        pending_html = ""
        if pending > 0:
            pending_html = (
                f"<br><span style='font-size:11px;color:#f57c00;'>"
                f"{safe_int(pending)} pending re-evaluation (expired rules)"
                f"</span>"
            )

        return f"""
<table style="width:100%;border-collapse:collapse;font-family:Arial,sans-serif;margin-bottom:8px;">
  <tr>
    <td style="padding:8px 12px;background:#FFF3E0;border-left:4px solid {rag_color};">
      <strong style="font-size:13px;">{self.DISPLAY_NAME}</strong><br>
      <span style="font-size:12px;">
        Accepted: {accepted}{acc_delta_str}
        &nbsp;&nbsp;|&nbsp;&nbsp;
        Recasted: {recasted}{rec_delta_str}
        &nbsp;&nbsp;|&nbsp;&nbsp;
        Rate: {rate_str}
      </span>
      {pending_html}<br>
      <em style="font-size:11px;color:#555;">{data.driver_narrative}</em>
    </td>
  </tr>
</table>"""

    # ------------------------------------------------------------------
    # render_analyst_tabs  (CONTRACT-02)
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, pd.DataFrame]]:
        """
        CONTRACT-02: analyst-detail workbook tabs.

        Returns ``data.analyst_rows`` when data is valid; ``[]`` on error.
        Rule Detail tab carries rule-level counts (finding_count), not per-finding
        rows (Pitfall 6c). filter_summary via _summarize_filter (no inline parse).
        """
        if data.error or not data.analyst_rows:
            return []
        return data.analyst_rows

    # ------------------------------------------------------------------
    # render_rag_strip_entry  (CONTRACT-03, honors data.rag_strip)
    # ------------------------------------------------------------------

    def render_rag_strip_entry(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict:
        """
        CONTRACT-03: cover-page RAG strip cell.

        Returns the pre-built ``data.rag_strip`` dict when present;
        falls back to a gray "No Data" cell if empty or error.
        """
        if data.error or not data.rag_strip:
            return build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            )
        return data.rag_strip
