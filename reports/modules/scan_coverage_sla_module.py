"""
reports/modules/scan_coverage_sla_module.py — Scan Coverage SLA metric module.

Computes the percentage of managed assets that received a licensed Tenable scan
within the last 30 days, with a per-owner breakdown.

Module ID:    scan_coverage_sla
Display Name: Scan Coverage SLA

SLA thresholds (board-defined):
    Green:  scan_coverage_pct >= 95%
    Yellow: scan_coverage_pct >= 90%  and < 95%
    Red:    scan_coverage_pct <  90%

Data source:  assets_df.last_licensed_scan_date  (fetch_all_assets cache)

Owner dimension:
    Derived from the Tenable tag category "Owner" via
    board_report_utils.extract_owner().  Assets without an
    Owner tag are grouped under "Unassigned".
"""

from __future__ import annotations

import html
import logging
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.modules.base import BaseModule, ModuleConfig, ModuleData
from reports.modules.registry import register_module
from reports.modules.board_pdf_layout import two_column_metric_section
from reports.modules.board_report_utils import (
    compute_per_bu_breakdown,
    deduplicate_assets_by_name,
    extract_owner,
    populate_rag_strip,  # noqa: F401  # re-exported for plan 03-02 contract surface
    sla_status_from_thresholds,
    ON_TIME_WINDOW_DAYS,
)
from reports.modules.chart_utils import draw_gauge
from reports.modules.format_utils import safe_int, safe_pct
from reports.modules.rag_utils import (
    build_rag_strip_entry,
    rag_status_from_value,
    STATUS_COLOR,
    STATUS_LABEL,
    STATUS_ICON,
    NO_DATA_DRIVER,
    NO_DATA_HEADLINE,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

_GREEN_THRESHOLD  = 95.0   # >= green
_YELLOW_THRESHOLD = 90.0   # >= yellow; < green
_DIRECTION        = "higher_is_better"

# draw_gauge threshold list: each tuple = (upper_bound, colour)
# Zones: 0–90 red | 90–95 amber | 95–100 green
# WR-04 fix — gauge amber band aligned with rag_utils.STATUS_COLOR['yellow']
# (#f57c00) so the gauge dial, the inline status badge, and the cover-page
# RAG strip cell all show the SAME orange. Previously the gauge used
# #fbc02d (the matplotlib chart_exporter 'Medium' palette) while the
# status badge and RAG strip used #f57c00 from the shared rag_utils
# palette — so a yellow score showed as warm yellow on the dial but
# orange on the chip, looking unprofessional in the PDF.
_GAUGE_THRESHOLDS = [
    (_YELLOW_THRESHOLD, "#d32f2f"),
    (_GREEN_THRESHOLD,  "#f57c00"),
    (100.0,             "#388e3c"),
]

# Status display properties
_STATUS_COLOR: dict[str, str] = {
    "green":   "#388e3c",
    "yellow":  "#f57c00",
    "red":     "#d32f2f",
    "no_data": "#757575",
}

_STATUS_LABEL: dict[str, str] = {
    "green":   "On Target",
    "yellow":  "At Risk",
    "red":     "Off Target",
    "no_data": "No Data",
}

# Excel fill colours (no leading #, openpyxl RGB format)
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")  # light green
_FILL_YELLOW = PatternFill("solid", fgColor="FFF9C4")  # light amber
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")  # light red
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")  # dark navy


# ===========================================================================
# Module class
# ===========================================================================

@register_module
class ScanCoverageSLAModule(BaseModule):
    """
    Percentage of assets scanned within the last 30 days, with per-BU breakdown.

    Supported options
    -----------------
    None — this module accepts no configurable options.
    """

    MODULE_ID         = "scan_coverage_sla"
    DISPLAY_NAME      = "Scan Coverage SLA"
    DESCRIPTION       = (
        "Percentage of licensed assets scanned within the last 30 days, "
        "with per-business-unit breakdown. "
        "Assets with no last_licensed_scan_date are excluded from both "
        "numerator and denominator."
    )
    REQUIRED_DATA     = ["assets"]
    SUPPORTED_OUTPUTS = ["pdf", "excel", "email"]
    VERSION           = "1.0.0"

    # ------------------------------------------------------------------
    # compute()
    # ------------------------------------------------------------------

    def compute(
        self,
        vulns_df:    pd.DataFrame,
        assets_df:   pd.DataFrame,
        report_date: Any,
        config:      ModuleConfig,
        **kwargs:    Any,
    ) -> ModuleData:
        """
        Compute scan coverage percentage and per-BU breakdown.

        Parameters
        ----------
        vulns_df : pd.DataFrame
            Not used by this module; accepted for interface compatibility.
        assets_df : pd.DataFrame
            Full (unfiltered) asset DataFrame from fetch_all_assets().
            Required columns: hostname, last_seen, last_licensed_scan_date,
            asset_uuid, tags.
        report_date : datetime
            UTC-aware datetime for the report run.  Used to compute the
            30-day recency cutoff.
        config : ModuleConfig
            Module configuration (no options consumed).

        Returns
        -------
        ModuleData
            ``error`` is None on success; set to an error string on failure.
        """
        logger.debug(
            "%s compute() — assets_df rows: %d",
            self._log_prefix(), len(assets_df),
        )

        try:
            # ---- Phase 3 D-07/D-15 — assets-empty early return path ----
            # Hands an explicit no-data ModuleData back rather than threading
            # an empty DataFrame through dedup / BU enrichment / breakdown
            # (each of which has its own zero-row edge case in older callers).
            # The Phase 3 contract fields are populated here so the cover
            # strip shows a gray cell, the email panel shows the placeholder,
            # and the analyst workbook contributes nothing.
            if assets_df is None or assets_df.empty:
                logger.warning(
                    "%s assets_df is empty — returning no-data ModuleData.",
                    self._log_prefix(),
                )
                computed_at = (
                    report_date.isoformat()
                    if hasattr(report_date, "isoformat")
                    else str(report_date)
                )
                rag_strip_payload = build_rag_strip_entry(
                    display_name       = self.DISPLAY_NAME,
                    headline_value_str = safe_pct(None),   # → "—"
                    status             = "no_data",
                )
                return ModuleData(
                    module_id    = self.MODULE_ID,
                    display_name = self.DISPLAY_NAME,
                    metrics      = {
                        "scan_coverage_pct":   None,
                        "scanned_on_time":     0,
                        "not_scanned_on_time": 0,
                        "total_licensed":      0,
                        "unlicensed_excluded": 0,
                        "status":              "no_data",
                    },
                    table_data   = [],
                    chart_data   = {
                        "value":      None,
                        "thresholds": {
                            "green":  _GREEN_THRESHOLD,
                            "yellow": _YELLOW_THRESHOLD,
                        },
                        "direction":  _DIRECTION,
                        "top_5":      [],
                    },
                    summary_text = (
                        "Scan coverage could not be computed — "
                        "no assets were provided."
                    ),
                    metadata     = {
                        "filter":              "No assets in scope",
                        "window":              f"Last {ON_TIME_WINDOW_DAYS} days from report_date",
                        "sla_source":          "Board-defined thresholds (Green ≥95%, Amber ≥90%, Red <90%)",
                        "computed_at":         computed_at,
                        "unlicensed_excluded": 0,
                        # Phase 3 D-15 — empty-data: no gauge for empty modules
                        "email_gauge_b64":     "",
                    },
                    # ── Phase 3 contract fields ──
                    driver_narrative = NO_DATA_DRIVER,
                    analyst_rows     = [],
                    rag_strip        = rag_strip_payload,
                    error            = None,
                )

            # ---- Step 1: deduplicate; separate licensed vs unlicensed ----
            # Deduplication runs on ALL assets first so that a hostname whose
            # most-recent record is unlicensed does not retain an older licensed
            # duplicate.  After deduplication the licensed split is clean.
            _lsd      = "last_licensed_scan_date"
            all_dedup = deduplicate_assets_by_name(assets_df).copy()
            if _lsd in all_dedup.columns:
                all_dedup.loc[:, _lsd] = pd.to_datetime(
                    all_dedup[_lsd], utc=True, errors="coerce"
                )
            else:
                # [Rule 1 - Bug] Pre-existing crash on a zero-row DataFrame:
                # `df.loc[:, col] = pd.NaT` raises "cannot set a frame with no
                # defined index and a scalar" when df is empty. Building the
                # column as an explicit Series sidesteps the empty-frame path
                # and produces the same result on populated frames.
                all_dedup[_lsd] = pd.Series(
                    pd.NaT, index=all_dedup.index, dtype="datetime64[ns, UTC]"
                )

            licensed_mask    = all_dedup[_lsd].notna()
            licensed         = all_dedup[licensed_mask].copy().reset_index(drop=True)
            unlicensed_count = int((~licensed_mask).sum())

            logger.debug(
                "%s dedup total=%d, licensed=%d, unlicensed=%d",
                self._log_prefix(), len(all_dedup), len(licensed), unlicensed_count,
            )

            # ---- Step 2: split licensed into on-time vs not-on-time ----
            if hasattr(report_date, "tzinfo") and report_date.tzinfo is not None:
                rd_ts = pd.Timestamp(report_date).tz_convert("UTC")
            else:
                rd_ts = pd.Timestamp(report_date, tz="UTC")
            cutoff = rd_ts - pd.Timedelta(days=ON_TIME_WINDOW_DAYS)

            if not licensed.empty:
                on_time_flag = licensed[_lsd] >= cutoff
                on_time      = licensed[on_time_flag].copy().reset_index(drop=True)
                not_on_time  = licensed[~on_time_flag].copy().reset_index(drop=True)
            else:
                on_time     = licensed.copy()
                not_on_time = licensed.copy()

            scanned_on_time     = len(on_time)
            not_scanned_on_time = len(not_on_time)
            total_licensed      = scanned_on_time + not_scanned_on_time

            # ---- Step 3: overall percentage + status ----
            if total_licensed == 0:
                scan_coverage_pct = None
                status = "no_data"
                logger.warning(
                    "%s no licensed assets found — returning no_data.",
                    self._log_prefix(),
                )
            else:
                scan_coverage_pct = round(
                    scanned_on_time / total_licensed * 100, 1
                )
                status = sla_status_from_thresholds(
                    scan_coverage_pct,
                    green_threshold  = _GREEN_THRESHOLD,
                    yellow_threshold = _YELLOW_THRESHOLD,
                    direction        = _DIRECTION,
                )

            # ---- Step 4: per-BU breakdown (licensed assets only) ----
            # Unlicensed assets are excluded from the denominator entirely.
            # [Rule 1 - Bug] extract_owner and compute_per_bu_breakdown
            # both call `df.loc[:, col] = scalar` patterns that crash on a
            # zero-row DataFrame. Short-circuit the owner computation when there
            # are no licensed assets in scope.
            if licensed.empty:
                enriched     = licensed.copy()
                bu_breakdown = pd.DataFrame(
                    columns=[
                        "owner", "numerator", "denominator",
                        "percentage",   "affected",
                    ],
                )
                table_data = []
            else:
                enriched = extract_owner(licensed)

                on_time_uuids   = set(on_time["asset_uuid"].dropna())
                on_time_mask_bu = enriched["asset_uuid"].isin(on_time_uuids)
                denom_mask      = pd.Series(True, index=enriched.index)

                bu_breakdown = compute_per_bu_breakdown(
                    enriched, on_time_mask_bu, denom_mask,
                    higher_is_better=True,
                )
                table_data = bu_breakdown.to_dict("records")

            # ---- Step 5: narrative summary ----
            if scan_coverage_pct is None:
                summary_text = (
                    "Scan coverage could not be computed — "
                    "no licensed assets were found in the asset inventory."
                )
            else:
                # WR-07 fix — use safe_pct() instead of an inline f-string
                # format spec on a possibly-None value. The early-return
                # guard above currently makes scan_coverage_pct non-None
                # here, but safe_pct() makes the rule mechanical: future
                # refactors that break the guard cannot crash this line.
                summary_text = (
                    f"Scan coverage is {safe_pct(scan_coverage_pct)} — "
                    f"{scanned_on_time:,} of {total_licensed:,} licensed assets "
                    f"were scanned within the last {ON_TIME_WINDOW_DAYS} days "
                    f"({unlicensed_count:,} unlicensed assets excluded). "
                    f"Status: {_STATUS_LABEL.get(status, status)}."
                )

            computed_at = (
                report_date.isoformat()
                if hasattr(report_date, "isoformat")
                else str(report_date)
            )

            # ---- Step 6: Phase 3 contract fields ----
            # Phase 3 D-10/D-11/D-13 — analyst rows for Scan Coverage SLA.
            # Source: not_on_time slice (licensed assets that did NOT meet the
            # ON_TIME_WINDOW_DAYS scan-recency window).
            if not not_on_time.empty:
                analyst_df = not_on_time.copy()
                # owner is already extracted into owner by extract_owner()
                # earlier in compute() (step 4); fall back defensively for the
                # not_on_time slice which is sourced from `licensed` pre-enrichment.
                if "owner" not in analyst_df.columns:
                    analyst_df = extract_owner(analyst_df)
                # days_since_licensed_scan
                analyst_df = analyst_df.assign(
                    days_since_licensed_scan = (
                        (rd_ts - analyst_df[_lsd])
                        .dt.days
                        .astype("Int64")
                    ),
                )
                # D-13 — apply asset-level dedup
                analyst_df = deduplicate_assets_by_name(analyst_df)
                # Project + reorder to the 6 contracted columns (D-10)
                analyst_df = analyst_df.reindex(columns=[
                    "hostname",
                    "ipv4",
                    "fqdn",
                    "last_licensed_scan_date",
                    "days_since_licensed_scan",
                    "owner",
                ])
                # D-11 — sort by days_since_licensed_scan desc; NaN last
                analyst_df = analyst_df.sort_values(
                    "days_since_licensed_scan",
                    ascending=False,
                    na_position="last",
                ).reset_index(drop=True)
                # T-03-02-02 — CSV-formula injection guard. Prepend a single
                # quote to any cell whose first char would trigger Excel
                # formula evaluation (=, +, -, @). Hostname/ipv4/fqdn are
                # external-source strings; owner is Tenable-normalised
                # but we apply the guard uniformly for defence-in-depth.
                for _col in ("hostname", "ipv4", "fqdn", "owner"):
                    if _col in analyst_df.columns:
                        analyst_df.loc[:, _col] = analyst_df[_col].astype("string").map(
                            lambda s: ("'" + s)
                            if isinstance(s, str) and len(s) > 0 and s[:1] in ("=", "+", "-", "@")
                            else s
                        )
                analyst_rows_payload: list[tuple[str, pd.DataFrame]] = [
                    ("Scan Coverage Detail", analyst_df),
                ]
            else:
                analyst_rows_payload = []

            # Phase 3 D-06 — Scan Coverage SLA driver narrative.
            # Template (locked in plan 03-02):
            #   "Best BU: {good_bu_name} at {good_bu_pct}; worst BU: {worst_bu_name}
            #    at {worst_bu_pct} ({overdue_count} of {total_count} licensed assets
            #    overdue)."
            # Sources:
            #   - good_bu_*  : bu_breakdown row with the highest percentage
            #                  (ties broken by alphabetical owner name)
            #   - worst_bu_* : bu_breakdown row with the lowest percentage
            #                  (ties broken by alphabetical owner name)
            #   - overdue_count = not_scanned_on_time
            #   - total_count   = total_licensed
            # W4 — `bu_breakdown` is the DataFrame produced earlier in compute()
            # at step 4 by compute_per_bu_breakdown(...). Its columns are
            # owner / numerator / denominator / percentage / affected.
            if total_licensed > 0 and not bu_breakdown.empty:
                sorted_bu_asc  = bu_breakdown.sort_values(
                    ["percentage", "owner"], ascending=[True,  True]
                )
                sorted_bu_desc = bu_breakdown.sort_values(
                    ["percentage", "owner"], ascending=[False, True]
                )
                worst_row = sorted_bu_asc.iloc[0]
                good_row  = sorted_bu_desc.iloc[0]
                driver = (
                    f"Best Owner: {good_row['owner']} at {safe_pct(good_row['percentage'])}; "
                    f"worst Owner: {worst_row['owner']} at {safe_pct(worst_row['percentage'])} "
                    f"({safe_int(not_scanned_on_time)} of {safe_int(total_licensed)} "
                    f"licensed assets overdue)."
                )
            else:
                # D-07 empty-data fallback
                driver = NO_DATA_DRIVER

            # Phase 3 D-04 — email gauge base64 for CID inline image.
            # Same draw_gauge call as the PDF section uses; we generate it ONCE
            # in compute() and stash it on data.metadata so render_email_panel
            # can reference cid only.
            if total_licensed > 0 and scan_coverage_pct is not None:
                try:
                    email_gauge_b64 = draw_gauge(
                        value      = scan_coverage_pct,
                        thresholds = _GAUGE_THRESHOLDS,
                        title      = self.DISPLAY_NAME,
                        unit       = "%",
                        figsize    = (2.4, 1.6),
                    )
                except Exception as _gauge_exc:  # noqa: BLE001
                    logger.warning(
                        "%s compute() draw_gauge for email failed: %s",
                        self._log_prefix(), _gauge_exc,
                    )
                    email_gauge_b64 = ""
            else:
                # D-15 empty-data — no gauge for empty modules
                email_gauge_b64 = ""

            # Phase 3 D-05/D-08/D-09 — RAG strip payload built via
            # pure-construction (option 2 from PATTERNS.md). Plan 03-02 locks
            # this shape so plans 03-03..05 can copy it directly.
            _rag_status = rag_status_from_value(
                scan_coverage_pct,
                green_threshold  = _GREEN_THRESHOLD,
                yellow_threshold = _YELLOW_THRESHOLD,
                direction        = _DIRECTION,
            )
            rag_strip_payload = build_rag_strip_entry(
                display_name       = self.DISPLAY_NAME,
                headline_value_str = safe_pct(scan_coverage_pct),
                status             = _rag_status,
            )

            return ModuleData(
                module_id    = self.MODULE_ID,
                display_name = self.DISPLAY_NAME,
                metrics      = {
                    "scan_coverage_pct":   scan_coverage_pct,
                    "scanned_on_time":     scanned_on_time,
                    "not_scanned_on_time": not_scanned_on_time,
                    "total_licensed":      total_licensed,
                    "unlicensed_excluded": unlicensed_count,
                    "status":              status,
                },
                table_data   = table_data,
                chart_data   = {
                    "value":      scan_coverage_pct,
                    "thresholds": {
                        "green":  _GREEN_THRESHOLD,
                        "yellow": _YELLOW_THRESHOLD,
                    },
                    "direction":  _DIRECTION,
                    # Pre-sliced top 5 for PDF render (worst performers first)
                    "top_5":      bu_breakdown.head(5).to_dict("records"),
                },
                summary_text = summary_text,
                metadata     = {
                    "filter":               (
                        "Deduplicated by hostname (most-recent last_seen retained); "
                        "unlicensed assets (null last_licensed_scan_date) excluded "
                        "from both numerator and denominator"
                    ),
                    "window":               f"Last {ON_TIME_WINDOW_DAYS} days from report_date",
                    "sla_source":           "Board-defined thresholds (Green ≥95%, Amber ≥90%, Red <90%)",
                    "computed_at":          computed_at,
                    "unlicensed_excluded":  unlicensed_count,
                    # Phase 3 D-04 — picked up by composer.collect_email_inline_images
                    "email_gauge_b64":      email_gauge_b64,
                },
                # ── Phase 3 contract fields ──
                driver_narrative = driver,
                analyst_rows     = analyst_rows_payload,
                rag_strip        = rag_strip_payload,
                error            = None,
            )

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s compute() failed: %s", self._log_prefix(), exc,
                exc_info=True,
            )
            return self._empty_result(str(exc), config)

    # ------------------------------------------------------------------
    # render_pdf_section()
    # ------------------------------------------------------------------

    def render_pdf_section(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Render a full-page PDF section for WeasyPrint.

        Layout (two-column row via ``two_column_metric_section`` to avoid the
        page bleed a stacked explanation caused):
        - Section heading (full width)
        - Left column: gauge with 95 / 90 colour zones, status badge showing
          coverage % and on-target label, and two bold support numbers
          (Scanned On Time | Not On Time)
        - Right column: explanatory paragraph (left-aligned)
        - Below, full width: Top-5 worst-performing BUs table (row-coloured by
          threshold band)

        Returns an error callout div if ``data.error`` is set.
        """
        if data.error:
            return (
                f'<div class="error-box">'
                f"<strong>{self.DISPLAY_NAME}</strong>: {data.error}"
                f"</div>"
            )

        m                   = data.metrics
        scan_coverage_pct   = m.get("scan_coverage_pct")
        scanned_on_time     = m.get("scanned_on_time", 0)
        not_scanned_on_time = m.get("not_scanned_on_time", 0)
        status              = m.get("status", "no_data")

        # ---- Gauge ----
        gauge_value = scan_coverage_pct if scan_coverage_pct is not None else 0.0
        try:
            gauge_b64  = draw_gauge(
                value      = gauge_value,
                min_val    = 0,
                max_val    = 100,
                thresholds = _GAUGE_THRESHOLDS,
                title      = "Scan Coverage %",
                unit       = "%",
                figsize    = (5, 3),
            )
            gauge_html = (
                f'<div style="text-align:center; margin-bottom:4mm;">'
                f'<img src="data:image/png;base64,{gauge_b64}" '
                f'style="width:46%; max-width:320px;" alt="Scan Coverage gauge">'
                f'</div>'
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("%s PDF gauge render failed: %s", self._log_prefix(), exc)
            pct_display = f"{gauge_value:.1f}%"  # safe: gauge_value guaranteed non-None by upstream coalescing (None coalesced to 0.0)
            gauge_html = (
                f'<p style="text-align:center; font-size:20pt; font-weight:bold; '
                f'color:{_STATUS_COLOR.get(status, "#333")};">{pct_display}</p>'
            )

        # ---- Status badge ----
        status_color = _STATUS_COLOR.get(status, "#757575")
        status_label = _STATUS_LABEL.get(status, status)
        pct_display  = (
            f"{scan_coverage_pct:.1f}%" if scan_coverage_pct is not None else "N/A"
        )
        status_html = (
            f'<p style="text-align:center; font-size:10pt; font-weight:bold; '
            f'color:{status_color}; margin:0 0 5mm 0;">'
            f'Coverage: {pct_display} &nbsp;&middot;&nbsp; {status_label}'
            f'</p>'
        )

        # ---- Supporting numbers ----
        not_scanned_color = "#d32f2f" if not_scanned_on_time > 0 else "#388e3c"
        support_html = f"""
<table style="width:52%; margin:0 auto 6mm auto; border-collapse:collapse;">
  <tr>
    <td style="text-align:center; padding:2mm 8mm;
               border-right:0.5pt solid #ddd; vertical-align:middle;">
      <span style="font-size:14pt; font-weight:bold; color:#1F3864;">{scanned_on_time:,}</span>
      <br><span style="font-size:7.5pt; color:#555;">Scanned On Time</span>
    </td>
    <td style="text-align:center; padding:2mm 8mm; vertical-align:middle;">
      <span style="font-size:14pt; font-weight:bold;
             color:{not_scanned_color};">{not_scanned_on_time:,}</span>
      <br><span style="font-size:7.5pt; color:#555;">Not On Time</span>
    </td>
  </tr>
</table>"""

        # ---- Top 5 BU table ----
        top5 = data.chart_data.get("top_5", [])
        if top5:
            rows_html = ""
            for row in top5:
                bu_pct  = float(row.get("percentage", 0.0))
                bu_name = str(row.get("owner", ""))
                bu_num  = int(row.get("numerator",    0))
                bu_den  = int(row.get("denominator",  0))
                row_bg  = _row_bg(bu_pct, _GREEN_THRESHOLD, _YELLOW_THRESHOLD,
                                  direction="higher_is_better")
                rows_html += (
                    f'<tr style="background:{row_bg};">'
                    f'<td style="padding:1.5mm 3mm;">{bu_name}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_num:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm;">{bu_den:,}</td>'
                    f'<td style="text-align:right; padding:1.5mm 3mm; '
                    f'font-weight:bold;">{bu_pct:.1f}%</td>'  # safe: bu_pct non-None per compute_per_bu_breakdown contract
                    f'</tr>'
                )
            bu_table_html = f"""
<h3 class="subsection-heading">Top 5 Worst-Performing Owners</h3>
<table class="data-table">
  <thead>
    <tr>
      <th>Owner</th>
      <th style="text-align:right;">Scanned On Time</th>
      <th style="text-align:right;">Licensed Assets</th>
      <th style="text-align:right;">Coverage %</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>"""
        else:
            bu_table_html = (
                '<p class="explanatory-text" style="color:#888; font-style:italic;">'
                'No owner breakdown available — '
                'assets may lack Owner tags.'
                '</p>'
            )

        # ---- Explanatory paragraph ----
        explain_html = f"""
<p class="explanatory-text">
  <strong>What this measures:</strong> The percentage of <em>licensed</em> assets
  that received a licensed Tenable scan within the past {ON_TIME_WINDOW_DAYS} days.
  Assets without a <em>last_licensed_scan_date</em> (unlicensed) are excluded from
  both the numerator and denominator.  Assets that go unscanned have unknown
  vulnerability posture &mdash; new or worsening findings cannot be detected until a
  scan completes.  Denominator is the deduplicated licensed asset inventory (one row
  per hostname; most-recent <em>last_seen</em> retained).  Board target is &ge;95%
  (green).  &ge;90% is at-risk (amber).  Below 90% is off-target (red).
  Owner breakdown uses the Tenable &ldquo;Owner&rdquo; tag category.
  Assets without an Owner tag are grouped under &ldquo;Unassigned&rdquo;.
</p>"""

        return two_column_metric_section(
            heading_html=f'<h2 class="section-heading">{self.DISPLAY_NAME}</h2>',
            left_html=f"{gauge_html}{status_html}{support_html}",
            explanation_html=explain_html,
            full_width_html=bu_table_html,
        )

    # ------------------------------------------------------------------
    # render_excel_tabs()
    # ------------------------------------------------------------------

    def render_excel_tabs(
        self,
        data:     ModuleData,
        workbook: Any,
        config:   ModuleConfig,
    ) -> list[str]:
        """
        Write the "Scan Coverage Summary" tab into ``workbook``.

        Tab contents:
        - Overall KPI block (coverage %, status, supporting numbers)
        - Full business-unit breakdown table with colour-coded Coverage % column

        Returns
        -------
        list[str]
            ``["Scan Coverage Summary"]`` on success, ``[]`` on error.
        """
        tab_name = "Scan Coverage Summary"
        try:
            ws = workbook.create_sheet(tab_name)

            if data.error:
                ws["A1"] = "Error"
                ws["B1"] = data.error
                return [tab_name]

            # D-16 — Phase 3 zero-row standardisation. If both metrics and
            # table_data are empty AND there is no error, emit a single
            # placeholder cell at A1 instead of a fully-empty sheet.
            # This is a behavior change from pre-Phase-3 — surfaced in the
            # 03-02 SUMMARY for live-Tenable Excel diff visibility.
            empty_metrics = not (
                data.metrics
                and any(v is not None for v in data.metrics.values())
            )
            empty_tables  = not data.table_data
            if empty_metrics and empty_tables:
                ws["A1"]      = "No data in scope"
                ws["A1"].font = Font(bold=True, color="666666")
                return [tab_name]

            m                   = data.metrics
            scan_coverage_pct   = m.get("scan_coverage_pct")
            scanned_on_time     = m.get("scanned_on_time", 0)
            not_scanned_on_time = m.get("not_scanned_on_time", 0)
            total_licensed      = m.get("total_licensed", 0)
            unlicensed_excluded = m.get("unlicensed_excluded", 0)
            status              = m.get("status", "no_data")

            # ---- Overall KPI block (rows 1–9) ----
            _xl_title(ws, "A1", "Scan Coverage SLA — Overall Summary")

            status_color_hex = _STATUS_COLOR.get(status, "#757575").lstrip("#")
            pct_str = (
                f"{scan_coverage_pct:.1f}%" if scan_coverage_pct is not None else "N/A"
            )

            _xl_kv(ws, 3, "Overall Coverage:", pct_str,
                   value_font=Font(bold=True, size=12, color=status_color_hex))
            _xl_kv(ws, 4, "Status:",
                   _STATUS_LABEL.get(status, status),
                   value_font=Font(bold=True, color=status_color_hex))
            _xl_kv(ws, 5, "Scanned On Time:",       f"{scanned_on_time:,}")
            _xl_kv(ws, 6, "Not On Time:",            f"{not_scanned_on_time:,}")
            _xl_kv(ws, 7, "Total Licensed Assets:",  f"{total_licensed:,}")
            _xl_kv(ws, 8, "Unlicensed Excluded:",    f"{unlicensed_excluded:,}")
            _xl_kv(ws, 9, "Window:",                 "Last 30 days (last_licensed_scan_date)")
            _xl_kv(ws, 10, "SLA Thresholds:",        "Green ≥95%  |  Amber ≥90%  |  Red <90%")

            # ---- Owner breakdown table (starts at row 12) ----
            header_row = 12
            headers    = ["Owner", "Scanned On Time", "Licensed Assets", "Coverage %"]

            for col_idx, header in enumerate(headers, start=1):
                cell             = ws.cell(row=header_row, column=col_idx, value=header)
                cell.font        = Font(bold=True, color="FFFFFF")
                cell.fill        = _FILL_HEADER
                cell.alignment   = Alignment(horizontal="center")

            for row_offset, row in enumerate(data.table_data or [], start=1):
                data_row = header_row + row_offset
                bu_pct   = float(row.get("percentage",  0.0))
                bu_fill  = _xl_fill(bu_pct, _GREEN_THRESHOLD, _YELLOW_THRESHOLD,
                                    direction="higher_is_better")

                ws.cell(row=data_row, column=1,
                        value=str(row.get("owner", ""))).alignment = (
                    Alignment(horizontal="left")
                )
                ws.cell(row=data_row, column=2,
                        value=int(row.get("numerator",   0))).alignment = (
                    Alignment(horizontal="right")
                )
                ws.cell(row=data_row, column=3,
                        value=int(row.get("denominator", 0))).alignment = (
                    Alignment(horizontal="right")
                )
                pct_cell           = ws.cell(row=data_row, column=4,
                                             value=f"{bu_pct:.1f}%")  # safe: bu_pct non-None per compute_per_bu_breakdown contract
                pct_cell.fill      = bu_fill
                pct_cell.font      = Font(bold=True)
                pct_cell.alignment = Alignment(horizontal="center")

            # ---- Column widths ----
            ws.column_dimensions[get_column_letter(1)].width = 32
            ws.column_dimensions[get_column_letter(2)].width = 18
            ws.column_dimensions[get_column_letter(3)].width = 14
            ws.column_dimensions[get_column_letter(4)].width = 14

            return [tab_name]

        except Exception as exc:  # noqa: BLE001
            logger.error(
                "%s render_excel_tabs() failed: %s",
                self._log_prefix(), exc, exc_info=True,
            )
            return []

    # ------------------------------------------------------------------
    # render_email_panel()  — Phase 3 D-02 horizontal split layout
    # ------------------------------------------------------------------

    def render_email_panel(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> str:
        """
        Build the per-module email panel — horizontal split layout (D-02).

        Layout: 620px-wide table, 150px gauge cell on the left + 430px text
        cell on the right. Inline CSS only. Outlook-safe ``<table>`` shell
        with explicit ``width=""`` attributes per project email conventions.

        Empty-data behavior (D-15): when ``data.error`` or no
        ``email_gauge_b64`` is available, returns the same 620px shell with
        a gray "No data" placeholder block where the gauge would be.

        Returns
        -------
        str
            Inline-CSS HTML fragment. Returns ``""`` only on a render
            exception (caught by the composer's per-module exception
            isolation pattern at composer.py:1232-1251).
        """
        try:
            # Empty-data placeholder per D-15
            b64 = (data.metadata or {}).get("email_gauge_b64", "")
            if data.error or not isinstance(b64, str) or not b64.strip():
                return self._render_empty_email_panel()

            pct       = data.metrics.get("scan_coverage_pct") if data.metrics else None
            headline  = safe_pct(pct)
            status    = (data.metrics or {}).get("status", "no_data")
            rag_color = STATUS_COLOR.get(status, STATUS_COLOR["no_data"])
            rag_label = STATUS_LABEL.get(status, STATUS_LABEL["no_data"])
            icon      = STATUS_ICON.get(status,  STATUS_ICON["no_data"])

            cid       = f"{self.MODULE_ID}_gauge"
            # T-03-02-01 — html-escape every module-supplied string
            label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
            driver_esc = html.escape(str(data.driver_narrative or ""), quote=True)

            return (
                '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
                'style="width:620px; max-width:620px; margin:8px 0; '
                'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
                '<tr>'
                '  <td width="150" style="padding:12px; vertical-align:middle; '
                '      text-align:center;">'
                f'    <img src="cid:{cid}" alt="" width="120" height="120" '
                '         style="display:block; margin:0 auto;" />'
                '  </td>'
                '  <td width="430" style="padding:12px; vertical-align:middle;">'
                f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
                f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{headline}</div>'
                f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
                f'{icon} {html.escape(rag_label)}</div>'
                f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
                f'{driver_esc}</div>'
                '  </td>'
                '</tr>'
                '</table>'
            )
        except Exception as exc:    # noqa: BLE001
            logger.error(
                "%s render_email_panel raised: %s",
                self._log_prefix(),
                exc,
            )
            return ""

    def _render_empty_email_panel(self) -> str:
        """Return the D-15 gray 'No Data' placeholder panel."""
        label_esc  = html.escape(str(self.DISPLAY_NAME), quote=True)
        driver_esc = html.escape(NO_DATA_DRIVER, quote=True)
        rag_color  = STATUS_COLOR["no_data"]
        rag_label  = STATUS_LABEL["no_data"]
        icon       = STATUS_ICON["no_data"]
        return (
            '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
            'style="width:620px; max-width:620px; margin:8px 0; '
            'border:1px solid #e0e0e0; border-collapse:separate; background:#ffffff;">'
            '<tr>'
            '  <td width="150" style="padding:12px; vertical-align:middle; '
            '      text-align:center; background:#f5f5f5; color:#999;">'
            '    <div style="font-size:10pt;">No data</div>'
            '  </td>'
            '  <td width="430" style="padding:12px; vertical-align:middle;">'
            f'    <div style="font-size:11pt; color:#666;">{label_esc}</div>'
            f'    <div style="font-size:24pt; font-weight:bold; color:#1a1a1a;">{NO_DATA_HEADLINE}</div>'
            f'    <div style="font-size:10pt; color:{rag_color}; font-weight:bold;">'
            f'{icon} {html.escape(rag_label)}</div>'
            f'    <div style="font-size:10pt; color:#444; margin-top:6px;">'
            f'{driver_esc}</div>'
            '  </td>'
            '</tr>'
            '</table>'
        )

    # ------------------------------------------------------------------
    # render_analyst_tabs()  — Phase 3 D-14 single-tab list
    # ------------------------------------------------------------------

    def render_analyst_tabs(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> list[tuple[str, "pd.DataFrame"]]:
        """
        Return the module's pre-built analyst rows (D-14 single-tab list).

        Source: ``data.analyst_rows`` populated inside ``compute()``.
        Empty-data: returns ``[]`` (no tab written for this module).
        """
        try:
            if data.error or not data.analyst_rows:
                return []
            return list(data.analyst_rows)
        except Exception as exc:   # noqa: BLE001
            logger.error("[%s] render_analyst_tabs raised: %s", self.MODULE_ID, exc)
            return []

    # ------------------------------------------------------------------
    # render_email_kpis()
    # ------------------------------------------------------------------

    def render_email_kpis(
        self,
        data:   ModuleData,
        config: ModuleConfig,
    ) -> dict[str, str]:
        """
        Return three KPI tiles for the HTML email body.

        Returns
        -------
        dict[str, str]
            Keys: "Scan Coverage", "Scanned On Time", "Not On Time".
            Returns empty dict if ``data.error`` is set.
        """
        if "email" not in self.SUPPORTED_OUTPUTS or data.error:
            return {}
        m   = data.metrics
        pct = m.get("scan_coverage_pct")
        return {
            "Scan Coverage":   f"{pct:.1f}%" if pct is not None else "N/A",
            "Scanned On Time": f"{m.get('scanned_on_time', 0):,}",
            "Not On Time":     f"{m.get('not_scanned_on_time', 0):,}",
        }

    # ------------------------------------------------------------------
    # get_audit_info()
    # ------------------------------------------------------------------

    def get_audit_info(self) -> dict:
        """Return calculation documentation for audit and runbook records."""
        return {
            **super().get_audit_info(),
            "calculations": {
                "scan_coverage_pct": (
                    "scanned_on_time / total_licensed × 100, rounded to 1 decimal. "
                    f"Numerator: licensed assets where last_licensed_scan_date "
                    f">= report_date − {ON_TIME_WINDOW_DAYS} days. "
                    "Denominator: licensed deduplicated assets "
                    "(last_licensed_scan_date IS NOT NULL after hostname dedup — "
                    "most-recent last_seen retained; unlicensed assets excluded entirely)."
                ),
                "status": (
                    "sla_status_from_thresholds(scan_coverage_pct, "
                    f"green_threshold={_GREEN_THRESHOLD}, "
                    f"yellow_threshold={_YELLOW_THRESHOLD}, "
                    "direction='higher_is_better'). "
                    "None → 'no_data'."
                ),
                "owner_breakdown": (
                    "compute_per_bu_breakdown(higher_is_better=True) applied to licensed "
                    "deduplicated assets enriched with owner from Owner tag. "
                    "Per-owner numerator = count of on-time assets for that owner; "
                    "denominator = licensed asset count for that owner. "
                    "Unlicensed assets are excluded from all owner buckets. "
                    "affected = denominator − numerator (assets NOT scanned on time). "
                    "Primary sort: affected DESC (most un-scanned assets first). "
                    "Secondary sort: percentage ASC (worst coverage % among ties). "
                    "Assets with no Owner tag → 'Unassigned' bucket."
                ),
                "deduplication": (
                    "deduplicate_assets_by_name() removes duplicate hostnames, "
                    "retaining the row with the most-recent last_seen. "
                    "Rows with empty/blank hostname are kept as-is."
                ),
            },
        }


# ===========================================================================
# Module-private rendering helpers
# ===========================================================================

def _row_bg(
    pct: float,
    green_threshold: float,
    yellow_threshold: float,
    direction: str = "higher_is_better",
) -> str:
    """Return a light HTML background-color for a BU table row."""
    if direction == "higher_is_better":
        if pct >= green_threshold:
            return "#E8F5E9"   # light green
        if pct >= yellow_threshold:
            return "#FFF8E1"   # light amber
        return "#FFEBEE"       # light red
    else:  # lower_is_better
        if pct <= green_threshold:
            return "#E8F5E9"
        if pct <= yellow_threshold:
            return "#FFF8E1"
        return "#FFEBEE"


def _xl_fill(
    pct: float,
    green_threshold: float,
    yellow_threshold: float,
    direction: str = "higher_is_better",
) -> PatternFill:
    """Return an openpyxl PatternFill for a coverage-% cell."""
    if direction == "higher_is_better":
        if pct >= green_threshold:
            return _FILL_GREEN
        if pct >= yellow_threshold:
            return _FILL_YELLOW
        return _FILL_RED
    else:  # lower_is_better
        if pct <= green_threshold:
            return _FILL_GREEN
        if pct <= yellow_threshold:
            return _FILL_YELLOW
        return _FILL_RED


def _xl_title(ws, cell_ref: str, value: str) -> None:
    """Write a bold, size-12 title cell."""
    ws[cell_ref]       = value
    ws[cell_ref].font  = Font(bold=True, size=12)


def _xl_kv(
    ws,
    row:         int,
    label:       str,
    value:       str,
    value_font:  Font | None = None,
) -> None:
    """Write a label in column A and its value in column B for the given row."""
    label_cell       = ws.cell(row=row, column=1, value=label)
    label_cell.font  = Font(bold=True)

    value_cell       = ws.cell(row=row, column=2, value=value)
    if value_font is not None:
        value_cell.font = value_font
