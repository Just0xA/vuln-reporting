"""
reports/ops_remediation.py — Operations Remediation Report.

Audience: IT Operations and remediation teams responsible for patching
specific asset groups. Designed to provide a clear, prioritized, and
actionable list of vulnerabilities to remediate, scoped by Tenable tag.

This report is NOT a replacement for sla_remediation.py — it has a
different structure and audience focus:
  - Vulnerabilities grouped by plugin (not per-row) for large-group usability
  - Four-state SLA status: Overdue / Urgent / Warning / On Track
  - Unscanned asset identification and exclusion from vuln counts
  - No per-asset breakdown in output — plugin + asset count is sufficient

Outputs
-------
- Excel: ops_remediation.xlsx  (6 tabs: Summary, Critical & High,
         Medium & Low, Overdue — All Severities, Unscanned Assets, Metadata)
- PDF:   ops_remediation.pdf   (cover, exec summary, overdue, urgent,
         unscanned, notes)
- Charts: none — returns empty list for run_all.py interface compatibility

CLI
---
python reports/ops_remediation.py
python reports/ops_remediation.py \\
    --tag-category "Operations" --tag-value "Server Operations"
python reports/ops_remediation.py --output-dir output/test/ --no-email
python reports/ops_remediation.py --cache-dir data/cache/2026-03-26/
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Allow running as a top-level script from any working directory
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import (
    CACHE_DIR,
    LOG_DIR,
    LOG_LEVEL,
    OUTPUT_DIR,
    SEVERITY_ORDER,
    SLA_DAYS,
    vpr_to_severity,
)
from data.fetchers import (
    enrich_vulns_with_assets,
    fetch_all_assets,
    fetch_all_vulnerabilities,
    fetch_recast_rules,
    filter_by_tag,
)
from utils.formatters import report_timestamp, safe_filename
from utils.sla_calculator import apply_sla_to_df

# ---------------------------------------------------------------------------
# Module constants
# ---------------------------------------------------------------------------
REPORT_NAME = "Operations Remediation Report"
REPORT_SLUG = "ops_remediation"

# Four-state SLA labels — used as cell text and in PDF/email body
OPS_SLA_OVERDUE  = "Overdue"
OPS_SLA_URGENT   = "Urgent - <=25% SLA Remaining"
OPS_SLA_WARNING  = "Warning - <=50% SLA Remaining"
OPS_SLA_ON_TRACK = "On Track"

# Ordered list for consistent display (worst → best)
OPS_SLA_STATE_ORDER: list[str] = [
    OPS_SLA_OVERDUE,
    OPS_SLA_URGENT,
    OPS_SLA_WARNING,
    OPS_SLA_ON_TRACK,
]

# Excel conditional formatting colors per SLA state (no leading #)
OPS_SLA_FILL: dict[str, str] = {
    OPS_SLA_OVERDUE:  "FF0000",   # red
    OPS_SLA_URGENT:   "FF6600",   # orange
    OPS_SLA_WARNING:  "FFC000",   # amber
    OPS_SLA_ON_TRACK: "E2EFDA",   # light green
}
OPS_SLA_FONT_COLOR: dict[str, str] = {
    OPS_SLA_OVERDUE:  "FFFFFF",   # white
    OPS_SLA_URGENT:   "FFFFFF",   # white
    OPS_SLA_WARNING:  "000000",   # black
    OPS_SLA_ON_TRACK: "000000",   # black
}
OPS_SLA_FONT_BOLD: dict[str, bool] = {
    OPS_SLA_OVERDUE:  True,
    OPS_SLA_URGENT:   True,
    OPS_SLA_WARNING:  False,
    OPS_SLA_ON_TRACK: False,
}

# Severity rank map for sort ordering (Critical = 0, sorts first)
SEVERITY_RANK: dict[str, int] = {s: i for i, s in enumerate(SEVERITY_ORDER)}

# Exploit code maturity rank (lower number = higher risk, sorts first).
# Values match the uppercase-normalized strings stored by fetch_all_vulnerabilities.
MATURITY_RANK: dict[str, int] = {
    "HIGH":       0,
    "FUNCTIONAL": 1,
    "POC":        2,
    "UNPROVEN":   3,
    "":           4,
}
_MATURITY_DISPLAY: dict[str, str] = {
    "HIGH":       "Yes (High)",
    "FUNCTIONAL": "Yes (Functional)",
    "POC":        "Yes (PoC)",
    "UNPROVEN":   "Yes (Unproven)",
}
_MATURITY_EXPLOITABLE = frozenset({"HIGH", "FUNCTIONAL"})

# Assets with no scan or scan older than this threshold are flagged as unscanned
UNSCANNED_THRESHOLD_DAYS: int = 30

# Note included at the top of every vulnerability tab and in the PDF
LARGE_GROUP_NOTE = (
    "Asset-level detail for each plugin is available in Tenable Vulnerability "
    "Management. Filter by Plugin ID to view all affected assets for your group."
)

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / "app.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)


# ===========================================================================
# Step 1 — Data preparation functions
# ===========================================================================


def _fetch_and_prepare(
    tio,
    cache_dir: Path,
    tag_category: Optional[str],
    tag_value: Optional[str],
    as_of: Optional[datetime] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Fetch, enrich, tag-filter, and SLA-annotate the vulnerability dataset.

    Uses the single unscoped parquet cache introduced in the fetcher refactor.
    Tag scoping is performed in-memory after the cache load, so all delivery
    groups running in the same ``run_all.py`` invocation share one API call.

    Unscanned asset exclusion is NOT performed here — call
    ``_identify_unscanned_assets()`` on the returned ``assets_df`` and filter
    ``vulns_df`` by the scanned asset IDs before passing to ``_group_by_plugin()``.

    Parameters
    ----------
    tio : TenableIO
        Authenticated Tenable client.
    cache_dir : Path
        Run-scoped parquet cache directory (shared with other reports).
    tag_category : str or None
        Tenable tag category to scope the report (e.g. "Operations").
    tag_value : str or None
        Tenable tag value to scope the report (e.g. "Server Operations").
    as_of : datetime, optional
        Reference timestamp for SLA age calculations.  Defaults to UTC now.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame]
        ``(vulns_df, assets_df)`` where both are scoped to the tag filter.
        ``vulns_df`` has full SLA columns from ``apply_sla_to_df()`` plus
        the four-state ``ops_sla_status`` column from ``_compute_sla_status()``.
        ``assets_df`` contains all scoped assets including unscanned ones.
    """
    if as_of is None:
        as_of = datetime.now(tz=timezone.utc)

    logger.info(
        "[%s] Fetching vulnerability data (tag=%s:%s)…",
        REPORT_NAME, tag_category, tag_value,
    )

    vulns_df  = fetch_all_vulnerabilities(tio, cache_dir)
    assets_df = fetch_all_assets(tio, cache_dir)

    if vulns_df.empty:
        logger.warning("[%s] No vulnerabilities returned — report will be empty.", REPORT_NAME)
        assets_df = filter_by_tag(assets_df, tag_category, tag_value)
        return vulns_df, assets_df

    logger.info("[%s] Enriching vulnerabilities with asset metadata…", REPORT_NAME)
    df = enrich_vulns_with_assets(vulns_df, assets_df)

    # Tag filtering happens after enrichment — the reliable tags column
    # comes from the asset export join, not the vuln export payload.
    df        = filter_by_tag(df, tag_category, tag_value)
    assets_df = filter_by_tag(assets_df, tag_category, tag_value)

    if df.empty:
        logger.warning(
            "[%s] No vulnerabilities match tag filter (%s=%s) — report will be empty.",
            REPORT_NAME, tag_category, tag_value,
        )
        return df, assets_df

    logger.info("[%s] Applying SLA calculations…", REPORT_NAME)
    df = apply_sla_to_df(df, as_of=as_of)
    df = _compute_sla_status(df)

    logger.info(
        "[%s] Prepared %d vulnerability records across %d unique assets.",
        REPORT_NAME, len(df), df["asset_uuid"].nunique(),
    )
    return df, assets_df


def _compute_sla_status(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add a four-state ``ops_sla_status`` column to a SLA-annotated vuln DataFrame.

    Requires the following columns produced by ``apply_sla_to_df()``:
        - ``is_overdue``     : bool
        - ``days_remaining`` : int  (negative when overdue)
        - ``sla_days``       : int  (None/NaN for info severity)
        - ``remediated``     : bool

    SLA state logic
    ---------------
    Overdue  : days_open > sla_days  (is_overdue == True)
    Urgent   : within SLA AND days_remaining ≤ 25% of sla_days
    Warning  : within SLA AND days_remaining ≤ 50% of sla_days (but > 25%)
    On Track : within SLA AND days_remaining > 50% of sla_days

    Rows with no applicable SLA (info severity, missing sla_days, or
    already remediated) receive ``OPS_SLA_ON_TRACK`` as a safe default.

    Parameters
    ----------
    df : pd.DataFrame
        Vulnerability DataFrame with SLA columns already applied.

    Returns
    -------
    pd.DataFrame
        A copy of ``df`` with the ``ops_sla_status`` column added.
    """
    if df.empty:
        df = df.copy()
        df["ops_sla_status"] = pd.Series(dtype="object")
        return df

    # Guard: sla_days must exist (comes from apply_sla_to_df)
    if "is_overdue" not in df.columns or "days_remaining" not in df.columns:
        raise ValueError(
            "_compute_sla_status requires apply_sla_to_df() to be called first. "
            "Missing columns: is_overdue and/or days_remaining."
        )

    # Compute thresholds — NaN where sla_days is NaN (info severity)
    urgent_threshold  = df["sla_days"] * 0.25
    warning_threshold = df["sla_days"] * 0.50

    # Ordered conditions evaluated top-down by np.select (first match wins)
    conditions = [
        df["is_overdue"],
        ~df["is_overdue"] & (df["days_remaining"] <= urgent_threshold),
        ~df["is_overdue"] & (df["days_remaining"] <= warning_threshold),
    ]
    choices = [OPS_SLA_OVERDUE, OPS_SLA_URGENT, OPS_SLA_WARNING]

    ops_status = np.select(conditions, choices, default=OPS_SLA_ON_TRACK)

    return df.assign(ops_sla_status=ops_status)


def _identify_unscanned_assets(
    assets_df: pd.DataFrame,
    as_of: Optional[datetime] = None,
    threshold_days: int = UNSCANNED_THRESHOLD_DAYS,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Identify assets that have not been scanned recently and split the asset
    list into scanned and unscanned subsets.

    Scan date resolution order (first non-null wins per asset):
      1. ``last_licensed_scan_date`` — Tenable licensed scan timestamp
      2. ``last_scan_time``          — general last-seen-by-scanner timestamp
      3. ``last_seen``               — last discovery event (fallback)

    An asset is considered **unscanned** when:
      - The resolved scan date is null / NaT  (never scanned), OR
      - The resolved scan date is more than ``threshold_days`` before ``as_of``

    Unscanned assets must be excluded from vulnerability counts and SLA
    calculations — their absence from scan data means we cannot make
    reliable vulnerability statements about them.

    Parameters
    ----------
    assets_df : pd.DataFrame
        Full scoped asset DataFrame from ``fetch_all_assets()`` after tag filter.
        Must contain ``asset_uuid``, ``hostname``, ``ipv4``, ``tags``.
        Should contain at least one of ``last_licensed_scan_date``,
        ``last_scan_time``, ``last_seen``.
    as_of : datetime, optional
        Reference timestamp for age calculation.  Defaults to UTC now.
    threshold_days : int
        Days since last scan to consider an asset unscanned.  Default: 30.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame]
        ``(scanned_df, unscanned_df)``

        *scanned_df* — assets with a recent scan; preserves all ``assets_df``
        columns plus ``last_scan_date`` (formatted string) and
        ``days_since_scan`` (int).

        *unscanned_df* — subset for unscanned assets with the same computed
        columns, sorted by ``days_since_scan`` descending (never-scanned
        first, then oldest scan first).
    """
    if as_of is None:
        as_of = datetime.now(tz=timezone.utc)

    _empty_cols = [
        "asset_uuid", "hostname", "ipv4", "tags",
        "mac_address", "last_seen", "last_licensed_scan_date",
        "source_name", "last_scan_date", "days_since_scan",
        "operating_system",
    ]
    if assets_df.empty:
        empty = pd.DataFrame(columns=_empty_cols)
        return empty.copy(), empty.copy()

    as_of_ts = pd.Timestamp(as_of)

    def _coerce_utc(col_name: str) -> "pd.Series":
        if col_name in assets_df.columns:
            return pd.to_datetime(assets_df[col_name], utc=True, errors="coerce")
        return pd.Series(pd.NaT, index=assets_df.index, dtype="datetime64[ns, UTC]")

    # Resolve best available scan date per asset
    licensed  = _coerce_utc("last_licensed_scan_date")
    scan_time = _coerce_utc("last_scan_time")
    last_seen = _coerce_utc("last_seen")

    best_scan = licensed.combine_first(scan_time).combine_first(last_seen)

    days_since = (as_of_ts - best_scan).dt.days
    df = assets_df.assign(
        _best_scan_date=best_scan,
        days_since_scan=days_since,
    )

    # Classify unscanned
    never_scanned  = df["_best_scan_date"].isna()
    stale_scan     = df["days_since_scan"] > threshold_days
    unscanned_mask = never_scanned | stale_scan

    # Add formatted last_scan_date column to all rows
    df = df.assign(
        last_scan_date=df["_best_scan_date"].apply(
            lambda ts: "Never" if pd.isna(ts) else ts.strftime("%Y-%m-%d")
        )
    ).drop(columns="_best_scan_date")

    # --- unscanned subset ------------------------------------------------
    unscanned_base = df[unscanned_mask]
    sort_key = unscanned_base["days_since_scan"].fillna(999_999)
    unscanned_result = (
        unscanned_base.assign(_sort_key=sort_key)
        .sort_values("_sort_key", ascending=False)
        .drop(columns="_sort_key")
        .reset_index(drop=True)
    )
    out_cols = [
        "asset_uuid", "hostname", "ipv4", "tags",
        "mac_address", "last_seen", "last_licensed_scan_date",
        "source_name", "last_scan_date", "days_since_scan",
        "operating_system",
    ]
    out_cols = [c for c in out_cols if c in unscanned_result.columns]
    unscanned_result = unscanned_result[out_cols]

    # --- scanned subset --------------------------------------------------
    scanned_result = df[~unscanned_mask].reset_index(drop=True)

    n_never = int(never_scanned.sum())
    n_stale = int(stale_scan.sum())
    logger.info(
        "[%s] Asset scan classification: %d scanned, %d unscanned "
        "(never scanned: %d, last scan >%d days ago: %d)",
        REPORT_NAME,
        len(scanned_result), len(unscanned_result),
        n_never, threshold_days, n_stale,
    )

    return scanned_result, unscanned_result


def _format_cves(series: pd.Series) -> str:
    """
    Collapse a per-plugin group of comma-separated CVE strings into one string.

    Takes the union of all CVEs across all rows in the group, sorts them,
    and formats as "CVE-YYYY-NNNNN, CVE-... +N more" when the count exceeds 3.

    Parameters
    ----------
    series : pd.Series
        Series of comma-separated CVE strings (e.g. the ``cve_list`` column
        for all rows sharing one plugin_id).

    Returns
    -------
    str
        Formatted CVE string, e.g. "CVE-2021-44228, CVE-2022-0001 +3 more",
        or empty string if no CVEs are present.
    """
    all_cves: set[str] = set()
    for raw in series.fillna(""):
        for cve in raw.split(","):
            cve = cve.strip()
            if cve:
                all_cves.add(cve)

    cves = sorted(all_cves)
    if not cves:
        return ""
    if len(cves) <= 3:
        return ", ".join(cves)
    return ", ".join(cves[:3]) + f" +{len(cves) - 3} more"


def _compute_exploitability_metrics(vulns_df: pd.DataFrame) -> dict:
    """
    Return exploitability counts keyed by unique plugin_id.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Scanned vulnerability rows (post-filter, post-SLA).

    Returns
    -------
    dict
        ``{known_exploit: int, functional: int, high_maturity: int}``
    """
    if vulns_df.empty:
        return {"known_exploit": 0, "functional": 0, "high_maturity": 0}

    plugins = vulns_df.drop_duplicates(subset=["plugin_id"]).copy()

    known_exploit = int(
        plugins["exploit_available"].fillna(False).astype(bool).sum()
        if "exploit_available" in plugins.columns else 0
    )
    functional_count = int(
        (plugins["exploit_code_maturity"] == "FUNCTIONAL").sum()
        if "exploit_code_maturity" in plugins.columns else 0
    )
    high_count = int(
        (plugins["exploit_code_maturity"] == "HIGH").sum()
        if "exploit_code_maturity" in plugins.columns else 0
    )
    return {
        "known_exploit": known_exploit,
        "functional":    functional_count,
        "high_maturity": high_count,
    }


def _get_top_priority_plugins(
    vulns_df: pd.DataFrame,
    n: int = 5,
) -> pd.DataFrame:
    """
    Return up to *n* plugins with FUNCTIONAL or HIGH exploit maturity,
    ranked by maturity (FUNCTIONAL first), then VPR score desc, then
    affected asset count desc.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Scanned vulnerability rows with ``exploit_code_maturity``,
        ``plugin_id``, ``plugin_name``, ``vpr_score``, ``asset_uuid``.

    Returns
    -------
    pd.DataFrame
        Columns: plugin_name, vpr_score, exploit_code_maturity,
        affected_assets.  Empty DataFrame if no qualifying plugins.
    """
    _RANK = {"FUNCTIONAL": 0, "HIGH": 1}

    if "exploit_code_maturity" not in vulns_df.columns or vulns_df.empty:
        return pd.DataFrame()

    exploitable = vulns_df[
        vulns_df["exploit_code_maturity"].isin(["FUNCTIONAL", "HIGH"])
    ].copy()

    if exploitable.empty:
        return pd.DataFrame()

    grouped = (
        exploitable
        .groupby(["plugin_id", "plugin_name", "exploit_code_maturity"])
        .agg(
            vpr_score       = ("vpr_score",   "max"),
            affected_assets = ("asset_uuid",  "nunique"),
        )
        .reset_index()
    )
    grouped["maturity_rank"] = (
        grouped["exploit_code_maturity"].map(_RANK).fillna(99)
    )
    return (
        grouped
        .sort_values(
            ["maturity_rank", "vpr_score", "affected_assets"],
            ascending=[True, False, False],
        )
        .head(n)
        [["plugin_name", "vpr_score", "exploit_code_maturity", "affected_assets"]]
        .reset_index(drop=True)
    )


def _extract_risk_modifications(
    vulns_df: pd.DataFrame,
    assets_df: pd.DataFrame,
    recast_rules_df: pd.DataFrame,
    as_of: datetime,
) -> pd.DataFrame:
    """
    Join active recast/accept rules with vulnerability data to produce one row
    per rule, enriched with plugin context from the vuln export.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Full vulnerability DataFrame (already filtered to scanned assets).
    assets_df : pd.DataFrame
        Asset DataFrame (used for affected asset count per plugin).
    recast_rules_df : pd.DataFrame
        Output of ``fetch_recast_rules()`` — one row per rule.
    as_of : datetime
        Report generation timestamp (UTC) — used to compute days_until_expiry.

    Returns
    -------
    pd.DataFrame
        Columns: Plugin ID, Plugin Name, Modification Type, Original Severity,
        Current Severity, VPR Score, Date Opened, Expiration Date,
        Days Until Expiry, Affected Assets.
        Sorted: Accepted first → Recast; then VPR desc; then Affected Assets desc.
        Empty DataFrame with same columns when no rules exist.
    """
    _OUTPUT_COLS = [
        "Plugin ID", "Plugin Name", "Modification Type", "Recast Reason",
        "Original Severity", "Current Severity", "VPR Score",
        "Date Opened", "Expiration Date", "Days Until Expiry",
        "Affected Assets", "Rule UUID",
    ]

    if vulns_df.empty or "severity_modification_type" not in vulns_df.columns:
        return pd.DataFrame(columns=_OUTPUT_COLS)

    # ------------------------------------------------------------------
    # Step 1: Filter vuln export to accepted / recast findings
    # ------------------------------------------------------------------
    _MOD_TYPES = {"recasted", "accepted"}
    mod_df = vulns_df[
        vulns_df["severity_modification_type"].str.lower().isin(_MOD_TYPES)
    ].copy()

    if mod_df.empty:
        return pd.DataFrame(columns=_OUTPUT_COLS)

    # Normalise plugin_id to int for grouping
    mod_df = mod_df.assign(
        plugin_id=pd.to_numeric(mod_df["plugin_id"], errors="coerce"),
    )

    # Group key: rule UUID + plugin_id (one output row per unique rule×plugin combo)
    uuid_col = "recast_rule_uuid" if "recast_rule_uuid" in mod_df.columns else None
    if uuid_col:
        mod_df = mod_df.assign(
            _rule_uuid=mod_df[uuid_col].fillna("").astype(str).str.strip()
        )
    else:
        mod_df = mod_df.assign(_rule_uuid="")

    agg = (
        mod_df
        .groupby(["_rule_uuid", "plugin_id"], as_index=False, dropna=False)
        .agg(
            plugin_name             = ("plugin_name",              "first"),
            modification_type_raw   = ("severity_modification_type", "first"),
            current_severity        = ("severity",                 "first"),
            vpr_score               = ("vpr_score",               "max"),
            date_opened             = ("first_found",              "min"),
            affected_assets         = ("asset_uuid",               "nunique"),
            recast_reason           = ("recast_reason",            "first")
                                      if "recast_reason" in mod_df.columns
                                      else ("plugin_name", lambda _: ""),
        )
        .reset_index(drop=True)
    )

    # ------------------------------------------------------------------
    # Step 2: Enrich with rules API metadata (expiry, original severity,
    #         filter scope) — optional; gracefully absent if rules API failed
    # ------------------------------------------------------------------
    as_of_utc = as_of if as_of.tzinfo is not None else as_of.replace(tzinfo=timezone.utc)

    rule_meta: dict[str, dict] = {}
    if recast_rules_df is not None and not recast_rules_df.empty:
        for _, r in recast_rules_df.iterrows():
            uid = str(r.get("rule_id") or "").strip()
            if uid:
                rule_meta[uid] = {
                    "original_severity": r.get("original_severity"),
                    "expires_at":        r.get("expires_at"),
                    "filter_summary":    r.get("filter_summary", ""),
                }

    def _days_until(exp) -> int | None:
        if exp is None or (isinstance(exp, str) and exp in ("", "Never")):
            return None
        try:
            ts = pd.Timestamp(exp)
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
            return (ts - as_of_utc).days
        except Exception:
            return None

    def _fmt_date(val) -> str:
        if val is None or (isinstance(val, str) and val.strip() in ("", "Never")):
            return "Never"
        if pd.isna(val):
            return "Never"
        try:
            if hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            return pd.Timestamp(val).strftime("%Y-%m-%d")
        except Exception:
            return str(val)[:10]

    def _clean_sev(val) -> str:
        s = str(val or "").strip().lower()
        return s.title() if s and s not in {"none", "null", "", "n/a"} else "N/A"

    def _mod_label(raw: str) -> str:
        return {"recasted": "Recast", "accepted": "Accepted"}.get(
            str(raw).lower(), str(raw).title()
        )

    # ------------------------------------------------------------------
    # Step 3: Build output rows
    # ------------------------------------------------------------------
    rows_out = []
    for _, row in agg.iterrows():
        uid     = str(row["_rule_uuid"]).strip()
        meta    = rule_meta.get(uid, {})
        exp_val = meta.get("expires_at")

        rows_out.append({
            "Plugin ID":         int(row["plugin_id"]) if pd.notna(row["plugin_id"]) else "",
            "Plugin Name":       str(row["plugin_name"] or ""),
            "Modification Type": _mod_label(row["modification_type_raw"]),
            "Recast Reason":     str(row.get("recast_reason") or ""),
            "Original Severity": _clean_sev(meta.get("original_severity")),
            "Current Severity":  _clean_sev(row["current_severity"]),
            "VPR Score":         round(float(row["vpr_score"]), 1)
                                 if pd.notna(row["vpr_score"]) else "",
            "Date Opened":       _fmt_date(row["date_opened"])
                                 if pd.notna(row["date_opened"]) else "",
            "Expiration Date":   _fmt_date(exp_val),
            "Days Until Expiry": _days_until(exp_val) if exp_val else "",
            "Affected Assets":   int(row["affected_assets"]),
            "Rule UUID":         uid if uid else "",
        })

    out = pd.DataFrame(rows_out, columns=_OUTPUT_COLS)

    # Sort: Accepted first → Recast, then VPR desc, then Affected Assets desc
    _mod_rank = {"Accepted": 0, "Recast": 1}
    out = out.assign(
        _mod_rank=out["Modification Type"].map(_mod_rank).fillna(2),
        _vpr_sort=pd.to_numeric(out["VPR Score"], errors="coerce").fillna(0),
    ).sort_values(
        ["_mod_rank", "_vpr_sort", "Affected Assets"],
        ascending=[True, False, False],
    ).drop(columns=["_mod_rank", "_vpr_sort"]).reset_index(drop=True)

    return out


def _extract_recurring_vulnerabilities(
    vulns_df: pd.DataFrame,
    assets_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Identify vulnerabilities that have resurfaced after a prior remediation.

    A vulnerability is recurring when ``resurfaced_date`` is not null
    (equivalently ``state == "REOPENED"``).  The approximate prior close date
    is derived as ``resurfaced_date - timedelta(seconds=time_taken_to_fix)``
    when ``time_taken_to_fix`` is available.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Full vulnerability DataFrame (already filtered to scanned assets).
    assets_df : pd.DataFrame
        Asset DataFrame used to enrich hostname / ipv4.

    Returns
    -------
    pd.DataFrame
        Columns: Plugin ID, Plugin Name, Asset Name, IP Address,
        Original First Found, Date Closed, Date Reopened, Last Seen,
        Current State, Severity, VPR Score, Exploit Available, Exploit Maturity.
        Sorted: severity rank → VPR desc → Date Reopened desc.
        Empty DataFrame with same columns when no recurring vulns exist.
    """
    from datetime import timedelta

    _OUTPUT_COLS = [
        "Plugin ID", "Plugin Name", "Asset Name", "IP Address",
        "Original First Found", "Date Closed", "Date Reopened", "Last Seen",
        "Current State", "Severity", "VPR Score", "Exploit Available",
        "Exploit Maturity",
    ]

    if vulns_df.empty or "resurfaced_date" not in vulns_df.columns:
        return pd.DataFrame(columns=_OUTPUT_COLS)

    recurring = vulns_df[vulns_df["resurfaced_date"].notna()].copy()
    if recurring.empty:
        return pd.DataFrame(columns=_OUTPUT_COLS)

    # Enrich with asset columns
    if not assets_df.empty and "asset_uuid" in assets_df.columns:
        _asset_cols = [c for c in ["asset_uuid", "hostname", "ipv4"] if c in assets_df.columns]
        _lookup = assets_df[_asset_cols].drop_duplicates("asset_uuid")
        drop_existing = [c for c in ["hostname", "ipv4"] if c in recurring.columns]
        recurring = (
            recurring.drop(columns=drop_existing)
            .merge(_lookup, on="asset_uuid", how="left")
        )

    # Compute approximate date_closed
    def _compute_date_closed(row):
        try:
            rd = row.get("resurfaced_date")
            ttf = row.get("time_taken_to_fix")
            if pd.isna(rd):
                return "Not Available"
            if ttf is not None and pd.notna(ttf) and float(ttf) > 0:
                closed = pd.Timestamp(rd) - timedelta(seconds=float(ttf))
                return closed.strftime("%Y-%m-%d")
            return "Not Available"
        except Exception:
            return "Not Available"

    recurring = recurring.assign(
        date_closed_str=recurring.apply(_compute_date_closed, axis=1)
    )

    # Format date columns
    def _fmt(val):
        if pd.isna(val):
            return ""
        try:
            if hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            return str(val)[:10]
        except Exception:
            return ""

    # Map fields
    recurring = recurring.assign(
        exploit_available_str=recurring["exploit_available"].apply(
            lambda x: "Yes" if x is True or str(x).lower() in ("true", "1", "yes") else "No"
        ) if "exploit_available" in recurring.columns else "Unknown",
        exploit_maturity_str=recurring["exploit_code_maturity"].apply(
            lambda x: str(x).replace("_", " ").title() if pd.notna(x) and x != "" else "N/A"
        ) if "exploit_code_maturity" in recurring.columns else "N/A",
        current_state=recurring["state"].apply(
            lambda x: "Open" if pd.notna(x) and str(x).upper() in ("OPEN", "REOPENED") else str(x).title()
        ) if "state" in recurring.columns else "Open",
    )

    out = pd.DataFrame({
        "Plugin ID":           recurring["plugin_id"].apply(
                                   lambda x: int(x) if pd.notna(x) else ""
                               ) if "plugin_id" in recurring.columns else "",
        "Plugin Name":         recurring["plugin_name"].fillna("Unknown") if "plugin_name" in recurring.columns else "",
        "Asset Name":          recurring["hostname"].fillna("Unknown") if "hostname" in recurring.columns else "",
        "IP Address":          recurring["ipv4"].fillna("") if "ipv4" in recurring.columns else "",
        "Original First Found": recurring["first_found"].apply(_fmt) if "first_found" in recurring.columns else "",
        "Date Closed":         recurring["date_closed_str"],
        "Date Reopened":       recurring["resurfaced_date"].apply(_fmt),
        "Last Seen":           recurring["last_found"].apply(_fmt) if "last_found" in recurring.columns else "",
        "Current State":       recurring["current_state"],
        "Severity":            recurring["severity"].str.title() if "severity" in recurring.columns else "",
        "VPR Score":           recurring["vpr_score"].apply(
                                   lambda x: round(float(x), 1) if pd.notna(x) else ""
                               ) if "vpr_score" in recurring.columns else "",
        "Exploit Available":   recurring["exploit_available_str"],
        "Exploit Maturity":    recurring["exploit_maturity_str"],
    })

    # Sort: severity rank → VPR desc → Date Reopened desc
    _sev_rank = {s: i for i, s in enumerate(["Critical", "High", "Medium", "Low", "Info"])}
    out = out.assign(
        _sev_rank=out["Severity"].map(_sev_rank).fillna(99),
        _vpr_sort=pd.to_numeric(out["VPR Score"], errors="coerce").fillna(0),
        _reopened_sort=pd.to_datetime(out["Date Reopened"], errors="coerce"),
    ).sort_values(
        ["_sev_rank", "_vpr_sort", "_reopened_sort"],
        ascending=[True, False, False],
    ).drop(columns=["_sev_rank", "_vpr_sort", "_reopened_sort"]).reset_index(drop=True)

    return out


def _group_by_plugin(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate the vulnerability DataFrame by plugin, producing one row per plugin.

    SLA status and days-remaining are taken from the **oldest instance** of
    each plugin (the row with the highest ``days_open``) — worst-case semantics
    that ensure the most urgent state for the group is surfaced.

    Sort order applied (as specified):
        1. Severity rank: Critical first → High → Medium → Low
        2. Within severity: highest VPR score first
        3. Within same VPR score: highest affected asset count first
        4. Within same asset count: most days open (oldest instance) first

    Parameters
    ----------
    df : pd.DataFrame
        Vulnerability DataFrame scoped to scanned assets only, with
        ``ops_sla_status`` column present (call ``_compute_sla_status()`` first).

    Returns
    -------
    pd.DataFrame
        One row per plugin with these columns:
            plugin_id, plugin_name, plugin_family, severity, vpr_score,
            cvss_score, cves, affected_asset_count,
            days_open_oldest, days_open_newest,
            sla_status, days_remaining, exploit_available
    """
    if df.empty:
        return pd.DataFrame(columns=[
            "plugin_id", "plugin_name", "plugin_family", "severity",
            "vpr_score", "cvss_score", "cves", "affected_asset_count",
            "days_open_oldest", "days_open_newest",
            "sla_status", "days_remaining", "exploit_available",
            "exploit_code_maturity",
        ])

    # ------------------------------------------------------------------
    # Step A: Identify worst-case (oldest) row per plugin for SLA state.
    # Sort descending on days_open so .first() within each plugin gives
    # the oldest instance.
    # ------------------------------------------------------------------
    worst_case_cols = ["ops_sla_status", "days_remaining"]
    worst_case = (
        df.sort_values("days_open", ascending=False, na_position="last")
        .groupby("plugin_id", sort=False)[worst_case_cols]
        .first()
        .reset_index()
        .rename(columns={"ops_sla_status": "sla_status"})
    )

    # ------------------------------------------------------------------
    # Step B: Standard per-plugin aggregations.
    # Add a numeric maturity rank column so we can take the min (worst)
    # maturity per plugin via a standard "min" aggregation.
    # ------------------------------------------------------------------
    df = df.assign(
        _maturity_rank=df["exploit_code_maturity"].map(MATURITY_RANK).fillna(len(MATURITY_RANK))
        if "exploit_code_maturity" in df.columns
        else len(MATURITY_RANK)
    )
    agg = df.groupby("plugin_id").agg(
        plugin_name          = ("plugin_name",        "first"),
        plugin_family        = ("plugin_family",       "first"),
        vpr_score            = ("vpr_score",           "max"),
        cvss_score           = ("cvss3_score",         "max"),
        affected_asset_count = ("asset_uuid",          "nunique"),
        days_open_oldest     = ("days_open",           "max"),
        days_open_newest     = ("days_open",           "min"),
        exploit_available    = ("exploit_available",   "any"),
        _maturity_rank_min   = ("_maturity_rank",      "min"),
    ).reset_index()

    # Map min maturity rank back to its label string; derive severity in same step
    _rank_to_maturity = {v: k for k, v in MATURITY_RANK.items()}
    agg = agg.assign(
        exploit_code_maturity=agg["_maturity_rank_min"].map(_rank_to_maturity).fillna(""),
        severity=agg["vpr_score"].apply(lambda v: vpr_to_severity(v, fallback="info")),
    )

    # ------------------------------------------------------------------
    # Step C: CVE aggregation — union of all CVEs per plugin
    # ------------------------------------------------------------------
    cve_agg = (
        df.groupby("plugin_id")["cve_list"]
        .apply(_format_cves)
        .reset_index()
        .rename(columns={"cve_list": "cves"})
    )

    # ------------------------------------------------------------------
    # Step E: Merge all aggregated components
    # ------------------------------------------------------------------
    result = (
        agg
        .merge(cve_agg,    on="plugin_id", how="left")
        .merge(worst_case, on="plugin_id", how="left")
    )

    # ------------------------------------------------------------------
    # Step F: Format exploit_available as a human-readable string that
    #         reflects the highest-risk maturity level per plugin.
    #         Priority: HIGH > FUNCTIONAL > POC > UNPROVEN > plain Yes > No
    # Step G: Apply sort order: severity → maturity rank → VPR → asset count
    #         → days open.  Maturity rank ensures functionally-exploitable
    #         plugins surface above theoretically-exploitable ones.
    # ------------------------------------------------------------------
    def _fmt_exploit(row) -> str:
        mat = row.get("exploit_code_maturity", "")
        if mat in _MATURITY_DISPLAY:
            return _MATURITY_DISPLAY[mat]
        return "Yes" if row.get("exploit_available") else "No"

    exploit_fmt = result.apply(_fmt_exploit, axis=1)
    sev_rank     = result["severity"].str.lower().map(SEVERITY_RANK).fillna(99).astype(int)
    mat_rank     = result["_maturity_rank_min"].fillna(len(MATURITY_RANK)).astype(int)
    result = result.assign(
        exploit_available=exploit_fmt,
        _sev_rank=sev_rank,
        _mat_rank=mat_rank,
    )
    result = (
        result
        .sort_values(
            ["_sev_rank", "_mat_rank", "vpr_score", "affected_asset_count", "days_open_oldest"],
            ascending=[True, True, False, False, False],
            na_position="last",
        )
        .drop(columns=["_sev_rank", "_mat_rank", "_maturity_rank_min"])
        .reset_index(drop=True)
    )

    # ------------------------------------------------------------------
    # Step H: Select and order final output columns
    # ------------------------------------------------------------------
    result = result[[
        "plugin_id",
        "plugin_name",
        "plugin_family",
        "severity",
        "vpr_score",
        "cvss_score",
        "cves",
        "affected_asset_count",
        "days_open_oldest",
        "days_open_newest",
        "sla_status",
        "days_remaining",
        "exploit_available",
        "exploit_code_maturity",
    ]]

    logger.info(
        "[%s] Grouped into %d plugins | %d unique assets affected.",
        REPORT_NAME, len(result), df["asset_uuid"].nunique(),
    )
    return result


def _compute_summary_metrics(
    vulns_df: pd.DataFrame,
    plugin_df: pd.DataFrame,
    assets_df: pd.DataFrame,
    unscanned_df: pd.DataFrame,
    tag_category: Optional[str],
    tag_value: Optional[str],
    as_of: Optional[datetime] = None,
) -> dict:
    """
    Compute all summary metrics used in the Excel Summary tab, PDF executive
    summary, and email KPI strip.

    Parameters
    ----------
    vulns_df : pd.DataFrame
        Full scoped and SLA-annotated vuln DataFrame (scanned assets only).
    plugin_df : pd.DataFrame
        Output of ``_group_by_plugin()`` for derived counts.
    assets_df : pd.DataFrame
        All scoped assets (scanned + unscanned).
    unscanned_df : pd.DataFrame
        Output of ``_identify_unscanned_assets()``.
    tag_category, tag_value : str or None
        Tag filter applied to the report.
    as_of : datetime, optional
        Reference timestamp for the report.

    Returns
    -------
    dict
        Keys match the metric names used by the email template and Excel
        Summary tab builder.  All count values are Python ints; rates are
        floats (0.0–1.0).
    """
    if as_of is None:
        as_of = datetime.now(tz=timezone.utc)

    tag_filter_str = (
        f"{tag_category} = {tag_value}"
        if tag_category and tag_value
        else "All Assets"
    )

    total_assets_in_scope = len(assets_df)
    total_unscanned       = len(unscanned_df)
    total_scanned         = total_assets_in_scope - total_unscanned

    if vulns_df.empty:
        return {
            "generated_at":         as_of.strftime("%Y-%m-%d %H:%M UTC"),
            "tag_filter":           tag_filter_str,
            "total_assets":         total_scanned,
            "total_unscanned":      total_unscanned,
            "open_critical":        0,
            "open_high":            0,
            "open_medium":          0,
            "open_low":             0,
            "count_overdue":        0,
            "count_urgent":         0,
            "count_warning":        0,
            "count_on_track":       0,
            "plugins_with_overdue": 0,
            "exploitable_plugins":  0,
            "sla_compliance_rate":  1.0,
            "top5_plugins":         [],
        }

    # Open vuln counts by severity (per finding row, not per plugin)
    open_by_sev: dict[str, int] = {
        sev: int((vulns_df["severity"].str.lower() == sev).sum())
        for sev in ("critical", "high", "medium", "low")
    }

    # SLA state counts (per finding row)
    state_counts: dict[str, int] = {
        state: int((vulns_df["ops_sla_status"] == state).sum())
        for state in OPS_SLA_STATE_ORDER
    }

    # SLA compliance rate: fraction of rows that are On Track (not Overdue/Urgent/Warning)
    total_rows = len(vulns_df)
    on_track_rows = state_counts.get(OPS_SLA_ON_TRACK, 0)
    sla_compliance_rate = round(on_track_rows / total_rows, 4) if total_rows > 0 else 1.0

    # Plugin-level derived metrics
    plugins_with_overdue = int(
        (plugin_df["sla_status"] == OPS_SLA_OVERDUE).sum()
    ) if not plugin_df.empty else 0

    if not plugin_df.empty and "exploit_code_maturity" in plugin_df.columns:
        exploitable_plugins = int(
            plugin_df["exploit_code_maturity"].isin(_MATURITY_EXPLOITABLE).sum()
        )
    elif not plugin_df.empty:
        exploitable_plugins = int(
            plugin_df["exploit_available"].str.startswith("Yes").sum()
        )
    else:
        exploitable_plugins = 0

    # Top 5 plugins by affected asset count
    top5 = (
        plugin_df
        .nlargest(5, "affected_asset_count", keep="first")
        [["plugin_name", "affected_asset_count"]]
        .to_dict("records")
    ) if not plugin_df.empty else []

    return {
        "generated_at":         as_of.strftime("%Y-%m-%d %H:%M UTC"),
        "tag_filter":           tag_filter_str,
        "total_assets":         total_scanned,
        "total_unscanned":      total_unscanned,
        "open_critical":        open_by_sev["critical"],
        "open_high":            open_by_sev["high"],
        "open_medium":          open_by_sev["medium"],
        "open_low":             open_by_sev["low"],
        "count_overdue":        state_counts.get(OPS_SLA_OVERDUE,  0),
        "count_urgent":         state_counts.get(OPS_SLA_URGENT,   0),
        "count_warning":        state_counts.get(OPS_SLA_WARNING,  0),
        "count_on_track":       state_counts.get(OPS_SLA_ON_TRACK, 0),
        "plugins_with_overdue": plugins_with_overdue,
        "exploitable_plugins":  exploitable_plugins,
        "sla_compliance_rate":  sla_compliance_rate,
        "top5_plugins":         top5,
    }


# ===========================================================================
# Steps 2–5 placeholders — to be implemented after Step 1 review
# ===========================================================================

def _apply_ops_sla_formatting(ws) -> None:
    """
    Apply four-state SLA conditional formatting to the ``sla_status`` column
    in an ops_remediation worksheet.

    States and fills:
      Overdue                    → red   (#FFCDD2 / bold red text)
      Urgent - <=25% SLA Remaining  → orange (#FFE0B2 / bold orange text)
      Warning - <=50% SLA Remaining → yellow (#FFF9C4 / amber text)
      On Track                   → green (#C8E6C9 / dark-green text)
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from exporters.excel_exporter import _col_letter_for  # reuse helper

    col = _col_letter_for(ws, "sla_status")
    if not col:
        return

    data_start = 2
    data_end   = ws.max_row
    rng        = f"{col}{data_start}:{col}{data_end}"

    rules = [
        (OPS_SLA_OVERDUE,  "FFCDD2", "B71C1C", True),
        (OPS_SLA_URGENT,   "FFE0B2", "E65100", True),
        (OPS_SLA_WARNING,  "FFF9C4", "F57F17", False),
        (OPS_SLA_ON_TRACK, "C8E6C9", "1B5E20", False),
    ]
    for label, bg, fg, bold in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=[f'"{label}"'],
                fill=PatternFill("solid", fgColor=bg),
                font=Font(bold=bold, color=fg, size=10, name="Calibri"),
            ),
        )


def _build_summary_sheet(wb, summary: dict) -> None:
    """Write the Summary tab with KPI metrics from the summary dict."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet(title="Summary", index=0)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    _title_font   = Font(bold=True, size=14, color="1F3864", name="Calibri")
    _section_font = Font(bold=True, size=11, color="1F3864", name="Calibri")
    _label_font   = Font(bold=True, size=10, name="Calibri")
    _value_font   = Font(size=10, name="Calibri")
    _head_fill    = PatternFill("solid", fgColor="E8EAF6")
    _alt_fill     = PatternFill("solid", fgColor="F5F7FA")
    _red_fill     = PatternFill("solid", fgColor="FFCDD2")
    _orange_fill  = PatternFill("solid", fgColor="FFE0B2")
    _yellow_fill  = PatternFill("solid", fgColor="FFF9C4")
    _green_fill   = PatternFill("solid", fgColor="C8E6C9")

    sev_fills = {
        "critical": PatternFill("solid", fgColor="FFCDD2"),
        "high":     PatternFill("solid", fgColor="FFE0B2"),
        "medium":   PatternFill("solid", fgColor="FFF9C4"),
        "low":      PatternFill("solid", fgColor="C8E6C9"),
    }
    sla_fills = {
        OPS_SLA_OVERDUE:  _red_fill,
        OPS_SLA_URGENT:   _orange_fill,
        OPS_SLA_WARNING:  _yellow_fill,
        OPS_SLA_ON_TRACK: _green_fill,
    }

    def _w(row, col, val, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font:  cell.font  = font
        if fill:  cell.fill  = fill
        if align: cell.alignment = align
        return cell

    row = 1
    _w(row, 1, "Operations Remediation Report — Summary",
       font=_title_font)
    ws.row_dimensions[row].height = 26
    row += 2

    # --- Scope & timestamp ---
    _w(row, 1, "Report Metadata", font=_section_font)
    row += 1
    _w(row, 1, "Generated (UTC)", font=_label_font, fill=_head_fill)
    _w(row, 2, summary.get("generated_at", ""), font=_value_font)
    row += 1
    _w(row, 1, "Scope / Tag Filter", font=_label_font, fill=_alt_fill)
    _w(row, 2, summary.get("tag_filter", "All Assets"), font=_value_font)
    row += 1
    _w(row, 1, "Total Assets in Scope", font=_label_font, fill=_head_fill)
    _w(row, 2, summary.get("total_assets", 0), font=_value_font)
    row += 1
    _w(row, 1, "Unscanned / Stale Assets (>30d)", font=_label_font, fill=_alt_fill)
    _w(row, 2, summary.get("total_unscanned", 0), font=_value_font)
    row += 2

    # --- Open vuln counts by severity ---
    _w(row, 1, "Open Vulnerability Counts by Severity", font=_section_font)
    row += 1
    for i, sev in enumerate(("critical", "high", "medium", "low")):
        label = sev.title()
        count = summary.get(f"open_{sev}", 0)
        fill  = sev_fills.get(sev, _alt_fill)
        _w(row, 1, label, font=_label_font, fill=fill)
        _w(row, 2, count, font=_value_font)
        row += 1
    row += 1

    # --- SLA state breakdown ---
    _w(row, 1, "SLA State Breakdown (by finding)", font=_section_font)
    row += 1
    state_map = [
        ("count_overdue",   OPS_SLA_OVERDUE,  _red_fill),
        ("count_urgent",    OPS_SLA_URGENT,   _orange_fill),
        ("count_warning",   OPS_SLA_WARNING,  _yellow_fill),
        ("count_on_track",  OPS_SLA_ON_TRACK, _green_fill),
    ]
    for key, label, fill in state_map:
        _w(row, 1, label, font=_label_font, fill=fill)
        _w(row, 2, summary.get(key, 0), font=_value_font)
        row += 1
    row += 1

    # --- Plugin-level metrics ---
    _w(row, 1, "Plugin-Level Metrics", font=_section_font)
    row += 1
    _w(row, 1, "Unique Plugins with Overdue Findings", font=_label_font, fill=_head_fill)
    _w(row, 2, summary.get("plugins_with_overdue", 0), font=_value_font)
    row += 1
    compliance_pct = round(summary.get("sla_compliance_rate", 0) * 100, 1)
    _w(row, 1, "SLA Compliance Rate (On Track %)", font=_label_font, fill=_head_fill)
    _w(row, 2, f"{compliance_pct}%", font=_value_font)
    row += 2

    # --- Exploitability section ---
    _navy_fill  = PatternFill("solid", fgColor="1F3864")
    _navy_font  = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    _exploit_orange_fill = PatternFill("solid", fgColor="FFC000")
    _exploit_red_fill    = PatternFill("solid", fgColor="FF0000")
    _exploit_hi_fill     = PatternFill("solid", fgColor="FF6600")
    _exploit_green_fill  = PatternFill("solid", fgColor="E2EFDA")
    _white_font          = Font(bold=False, size=10, color="FFFFFF", name="Calibri")

    _w(row, 1, "Exploitability", font=_navy_font, fill=_navy_fill)
    _w(row, 2, "",               font=_navy_font, fill=_navy_fill)
    row += 1

    known = summary.get("known_exploit", 0)
    _w(row, 1, "Plugins with Known Exploit", font=_label_font, fill=_head_fill)
    known_cell = _w(row, 2, known, font=_value_font,
                    fill=_exploit_orange_fill if known > 0 else _exploit_green_fill)
    row += 1

    func = summary.get("exploit_functional", 0)
    _w(row, 1, "Exploit Maturity — Functional", font=_label_font, fill=_alt_fill)
    func_fill = _exploit_red_fill if func > 0 else _exploit_green_fill
    func_font = _white_font if func > 0 else _value_font
    _w(row, 2, func, font=func_font, fill=func_fill)
    row += 1

    high = summary.get("exploit_high", 0)
    _w(row, 1, "Exploit Maturity — High", font=_label_font, fill=_head_fill)
    high_fill = _exploit_hi_fill if high > 0 else _exploit_green_fill
    high_font = _white_font if high > 0 else _value_font
    _w(row, 2, high, font=high_font, fill=high_fill)
    row += 2

    # --- Top 5 Priority Plugins (Functional & High exploit maturity) ---
    priority = summary.get("priority_plugins", [])
    _w(row, 1, "Top 5 Priority Plugins (Functional & High Exploit Maturity)",
       font=_section_font)
    row += 1
    if priority:
        _w(row, 1, "Plugin Name",      font=_label_font, fill=_head_fill)
        _w(row, 2, "VPR Score",        font=_label_font, fill=_head_fill)
        _w(row, 3, "Exploit Maturity", font=_label_font, fill=_head_fill)
        _w(row, 4, "Affected Assets",  font=_label_font, fill=_head_fill)
        row += 1
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 16
        for i, p in enumerate(priority):
            fill = _alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            vpr_str = (
                f"{float(p.get('vpr_score', 0)):.1f}"
                if p.get("vpr_score") is not None else "N/A"
            )
            mat_display = (
                "Functional" if p.get("exploit_code_maturity") == "FUNCTIONAL"
                else "High" if p.get("exploit_code_maturity") == "HIGH"
                else p.get("exploit_code_maturity", "")
            )
            _w(row, 1, p.get("plugin_name", ""),     font=_value_font, fill=fill)
            _w(row, 2, vpr_str,                      font=_value_font, fill=fill)
            _w(row, 3, mat_display,                  font=_value_font, fill=fill)
            _w(row, 4, p.get("affected_assets", 0),  font=_value_font, fill=fill)
            row += 1
        n_shown = len(priority)
        if n_shown < 5:
            _w(row, 1,
               f"Showing {n_shown} of {n_shown} qualifying plugins.",
               font=_value_font)
            row += 1
    else:
        _w(row, 1,
           "No plugins with Functional or High exploit maturity found "
           "in this reporting period.",
           font=_value_font)
        row += 1

    row += 1

    # --- Risk Management section ---
    _w(row, 1, "Risk Management", font=_navy_font, fill=_navy_fill)
    _w(row, 2, "",                font=_navy_font, fill=_navy_fill)
    row += 1

    accepted  = summary.get("count_risk_accepted", 0)
    recast    = summary.get("count_risk_recast", 0)
    expiring  = summary.get("count_expiring_soon", 0)
    expired   = summary.get("count_expired", 0)
    recurring = summary.get("count_recurring", 0)

    _w(row, 1, "Accepted Findings (suppressed from counts)", font=_label_font, fill=_head_fill)
    _w(row, 2, accepted, font=_value_font,
       fill=_exploit_orange_fill if accepted > 0 else _exploit_green_fill)
    row += 1

    _w(row, 1, "Recast Findings (severity changed)", font=_label_font, fill=_alt_fill)
    _w(row, 2, recast, font=_value_font,
       fill=_exploit_orange_fill if recast > 0 else _exploit_green_fill)
    row += 1

    _w(row, 1, "Rules Expiring Within 30 Days", font=_label_font, fill=_head_fill)
    _w(row, 2, expiring, font=Font(bold=True if expiring > 0 else False, size=10, name="Calibri"),
       fill=_orange_fill if expiring > 0 else _exploit_green_fill)
    row += 1

    _w(row, 1, "Expired Rules (past expiration date)", font=_label_font, fill=_alt_fill)
    _w(row, 2, expired, font=Font(bold=True if expired > 0 else False, color="B71C1C" if expired > 0 else "000000", size=10, name="Calibri"),
       fill=_red_fill if expired > 0 else _exploit_green_fill)
    row += 1

    _w(row, 1, "Recurring Vulnerabilities (resurfaced after fix)", font=_label_font, fill=_head_fill)
    _w(row, 2, recurring, font=Font(bold=True if recurring > 0 else False, color="E65100" if recurring > 0 else "000000", size=10, name="Calibri"),
       fill=_orange_fill if recurring > 0 else _exploit_green_fill)
    row += 1


def _extend_metadata_tab(
    wb,
    risk_mods_df: Optional[pd.DataFrame],
    recurring_df: Optional[pd.DataFrame],
) -> None:
    """
    Append tab reference and field availability notes to the "Report Info"
    worksheet created by ``write_metadata_tab()``.

    Adds two sections:
      - Workbook Tab Reference: one row per tab with name + description
      - Field Availability Notes: explains which fields may be absent and why
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb["Report Info"]
    if ws is None:
        return

    _section_font = Font(bold=True, size=11, color="1F3864", name="Calibri")
    _label_font   = Font(bold=True, size=10, name="Calibri")
    _value_font   = Font(size=10, name="Calibri")
    _note_font    = Font(italic=True, size=9, color="757575", name="Calibri")
    _head_fill    = PatternFill("solid", fgColor="E8EAF6")
    _alt_fill     = PatternFill("solid", fgColor="F5F7FA")
    _white_fill   = PatternFill("solid", fgColor="FFFFFF")

    def _w(row, col, val, font=None, fill=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font: cell.font  = font
        if fill: cell.fill  = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return cell

    # Find the next empty row (leave one blank spacer row after existing content)
    row = ws.max_row + 2

    # --- Workbook Tab Reference ---
    _w(row, 1, "Workbook Tab Reference", font=_section_font)
    ws.row_dimensions[row].height = 20
    row += 1

    _w(row, 1, "Tab Name",    font=_label_font, fill=_head_fill)
    _w(row, 2, "Description", font=_label_font, fill=_head_fill)
    row += 1

    tab_ref = [
        ("Summary",
         "KPI metrics and state breakdown for the full reporting scope. "
         "Includes open vuln counts by severity, SLA state distribution, "
         "exploitability, risk management counts, and top priority plugins."),
        ("Plugins",
         "One row per unique plugin with open findings. Shows severity, VPR score, "
         "affected asset count, oldest/newest days open, CVEs, exploit availability, "
         "and four-state SLA status (Overdue / Urgent / Warning / On Track)."),
        ("Overdue Detail",
         "Individual vulnerability rows where SLA has been breached (status = Overdue). "
         "Sorted by severity then days open descending. Each row represents one "
         "finding on one asset."),
        ("Unscanned Assets",
         "Assets in scope that have not received a licensed scan within the last 30 days. "
         "Includes last seen date, last licensed scan date, days since scan, and source. "
         "These assets are excluded from all vulnerability counts."),
        ("Risk Acceptances & Recasts",
         "Active HOST-scoped risk rules from Tenable. Accepted findings are suppressed "
         "from open vuln counts. Recast findings have had their severity changed from "
         "the original value. Includes expiration dates and days until expiry. "
         "Expired rules are highlighted red; rules expiring within 30 days are orange."),
        ("Recurring Vulnerabilities",
         "Findings that were previously remediated and have resurfaced. Detected via "
         "the resurfaced_date field in the Tenable export. Date Closed is an "
         "approximation based on resurfaced_date minus time_taken_to_fix. "
         "Recurring findings indicate a systemic patching gap on the affected asset."),
        ("Report Info",
         "This tab. Contains report metadata, SLA definitions, VPR score ranges, "
         "tab descriptions, and field availability notes."),
    ]

    for i, (tab_name, desc) in enumerate(tab_ref):
        fill = _alt_fill if i % 2 == 0 else _white_fill
        _w(row, 1, tab_name, font=_label_font, fill=fill)
        _w(row, 2, desc,     font=_value_font, fill=fill)
        ws.row_dimensions[row].height = 40
        row += 1

    ws.column_dimensions["B"].width = 72
    row += 1

    # --- Field Availability Notes ---
    _w(row, 1, "Field Availability Notes", font=_section_font)
    ws.row_dimensions[row].height = 20
    row += 1

    has_risk_mods  = risk_mods_df is not None and not risk_mods_df.empty
    has_recurring  = recurring_df is not None and not recurring_df.empty

    field_notes = [
        ("Risk Acceptances & Recasts tab",
         "Populated from POST /v1/recast/rules/search (Tenable IO). "
         "Only HOST-scoped rules with action RECAST or ACCEPT are included. "
         "Rules for Host Audits (CHANGE_RESULT / ACCEPT_RESULT) and Web App scans "
         "are excluded. Currently shows "
         + (f"{len(risk_mods_df)} rule(s)." if has_risk_mods
            else "0 rules — either no active rules exist for this scope, "
                 "or the accepted state filter may need to include 'accepted' "
                 "findings in the vulnerability export.")),
        ("Date Closed (Recurring Vulnerabilities tab)",
         "Approximated as resurfaced_date minus time_taken_to_fix (seconds). "
         "Shown as 'Not Available' when time_taken_to_fix is absent or zero. "
         "Currently shows "
         + (f"{len(recurring_df)} recurring finding(s)." if has_recurring
            else "0 recurring findings — no resurfaced_date values found in export.")),
        ("VPR Score",
         "Vulnerability Priority Rating from Tenable. Used to derive severity "
         "(Critical 9.0–10.0, High 7.0–8.9, Medium 4.0–6.9, Low 0.1–3.9). "
         "Findings with no VPR score fall back to native Tenable severity."),
        ("Days Open",
         "Calculated as report generation date minus first_found date (UTC). "
         "Reflects calendar days, not business days."),
        ("Affected Assets (Risk Acceptances & Recasts tab)",
         "Count of distinct assets in this scope with an open finding matching "
         "the plugin ID of the rule. Assets outside the tag filter are excluded."),
    ]

    _w(row, 1, "Field",       font=_label_font, fill=_head_fill)
    _w(row, 2, "Note",        font=_label_font, fill=_head_fill)
    row += 1

    for i, (field, note) in enumerate(field_notes):
        fill = _alt_fill if i % 2 == 0 else _white_fill
        _w(row, 1, field, font=Font(bold=True, size=9, name="Calibri"), fill=fill)
        _w(row, 2, note,  font=_note_font, fill=fill)
        ws.row_dimensions[row].height = 45
        row += 1


def _build_risk_mods_sheet(
    wb,
    risk_mods_df: Optional[pd.DataFrame],
    tag_filter_str: str,
) -> None:
    """
    Write the 'Risk Acceptances & Recasts' tab to *wb*.

    Columns: Plugin ID, Plugin Name, Modification Type, Original Severity,
    Current Severity, VPR Score, Date Opened, Expiration Date,
    Days Until Expiry, Affected Assets.

    Conditional formatting:
      - Expiration Date / Days Until Expiry: red if expired (<0), orange if
        expiring within 30 days, green if >30 days or no expiry.
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from exporters.excel_exporter import _col_letter_for, write_dataframe_to_sheet

    SHEET_NAME = "Risk Acceptances & Recasts"
    NOTE = (
        "Accepted: finding is acknowledged and suppressed from counts. "
        "Recast: finding severity has been changed from its original value. "
        "Expiration Date shows when the rule lapses — expired rules are highlighted red."
    )

    if risk_mods_df is None or risk_mods_df.empty:
        ws = wb.create_sheet(title=SHEET_NAME)
        ws.cell(row=1, column=1, value=NOTE)
        ws.cell(row=3, column=1, value="No risk acceptances or recasts found for this scope.")
        ws.column_dimensions["A"].width = 80
        return

    write_dataframe_to_sheet(
        wb, risk_mods_df,
        sheet_name=SHEET_NAME,
        title_row=f"Risk Acceptances & Recasts — {tag_filter_str}",
        severity_col="Current Severity",
    )
    ws = wb[SHEET_NAME]

    # Insert the explanatory note above the title row
    ws.insert_rows(1)
    note_cell = ws.cell(row=1, column=1, value=NOTE)
    note_cell.font = Font(italic=True, size=9, color="757575", name="Calibri")
    if len(risk_mods_df.columns) > 1:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(risk_mods_df.columns),
        )
    ws.row_dimensions[1].height = 30

    # Conditional formatting on Days Until Expiry column
    exp_col = _col_letter_for(ws, "Days Until Expiry")
    if exp_col:
        data_start = 4   # row 1=note, row 2=title, row 3=header, row 4+ data
        data_end   = ws.max_row
        rng        = f"{exp_col}{data_start}:{exp_col}{data_end}"
        _red    = PatternFill("solid", fgColor="FFCDD2")
        _orange = PatternFill("solid", fgColor="FFE0B2")
        _green  = PatternFill("solid", fgColor="C8E6C9")
        _red_font    = Font(bold=True, color="B71C1C", size=10, name="Calibri")
        _orange_font = Font(bold=True, color="E65100", size=10, name="Calibri")
        _green_font  = Font(color="1B5E20", size=10, name="Calibri")
        # Expired (negative)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["0"],
                       fill=_red, font=_red_font),
        )
        # Expiring within 30 days
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["0", "30"],
                       fill=_orange, font=_orange_font),
        )
        # Safe (>30 days)
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["30"],
                       fill=_green, font=_green_font),
        )

    ws.freeze_panes = "A4"


def _build_recurring_sheet(
    wb,
    recurring_df: Optional[pd.DataFrame],
    tag_filter_str: str,
) -> None:
    """
    Write the 'Recurring Vulnerabilities' tab to *wb*.

    Columns: Plugin ID, Plugin Name, Asset Name, IP Address,
    Original First Found, Date Closed, Date Reopened, Last Seen,
    Current State, Severity, VPR Score, Exploit Available, Exploit Maturity.

    Conditional formatting:
      - Severity column: standard severity fill colours.
      - Exploit Available column: red fill when "Yes".
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from exporters.excel_exporter import _col_letter_for, write_dataframe_to_sheet

    SHEET_NAME = "Recurring Vulnerabilities"
    NOTE = (
        "Recurring vulnerabilities were previously remediated but have resurfaced. "
        "Date Closed is approximated from the resurfaced date minus the recorded fix duration. "
        "Prioritise recurring findings — repeated remediation failures indicate a systemic patching gap."
    )

    if recurring_df is None or recurring_df.empty:
        ws = wb.create_sheet(title=SHEET_NAME)
        ws.cell(row=1, column=1, value=NOTE)
        ws.cell(row=3, column=1, value="No recurring vulnerabilities found for this scope.")
        ws.column_dimensions["A"].width = 80
        return

    write_dataframe_to_sheet(
        wb, recurring_df,
        sheet_name=SHEET_NAME,
        title_row=f"Recurring Vulnerabilities — {tag_filter_str}",
        severity_col="Severity",
    )
    ws = wb[SHEET_NAME]

    # Insert the explanatory note above the title row
    ws.insert_rows(1)
    note_cell = ws.cell(row=1, column=1, value=NOTE)
    note_cell.font = Font(italic=True, size=9, color="757575", name="Calibri")
    if len(recurring_df.columns) > 1:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(recurring_df.columns),
        )
    ws.row_dimensions[1].height = 30

    # Conditional formatting on Exploit Available column
    ea_col = _col_letter_for(ws, "Exploit Available")
    if ea_col:
        data_start = 4
        data_end   = ws.max_row
        rng        = f"{ea_col}{data_start}:{ea_col}{data_end}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=['"Yes"'],
                fill=PatternFill("solid", fgColor="FFCDD2"),
                font=Font(bold=True, color="B71C1C", size=10, name="Calibri"),
            ),
        )

    ws.freeze_panes = "A4"


def _build_excel(
    plugin_df: pd.DataFrame,
    overdue_df: pd.DataFrame,
    unscanned_df: pd.DataFrame,
    summary: dict,
    output_path: Path,
    tag_category: Optional[str],
    tag_value: Optional[str],
    risk_mods_df: Optional[pd.DataFrame] = None,
    recurring_df: Optional[pd.DataFrame] = None,
) -> Path:
    """
    Build ops_remediation.xlsx with seven worksheets:

    1. Summary                   — KPI metrics from the summary dict
    2. Plugins                   — plugin_df with four-state SLA conditional formatting
    3. Overdue Detail             — raw vuln rows that are Overdue
    4. Unscanned Assets           — assets with no recent scan
    5. Risk Acceptances & Recasts — accepted / recast rules with expiry tracking
    6. Recurring Vulnerabilities  — vulns that resurfaced after prior remediation
    7. Report Info                — metadata tab

    Parameters
    ----------
    plugin_df : pd.DataFrame
        Output of ``_group_by_plugin()``.
    overdue_df : pd.DataFrame
        Vuln-level rows filtered to ``ops_sla_status == OPS_SLA_OVERDUE``.
    unscanned_df : pd.DataFrame
        Output of ``_identify_unscanned_assets()``.
    summary : dict
        Output of ``_compute_summary_metrics()``.
    output_path : Path
        Destination file path (must end in .xlsx).
    tag_category, tag_value : str or None
        Tag filter — forwarded to the Report Info tab.
    risk_mods_df : pd.DataFrame, optional
        Output of ``_extract_risk_modifications()``.
    recurring_df : pd.DataFrame, optional
        Output of ``_extract_recurring_vulnerabilities()``.

    Returns
    -------
    Path
        The written file path (same as *output_path*).
    """
    from openpyxl import Workbook
    from exporters.excel_exporter import (
        write_dataframe_to_sheet,
        write_metadata_tab,
    )

    tag_filter_str = (
        f"{tag_category} = {tag_value}"
        if tag_category and tag_value
        else "All Assets"
    )

    wb = Workbook()
    # Remove default sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Tab 1: Summary ---
    _build_summary_sheet(wb, summary)

    # --- Tab 2: Plugins ---
    plugin_display = plugin_df.copy()
    # Rename columns to title-case display names for readability
    plugin_display = plugin_display.rename(columns={
        "plugin_id":           "Plugin ID",
        "plugin_name":         "Plugin Name",
        "plugin_family":       "Plugin Family",
        "severity":            "Severity",
        "vpr_score":           "VPR Score",
        "cvss_score":          "CVSS Score",
        "cves":                "CVEs",
        "affected_asset_count":"Affected Assets",
        "days_open_oldest":    "Days Open (Oldest)",
        "days_open_newest":    "Days Open (Newest)",
        "sla_status":            "SLA Status",
        "days_remaining":        "Days Remaining",
        "exploit_available":     "Exploit Available",
        "exploit_code_maturity": "Exploit Maturity",
    })
    write_dataframe_to_sheet(
        wb, plugin_display,
        sheet_name="Plugins",
        title_row=f"Plugin Summary — {tag_filter_str}",
        severity_col="Severity",
    )
    # Apply four-state SLA formatting — must rename col to match internal helper
    ws_plugins = wb["Plugins"]
    # Temporarily map display header back for helper lookup
    for cell in ws_plugins[2]:  # row 2 is header (row 1 is title)
        if cell.value == "SLA Status":
            cell.value = "sla_status"
            break
    _apply_ops_sla_formatting(ws_plugins)
    # Restore display header
    for cell in ws_plugins[2]:
        if cell.value == "sla_status":
            cell.value = "SLA Status"
            break

    # --- Tab 3: Overdue Detail ---
    if not overdue_df.empty:
        overdue_display = overdue_df[[
            c for c in [
                "asset_uuid", "hostname", "mac_address", "ipv4",
                "plugin_id", "plugin_name", "severity", "vpr_score",
                "cve", "days_open", "days_remaining", "ops_sla_status",
                "first_found",
            ] if c in overdue_df.columns
        ]].copy()
        overdue_display = overdue_display.rename(columns={
            "asset_uuid":     "Asset UUID",
            "hostname":       "Asset Name",
            "mac_address":    "MAC Address",
            "ipv4":           "IP Address",
            "plugin_id":      "Plugin ID",
            "plugin_name":    "Plugin Name",
            "severity":       "Severity",
            "vpr_score":      "VPR Score",
            "cve":            "CVE",
            "days_open":      "Days Open",
            "days_remaining": "Days Remaining",
            "ops_sla_status": "SLA Status",
            "first_found":    "First Found",
        })
        write_dataframe_to_sheet(
            wb, overdue_display,
            sheet_name="Overdue Detail",
            title_row=f"Overdue Findings — {tag_filter_str}",
            severity_col="Severity",
        )
        ws_overdue = wb["Overdue Detail"]
        for cell in ws_overdue[2]:
            if cell.value == "SLA Status":
                cell.value = "sla_status"
                break
        _apply_ops_sla_formatting(ws_overdue)
        for cell in ws_overdue[2]:
            if cell.value == "sla_status":
                cell.value = "SLA Status"
                break
    else:
        ws_empty = wb.create_sheet(title="Overdue Detail")
        ws_empty.cell(row=1, column=1, value="No overdue findings for this scope.")

    # --- Tab 4: Unscanned Assets ---
    if not unscanned_df.empty:
        unscanned_excel = unscanned_df[[
            c for c in [
                "hostname", "ipv4", "mac_address",
                "last_seen", "last_licensed_scan_date",
                "days_since_scan", "source_name",
            ] if c in unscanned_df.columns
        ]].copy()
        # Format raw datetime columns as YYYY-MM-DD strings
        for _dt_col in ("last_seen", "last_licensed_scan_date"):
            if _dt_col in unscanned_excel.columns:
                unscanned_excel[_dt_col] = pd.to_datetime(
                    unscanned_excel[_dt_col], utc=True, errors="coerce"
                ).apply(
                    lambda ts: ts.strftime("%Y-%m-%d") if pd.notna(ts) else "Never"
                )
        unscanned_excel = unscanned_excel.rename(columns={
            "hostname":                "Hostname",
            "ipv4":                    "IP Address",
            "mac_address":             "MAC Address",
            "last_seen":               "Last Seen",
            "last_licensed_scan_date": "Last Licensed Scan",
            "days_since_scan":         "Days Since Last Seen",
            "source_name":             "Source",
        })
        write_dataframe_to_sheet(
            wb, unscanned_excel,
            sheet_name="Unscanned Assets",
            title_row=f"Unscanned / Stale Assets (>30 days) — {tag_filter_str}",
            severity_col=None,
        )
    else:
        ws_empty2 = wb.create_sheet(title="Unscanned Assets")
        ws_empty2.cell(row=1, column=1, value="No unscanned assets for this scope.")

    # --- Tab 5: Risk Acceptances & Recasts ---
    _build_risk_mods_sheet(wb, risk_mods_df, tag_filter_str)

    # --- Tab 6: Recurring Vulnerabilities ---
    _build_recurring_sheet(wb, recurring_df, tag_filter_str)

    # --- Tab 7: Report Info ---
    from datetime import datetime, timezone
    generated_at = datetime.now(tz=timezone.utc)
    write_metadata_tab(wb, REPORT_NAME, tag_filter_str, generated_at)
    _extend_metadata_tab(wb, risk_mods_df, recurring_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("[%s] Excel workbook saved: %s", REPORT_NAME, output_path)
    return output_path


def _kpi_html(summary: dict) -> str:
    """
    Render the summary dict as an inline-CSS HTML block suitable for embedding
    in a PDF section.  Produces a two-column tile grid (label | value) followed
    by a top-5 plugins mini-table.
    """
    accent      = "#1F3864"
    text        = "#212121"
    muted       = "#757575"
    border      = "#DDDFE2"
    white       = "#FFFFFF"
    alt         = "#E8EAF6"

    sla_colors = {
        OPS_SLA_OVERDUE:  ("#FFCDD2", "#B71C1C"),
        OPS_SLA_URGENT:   ("#FFE0B2", "#E65100"),
        OPS_SLA_WARNING:  ("#FFF9C4", "#F57F17"),
        OPS_SLA_ON_TRACK: ("#C8E6C9", "#1B5E20"),
    }
    sev_bg = {"critical": "#FFCDD2", "high": "#FFE0B2", "medium": "#FFF9C4", "low": "#C8E6C9"}
    sev_fg = {"critical": "#B71C1C", "high": "#E65100", "medium": "#F57F17", "low": "#1B5E20"}

    td_l = (
        f"padding: 5px 10px; font-family: Arial, sans-serif; font-size: 9pt; "
        f"font-weight: bold; color: {muted}; border: 1px solid {border}; "
        f"background-color: {alt}; width: 56%;"
    )
    td_r = (
        f"padding: 5px 10px; font-family: Arial, sans-serif; font-size: 10pt; "
        f"font-weight: bold; color: {text}; border: 1px solid {border}; "
        f"background-color: {white}; width: 37%; text-align: right;"
    )

    compliance_pct = round(summary.get("sla_compliance_rate", 0) * 100, 1)

    # SLA state rows with per-state background colour
    def _state_row(label: str, key: str) -> str:
        bg, fg = sla_colors.get(label, (white, text))
        val = summary.get(key, 0)
        return (
            f"<tr>"
            f'<td style="{td_l} background-color: {bg}; color: {fg};">{label}</td>'
            f'<td style="{td_r} background-color: {bg}; color: {fg};">{val:,}</td>'
            f"</tr>\n"
        )

    # Severity count rows
    def _sev_row(sev: str) -> str:
        bg = sev_bg.get(sev, white)
        fg = sev_fg.get(sev, text)
        val = summary.get(f"open_{sev}", 0)
        return (
            f"<tr>"
            f'<td style="{td_l} background-color: {bg}; color: {fg};">'
            f"Open {sev.title()}</td>"
            f'<td style="{td_r} background-color: {bg}; color: {fg};">{val:,}</td>'
            f"</tr>\n"
        )

    rows = (
        f"<tr><td style='{td_l}'>Generated (UTC)</td>"
        f"<td style='{td_r}'>{summary.get('generated_at', '')}</td></tr>\n"
        f"<tr><td style='{td_l}'>Scope</td>"
        f"<td style='{td_r}'>{summary.get('tag_filter', 'All Assets')}</td></tr>\n"
        f"<tr><td style='{td_l}'>Total Assets in Scope</td>"
        f"<td style='{td_r}'>{summary.get('total_assets', 0):,}</td></tr>\n"
        f"<tr><td style='{td_l}'>Unscanned / Stale Assets (&gt;30d)</td>"
        f"<td style='{td_r}'>{summary.get('total_unscanned', 0):,}</td></tr>\n"
    )
    for sev in ("critical", "high", "medium", "low"):
        rows += _sev_row(sev)
    rows += _state_row(OPS_SLA_OVERDUE,  "count_overdue")
    rows += _state_row(OPS_SLA_URGENT,   "count_urgent")
    rows += _state_row(OPS_SLA_WARNING,  "count_warning")
    rows += _state_row(OPS_SLA_ON_TRACK, "count_on_track")
    rows += (
        f"<tr><td style='{td_l}'>Unique Plugins with Overdue Findings</td>"
        f"<td style='{td_r}'>{summary.get('plugins_with_overdue', 0):,}</td></tr>\n"
        f"<tr><td style='{td_l}'>SLA Compliance Rate (On Track)</td>"
        f"<td style='{td_r}'>{compliance_pct}%</td></tr>\n"
    )

    # Exploitability section rows
    _navy_hdr = (
        f"padding: 5px 10px; font-family: Arial, sans-serif; font-size: 9pt; "
        f"font-weight: bold; color: #FFFFFF; background-color: #1F3864; "
        f"border: 1px solid {border};"
    )
    rows += (
        f"<tr>"
        f'<td colspan="2" style="{_navy_hdr}">Exploitability</td>'
        f"</tr>\n"
    )

    def _exploit_row(label: str, val: int, val_color: str) -> str:
        v_style = (
            f"padding: 5px 10px; font-family: Arial, sans-serif; font-size: 10pt; "
            f"font-weight: bold; color: {val_color}; border: 1px solid {border}; "
            f"background-color: {white}; width: 37%; text-align: right;"
        )
        return (
            f"<tr>"
            f'<td style="{td_l}">{label}</td>'
            f'<td style="{v_style}">{val:,}</td>'
            f"</tr>\n"
        )

    known = summary.get("known_exploit", 0)
    func  = summary.get("exploit_functional", 0)
    high  = summary.get("exploit_high", 0)
    rows += _exploit_row(
        "Plugins with Known Exploit",
        known,
        "#FF6600" if known > 0 else "#212121",
    )
    rows += _exploit_row(
        "Exploit Maturity — Functional",
        func,
        "#C00000" if func > 0 else "#212121",
    )
    rows += _exploit_row(
        "Exploit Maturity — High",
        high,
        "#FF6600" if high > 0 else "#212121",
    )

    # Risk Management section
    rows += (
        f"<tr>"
        f'<td colspan="2" style="{_navy_hdr}">Risk Management</td>'
        f"</tr>\n"
    )

    def _risk_row(label: str, val: int, val_color: str) -> str:
        v_style = (
            f"padding: 5px 10px; font-family: Arial, sans-serif; font-size: 10pt; "
            f"font-weight: bold; color: {val_color}; border: 1px solid {border}; "
            f"background-color: {white}; width: 37%; text-align: right;"
        )
        return (
            f"<tr>"
            f'<td style="{td_l}">{label}</td>'
            f'<td style="{v_style}">{val:,}</td>'
            f"</tr>\n"
        )

    accepted  = summary.get("count_risk_accepted", 0)
    recast    = summary.get("count_risk_recast", 0)
    expiring  = summary.get("count_expiring_soon", 0)
    expired   = summary.get("count_expired", 0)
    recurring = summary.get("count_recurring", 0)

    rows += _risk_row("Accepted Findings",             accepted,  "#FF6600" if accepted  > 0 else "#212121")
    rows += _risk_row("Recast Findings",               recast,    "#FF6600" if recast    > 0 else "#212121")
    rows += _risk_row("Rules Expiring Within 30 Days", expiring,  "#FF6600" if expiring  > 0 else "#212121")
    rows += _risk_row("Expired Rules",                 expired,   "#C00000" if expired   > 0 else "#212121")
    rows += _risk_row("Recurring Vulnerabilities",     recurring, "#FF6600" if recurring > 0 else "#212121")

    kpi_table = (
        f'<table style="border-collapse: collapse; width: 100%; margin: 8px 0;">'
        f"{rows}</table>"
    )

    # Top 5 Priority Plugins mini-table
    priority = summary.get("priority_plugins", [])
    priority_html = ""
    th = (
        f"font-family: Arial, sans-serif; font-size: 9pt; font-weight: bold; "
        f"color: {white}; background-color: {accent}; "
        f"border: 1px solid {border}; padding: 5px 8px; text-align: left;"
    )
    td_cell = (
        f"font-family: Arial, sans-serif; font-size: 9pt; color: {text}; "
        f"border: 1px solid {border}; padding: 4px 8px; "
        f"word-wrap: break-word; overflow-wrap: break-word; vertical-align: top;"
    )
    priority_html += (
        f'<p style="font-weight: bold; color: {accent}; margin: 14px 0 4px 0; '
        f'page-break-before: always;">'
        f"Top 5 Priority Plugins (Functional &amp; High Exploit Maturity)</p>"
    )
    if priority:
        t_rows = ""
        for i, p in enumerate(priority):
            bg = alt if i % 2 == 0 else white
            vpr_str = (
                f"{float(p.get('vpr_score', 0)):.1f}"
                if p.get("vpr_score") is not None else "N/A"
            )
            mat_display = (
                "Functional" if p.get("exploit_code_maturity") == "FUNCTIONAL"
                else "High" if p.get("exploit_code_maturity") == "HIGH"
                else p.get("exploit_code_maturity", "")
            )
            name = p.get("plugin_name", "")
            if len(name) > 55:
                name = name[:52] + "..."
            t_rows += (
                f'<tr style="background-color: {bg};">'
                f'<td style="{td_cell}">{name}</td>'
                f'<td style="{td_cell}; text-align: right;">{vpr_str}</td>'
                f'<td style="{td_cell}">{mat_display}</td>'
                f'<td style="{td_cell}; text-align: right;">'
                f'{p.get("affected_assets", 0):,}</td>'
                f"</tr>\n"
            )
        n_shown = len(priority)
        note = (
            f'<p style="font-family: Arial, sans-serif; font-size: 8pt; '
            f'color: {muted}; margin: 2px 0 0 0;">'
            f"Showing {n_shown} of {n_shown} qualifying plugins.</p>"
            if n_shown < 5 else ""
        )
        priority_html += (
            f'<table style="width: 100%; table-layout: fixed; '
            f'border-collapse: collapse; margin: 4px 0;">'
            f"<thead><tr>"
            f'<th style="{th}">Plugin Name</th>'
            f'<th style="{th}; text-align: right;">VPR Score</th>'
            f'<th style="{th}">Exploit Maturity</th>'
            f'<th style="{th}; text-align: right;">Affected Assets</th>'
            f"</tr></thead>"
            f"<tbody>{t_rows}</tbody></table>{note}"
        )
    else:
        priority_html += (
            f'<p style="font-family: Arial, sans-serif; font-size: 9pt; '
            f'color: {muted}; font-style: italic;">'
            f"No plugins with Functional or High exploit maturity found "
            f"in this reporting period.</p>"
        )

    return kpi_table + priority_html


def _truncate_for_pdf(df: pd.DataFrame) -> pd.DataFrame:
    """
    Truncate long text fields for PDF rendering only.

    Keeps plugin names and hostnames within the fixed column widths used by the
    WeasyPrint tables.  Never call this before build_excel() — Excel must always
    receive full untruncated values.
    """
    df = df.copy()
    if "Plugin Name" in df.columns:
        df["Plugin Name"] = df["Plugin Name"].apply(
            lambda x: (x[:47] + "...") if isinstance(x, str) and len(x) > 50 else (x or "")
        )
    if "Hostname" in df.columns:
        df["Hostname"] = df["Hostname"].apply(
            lambda x: (x[:37] + "...") if isinstance(x, str) and len(x) > 40 else (x or "")
        )
    return df


# Column width maps for WeasyPrint table-layout: fixed.
# Keys must match the renamed (display) column headers passed to build_pdf().
# Percentages sum to ≈80% (9-col), ≈82% (8-col), ≈90% (4-col) to prevent
# WeasyPrint overflow: th padding (5px 7px) and border are content-box additions
# on top of the percentage width.  Formula: available = (680px - N×14px - (N+1)px) / 680px.
_PDF_COL_WIDTHS_VULN = {
    "Asset Name":      "15%",
    "IP Address":      "11%",
    "Plugin ID":        "6%",
    "Plugin Name":     "20%",
    "Severity":         "7%",
    "VPR Score":        "6%",
    "Days Open":        "6%",
    "Days Overdue":     "5%",   # overdue table: days_remaining (negative = days past SLA)
    "Days Remaining":   "5%",   # urgent table:  days_remaining (positive = days left)
}
# Active 8 cols per table sum = 76% (well within ~82% max for 8-col with 5px 7px th padding)

_PDF_COL_WIDTHS_PLUGIN = {
    "Severity":           "7%",
    "Plugin ID":          "6%",
    "Plugin Name":       "19%",
    "VPR Score":          "6%",
    "Affected Assets":    "8%",
    "Days Open (Oldest)": "8%",
    "Days Open (Newest)": "8%",
    "CVEs":              "12%",
    "Exploit Available":  "6%",
}
# 9 cols sum = 80%

_PDF_COL_WIDTHS_UNSCANNED = {
    "Hostname":            "19%",
    "IP Address":          "12%",
    "Last Seen":           "14%",
    "Last Licensed Scan":  "14%",
    "Days Since Scan":     "12%",
    "Source":              "15%",
}
# 6 cols sum = 86% (max ~87% for 6-col with 5px 7px th padding)


def _build_pdf(
    plugin_df: pd.DataFrame,
    overdue_df: pd.DataFrame,
    urgent_df: pd.DataFrame,
    unscanned_df: pd.DataFrame,
    summary: dict,
    output_path: Path,
    tag_category: Optional[str],
    tag_value: Optional[str],
) -> Path:
    """
    Build ops_remediation.pdf via WeasyPrint with five sections:

    1. Executive Summary  — KPI tiles from the summary dict
    2. Overdue Findings   — overdue_df (top 200, sorted severity → days_open desc)
    3. Urgent Findings    — urgent_df (<=25% SLA remaining, sorted severity → days_remaining)
    4. Plugin Summary     — plugin_df (all plugins, same sort as _group_by_plugin)
    5. Unscanned Assets   — unscanned_df (only if non-empty)

    The cover page (generated by build_pdf) shows report title, scope banner,
    SLA definitions, and VPR score ranges.

    Parameters
    ----------
    plugin_df : pd.DataFrame
        Output of ``_group_by_plugin()``.
    overdue_df : pd.DataFrame
        Vuln-level rows with ``ops_sla_status == OPS_SLA_OVERDUE``.
    urgent_df : pd.DataFrame
        Vuln-level rows with ``ops_sla_status == OPS_SLA_URGENT``.
    unscanned_df : pd.DataFrame
        Output of ``_identify_unscanned_assets()``.
    summary : dict
        Output of ``_compute_summary_metrics()``.
    output_path : Path
        Destination .pdf file path.
    tag_category, tag_value : str or None
        Tag filter applied to the report.

    Returns
    -------
    Path
        Absolute path of the written PDF.
    """
    from exporters.pdf_exporter import build_pdf

    tag_filter_str = (
        f"{tag_category} = {tag_value}"
        if tag_category and tag_value
        else "All Assets"
    )

    # ------------------------------------------------------------------
    # Section 1: Executive Summary KPIs (rendered as styled HTML, no table)
    # ------------------------------------------------------------------
    sections: list[dict] = [
        {
            "heading":  "Executive Summary",
            "text":     _kpi_html(summary),
        },
    ]

    # ------------------------------------------------------------------
    # Section 2: Overdue Findings (vuln-level, top 200)
    # ------------------------------------------------------------------
    overdue_count = len(overdue_df)
    overdue_cols  = [c for c in [
        "hostname", "ipv4", "plugin_id", "plugin_name",
        "severity", "vpr_score", "days_open", "days_remaining",
    ] if c in overdue_df.columns]

    overdue_display = overdue_df[overdue_cols].head(200).copy()
    overdue_display = overdue_display.rename(columns={
        "hostname":       "Asset Name",
        "ipv4":           "IP Address",
        "plugin_id":      "Plugin ID",
        "plugin_name":    "Plugin Name",
        "severity":       "Severity",
        "vpr_score":      "VPR Score",
        "days_open":      "Days Open",
        "days_remaining": "Days Overdue",
    })

    overdue_note = ""
    if overdue_count > 200:
        overdue_note = (
            f"Showing top 200 of {overdue_count:,} overdue findings. "
            f"Full detail available in the Excel attachment."
        )

    sections.append({
        "heading":      f"Overdue Findings ({overdue_count:,} total)",
        "text":         overdue_note or None,
        "dataframe":    _truncate_for_pdf(overdue_display) if not overdue_df.empty else None,
        "severity_col": "Severity",
        "col_widths":   _PDF_COL_WIDTHS_VULN,
    })

    # ------------------------------------------------------------------
    # Section 3: Urgent Findings (<=25% SLA remaining)
    # ------------------------------------------------------------------
    urgent_count = len(urgent_df)
    if not urgent_df.empty:
        urgent_cols = [c for c in [
            "hostname", "ipv4", "plugin_id", "plugin_name",
            "severity", "vpr_score", "days_open", "days_remaining",
        ] if c in urgent_df.columns]

        urgent_display = urgent_df[urgent_cols].head(200).copy()
        urgent_display = urgent_display.rename(columns={
            "hostname":       "Asset Name",
            "ipv4":           "IP Address",
            "plugin_id":      "Plugin ID",
            "plugin_name":    "Plugin Name",
            "severity":       "Severity",
            "vpr_score":      "VPR Score",
            "days_open":      "Days Open",
            "days_remaining": "Days Remaining",
        })
        sections.append({
            "heading":      f"Urgent Findings — <=25% SLA Remaining ({urgent_count:,} total)",
            "text":         (
                f"These {urgent_count:,} findings are not yet overdue but have fewer than "
                f"25% of their SLA window remaining. Prioritize alongside overdue items."
            ),
            "dataframe":    _truncate_for_pdf(urgent_display),
            "severity_col": "Severity",
            "col_widths":   _PDF_COL_WIDTHS_VULN,
        })

    # ------------------------------------------------------------------
    # Section 4: Plugin Summary
    # ------------------------------------------------------------------
    plugin_count = len(plugin_df)
    plugin_cols  = [c for c in [
        "severity", "plugin_id", "plugin_name",
        "vpr_score", "affected_asset_count",
        "days_open_oldest", "days_open_newest",
        "cves", "exploit_available",
    ] if c in plugin_df.columns]

    plugin_display = plugin_df[plugin_cols].head(200).copy()
    if "vpr_score" in plugin_display.columns:
        plugin_display["vpr_score"] = plugin_display["vpr_score"].apply(
            lambda v: f"{float(v):.1f}" if pd.notna(v) else ""
        )
    plugin_display = plugin_display.rename(columns={
        "severity":            "Severity",
        "plugin_id":           "Plugin ID",
        "plugin_name":         "Plugin Name",
        "vpr_score":           "VPR Score",
        "affected_asset_count":"Affected Assets",
        "days_open_oldest":    "Days Open (Oldest)",
        "days_open_newest":    "Days Open (Newest)",
        "cves":                "CVEs",
        "exploit_available":   "Exploit Available",
    })

    plugin_note = ""
    if plugin_count > 200:
        plugin_note = (
            f"Showing top 200 of {plugin_count:,} plugins (sorted by severity, VPR score, "
            f"affected asset count). Full list in the Excel attachment."
        )

    sections.append({
        "heading":      f"Plugin Summary ({plugin_count:,} unique plugins)",
        "text":         plugin_note or None,
        "dataframe":    _truncate_for_pdf(plugin_display),
        "severity_col": "Severity",
        "col_widths":   _PDF_COL_WIDTHS_PLUGIN,
    })

    # ------------------------------------------------------------------
    # Section 5: Unscanned Assets (omit section if empty)
    # ------------------------------------------------------------------
    if not unscanned_df.empty:
        unscanned_cols = [c for c in [
            "hostname", "ipv4", "last_seen", "last_licensed_scan_date",
            "days_since_scan", "source_name",
        ] if c in unscanned_df.columns]

        unscanned_display = unscanned_df[unscanned_cols].copy()
        for _dt_col in ("last_seen", "last_licensed_scan_date"):
            if _dt_col in unscanned_display.columns:
                unscanned_display[_dt_col] = pd.to_datetime(
                    unscanned_display[_dt_col], utc=True, errors="coerce"
                ).apply(lambda ts: ts.strftime("%Y-%m-%d") if pd.notna(ts) else "—")
        unscanned_display = unscanned_display.rename(columns={
            "hostname":                "Hostname",
            "ipv4":                    "IP Address",
            "last_seen":               "Last Seen",
            "last_licensed_scan_date": "Last Licensed Scan",
            "days_since_scan":         "Days Since Scan",
            "source_name":             "Source",
        })

        sections.append({
            "heading":      f"Unscanned / Stale Assets ({len(unscanned_df):,} assets)",
            "text":         (
                "Assets below have not been scanned in the last 30 days. "
                "Vulnerability counts for these assets may be incomplete."
            ),
            "dataframe":    _truncate_for_pdf(unscanned_display),
            "severity_col": None,
            "col_widths":   _PDF_COL_WIDTHS_UNSCANNED,
        })

    # ------------------------------------------------------------------
    # Render
    # ------------------------------------------------------------------
    output_path = Path(output_path)
    result = build_pdf(
        report_title=REPORT_NAME,
        scope_str=tag_filter_str,
        sections=sections,
        output_path=output_path,
    )
    logger.info("[%s] PDF written: %s", REPORT_NAME, result)
    return result


def _build_email_summary(
    summary: dict,
    tag_category: Optional[str],
    tag_value: Optional[str],
) -> dict:
    """
    Build the ``metrics`` sub-dict that goes into the ``report_outputs`` entry
    for this report.  ``build_kpi_metrics()`` in ``delivery/email_template.py``
    reads this to construct the four KPI tiles shown in the email body.

    The returned dict has two top-level sections:

    ``kpi_tiles``
        Pre-built list of tile dicts ``{label, value, color, sub_label?}``
        consumed directly by the email template renderer.  Avoids duplicating
        the extraction logic that other reports embed in their own ``metrics``
        structures.

    ``raw``
        Flat copy of the summary dict for any downstream consumer that wants
        the underlying numbers (e.g. dashboards, audit scripts).

    Parameters
    ----------
    summary : dict
        Output of ``_compute_summary_metrics()``.
    tag_category, tag_value : str or None
        Tag filter — included in ``raw`` for context.

    Returns
    -------
    dict
        ``{kpi_tiles: list[dict], raw: dict}``
    """
    _RED    = "#d32f2f"
    _ORANGE = "#f57c00"
    _AMBER  = "#fbc02d"
    _GREEN  = "#388e3c"
    _GREY   = "#757575"

    # ------------------------------------------------------------------
    # Tile 1: Open Criticals
    # ------------------------------------------------------------------
    open_crit = summary.get("open_critical", 0)
    tile_crit = {
        "label": "Open Criticals",
        "value": f"{open_crit:,}",
        "color": _RED if open_crit > 0 else _GREEN,
    }

    # ------------------------------------------------------------------
    # Tile 2: Overdue Findings (all severities)
    # ------------------------------------------------------------------
    count_overdue = summary.get("count_overdue", 0)
    tile_overdue = {
        "label": "Overdue Findings",
        "value": f"{count_overdue:,}",
        "color": _RED if count_overdue > 0 else _GREEN,
        "sub_label": "across all severities",
    }

    # ------------------------------------------------------------------
    # Tile 3: SLA Compliance Rate
    # ------------------------------------------------------------------
    rate = summary.get("sla_compliance_rate", 0.0)
    rate_pct = round(rate * 100, 1)
    if rate >= 0.80:
        compliance_color = _GREEN
    elif rate >= 0.60:
        compliance_color = _AMBER
    else:
        compliance_color = _RED
    tile_sla = {
        "label":     "SLA Compliance",
        "value":     f"{rate_pct}%",
        "color":     compliance_color,
        "sub_label": "findings on track",
    }

    # ------------------------------------------------------------------
    # Tile 4: Exploitable Plugins
    # ------------------------------------------------------------------
    exploitable = summary.get("exploitable_plugins", 0)
    tile_exploit = {
        "label": "Exploitable Plugins",
        "value": f"{exploitable:,}",
        "color": _RED if exploitable > 0 else _GREEN,
        "sub_label": "with known exploit",
    }

    # ------------------------------------------------------------------
    # Tile 5: Recurring Vulnerabilities
    # ------------------------------------------------------------------
    recurring = summary.get("count_recurring", 0)
    tile_recurring = {
        "label": "Recurring Findings",
        "value": f"{recurring:,}",
        "color": _ORANGE if recurring > 0 else _GREEN,
        "sub_label": "resurfaced after prior fix",
    }

    kpi_tiles = [tile_crit, tile_overdue, tile_sla, tile_exploit, tile_recurring]

    # ------------------------------------------------------------------
    # Plain-language narrative sentences for risk management context
    # ------------------------------------------------------------------
    accepted = summary.get("count_risk_accepted", 0)
    recast   = summary.get("count_risk_recast", 0)
    expiring = summary.get("count_expiring_soon", 0)
    expired  = summary.get("count_expired", 0)

    narrative_sentences: list[str] = []

    if recurring > 0:
        narrative_sentences.append(
            f"{recurring:,} finding{'s' if recurring != 1 else ''} have resurfaced after a prior remediation — "
            f"review the Recurring Vulnerabilities tab for systemic patching gaps."
        )
    if expired > 0:
        narrative_sentences.append(
            f"{expired:,} risk acceptance or recast rule{'s have' if expired != 1 else ' has'} passed "
            f"its expiration date and should be reviewed or renewed."
        )
    if expiring > 0:
        narrative_sentences.append(
            f"{expiring:,} risk rule{'s are' if expiring != 1 else ' is'} expiring within the next 30 days — "
            f"confirm whether renewal or remediation is required."
        )
    if accepted > 0 or recast > 0:
        parts = []
        if accepted > 0:
            parts.append(f"{accepted:,} accepted")
        if recast > 0:
            parts.append(f"{recast:,} recast")
        narrative_sentences.append(
            f"{' and '.join(parts)} finding{'s are' if (accepted + recast) != 1 else ' is'} currently "
            f"suppressed or severity-modified — see the Risk Acceptances & Recasts tab for details."
        )

    return {
        "kpi_tiles": kpi_tiles,
        "raw": {
            "tag_filter":           summary.get("tag_filter", "All Assets"),
            "generated_at":         summary.get("generated_at", ""),
            "total_assets":         summary.get("total_assets", 0),
            "total_unscanned":      summary.get("total_unscanned", 0),
            "open_critical":        summary.get("open_critical", 0),
            "open_high":            summary.get("open_high", 0),
            "open_medium":          summary.get("open_medium", 0),
            "open_low":             summary.get("open_low", 0),
            "count_overdue":        summary.get("count_overdue", 0),
            "count_urgent":         summary.get("count_urgent", 0),
            "count_warning":        summary.get("count_warning", 0),
            "count_on_track":       summary.get("count_on_track", 0),
            "plugins_with_overdue": summary.get("plugins_with_overdue", 0),
            "exploitable_plugins":  summary.get("exploitable_plugins", 0),
            "sla_compliance_rate":  summary.get("sla_compliance_rate", 0.0),
            "count_risk_accepted":  summary.get("count_risk_accepted", 0),
            "count_risk_recast":    summary.get("count_risk_recast", 0),
            "count_expiring_soon":  summary.get("count_expiring_soon", 0),
            "count_expired":        summary.get("count_expired", 0),
            "count_recurring":      summary.get("count_recurring", 0),
            "narrative_sentences":  narrative_sentences,
        },
    }


def run_report(
    tio,
    run_id: str,
    tag_category: Optional[str] = None,
    tag_value: Optional[str] = None,
    output_dir: Optional[Path] = None,
    generated_at: Optional[datetime] = None,
    cache_dir: Optional[Path] = None,
) -> dict:
    """
    Main entry point called by ``run_all.py`` and ``scheduler.py``.

    Signature matches the convention used by all other report modules so
    ``run_all.py`` can call it uniformly.

    Parameters
    ----------
    tio : TenableIO
        Authenticated Tenable client.
    run_id : str
        Date string (``YYYY-MM-DD``) used to name the default output directory
        when *output_dir* is not provided.
    tag_category, tag_value : str, optional
        Tag filter for asset scoping.
    output_dir : Path, optional
        Where to write Excel and PDF files.  Created if it does not exist.
        Defaults to ``OUTPUT_DIR / run_id / "ops_remediation"``.
    generated_at : datetime, optional
        Report timestamp.  Defaults to UTC now.
    cache_dir : Path, optional
        Parquet cache directory shared across the batch run.
        Defaults to ``CACHE_DIR / run_id``.

    Returns
    -------
    dict
        ``{"pdf": Path, "excel": Path, "charts": [], "metrics": dict}``
        compatible with ``run_all.py`` / ``email_sender.py``.
    """
    if generated_at is None:
        generated_at = datetime.now(tz=timezone.utc)
    if cache_dir is None:
        cache_dir = CACHE_DIR / (run_id or generated_at.strftime("%Y-%m-%d"))
    if output_dir is None:
        output_dir = OUTPUT_DIR / (run_id or generated_at.strftime("%Y-%m-%d")) / REPORT_SLUG
    output_dir = Path(output_dir)
    cache_dir  = Path(cache_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)

    logger.info(
        "[%s] Starting | filter=%s=%s | output=%s",
        REPORT_NAME,
        tag_category or "*",
        tag_value    or "*",
        output_dir,
    )

    # ------------------------------------------------------------------
    # Step 1: Fetch and prepare data
    # ------------------------------------------------------------------
    vulns_df, assets_df = _fetch_and_prepare(
        tio          = tio,
        cache_dir    = cache_dir,
        tag_category = tag_category,
        tag_value    = tag_value,
        as_of        = generated_at,
    )

    recast_rules_df = fetch_recast_rules(tio, cache_dir)

    scanned_df, unscanned_df = _identify_unscanned_assets(assets_df, as_of=generated_at)

    scanned_ids   = set(scanned_df["asset_uuid"])
    vulns_scanned = vulns_df[vulns_df["asset_uuid"].isin(scanned_ids)]

    plugin_df = _group_by_plugin(vulns_scanned)

    summary = _compute_summary_metrics(
        vulns_df     = vulns_scanned,
        plugin_df    = plugin_df,
        assets_df    = assets_df,
        unscanned_df = unscanned_df,
        tag_category = tag_category,
        tag_value    = tag_value,
        as_of        = generated_at,
    )

    _exploit_metrics = _compute_exploitability_metrics(vulns_scanned)
    summary["known_exploit"]      = _exploit_metrics["known_exploit"]
    summary["exploit_functional"] = _exploit_metrics["functional"]
    summary["exploit_high"]       = _exploit_metrics["high_maturity"]

    _priority_df = _get_top_priority_plugins(vulns_scanned)
    summary["priority_plugins"] = _priority_df.to_dict("records")

    risk_mods_df  = _extract_risk_modifications(
        vulns_df        = vulns_scanned,
        assets_df       = assets_df,
        recast_rules_df = recast_rules_df,
        as_of           = generated_at,
    )
    recurring_df = _extract_recurring_vulnerabilities(
        vulns_df  = vulns_scanned,
        assets_df = assets_df,
    )
    summary["count_risk_accepted"]  = int((risk_mods_df["Modification Type"] == "Accepted").sum()) if not risk_mods_df.empty else 0
    summary["count_risk_recast"]    = int((risk_mods_df["Modification Type"] == "Recast").sum())   if not risk_mods_df.empty else 0
    summary["count_expiring_soon"]  = int(
        risk_mods_df["Days Until Expiry"].apply(
            lambda x: isinstance(x, int) and 0 <= x <= 30
        ).sum()
    ) if not risk_mods_df.empty else 0
    summary["count_expired"]        = int(
        risk_mods_df["Days Until Expiry"].apply(
            lambda x: isinstance(x, int) and x < 0
        ).sum()
    ) if not risk_mods_df.empty else 0
    summary["count_recurring"]      = len(recurring_df)

    # ------------------------------------------------------------------
    # Step 2: Build Excel
    # ------------------------------------------------------------------
    slug_str   = safe_filename(f"{tag_category or 'all'}_{tag_value or 'assets'}")
    excel_path = output_dir / f"ops_remediation_{slug_str}.xlsx"

    # Enrich overdue/urgent with authoritative asset columns from the asset export.
    # Even though enrich_vulns_with_assets already joined these, we re-join here
    # to guarantee hostname, mac_address, ipv4 reflect the asset export's values.
    _asset_lookup = (
        assets_df[["asset_uuid", "hostname", "mac_address", "ipv4"]]
        .drop_duplicates("asset_uuid")
        if all(c in assets_df.columns for c in ["asset_uuid", "hostname", "mac_address", "ipv4"])
        else None
    )

    def _enrich_with_assets(df: pd.DataFrame) -> pd.DataFrame:
        if _asset_lookup is None or df.empty:
            return df
        drop_cols = [c for c in ["hostname", "mac_address", "ipv4"] if c in df.columns]
        return (
            df.drop(columns=drop_cols)
            .merge(_asset_lookup, on="asset_uuid", how="left")
        )

    overdue_df = _enrich_with_assets(
        vulns_scanned[vulns_scanned["ops_sla_status"] == OPS_SLA_OVERDUE]
    ).sort_values(
        ["severity", "days_open"],
        ascending=[True, False],
        na_position="last",
        key=lambda s: s.map(SEVERITY_RANK) if s.name == "severity" else s,
    )

    _build_excel(
        plugin_df    = plugin_df,
        overdue_df   = overdue_df,
        unscanned_df = unscanned_df,
        summary      = summary,
        output_path  = excel_path,
        tag_category = tag_category,
        tag_value    = tag_value,
        risk_mods_df = risk_mods_df,
        recurring_df = recurring_df,
    )

    # ------------------------------------------------------------------
    # Step 3: Build PDF
    # ------------------------------------------------------------------
    pdf_path = output_dir / f"ops_remediation_{slug_str}.pdf"

    urgent_df = _enrich_with_assets(
        vulns_scanned[vulns_scanned["ops_sla_status"] == OPS_SLA_URGENT]
    ).sort_values(
        ["severity", "days_remaining"],
        ascending=[True, True],
        na_position="last",
        key=lambda s: s.map(SEVERITY_RANK) if s.name == "severity" else s,
    )

    _build_pdf(
        plugin_df    = plugin_df,
        overdue_df   = overdue_df,
        urgent_df    = urgent_df,
        unscanned_df = unscanned_df,
        summary      = summary,
        output_path  = pdf_path,
        tag_category = tag_category,
        tag_value    = tag_value,
    )

    # ------------------------------------------------------------------
    # Step 4: Build email metrics
    # ------------------------------------------------------------------
    email_metrics = _build_email_summary(
        summary      = summary,
        tag_category = tag_category,
        tag_value    = tag_value,
    )

    logger.info(
        "[%s] Complete | excel=%s | pdf=%s",
        REPORT_NAME, excel_path.name, pdf_path.name,
    )

    return {
        "pdf":     pdf_path,
        "excel":   excel_path,
        "charts":  [],
        "metrics": email_metrics,
    }


# ===========================================================================
# CLI entry point
# ===========================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Operations Remediation Report — prioritized, tag-scoped vuln list for IT ops teams.",
    )
    parser.add_argument("--tag-category", default=None, help='Tenable tag category (e.g. "Operations")')
    parser.add_argument("--tag-value",    default=None, help='Tenable tag value    (e.g. "Server Operations")')
    parser.add_argument("--output-dir",   default=None, help="Output directory (default: output/)")
    parser.add_argument("--cache-dir",    default=None, help="Parquet cache directory (default: data/cache/<today>)")
    parser.add_argument("--no-email",     action="store_true", help="Generate reports but skip email delivery")
    args = parser.parse_args()

    _cache_dir = (
        Path(args.cache_dir)
        if args.cache_dir
        else CACHE_DIR / datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")
    )
    _output_dir = Path(args.output_dir) if args.output_dir else OUTPUT_DIR

    from tenable_client import get_client
    _tio = get_client()

    _as_of = datetime.now(tz=timezone.utc)

    # -----------------------------------------------------------------------
    # Step 1 smoke-test — exercises all data prep functions and prints a
    # summary.  run_report() will replace this block once Steps 2-5 are done.
    # -----------------------------------------------------------------------

    # Pre-filter cache inspection: load raw parquets and show tag samples
    # so we can verify the tag column format before any filtering is applied.
    _raw_assets_cache = _cache_dir / "assets_all.parquet"
    _raw_vulns_cache  = _cache_dir / "vulns_all.parquet"
    print("\n=== Cache Inspection (pre-filter) ===")
    if _raw_assets_cache.exists():
        _raw_assets = pd.read_parquet(_raw_assets_cache, engine="fastparquet")
        _non_empty_asset_tags = _raw_assets["tags"].fillna("").loc[lambda s: s != ""]
        print(f"  assets_all rows        : {len(_raw_assets)}")
        print(f"  assets with tags       : {len(_non_empty_asset_tags)}/{len(_raw_assets)}")
        print(f"  sample tag values      : {_non_empty_asset_tags.head(5).tolist()}")
    else:
        print(f"  assets_all.parquet not found at {_raw_assets_cache}")
    if _raw_vulns_cache.exists():
        _raw_vulns = pd.read_parquet(_raw_vulns_cache, engine="fastparquet")
        _vulns_with_tags = _raw_vulns["tags"].fillna("").loc[lambda s: s != ""] if "tags" in _raw_vulns.columns else pd.Series(dtype=str)
        print(f"  vulns_all rows         : {len(_raw_vulns)}")
        print(f"  vulns with tags col    : {'yes' if 'tags' in _raw_vulns.columns else 'no (expected)'}")
        print(f"  vuln asset_uuid sample   : {_raw_vulns['asset_uuid'].dropna().head(3).tolist()}")
        print(f"  asset asset_uuid sample  : {_raw_assets['asset_uuid'].dropna().head(3).tolist() if _raw_assets_cache.exists() else 'N/A'}")
    else:
        print(f"  vulns_all.parquet not found at {_raw_vulns_cache}")
    print()

    _vulns_df, _assets_df = _fetch_and_prepare(
        tio          = _tio,
        cache_dir    = _cache_dir,
        tag_category = args.tag_category,
        tag_value    = args.tag_value,
        as_of        = _as_of,
    )

    _recast_rules_df = fetch_recast_rules(_tio, _cache_dir)

    _scanned_df, _unscanned_df = _identify_unscanned_assets(_assets_df, as_of=_as_of)

    _scanned_ids   = set(_scanned_df["asset_uuid"])
    _vulns_scanned = _vulns_df[_vulns_df["asset_uuid"].isin(_scanned_ids)]

    _plugin_df = _group_by_plugin(_vulns_scanned)

    _summary = _compute_summary_metrics(
        vulns_df     = _vulns_scanned,
        plugin_df    = _plugin_df,
        assets_df    = _assets_df,
        unscanned_df = _unscanned_df,
        tag_category = args.tag_category,
        tag_value    = args.tag_value,
        as_of        = _as_of,
    )

    _exploit_metrics_cli = _compute_exploitability_metrics(_vulns_scanned)
    _summary["known_exploit"]      = _exploit_metrics_cli["known_exploit"]
    _summary["exploit_functional"] = _exploit_metrics_cli["functional"]
    _summary["exploit_high"]       = _exploit_metrics_cli["high_maturity"]

    _priority_df_cli = _get_top_priority_plugins(_vulns_scanned)
    _summary["priority_plugins"] = _priority_df_cli.to_dict("records")

    _risk_mods_df_cli = _extract_risk_modifications(
        vulns_df        = _vulns_scanned,
        assets_df       = _assets_df,
        recast_rules_df = _recast_rules_df,
        as_of           = _as_of,
    )
    _recurring_df_cli = _extract_recurring_vulnerabilities(
        vulns_df  = _vulns_scanned,
        assets_df = _assets_df,
    )
    _summary["count_risk_accepted"] = int((_risk_mods_df_cli["Modification Type"] == "Accepted").sum()) if not _risk_mods_df_cli.empty else 0
    _summary["count_risk_recast"]   = int((_risk_mods_df_cli["Modification Type"] == "Recast").sum())   if not _risk_mods_df_cli.empty else 0
    _summary["count_expiring_soon"] = int(
        _risk_mods_df_cli["Days Until Expiry"].apply(
            lambda x: isinstance(x, int) and 0 <= x <= 30
        ).sum()
    ) if not _risk_mods_df_cli.empty else 0
    _summary["count_expired"]       = int(
        _risk_mods_df_cli["Days Until Expiry"].apply(
            lambda x: isinstance(x, int) and x < 0
        ).sum()
    ) if not _risk_mods_df_cli.empty else 0
    _summary["count_recurring"]     = len(_recurring_df_cli)

    print("\n=== Step 1 — Data Preparation Summary ===")
    for k, v in _summary.items():
        if k == "top5_plugins":
            print(f"  top5_plugins:")
            for p in v:
                print(f"    {p['plugin_name'][:60]:<60}  {p['affected_asset_count']} assets")
        else:
            print(f"  {k:<28}: {v}")

    print(f"\n  plugin_df rows       : {len(_plugin_df)}")
    print(f"  unscanned_df rows    : {len(_unscanned_df)}")
    print(f"  risk_mods rows       : {len(_risk_mods_df_cli)}")
    print(f"  recurring vuln rows  : {len(_recurring_df_cli)}")
    if not _plugin_df.empty:
        print(f"\n  SLA state distribution (by plugin row):")
        for state in OPS_SLA_STATE_ORDER:
            n = int((_plugin_df["sla_status"] == state).sum())
            print(f"    {state:<35}: {n}")

    # --- Asset enrichment for overdue/urgent display ---
    _asset_lookup_cli = (
        _assets_df[["asset_uuid", "hostname", "mac_address", "ipv4"]]
        .drop_duplicates("asset_uuid")
        if all(c in _assets_df.columns for c in ["asset_uuid", "hostname", "mac_address", "ipv4"])
        else None
    )

    def _enrich_cli(df: pd.DataFrame) -> pd.DataFrame:
        if _asset_lookup_cli is None or df.empty:
            return df
        drop_cols = [c for c in ["hostname", "mac_address", "ipv4"] if c in df.columns]
        return (
            df.drop(columns=drop_cols)
            .merge(_asset_lookup_cli, on="asset_uuid", how="left")
        )

    # --- Step 2: Build Excel ---
    _overdue_df = _enrich_cli(
        _vulns_scanned[_vulns_scanned["ops_sla_status"] == OPS_SLA_OVERDUE]
    ).sort_values(
        ["severity", "days_open"],
        ascending=[True, False],
        na_position="last",
        key=lambda s: s.map(SEVERITY_RANK) if s.name == "severity" else s,
    )
    _output_dir.mkdir(parents=True, exist_ok=True)
    _slug = safe_filename(
        f"{args.tag_category or 'all'}_{args.tag_value or 'assets'}"
    )
    _excel_path = _output_dir / f"ops_remediation_{_slug}.xlsx"
    _build_excel(
        plugin_df    = _plugin_df,
        overdue_df   = _overdue_df,
        unscanned_df = _unscanned_df,
        summary      = _summary,
        output_path  = _excel_path,
        tag_category = args.tag_category,
        tag_value    = args.tag_value,
        risk_mods_df = _risk_mods_df_cli,
        recurring_df = _recurring_df_cli,
    )
    print(f"\n  Excel written: {_excel_path}")

    # --- Step 3: Build PDF ---
    _urgent_df = _enrich_cli(
        _vulns_scanned[_vulns_scanned["ops_sla_status"] == OPS_SLA_URGENT]
    ).sort_values(
        ["severity", "days_remaining"],
        ascending=[True, True],
        na_position="last",
        key=lambda s: s.map(SEVERITY_RANK) if s.name == "severity" else s,
    )
    _pdf_path = _output_dir / f"ops_remediation_{_slug}.pdf"
    _build_pdf(
        plugin_df    = _plugin_df,
        overdue_df   = _overdue_df,
        urgent_df    = _urgent_df,
        unscanned_df = _unscanned_df,
        summary      = _summary,
        output_path  = _pdf_path,
        tag_category = args.tag_category,
        tag_value    = args.tag_value,
    )
    print(f"  PDF written:  {_pdf_path}")

    # --- Step 4: Build email summary ---
    _email_metrics = _build_email_summary(
        summary      = _summary,
        tag_category = args.tag_category,
        tag_value    = args.tag_value,
    )
    print(f"\n=== Step 4 - Email Summary ===")
    print(f"  kpi_tiles ({len(_email_metrics['kpi_tiles'])} tiles):")
    for tile in _email_metrics["kpi_tiles"]:
        sub = f"  ({tile['sub_label']})" if tile.get("sub_label") else ""
        print(f"    [{tile['label']}] {tile['value']}{sub}  color={tile['color']}")
    print(f"  raw keys: {list(_email_metrics['raw'].keys())}")
