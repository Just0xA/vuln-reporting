"""
exporters/excel_exporter.py — Excel (.xlsx) report generation with openpyxl.

Every sheet produced by this module:
  - Has auto-sized columns (capped at 60 chars wide)
  - Has the header row frozen (freeze_panes)
  - Has the header row styled with a dark band and white bold text
  - Can receive SLA conditional formatting via apply_sla_conditional_formatting()

Exported API
------------
write_dataframe_to_sheet()         — write a DataFrame to a named worksheet
apply_sla_conditional_formatting() — red/yellow/green rules on SLA status column
write_metadata_tab()               — standard "Report Info" sheet
export_to_excel()                  — top-level function: assemble full workbook
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.differential import DifferentialStyle

from config import (
    SEVERITY_FILL_COLORS,
    SLA_DAYS,
    VPR_SEVERITY_MAP,
    SEVERITY_LABELS,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
_BODY_FONT   = Font(size=10, name="Calibri")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_ALIGN   = Alignment(vertical="center", wrap_text=False)
_THIN_BORDER  = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# SLA conditional format fills
_FILL_RED    = PatternFill("solid", fgColor="FFCDD2")   # overdue
_FILL_YELLOW = PatternFill("solid", fgColor="FFF9C4")   # near SLA
_FILL_GREEN  = PatternFill("solid", fgColor="C8E6C9")   # healthy

_HEADER_ROW_HEIGHT = 28  # pts
_BODY_ROW_HEIGHT   = 16  # pts
_MAX_COL_WIDTH     = 60  # chars


# ===========================================================================
# Internal helpers
# ===========================================================================

def _apply_header_style(ws, header_row: int = 1) -> None:
    """Apply dark-band style to every cell in the header row."""
    for cell in ws[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[header_row].height = _HEADER_ROW_HEIGHT


def _apply_body_style(ws, start_row: int = 2) -> None:
    """Apply body font and zebra-stripe fills to all data rows."""
    _zebra_even = PatternFill("solid", fgColor="F5F7FA")
    _zebra_odd  = PatternFill("solid", fgColor="FFFFFF")
    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        fill = _zebra_even if i % 2 == 0 else _zebra_odd
        for cell in row:
            cell.font = _BODY_FONT
            cell.alignment = _BODY_ALIGN
            cell.fill = fill
            cell.border = _THIN_BORDER
        ws.row_dimensions[start_row + i].height = _BODY_ROW_HEIGHT


def _autosize_columns(ws, min_width: int = 8) -> None:
    """Set each column width to the maximum content length, capped at _MAX_COL_WIDTH."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, _MAX_COL_WIDTH)


def _col_letter_for(ws, header_name: str) -> Optional[str]:
    """
    Return the column letter for the column whose header matches *header_name*.
    Returns None if not found.
    """
    for cell in ws[1]:
        if str(cell.value).lower() == header_name.lower():
            return get_column_letter(cell.column)
    return None


# ===========================================================================
# Public API
# ===========================================================================

def write_dataframe_to_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    title_row: Optional[str] = None,
    severity_col: Optional[str] = "severity",
) -> None:
    """
    Write a DataFrame to a new worksheet in *wb* with full formatting applied.

    Parameters
    ----------
    wb : Workbook
        Destination openpyxl workbook.
    df : pd.DataFrame
        Data to write.  All columns are written as-is; datetime columns are
        formatted as ISO date strings.
    sheet_name : str
        Tab name.  Truncated to 31 chars (Excel limit).
    title_row : str, optional
        If provided, a merged title row is inserted above the header.
    severity_col : str, optional
        If the DataFrame has a column with this name, cells in that column
        receive the matching severity fill color from config.SEVERITY_FILL_COLORS.
        Pass None to skip severity coloring.
    """
    ws = wb.create_sheet(title=sheet_name[:31])

    header_offset = 0

    # Optional merged title row
    if title_row:
        ws.append([title_row])
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=13, color="1F3864", name="Calibri")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        if len(df.columns) > 1:
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=len(df.columns),
            )
        header_offset = 1

    # Coerce datetime columns to strings for Excel compatibility
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d").fillna("")

    # Write header + data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)

    # Adjust row references to account for optional title row
    header_row_num = 1 + header_offset

    _apply_header_style(ws, header_row=header_row_num)
    _apply_body_style(ws, start_row=header_row_num + 1)
    _autosize_columns(ws)

    # Freeze header (and title if present)
    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=1)

    # Severity cell coloring
    if severity_col and severity_col in df.columns:
        sev_letter = _col_letter_for(ws, severity_col)
        if sev_letter:
            for row in ws.iter_rows(
                min_row=header_row_num + 1,
                max_row=ws.max_row,
                min_col=ws[sev_letter + str(header_row_num + 1)].column,
                max_col=ws[sev_letter + str(header_row_num + 1)].column,
            ):
                for cell in row:
                    sev_key = str(cell.value).lower() if cell.value else ""
                    fill_hex = SEVERITY_FILL_COLORS.get(sev_key)
                    if fill_hex:
                        cell.fill = PatternFill("solid", fgColor=fill_hex)

    logger.debug("Wrote sheet '%s' (%d rows, %d cols)", sheet_name, len(df), len(df.columns))


def apply_sla_conditional_formatting(
    ws,
    sla_status_col: str = "sla_status",
) -> None:
    """
    Apply red / yellow / green conditional formatting to the SLA status column.

    Rules applied:
      - "Overdue"     → light red fill  (#FFCDD2)
      - "Within SLA"  → light green fill (#C8E6C9)
      - "Remediated"  → no special fill (cleared by body zebra stripe)

    The "near SLA" yellow band (within 20% of SLA window remaining) requires
    numeric days_remaining data; when that column is present it is also
    highlighted.  This is applied as cell-value rules on the sla_status column
    plus a formula rule on the days_remaining column if found.

    Parameters
    ----------
    ws : Worksheet
        The openpyxl worksheet to format (must already have data).
    sla_status_col : str
        Header name of the SLA status text column.
    """
    status_letter = _col_letter_for(ws, sla_status_col)
    if not status_letter:
        logger.debug("apply_sla_conditional_formatting: column '%s' not found in sheet", sla_status_col)
        return

    data_start = 2  # row 1 is header
    data_end = ws.max_row
    status_range = f"{status_letter}{data_start}:{status_letter}{data_end}"

    # Overdue → red
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal",
            formula=['"Overdue"'],
            fill=_FILL_RED,
            font=Font(bold=True, color="B71C1C", size=10, name="Calibri"),
        ),
    )

    # Within SLA → green
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal",
            formula=['"Within SLA"'],
            fill=_FILL_GREEN,
            font=Font(color="1B5E20", size=10, name="Calibri"),
        ),
    )

    # Remediated → blue-tint (informational)
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal",
            formula=['"Remediated"'],
            fill=PatternFill("solid", fgColor="BBDEFB"),
            font=Font(color="0D47A1", size=10, name="Calibri"),
        ),
    )

    # days_remaining column: yellow when 0 < days_remaining <= 20% of SLA max (30d)
    dr_letter = _col_letter_for(ws, "days_remaining")
    if dr_letter:
        dr_range = f"{dr_letter}{data_start}:{dr_letter}{data_end}"
        # Near SLA: positive but small (within 6 days of a 30-day SLA as default proxy)
        ws.conditional_formatting.add(
            dr_range,
            CellIsRule(
                operator="between",
                formula=["0", "6"],
                fill=_FILL_YELLOW,
                font=Font(color="F57F17", size=10, name="Calibri"),
            ),
        )

    logger.debug("Applied SLA conditional formatting to sheet column %s", status_letter)


def write_metadata_tab(
    wb: Workbook,
    report_name: str,
    tag_filter: str,
    generated_at: Optional[datetime] = None,
) -> None:
    """
    Insert a "Report Info" worksheet at position 0 with report metadata.

    Sheet contents:
      - Report name, generation timestamp, tag filter / scope
      - SLA definitions table (severity → SLA days)
      - VPR severity score ranges table

    Parameters
    ----------
    wb : Workbook
    report_name : str
    tag_filter : str
        Human-readable scope string, e.g. "Environment = Production"
        or "All Assets".
    generated_at : datetime, optional
        Defaults to UTC now.
    """
    if generated_at is None:
        generated_at = datetime.now(tz=timezone.utc)

    ws = wb.create_sheet(title="Report Info", index=0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    _section_font  = Font(bold=True, size=11, color="1F3864", name="Calibri")
    _label_font    = Font(bold=True, size=10, name="Calibri")
    _value_font    = Font(size=10, name="Calibri")
    _heading_fill  = PatternFill("solid", fgColor="E8EAF6")
    _alt_fill      = PatternFill("solid", fgColor="F5F7FA")

    def _write(row, col, value, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        return cell

    row = 1

    # --- Report header ---
    _write(row, 1, "Report Metadata", font=Font(bold=True, size=14, color="1F3864", name="Calibri"))
    ws.row_dimensions[row].height = 24
    row += 1

    meta_rows = [
        ("Report Name",          report_name),
        ("Generated (UTC)",      generated_at.strftime("%Y-%m-%d %H:%M UTC")),
        ("Scope / Tag Filter",   tag_filter),
    ]
    for label, value in meta_rows:
        _write(row, 1, label, font=_label_font)
        _write(row, 2, value, font=_value_font)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1  # spacer

    # --- SLA definitions ---
    _write(row, 1, "SLA Definitions", font=_section_font)
    ws.row_dimensions[row].height = 20
    row += 1

    _write(row, 1, "Severity",  font=_label_font, fill=_heading_fill)
    _write(row, 2, "SLA (Days to Remediate)", font=_label_font, fill=_heading_fill)
    row += 1

    for i, (sev, days) in enumerate(SLA_DAYS.items()):
        fill = _alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        sev_fill = PatternFill("solid", fgColor=SEVERITY_FILL_COLORS.get(sev, "FFFFFF"))
        _write(row, 1, SEVERITY_LABELS.get(sev, sev.title()), font=_value_font, fill=sev_fill)
        _write(row, 2, f"{days} days", font=_value_font, fill=fill)
        row += 1

    row += 1  # spacer

    # --- VPR ranges ---
    _write(row, 1, "VPR Severity Score Ranges", font=_section_font)
    ws.row_dimensions[row].height = 20
    row += 1

    _write(row, 1, "Severity",    font=_label_font, fill=_heading_fill)
    _write(row, 2, "VPR Score Range", font=_label_font, fill=_heading_fill)
    row += 1

    for i, (lo, hi, label) in enumerate(VPR_SEVERITY_MAP):
        fill = _alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        sev_fill = PatternFill("solid", fgColor=SEVERITY_FILL_COLORS.get(label, "FFFFFF"))
        _write(row, 1, SEVERITY_LABELS.get(label, label.title()), font=_value_font, fill=sev_fill)
        _write(row, 2, f"{lo} – {hi}", font=_value_font, fill=fill)
        row += 1

    logger.debug("Wrote metadata tab for report '%s'", report_name)


def export_to_excel(
    sheets: list[dict],
    output_path: str | Path,
    report_name: str,
    tag_filter: str = "All Assets",
    generated_at: Optional[datetime] = None,
) -> Path:
    """
    Assemble a complete .xlsx workbook from a list of sheet descriptors.

    Parameters
    ----------
    sheets : list[dict]
        Each dict describes one worksheet:

        .. code-block:: python

            {
                "name": "SLA Summary",          # tab name (required)
                "df": pd.DataFrame(...),         # data (required)
                "title": "SLA Status — Finance", # optional merged title row
                "severity_col": "severity",      # column to severity-color (default "severity")
                "sla_formatting": True,          # apply SLA conditional formatting (default False)
                "sla_status_col": "sla_status",  # which col to format (default "sla_status")
            }

    output_path : str or Path
        Destination file path (including .xlsx extension).
    report_name : str
        Used in the metadata tab.
    tag_filter : str
        Human-readable scope string for the metadata tab.
    generated_at : datetime, optional
        Defaults to UTC now.

    Returns
    -------
    Path
        Absolute path of the written file.
    """
    if generated_at is None:
        generated_at = datetime.now(tz=timezone.utc)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Metadata tab always first
    write_metadata_tab(wb, report_name, tag_filter, generated_at)

    for sheet_cfg in sheets:
        name        = sheet_cfg["name"]
        df          = sheet_cfg["df"]
        title       = sheet_cfg.get("title")
        sev_col     = sheet_cfg.get("severity_col", "severity")
        do_sla_fmt  = sheet_cfg.get("sla_formatting", False)
        sla_col     = sheet_cfg.get("sla_status_col", "sla_status")

        if df is None or df.empty:
            logger.warning("Sheet '%s' has no data — writing empty placeholder.", name)
            df = pd.DataFrame({"(no data)": ["No records matched the current filter."]})
            sev_col = None

        write_dataframe_to_sheet(wb, df, name, title_row=title, severity_col=sev_col)

        if do_sla_fmt:
            ws = wb[name[:31]]
            apply_sla_conditional_formatting(ws, sla_status_col=sla_col)

    wb.save(output_path)
    logger.info("Excel workbook written: %s (%d sheet(s))", output_path, len(wb.sheetnames))
    return output_path.resolve()


if __name__ == "__main__":
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import OUTPUT_DIR

    out = OUTPUT_DIR / "excel_test"
    out.mkdir(parents=True, exist_ok=True)

    sample_df = pd.DataFrame({
        "asset_hostname": ["web-01.corp", "db-02.corp", "app-03.corp"],
        "severity":       ["critical",    "high",       "medium"],
        "plugin_name":    ["OpenSSL RCE", "Log4j",      "PHP XSS"],
        "days_open":      [20,            35,           60],
        "sla_days":       [15,            30,           90],
        "days_remaining": [-5,            -5,           30],
        "sla_status":     ["Overdue",     "Overdue",    "Within SLA"],
    })

    path = export_to_excel(
        sheets=[{
            "name": "SLA Status",
            "df": sample_df,
            "title": "SLA Remediation — All Assets",
            "sla_formatting": True,
        }],
        output_path=out / "test_report.xlsx",
        report_name="SLA Remediation Test",
        tag_filter="All Assets",
    )
    print(f"Written: {path}")
