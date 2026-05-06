"""
tests/test_phase2_composer_pipeline.py — Phase 2 regression + isolation bar.

Exercises the four-channel `ReportComposer` orchestration introduced in
Phase 2 (Plans 02-01..02-04) against COMPOSER-01..04 requirements and
CONTEXT.md decisions D-22..D-29.

Seven checks run sequentially:

1. Bundle-shape regression — `run_full_pipeline()` returns EXACTLY the
   seven keys documented in D-22.
2. Module-configs ordering — channels iterate in `_module_configs`
   order (D-27).
3. Page-2 RAG strip presence + cover stability (D-29) — generated PDF
   HTML carries `<div class="rag-strip">` AFTER `<div class="report-cover">`,
   and the cover region is hash-stable across two runs with identical
   inputs.
4a. Excel CONTENT hash stability (D-29 intent) — saved main-Excel hashes
    equally on two equivalent runs when hashed over per-cell values.
4b. Excel BYTE hash stability with mtime normalization (D-29 literal) —
    saved main-Excel bytes hash equally on two equivalent runs AFTER
    rewriting the .xlsx ZIP with all internal entry mtimes forced to
    the ZIP epoch (1980-01-01 00:00:00). Defeats openpyxl's per-entry
    timestamp non-determinism. If the rewrite fails on this openpyxl /
    Python version, the check degrades to [SKIP] (logged, not failed) —
    Check 4a alone still satisfies D-29's intent.
5. Email-panel exception isolation (D-28) — raising stub produces a
   placeholder `<div>` with the error text; other modules' panels
   still render.
6. Analyst-tabs exception isolation (D-28) — raising stub is recorded
   in the `_Metadata` Failures subsection; other modules' tabs still
   land in the workbook.

CLI
---
python tests/test_phase2_composer_pipeline.py
python tests/test_phase2_composer_pipeline.py --verbose
"""

from __future__ import annotations

import argparse
import hashlib
import io
import sys
import tempfile
import traceback
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd
import openpyxl

# Allow running as a top-level script from any working directory
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from reports.modules import (  # noqa: E402
    ReportComposer, BaseModule, ModuleConfig, ModuleData,
    register_module,
)

# ── Phase 3 (Plan 03-06) — real migrated module classes ──
from reports.modules.scan_coverage_sla_module       import ScanCoverageSLAModule  # noqa: E402
from reports.modules.critical_remediation_sla_module import CriticalRemediationSLAModule  # noqa: E402
from reports.modules.high_risk_assets_module        import HighRiskAssetsModule  # noqa: E402
from reports.modules.aged_vulns_assets_module       import AgedVulnsAssetsModule  # noqa: E402
from reports.modules.rag_utils                      import (  # noqa: E402
    NO_DATA_HEADLINE,
    NO_DATA_DRIVER,
    STATUS_COLOR,
)
# W7 — fixture code uses safe_pct for headline_value_str (CLAUDE.md Empty-data
# guard pattern rule 1: never use inline f-string format specs like the
# percent-precision spec on a possibly-None value, even in tests).
from reports.modules.format_utils                   import safe_pct  # noqa: E402


# ===========================================================================
# Stub modules — registered fresh per script run; namespaced with
# _phase2_test_* to avoid colliding with the 8 production modules.
# ===========================================================================

@register_module
class _Phase2TestPanelA(BaseModule):
    MODULE_ID    = "_phase2_test_panel_a"
    DISPLAY_NAME = "Phase2 Test Panel A"

    def compute(self, vulns_df, assets_df, report_date, config, **k):
        return ModuleData(
            module_id=self.MODULE_ID, display_name=self.DISPLAY_NAME,
            metrics={}, table_data=[], chart_data={}, summary_text="",
            metadata={},
        )

    def render_email_panel(self, data, config):
        return '<table data-stub="A"><tr><td>PANEL_A_BODY</td></tr></table>'

    def render_analyst_tabs(self, data, config):
        return [("AnalystTabA", pd.DataFrame({"col1": [1, 2], "col2": ["x", "y"]}))]

    def render_rag_strip_entry(self, data, config):
        from reports.modules.rag_utils import build_rag_strip_entry  # noqa: PLC0415
        return build_rag_strip_entry(self.DISPLAY_NAME, "97.2%", "green")


@register_module
class _Phase2TestPanelB(BaseModule):
    MODULE_ID    = "_phase2_test_panel_b"
    DISPLAY_NAME = "Phase2 Test Panel B"

    def compute(self, vulns_df, assets_df, report_date, config, **k):
        return ModuleData(
            module_id=self.MODULE_ID, display_name=self.DISPLAY_NAME,
            metrics={}, table_data=[], chart_data={}, summary_text="",
            metadata={},
        )

    def render_email_panel(self, data, config):
        return '<table data-stub="B"><tr><td>PANEL_B_BODY</td></tr></table>'

    def render_analyst_tabs(self, data, config):
        return [("AnalystTabB", pd.DataFrame({"col1": [3, 4]}))]

    def render_rag_strip_entry(self, data, config):
        from reports.modules.rag_utils import build_rag_strip_entry  # noqa: PLC0415
        return build_rag_strip_entry(self.DISPLAY_NAME, "12 assets", "yellow")


@register_module
class _Phase2TestPanelBoom(BaseModule):
    """Stub that raises in BOTH render_email_panel and render_analyst_tabs.
    Used to verify D-28 isolation: one module's failure must not abort
    assembly of other modules' contributions."""
    MODULE_ID    = "_phase2_test_panel_boom"
    DISPLAY_NAME = "Phase2 Test Panel Boom"

    def compute(self, vulns_df, assets_df, report_date, config, **k):
        return ModuleData(
            module_id=self.MODULE_ID, display_name=self.DISPLAY_NAME,
            metrics={}, table_data=[], chart_data={}, summary_text="",
            metadata={},
        )

    def render_email_panel(self, data, config):
        raise RuntimeError("phase2-email-boom")

    def render_analyst_tabs(self, data, config):
        raise RuntimeError("phase2-analyst-boom")


# ===========================================================================
# Test helpers
# ===========================================================================

# Sentinel returned by Check 4b when mtime normalization is infeasible.
class _SkipCheck(Exception):
    """Raise to mark a check as [SKIP] rather than [PASS]/[FAIL]."""
    pass


def _make_composer(*module_ids: str) -> ReportComposer:
    """Build a ReportComposer with frozen inputs so two runs are byte-equal."""
    return ReportComposer(
        vulns_df=pd.DataFrame(),
        assets_df=pd.DataFrame(),
        report_date="2026-05-06",   # frozen string keeps timestamps deterministic
        module_configs=[ModuleConfig(mid) for mid in module_ids],
    )


def _make_results(*module_ids: str) -> list[ModuleData]:
    """Build successful ModuleData entries for the given module IDs."""
    return [
        ModuleData(
            module_id=mid, display_name=mid.replace("_", " ").title(),
            metrics={}, table_data=[], chart_data={}, summary_text="",
            metadata={},
        )
        for mid in module_ids
    ]


def _save_workbook_bytes(wb: openpyxl.Workbook) -> bytes:
    """Save the workbook to an in-memory buffer and return the raw bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sha256(data: bytes | str) -> str:
    if isinstance(data, str):
        data = data.encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def _normalize_xlsx_mtimes(xlsx_bytes: bytes) -> bytes:
    """Rewrite an .xlsx (ZIP) container with every ZipInfo.date_time
    forced to the ZIP epoch (1980-01-01 00:00:00). This defeats
    openpyxl's per-entry mtime non-determinism so two equivalent runs
    can produce byte-identical output.

    Raises _SkipCheck if the input is not a valid ZIP / openpyxl emits
    something the rewrite cannot losslessly reconstruct (extras,
    encryption, etc.).
    """
    EPOCH = (1980, 1, 1, 0, 0, 0)
    try:
        out_buf = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as z_in:
            # Sort entries so the order is deterministic too — openpyxl
            # writes them in a consistent order today, but explicit
            # sorting is cheap insurance.
            infos = sorted(z_in.infolist(), key=lambda i: i.filename)
            with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as z_out:
                for info in infos:
                    raw = z_in.read(info.filename)
                    new_info = zipfile.ZipInfo(filename=info.filename, date_time=EPOCH)
                    new_info.compress_type = info.compress_type
                    # Preserve external_attr so file vs. dir bits stay correct.
                    new_info.external_attr = info.external_attr
                    z_out.writestr(new_info, raw)
        return out_buf.getvalue()
    except (zipfile.BadZipFile, RuntimeError, NotImplementedError) as exc:
        raise _SkipCheck(f"mtime-normalization rewrite failed: {exc!r}") from exc


# ===========================================================================
# The 7 checks
# ===========================================================================

def check_1_bundle_shape() -> None:
    """D-22: run_full_pipeline returns exactly the documented bundle keys.

    Plan 03-01 (Phase 3 D-04) adds ``email_inline_images`` to the bundle —
    expected set is updated accordingly. Plan 03-06 owns the broader
    regression-snapshot extension; this is the minimum required to keep
    the bundle-shape regression bar green at Plan 03-01's commit.
    """
    composer = _make_composer("_phase2_test_panel_a", "_phase2_test_panel_b")
    results  = _make_results("_phase2_test_panel_a", "_phase2_test_panel_b")
    with tempfile.TemporaryDirectory() as td:
        bundle = composer.run_full_pipeline(
            results, td,
            slug="phase2_test", report_date="2026-05-06",
            pdf_title="P2 Test", pdf_subtitle="Scope: All Assets",
            scope_label="All Assets",
        )
    expected = {
        "pdf_html", "excel_workbook", "analyst_workbook_path",
        "email_body_html", "email_inline_images",  # Plan 03-01 D-04
        "email_kpis", "metrics", "errors",
    }
    actual = set(bundle.keys())
    assert actual == expected, (
        f"Bundle keys mismatch — expected {sorted(expected)}, got {sorted(actual)}"
    )


def check_2_ordering() -> None:
    """D-27: channels iterate in _module_configs order."""
    composer = _make_composer("_phase2_test_panel_b", "_phase2_test_panel_a")  # B before A
    results  = _make_results("_phase2_test_panel_b", "_phase2_test_panel_a")
    with tempfile.TemporaryDirectory() as td:
        bundle = composer.run_full_pipeline(
            results, td, slug="phase2_test", report_date="2026-05-06",
        )
        # Email body: PANEL_B_BODY appears before PANEL_A_BODY
        body = bundle["email_body_html"]
        assert "PANEL_B_BODY" in body and "PANEL_A_BODY" in body, body
        assert body.index("PANEL_B_BODY") < body.index("PANEL_A_BODY"), (
            "Email body did not preserve _module_configs order — "
            f"got: {body!r}"
        )
        # Analyst workbook: AnalystTabB appears before AnalystTabA
        # (Read inside the TemporaryDirectory context — Rule 1 fix vs.
        # plan's original code which exited the context before the read.)
        wb_path = bundle["analyst_workbook_path"]
        assert wb_path is not None
        wb = openpyxl.load_workbook(str(wb_path))
        sheet_names = [n for n in wb.sheetnames if n != "_Metadata"]
        assert sheet_names == ["AnalystTabB", "AnalystTabA"], (
            f"Analyst workbook tab order != module_configs order: {sheet_names}"
        )


def check_3_page2_strip_and_cover_stability() -> None:
    """D-29: page-2 RAG strip is present, appears AFTER the cover, and the
    cover-region HTML is hash-stable across two equivalent runs."""
    def _run() -> str:
        composer = _make_composer("_phase2_test_panel_a", "_phase2_test_panel_b")
        results  = _make_results("_phase2_test_panel_a", "_phase2_test_panel_b")
        return composer.assemble_pdf(
            results, title="P2 Stability", subtitle="Scope: All Assets",
        )

    html_a = _run()
    html_b = _run()

    # 1. Both runs render an identical PDF HTML (deterministic)
    assert _sha256(html_a) == _sha256(html_b), (
        "PDF HTML differs between two equivalent runs — "
        "non-determinism in assemble_pdf would defeat the regression bar."
    )

    # 2. Page-2 RAG strip is present and AFTER the cover
    assert '<div class="report-cover">' in html_a, "cover div missing"
    assert '<div class="rag-strip">' in html_a, "page-2 strip div missing"
    assert "Risk Status Summary" in html_a, "page-2 header missing (D-05)"
    assert html_a.index('<div class="rag-strip">') > html_a.index(
        '<div class="report-cover">'
    ), "page-2 strip rendered BEFORE cover — wrong placement"

    # 3. Cover-region substring (page 1) is byte-stable
    cover_start = html_a.index('<div class="report-cover">')
    cover_end   = html_a.index("</div>", cover_start)
    cover_a = html_a[cover_start:cover_end]
    cover_b = html_b[cover_start:cover_end]
    assert _sha256(cover_a) == _sha256(cover_b), (
        "Cover region hash drifted between two equivalent runs"
    )


def _excel_run_bytes() -> bytes:
    """Run the main-Excel assembly once and return the saved .xlsx bytes."""
    composer = _make_composer("_phase2_test_panel_a", "_phase2_test_panel_b")
    results  = _make_results("_phase2_test_panel_a", "_phase2_test_panel_b")
    wb = openpyxl.Workbook()
    if wb.worksheets:
        wb.remove(wb.worksheets[0])
    composer.assemble_excel(results, wb)
    return _save_workbook_bytes(wb)


def check_4a_excel_content_hash_stability() -> None:
    """D-29 intent: Saved main-Excel CONTENT hashes equally across two
    equivalent runs (hashes per-cell values, not raw bytes — defeats
    openpyxl's non-deterministic ZipInfo mtimes)."""
    bytes_a = _excel_run_bytes()
    bytes_b = _excel_run_bytes()

    def _content_hash(b: bytes) -> str:
        wb = openpyxl.load_workbook(io.BytesIO(b))
        m = hashlib.sha256()
        for ws_name in sorted(wb.sheetnames):
            ws = wb[ws_name]
            m.update(ws_name.encode("utf-8"))
            for row in ws.iter_rows(values_only=True):
                for val in row:
                    m.update(repr(val).encode("utf-8"))
        return m.hexdigest()
    assert _content_hash(bytes_a) == _content_hash(bytes_b), (
        "Main Excel content hash differs between two equivalent runs — "
        "regression bar would fail."
    )


def check_4b_excel_byte_hash_stability_mtime_normalized() -> None:
    """D-29 literal: Saved main-Excel BYTES hash equally across two
    equivalent runs AFTER ZIP-entry mtimes are normalized to the ZIP
    epoch (1980-01-01 00:00:00). If the openpyxl ZIP layout cannot be
    losslessly rewritten (extras / encryption / runtime drift), this
    check degrades to [SKIP] — Check 4a alone still satisfies D-29's
    content-equivalence intent.

    Records the deviation (if it skips) by raising _SkipCheck which the
    driver renders as [SKIP] rather than [FAIL]."""
    bytes_a = _excel_run_bytes()
    bytes_b = _excel_run_bytes()

    norm_a = _normalize_xlsx_mtimes(bytes_a)   # may raise _SkipCheck
    norm_b = _normalize_xlsx_mtimes(bytes_b)

    assert _sha256(norm_a) == _sha256(norm_b), (
        "Main Excel mtime-normalized byte hash differs between two equivalent "
        "runs. openpyxl is emitting non-mtime non-determinism (entry order, "
        "compression flags, internal XML attribute ordering, etc.) — investigate "
        "before claiming D-29 byte-for-byte stability."
    )


def check_5_email_panel_isolation() -> None:
    """D-28: One module's render_email_panel exception MUST NOT abort
    other modules' panels. The raising module's slot becomes a visible
    error placeholder div."""
    composer = _make_composer(
        "_phase2_test_panel_a",
        "_phase2_test_panel_boom",
        "_phase2_test_panel_b",
    )
    results = _make_results(
        "_phase2_test_panel_a",
        "_phase2_test_panel_boom",
        "_phase2_test_panel_b",
    )
    body = composer.assemble_email_body(results)

    # Stubs A and B still rendered
    assert "PANEL_A_BODY" in body, "Panel A body missing — isolation broken"
    assert "PANEL_B_BODY" in body, "Panel B body missing — isolation broken"
    # Boom became a visible placeholder containing the exception text
    assert "phase2-email-boom" in body, (
        f"Email panel placeholder did not contain the exception text. body={body!r}"
    )
    assert "email panel render failed" in body or "email panel" in body, (
        "Email panel placeholder did not contain a 'failed' marker"
    )
    # Order preserved: A → placeholder → B
    assert body.index("PANEL_A_BODY") < body.index("phase2-email-boom") < body.index("PANEL_B_BODY"), (
        "Email panel ordering broken around the failed module"
    )


def check_6_analyst_tabs_isolation() -> None:
    """D-28: One module's render_analyst_tabs exception MUST be recorded
    in the _Metadata Failures subsection; other modules' tabs still land."""
    composer = _make_composer(
        "_phase2_test_panel_a",
        "_phase2_test_panel_boom",
        "_phase2_test_panel_b",
    )
    results = _make_results(
        "_phase2_test_panel_a",
        "_phase2_test_panel_boom",
        "_phase2_test_panel_b",
    )
    with tempfile.TemporaryDirectory() as td:
        target = Path(td) / "phase2_isolation_analyst.xlsx"
        out = composer.assemble_analyst_workbook(
            results, target,
            slug="phase2_test", scope_label="All Assets",
        )
        assert out == target and target.exists(), (
            "Analyst workbook should still be written when only ONE module raises"
        )
        wb = openpyxl.load_workbook(str(target))
        # Surviving modules' tabs are present
        assert "AnalystTabA" in wb.sheetnames
        assert "AnalystTabB" in wb.sheetnames
        # _Metadata Failures records the boom
        assert "_Metadata" in wb.sheetnames
        md = wb["_Metadata"]
        assert md["A6"].value == "Failures", "Failures section missing"
        rows_text = []
        for row in md.iter_rows(min_row=8, max_col=2, values_only=True):
            rows_text.append(repr(row))
        joined = " ".join(rows_text)
        assert "_phase2_test_panel_boom" in joined, (
            f"Failed module ID not recorded in _Metadata. rows={rows_text}"
        )
        assert "phase2-analyst-boom" in joined, (
            f"Failure detail text not recorded in _Metadata. rows={rows_text}"
        )


# ===========================================================================
# Phase 3 (Plan 03-06) — QUALITY-02 zero-row integration coverage
# ===========================================================================

def check_8_phase3_zero_row_render_methods() -> None:
    """
    QUALITY-02: every migrated module's render methods MUST handle a
    zero-row ModuleData without raising and MUST return the contracted
    empty/N-A representations.

    For each of the four migrated module classes, build an empty
    ModuleData fixture (empty metrics + empty table_data + empty
    analyst_rows + driver=NO_DATA_DRIVER + rag_strip with status=no_data
    + metadata['email_gauge_b64']="") and assert:
      - render_email_panel returns a string containing 'No data' AND
        the NO_DATA_HEADLINE em-dash AND no 'cid:' reference
      - render_analyst_tabs returns []
      - render_rag_strip_entry returns a dict whose 'rag_color' is the
        gray STATUS_COLOR['no_data'] (the BaseModule default honors
        data.rag_strip when populated)
      - No exception raised by any of the three calls
    """
    from reports.modules.base import ModuleData, ModuleConfig  # noqa: PLC0415
    from reports.modules.rag_utils import build_rag_strip_entry  # noqa: PLC0415

    module_classes = [
        ScanCoverageSLAModule,
        CriticalRemediationSLAModule,
        HighRiskAssetsModule,
        AgedVulnsAssetsModule,
    ]

    for cls in module_classes:
        instance = cls()
        cfg      = ModuleConfig(module_id=instance.MODULE_ID)
        empty_md = ModuleData(
            module_id        = instance.MODULE_ID,
            display_name     = instance.DISPLAY_NAME,
            metrics          = {},
            table_data       = [],
            chart_data       = {},
            summary_text     = "",
            metadata         = {"email_gauge_b64": ""},
            error            = None,
            driver_narrative = NO_DATA_DRIVER,
            analyst_rows     = [],
            rag_strip        = build_rag_strip_entry(
                display_name       = instance.DISPLAY_NAME,
                headline_value_str = NO_DATA_HEADLINE,
                status             = "no_data",
            ),
        )

        # render_email_panel
        panel = instance.render_email_panel(empty_md, cfg)
        assert isinstance(panel, str), f"{cls.__name__}: render_email_panel must return str"
        assert "No data" in panel, f"{cls.__name__}: empty panel missing 'No data' text"
        assert NO_DATA_HEADLINE in panel, f"{cls.__name__}: empty panel missing em-dash"
        assert "cid:" not in panel, f"{cls.__name__}: empty panel must NOT reference cid (no gauge for empty)"

        # render_analyst_tabs
        tabs = instance.render_analyst_tabs(empty_md, cfg)
        assert tabs == [], f"{cls.__name__}: render_analyst_tabs on empty must return []"

        # render_rag_strip_entry — inherited default honors data.rag_strip
        cell = instance.render_rag_strip_entry(empty_md, cfg)
        assert isinstance(cell, dict), f"{cls.__name__}: render_rag_strip_entry must return dict"
        assert cell.get("rag_color") == STATUS_COLOR["no_data"], (
            f"{cls.__name__}: empty rag cell color must be {STATUS_COLOR['no_data']!r}, "
            f"got {cell.get('rag_color')!r}"
        )


def check_9_phase3_populated_render_methods() -> None:
    """
    Phase 3 happy-path: each migrated module produces a non-empty
    panel + at-least-one analyst tab + a non-gray rag cell when given a
    populated ModuleData fixture mirroring its compute() shape.

    W7 — headline_value_str is built via ``safe_pct(pct)`` (the
    CLAUDE.md Empty-data guard pattern rule 1). Even though `pct` is a
    known float in this test, the fixture code is what new contributors
    will copy when adding tests for new modules; modeling the wrong
    pattern (an inline percent-precision f-string spec) would propagate
    the bug.
    """
    import pandas as pd  # noqa: PLC0415
    from reports.modules.base import ModuleData, ModuleConfig  # noqa: PLC0415
    from reports.modules.rag_utils import build_rag_strip_entry  # noqa: PLC0415

    cases = [
        # (cls, metrics_key, headline_pct, status, sample_b64, sheet_name)
        (ScanCoverageSLAModule,        "scan_coverage_pct",   97.3, "green",  "iVBORw0KGgo=", "Scan Coverage Detail"),
        (CriticalRemediationSLAModule, "remediation_sla_pct", 92.4, "yellow", "iVBORw0KGgo=", "Critical Remediation Detail"),
        (HighRiskAssetsModule,         "high_risk_pct",       0.4,  "green",  "iVBORw0KGgo=", "High-Risk Assets Detail"),
        (AgedVulnsAssetsModule,        "aged_assets_pct",     2.7,  "yellow", "iVBORw0KGgo=", "Aged Vulns Detail"),
    ]

    for cls, metric_key, pct, status, b64, sheet_name in cases:
        instance = cls()
        cfg      = ModuleConfig(module_id=instance.MODULE_ID)
        sample_df = pd.DataFrame({"col": [1, 2]})
        populated_md = ModuleData(
            module_id        = instance.MODULE_ID,
            display_name     = instance.DISPLAY_NAME,
            metrics          = {metric_key: pct, "status": status},
            table_data       = [],
            chart_data       = {},
            summary_text     = "",
            metadata         = {"email_gauge_b64": b64},
            error            = None,
            driver_narrative = f"Test driver line for {instance.MODULE_ID}.",
            analyst_rows     = [(sheet_name, sample_df)],
            rag_strip        = build_rag_strip_entry(
                display_name       = instance.DISPLAY_NAME,
                # W7 — None/NaN-safe formatter (NOT an inline
                # percent-precision spec). See CLAUDE.md Empty-data
                # guard pattern rule 1. safe_pct returns
                # NO_DATA_HEADLINE for None/NaN; for a float it
                # produces an equivalent "{:.1f}%" output but it is
                # the correct pattern for fixture code.
                headline_value_str = safe_pct(pct),
                status             = status,
            ),
        )

        # Panel
        panel = instance.render_email_panel(populated_md, cfg)
        expected_cid = f"cid:{instance.MODULE_ID}_gauge"
        assert expected_cid in panel, f"{cls.__name__}: populated panel must reference {expected_cid!r}"
        assert "width:620px" in panel, f"{cls.__name__}: panel missing 620px width"
        assert "Test driver line" in panel, f"{cls.__name__}: driver narrative not interpolated"
        assert "<style" not in panel.lower(), f"{cls.__name__}: panel must use inline CSS only — no <style> blocks"

        # Analyst tabs
        tabs = instance.render_analyst_tabs(populated_md, cfg)
        assert len(tabs) == 1, f"{cls.__name__}: expected 1 analyst tab, got {len(tabs)}"
        got_name, got_df = tabs[0]
        assert got_name == sheet_name, f"{cls.__name__}: tab name {got_name!r} != {sheet_name!r}"
        assert got_df is sample_df or got_df.equals(sample_df), f"{cls.__name__}: tab DataFrame round-trip"

        # RAG cell
        cell = instance.render_rag_strip_entry(populated_md, cfg)
        assert cell.get("rag_color") != STATUS_COLOR["no_data"], (
            f"{cls.__name__}: populated cell must NOT be gray; got {cell}"
        )


def check_10_phase3_bundle_email_inline_images_key() -> None:
    """
    Plan 03-01 contract: ReportComposer.run_full_pipeline returns a
    bundle dict that includes the new ``email_inline_images`` key
    alongside the Phase 2 seven keys.
    """
    import inspect  # noqa: PLC0415
    from reports.modules.composer import ReportComposer  # noqa: PLC0415

    src = inspect.getsource(ReportComposer.run_full_pipeline)
    assert "email_inline_images" in src, (
        "run_full_pipeline must populate bundle['email_inline_images'] "
        "(Plan 03-01 contract)"
    )

    # Best-effort runtime check via collect_email_inline_images accessor
    rc = _make_composer()  # empty module_configs; same constructor used by checks 1-7
    entries = rc.collect_email_inline_images([])
    assert entries == [], f"empty results -> empty entries; got {entries!r}"


# ===========================================================================
# Driver
# ===========================================================================

CHECKS = [
    ("D-22 bundle shape",                              check_1_bundle_shape),
    ("D-27 module ordering across channels",           check_2_ordering),
    ("D-29 page-2 strip + cover stability",            check_3_page2_strip_and_cover_stability),
    ("D-29 main-Excel content hash stability",         check_4a_excel_content_hash_stability),
    ("D-29 main-Excel mtime-normalized byte stability", check_4b_excel_byte_hash_stability_mtime_normalized),
    ("D-28 email panel exception isolation",           check_5_email_panel_isolation),
    ("D-28 analyst tabs exception isolation",          check_6_analyst_tabs_isolation),
    ("Phase3 QUALITY-02 zero-row render methods",      check_8_phase3_zero_row_render_methods),
    ("Phase3 populated render methods",                check_9_phase3_populated_render_methods),
    ("Phase3 bundle email_inline_images key",          check_10_phase3_bundle_email_inline_images_key),
]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Phase 2 ReportComposer regression + isolation bar."
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Print full traceback for any failed check.",
    )
    args = parser.parse_args()

    print("=" * 70)
    print("Phase 2 ReportComposer Pipeline — regression + isolation checks")
    print("=" * 70)

    failed  = 0
    skipped = 0
    for name, fn in CHECKS:
        try:
            fn()
            print(f"[PASS] {name}")
        except _SkipCheck as exc:
            skipped += 1
            print(f"[SKIP] {name}: {exc}")
        except AssertionError as exc:
            failed += 1
            print(f"[FAIL] {name}: {exc}")
            if args.verbose:
                traceback.print_exc()
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"[ERROR] {name}: {type(exc).__name__}: {exc}")
            if args.verbose:
                traceback.print_exc()

    print("-" * 70)
    passed = len(CHECKS) - failed - skipped
    print(f"Result: {passed}/{len(CHECKS)} passed, {skipped} skipped, {failed} failed.")

    # SKIP is non-fatal (D-29 deviation is documented in 02-05 must_haves)
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
