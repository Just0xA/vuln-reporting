"""
tests/unit/test_vuln_type_distribution.py — Unit tests for VulnTypeDistributionModule.

Covers:
- VTD-01 classify() function: family-override-first, CPE part a>o>h precedence,
  "Other" fallback, labelled samples from the spec.
- compute(): Hardware hidden when count == 0; within-tag pct math; /0 guard.
- Empty-data guard: zero-row input does not raise in any channel.
"""
from __future__ import annotations

from datetime import datetime, timezone

import pandas as pd
import pytest

from reports.modules.vuln_type_distribution_module import classify, VulnTypeDistributionModule
from reports.modules.base import ModuleConfig, ModuleData

pytestmark = pytest.mark.unit

_NOW = datetime.now(tz=timezone.utc)


# ===========================================================================
# classify() — labelled samples
# ===========================================================================

@pytest.mark.parametrize("plugin_family,cpe,expected", [
    # ── Family override: Linux distro families → OS even with app CPE ─────
    ("Red Hat Local Security Checks",  "cpe:/a:redhat:foo",         "OS"),
    ("Local Security Checks",          "cpe:/a:openssl:openssl:1.0","OS"),
    ("Ubuntu Local Security Checks",   "cpe:/a:canonical:ubuntu",   "OS"),
    ("Debian Local Security Checks",   "cpe:/a:debian:python",      "OS"),
    ("CentOS Local Security Checks",   "cpe:/a:centos:bash",        "OS"),
    ("Oracle Linux Local Security Checks","cpe:/a:oracle:binutils", "OS"),
    ("Rocky Linux Local Security Checks","cpe:/a:rocky:kernel",     "OS"),
    ("Alma Linux Local Security Checks","cpe:/a:alma:glibc",        "OS"),
    ("Fedora Local Security Checks",   "cpe:/a:fedora:python",      "OS"),
    ("Amazon Linux AMI Local Security Checks","cpe:/a:amazon:bash", "OS"),
    ("SuSE Local Security Checks",     "cpe:/a:suse:openssl",       "OS"),
    # ── Family override: Microsoft Bulletins → OS ─────────────────────────
    ("Microsoft Bulletin",             "cpe:/a:microsoft:office",   "OS"),
    ("Microsoft Bulletins",            "cpe:/a:microsoft:word",     "OS"),
    # ── Windows third-party apps: family does NOT match OS_FAMILY override ─
    ("Windows",                        "cpe:/a:adobe:reader",       "Application"),
    ("Windows",                        "cpe:/a:google:chrome",      "Application"),
    ("SMTP problems",                  "cpe:/a:sendmail:sendmail",  "Application"),
    # ── CPE 2.2 and 2.3 formats for Application ───────────────────────────
    ("",                               "cpe:/a:openssl:openssl",    "Application"),
    ("",                               "cpe:2.3:a:openssl:openssl:1.0:*:*:*:*:*:*:*", "Application"),
    # ── CPE 2.2 and 2.3 formats for OS ────────────────────────────────────
    ("",                               "cpe:/o:linux:linux_kernel", "OS"),
    ("",                               "cpe:2.3:o:linux:linux_kernel:5.0:*:*:*:*:*:*:*", "OS"),
    # ── Hardware CPE ──────────────────────────────────────────────────────
    ("",                               "cpe:/h:cisco:router",       "Hardware"),
    ("",                               "cpe:2.3:h:cisco:cisco_ios:*:*:*:*:*:*:*:*", "Hardware"),
    # ── Missing / unparseable CPE → Other ─────────────────────────────────
    ("",                               None,                        "Other"),
    ("",                               "",                          "Other"),
    ("Web Application Scanning",       "",                          "Other"),
    ("General",                        "not-a-cpe",                 "Other"),
    # ── Mixed CPE: a > o > h precedence ───────────────────────────────────
    ("",   "cpe:/a:openssl:foo,cpe:/o:linux:kernel",     "Application"),  # a wins
    ("",   "cpe:/o:linux:kernel,cpe:/h:cisco:router",    "OS"),           # o wins over h
    ("",   "cpe:/h:cisco:router,cpe:/o:linux:kernel",    "OS"),           # o wins over h (order irrelevant)
])
def test_classify(plugin_family, cpe, expected):
    result = classify(plugin_family, cpe)
    assert result == expected, (
        f"classify({plugin_family!r}, {cpe!r}) expected {expected!r}, got {result!r}"
    )


def test_classify_none_family_none_cpe():
    """Both None inputs must yield 'Other' without raising."""
    assert classify(None, None) == "Other"


def test_classify_returns_other_not_unclassified():
    """Fallback label must be 'Other', NOT 'Unclassified' (spec D4)."""
    result = classify("", "")
    assert result == "Other"
    assert result != "Unclassified"


# ===========================================================================
# compute() — within-tag math and Hardware-hidden rule
# ===========================================================================

def _make_vulns_df(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _run_compute(vulns_df: pd.DataFrame) -> ModuleData:
    inst = VulnTypeDistributionModule()
    return inst.compute(
        vulns_df,
        pd.DataFrame(),
        _NOW,
        ModuleConfig("vuln_type_distribution"),
    )


class TestComputeMath:

    def test_within_tag_pct(self):
        """pct = count / tag_total * 100; sum = 100% (within-tag, not env)."""
        df = _make_vulns_df([
            {"state": "open",     "plugin_family": "Windows",   "cpe": "cpe:/a:adobe:reader"},   # App
            {"state": "open",     "plugin_family": "Windows",   "cpe": "cpe:/a:google:chrome"},  # App
            {"state": "open",     "plugin_family": "Red Hat Local Security Checks", "cpe": "cpe:/a:redhat:bash"},  # OS
            {"state": "open",     "plugin_family": "",          "cpe": "cpe:/o:linux:kernel"},   # OS
            {"state": "reopened", "plugin_family": "",          "cpe": ""},                       # Other
            {"state": "fixed",    "plugin_family": "Windows",   "cpe": "cpe:/a:adobe:acrobat"},  # excluded
        ])
        data = _run_compute(df)
        assert data.error is None
        m = data.metrics
        assert m["tag_total"] == 5
        assert m["application_count"] == 2
        assert m["os_count"]          == 2
        assert m["other_count"]       == 1
        assert m["hardware_count"]    == 0
        assert pytest.approx(m["application_pct"], rel=1e-5) == 40.0
        assert pytest.approx(m["os_pct"],          rel=1e-5) == 40.0
        assert pytest.approx(m["other_pct"],       rel=1e-5) == 20.0

    def test_hardware_hidden_when_zero(self):
        """hide_hardware must be True when Hardware count == 0."""
        df = _make_vulns_df([
            {"state": "open", "plugin_family": "Windows", "cpe": "cpe:/a:adobe:reader"},
        ])
        data = _run_compute(df)
        assert data.metrics["hide_hardware"] is True

    def test_hardware_not_hidden_when_nonzero(self):
        """hide_hardware must be False when Hardware count > 0."""
        df = _make_vulns_df([
            {"state": "open", "plugin_family": "", "cpe": "cpe:/h:cisco:router"},
        ])
        data = _run_compute(df)
        assert data.metrics["hide_hardware"] is False

    def test_divide_by_zero_guard_empty_df(self):
        """Zero-row input must not raise; all pcts must be 0.0."""
        data = _run_compute(pd.DataFrame(columns=["state", "plugin_family", "cpe"]))
        assert data.error is None
        for bucket in ("application", "os", "hardware", "other"):
            assert data.metrics[f"{bucket}_pct"] == 0.0

    def test_reopened_included(self):
        """'reopened' state must be included; 'fixed' excluded."""
        df = _make_vulns_df([
            {"state": "open",     "plugin_family": "Windows", "cpe": "cpe:/a:adobe:reader"},
            {"state": "reopened", "plugin_family": "Windows", "cpe": "cpe:/a:google:chrome"},
            {"state": "fixed",    "plugin_family": "Windows", "cpe": "cpe:/a:ms:word"},
        ])
        data = _run_compute(df)
        assert data.metrics["tag_total"] == 2
        assert data.metrics["application_count"] == 2


# ===========================================================================
# render_pdf_section — Hardware row omission
# ===========================================================================

class TestRenderHardwareOmission:

    def test_hardware_row_omitted_in_pdf_when_zero(self):
        """When hide_hardware is True the PDF section must not include 'Hardware'."""
        df = _make_vulns_df([
            {"state": "open", "plugin_family": "Windows", "cpe": "cpe:/a:adobe:reader"},
        ])
        inst = VulnTypeDistributionModule()
        data = inst.compute(df, pd.DataFrame(), _NOW, ModuleConfig("vuln_type_distribution"))
        cfg  = ModuleConfig("vuln_type_distribution")
        html = inst.render_pdf_section(data, cfg)
        # The word "Hardware" should only appear in the explanatory note, not as a table row label
        # (allow it once for "Hardware omitted" notice but not twice as a cell value)
        assert data.metrics["hide_hardware"] is True
        # The row label "Hardware" must not appear as a table data cell
        assert "<td" not in html or html.count(">Hardware<") == 0

    def test_hardware_row_present_in_excel_when_nonzero(self):
        """When Hardware count > 0 the Excel tab must include the Hardware row."""
        from openpyxl import Workbook
        df = _make_vulns_df([
            {"state": "open", "plugin_family": "", "cpe": "cpe:/h:cisco:router"},
        ])
        inst = VulnTypeDistributionModule()
        data = inst.compute(df, pd.DataFrame(), _NOW, ModuleConfig("vuln_type_distribution"))
        cfg  = ModuleConfig("vuln_type_distribution")
        wb   = Workbook()
        tabs = inst.render_excel_tabs(data, wb, cfg)
        assert len(tabs) == 1
        ws = wb[tabs[0]]
        values = [ws.cell(row=r, column=1).value for r in range(1, 10) if ws.cell(row=r, column=1).value]
        assert "Hardware" in values


# ===========================================================================
# Empty-data guard — all channels, zero-row input
# ===========================================================================

class TestEmptyDataGuard:

    def test_all_channels_no_raise(self):
        """Zero-row input must not raise in any channel."""
        from openpyxl import Workbook

        inst     = VulnTypeDistributionModule()
        empty_df = pd.DataFrame(columns=["state", "plugin_family", "cpe"])
        data     = inst.compute(empty_df, pd.DataFrame(), _NOW, ModuleConfig("vuln_type_distribution"))
        cfg      = ModuleConfig("vuln_type_distribution")

        pdf   = inst.render_pdf_section(data, cfg)
        tabs  = inst.render_excel_tabs(data, Workbook(), cfg)
        panel = inst.render_email_panel(data, cfg)
        atabs = inst.render_analyst_tabs(data, cfg)
        strip = inst.render_rag_strip_entry(data, cfg)

        assert isinstance(pdf,   str)
        assert isinstance(tabs,  list)
        assert isinstance(panel, str)
        assert isinstance(atabs, list)
        assert isinstance(strip, dict)
        assert {"label", "headline_value", "rag_color", "rag_label"} <= set(strip)

    def test_empty_strip_is_no_data(self):
        """Zero-row input must produce a gray 'No Data' RAG strip cell."""
        inst     = VulnTypeDistributionModule()
        empty_df = pd.DataFrame(columns=["state", "plugin_family", "cpe"])
        data     = inst.compute(empty_df, pd.DataFrame(), _NOW, ModuleConfig("vuln_type_distribution"))
        strip    = inst.render_rag_strip_entry(data, ModuleConfig("vuln_type_distribution"))
        assert strip["rag_label"] == "No Data"
        assert strip["headline_value"] == "—"
