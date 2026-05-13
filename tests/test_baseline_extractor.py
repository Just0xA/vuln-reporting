"""
tests/test_baseline_extractor.py — Phase 4 BOARD-08 / D-04-05 (revised) / D-04-08.

Locks tests/baseline_utils.py behavior against the STRUCTURAL-ONLY snapshot
schema. Per revised D-04-05, snapshots contain NO metric values (which drift
daily with vulnerability churn) and NO row-level data — only deterministic
structural shape (counts, booleans, sorted name lists).

This file is committed RED (negative-control proof): without baseline_utils
present, every test fails on import. Task 2's implementation flips them
GREEN. The commit-sequence diff IS the proof that the structural snapshot
+ PII guard work — there is no manual revert-and-retest required.

HTML markers verified against actual production source (composer.py:677,
871; scan_coverage_sla_module.py et al.):
  - email panels emit <table role="presentation" ...> exactly once per panel
  - PDF cover page is NOT separated from module 1 by an HTML page-break div;
    cover separation relies on CSS page-break-after on .report-cover, so the
    byte-level PDF page count is the authoritative number.

Run: python tests/test_baseline_extractor.py
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pandas as pd  # noqa: E402

from tests.baseline_utils import (    # noqa: E402   — RED until Task 2 lands
    extract_structural_snapshot,
    compare_snapshots,
    _PII_EXACT,
    _PII_SUBSTRING,
    _is_pii_column,
)

FAILED: list[str] = []

EXPECTED_KEYS = frozenset({
    "schema_version",
    "group_slug",
    "pdf_page_count",
    "pdf_has_risk_status_summary_header",
    "pdf_rag_cell_count",
    "excel_tab_names_sorted",
    "email_panel_count",
    "email_inline_image_cids_per_module",
    "bundle_keys_present",
    "analyst_excel_present",
    "rag_cells_all_no_data",
    "panel_drivers_all_no_data_in_scope",
})

FORBIDDEN_KEYS = frozenset({
    "headline_metrics",
    "scan_coverage_pct",
    "critical_remediation_sla_pct",
    "high_risk_assets_pct",
    "aged_vulns_assets_pct",
    "excel_tabs",          # the row-count-per-tab mapping is forbidden
    "row_count",
    "tab_row_counts",
})


def _check(name: str, cond: bool, hint: str = "") -> None:
    if cond:
        print(f"PASS  {name}")
    else:
        FAILED.append(name)
        print(f"FAIL  {name}{(': ' + hint) if hint else ''}")


def _build_pdf_html_n_pages(n_pages: int, *, all_no_data: bool = False) -> str:
    """Build a WeasyPrint-renderable HTML for `n_pages` actual rendered pages.

    Uses CSS `page-break-before: always` on sibling divs to force page breaks
    so the rendered byte stream has exactly `n_pages` pages. The cover page
    has the "Key Performance Metrics" header and 4 RAG cells (matching the
    production board_summary cover); module pages each have a body div.
    """
    # Use the actual STATUS_COLOR['no_data'] sentinel so the extractor's
    # rag_cells_all_no_data detector matches against the real color.
    try:
        from reports.modules.rag_utils import STATUS_COLOR
        no_data_color = STATUS_COLOR.get("no_data", "#cccccc")
    except Exception:
        no_data_color = "#cccccc"
    rag_color = no_data_color if all_no_data else "#22aa55"
    rag_cells = (
        f'<div class="rag-cell" style="background-color: {rag_color}"></div>'
        * 4
    )
    cover = (
        '<div class="report-cover" style="page-break-after: always;">'
        '<h2>Key Performance Metrics</h2>'
        f'{rag_cells}'
        '</div>'
    )
    modules = []
    for i in range(n_pages - 1):
        modules.append(
            f'<div style="page-break-before: always;"><h3>Module {i+1}</h3></div>'
        )
    return (
        '<!DOCTYPE html><html><head><meta charset="utf-8"></head><body>'
        + cover + "".join(modules) +
        '</body></html>'
    )


def _build_email_body_n_panels(n_panels: int, *, all_no_data: bool = False) -> str:
    """Build email_body_html with N `role="presentation"` table panels.

    Mirrors scan_coverage_sla_module.py panel rendering — each panel is
    exactly one `<table role="presentation" ...>` with a driver-narrative div.
    """
    driver = "No data in scope." if all_no_data else "Coverage at 95%."
    panels = []
    for _ in range(n_panels):
        panels.append(
            '<table role="presentation" cellpadding="0" cellspacing="0" border="0" '
            'style="width:620px; max-width:620px;">'
            f'<tr><td>{driver}</td></tr></table>'
        )
    return "\n".join(panels)


def _minimal_bundle(*, analyst_excel=None, all_no_data: bool = False,
                    extra_df_in_module: pd.DataFrame | None = None,
                    n_pages: int = 5, n_panels: int = 4) -> dict:
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Scan Coverage Detail")
    wb.create_sheet("_Metadata")

    pdf_html = _build_pdf_html_n_pages(n_pages, all_no_data=all_no_data)
    email_body = _build_email_body_n_panels(n_panels, all_no_data=all_no_data)

    class MD:
        module_id = "scan_coverage_sla"
        display_name = "Scan Coverage SLA"
        metrics = {"scan_coverage_pct": 95.5}
        analyst_rows = []
        if extra_df_in_module is not None:
            analyst_rows = [("Sheet With PII", extra_df_in_module)]

    return {
        "pdf_html": pdf_html,
        "excel_workbook": wb,
        "analyst_workbook_path": analyst_excel,
        "analyst_excel": analyst_excel,
        "email_body_html": email_body,
        "email_kpis": [],
        "email_inline_images": [
            {"cid": "scan_coverage_sla_gauge", "b64_png": "x"},
            {"cid": "critical_remediation_sla_gauge", "b64_png": "x"},
            {"cid": "high_risk_assets_gauge", "b64_png": "x"},
            {"cid": "aged_vulns_assets_gauge", "b64_png": "x"},
        ],
        "errors": [],
        "module_results": [MD()],
    }


def main() -> int:
    # A — keyset is exactly the structural-only set (no metric values)
    snap = extract_structural_snapshot(_minimal_bundle(), "test_pull")
    actual_keys = set(snap.keys())
    _check(
        "A_keyset_exact",
        actual_keys == EXPECTED_KEYS,
        hint=f"missing={EXPECTED_KEYS - actual_keys} extra={actual_keys - EXPECTED_KEYS}",
    )
    _check(
        "A_no_forbidden_keys",
        not (actual_keys & FORBIDDEN_KEYS),
        hint=f"forbidden present: {actual_keys & FORBIDDEN_KEYS}",
    )

    # B — PII guard
    for col in ("hostname", "ipv4", "fqdn", "plugin_name"):
        bad = pd.DataFrame({col: ["x"]})
        try:
            extract_structural_snapshot(_minimal_bundle(extra_df_in_module=bad), "test_pull")
            _check(f"B_pii_{col}", False, "no exception raised")
        except ValueError as exc:
            _check(f"B_pii_{col}", col in str(exc).lower(), hint=str(exc))

    # B substring backstop — host_asset_name rejected
    bad = pd.DataFrame({"host_asset_name": ["x"]})
    try:
        extract_structural_snapshot(_minimal_bundle(extra_df_in_module=bad), "test_pull")
        _check("B_pii_substring_asset_name", False, "no exception raised")
    except ValueError as exc:
        _check("B_pii_substring_asset_name", "asset_name" in str(exc).lower(), hint=str(exc))

    # B substring no-false-positive — tag_name should NOT trigger
    _check(
        "B_pii_no_false_positive_tag_name",
        not _is_pii_column("tag_name"),
        hint="_is_pii_column('tag_name') returned True (false positive)",
    )
    _check(
        "B_pii_no_false_positive_email_panel_count",
        not _is_pii_column("email_panel_count"),
        hint="_is_pii_column('email_panel_count') returned True (false positive)",
    )

    # C — analyst_off shape
    snap = extract_structural_snapshot(_minimal_bundle(analyst_excel=None), "test_pull_analyst_off")
    _check(
        "C_analyst_off_present_false",
        snap["analyst_excel_present"] is False,
        hint=str(snap.get("analyst_excel_present")),
    )

    # D — zero-match shape
    snap = extract_structural_snapshot(_minimal_bundle(all_no_data=True), "test_pull_zero_match")
    _check(
        "D_rag_cells_all_no_data_true",
        snap["rag_cells_all_no_data"] is True,
        hint=str(snap.get("rag_cells_all_no_data")),
    )
    _check(
        "D_panel_drivers_all_no_data_in_scope_true",
        snap["panel_drivers_all_no_data_in_scope"] is True,
        hint=str(snap.get("panel_drivers_all_no_data_in_scope")),
    )

    # E — compare match
    d = extract_structural_snapshot(_minimal_bundle(), "test_pull")
    _check("E_compare_match", compare_snapshots(d, d) == [], hint=str(compare_snapshots(d, d)))

    # F — compare drift on pdf_page_count
    d2 = dict(d)
    d2["pdf_page_count"] = d["pdf_page_count"] + 1
    diffs = compare_snapshots(d, d2)
    _check(
        "F_compare_drift_pdf_page_count",
        any("pdf_page_count" in line for line in diffs),
        hint=str(diffs),
    )

    # G — compare detects extra key in actual
    d_extra = dict(d)
    d_extra["surprise_new_key"] = 42
    diffs = compare_snapshots(d_extra, d)
    _check(
        "G_compare_extra_actual_key",
        any("surprise_new_key" in line and "missing from baseline" in line for line in diffs),
        hint=str(diffs),
    )

    # H — pdf_page_count is byte-authoritative (5 pages for default fixture)
    snap = extract_structural_snapshot(_minimal_bundle(n_pages=5), "test_pull")
    _check(
        "H_pdf_page_count_5",
        snap["pdf_page_count"] == 5,
        hint=f"actual={snap.get('pdf_page_count')} expected=5 (cover + 4 modules; byte-stream count)",
    )

    # I — email_panel_count uses role="presentation" not class="module-panel"
    snap = extract_structural_snapshot(_minimal_bundle(n_panels=4), "test_pull")
    _check(
        "I_email_panel_count_4",
        snap["email_panel_count"] == 4,
        hint=f"actual={snap.get('email_panel_count')} expected=4 (role=\"presentation\" per panel)",
    )
    # Negative control: a body string with class="module-panel" but NO
    # role="presentation" should yield 0, proving the legacy literal is not used.
    legacy_body_bundle = _minimal_bundle()
    legacy_body_bundle["email_body_html"] = (
        '<div class="module-panel">x</div>' * 4
    )
    snap = extract_structural_snapshot(legacy_body_bundle, "test_pull")
    _check(
        "I_email_panel_count_legacy_class_not_counted",
        snap["email_panel_count"] == 0,
        hint=f"legacy class=\"module-panel\" should NOT count; got {snap.get('email_panel_count')}",
    )

    if FAILED:
        print(f"\n{len(FAILED)} check(s) failed: {FAILED}")
        return 1
    print("\nAll checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
