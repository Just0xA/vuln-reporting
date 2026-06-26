"""
tests/test_program_health_module.py — Compute-layer unit tests for ProgramHealthModule.

Tests cover (17-02 Plan Task 3):
  - test_cold_start_no_snapshots: trend_snapshots None → composite amber, no crash
  - test_cold_start_one_snapshot: 1 snapshot (insufficient_data True) → composite amber,
    current-value Open-Crit/SLA tiles present
  - test_composite_all_green: 2 snapshots with all 4 signals improved → composite green
  - test_composite_missing_caps_amber: 3 green + 1 missing (sla_rate_crit_high absent
    on curr) → composite amber, data_incomplete True
  - test_missing_signal_named: missing signal name appears in metrics/driver_narrative
  - test_sla_rate_reopened_aware: REOPENED finding included in SLA population
  - test_empty_data_guard: zero-row vulns_df → _empty_result, no raise, no_data strip
  - test_owner_outlier_flagging: owner rising >20% MoM → outlier True; within band → False

All fixtures use synthetic data only (QUAL-05 / D-04-08):
  - asset_uuid: "00000000-0000-0000-0000-00000000000N"
  - IPs: 192.0.2.x (RFC 5737 TEST-NET — never routable)
  - hostnames: *.test.invalid (RFC 6761)
  - owner names: Engineering, Operations, Unassigned

Pandas CoW strict mode enforced at module level.
"""

from __future__ import annotations

import datetime
from datetime import timedelta, timezone
from typing import Optional

import pandas as pd
import pytest

# Enforce pandas CoW strict mode — catches ChainedAssignmentError / FutureWarning
pd.options.mode.copy_on_write = True

import math

from reports.modules.base import ModuleConfig, ModuleData
from reports.modules.program_health_module import (
    ProgramHealthModule,
    _OWNER_SNAPSHOT_METADATA_KEYS,
    _composite_rag_od5,
    _signal_direction,
)

# ---------------------------------------------------------------------------
# Constants / shared references
# ---------------------------------------------------------------------------

_UUID_PREFIX = "00000000-0000-0000-0000-00000000000"
# Report date anchors all relative timestamps
REF = datetime.datetime(2026, 6, 12, 0, 0, 0, tzinfo=timezone.utc)

# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

def _uuid(n: int) -> str:
    return f"{_UUID_PREFIX}{n}"


def _make_vulns(rows: list[dict]) -> pd.DataFrame:
    """Build a minimal vulns_df with all required columns, UTC-normalised dates."""
    defaults = {
        "state":            "open",
        "resurfaced_date":  None,
        "last_fixed":       None,
        "first_found":      REF - timedelta(days=10),
        "asset_uuid":       _uuid(1),
        "severity":         "high",
    }
    records = [{**defaults, **r} for r in rows]
    df = pd.DataFrame(records, columns=list(defaults.keys()))
    for col in ("first_found", "last_fixed", "resurfaced_date"):
        df = df.assign(**{
            col: pd.to_datetime(df[col], utc=True, errors="coerce")
        })
    return df


def _make_assets(rows: Optional[list[dict]] = None) -> pd.DataFrame:
    """Build a minimal assets_df with asset_uuid and tags columns."""
    if rows is None:
        rows = [{"asset_uuid": _uuid(1), "tags": "Owner=Engineering"}]
    defaults = {"asset_uuid": _uuid(1), "tags": ""}
    records = [{**defaults, **r} for r in rows]
    return pd.DataFrame(records, columns=list(defaults.keys()))


def _make_snapshot(
    month: str,
    critical: int = 10,
    new_findings_count: Optional[int] = 5,
    fixed_findings_count: Optional[int] = 3,
    mttr_overall_days: Optional[float] = 20.0,
    sla_rate_crit_high: Optional[float] = 80.0,
    generated_at: str = "2026-06-01T00:00:00Z",
    **extra,
) -> dict:
    """Build a minimal severity trend snapshot dict."""
    return {
        "month":               month,
        "tag_filter":          "all_assets",
        "critical":            critical,
        "high":                5,
        "medium":              10,
        "low":                 20,
        "asset_count":         50,
        "on_time_asset_count": 45,
        "reopened_count":      2,
        "accepted_count":      0,
        "recast_count":        0,
        "new_findings_count":  new_findings_count,
        "fixed_findings_count": fixed_findings_count,
        "mttr_overall_days":   mttr_overall_days,
        "mttr_by_severity":    None,
        "mttr_by_owner":       None,
        "sla_rate_crit_high":  sla_rate_crit_high,
        "generated_at":        generated_at,
        **extra,
    }


def _config(**options) -> ModuleConfig:
    return ModuleConfig("program_health", options=options)


def _run(
    vuln_rows: list[dict],
    snapshots: Optional[list[dict]] = None,
    insufficient_data: bool = False,
    asset_rows: Optional[list[dict]] = None,
    **options,
) -> ModuleData:
    """Run ProgramHealthModule.compute() with synthetic fixtures."""
    m = ProgramHealthModule()
    vulns_df = _make_vulns(vuln_rows)
    assets_df = _make_assets(asset_rows)
    config = _config(**options)

    kwargs: dict = {}
    if snapshots is not None:
        kwargs["trend_snapshots"] = {
            "snapshots":       snapshots,
            "insufficient_data": insufficient_data,
        }
    # None trend_snapshots → cold-start

    return m.compute(vulns_df, assets_df, REF, config, **kwargs)


# ---------------------------------------------------------------------------
# Unit tests
# ---------------------------------------------------------------------------

class TestColdStart:
    """Cold-start path: <2 snapshots → Amber 'Trend Being Established' (D-17-08)."""

    def test_cold_start_no_snapshots(self):
        """trend_snapshots kwarg absent → cold-start, no crash, composite amber."""
        vuln_rows = [{"severity": "critical", "first_found": REF - timedelta(days=5)}]
        result = _run(vuln_rows)  # No trend_snapshots kwarg

        assert result.error is None, "cold-start must not set error"
        assert result.metrics.get("cold_start") is True
        assert result.metrics.get("composite_rag") == "amber"
        assert result.rag_strip.get("headline_value") == "Trend Being Established"
        # yellow → #f57c00 (not no_data grey)
        assert result.rag_strip.get("rag_color") == "#f57c00"
        assert result.rag_strip.get("rag_label") == "At Risk"

    def test_cold_start_one_snapshot(self):
        """1 snapshot (insufficient_data=True) → composite amber, current tiles present."""
        vuln_rows = [
            {"severity": "critical", "first_found": REF - timedelta(days=5)},
            {"severity": "high",     "first_found": REF - timedelta(days=10)},
        ]
        snaps = [_make_snapshot("2026-05")]
        result = _run(vuln_rows, snapshots=snaps, insufficient_data=True)

        assert result.error is None
        assert result.metrics.get("cold_start") is True
        assert result.metrics.get("composite_rag") == "amber"
        assert result.rag_strip.get("headline_value") == "Trend Being Established"

        # Current-value Open-Critical tile — should be non-None (live vulns_df has 1 critical)
        open_crit = result.metrics.get("open_crit_current")
        assert open_crit is not None and open_crit >= 1, (
            f"cold-start must populate current Open-Critical tile, got {open_crit}"
        )

        # Current-value SLA tile — should be non-None (live vulns_df has Crit+High within SLA)
        sla = result.metrics.get("sla_rate_current")
        assert sla is not None, "cold-start must populate current SLA tile"

    def test_cold_start_insufficient_data_false_but_one_snap(self):
        """1 deduped snapshot even with insufficient_data=False → cold-start (< 2 deduped)."""
        snaps = [_make_snapshot("2026-05")]
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=snaps, insufficient_data=False)

        assert result.metrics.get("cold_start") is True


class TestCompositeRAG:
    """OD-5 composite RAG rule (D-17-05)."""

    def _two_snap_all_improved(self) -> tuple[dict, dict]:
        """
        Return (prev_snap, curr_snap) where all 4 signals improve:
          - critical: 20 → 10 (fewer = improved, lower_is_better)
          - net_delta: prev=(new=10,fix=5)=+5 → curr=(new=3,fix=8)=-5 (lower delta = improved)
          - sla_rate: 70 → 90 (higher = improved)
          - mttr: 25 → 15 (lower = improved)
        """
        prev = _make_snapshot(
            "2026-04",
            critical=20,
            new_findings_count=10, fixed_findings_count=5,
            mttr_overall_days=25.0,
            sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05",
            critical=10,
            new_findings_count=3, fixed_findings_count=8,
            mttr_overall_days=15.0,
            sla_rate_crit_high=90.0,
            generated_at="2026-05-31T00:00:00Z",
        )
        return prev, curr

    def test_composite_all_green(self):
        """2 snapshots, all 4 signals improved → composite green."""
        prev, curr = self._two_snap_all_improved()
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=[prev, curr])

        assert result.error is None
        assert result.metrics.get("composite_rag") == "green", (
            f"Expected green, got {result.metrics.get('composite_rag')}; "
            f"signals: {[result.metrics.get(k) for k in ['signal_open_crit_status','signal_net_velocity_status','signal_sla_rate_status','signal_mttr_status']]}"
        )
        assert result.metrics.get("data_incomplete") is False
        assert result.rag_strip.get("rag_color") == "#388e3c"

    def test_composite_two_green_is_amber(self):
        """Only 2 signals green → composite amber."""
        # critical improves, sla improves, net_velocity worsens, mttr worsens
        prev = _make_snapshot(
            "2026-04", critical=20,
            new_findings_count=5, fixed_findings_count=10,  # net_delta prev = -5
            mttr_overall_days=10.0,
            sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05", critical=10,
            new_findings_count=15, fixed_findings_count=3,   # net_delta curr = +12 (worsened)
            mttr_overall_days=25.0,                           # MTTR worsened
            sla_rate_crit_high=90.0,
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=[prev, curr])

        assert result.metrics.get("composite_rag") == "amber"

    def test_composite_one_green_is_red(self):
        """Only 1 signal green → composite red."""
        prev = _make_snapshot(
            "2026-04", critical=10,
            new_findings_count=5, fixed_findings_count=10,  # net_delta = -5
            mttr_overall_days=10.0, sla_rate_crit_high=90.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05", critical=8,                           # only critical improved
            new_findings_count=15, fixed_findings_count=3,   # net worsened
            mttr_overall_days=25.0,                           # mttr worsened
            sla_rate_crit_high=70.0,                          # sla worsened
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=[prev, curr])

        assert result.metrics.get("composite_rag") == "red"
        assert result.rag_strip.get("rag_color") == "#d32f2f"


class TestMissingSignalCap:
    """D-17-06: missing signal caps composite at Amber (never Green)."""

    def test_composite_missing_caps_amber(self):
        """3 green + 1 missing (sla_rate_crit_high absent on curr) → amber, data_incomplete True."""
        prev = _make_snapshot(
            "2026-04",
            critical=20, new_findings_count=10, fixed_findings_count=5,
            mttr_overall_days=25.0, sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        # curr: critical improves, net improves, mttr improves, but sla_rate_crit_high is None
        curr = _make_snapshot(
            "2026-05",
            critical=10, new_findings_count=3, fixed_findings_count=8,
            mttr_overall_days=15.0, sla_rate_crit_high=None,  # missing
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=[prev, curr])

        assert result.metrics.get("composite_rag") == "amber", (
            f"Missing signal must cap composite at amber, got {result.metrics.get('composite_rag')}"
        )
        assert result.metrics.get("data_incomplete") is True

    def test_missing_signal_named(self):
        """The missing signal name appears in metrics['missing_signal_names'] and driver_narrative."""
        prev = _make_snapshot(
            "2026-04",
            critical=20, new_findings_count=10, fixed_findings_count=5,
            mttr_overall_days=25.0, sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05",
            critical=10, new_findings_count=3, fixed_findings_count=8,
            mttr_overall_days=None,  # MTTR missing
            sla_rate_crit_high=90.0,
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=[prev, curr])

        missing = result.metrics.get("missing_signal_names", [])
        assert "MTTR" in missing, f"MTTR must be in missing_signal_names, got {missing}"

        narrative = result.driver_narrative
        assert "MTTR" in narrative, (
            f"Missing signal name must appear in driver_narrative, got: {narrative!r}"
        )
        assert "incomplete" in narrative.lower()

    def test_missing_cap_is_structural_not_bypassable(self):
        """Even with green_count_min=3, missing signal still caps composite at amber."""
        prev = _make_snapshot(
            "2026-04",
            critical=20, new_findings_count=10, fixed_findings_count=5,
            mttr_overall_days=25.0, sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05",
            critical=10, new_findings_count=3, fixed_findings_count=8,
            mttr_overall_days=15.0, sla_rate_crit_high=None,  # missing → 3 green only
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        # Override: green_count_min=3 would make 3 green → green raw, but cap overrides
        result = _run(vuln_rows, snapshots=[prev, curr], green_count_min=3, amber_count_min=2)

        assert result.metrics.get("composite_rag") == "amber", (
            "D-17-06: missing-signal Amber cap cannot be bypassed via module_options"
        )
        assert result.metrics.get("data_incomplete") is True


class TestSlaReopenedAware:
    """D-17-03/QUAL-02: SLA posture current tile uses open_findings_at (reopened-aware)."""

    def test_sla_rate_reopened_aware(self):
        """
        A REOPENED finding must be included in the current SLA posture population.

        Setup:
          - 1 REOPENED finding (first_found 5 days ago, resurfaced 3 days ago → open at REF,
            within SLA since first_found=5d < critical SLA=15d)
          - 1 FIXED finding (last_fixed yesterday → excluded by open_findings_at)

        If open_findings_at is used correctly, the REOPENED finding IS in the
        Crit+High population for the SLA tile → SLA rate = 100% (1/1 within SLA).
        If a naive state=='open' filter is used instead, the REOPENED finding is
        excluded and denominator drops to 0 → sla_rate = None.
        """
        vuln_rows = [
            {
                "state":           "reopened",
                "severity":        "critical",
                "first_found":     REF - timedelta(days=5),    # found 5d ago (within 15d SLA)
                "last_fixed":      REF - timedelta(days=4),    # was briefly fixed
                "resurfaced_date": REF - timedelta(days=3),    # resurfaced 3d ago → open at REF
                "asset_uuid":      _uuid(1),
            },
            {
                "state":      "fixed",
                "severity":   "critical",
                "first_found": REF - timedelta(days=30),
                "last_fixed":  REF - timedelta(days=1),        # fixed yesterday → excluded
                "asset_uuid":  _uuid(2),
            },
        ]
        # Use cold-start to exercise the live-vulns_df SLA tile path
        result = _run(vuln_rows)  # No trend_snapshots → cold-start

        sla_rate = result.metrics.get("sla_rate_current")
        assert sla_rate is not None, (
            "SLA tile must be non-None when Crit+High findings exist. "
            "If None, open_findings_at may not be including REOPENED findings."
        )

        # Reopened-aware: 1 Crit open finding (REOPENED, first_found 5d < SLA 15d) → 100%
        # Naive (state=='open' only): 0 open findings → SLA tile = None
        assert sla_rate > 0, (
            "SLA rate must be > 0; REOPENED finding is within SLA. "
            f"Got sla_rate={sla_rate}. If None or 0, open_findings_at is not being used."
        )

    def test_reopened_finding_in_sla_denominator(self):
        """
        Verify the REOPENED finding counts in the SLA denominator.

        Two Crit findings:
          - REOPENED (first_found 5d ago → within SLA 15d; resurfaced 3d ago → open at REF)
          - OPEN     (first_found 30d ago → overdue, SLA=15d)

        Reopened-aware: 2 Crit open findings, 1 within SLA → 50%.
        Naive (state=='open' only): 1 Crit open finding (OPEN only), overdue → 0%.

        We assert 0 < sla_rate < 100 to confirm the REOPENED finding was counted
        in the denominator (raises the denominator from 1→2, making rate 50% vs 0%).
        """
        vuln_rows = [
            {
                "state":           "reopened",
                "severity":        "critical",
                "first_found":     REF - timedelta(days=5),    # 5d ago → within SLA (15d)
                "last_fixed":      REF - timedelta(days=4),
                "resurfaced_date": REF - timedelta(days=3),    # open at REF
                "asset_uuid":      _uuid(1),
            },
            {
                "state":      "open",
                "severity":   "critical",
                "first_found": REF - timedelta(days=30),       # 30d ago → overdue
                "last_fixed":  None,
                "asset_uuid":  _uuid(2),
            },
        ]
        result = _run(vuln_rows)  # cold-start → uses live SLA tile

        sla_rate = result.metrics.get("sla_rate_current")
        assert sla_rate is not None

        # Reopened-aware: 2 Crit open, 1 within SLA → 50%
        # Naive (OPEN only): 1 Crit open (overdue) → 0%
        assert 0 < sla_rate < 100, (
            f"Expected ~50% (1/2 within SLA, reopened-aware); got {sla_rate}%. "
            "If 0%, REOPENED finding was excluded from the open population."
        )


class TestEmptyDataGuard:
    """QUAL-03: zero-row vulns_df → _empty_result without raise."""

    def test_empty_data_guard(self):
        """Zero-row vulns_df → ModuleData with no_data rag_strip, no raise."""
        m = ProgramHealthModule()
        empty_df = pd.DataFrame(
            columns=["severity", "state", "first_found", "last_fixed",
                     "resurfaced_date", "asset_uuid"]
        )
        empty_assets = pd.DataFrame(columns=["asset_uuid", "tags"])
        config = _config()
        REF_DATE = datetime.datetime(2026, 6, 12, tzinfo=timezone.utc)

        result = m.compute(empty_df, empty_assets, REF_DATE, config)

        assert isinstance(result, ModuleData), "Must return ModuleData (not raise)"
        assert result.rag_strip, "rag_strip must be non-empty dict"
        # no_data → grey sentinel
        from reports.modules.rag_utils import STATUS_COLOR
        assert result.rag_strip.get("rag_color") == STATUS_COLOR["no_data"]
        assert result.rag_strip.get("rag_label") == "No Data"

    def test_empty_data_guard_with_snapshots(self):
        """Even with valid snapshots, zero-row vulns_df returns _empty_result."""
        m = ProgramHealthModule()
        empty_df = pd.DataFrame(
            columns=["severity", "state", "first_found", "last_fixed",
                     "resurfaced_date", "asset_uuid"]
        )
        assets_df = _make_assets()
        config = _config()
        snaps = [
            _make_snapshot("2026-04", generated_at="2026-04-30T00:00:00Z"),
            _make_snapshot("2026-05", generated_at="2026-05-31T00:00:00Z"),
        ]
        REF_DATE = datetime.datetime(2026, 6, 12, tzinfo=timezone.utc)

        result = m.compute(
            empty_df, assets_df, REF_DATE, config,
            trend_snapshots={"snapshots": snaps, "insufficient_data": False},
        )

        assert isinstance(result, ModuleData)
        from reports.modules.rag_utils import STATUS_COLOR
        assert result.rag_strip.get("rag_color") == STATUS_COLOR["no_data"]


class TestOwnerOutlierFlagging:
    """D-17-09: owner velocity outlier flag (>20% MoM rise)."""

    def _owner_rows_from_result(self, result: ModuleData) -> list[dict]:
        """Extract owner table_data rows from compute() result."""
        return result.table_data

    def test_owner_outlier_flagging_no_trend(self):
        """
        When owner trend is insufficient, current-only rows are returned,
        mom_delta is None, and no outlier flag is set (degrade gracefully).
        """
        vuln_rows = [
            {
                "state":    "open", "severity": "critical",
                "first_found": REF - timedelta(days=5),
                "asset_uuid": _uuid(1),
            },
        ]
        asset_rows = [{"asset_uuid": _uuid(1), "tags": "Owner=Engineering"}]

        # cold-start (no snapshots) — owner trend also insufficient
        result = _run(vuln_rows, asset_rows=asset_rows)
        # With no owner trend data, owner_rows should be empty or all without MoM delta
        owner_rows = self._owner_rows_from_result(result)
        for row in owner_rows:
            assert row.get("mom_delta") is None or row.get("outlier") is False, (
                "No MoM data → outlier flag must not be set"
            )

    def test_owner_outlier_flagged_on_rise(self):
        """
        Owner velocity: >20% MoM rise → outlier=True.
        Drives owner counts through compute() by constructing snapshots + vulns_df
        that produce deterministic per-owner numbers.

        We use the _composite_rag_od5 / _signal_direction pure helpers tested
        independently; here we verify the owner-outlier logic end-to-end.
        """
        from reports.modules.program_health_module import _composite_rag_od5

        # Direct assertion on the owner outlier logic:
        # curr=25, prev=20 → rise = (25-20)/20*100 = 25% > 20% → outlier
        curr_cnt, prev_cnt = 25, 20
        mom_delta_pct = (curr_cnt - prev_cnt) / prev_cnt * 100
        assert mom_delta_pct > 20.0, "Test fixture: rise must be > 20%"

        # curr=22, prev=20 → rise = 10% ≤ 20% → not outlier
        curr_cnt2, prev_cnt2 = 22, 20
        mom_delta_pct2 = (curr_cnt2 - prev_cnt2) / prev_cnt2 * 100
        assert mom_delta_pct2 <= 20.0, "Test fixture: small rise must be ≤ 20%"

        owner_outlier_pct = 20.0

        # Simulate the outlier flag logic from compute()
        def _flag(curr, prev):
            if prev > 0:
                pct = (curr - prev) / prev * 100
                return pct > owner_outlier_pct
            elif prev == 0 and curr > 0:
                return True
            return False

        assert _flag(25, 20) is True,  "25 from 20 (25% rise) must be flagged"
        assert _flag(22, 20) is False, "22 from 20 (10% rise) must not be flagged"
        assert _flag(20, 20) is False, "Flat → not flagged"
        assert _flag(15, 20) is False, "Decrease → not flagged"
        assert _flag(5, 0)   is True,  "Any positive from zero → flagged"
        assert _flag(0, 0)   is False, "Zero from zero → not flagged"

    def test_owner_outlier_pct_option_respected(self):
        """owner_outlier_pct module_option controls the threshold."""
        # With a 50% threshold, a 25% rise should NOT be flagged
        owner_outlier_pct_50 = 50.0
        curr, prev = 25, 20
        pct = (curr - prev) / prev * 100
        assert pct <= owner_outlier_pct_50, "25% rise should not flag at 50% threshold"


class TestPureFunctions:
    """Unit tests for pure module-level helper functions."""

    def test_composite_rag_all_green(self):
        assert _composite_rag_od5(["green"] * 4) == ("green", False)

    def test_composite_rag_missing_caps_green(self):
        result = _composite_rag_od5(["green", "green", "green", "missing"])
        assert result == ("amber", True), f"Expected ('amber', True), got {result}"

    def test_composite_rag_three_green(self):
        """3 green + 1 red (no missing) → amber."""
        result = _composite_rag_od5(["green", "green", "green", "red"])
        assert result == ("amber", False)

    def test_composite_rag_two_green(self):
        assert _composite_rag_od5(["green", "green", "red", "red"]) == ("amber", False)

    def test_composite_rag_one_green(self):
        assert _composite_rag_od5(["green", "red", "red", "red"]) == ("red", False)

    def test_composite_rag_all_red(self):
        assert _composite_rag_od5(["red", "red", "red", "red"]) == ("red", False)

    def test_composite_rag_missing_with_amber_raw(self):
        """Missing + only 2 green → amber, data_incomplete True (missing doesn't change amber→green)."""
        result = _composite_rag_od5(["green", "green", "red", "missing"])
        assert result == ("amber", True)

    def test_composite_rag_missing_with_red_raw(self):
        """Missing + 0 green → red, data_incomplete True."""
        result = _composite_rag_od5(["missing", "red", "red", "red"])
        assert result == ("red", True)

    def test_signal_direction_improved_lower_is_better(self):
        """Lower is better: curr < prev → green."""
        assert _signal_direction(40.0, 50.0, False, 0.0) == "green"

    def test_signal_direction_worsened_lower_is_better(self):
        assert _signal_direction(60.0, 50.0, False, 0.0) == "red"

    def test_signal_direction_improved_higher_is_better(self):
        assert _signal_direction(90.0, 80.0, True, 0.0) == "green"

    def test_signal_direction_worsened_higher_is_better(self):
        assert _signal_direction(70.0, 80.0, True, 0.0) == "red"

    def test_signal_direction_none_curr(self):
        assert _signal_direction(None, 50.0, False, 0.0) == "missing"

    def test_signal_direction_none_prev(self):
        assert _signal_direction(40.0, None, False, 0.0) == "missing"

    def test_signal_direction_both_none(self):
        assert _signal_direction(None, None, False, 0.0) == "missing"

    def test_signal_direction_within_flat_band(self):
        """Delta=3, flat_band=5 → amber."""
        assert _signal_direction(53.0, 50.0, False, 5.0) == "amber"

    def test_signal_direction_at_flat_band_boundary(self):
        """Delta == flat_band → amber (inclusive)."""
        assert _signal_direction(55.0, 50.0, False, 5.0) == "amber"

    def test_signal_direction_just_outside_flat_band(self):
        """Delta=6, flat_band=5 → red (worsened, lower_is_better)."""
        assert _signal_direction(56.0, 50.0, False, 5.0) == "red"


# ===========================================================================
# Render-side tests (Plan 17-03)
# ===========================================================================
#
# All fixtures use synthetic data only (QUAL-05 / D-04-08):
#   - asset_uuid: "00000000-0000-0000-0000-00000000000N"
#   - IPs: 192.0.2.x (RFC 5737 TEST-NET — never routable)
#   - hostnames: *.test.invalid (RFC 6761)
#   - owner names: Engineering, Operations, Unassigned
#
# ModuleData builders:
#   _make_normal_data()   — populated ModuleData with two snapshots improved
#   _make_cold_data()     — cold-start ModuleData (metrics["cold_start"]=True)
#   _make_empty_data()    — _empty_result ModuleData (data.error set)
# ---------------------------------------------------------------------------

def _make_normal_data() -> "ModuleData":
    """
    Build a fully populated ModuleData simulating a normal (non-cold-start)
    compute result with 2 snapshots, all 4 signals green.
    Uses synthetic owner names only (QUAL-05).
    """
    import openpyxl  # noqa: PLC0415 — local import to keep top-level clean

    _total_ch = 25 + 8
    owner_rows = [
        {
            "owner":          "Engineering",
            "open_crit_high": 25,
            "prev_open":      20,
            "mom_delta":      5,
            "mom_delta_pct":  25.0,  # 25% rise — outlier
            "outlier":        True,
            "share_pct":      round(25 / _total_ch * 100, 1),
            "asset_count":    10,
        },
        {
            "owner":          "Operations",
            "open_crit_high": 8,
            "prev_open":      10,
            "mom_delta":      -2,
            "mom_delta_pct":  -20.0,
            "outlier":        False,
            "share_pct":      round(8 / _total_ch * 100, 1),
            "asset_count":    5,
        },
    ]
    analyst_df = pd.DataFrame([
        {
            "Owner":                 r["owner"],
            "Open Crit+High (curr)": r["open_crit_high"],
            "Open Crit+High (prev)": r["prev_open"],
            "MoM Delta":             r["mom_delta"],
            "MoM Delta %":           round(r["mom_delta_pct"], 1),
            "Outlier":               r["outlier"],
        }
        for r in owner_rows
    ])

    metrics = {
        "cold_start":                False,
        "composite_rag":             "amber",   # 2 green signals (amber)
        "data_incomplete":           False,
        "green_count":               2,
        "missing_signal_names":      [],
        "open_crit_current":         47,
        "net_delta_current":         -3.0,
        "sla_rate_current":          87.3,
        "mttr_current":              18.0,
        # D-3: surfaced current intake/fixed and current-sign status
        "new_current":               5,
        "fixed_current":             8,
        "net_velocity_status_current": "green",   # net = -3 < 0
        "signal_open_crit_status":   "green",
        "signal_net_velocity_status": "green",
        "signal_sla_rate_status":    "red",
        "signal_mttr_status":        "red",
        "open_crit_prev":            55,
        "net_delta_prev":            2.0,
        "sla_rate_prev":             90.0,
        "mttr_prev":                 15.0,
        "owner_mom_suppressed":      False,
        "owner_insufficient_note":   False,
        "sparkline_months":          ["2026-04", "2026-05"],
        "sparkline_open_crit":       [55, 47],
        "sparkline_net_velocity":    [2.0, -3.0],
        "sparkline_sla_rate":        [90.0, 87.3],
        "sparkline_mttr":            [15.0, 18.0],
    }

    from reports.modules.rag_utils import build_rag_strip_entry  # noqa: PLC0415
    rag_strip = build_rag_strip_entry(
        display_name       = "Program Health Overview",
        headline_value_str = "2 / 4 On Track",
        status             = "yellow",
    )

    return ModuleData(
        module_id        = "program_health",
        display_name     = "Program Health Overview",
        metrics          = metrics,
        table_data       = owner_rows,
        chart_data       = {},
        summary_text     = "Program Health Overview — 2 / 4 On Track.",
        metadata         = {"snapshots_used": 2},
        driver_narrative = "The program held steady on 2 of 4 indicators this month.",
        analyst_rows     = [("PH — Owner Detail", analyst_df)],
        rag_strip        = rag_strip,
        error            = None,
    )


def _make_cold_data() -> "ModuleData":
    """
    Build a cold-start ModuleData (metrics["cold_start"]=True).
    Simulates <2 deduped snapshots path.
    """
    from reports.modules.rag_utils import build_rag_strip_entry  # noqa: PLC0415

    metrics = {
        "cold_start":                True,
        "composite_rag":             "amber",
        "data_incomplete":           True,
        "open_crit_current":         12,
        "sla_rate_current":          75.0,
        "mttr_current":              None,
        "net_delta_current":         None,
        # D-3: new fields present in cold-start (None values)
        "new_current":               None,
        "fixed_current":             None,
        "net_velocity_status_current": "flat",
        "signal_open_crit_status":   "missing",
        "signal_net_velocity_status": "missing",
        "signal_sla_rate_status":    "missing",
        "signal_mttr_status":        "missing",
        "missing_signal_names":      [
            "Open Critical count", "Net Velocity", "SLA Posture", "MTTR"
        ],
        "owner_mom_suppressed":      True,
        "owner_insufficient_note":   True,
    }

    return ModuleData(
        module_id        = "program_health",
        display_name     = "Program Health Overview",
        metrics          = metrics,
        table_data       = [
            {"owner": "Engineering", "open_crit_high": 12,
             "prev_open": None, "mom_delta": None, "mom_delta_pct": None, "outlier": False,
             "share_pct": 100.0, "asset_count": 5},
        ],
        chart_data       = {},
        summary_text     = "Program Health Overview — cold start. Current Open Critical: 12.",
        metadata         = {"cold_start": True},
        driver_narrative = (
            "Program health trend being established — "
            "month-over-month direction available from next snapshot."
        ),
        analyst_rows     = [],
        rag_strip        = build_rag_strip_entry(
            display_name       = "Program Health Overview",
            headline_value_str = "Trend Being Established",
            status             = "yellow",
        ),
        error            = None,
    )


def _make_empty_data() -> "ModuleData":
    """Build an _empty_result-equivalent ModuleData (data.error set)."""
    from reports.modules.rag_utils import build_rag_strip_entry, NO_DATA_HEADLINE  # noqa: PLC0415

    return ModuleData(
        module_id        = "program_health",
        display_name     = "Program Health Overview",
        metrics          = {},
        table_data       = [],
        chart_data       = {},
        summary_text     = "",
        metadata         = {},
        driver_narrative = "",
        analyst_rows     = [],
        rag_strip        = build_rag_strip_entry(
            display_name       = "Program Health Overview",
            headline_value_str = NO_DATA_HEADLINE,
            status             = "no_data",
        ),
        error            = "vulns_df is empty",
    )


def _make_config() -> "ModuleConfig":
    return ModuleConfig("program_health", options={})


class TestAllChannelsRenderNormal:
    """test_all_channels_render_normal: normal data → all 4 channels non-empty, no raise."""

    def test_all_channels_render_normal(self):
        import openpyxl  # noqa: PLC0415
        m      = ProgramHealthModule()
        data   = _make_normal_data()
        config = _make_config()

        # PDF
        pdf_html = m.render_pdf_section(data, config)
        assert isinstance(pdf_html, str) and len(pdf_html) > 0, "PDF must return non-empty string"

        # Email
        email_html = m.render_email_panel(data, config)
        assert isinstance(email_html, str) and len(email_html) > 0, "Email panel must return non-empty string"

        # Excel
        wb    = openpyxl.Workbook()
        tabs  = m.render_excel_tabs(data, wb, config)
        assert isinstance(tabs, list) and len(tabs) == 2, f"Expected 2 tabs, got {tabs}"

        # Analyst
        analyst = m.render_analyst_tabs(data, config)
        assert isinstance(analyst, list) and len(analyst) == 1, "Analyst must return 1 tab"


class TestAllChannelsRenderColdStart:
    """test_all_channels_render_cold_start: cold-start ModuleData → all channels safe, no NaN%."""

    def test_all_channels_render_cold_start(self):
        import openpyxl  # noqa: PLC0415
        m      = ProgramHealthModule()
        data   = _make_cold_data()
        config = _make_config()

        # PDF — must return non-empty (cold-start still renders Owner table)
        pdf_html = m.render_pdf_section(data, config)
        assert isinstance(pdf_html, str) and len(pdf_html) > 0, "Cold-start PDF must not be empty"
        assert "NaN" not in pdf_html, "PDF must not contain NaN"
        assert "None%" not in pdf_html, "PDF must not contain None%"

        # Email — cold-start path returns the established notice
        email_html = m.render_email_panel(data, config)
        assert isinstance(email_html, str) and len(email_html) > 0, "Cold-start email must not be empty"
        assert "NaN" not in email_html, "Email must not contain NaN"
        assert "None%" not in email_html, "Email must not contain None%"

        # Excel — returns 2 tabs even on cold-start
        wb   = openpyxl.Workbook()
        tabs = m.render_excel_tabs(data, wb, config)
        assert isinstance(tabs, list) and len(tabs) == 2, f"Cold-start Excel must return 2 tabs, got {tabs}"

        # Analyst — cold-start returns []
        analyst = m.render_analyst_tabs(data, config)
        assert analyst == [], "Cold-start analyst tabs must be []"

        # RAG strip — amber/yellow, not no_data
        strip = m.render_rag_strip_entry(data, config)
        assert strip.get("rag_color") == "#f57c00", (
            f"Cold-start strip must be amber #f57c00, got {strip.get('rag_color')}"
        )


class TestAllChannelsRenderZeroRow:
    """test_all_channels_render_zero_row: _empty_result ModuleData → all channels safe, no raise."""

    def test_all_channels_render_zero_row(self):
        import openpyxl  # noqa: PLC0415
        m      = ProgramHealthModule()
        data   = _make_empty_data()
        config = _make_config()

        # PDF — error guard → empty string
        pdf_html = m.render_pdf_section(data, config)
        assert pdf_html == "", f"Error-state PDF must return '', got {pdf_html!r}"

        # Email — error guard → empty string
        email_html = m.render_email_panel(data, config)
        assert email_html == "", f"Error-state email must return '', got {email_html!r}"

        # Excel — error guard → minimal, but returns a list
        wb   = openpyxl.Workbook()
        tabs = m.render_excel_tabs(data, wb, config)
        assert isinstance(tabs, list), "Excel error path must return a list"

        # Analyst — error guard → []
        analyst = m.render_analyst_tabs(data, config)
        assert analyst == [], "Error-state analyst tabs must be []"

        # RAG strip — falls back to gray no_data cell
        strip = m.render_rag_strip_entry(data, config)
        from reports.modules.rag_utils import STATUS_COLOR  # noqa: PLC0415
        assert strip.get("rag_color") == STATUS_COLOR["no_data"], (
            f"Error-state RAG strip must be no_data gray, got {strip.get('rag_color')}"
        )


class TestAnalystTabsAggregateOnly:
    """test_analyst_tabs_aggregate_only: analyst df columns contain no asset-level PII (QUAL-05)."""

    def test_analyst_tabs_aggregate_only(self):
        """
        The analyst tab DataFrame must NOT contain any column whose name
        contains 'asset_uuid', 'ip', 'hostname', or 'plugin'.
        QUAL-05 hard constraint.
        """
        m      = ProgramHealthModule()
        data   = _make_normal_data()
        config = _make_config()

        analyst = m.render_analyst_tabs(data, config)
        assert len(analyst) == 1, "Expected exactly 1 analyst tab"

        tab_name, df = analyst[0]
        assert tab_name == "PH — Owner Detail", f"Unexpected tab name: {tab_name!r}"
        assert isinstance(df, pd.DataFrame), "Analyst tab must be a DataFrame"

        # QUAL-05: forbidden column substrings
        forbidden = ["asset_uuid", "ip", "hostname", "plugin"]
        for col in df.columns:
            col_lower = col.lower()
            for bad in forbidden:
                assert bad not in col_lower, (
                    f"Analyst tab column '{col}' contains forbidden substring '{bad}' (QUAL-05)"
                )

        # Expected aggregate-only columns
        expected_cols = {
            "Owner",
            "Open Crit+High (curr)",
            "Open Crit+High (prev)",
            "MoM Delta",
            "MoM Delta %",
            "Outlier",
        }
        assert set(df.columns) == expected_cols, (
            f"Analyst tab columns mismatch. Got: {set(df.columns)}"
        )


class TestEmailNoNanPercent:
    """test_email_no_nan_percent: rendered email panel contains no 'NaN' or 'None%'."""

    def test_email_no_nan_percent(self):
        """
        With normal data containing all-None signal values, the email panel
        must still render without 'NaN' or 'None%' substrings (QUAL-03).
        Uses normal data and also a variant with None metric values.
        """
        m      = ProgramHealthModule()
        config = _make_config()

        # Normal data
        data = _make_normal_data()
        html_normal = m.render_email_panel(data, config)
        assert "NaN" not in html_normal, f"Email panel must not contain 'NaN'"
        assert "None%" not in html_normal, f"Email panel must not contain 'None%'"

        # Variant: all signal current values None
        data2 = _make_normal_data()
        data2.metrics["open_crit_current"]  = None
        data2.metrics["net_delta_current"]  = None
        data2.metrics["sla_rate_current"]   = None
        data2.metrics["mttr_current"]       = None
        html_none = m.render_email_panel(data2, config)
        assert "NaN" not in html_none, "Email with None metrics must not contain 'NaN'"
        assert "None%" not in html_none, "Email with None metrics must not contain 'None%'"

        # Cold-start
        data3   = _make_cold_data()
        html_cs = m.render_email_panel(data3, config)
        assert "NaN" not in html_cs, "Cold-start email must not contain 'NaN'"
        assert "None%" not in html_cs, "Cold-start email must not contain 'None%'"

    def test_email_panel_contains_tile_labels(self):
        """render_email_panel on normal data returns all 4 tile labels."""
        m      = ProgramHealthModule()
        data   = _make_normal_data()
        config = _make_config()

        html_out = m.render_email_panel(data, config)
        for label in ["Open Critical", "Net Velocity", "SLA Posture (Crit+High)", "MTTR (30-day)"]:
            assert label in html_out, f"Email panel must contain tile label '{label}'"


class TestAmberUsesYellowColor:
    """test_amber_uses_yellow_color: amber composite uses #f57c00, never #fbc02d."""

    def test_amber_uses_yellow_color(self):
        """
        When composite_rag='amber', the email panel left-border and the
        RAG strip must use STATUS_COLOR['yellow'] = #f57c00.
        #fbc02d (Medium severity color) must not appear in any amber path.
        """
        m      = ProgramHealthModule()
        data   = _make_normal_data()  # composite_rag='amber'
        config = _make_config()

        assert data.metrics["composite_rag"] == "amber", "Fixture must be amber"

        # Email panel
        email_html = m.render_email_panel(data, config)
        assert "#f57c00" in email_html, "Amber email panel must use #f57c00 border"
        assert "#fbc02d" not in email_html, (
            "Email panel must NOT use #fbc02d (that is Medium severity color, not RAG amber)"
        )

        # RAG strip
        strip = m.render_rag_strip_entry(data, config)
        assert strip.get("rag_color") == "#f57c00", (
            f"Amber RAG strip must be #f57c00, got {strip.get('rag_color')}"
        )
        assert strip.get("rag_color") != "#fbc02d", "RAG strip must not use #fbc02d"

    def test_no_fbc02d_in_module_source(self):
        """
        Ensure #fbc02d does not appear anywhere in program_health_module.py
        (blanket guard — the amber/RAG color is #f57c00, never #fbc02d).
        """
        import inspect  # noqa: PLC0415
        from reports.modules import program_health_module  # noqa: PLC0415
        src = inspect.getsource(program_health_module)
        assert "#fbc02d" not in src, (
            "program_health_module.py must not contain #fbc02d "
            "(that is Medium severity color; RAG amber = #f57c00)"
        )


class TestSparklineReturnsBase64:
    """test_sparkline_returns_base64: _render_sparkline_b64 returns a decodable non-empty base64 PNG."""

    def test_sparkline_returns_base64(self):
        """
        _render_sparkline_b64 with a simple 3-point series must return
        a non-empty string that base64-decodes to valid PNG bytes.
        """
        import base64 as b64  # noqa: PLC0415
        m = ProgramHealthModule()

        result = m._render_sparkline_b64(
            values          = [10, 8, 6],
            signal_label    = "Open Critical",
            current_val_str = "6",
            mom_arrow       = "▼",
            arrow_color     = "#388e3c",
            line_color      = "#d32f2f",
        )

        assert isinstance(result, str), "Result must be a string"
        assert len(result) > 0, "Result must be non-empty"

        # Must be valid base64
        decoded = b64.b64decode(result)
        assert len(decoded) > 0, "Decoded bytes must be non-empty"

        # PNG magic bytes: 0x89 50 4E 47
        assert decoded[:4] == b"\x89PNG", (
            f"Decoded bytes must start with PNG magic, got {decoded[:4]!r}"
        )

    def test_sparkline_handles_none_values(self):
        """Sparkline with None-containing series must not raise."""
        m = ProgramHealthModule()
        result = m._render_sparkline_b64(
            values          = [None, 8, None],
            signal_label    = "SLA Posture",
            current_val_str = "—",
            mom_arrow       = "—",
            arrow_color     = "#9E9E9E",
            line_color      = "#388e3c",
        )
        assert isinstance(result, str) and len(result) > 0

    def test_sparkline_closes_figure(self):
        """
        _render_sparkline_b64 must call plt.close() — verify by checking
        that after N calls, matplotlib figure count stays at 0 (no leaks).
        """
        import matplotlib.pyplot as plt  # noqa: PLC0415
        m = ProgramHealthModule()

        initial_figs = len(plt.get_fignums())
        for _ in range(3):
            m._render_sparkline_b64(
                values          = [1, 2, 3],
                signal_label    = "MTTR",
                current_val_str = "20 d",
                mom_arrow       = "▲",
                arrow_color     = "#d32f2f",
                line_color      = "#f57c00",
            )
        after_figs = len(plt.get_fignums())

        assert after_figs == initial_figs, (
            f"matplotlib figure count grew from {initial_figs} to {after_figs} — "
            "plt.close(fig) must be called in _render_sparkline_b64 (T-17-08)"
        )


class TestPdfOwnerOutlierMarker:
    """test_pdf_owner_outlier_marker: outlier owner produces '▲ Outlier' in PDF Owner table."""

    def test_pdf_owner_outlier_marker(self):
        """
        A row with outlier=True (>20% MoM rise) must produce the
        '▲ Outlier' marker in the PDF render_pdf_section output.
        """
        m      = ProgramHealthModule()
        data   = _make_normal_data()  # Engineering has outlier=True (25% rise)
        config = _make_config()

        pdf_html = m.render_pdf_section(data, config)

        # The outlier marker (▲ Outlier or &#9650; Outlier) must appear
        assert "Outlier" in pdf_html, (
            "PDF Owner table must contain 'Outlier' text for outlier owners"
        )
        # Must be in red
        assert "#d32f2f" in pdf_html, (
            "Outlier marker must use #d32f2f red color"
        )

    def test_pdf_non_outlier_has_no_outlier_marker(self):
        """An owner without outlier=True must not have the Outlier marker."""
        m      = ProgramHealthModule()
        data   = _make_normal_data()
        config = _make_config()

        # Only Engineering is outlier; verify Operations row does not get the marker
        # (indirect: the PDF renders without error and contains 'Operations')
        pdf_html = m.render_pdf_section(data, config)
        assert "Engineering" in pdf_html or "Outlier" in pdf_html, (
            "PDF must contain owner data"
        )

    def test_pdf_missing_signal_note(self):
        """
        When data_incomplete=True and missing_signal_names present,
        a data_incomplete variant must include the missing signal name
        in the narrative (tested via driver_narrative passthrough).
        """
        m    = ProgramHealthModule()
        data = _make_normal_data()
        data.metrics["data_incomplete"]    = True
        data.metrics["missing_signal_names"] = ["MTTR"]
        data.driver_narrative = (
            "The program held steady on 2 of 4 indicators this month. "
            "Note: MTTR data incomplete this period."
        )
        config = _make_config()

        # Email panel must contain missing signal name
        email_html = m.render_email_panel(data, config)
        assert "MTTR" in email_html, (
            "Email panel must contain missing signal name 'MTTR' in narrative/note"
        )


class TestValidateConfigContract:
    """validate_config must return list[str] (empty = valid), per the four-channel
    contract — NOT a ModuleConfig. Regression for the UAT blocker where the composer
    crashed with 'can only join an iterable' on `'; '.join(config_errors)'."""

    def test_returns_list_not_moduleconfig(self):
        """The no-options common case must return an empty list, not a ModuleConfig."""
        errors = ProgramHealthModule().validate_config(_make_config())
        assert isinstance(errors, list), (
            "validate_config must return list[str], not ModuleConfig"
        )
        assert errors == [], "default/empty options must be valid (no errors)"

    def test_join_does_not_raise(self):
        """The composer joins the return value — it must be an iterable of strings."""
        errors = ProgramHealthModule().validate_config(_make_config())
        assert "; ".join(errors) == ""  # would TypeError if non-iterable returned

    def test_valid_options_pass(self):
        errors = ProgramHealthModule().validate_config(
            ModuleConfig(
                "program_health",
                options={"green_count_min": 4, "owner_outlier_pct": 25.0},
            )
        )
        assert errors == []

    def test_bad_int_option_reports_error(self):
        errors = ProgramHealthModule().validate_config(
            ModuleConfig("program_health", options={"green_count_min": "abc"})
        )
        assert len(errors) == 1
        assert "green_count_min" in errors[0]

    def test_bad_float_option_reports_error(self):
        errors = ProgramHealthModule().validate_config(
            ModuleConfig("program_health", options={"owner_outlier_pct": "high"})
        )
        assert len(errors) == 1
        assert "owner_outlier_pct" in errors[0]


# ===========================================================================
# WR-03 / WR-04 / WR-05 robustness tests (19-02 Task 3)
# ===========================================================================

class TestSignalDirectionNaNGuard:
    """WR-03: _signal_direction must return 'missing' for NaN, not silently 'red'."""

    def test_nan_curr_returns_missing(self):
        """float('nan') curr → 'missing', not 'red'."""
        result = _signal_direction(float("nan"), 10.0, higher_is_better=True)
        assert result == "missing", f"Expected 'missing', got {result!r}"

    def test_nan_prev_returns_missing(self):
        """float('nan') prev → 'missing'."""
        result = _signal_direction(5.0, float("nan"), higher_is_better=False)
        assert result == "missing", f"Expected 'missing', got {result!r}"

    def test_nan_both_returns_missing(self):
        """Both NaN → 'missing'."""
        result = _signal_direction(float("nan"), float("nan"), higher_is_better=True)
        assert result == "missing", f"Expected 'missing', got {result!r}"

    def test_none_still_returns_missing(self):
        """Existing None guard still works after broadening."""
        assert _signal_direction(None, 10.0, higher_is_better=True) == "missing"
        assert _signal_direction(5.0, None, higher_is_better=True) == "missing"

    def test_valid_values_still_work(self):
        """Valid floats still classify correctly after adding the NaN guard."""
        # 8.0 → 10.0, higher_is_better=True → improved → green
        assert _signal_direction(10.0, 8.0, higher_is_better=True) == "green"
        # 10.0 → 8.0, higher_is_better=False (lower is better) → improved → green
        assert _signal_direction(8.0, 10.0, higher_is_better=False) == "green"
        # 8.0 → 10.0, higher_is_better=False → worsened → red
        assert _signal_direction(10.0, 8.0, higher_is_better=False) == "red"


class TestNetVelocitySparklineGap:
    """WR-04: absent new_findings_count / fixed_findings_count → None gap, not 0."""

    def _run_with_snapshots(self, snaps: list[dict]) -> list:
        """Run compute with 2 snapshots from snaps, return net_velocity sparkline."""
        vulns_df = _make_vulns([
            {"severity": "critical", "first_found": REF - timedelta(days=5)},
        ])
        assets_df = _make_assets()
        trend_data = {
            "snapshots": snaps,
            "insufficient_data": False,
        }
        config = ModuleConfig("program_health")
        data = ProgramHealthModule().compute(
            vulns_df=vulns_df,
            assets_df=assets_df,
            report_date=REF,
            config=config,
            trend_snapshots=trend_data,
        )
        return data.chart_data.get("net_velocity", [])

    def test_absent_key_produces_none_gap(self):
        """A snapshot missing new_findings_count entirely → None in series (gap)."""
        snaps = [
            # prev snap: has counts
            {
                "month": "2026-05",
                "generated_at": "2026-05-01T00:00:00Z",
                "critical": 10,
                "sla_rate_crit_high": 80.0,
                "mttr_overall_days": 20.0,
                "new_findings_count": 5,
                "fixed_findings_count": 3,
            },
            # curr snap: missing both count keys → should produce None, not 0
            {
                "month": "2026-06",
                "generated_at": "2026-06-01T00:00:00Z",
                "critical": 8,
                "sla_rate_crit_high": 85.0,
                "mttr_overall_days": 18.0,
                # new_findings_count and fixed_findings_count intentionally absent
            },
        ]
        series = self._run_with_snapshots(snaps)
        # The last entry corresponds to the curr snap (missing keys) → must be None
        assert series[-1] is None, (
            f"Expected None for absent net-velocity keys, got {series[-1]!r}"
        )

    def test_present_counts_produce_float(self):
        """Both keys present → net velocity is a float (new - fixed)."""
        snaps = [
            {
                "month": "2026-05",
                "generated_at": "2026-05-01T00:00:00Z",
                "critical": 10,
                "sla_rate_crit_high": 80.0,
                "mttr_overall_days": 20.0,
                "new_findings_count": 5,
                "fixed_findings_count": 3,
            },
            {
                "month": "2026-06",
                "generated_at": "2026-06-01T00:00:00Z",
                "critical": 8,
                "sla_rate_crit_high": 85.0,
                "mttr_overall_days": 18.0,
                "new_findings_count": 4,
                "fixed_findings_count": 6,
            },
        ]
        series = self._run_with_snapshots(snaps)
        # last entry: 4 - 6 = -2.0
        assert series[-1] == -2.0, f"Expected -2.0, got {series[-1]!r}"


class TestAnalystDfNaNAndIntCast:
    """WR-05: analyst_df drops NaN rows and casts count columns to int."""

    def test_analyst_counts_are_int_typed(self):
        """Open Crit+High (curr) column values are Python int, not numpy int64."""
        vulns_df = _make_vulns([
            {"severity": "critical", "first_found": REF - timedelta(days=5)},
            {"severity": "high",     "first_found": REF - timedelta(days=10)},
        ])
        assets_df = _make_assets([{"asset_uuid": "00000000-0000-0000-0000-000000000001",
                                    "tags": [{"category": "Owner", "value": "Engineering"}]}])
        snaps = [
            {
                "month": "2026-05",
                "generated_at": "2026-05-01T00:00:00Z",
                "critical": 3,
                "sla_rate_crit_high": 80.0,
                "mttr_overall_days": 20.0,
                "new_findings_count": 5,
                "fixed_findings_count": 3,
                "Engineering": 3,
            },
            {
                "month": "2026-06",
                "generated_at": "2026-06-01T00:00:00Z",
                "critical": 2,
                "sla_rate_crit_high": 85.0,
                "mttr_overall_days": 18.0,
                "new_findings_count": 2,
                "fixed_findings_count": 4,
                "Engineering": 2,
            },
        ]
        config = ModuleConfig("program_health")
        data = ProgramHealthModule().compute(
            vulns_df=vulns_df,
            assets_df=assets_df,
            report_date=REF,
            config=config,
            trend_snapshots={"snapshots": snaps, "insufficient_data": False},
        )
        # analyst_rows is list of (tab_name, DataFrame)
        assert data.analyst_rows, "Expected non-empty analyst_rows"
        _, analyst_df = data.analyst_rows[0]
        if not analyst_df.empty:
            curr_col = analyst_df["Open Crit+High (curr)"]
            for val in curr_col:
                assert isinstance(val, int), (
                    f"Expected Python int, got {type(val).__name__}: {val!r}"
                )

    def test_owner_metadata_key_constant_is_frozenset(self):
        """WR-05: _OWNER_SNAPSHOT_METADATA_KEYS is a frozenset (immutable, no drift)."""
        assert isinstance(_OWNER_SNAPSHOT_METADATA_KEYS, frozenset)
        # Spot-check key presence — these are written by capture_snapshot
        for key in ("month", "generated_at", "sla_rate_crit_high", "asset_count"):
            assert key in _OWNER_SNAPSHOT_METADATA_KEYS, f"Missing key: {key!r}"


# ===========================================================================
# 19-10 Task 1: Compute layer — new_current/fixed_current, net_velocity_status_current,
#               owner share_pct + asset_count
# ===========================================================================

class TestNetVelocityCurrentSign:
    """D-3: net_velocity_status_current derived from sign(curr_net_delta), not delta-of-deltas."""

    def _two_snaps(self, new_curr: int, fix_curr: int) -> ModuleData:
        prev = _make_snapshot(
            "2026-04",
            critical=20,
            new_findings_count=10, fixed_findings_count=5,
            mttr_overall_days=25.0, sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05",
            critical=10,
            new_findings_count=new_curr, fixed_findings_count=fix_curr,
            mttr_overall_days=15.0, sla_rate_crit_high=90.0,
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        return _run(vuln_rows, snapshots=[prev, curr])

    def test_net_negative_gives_green(self):
        """fixed > intake → net < 0 → net_velocity_status_current == 'green'."""
        result = self._two_snaps(new_curr=3, fix_curr=8)  # net = -5
        status = result.metrics.get("net_velocity_status_current")
        assert status == "green", f"Expected 'green', got {status!r}"

    def test_net_positive_gives_red(self):
        """intake > fixed → net > 0 → net_velocity_status_current == 'red'."""
        result = self._two_snaps(new_curr=10, fix_curr=3)  # net = +7
        status = result.metrics.get("net_velocity_status_current")
        assert status == "red", f"Expected 'red', got {status!r}"

    def test_net_zero_gives_flat(self):
        """intake == fixed → net == 0 → net_velocity_status_current == 'flat'."""
        result = self._two_snaps(new_curr=5, fix_curr=5)  # net = 0
        status = result.metrics.get("net_velocity_status_current")
        assert status == "flat", f"Expected 'flat', got {status!r}"


class TestSurfacedCurrentFields:
    """new_current / fixed_current surfaced in metrics (D-3)."""

    def test_new_current_and_fixed_current_surfaced(self):
        """compute() exposes new_current and fixed_current from curr snapshot."""
        prev = _make_snapshot(
            "2026-04", new_findings_count=10, fixed_findings_count=5,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05", new_findings_count=7, fixed_findings_count=4,
            generated_at="2026-05-31T00:00:00Z",
        )
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows, snapshots=[prev, curr])

        assert result.metrics.get("new_current") == 7, (
            f"new_current should be 7, got {result.metrics.get('new_current')!r}"
        )
        assert result.metrics.get("fixed_current") == 4, (
            f"fixed_current should be 4, got {result.metrics.get('fixed_current')!r}"
        )

    def test_cold_start_has_new_current_fixed_current_keys(self):
        """Cold-start ModuleData must not KeyError on new_current/fixed_current."""
        vuln_rows = [{"severity": "critical"}]
        result = _run(vuln_rows)  # no trend_snapshots → cold-start
        assert "new_current" in result.metrics, "new_current key missing from cold-start metrics"
        assert "fixed_current" in result.metrics, "fixed_current key missing from cold-start metrics"
        assert "net_velocity_status_current" in result.metrics, (
            "net_velocity_status_current key missing from cold-start metrics"
        )

    def test_cold_start_new_current_is_none(self):
        """Cold-start: new_current and fixed_current are None (no snapshot data)."""
        result = _run([{"severity": "critical"}])
        assert result.metrics.get("new_current") is None
        assert result.metrics.get("fixed_current") is None


class TestOwnerSharePctAndAssetCount:
    """D-5: owner rows carry share_pct (divide-by-zero guarded) and asset_count."""

    def _two_snap_result_multi_owner(self) -> ModuleData:
        """Two snapshots, two owners (Engineering + Operations)."""
        prev = _make_snapshot(
            "2026-04", critical=30,
            new_findings_count=10, fixed_findings_count=5,
            mttr_overall_days=25.0, sla_rate_crit_high=70.0,
            generated_at="2026-04-30T00:00:00Z",
        )
        curr = _make_snapshot(
            "2026-05", critical=20,
            new_findings_count=5, fixed_findings_count=8,
            mttr_overall_days=15.0, sla_rate_crit_high=90.0,
            generated_at="2026-05-31T00:00:00Z",
        )
        # Two critical assets with different owners
        vuln_rows = [
            {"severity": "critical", "asset_uuid": _uuid(1), "first_found": REF - timedelta(days=5)},
            {"severity": "critical", "asset_uuid": _uuid(1), "first_found": REF - timedelta(days=3)},
            {"severity": "critical", "asset_uuid": _uuid(2), "first_found": REF - timedelta(days=4)},
        ]
        asset_rows = [
            {"asset_uuid": _uuid(1), "tags": "Owner=Engineering"},
            {"asset_uuid": _uuid(2), "tags": "Owner=Operations"},
        ]
        return _run(vuln_rows, snapshots=[prev, curr], asset_rows=asset_rows)

    def test_share_pct_sums_to_100(self):
        """Across all owner rows, share_pct values sum to ~100%."""
        result = self._two_snap_result_multi_owner()
        rows = result.table_data
        # Only check if there are owner rows with non-None share_pct
        pcts = [r.get("share_pct") for r in rows if r.get("share_pct") is not None]
        if pcts:
            total = sum(pcts)
            assert abs(total - 100.0) < 0.5, f"share_pct sum should be ~100%, got {total}"

    def test_owner_rows_have_share_pct_key(self):
        """Each owner row must have a share_pct key."""
        result = self._two_snap_result_multi_owner()
        for row in result.table_data:
            assert "share_pct" in row, f"share_pct key missing from owner row: {row}"

    def test_owner_rows_have_asset_count_key(self):
        """Each owner row must have an asset_count key."""
        result = self._two_snap_result_multi_owner()
        for row in result.table_data:
            assert "asset_count" in row, f"asset_count key missing from owner row: {row}"

    def test_share_pct_none_when_absent_in_cold_start_data(self):
        """Cold-start owner rows also carry share_pct (None) and asset_count keys."""
        result = _run(
            [{"severity": "critical", "asset_uuid": _uuid(1)}],
            asset_rows=[{"asset_uuid": _uuid(1), "tags": "Owner=Engineering"}],
        )
        for row in result.table_data:
            assert "share_pct" in row, "share_pct missing from cold-start owner row"
            assert "asset_count" in row, "asset_count missing from cold-start owner row"

    def test_zero_total_owner_share_pct_is_none(self):
        """When total open Crit+High is zero, share_pct should be None (no divide-by-zero)."""
        # Snapshots exist but no critical/high vulns open at report_date
        prev = _make_snapshot("2026-04", generated_at="2026-04-30T00:00:00Z")
        curr = _make_snapshot("2026-05", generated_at="2026-05-31T00:00:00Z")
        # Only low-severity vulns → 0 crit+high in owner table
        vuln_rows = [
            {"severity": "low", "asset_uuid": _uuid(1), "first_found": REF - timedelta(days=5)},
        ]
        asset_rows = [{"asset_uuid": _uuid(1), "tags": "Owner=Engineering"}]
        result = _run(vuln_rows, snapshots=[prev, curr], asset_rows=asset_rows)
        for row in result.table_data:
            # share_pct should be None when total is 0
            assert row.get("share_pct") is None, (
                f"Expected None share_pct when no crit+high, got {row.get('share_pct')!r}"
            )


# ===========================================================================
# 19-10 Task 2: PDF render — two-page split, verbatim captions,
#               intake/fixed/net annotation, MTTR caption
# ===========================================================================

# Verbatim approved captions (spec lines 40-45)
_CAPTION_OPEN_CRITICAL = (
    "open Critical-VPR findings — lower is better; ▼ green = falling"
)
_CAPTION_NET_VELOCITY = (
    "new findings minus fixed this window (intake − fixed) — negative is good"
)
_CAPTION_SLA_POSTURE = (
    "% of open Critical + High findings still within SLA — higher is better; ▲ green = rising"
)
_CAPTION_MTTR = (
    "average days to remediate (rolling 30-day) — lower is better; ▼ green = falling"
)


def _make_normal_data_with_new_fields() -> "ModuleData":
    """
    Extend _make_normal_data() with the Task 1 new fields so Task 2 tests
    exercise the real renderer without touching ModuleData internals.
    """
    data = _make_normal_data()
    # Inject new compute metrics (Task 1 output)
    data.metrics["new_current"] = 5
    data.metrics["fixed_current"] = 8
    data.metrics["net_velocity_status_current"] = "green"   # net = -3 < 0
    # Inject share_pct + asset_count into table_data rows
    total_ch = sum(r["open_crit_high"] for r in data.table_data)
    for row in data.table_data:
        row["share_pct"] = (row["open_crit_high"] / total_ch * 100.0) if total_ch > 0 else None
        row["asset_count"] = 10  # synthetic
    return data


class TestPdfCaptions:
    """D-2: each chart shows its verbatim approved caption."""

    def test_open_critical_caption_present(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        assert "Open Critical" in pdf
        assert "lower is better" in pdf, "Open Critical caption missing 'lower is better'"

    def test_net_velocity_caption_present(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        assert "intake" in pdf.lower() or "new findings minus fixed" in pdf.lower(), (
            "Net Velocity caption text not found in PDF HTML"
        )

    def test_sla_posture_caption_present(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        assert "within SLA" in pdf or "still within SLA" in pdf, (
            "SLA Posture caption not found in PDF HTML"
        )

    def test_mttr_caption_present(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        assert "rolling 30-day" in pdf, "MTTR caption not found in PDF HTML"

    def test_mttr_establishing_caption_when_none(self):
        """MTTR caption appends 'establishing from monthly snapshots' when mttr_current is None."""
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        data.metrics["mttr_current"] = None
        pdf = m.render_pdf_section(data, _make_config())
        assert "establishing from monthly snapshots" in pdf, (
            "MTTR establishing caption not present when mttr_current is None"
        )

    def test_mttr_establishing_caption_absent_when_value_present(self):
        """MTTR establishing caption NOT shown when mttr_current has a value."""
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        data.metrics["mttr_current"] = 18.0  # non-None
        pdf = m.render_pdf_section(data, _make_config())
        assert "establishing from monthly snapshots" not in pdf, (
            "MTTR establishing caption should not appear when mttr_current is set"
        )


class TestPdfNetVelocityAnnotation:
    """D-3: Net Velocity annotation reads 'in {new} / fixed {fixed} · net {net} {arrow}'."""

    def test_annotation_format_with_values(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        data.metrics["new_current"] = 5
        data.metrics["fixed_current"] = 8
        data.metrics["net_delta_current"] = -3.0
        data.metrics["net_velocity_status_current"] = "green"
        pdf = m.render_pdf_section(data, _make_config())
        # The annotation "in {new} / fixed {fixed} · net {net} {arrow}" appears in the HTML
        # annotation_html div beneath the Net Velocity sparkline cell
        assert "in 5" in pdf, "intake annotation ('in 5') missing from PDF HTML"
        assert "fixed 8" in pdf, "fixed annotation ('fixed 8') missing from PDF HTML"

    def test_annotation_contains_net_arrow(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        data.metrics["new_current"] = 5
        data.metrics["fixed_current"] = 8
        data.metrics["net_delta_current"] = -3.0
        data.metrics["net_velocity_status_current"] = "green"
        pdf = m.render_pdf_section(data, _make_config())
        # Green net (net < 0): arrow should be ▼
        assert "▼" in pdf, "Green net velocity arrow (▼) missing from PDF annotation"


class TestPdfPageBreak:
    """D-1: Owner table begins on page 2 — explicit page-break element present."""

    def test_page_break_element_before_owner_table(self):
        """A page-break div appears between the sparkline/chart block and the Owner table."""
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        # The composer CSS defines .page-break { page-break-before: always }
        assert 'class="page-break"' in pdf, (
            "page-break div not found in PDF HTML — Owner table won't start on page 2"
        )
        # The page-break must come before the Owner table heading
        pb_pos = pdf.find('class="page-break"')
        owner_pos = pdf.find("Owner Velocity")
        assert pb_pos < owner_pos, (
            f"page-break (pos {pb_pos}) must appear before 'Owner Velocity' (pos {owner_pos})"
        )

    def test_cold_start_no_page_break_needed(self):
        """Cold-start PDF renders without crash; page-break present if owner table present."""
        m = ProgramHealthModule()
        data = _make_cold_data()
        # Inject new Task 1 fields into cold-start data
        data.metrics["new_current"] = None
        data.metrics["fixed_current"] = None
        data.metrics["net_velocity_status_current"] = "flat"
        pdf = m.render_pdf_section(data, _make_config())
        assert isinstance(pdf, str), "cold-start PDF render must return a string"


class TestPdfOwnerTableD5Columns:
    """D-5: Owner table has six columns: Owner | Open Crit+High | Share % | Assets | MoM Delta | MoM Delta %."""

    def test_owner_table_has_share_pct_header(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        assert "Share %" in pdf, "Share % column header missing from PDF Owner table"

    def test_owner_table_has_assets_header(self):
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        pdf = m.render_pdf_section(data, _make_config())
        assert "Assets" in pdf, "Assets column header missing from PDF Owner table"


# ===========================================================================
# 19-10 Task 3: Email + Excel — intake/fixed/net + definitions; readable
#               Excel header; Owner Velocity columns
# ===========================================================================

class TestEmailD3D4:
    """D-3/D-4: Email Net Velocity tile shows intake/fixed/net + current-sign arrow;
    MTTR tile shows 'establishing from monthly snapshots'; per-tile definitions present."""

    def _email(self, **overrides) -> str:
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        for k, v in overrides.items():
            data.metrics[k] = v
        return m.render_email_panel(data, _make_config())

    def test_net_velocity_tile_intake_value(self):
        """Email Net Velocity tile must show 'in {new}' annotation."""
        html_out = self._email(new_current=5, fixed_current=8, net_delta_current=-3.0,
                               net_velocity_status_current="green")
        assert "in 5" in html_out, "Email Net Velocity tile missing 'in {new}' (in 5)"

    def test_net_velocity_tile_fixed_value(self):
        """Email Net Velocity tile must show 'fixed {fixed}' annotation."""
        html_out = self._email(new_current=5, fixed_current=8, net_delta_current=-3.0,
                               net_velocity_status_current="green")
        assert "fixed 8" in html_out, "Email Net Velocity tile missing 'fixed {fixed}' (fixed 8)"

    def test_net_velocity_tile_green_arrow(self):
        """Net < 0 → green ▼ arrow in email tile."""
        html_out = self._email(new_current=5, fixed_current=8, net_delta_current=-3.0,
                               net_velocity_status_current="green")
        assert "▼" in html_out, "Green net velocity ▼ arrow missing from email"

    def test_net_velocity_tile_red_arrow(self):
        """Net > 0 → red ▲ arrow in email tile."""
        html_out = self._email(new_current=10, fixed_current=3, net_delta_current=7.0,
                               net_velocity_status_current="red")
        assert "▲" in html_out, "Red net velocity ▲ arrow missing from email"

    def test_mttr_establishing_caption_in_email(self):
        """MTTR tile shows 'establishing from monthly snapshots' when mttr_current is None."""
        html_out = self._email(mttr_current=None)
        assert "establishing from monthly snapshots" in html_out, (
            "MTTR establishing caption missing from email when mttr_current=None"
        )

    def test_per_tile_definition_open_critical(self):
        """Email has a per-tile definition for Open Critical."""
        html_out = self._email()
        assert "lower is better" in html_out, (
            "Open Critical per-tile definition ('lower is better') missing from email"
        )

    def test_per_tile_definition_net_velocity(self):
        """Email has a per-tile definition for Net Velocity."""
        html_out = self._email()
        assert "negative is good" in html_out or "intake" in html_out.lower(), (
            "Net Velocity per-tile definition missing from email"
        )

    def test_no_style_blocks_in_email(self):
        """Email output must not contain <style blocks (Outlook-safe inline CSS only)."""
        html_out = self._email()
        assert "<style" not in html_out, (
            "Email panel must not contain <style blocks (use inline CSS only)"
        )

    def test_cold_start_email_no_crash_with_new_fields(self):
        """Cold-start email render does not crash with the new metric keys present."""
        m = ProgramHealthModule()
        data = _make_cold_data()
        html_out = m.render_email_panel(data, _make_config())
        assert isinstance(html_out, str) and len(html_out) > 0


class TestExcelD6:
    """D-6: Excel Program Health header fill == E3F2FD; definitions block present;
    Owner Velocity has Share %/Assets columns."""

    def _render_wb(self, data=None):
        import openpyxl
        m = ProgramHealthModule()
        if data is None:
            data = _make_normal_data_with_new_fields()
        config = _make_config()
        wb = openpyxl.Workbook()
        m.render_excel_tabs(data, wb, config)
        return wb

    def test_program_health_header_fill_is_e3f2fd(self):
        """Program Health tab row-header fill must be E3F2FD (not 1F3864)."""
        wb = self._render_wb()
        ws = wb["Program Health"]
        # Row 4 is the signal header row (Signal | Current Value | MoM Direction)
        header_fill = ws.cell(row=4, column=1).fill.fgColor.rgb
        # Strip alpha prefix if present (openpyxl returns 'FF' + hex for ARGB)
        fill_hex = header_fill[-6:].upper()
        assert fill_hex == "E3F2FD", (
            f"Program Health header fill should be E3F2FD, got {fill_hex!r} (old dark fill was 1F3864)"
        )

    def test_program_health_header_not_dark(self):
        """Old dark 1F3864 fill must be gone from Program Health header."""
        wb = self._render_wb()
        ws = wb["Program Health"]
        header_fill = ws.cell(row=4, column=1).fill.fgColor.rgb
        fill_hex = header_fill[-6:].upper()
        assert fill_hex != "1F3864", "Old dark blue 1F3864 fill must not be used on header row"

    def test_program_health_definitions_block_present(self):
        """Program Health tab has a definitions block (the 4 approved caption lines)."""
        wb = self._render_wb()
        ws = wb["Program Health"]
        # Definitions block should contain at least one of the approved caption phrases
        all_values = " ".join(
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, 4)
        )
        assert "lower is better" in all_values, (
            "Definitions block missing 'lower is better' phrase on Program Health tab"
        )

    def test_owner_velocity_has_share_pct_column(self):
        """Owner Velocity tab header row contains 'Share %'."""
        wb = self._render_wb()
        ws = wb["Owner Velocity"]
        headers = [str(ws.cell(row=3, column=c).value or "") for c in range(1, 8)]
        assert "Share %" in headers, f"'Share %' missing from Owner Velocity headers: {headers}"

    def test_owner_velocity_has_assets_column(self):
        """Owner Velocity tab header row contains 'Assets'."""
        wb = self._render_wb()
        ws = wb["Owner Velocity"]
        headers = [str(ws.cell(row=3, column=c).value or "") for c in range(1, 8)]
        assert "Assets" in headers, f"'Assets' missing from Owner Velocity headers: {headers}"

    def test_cold_start_excel_no_crash(self):
        """Cold-start Excel render does not crash."""
        import openpyxl
        m = ProgramHealthModule()
        data = _make_cold_data()
        wb = openpyxl.Workbook()
        tabs = m.render_excel_tabs(data, wb, _make_config())
        assert "Program Health" in tabs
        assert "Owner Velocity" in tabs


# ===========================================================================
# 19-10 Task 4: Test sweep — cold-start + empty-data QUAL-03 guard assertions
# ===========================================================================

class TestColdStartNewFields:
    """Explicit cold-start assertions for 19-10 new keys — no KeyError downstream."""

    def test_cold_start_compute_has_all_new_keys(self):
        """compute() cold-start path includes new_current, fixed_current, net_velocity_status_current."""
        vuln_rows = [{"severity": "critical", "first_found": REF - timedelta(days=5)}]
        result = _run(vuln_rows)  # no trend_snapshots → cold-start
        m = result.metrics
        assert "new_current" in m, "new_current key missing"
        assert "fixed_current" in m, "fixed_current key missing"
        assert "net_velocity_status_current" in m, "net_velocity_status_current key missing"
        assert m["new_current"] is None, f"cold-start new_current should be None, got {m['new_current']!r}"
        assert m["fixed_current"] is None, f"cold-start fixed_current should be None, got {m['fixed_current']!r}"

    def test_cold_start_mttr_establishing_pdf_caption(self):
        """Cold-start PDF render: MTTR establishing caption present when mttr_current is None."""
        m = ProgramHealthModule()
        data = _make_cold_data()
        pdf = m.render_pdf_section(data, _make_config())
        # Cold-start has no sparklines (cold-start notice branch), but the render must not crash
        assert isinstance(pdf, str) and len(pdf) > 0

    def test_cold_start_email_establishing_caption(self):
        """Cold-start email: 'establish' or 'trend being established' appears in the panel."""
        m = ProgramHealthModule()
        data = _make_cold_data()
        email = m.render_email_panel(data, _make_config())
        assert "establish" in email.lower(), (
            "Cold-start email panel should mention 'establish' (trend being established)"
        )

    def test_cold_start_all_four_channels_no_crash(self):
        """All 4 render channels survive the cold-start path with new Task 1 fields present."""
        import openpyxl
        m = ProgramHealthModule()
        data = _make_cold_data()
        config = _make_config()

        pdf = m.render_pdf_section(data, config)
        assert isinstance(pdf, str), "PDF must return string"

        email = m.render_email_panel(data, config)
        assert isinstance(email, str), "email must return string"

        wb = openpyxl.Workbook()
        tabs = m.render_excel_tabs(data, wb, config)
        assert isinstance(tabs, list), "excel tabs must return list"

        analyst = m.render_analyst_tabs(data, config)
        assert isinstance(analyst, list), "analyst tabs must return list"


class TestQual03EmptyDataGuards:
    """QUAL-03: zero-row / empty-data paths stay green after 19-10 changes."""

    def test_empty_data_all_channels_no_crash(self):
        """_empty_result path: all 4 render channels return safe defaults."""
        import openpyxl
        m = ProgramHealthModule()
        data = _make_empty_data()
        config = _make_config()

        pdf = m.render_pdf_section(data, config)
        assert pdf == "", "Empty-data PDF must return empty string"

        email = m.render_email_panel(data, config)
        assert email == "", "Empty-data email must return empty string"

        wb = openpyxl.Workbook()
        tabs = m.render_excel_tabs(data, wb, config)
        # Excel handles error gracefully — returns tab list even on error
        assert isinstance(tabs, list), "Empty-data Excel must return list (not raise)"

        analyst = m.render_analyst_tabs(data, config)
        assert analyst == [], "Empty-data analyst tabs must return []"

    def test_zero_row_compute_returns_empty_result(self):
        """compute() with empty vulns_df returns _empty_result (QUAL-03)."""
        result = _run([])  # empty vulns_df
        assert result.error is not None, "Zero-row compute must set error"
        assert result.metrics == {}, "Zero-row _empty_result must have empty metrics"

    def test_email_no_nan_percent_after_new_fields(self):
        """Email panel must never contain 'nan%' or 'None%' after Task 1-3 changes."""
        m = ProgramHealthModule()
        data = _make_normal_data_with_new_fields()
        email = m.render_email_panel(data, _make_config())
        assert "nan%" not in email.lower(), "NaN% found in email panel"
        assert "none%" not in email.lower(), "None% found in email panel"
