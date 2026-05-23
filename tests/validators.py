"""
tests/validators.py — structural assertions for produced artifacts.

assert_valid_pdf      — pypdf opens it and it has >=1 page.
assert_valid_xlsx     — openpyxl loads it; optional expected-tab check.
assert_well_formed_html — html.parser consumes it without raising.
assert_email_cids_resolve — every <img src="cid:X"> has a matching
                            Content-ID part in the MIME message.
"""
from __future__ import annotations

import re
from email.message import Message
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def assert_valid_pdf(path: Path) -> None:
    assert Path(path).exists(), f"PDF missing: {path}"
    reader = PdfReader(str(path))
    assert len(reader.pages) >= 1, f"PDF has no pages: {path}"


def assert_valid_xlsx(path: Path, expected_tabs: list[str] | None = None) -> None:
    assert Path(path).exists(), f"Excel missing: {path}"
    wb = load_workbook(str(path), read_only=True)
    if expected_tabs:
        missing = [t for t in expected_tabs if t not in wb.sheetnames]
        assert not missing, f"Excel {path} missing tabs: {missing}"


class _StrictHTML(HTMLParser):
    def error(self, message):  # pragma: no cover - parser rarely calls this
        raise ValueError(message)


def assert_well_formed_html(html: str) -> None:
    assert html and html.strip(), "HTML body is empty"
    _StrictHTML().feed(html)  # raises on malformed markup


def assert_email_cids_resolve(msg: Message) -> None:
    """Every cid: referenced in the HTML body must have a matching part."""
    html_parts = [
        p.get_payload(decode=True).decode("utf-8", "replace")
        for p in msg.walk()
        if p.get_content_type() == "text/html"
    ]
    referenced = set()
    for body in html_parts:
        referenced.update(re.findall(r'src=["\']cid:([^"\']+)["\']', body))

    available = set()
    for part in msg.walk():
        cid = part.get("Content-ID", "")
        if cid:
            available.add(cid.strip("<>"))

    missing = referenced - available
    assert not missing, f"Unresolved inline CIDs: {missing} (have: {available})"
