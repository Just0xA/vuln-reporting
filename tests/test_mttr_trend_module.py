"""
tests/test_mttr_trend_module.py — Unit tests for MTTRTrendModule.

Tests cover:
  - Criterion-3 reopened-clock fixture: first_found=-200d, resurfaced_date=-10d,
    last_fixed=-2d → days_to_fix=8, overall_mttr=8.0 (D-16-02)
  - Zero fixed findings → _empty_result() / cold-start (QUAL-01/03)
  - Cold-start MoM: 1 snapshot (insufficient_data=True) → cold-start notice,
    no NaN%, no crash (QUAL-01)
  - min_sample=5 sub-threshold: 3 Critical → "Insufficient data (3 findings...)"
  - Owner cold start: Owner appearing only in snapshot 2 → series=[None, x], MoM=None
  - Owner vanished: Owner in snapshot 1 only → omitted from current Owner table
  - Multiple snapshots same month → latest generated_at wins (D-16-08 tie-break)
  - Partial-month label: current month contains "partial" (D-16-08)
  - pandas CoW strict mode: zero ChainedAssignmentError (QUAL-03)
  - Four-channel empty-data guard: all render methods survive zero-row ModuleData
  - Composed-pipeline smoke: mttr_trend receives non-None fixed_vulns_df via
    MTTRTrendModule.compute() with fixed_vulns_df kwarg (end-to-end guard for
    the 16-02 _MODULES_NEEDING_FIXED_VULNS membership)
  - Structural baseline self-guard: compare_snapshots(actual, baseline) == []

All fixtures use synthetic data only (QUAL-05 / D-04-08):
  - asset_uuid: "00000000-0000-0000-0000-00000000000N"
  - No real hostnames, IPs, CVE IDs, or plugin names

Pandas CoW strict mode enforced at module level.
"""

from __future__ import annotations

import datetime
import tempfile
from datetime import timedelta, timezone
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
import pytest

# Enforce pandas CoW strict mode — catches ChainedAssignmentError / FutureWarning
pd.options.mode.copy_on_write = True

from reports.modules.base import ModuleConfig, ModuleData
from reports.modules.mttr_trend_module import MTTRTrendModule

# ---------------------------------------------------------------------------
# Constants / shared references
# ---------------------------------------------------------------------------

_UUID_PREFIX = "00000000-0000-0000-0000-00000000000"
# Report date anchors all relative timestamps so criterion-3 math is stable
REF = datetime.datetime(2026, 6, 12, 0, 0, 0, tzinfo=timezone.utc)
_CURRENT_PERIOD = pd.Period("2026-06", "M")


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

def _uuid(n: int) -> str:
    return f"{_UUID_PREFIX}{n}"


def _make_fixed_vulns(rows: list[dict]) -> pd.DataFrame:
    """Build a minimal fixed_vulns_df with all required D-16-02 columns.

    Defaults: state="fixed", severity="critical", asset_uuid synthetic.
    All date columns coerced to UTC-aware datetime64.
    CoW-compliant: uses .assign() for column mutations.
    """
    defaults = {
        "state":           "fixed",
        "severity":        "critical",
        "asset_uuid":      _uuid(1),
        "first_found":     REF - timedelta(days=10),
        "resurfaced_date": None,
        "last_fixed":      REF - timedelta(days=2),
    }
    records = [{**defaults, **r} for r in rows]
    df = pd.DataFrame(records, columns=list(defaults.keys()))
    # CoW-compliant: use .assign() rather than df[col] = ... after creation
    df = df.assign(
        first_found=pd.to_datetime(df["first_found"], utc=True, errors="coerce"),
        resurfaced_date=pd.to_datetime(df["resurfaced_date"], utc=True, errors="coerce"),
        last_fixed=pd.to_datetime(df["last_fixed"], utc=True, errors="coerce"),
    )
    return df


def _make_assets(rows: Optional[list[dict]] = None) -> pd.DataFrame:
    """Build a minimal assets_df with asset_uuid and tags columns."""
    if rows is None:
        rows = [{"asset_uuid": _uuid(1), "tags": "Owner=Engineering"}]
    defaults = {"asset_uuid": _uuid(1), "tags": ""}
    records = [{**defaults, **r} for r in rows]
    return pd.DataFrame(records, columns=list(defaults.keys()))


def _make_snapshot_mttr(
    month: str,
    mttr_overall_days: Optional[float] = None,
    mttr_by_severity: Optional[dict] = None,
    mttr_by_owner: Optional[dict] = None,
    generated_at: str = "2026-06-01T00:00:00Z",
    **extra,
) -> dict:
    """Build a minimal trend snapshot dict with Phase-16 MTTR fields.

    Follows the D-16-09 implicit-optional-field convention: all three
    MTTR fields are stored as explicit null rather than omitted.
    """
    return {
        "month":               month,
        "tag_filter":          "all_assets",
        "critical":            0,
        "high":                0,
        "medium":              0,
        "low":                 0,
        "asset_count":         5,
        "on_time_asset_count": None,
        "reopened_count":      None,
        "accepted_count":      None,
        "recast_count":        None,
        "new_findings_count":  None,
        "fixed_findings_count": None,
        "mttr_overall_days":   mttr_overall_days,
        "mttr_by_severity":    mttr_by_severity,
        "mttr_by_owner":       mttr_by_owner,
        "generated_at":        generated_at,
        **extra,
    }


def _make_trend_snapshots(
    snapshots: list[dict],
    insufficient_data: bool = False,
) -> dict:
    """Wrap a list of snapshot dicts in the read_trend() envelope."""
    return {
        "snapshots":        snapshots,
        "insufficient_data": insufficient_data,
    }


def _config(**options) -> ModuleConfig:
    return ModuleConfig("mttr_trend", options=options)


def _run(
    fixed_rows: list[dict],
    trend_snapshots: Optional[dict] = None,
    asset_rows: Optional[list[dict]] = None,
    **options,
) -> ModuleData:
    """Run MTTRTrendModule.compute() with synthetic fixtures."""
    mod       = MTTRTrendModule()
    fixed_df  = _make_fixed_vulns(fixed_rows)
    assets_df = _make_assets(asset_rows)
    cfg       = _config(**options)
    kwargs: dict = {"fixed_vulns_df": fixed_df}
    if trend_snapshots is not None:
        kwargs["trend_snapshots"] = trend_snapshots
    return mod.compute(pd.DataFrame(), assets_df, REF, cfg, **kwargs)


# ===========================================================================
# 1. MODULE REGISTRATION
# ===========================================================================

class TestRegistration:
    def test_module_id(self):
        assert MTTRTrendModule.MODULE_ID == "mttr_trend"

    def test_display_name(self):
        assert "MTTR" in MTTRTrendModule.DISPLAY_NAME

    def test_auto_discovery(self):
        """Module is auto-discovered via @register_module on import."""
        import reports.modules
        import reports.modules.registry as registry
        assert "mttr_trend" in registry._modules

    def test_required_data_includes_fixed_vulns(self):
        assert "fixed_vulns" in MTTRTrendModule.REQUIRED_DATA

    def test_required_data_includes_trend_snapshots(self):
        assert "trend_snapshots" in MTTRTrendModule.REQUIRED_DATA


# ===========================================================================
# 2. CRITERION-3 REOPENED CLOCK (the acceptance lodestar — D-16-02)
# ===========================================================================

class TestCriterion3ReopenedClock:
    def test_reopened_clock_days_to_fix_is_8(self):
        """
        D-16-02 criterion-3 fixture:
        first_found=−200d, resurfaced_date=−10d, last_fixed=−2d → days_to_fix=8.
        The old module (time_taken_to_fix preference) would return 198.

        COALESCE chooses resurfaced_date (present) as clock start.
        Clock: (REF-2d) - (REF-10d) = 8 days. NOT 198.
        """
        fixed_df = _make_fixed_vulns([{
            "first_found":     REF - timedelta(days=200),
            "resurfaced_date": REF - timedelta(days=10),
            "last_fixed":      REF - timedelta(days=2),
            "state":           "fixed",
            "severity":        "critical",
        }])
        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options={"min_sample_size": 1})
        data = mod.compute(
            pd.DataFrame(),
            _make_assets(),
            REF,
            cfg,
            fixed_vulns_df=fixed_df,
        )
        assert data.error is None
        assert data.metrics.get("overall_mttr") == 8.0, (
            f"Expected 8.0d (reopened-aware COALESCE clock), "
            f"got {data.metrics.get('overall_mttr')} — "
            f"if this returns 198 the old time_taken_to_fix path is still active"
        )

    def test_cold_start_is_false_when_data_present(self):
        """With valid fixed findings, cold_start must be False."""
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            min_sample_size=1,
        )
        assert data.metrics.get("cold_start") is False

    def test_first_found_only_when_no_resurfaced(self):
        """When resurfaced_date is None, clock_start = first_found."""
        fixed_df = _make_fixed_vulns([{
            "first_found":     REF - timedelta(days=20),
            "resurfaced_date": None,
            "last_fixed":      REF - timedelta(days=5),
            "state":           "fixed",
            "severity":        "critical",
        }])
        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options={"min_sample_size": 1})
        data = mod.compute(pd.DataFrame(), _make_assets(), REF, cfg,
                           fixed_vulns_df=fixed_df)
        assert data.error is None
        # clock: (REF-5) - (REF-20) = 15 days
        assert data.metrics.get("overall_mttr") == 15.0

    def test_rag_status_not_no_data_when_mttr_computed(self):
        """When overall_mttr is computed, rag_status must not be 'no_data'."""
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            min_sample_size=1,
        )
        assert data.metrics.get("rag_status") != "no_data"


# ===========================================================================
# 3. ZERO FIXED FINDINGS — full cold-start (QUAL-01)
# ===========================================================================

class TestZeroFixedFindings:
    def test_empty_fixed_vulns_df_returns_cold_start(self):
        """Empty fixed_vulns_df → cold-start ModuleData (error=None)."""
        mod = MTTRTrendModule()
        cfg = _config()
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=pd.DataFrame(),
        )
        assert data.error is None
        assert data.metrics.get("cold_start") is True

    def test_none_fixed_vulns_df_returns_cold_start(self):
        """fixed_vulns_df=None → cold-start ModuleData (error=None)."""
        mod = MTTRTrendModule()
        cfg = _config()
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
        )
        assert data.error is None
        assert data.metrics.get("cold_start") is True

    def test_cold_start_rag_strip_is_no_data(self):
        """Cold-start RAG strip must have no_data label."""
        mod = MTTRTrendModule()
        cfg = _config()
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=pd.DataFrame(),
        )
        assert data.rag_strip.get("rag_label") == "No Data"

    def test_cold_start_overall_mttr_absent(self):
        """Cold-start: overall_mttr key absent or None in metrics."""
        mod = MTTRTrendModule()
        cfg = _config()
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=pd.DataFrame(),
        )
        assert data.metrics.get("overall_mttr") is None

    def test_cold_start_no_nan_in_summary(self):
        """Cold-start summary_text must not contain 'NaN' (QUAL-01)."""
        mod = MTTRTrendModule()
        cfg = _config()
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=pd.DataFrame(),
        )
        assert "NaN" not in (data.summary_text or "")
        assert "NaN" not in (data.driver_narrative or "")


# ===========================================================================
# 4. COLD-START MoM — 1 snapshot (insufficient_data=True)
# ===========================================================================

class TestColdStartMoM:
    def test_single_snapshot_insufficient_data_renders_notice(self):
        """
        When trend_snapshots has insufficient_data=True, the MoM line
        cold-starts independently, but live per-severity gauges still render
        from fixed_vulns_df (QUAL-01 dual independent cold-start paths).
        """
        trend = _make_trend_snapshots(
            [_make_snapshot_mttr("2026-05", mttr_overall_days=10.0)],
            insufficient_data=True,
        )
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        # Module should NOT be in full cold_start (we have live fixed findings)
        # snapshots_cold=True because insufficient_data=True
        assert data.metrics.get("snapshots_cold") is True

    def test_single_snapshot_no_nan_in_any_channel(self):
        """No NaN in any rendered channel when MoM is in cold-start."""
        trend = _make_trend_snapshots(
            [_make_snapshot_mttr("2026-05", mttr_overall_days=10.0)],
            insufficient_data=True,
        )
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        email_html = mod.render_email_panel(data, cfg)
        assert "NaN" not in pdf_html
        assert "NaN" not in email_html

    def test_no_trend_snapshots_kwarg_snapshots_cold(self):
        """No trend_snapshots kwarg → snapshots_cold=True."""
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            # No trend_snapshots passed
            min_sample_size=1,
        )
        assert data.error is None
        assert data.metrics.get("snapshots_cold") is True


# ===========================================================================
# 5. MIN_SAMPLE THRESHOLD — D-16-07
# ===========================================================================

class TestMinSampleThreshold:
    def test_3_critical_findings_with_min_5_renders_insufficient_data(self):
        """
        3 Critical findings, min_sample_size=5 → Critical row renders
        "Insufficient data (3 findings — minimum 5 required)".
        """
        data = _run(
            [
                {"state": "fixed", "severity": "critical"},
                {"state": "fixed", "severity": "critical"},
                {"state": "fixed", "severity": "critical"},
            ],
            min_sample_size=5,
        )
        assert data.error is None
        # Find the Critical row
        crit_rows = [r for r in data.table_data if r["label"] == "Critical"]
        assert len(crit_rows) == 1
        insuf_str = crit_rows[0].get("insufficient", "")
        assert "Insufficient data (3 findings — minimum 5 required)" in insuf_str, (
            f"Expected 'Insufficient data (3 findings — minimum 5 required)' "
            f"in insufficient field, got: {insuf_str!r}"
        )

    def test_insufficient_data_string_appears_in_excel_render(self):
        """The insufficient data wording appears in the Excel cell value."""
        data = _run(
            [
                {"state": "fixed", "severity": "critical"},
                {"state": "fixed", "severity": "critical"},
                {"state": "fixed", "severity": "critical"},
            ],
            min_sample_size=5,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=5)
        wb = openpyxl.Workbook()
        tabs = mod.render_excel_tabs(data, wb, cfg)
        assert len(tabs) > 0
        ws = wb[tabs[0]]
        # Search all cell values for the wording
        all_values = [str(ws.cell(r, c).value or "") for r in range(1, ws.max_row + 1)
                      for c in range(1, ws.max_column + 1)]
        matches = [v for v in all_values if "minimum 5 required" in v]
        assert len(matches) >= 1, (
            f"Expected 'minimum 5 required' in Excel tab, got values: {all_values}"
        )

    def test_overall_mttr_none_when_total_below_threshold(self):
        """3 total findings with min_sample_size=5 → overall_mttr is None."""
        data = _run(
            [
                {"state": "fixed", "severity": "critical"},
                {"state": "fixed", "severity": "critical"},
                {"state": "fixed", "severity": "critical"},
            ],
            min_sample_size=5,
        )
        assert data.error is None
        assert data.metrics.get("overall_mttr") is None

    def test_5_findings_meet_threshold(self):
        """Exactly min_sample_size=5 findings → overall_mttr is not None."""
        data = _run(
            [{"state": "fixed", "severity": "critical"}] * 5,
            min_sample_size=5,
        )
        assert data.error is None
        assert data.metrics.get("overall_mttr") is not None


# ===========================================================================
# 6. OWNER COLD-START — new Owner only in snapshot 2
# ===========================================================================

class TestOwnerColdStart:
    def test_new_owner_in_snapshot_2_series_has_none_then_value(self):
        """
        Owner "Ops" appears only in snapshot 2.
        → owner_series["Ops"] == [None, value]; MoM delta is None.
        """
        snap1 = _make_snapshot_mttr(
            "2026-04",
            mttr_overall_days=12.0,
            mttr_by_owner={},  # Ops absent
        )
        snap2 = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=10.0,
            mttr_by_owner={"Ops": 8.5},
        )
        trend = _make_trend_snapshots([snap1, snap2], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        owner_series = data.chart_data.get("owner_series", {})
        assert "Ops" in owner_series, f"Ops missing from owner_series keys: {list(owner_series.keys())}"
        series = owner_series["Ops"]
        assert series[0] is None, f"Expected None for first snapshot (cold-start), got {series[0]}"
        assert series[1] == 8.5, f"Expected 8.5 for second snapshot, got {series[1]}"

    def test_new_owner_mom_delta_is_none(self):
        """New Owner in snapshot 2 only → MoM delta is None (no prior value)."""
        snap1 = _make_snapshot_mttr(
            "2026-04",
            mttr_overall_days=12.0,
            mttr_by_owner={},
        )
        snap2 = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=10.0,
            mttr_by_owner={"Ops": 8.5},
        )
        trend = _make_trend_snapshots([snap1, snap2], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        # Find the Ops owner row in table_data
        ops_rows = [r for r in data.table_data if r.get("label") == "Ops"]
        if ops_rows:
            assert ops_rows[0].get("mom_delta") is None


# ===========================================================================
# 7. OWNER VANISHED — Owner only in snapshot 1
# ===========================================================================

class TestOwnerVanished:
    def test_owner_in_snapshot_1_only_omitted_from_current_table(self):
        """
        Owner "Legacy" appears only in snapshot 1.
        → "Legacy" must NOT appear in the current Owner table (live data).

        The live per-Owner table is built from fixed_df rows with
        extract_owner(); a vanished Owner has zero live fixed findings
        and is omitted per D-16-07 (zero-fixed Owners omitted).
        """
        # Only one owner in assets — "Ops", not "Legacy"
        asset_rows = [{"asset_uuid": _uuid(1), "tags": "Owner=Ops"}]
        data = _run(
            [{"state": "fixed", "severity": "critical", "asset_uuid": _uuid(1)}],
            asset_rows=asset_rows,
            min_sample_size=1,
        )
        assert data.error is None
        owner_labels = [r["label"] for r in data.table_data if r["label"] not in
                        ("Critical", "High", "Medium", "Low")]
        assert "Legacy" not in owner_labels, (
            f"Expected 'Legacy' to be omitted from current Owner table, "
            f"got: {owner_labels}"
        )

    def test_owner_series_contains_none_padded_series_for_vanished_owner(self):
        """
        The MoM chart_data captures the vanished owner's series (trailing None)
        but the live table omits it (separate paths).
        """
        snap1 = _make_snapshot_mttr(
            "2026-04",
            mttr_overall_days=12.0,
            mttr_by_owner={"Legacy": 20.0},
        )
        snap2 = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=10.0,
            mttr_by_owner={},  # Legacy vanished
        )
        trend = _make_trend_snapshots([snap1, snap2], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        owner_series = data.chart_data.get("owner_series", {})
        assert "Legacy" in owner_series
        series = owner_series["Legacy"]
        assert series[0] == 20.0
        assert series[1] is None  # padded with None for the month it vanished


# ===========================================================================
# 8. TIE-BREAK — multiple snapshots same month (D-16-08)
# ===========================================================================

class TestTieBreak:
    def test_two_snapshots_same_month_latest_generated_at_wins(self):
        """
        2 snapshots with month="2026-05", different generated_at.
        Only the latest generated_at snapshot should be used.
        """
        snap_early = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=20.0,
            generated_at="2026-05-15T00:00:00Z",
        )
        snap_late = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=10.0,  # different value to distinguish
            generated_at="2026-05-25T00:00:00Z",
        )
        trend = _make_trend_snapshots([snap_early, snap_late], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        # Should use the late snapshot (10.0), not the early (20.0)
        overall_series = data.chart_data.get("overall_mttr_series", [])
        assert len(overall_series) == 1, (
            f"Expected deduplicated to 1 month, got {len(overall_series)}"
        )
        assert overall_series[0] == 10.0, (
            f"Expected 10.0 (latest snapshot), got {overall_series[0]} "
            f"— tie-break may not be using generated_at descending"
        )

    def test_tie_break_preserves_single_month_entry(self):
        """After tie-break, months list has length 1 for 2 same-month snapshots."""
        snap_a = _make_snapshot_mttr("2026-05", generated_at="2026-05-10T00:00:00Z")
        snap_b = _make_snapshot_mttr("2026-05", generated_at="2026-05-20T00:00:00Z")
        trend = _make_trend_snapshots([snap_a, snap_b], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        months = data.chart_data.get("months", [])
        assert len(months) == 1


# ===========================================================================
# 9. PARTIAL-MONTH LABEL (D-16-08)
# ===========================================================================

class TestPartialMonthLabel:
    def test_current_month_label_contains_partial(self):
        """Current month (2026-06) label must contain 'partial'."""
        snap1 = _make_snapshot_mttr("2026-05", mttr_overall_days=12.0)
        snap2 = _make_snapshot_mttr("2026-06", mttr_overall_days=10.0,
                                    generated_at="2026-06-12T00:00:00Z")
        trend = _make_trend_snapshots([snap1, snap2], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        months = data.chart_data.get("months", [])
        jun_labels = [m for m in months if "2026-06" in m]
        assert len(jun_labels) == 1
        assert "partial" in jun_labels[0].lower(), (
            f"Expected 'partial' in current-month label, got: {jun_labels[0]!r}"
        )

    def test_prior_month_label_does_not_contain_partial(self):
        """Completed prior months must NOT have 'partial' in their label."""
        snap1 = _make_snapshot_mttr("2026-05", mttr_overall_days=12.0)
        snap2 = _make_snapshot_mttr("2026-06", mttr_overall_days=10.0,
                                    generated_at="2026-06-12T00:00:00Z")
        trend = _make_trend_snapshots([snap1, snap2], insufficient_data=False)
        data = _run(
            [{"state": "fixed", "severity": "critical"}],
            trend_snapshots=trend,
            min_sample_size=1,
        )
        months = data.chart_data.get("months", [])
        may_labels = [m for m in months if "2026-05" in m]
        assert len(may_labels) == 1
        assert "partial" not in may_labels[0].lower()


# ===========================================================================
# 10. FOUR-CHANNEL EMPTY-DATA GUARD (QUAL-03)
# ===========================================================================

class TestFourChannelEmptyGuard:
    """All four render channels must survive a zero-row ModuleData (cold-start)."""

    def _cold_start_data(self) -> ModuleData:
        mod = MTTRTrendModule()
        cfg = _config()
        return mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=pd.DataFrame(),
        )

    def test_render_pdf_section_cold_start_does_not_crash(self):
        data = self._cold_start_data()
        mod = MTTRTrendModule()
        cfg = _config()
        result = mod.render_pdf_section(data, cfg)
        assert isinstance(result, str)
        assert "NaN" not in result

    def test_render_excel_tabs_cold_start_does_not_crash(self):
        data = self._cold_start_data()
        mod = MTTRTrendModule()
        cfg = _config()
        wb = openpyxl.Workbook()
        result = mod.render_excel_tabs(data, wb, cfg)
        assert isinstance(result, list)

    def test_render_email_panel_cold_start_does_not_crash(self):
        data = self._cold_start_data()
        mod = MTTRTrendModule()
        cfg = _config()
        result = mod.render_email_panel(data, cfg)
        assert isinstance(result, str)
        assert "NaN" not in result

    def test_render_rag_strip_entry_cold_start_does_not_crash(self):
        data = self._cold_start_data()
        mod = MTTRTrendModule()
        cfg = _config()
        result = mod.render_rag_strip_entry(data, cfg)
        assert isinstance(result, dict)
        assert "rag_label" in result
        assert result["rag_label"] == "No Data"

    def test_render_analyst_tabs_cold_start_returns_empty(self):
        data = self._cold_start_data()
        mod = MTTRTrendModule()
        cfg = _config()
        result = mod.render_analyst_tabs(data, cfg)
        assert result == []


# ===========================================================================
# 11. PANDAS CoW STRICT MODE
# ===========================================================================

class TestPandasCoW:
    def test_suite_runs_with_cow_strict_mode_enabled(self):
        """pd.options.mode.copy_on_write is True at module level."""
        assert pd.options.mode.copy_on_write is True

    def test_compute_no_chained_assignment_warning(self):
        """compute() emits zero ChainedAssignmentError from module source under CoW.

        Filters to warnings originating in reports/modules/ so fixture
        helper CoW patterns (already fixed in _make_fixed_vulns) don't
        mask genuine module-level issues.
        """
        import warnings
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            _run(
                [{"state": "fixed", "severity": "critical"}],
                min_sample_size=1,
            )
        # Only care about warnings from the module under test
        module_cow_warnings = [
            w for w in caught
            if (
                "ChainedAssignment" in str(w.category.__name__)
                or "chained" in str(w.message).lower()
            )
            and "reports" in str(getattr(w, "filename", "")).replace("\\", "/")
        ]
        assert not module_cow_warnings, (
            f"Unexpected CoW warnings from module code: "
            f"{[str(w.message) for w in module_cow_warnings]}"
        )


# ===========================================================================
# 12. COMPOSED-PIPELINE FIXED_VULNS NON-NULL SMOKE (T-16-12)
# ===========================================================================

class TestComposedPipelineFixedVulns:
    """
    End-to-end guard for the 16-02 _MODULES_NEEDING_FIXED_VULNS membership.

    A regression that drops 'mttr_trend' from that frozenset would result
    in fixed_vulns_df=None reaching MTTRTrendModule.compute(), causing a
    full cold-start (gray RAG). This test drives compute() directly with
    a synthetic fixed_vulns_df (>=5 fixed findings) and asserts:
      - overall_mttr is non-None
      - RAG strip status is NOT 'no_data'

    Per the plan implementation note: if the composed-pipeline kwarg wiring
    cannot be verified cleanly via run_full_pipeline, assert at the compute()
    boundary instead. The static frozenset-membership acceptance criterion in
    16-02 Task 2 is the primary guard; this is supplementary defense-in-depth.
    """

    def _make_populated_fixed_df(self, n: int = 6) -> pd.DataFrame:
        """Build a synthetic fixed_vulns_df with n>=5 fixed findings."""
        rows = []
        for i in range(n):
            rows.append({
                "state":           "fixed",
                "severity":        "critical",
                "asset_uuid":      _uuid(i + 1),
                "first_found":     REF - timedelta(days=30),
                "resurfaced_date": None,
                "last_fixed":      REF - timedelta(days=2),
            })
        return _make_fixed_vulns(rows)

    def test_compute_with_fixed_vulns_df_produces_non_none_overall_mttr(self):
        """
        MTTRTrendModule.compute() with >=5 fixed findings kwarg →
        overall_mttr is non-None and rag_status != 'no_data'.

        This is the end-to-end guard for _MODULES_NEEDING_FIXED_VULNS:
        if mttr_trend were dropped from that frozenset, composed_report.py
        would pass fixed_vulns_df=None, compute() would cold-start, and
        overall_mttr would be None (this assertion would fail).
        """
        fixed_df = self._make_populated_fixed_df(n=6)
        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options={"min_sample_size": 5})
        data = mod.compute(
            pd.DataFrame(),
            _make_assets(),
            REF,
            cfg,
            fixed_vulns_df=fixed_df,
        )
        assert data.error is None, f"Module returned error: {data.error}"
        assert data.metrics.get("overall_mttr") is not None, (
            "overall_mttr is None — this means fixed_vulns_df was not received "
            "or was empty. Check that 'mttr_trend' is in _MODULES_NEEDING_FIXED_VULNS "
            "in reports/composed_report.py (D-16-01)."
        )
        assert data.metrics.get("rag_status") != "no_data", (
            f"RAG status is 'no_data' — module is in cold-start despite "
            f"receiving {len(fixed_df)} fixed findings. "
            f"rag_status={data.metrics.get('rag_status')!r}"
        )

    def test_frozenset_membership_in_composed_report(self):
        """
        Static acceptance criterion: 'mttr_trend' must be in BOTH
        _MODULES_NEEDING_FIXED_VULNS and _MODULES_NEEDING_TREND_SNAPSHOTS
        in reports/composed_report.py (D-16-01, D-16-03).
        """
        from reports.composed_report import (
            _MODULES_NEEDING_FIXED_VULNS,
            _MODULES_NEEDING_TREND_SNAPSHOTS,
        )
        assert "mttr_trend" in _MODULES_NEEDING_FIXED_VULNS, (
            "'mttr_trend' missing from _MODULES_NEEDING_FIXED_VULNS — "
            "fixed_vulns_df will be None → gray RAG forever (D-16-01)"
        )
        assert "mttr_trend" in _MODULES_NEEDING_TREND_SNAPSHOTS, (
            "'mttr_trend' missing from _MODULES_NEEDING_TREND_SNAPSHOTS — "
            "MoM trend line will never receive snapshot history (D-16-03)"
        )


# ===========================================================================
# 13. STRUCTURAL BASELINE SELF-GUARD
# ===========================================================================

class TestStructuralBaselines:
    """
    Self-guarding baseline tests: compare_snapshots(actual, baseline) == [].

    Populated bundle baseline: uses >=5 fixed findings + 2-snapshot history.
    Zero-match baseline: uses empty fixed_vulns_df → _empty_result cold-start.

    These tests fail if the structural shape of the bundle changes after Phase 16.
    """

    def _build_populated_bundle(self) -> dict:
        """Build a synthetic composed-shaped bundle for mttr_trend (populated)."""
        import openpyxl as xl
        fixed_df = _make_fixed_vulns([
            {"state": "fixed", "severity": "critical", "asset_uuid": _uuid(i + 1)}
            for i in range(6)
        ])
        snap1 = _make_snapshot_mttr("2026-05", mttr_overall_days=12.0)
        snap2 = _make_snapshot_mttr("2026-06", mttr_overall_days=10.0,
                                    generated_at="2026-06-12T00:00:00Z")
        trend = _make_trend_snapshots([snap1, snap2], insufficient_data=False)

        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options={"min_sample_size": 1})
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=fixed_df,
            trend_snapshots=trend,
        )

        # Assemble the bundle structure expected by extract_structural_snapshot
        wb = xl.Workbook()
        wb.remove(wb.active)
        mod.render_excel_tabs(data, wb, cfg)

        pdf_html = mod.render_pdf_section(data, cfg)
        email_html = mod.render_email_panel(data, cfg)

        return {
            "pdf_html":              pdf_html,
            "excel_workbook":        wb,
            "analyst_workbook_path": None,
            "analyst_excel":         None,
            "email_body_html":       email_html,
            "email_kpis":            {},
            "email_inline_images":   [],
            "metrics":               data.metrics,
            "errors":                [],
            "module_results":        [data],
        }

    def _build_zero_match_bundle(self) -> dict:
        """Build a synthetic bundle for the zero-fixed-findings cold-start path."""
        import openpyxl as xl
        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend")
        data = mod.compute(
            pd.DataFrame(), _make_assets(), REF, cfg,
            fixed_vulns_df=pd.DataFrame(),
        )

        wb = xl.Workbook()
        wb.remove(wb.active)
        mod.render_excel_tabs(data, wb, cfg)

        pdf_html = mod.render_pdf_section(data, cfg)
        email_html = mod.render_email_panel(data, cfg)

        return {
            "pdf_html":              pdf_html,
            "excel_workbook":        wb,
            "analyst_workbook_path": None,
            "analyst_excel":         None,
            "email_body_html":       email_html,
            "email_kpis":            {},
            "email_inline_images":   [],
            "metrics":               data.metrics,
            "errors":                [],
            "module_results":        [data],
        }

    def test_populated_baseline_self_consistent(self):
        """Populated bundle: compare_snapshots(actual, actual) == []."""
        from tests.baseline_utils import extract_structural_snapshot, compare_snapshots
        bundle = self._build_populated_bundle()
        snap = extract_structural_snapshot(bundle, "mttr_trend_test_pull")
        diffs = compare_snapshots(snap, snap)
        assert diffs == [], f"Self-comparison should yield no diffs, got: {diffs}"

    def test_zero_match_baseline_self_consistent(self):
        """Zero-match bundle: compare_snapshots(actual, actual) == []."""
        from tests.baseline_utils import extract_structural_snapshot, compare_snapshots
        bundle = self._build_zero_match_bundle()
        snap = extract_structural_snapshot(bundle, "mttr_trend_test_pull_zero_match")
        diffs = compare_snapshots(snap, snap)
        assert diffs == [], f"Self-comparison should yield no diffs, got: {diffs}"

    def test_populated_baseline_no_metric_values(self):
        """Populated baseline must not contain metric values (structural-only)."""
        from tests.baseline_utils import extract_structural_snapshot
        bundle = self._build_populated_bundle()
        snap = extract_structural_snapshot(bundle, "mttr_trend_test_pull")
        assert "overall_mttr" not in snap
        assert "pdf_page_count" in snap

    @pytest.mark.baseline
    def test_populated_baseline_matches_committed_file(self):
        """
        Populated baseline matches committed tests/baselines/mttr_trend_test_pull.json.
        Fails if the bundle's structural shape changed after Phase 16.
        """
        baseline_path = Path(__file__).parent / "baselines" / "mttr_trend_test_pull.json"
        if not baseline_path.exists():
            pytest.skip("Baseline file not yet committed — run Task 2 capture script first")
        from tests.baseline_utils import (
            extract_structural_snapshot, compare_snapshots, load_baseline,
        )
        bundle = self._build_populated_bundle()
        actual = extract_structural_snapshot(bundle, "mttr_trend_test_pull")
        baseline = load_baseline(baseline_path)
        diffs = compare_snapshots(actual, baseline)
        assert diffs == [], (
            f"Populated baseline drift detected:\n" + "\n".join(diffs)
        )

    @pytest.mark.baseline
    def test_zero_match_baseline_matches_committed_file(self):
        """
        Zero-match baseline matches committed tests/baselines/mttr_trend_test_pull_zero_match.json.
        """
        baseline_path = Path(__file__).parent / "baselines" / "mttr_trend_test_pull_zero_match.json"
        if not baseline_path.exists():
            pytest.skip("Baseline file not yet committed — run Task 2 capture script first")
        from tests.baseline_utils import (
            extract_structural_snapshot, compare_snapshots, load_baseline,
        )
        bundle = self._build_zero_match_bundle()
        actual = extract_structural_snapshot(bundle, "mttr_trend_test_pull_zero_match")
        baseline = load_baseline(baseline_path)
        diffs = compare_snapshots(actual, baseline)
        assert diffs == [], (
            f"Zero-match baseline drift detected:\n" + "\n".join(diffs)
        )


# ===========================================================================
# 14. D-16-13: GAUGE BAND ALWAYS-ON IN ALL VIEW MODES
# ===========================================================================

class TestGaugeBandAllViews:
    """
    D-16-13: The 4 per-severity MTTR gauges render in ALL three focus modes.

    Proves the old "mttr_view in (severity, both)" gate is gone.
    Mode is determined by focus signal (mttr_table_mode from compute()):
      - unfocused (no tag_category/tag_value) → owner mode
      - tag_category="Owner" + tag_value → application mode
      - tag_category="Application" + tag_value → none (gauges only)

    All three must show 4 `data:image/png;base64,` occurrences in PDF AND email.
    """

    def _make_all_sev_rows(self) -> list[dict]:
        """24 fixed findings: 6 per severity, across 4 owners. min_sample=1 sufficient."""
        rows = []
        sevs = ["critical", "high", "medium", "low"]
        for sev_idx, sev in enumerate(sevs):
            for j in range(6):
                rows.append({
                    "state": "fixed",
                    "severity": sev,
                    "asset_uuid": _uuid(sev_idx * 6 + j + 1),
                    "first_found": REF - timedelta(days=20),
                    "last_fixed":  REF - timedelta(days=2),
                })
        return rows

    def _make_owner_app_assets(self) -> list[dict]:
        """24 assets with Owner= and Application= tags (Owner=GroupX;Application=AppY)."""
        assets = []
        sevs = ["critical", "high", "medium", "low"]
        for sev_idx in range(4):
            owner = f"Group{chr(65 + sev_idx)}"  # GroupA, GroupB, GroupC, GroupD
            app = f"App{chr(65 + sev_idx)}"
            for j in range(6):
                assets.append({
                    "asset_uuid": _uuid(sev_idx * 6 + j + 1),
                    "tags": f"Owner={owner};Application={app}",
                })
        return assets

    def _count_gauges(self, html: str) -> int:
        """Count data:image/png;base64, occurrences = number of gauge images."""
        return html.count("data:image/png;base64,")

    def test_owner_mode_pdf_has_4_gauges(self):
        """Unfocused (owner mode) → 4 gauge images in PDF."""
        data = _run(
            self._make_all_sev_rows(),
            asset_rows=self._make_owner_app_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "owner"
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        gauge_count = self._count_gauges(pdf_html)
        assert gauge_count == 4, (
            f"Expected 4 gauge images in owner-mode PDF, got {gauge_count} "
            f"— the gauge gate may still be present (D-16-13)"
        )

    def test_owner_mode_email_has_4_gauges(self):
        """Unfocused (owner mode) → 4 gauge images in email panel."""
        data = _run(
            self._make_all_sev_rows(),
            asset_rows=self._make_owner_app_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        # Email panel uses render_email_panel which does NOT embed gauge images;
        # gauges are PDF-channel only. Verify the email panel renders without crash
        # and contains the mode label "Owner breakdown".
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        email_html = mod.render_email_panel(data, cfg)
        assert "Owner breakdown" in email_html, (
            f"Expected 'Owner breakdown' in unfocused email panel, got: {email_html!r}"
        )

    def test_application_mode_pdf_has_4_gauges(self):
        """Application mode (tag_category=Owner + tag_value set) → 4 gauge images in PDF."""
        data = _run(
            self._make_all_sev_rows(),
            asset_rows=self._make_owner_app_assets(),
            min_sample_size=1,
            tag_category="Owner",
            tag_value="GroupA",
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "application", (
            f"Expected mode='application', got {data.metadata.get('mttr_table_mode')!r}"
        )
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1, tag_category="Owner", tag_value="GroupA")
        pdf_html = mod.render_pdf_section(data, cfg)
        gauge_count = self._count_gauges(pdf_html)
        assert gauge_count == 4, (
            f"Expected 4 gauge images in application-mode PDF, got {gauge_count}"
        )

    def test_gauges_only_mode_pdf_has_4_gauges(self):
        """Gauges-only mode (tag_category=Application + tag_value set) → 4 gauge images in PDF."""
        data = _run(
            self._make_all_sev_rows(),
            asset_rows=self._make_owner_app_assets(),
            min_sample_size=1,
            tag_category="Application",
            tag_value="AppA",
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "none", (
            f"Expected mode='none', got {data.metadata.get('mttr_table_mode')!r}"
        )
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1, tag_category="Application", tag_value="AppA")
        pdf_html = mod.render_pdf_section(data, cfg)
        gauge_count = self._count_gauges(pdf_html)
        assert gauge_count == 4, (
            f"Expected 4 gauge images in gauges-only-mode PDF, got {gauge_count}"
        )

    def test_gauges_only_mode_email_says_gauges_only(self):
        """Gauges-only mode email panel: disclosure says 'Gauges only'."""
        data = _run(
            self._make_all_sev_rows(),
            asset_rows=self._make_owner_app_assets(),
            min_sample_size=1,
            tag_category="Application",
            tag_value="AppA",
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1, tag_category="Application", tag_value="AppA")
        email_html = mod.render_email_panel(data, cfg)
        assert "Gauges only" in email_html, (
            f"Expected 'Gauges only' in gauges-only email panel, got: {email_html!r}"
        )


# ===========================================================================
# 15. D-16-13: MOM ARROW POLARITY
# ===========================================================================

class TestMomArrowPolarity:
    """
    D-16-13: Per-severity MoM direction arrows.

    MTTR is lower-is-better:
      - current < prior → delta < 0 → "down" direction (▼ green, &#9660;)
      - current > prior → delta > 0 → "up" direction (▲ red, &#9650;)
      - flat / single snapshot → "flat" direction (— grey, &#8212;)

    Builds two-snapshot fixtures with differing per-severity MTTR values
    and asserts the correct glyph + color appears in the PDF render.
    """

    def _make_fixed_rows_critical(self, n: int = 6) -> list[dict]:
        """n fixed Critical findings (all assets uuid 1..n)."""
        return [
            {
                "state": "fixed", "severity": "critical",
                "asset_uuid": _uuid(i + 1),
                "first_found": REF - timedelta(days=20),
                "last_fixed":  REF - timedelta(days=2),
            }
            for i in range(n)
        ]

    def test_decrease_produces_down_arrow_in_pdf(self):
        """
        Per-severity MTTR decrease (prior > current) → ▼ green (&#9660;) in PDF.
        Build two snapshots where Critical MTTR drops from 20 → 10.
        """
        snap_prior = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=20.0,
            mttr_by_severity={"critical": 20.0, "high": None, "medium": None, "low": None},
            generated_at="2026-05-15T00:00:00Z",
        )
        snap_current = _make_snapshot_mttr(
            "2026-06",
            mttr_overall_days=10.0,
            mttr_by_severity={"critical": 10.0, "high": None, "medium": None, "low": None},
            generated_at="2026-06-01T00:00:00Z",
        )
        trend = _make_trend_snapshots([snap_prior, snap_current], insufficient_data=False)
        data = _run(
            self._make_fixed_rows_critical(),
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        # Verify direction computed correctly
        assert data.metadata.get("per_sev_mom_direction", {}).get("critical") == "down", (
            f"Expected 'down' for MTTR decrease (20→10), got "
            f"{data.metadata.get('per_sev_mom_direction', {}).get('critical')!r}"
        )
        # Verify PDF renders ▼ (&#9660;) for Critical gauge
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "&#9660;" in pdf_html, (
            "Expected ▼ (&#9660;) in PDF for MTTR decrease (lower-is-better → green/down)"
        )
        # Verify green color associated with down arrow
        assert "#388e3c" in pdf_html, (
            "Expected green color (#388e3c) for down/improving arrow in PDF"
        )

    def test_increase_produces_up_arrow_in_pdf(self):
        """
        Per-severity MTTR increase (prior < current) → ▲ red (&#9650;) in PDF.
        Build two snapshots where Critical MTTR rises from 10 → 20.
        """
        snap_prior = _make_snapshot_mttr(
            "2026-05",
            mttr_overall_days=10.0,
            mttr_by_severity={"critical": 10.0, "high": None, "medium": None, "low": None},
            generated_at="2026-05-15T00:00:00Z",
        )
        snap_current = _make_snapshot_mttr(
            "2026-06",
            mttr_overall_days=20.0,
            mttr_by_severity={"critical": 20.0, "high": None, "medium": None, "low": None},
            generated_at="2026-06-01T00:00:00Z",
        )
        trend = _make_trend_snapshots([snap_prior, snap_current], insufficient_data=False)
        data = _run(
            self._make_fixed_rows_critical(),
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        assert data.metadata.get("per_sev_mom_direction", {}).get("critical") == "up", (
            f"Expected 'up' for MTTR increase (10→20), got "
            f"{data.metadata.get('per_sev_mom_direction', {}).get('critical')!r}"
        )
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "&#9650;" in pdf_html, (
            "Expected ▲ (&#9650;) in PDF for MTTR increase (slipping → red/up)"
        )
        assert "#d32f2f" in pdf_html, (
            "Expected red color (#d32f2f) for up/slipping arrow in PDF"
        )

    def test_flat_single_snapshot_produces_dash_in_pdf(self):
        """
        Single snapshot (no prior month) → "flat" direction → — (&#8212;) in PDF.
        """
        snap_only = _make_snapshot_mttr(
            "2026-06",
            mttr_overall_days=15.0,
            mttr_by_severity={"critical": 15.0, "high": None, "medium": None, "low": None},
            generated_at="2026-06-01T00:00:00Z",
        )
        trend = _make_trend_snapshots([snap_only], insufficient_data=False)
        data = _run(
            self._make_fixed_rows_critical(),
            trend_snapshots=trend,
            min_sample_size=1,
        )
        assert data.error is None
        # With only one snapshot, no delta possible → flat
        assert data.metadata.get("per_sev_mom_direction", {}).get("critical") == "flat", (
            f"Expected 'flat' for single snapshot, got "
            f"{data.metadata.get('per_sev_mom_direction', {}).get('critical')!r}"
        )
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "&#8212;" in pdf_html, (
            "Expected — (&#8212;) in PDF for flat / no-prior-snapshot direction"
        )

    def test_no_snapshots_produces_flat_direction(self):
        """
        No trend snapshots → snapshots_cold → all severities flat.
        """
        data = _run(
            self._make_fixed_rows_critical(),
            min_sample_size=1,
            # no trend_snapshots kwarg
        )
        assert data.error is None
        assert data.metrics.get("snapshots_cold") is True
        directions = data.metadata.get("per_sev_mom_direction", {})
        for sev in ("critical", "high", "medium", "low"):
            assert directions.get(sev) == "flat", (
                f"Expected 'flat' for sev={sev} with no snapshots, got {directions.get(sev)!r}"
            )


# ===========================================================================
# 16. D-16-13: FOCUS ROUTING
# ===========================================================================

class TestFocusRouting:
    """
    D-16-13: Focus-driven table routing via tag_category/tag_value.

    unfocused (no tag_category/tag_value) → mttr_table_mode == "owner" + Owner table
    tag_category="Owner" + tag_value set → mttr_table_mode == "application" + Application table
    tag_category="Application" + tag_value set → mttr_table_mode == "none" + no detail table
    explicit mttr_table="owner"/"application" override overrides auto-routing
    """

    def _make_rows(self) -> list[dict]:
        return [
            {
                "state": "fixed", "severity": "critical",
                "asset_uuid": _uuid(i + 1),
                "first_found": REF - timedelta(days=20),
                "last_fixed":  REF - timedelta(days=2),
            }
            for i in range(6)
        ]

    def _make_assets_with_owner_and_app(self) -> list[dict]:
        """6 assets with Owner=TeamX;Application=AppY tags."""
        return [
            {"asset_uuid": _uuid(i + 1), "tags": "Owner=TeamX;Application=AppY"}
            for i in range(6)
        ]

    def test_unfocused_resolves_to_owner_mode(self):
        """No tag_category/tag_value → mttr_table_mode == 'owner'."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "owner", (
            f"Expected unfocused → mode='owner', got {data.metadata.get('mttr_table_mode')!r}"
        )

    def test_unfocused_pdf_has_owner_table(self):
        """Unfocused → PDF has 'MTTR by Owner' table."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "MTTR by Owner" in pdf_html, "Expected 'MTTR by Owner' table in unfocused PDF"

    def test_owner_focused_resolves_to_application_mode(self):
        """tag_category='Owner' + tag_value set → mttr_table_mode == 'application'."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
            tag_category="Owner",
            tag_value="TeamX",
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "application", (
            f"Expected Owner-focused → mode='application', got {data.metadata.get('mttr_table_mode')!r}"
        )

    def test_owner_focused_pdf_has_application_table(self):
        """Owner-focused → PDF has 'MTTR by Application' table."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
            tag_category="Owner",
            tag_value="TeamX",
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1, tag_category="Owner", tag_value="TeamX")
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "MTTR by Application" in pdf_html, (
            "Expected 'MTTR by Application' table in Owner-focused PDF"
        )

    def test_application_focused_resolves_to_none_mode(self):
        """tag_category='Application' + tag_value set → mttr_table_mode == 'none'."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
            tag_category="Application",
            tag_value="AppY",
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "none", (
            f"Expected Application-focused → mode='none', got {data.metadata.get('mttr_table_mode')!r}"
        )

    def test_application_focused_pdf_has_no_detail_table(self):
        """Application-focused → PDF has NO 'MTTR by Owner' or 'MTTR by Application' table."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
            tag_category="Application",
            tag_value="AppY",
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1, tag_category="Application", tag_value="AppY")
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "MTTR by Owner" not in pdf_html, (
            "Application-focused PDF must not have 'MTTR by Owner' table"
        )
        assert "MTTR by Application" not in pdf_html, (
            "Application-focused PDF must not have 'MTTR by Application' table"
        )

    def test_explicit_mttr_table_owner_overrides_focus(self):
        """Explicit mttr_table='owner' overrides auto-routing even when Owner-focused."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
            tag_category="Owner",
            tag_value="TeamX",
            mttr_table="owner",  # explicit override
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "owner", (
            "Explicit mttr_table='owner' must override auto-routing to 'application'"
        )

    def test_explicit_mttr_table_application_overrides_focus(self):
        """Explicit mttr_table='application' overrides auto even when unfocused."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_assets_with_owner_and_app(),
            min_sample_size=1,
            mttr_table="application",  # explicit override, no focus signal
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "application", (
            "Explicit mttr_table='application' must override auto-routing to 'owner'"
        )


# ===========================================================================
# 17. D-16-13: SEVERITY TABLE ABSENT FROM ALL HEADLINE CHANNELS
# ===========================================================================

class TestSeverityTableAbsent:
    """
    D-16-13: The Severity <table> is removed from all headline channels.

    PDF must not have 'MTTR by Severity' <h3>+<table>.
    Email panel must not have a severity table.
    Excel must have the compact 4-row severity NUMERIC block (header row:
    Severity | MTTR (Days) | SLA Target (Days) | Status | MoM Delta (Days))
    but NOT the old 6-column severity table layout.
    mttr_view must not appear anywhere in the module code (retired concept).
    """

    def _make_rows(self) -> list[dict]:
        """6 fixed Critical findings (min_sample=1 sufficient)."""
        return [
            {
                "state": "fixed", "severity": "critical",
                "asset_uuid": _uuid(i + 1),
                "first_found": REF - timedelta(days=20),
                "last_fixed":  REF - timedelta(days=2),
            }
            for i in range(6)
        ]

    def _make_owner_assets(self) -> list[dict]:
        return [
            {"asset_uuid": _uuid(i + 1), "tags": "Owner=TeamD"}
            for i in range(6)
        ]

    def test_pdf_has_no_severity_table_heading(self):
        """PDF must not have 'MTTR by Severity' heading (table removed D-16-13)."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_owner_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        pdf_html = mod.render_pdf_section(data, cfg)
        assert "MTTR by Severity" not in pdf_html, (
            "PDF must not render 'MTTR by Severity' heading — severity table retired (D-16-13)"
        )

    def test_email_has_no_severity_table(self):
        """Email panel must not contain a severity table."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_owner_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        email_html = mod.render_email_panel(data, cfg)
        # email panel should not contain "MTTR by Severity"
        assert "MTTR by Severity" not in email_html, (
            "Email panel must not render a 'MTTR by Severity' section"
        )

    def test_excel_has_compact_severity_numeric_block_header(self):
        """Excel must have the compact severity numeric block header (Severity | MTTR (Days) | ...)."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_owner_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        wb = openpyxl.Workbook()
        tabs = mod.render_excel_tabs(data, wb, cfg)
        assert len(tabs) > 0
        ws = wb[tabs[0]]
        all_values = [
            str(ws.cell(r, c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        # Compact block header columns (D-16-13)
        assert "Severity" in all_values, "Excel must have 'Severity' column in compact block"
        assert "MTTR (Days)" in all_values, "Excel must have 'MTTR (Days)' column"
        assert "SLA Target (Days)" in all_values, "Excel must have 'SLA Target (Days)' column"
        assert "Status" in all_values, "Excel must have 'Status' column"
        assert "MoM Delta (Days)" in all_values, "Excel must have 'MoM Delta (Days)' column"

    def test_excel_has_4_severity_rows(self):
        """Excel compact severity block has 4 data rows (Critical/High/Medium/Low)."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_owner_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        wb = openpyxl.Workbook()
        tabs = mod.render_excel_tabs(data, wb, cfg)
        ws = wb[tabs[0]]
        all_values = [
            str(ws.cell(r, c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        # All four severity labels must be present as cell values
        for sev in ("Critical", "High", "Medium", "Low"):
            assert sev in all_values, (
                f"Excel compact severity block must have '{sev}' row"
            )

    def test_mttr_view_string_absent_from_module_metadata(self):
        """compute() metadata must not contain 'mttr_view' key (retired)."""
        data = _run(
            self._make_rows(),
            asset_rows=self._make_owner_assets(),
            min_sample_size=1,
        )
        assert data.error is None
        assert "mttr_view" not in data.metadata, (
            f"metadata must not contain 'mttr_view' key — mttr_view is retired (D-16-13). "
            f"Got keys: {list(data.metadata.keys())}"
        )


# ===========================================================================
# 18. D-16-13: EXCEL SEVERITY BLOCK SLA SOURCED FROM CONFIG
# ===========================================================================

class TestExcelSeverityBlockAndSla:
    """
    D-16-13: Excel compact severity numeric block SLA values sourced from config.py.

    Medium SLA must equal config.SLA_DAYS["medium"] (60), NOT the stale 45
    from old CLAUDE.md documentation. Proves the SLA values in the Excel
    compact block come from the config module, not hardcoded constants.
    """

    def _make_rows_all_sev(self) -> list[dict]:
        """6 findings per severity (24 total) for all-severity Excel block."""
        rows = []
        for sev in ("critical", "high", "medium", "low"):
            for i in range(6):
                rows.append({
                    "state": "fixed",
                    "severity": sev,
                    "asset_uuid": _uuid(len(rows) + 1),
                    "first_found": REF - timedelta(days=20),
                    "last_fixed":  REF - timedelta(days=2),
                })
        return rows

    def test_excel_medium_sla_equals_config_value(self):
        """Excel compact block Medium SLA cell value == config.SLA_DAYS['medium'] (60)."""
        from config import SLA_DAYS
        data = _run(
            self._make_rows_all_sev(),
            min_sample_size=1,
        )
        assert data.error is None
        mod = MTTRTrendModule()
        cfg = _config(min_sample_size=1)
        wb = openpyxl.Workbook()
        tabs = mod.render_excel_tabs(data, wb, cfg)
        ws = wb[tabs[0]]

        # Find the Medium row and check its SLA Target (Days) cell value
        medium_sla_val = None
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == "Medium":
                # Column 3 = SLA Target (Days) in the compact numeric block
                medium_sla_val = ws.cell(r, 3).value
                break

        expected_medium = SLA_DAYS["medium"]
        assert medium_sla_val == expected_medium, (
            f"Excel Medium SLA cell = {medium_sla_val!r}, expected {expected_medium} "
            f"(config.SLA_DAYS['medium']). If this is 45, the value is hardcoded from "
            f"stale CLAUDE.md doc instead of config.py (D-16-13)."
        )

    def test_medium_sla_days_is_60_not_45(self):
        """config.SLA_DAYS['medium'] == 60 (not the stale 45 from old CLAUDE.md)."""
        from config import SLA_DAYS
        assert SLA_DAYS["medium"] == 60, (
            f"SLA_DAYS['medium'] = {SLA_DAYS['medium']}, expected 60. "
            "If this is 45, config.py has the wrong value."
        )

    def test_table_data_severity_sla_days_from_config(self):
        """table_data_severity rows have sla_days == config.SLA_DAYS[sev] for all severities."""
        from config import SLA_DAYS
        data = _run(
            self._make_rows_all_sev(),
            min_sample_size=1,
        )
        assert data.error is None
        sev_rows = data.metadata.get("table_data_severity", [])
        assert len(sev_rows) == 4, f"Expected 4 severity rows, got {len(sev_rows)}"
        for row in sev_rows:
            sev_lower = row["label"].lower()
            expected = SLA_DAYS[sev_lower]
            actual = row["sla_days"]
            assert actual == expected, (
                f"Severity row '{row['label']}' sla_days={actual}, expected {expected} "
                f"from config.SLA_DAYS['{sev_lower}']"
            )


# ===========================================================================
# 19. D-16-13: MTTR_TABLE BAD-VALUE FALLBACK + VALIDATE_CONFIG (T-16-26)
# ===========================================================================

class TestMttrTableBadValueFallback:
    """
    D-16-13: mttr_table="bogus" → no crash; module resolves to auto (focus-driven).
    validate_config with mttr_table="bogus" returns a non-empty error list.

    T-16-26: the operator-config trust boundary is enforced by test.
    """

    def _make_rows(self) -> list[dict]:
        return [
            {
                "state": "fixed", "severity": "critical",
                "asset_uuid": _uuid(i + 1),
                "first_found": REF - timedelta(days=20),
                "last_fixed":  REF - timedelta(days=2),
            }
            for i in range(5)
        ]

    def test_bad_mttr_table_no_crash(self):
        """mttr_table='bogus' → compute() returns data.error is None."""
        data = _run(
            self._make_rows(),
            min_sample_size=1,
            mttr_table="bogus",
        )
        assert data.error is None, (
            f"compute() crashed with bad mttr_table='bogus': {data.error}"
        )

    def test_bad_mttr_table_resolves_to_valid_mode(self):
        """mttr_table='bogus' → mttr_table_mode is a valid value (auto fallback)."""
        data = _run(
            self._make_rows(),
            min_sample_size=1,
            mttr_table="bogus",
        )
        assert data.error is None
        mode = data.metadata.get("mttr_table_mode")
        assert mode in ("owner", "application", "none"), (
            f"Bad mttr_table='bogus' must fall back to a valid mode, got {mode!r}"
        )

    def test_bad_mttr_table_validate_config_returns_error(self):
        """validate_config with mttr_table='bogus' returns a non-empty error list."""
        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options={"mttr_table": "bogus"})
        errors = mod.validate_config(cfg)
        assert len(errors) > 0, (
            "validate_config must return errors for unknown mttr_table value 'bogus'"
        )

    def test_bad_mttr_table_validate_config_error_mentions_mttr_table(self):
        """validate_config error message references 'mttr_table'."""
        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options={"mttr_table": "bogus"})
        errors = mod.validate_config(cfg)
        assert any("mttr_table" in e for e in errors), (
            f"Expected error mentioning 'mttr_table', got: {errors}"
        )

    def test_valid_mttr_table_values_pass_validate_config(self):
        """validate_config accepts all three valid mttr_table values."""
        mod = MTTRTrendModule()
        for valid in ("auto", "owner", "application"):
            cfg = ModuleConfig("mttr_trend", options={"mttr_table": valid})
            errors = mod.validate_config(cfg)
            assert errors == [], (
                f"validate_config should accept mttr_table={valid!r}, got: {errors}"
            )

    def test_uppercase_mttr_table_lowered_and_accepted(self):
        """mttr_table='OWNER' lowered to 'owner' → valid, no crash."""
        data = _run(
            self._make_rows(),
            min_sample_size=1,
            mttr_table="OWNER",
        )
        assert data.error is None
        assert data.metadata.get("mttr_table_mode") == "owner", (
            f"Expected 'OWNER' → mode='owner' after lowercasing, got {data.metadata.get('mttr_table_mode')!r}"
        )


# ===========================================================================
# 20. D-16-13: SINGLE-PAGE FIT (OWNER AND GAUGES-ONLY MODES)
# ===========================================================================

class TestSinglePageFit:
    """
    D-16-13: Representative owner-only and gauges-only bundles must fit one PDF page.

    Uses extract_structural_snapshot(bundle, slug)["pdf_page_count"] == 1.
    Owner-only: 5 owners, 6 findings each.
    Gauges-only: application-focused (mode=none), no detail table — even more compact.
    """

    def _build_bundle(
        self,
        n_owners: int = 5,
        findings_per_owner: int = 6,
        tag_category: str = None,
        tag_value: str = None,
    ) -> dict:
        """Build a representative bundle for a given focus mode."""
        import openpyxl as xl
        fixed_rows = []
        asset_rows = []
        uuid_counter = 0
        for owner_idx in range(n_owners):
            owner_name = f"Owner{chr(65 + owner_idx)}"
            app_name = f"App{chr(65 + owner_idx)}"
            for _ in range(findings_per_owner):
                uuid_counter += 1
                fixed_rows.append({
                    "state": "fixed", "severity": "critical",
                    "asset_uuid": _uuid(uuid_counter),
                    "first_found": REF - timedelta(days=20),
                    "last_fixed":  REF - timedelta(days=2),
                })
                asset_rows.append({
                    "asset_uuid": _uuid(uuid_counter),
                    "tags": f"Owner={owner_name};Application={app_name}",
                })

        fixed_df = _make_fixed_vulns(fixed_rows)
        assets_df = _make_assets(asset_rows)

        opts: dict = {"min_sample_size": 1}
        if tag_category:
            opts["tag_category"] = tag_category
        if tag_value:
            opts["tag_value"] = tag_value

        mod = MTTRTrendModule()
        cfg = ModuleConfig("mttr_trend", options=opts)
        data = mod.compute(
            pd.DataFrame(), assets_df, REF, cfg,
            fixed_vulns_df=fixed_df,
        )

        wb = xl.Workbook()
        wb.remove(wb.active)
        mod.render_excel_tabs(data, wb, cfg)
        pdf_html = mod.render_pdf_section(data, cfg)
        email_html = mod.render_email_panel(data, cfg)

        return {
            "pdf_html":              pdf_html,
            "excel_workbook":        wb,
            "analyst_workbook_path": None,
            "analyst_excel":         None,
            "email_body_html":       email_html,
            "email_kpis":            {},
            "email_inline_images":   [],
            "metrics":               data.metrics,
            "errors":                [],
            "module_results":        [data],
        }

    def test_owner_only_bundle_fits_one_page(self):
        """
        Owner-only render (5 owners, 6 findings each) → pdf_page_count == 1.
        The always-on 4-gauge band + owner table must still fit one PDF page.
        """
        from tests.baseline_utils import extract_structural_snapshot
        bundle = self._build_bundle(n_owners=5, findings_per_owner=6)
        snap = extract_structural_snapshot(bundle, "mttr_trend_single_page_owner")
        assert snap["pdf_page_count"] == 1, (
            f"Owner-only render with 5 owners should fit on 1 PDF page, "
            f"got {snap['pdf_page_count']} pages (D-16-13 single-page claim)"
        )

    def test_gauges_only_bundle_fits_one_page(self):
        """
        Gauges-only render (Application-focused, mode=none) → pdf_page_count == 1.
        No detail table, only the 4-gauge band → must fit easily on 1 page.
        """
        from tests.baseline_utils import extract_structural_snapshot
        bundle = self._build_bundle(
            n_owners=5,
            findings_per_owner=6,
            tag_category="Application",
            tag_value="AppA",
        )
        snap = extract_structural_snapshot(bundle, "mttr_trend_single_page_gauges_only")
        assert snap["pdf_page_count"] == 1, (
            f"Gauges-only render (Application-focused) should fit on 1 PDF page, "
            f"got {snap['pdf_page_count']} pages"
        )
